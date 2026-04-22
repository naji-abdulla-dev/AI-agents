"""
The Trade Desk Inc. (TTD) Financial Analysis - Excel Generator
Data as of April 2026 | FY2025 Annual (ended December 31, 2025)
Largest independent demand-side platform (DSP) in programmatic advertising
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

# ── The Trade Desk Color Palette ──────────────────────────────────────────────
TTD_BLUE     = "0095D9"   # brand teal-blue
TTD_DARK     = "003366"   # dark navy
TTD_MID      = "0066A1"   # mid blue
HEADER_BG    = "003366"
HEADER_FG    = "FFFFFF"
SUBHDR_BG    = "D6EAF8"
SUBHDR_FG    = "003366"
ALT_ROW      = "EBF5FB"
WHITE        = "FFFFFF"
LIGHT_GRAY   = "F8F8F8"
GREEN        = "27AE60"
RED          = "C0392B"
ORANGE       = "E67E22"
GOLD         = "F39C12"
DARK_GRAY    = "555555"
LIGHT_BLUE   = "EBF5FB"

FONT_SIZE = 14

def mf(bold=False, size=FONT_SIZE, color="000000", italic=False):
    return Font(name="Calibri", bold=bold, size=size, color=color, italic=italic)

def mfill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def mborder(color="AAAAAA"):
    s = Side(border_style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def cal(h="left", wrap=False):
    return Alignment(horizontal=h, vertical="center", wrap_text=wrap)

def wc(ws, row, col, value, bold=False, size=FONT_SIZE, fg="000000", bg=None,
       align="left", wrap=False, border=False, italic=False, num_fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = mf(bold=bold, size=size, color=fg, italic=italic)
    if bg:
        c.fill = mfill(bg)
    c.alignment = cal(align, wrap)
    if border:
        c.border = mborder()
    if num_fmt:
        c.number_format = num_fmt
    return c

def header_row(ws, row, cols, labels, bg=HEADER_BG, fg=HEADER_FG):
    for col, label in zip(cols, labels):
        wc(ws, row, col, label, bold=True, fg=fg, bg=bg, align="center", border=True)

def section_header(ws, row, col, label, span=1, bg=SUBHDR_BG, fg=SUBHDR_FG):
    wc(ws, row, col, label, bold=True, fg=fg, bg=bg, align="left", border=True)
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col + span - 1)

def set_col_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

def freeze(ws, cell="B3"):
    ws.freeze_panes = cell

def data_row(ws, row, col_start, label, values, bold=False, num_fmt=None,
             label_bg=None, val_bg=None, pct=False, alt=False):
    bg = ALT_ROW if alt else None
    if label_bg:
        bg_l = label_bg
    else:
        bg_l = bg
    wc(ws, row, col_start, label, bold=bold, bg=bg_l, border=True)
    for i, v in enumerate(values):
        fmt = num_fmt
        if pct and fmt is None:
            fmt = "0.0%"
        wc(ws, row, col_start + 1 + i, v, bold=bold, bg=val_bg if val_bg else bg,
           align="right", border=True, num_fmt=fmt)


# ══════════════════════════════════════════════════════════════════════════════
# TAB 1 – COVER
# ══════════════════════════════════════════════════════════════════════════════
def build_cover(wb):
    ws = wb.create_sheet("1. Cover")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, {1: 5, 2: 38, 3: 28, 4: 22, 5: 22, 6: 18})

    # Title block
    ws.row_dimensions[2].height = 65
    ws.merge_cells("B2:F2")
    c = ws["B2"]
    c.value = "THE TRADE DESK, INC. (NASDAQ: TTD)"
    c.font = Font(name="Calibri", bold=True, size=34, color=HEADER_FG)
    c.fill = mfill(TTD_DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[3].height = 32
    ws.merge_cells("B3:F3")
    c = ws["B3"]
    c.value = "Comprehensive Financial Analysis — April 2026"
    c.font = Font(name="Calibri", bold=False, size=18, color=HEADER_FG)
    c.fill = mfill(TTD_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[4].height = 22
    ws.merge_cells("B4:F4")
    c = ws["B4"]
    c.value = "Largest Independent Demand-Side Platform (DSP) in Programmatic Advertising"
    c.font = Font(name="Calibri", bold=False, size=14, color=TTD_DARK, italic=True)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # ── Key facts ──────────────────────────────────────────────────────────
    facts_title_row = 6
    ws.row_dimensions[facts_title_row].height = 26
    ws.merge_cells(f"B{facts_title_row}:F{facts_title_row}")
    wc(ws, facts_title_row, 2, "KEY COMPANY FACTS", bold=True, fg=HEADER_FG,
       bg=TTD_MID, align="center", size=16)

    facts = [
        ("Sector",             "Technology — AdTech / Programmatic Advertising"),
        ("Exchange",           "NASDAQ: TTD"),
        ("Headquarters",       "Ventura, California, USA"),
        ("Founded",            "2009 by Jeff Green & Dave Pickles"),
        ("CEO",                "Jeff Green (Co-Founder, ~11.1% insider ownership)"),
        ("Platform",           "Kokai AI-Powered DSP (Demand-Side Platform)"),
        ("Business Model",     "SaaS take-rate (~21%) on digital ad spend through the platform"),
        ("Customers",          "Agencies & brands globally; >95% retention for 11 consecutive years"),
        ("FY 2025 Revenue",    "$2.90 Billion (+18.5% YoY)"),
        ("FY 2025 Net Income", "$443.3 Million (15.3% margin)"),
        ("FY 2025 Adj EBITDA", "$1.20 Billion (41.4% margin)"),
        ("FY 2025 Free CF",    "$795.7 Million (27.5% FCF margin)"),
        ("Stock Price (Apr 13)","~$21.00 per share"),
        ("Market Cap",         "~$9.9 Billion"),
        ("52-Week Range",      "$19.74 – $91.45"),
        ("Analyst Avg Target", "$46.63 (32 analysts — consensus Buy)"),
        ("Key Risk",           "Publicis split + agency fee disputes; CFO departure; walled garden competition"),
        ("Rule of 50",         "18.5% growth + 27.5% FCF margin = 46.0 (SaaS benchmark; near threshold)"),
    ]

    for i, (label, value) in enumerate(facts):
        r = facts_title_row + 1 + i
        ws.row_dimensions[r].height = 22
        alt = i % 2 == 0
        bg = ALT_ROW if alt else WHITE
        wc(ws, r, 2, label, bold=True, fg=TTD_DARK, bg=bg, border=True)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
        wc(ws, r, 3, value, bold=False, fg="222222", bg=bg, border=True, wrap=True)

    # ── Investment Thesis ──────────────────────────────────────────────────
    thesis_start = facts_title_row + len(facts) + 2
    ws.row_dimensions[thesis_start].height = 26
    ws.merge_cells(f"B{thesis_start}:F{thesis_start}")
    wc(ws, thesis_start, 2, "INVESTMENT THESIS SUMMARY", bold=True, fg=HEADER_FG,
       bg=TTD_DARK, align="center", size=16)

    thesis = [
        ("BULL CASE",   "TTD owns the open internet's buy-side infrastructure. At $21, it trades at ~3x EV/Revenue "
                        "and <12x FCF — historically cheap for a company with 95%+ retention, 80% gross margins, "
                        "and $796M free cash flow. CTV and retail media are secular tailwinds. CEO bought $148M in shares."),
        ("BEAR CASE",   "Walled gardens (Google, Meta, Amazon) continue to capture disproportionate ad budgets. "
                        "Publicis split signals fee disputes may spread to WPP/Dentsu. CFO departure adds execution "
                        "risk. Growth decelerating from 30%+ to 18%. High SBC ($495M) dilutes GAAP profitability."),
        ("FAIR VALUE",  "Base-case DCF at 10% WACC implies ~$50/share intrinsic value (140% upside). Bear-case at "
                        "10% revenue CAGR still suggests ~$33/share. Stock appears to price in near-worst-case. "
                        "Margin of safety appears significant at current levels for long-term investors."),
    ]

    for i, (t, body) in enumerate(thesis):
        r = thesis_start + 1 + i
        ws.row_dimensions[r].height = 60
        colors = [GREEN, RED, TTD_BLUE]
        bg = colors[i] if i < len(colors) else TTD_BLUE
        wc(ws, r, 2, t, bold=True, fg=WHITE, bg=bg, border=True, size=14)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
        wc(ws, r, 3, body, bold=False, fg="111111", bg=ALT_ROW if i % 2 == 0 else WHITE,
           border=True, wrap=True, size=13)

    # Footer
    r_foot = thesis_start + len(thesis) + 2
    ws.merge_cells(f"B{r_foot}:F{r_foot}")
    wc(ws, r_foot, 2,
       "Disclaimer: This analysis is for informational purposes only. Not investment advice. Data sourced from "
       "SEC filings, investor relations, and financial databases as of April 2026.",
       italic=True, fg="888888", size=11, wrap=True)


# ══════════════════════════════════════════════════════════════════════════════
# TAB 2 – BUSINESS OVERVIEW
# ══════════════════════════════════════════════════════════════════════════════
def build_business(wb):
    ws = wb.create_sheet("2. Business Overview")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, {1: 4, 2: 32, 3: 65})

    r = 2
    ws.merge_cells("B2:C2")
    wc(ws, r, 2, "THE TRADE DESK — BUSINESS OVERVIEW", bold=True, fg=HEADER_FG,
       bg=TTD_DARK, align="center", size=18)

    sections = [
        ("WHAT THE COMPANY DOES", [
            ("Core Business", "The Trade Desk is the world's largest independent demand-side platform (DSP). "
             "It enables advertisers and agencies to buy digital advertising programmatically across display, "
             "video, audio, Connected TV (CTV), mobile, digital out-of-home (DOOH), and native formats."),
            ("Independence", "Unlike Google DV360 or Amazon DSP, TTD does NOT own media properties — it is "
             "purely buy-side, meaning it is not conflicted by publisher interests. This neutrality is a core value proposition."),
            ("Platform: Kokai", "The Kokai platform (2023+) is TTD's AI-powered DSP successor to its legacy platform. "
             "It uses AI (Koa™) to optimize bids in real time. By Q3 2025, ~75% of client spend was on Kokai, "
             "with full adoption expected by end of 2025."),
            ("Revenue Model", "TTD charges a take-rate (fee) on the total media spend managed through its platform. "
             "In FY2025, revenue of $2.90B represents ~21% of total managed ad spend on the platform. "
             "Customers pay separately for data, inventory, and technology fees."),
        ]),
        ("PRODUCTS & CHANNELS", [
            ("Connected TV (CTV)", "The fastest-growing segment. TTD connects buyers to streaming inventory on "
             "Hulu, Peacock, Paramount+, Disney+, and others. CTV is structurally moving from linear TV, "
             "and TTD is positioned as the dominant buy-side gateway to this shift."),
            ("Retail Media", "Emerging high-margin business unit. TTD partners with retailers (Walmart, Target, "
             "Kroger) to allow brands to advertise against shoppers' purchase intent data. Retail Media segment "
             "created as dedicated business unit in December 2025 reorganization."),
            ("Digital Audio", "Programmatic audio advertising across Spotify, podcasts, and streaming music. "
             "Programmatic audio market expected to grow from $1.78B (2024) to $2.26B (2025)."),
            ("DOOH", "Digital-Out-of-Home advertising (billboards, transit, retail screens) — growing segment "
             "benefiting from real-time programmatic bidding."),
            ("Display & Video", "Core business. Programmatic banner, rich media, and video ads across the open web."),
            ("OpenPath", "Direct publisher connection initiative launched in 2022. Bypasses exchanges to connect "
             "directly with major publishers, improving margins for both sides. Freestar saw 3x fill rate +27% revenue."),
        ]),
        ("REVENUE BREAKDOWN (Estimated)", [
            ("By Channel (2025E)", "CTV: ~45% | Display/Video: ~30% | Audio: ~8% | Mobile: ~10% | DOOH: ~4% | Other: ~3%"),
            ("By Client Type", "Agency Holding Companies: ~30% | Independent Agencies: ~35% | Direct Brands: ~35%"),
            ("By Geography", "North America: ~85% | International: ~15% (growing focus on APAC and Europe)"),
            ("Seasonality", "Q4 is strongest (holiday ad spend); Q1 is weakest. Ad spending follows typical media industry seasonality."),
        ]),
        ("KEY CLIENTS & VALUE PROPOSITION", [
            ("Client Profile", "Fortune 500 brands and large agencies including Omnicom, IPG, WPP, and independents. "
             "Top 10 clients represent a modest share of revenue — client base is broad."),
            ("Value Proposition", "1) Transparency: full auction log reporting; 2) Independence: no media ownership conflicts; "
             "3) Scale: access to 1T+ daily bid requests; 4) Data: UID2 identity solution; "
             "5) AI: Kokai optimization engine; 6) Cross-channel: single platform for all digital media."),
            ("Buying Process", "Self-service SaaS model. Agency traders or in-house programmatic teams use the platform to: "
             "1) Define audience segments using UID2/first-party data; 2) Set campaign parameters and bid strategies; "
             "3) Buy inventory via RTB auction; 4) Measure attribution and optimize."),
            ("Customer Retention", "95%+ retention rate for 11 consecutive years — among the highest in enterprise SaaS. "
             "Speaks to high switching costs once agency workflows are built on the platform."),
        ]),
        ("COMPETITIVE LANDSCAPE", [
            ("Google DV360", "Largest DSP with ~41% market share. Integrated with Google's vast data and inventory. "
             "Conflict of interest as Google also operates the sell-side (GAM/AdX)."),
            ("Amazon DSP", "Growing fast (~20% share). Unique retail data is powerful for CPG/e-commerce advertisers. "
             "Major threat to TTD in retail-adjacent categories."),
            ("The Trade Desk", "~20% DSP market share. Wins on independence, transparency, and open internet breadth. "
             "Gartner rated 4.6 stars vs Amazon DSP's 4.4. Only pure-play independent."),
            ("Smaller DSPs", "MediaMath (bankrupt 2023), Xandr (AT&T→Microsoft), Yahoo DSP — all losing share."),
        ]),
        ("IDENTITY & DATA STRATEGY", [
            ("UID2", "Unified ID 2.0 — TTD's open-source alternative to third-party cookies. Uses hashed emails "
             "as durable identifiers. Over 200 partners (publishers, data companies) have adopted UID2. "
             "This positions TTD to thrive in a post-cookie world."),
            ("OpenSincera", "Announced May 2025. API providing metadata on advertising quality and supply chain health. "
             "Moves TTD into data transparency as a service — extends monetization surface."),
            ("Audience Unlimited", "AI-powered audience scoring tool launched 2025 — scores segments from 100s of "
             "privacy-safe third-party data providers. Full rollout expected early 2026."),
        ]),
    ]

    for sec_title, items in sections:
        r += 1
        ws.row_dimensions[r].height = 26
        ws.merge_cells(f"B{r}:C{r}")
        wc(ws, r, 2, sec_title, bold=True, fg=HEADER_FG, bg=TTD_MID, align="left", size=15)
        r += 1
        for i, (label, desc) in enumerate(items):
            ws.row_dimensions[r].height = 55
            alt = i % 2 == 0
            bg = ALT_ROW if alt else WHITE
            wc(ws, r, 2, label, bold=True, fg=TTD_DARK, bg=bg, border=True, wrap=True)
            wc(ws, r, 3, desc, bold=False, fg="222222", bg=bg, border=True, wrap=True)
            r += 1
        r += 1


# ══════════════════════════════════════════════════════════════════════════════
# TAB 3 – MOAT ANALYSIS
# ══════════════════════════════════════════════════════════════════════════════
def build_moat(wb):
    ws = wb.create_sheet("3. Moat Analysis")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, {1: 4, 2: 30, 3: 16, 4: 60})

    r = 2
    ws.merge_cells("B2:D2")
    wc(ws, r, 2, "COMPETITIVE MOAT ANALYSIS — THE TRADE DESK", bold=True, fg=HEADER_FG,
       bg=TTD_DARK, align="center", size=18)

    r += 1
    ws.merge_cells("B3:D3")
    wc(ws, r, 2, "Morningstar revised moat to 'None' in 2025-26 due to agency fee disputes — we disagree and analyze below",
       italic=True, fg=RED, size=13, wrap=True)

    moat_sections = [
        ("MOAT SOURCES", [
            ("Switching Costs", "STRONG", "Once an agency builds workflows, integrations, data taxonomies, and "
             "reporting around TTD's platform, switching to a competitor requires: retraining teams, rebuilding "
             "audience segments, renegotiating publisher deals, and accepting performance regression during migration. "
             "This is WHY 95%+ retention holds even during disputes."),
            ("Network Effects", "MODERATE", "TTD's data network improves with scale: more bidding data → better Kokai "
             "AI predictions → better ROI for advertisers → more spend routed to TTD. UID2 adoption creates ecosystem "
             "lock-in: more publishers adopt UID2 because buyers prefer it; more buyers use it because publishers support it."),
            ("Scale Advantage", "STRONG", "TTD processes 1T+ bid requests daily — scale that creates data advantages "
             "and negotiating leverage with publishers. OpenPath gives preferential inventory access. Smaller DSPs "
             "cannot replicate this at the same cost structure."),
            ("Kokai / AI IP", "MODERATE", "Proprietary AI models trained on years of bidding data. Koa™ AI makes "
             "real-time optimizations across 1T+ daily signals. Difficult to replicate overnight, though AI capabilities "
             "are commoditizing industry-wide."),
            ("UID2 Ecosystem", "EMERGING", "As cookie-less identity becomes critical, TTD's UID2 standard — now "
             "supported by 200+ publishers — creates a platform dependency. Advertisers using UID2 audience segments "
             "are tied to TTD's ecosystem."),
            ("Brand / Trust", "MODERATE", "TTD's brand = transparency and independence. This is increasingly valuable "
             "as advertiser trust in walled gardens erodes. However, the Publicis audit dispute has damaged this narrative "
             "in the short term."),
        ]),
        ("MOAT RISKS & CHALLENGES", [
            ("Agency Fee Disputes", "HIGH RISK", "Publicis Groupe (representing ~30% of agency spend on TTD) ended "
             "its preferred DSP recommendation after a third-party audit found fee irregularities. WPP and Dentsu have "
             "also raised concerns. This is the most acute near-term risk."),
            ("Walled Garden Growth", "HIGH RISK", "Meta, Google, and Amazon continue to grow faster than the open web. "
             "Advertisers who move budgets to these closed ecosystems bypass TTD entirely. TAM accessible to TTD "
             "may shrink in relative terms."),
            ("DSP Commoditization", "MEDIUM RISK", "The core DSP function (bid management, targeting, frequency capping) "
             "is increasingly seen as infrastructure. Margin compression may emerge as agencies in-house more capabilities."),
            ("Google's Dual Role", "MEDIUM RISK", "Google's dominance of both buy-side (DV360) and sell-side (GAM/AdX) "
             "creates structural disadvantage for independent DSPs. Regulatory actions (DOJ antitrust case) could help or hurt."),
            ("CFO Departure", "MEDIUM RISK", "CFO Blake Grayson's departure in 2025 adds uncertainty. Strong CFO "
             "leadership is critical for capital allocation discipline at this stage of growth."),
        ]),
        ("MOAT VERDICT", [
            ("Overall Rating", "NARROW MOAT", "TTD has genuine switching costs and scale advantages, but the Publicis "
             "dispute reveals that the moat is narrower than previously believed. Independence is a real differentiator, "
             "but not an insurmountable one. We rate it 'Narrow' (vs Morningstar's 'None')."),
            ("Moat Durability", "5-10 YEARS", "The CTV and identity (UID2) transitions offer a 5-10 year window "
             "of advantage. Beyond that, commoditization risk is real. Winners will be those who own the data layer."),
            ("Key Watch Variable", "AGENCY RETENTION", "Monitor Q1 2026 revenue guidance: if revenue growth holds "
             "above 15% despite Publicis exit, moat is intact. Below 12% growth would validate moat deterioration thesis."),
        ]),
    ]

    for sec_title, items in moat_sections:
        r += 1
        ws.row_dimensions[r].height = 26
        ws.merge_cells(f"B{r}:D{r}")
        wc(ws, r, 2, sec_title, bold=True, fg=HEADER_FG, bg=TTD_MID, align="left", size=15)
        r += 1

        if sec_title == "MOAT SOURCES":
            header_row(ws, r, [2, 3, 4], ["Moat Source", "Strength", "Analysis"])
            r += 1
            for i, (src, strength, analysis) in enumerate(items):
                ws.row_dimensions[r].height = 60
                alt = i % 2 == 0
                bg = ALT_ROW if alt else WHITE
                strength_color = GREEN if "STRONG" in strength else (ORANGE if "MODERATE" in strength else TTD_BLUE)
                wc(ws, r, 2, src, bold=True, fg=TTD_DARK, bg=bg, border=True, wrap=True)
                wc(ws, r, 3, strength, bold=True, fg=WHITE, bg=strength_color, align="center", border=True)
                wc(ws, r, 4, analysis, bold=False, fg="222222", bg=bg, border=True, wrap=True)
                r += 1
        elif sec_title == "MOAT RISKS & CHALLENGES":
            header_row(ws, r, [2, 3, 4], ["Risk Factor", "Severity", "Detail"])
            r += 1
            for i, (src, strength, analysis) in enumerate(items):
                ws.row_dimensions[r].height = 60
                alt = i % 2 == 0
                bg = ALT_ROW if alt else WHITE
                risk_color = RED if "HIGH" in strength else ORANGE
                wc(ws, r, 2, src, bold=True, fg=TTD_DARK, bg=bg, border=True, wrap=True)
                wc(ws, r, 3, strength, bold=True, fg=WHITE, bg=risk_color, align="center", border=True)
                wc(ws, r, 4, analysis, bold=False, fg="222222", bg=bg, border=True, wrap=True)
                r += 1
        else:
            header_row(ws, r, [2, 3, 4], ["Topic", "Verdict", "Rationale"])
            r += 1
            for i, (src, strength, analysis) in enumerate(items):
                ws.row_dimensions[r].height = 55
                alt = i % 2 == 0
                bg = ALT_ROW if alt else WHITE
                wc(ws, r, 2, src, bold=True, fg=TTD_DARK, bg=bg, border=True, wrap=True)
                wc(ws, r, 3, strength, bold=True, fg=WHITE, bg=TTD_DARK, align="center", border=True)
                wc(ws, r, 4, analysis, bold=False, fg="222222", bg=bg, border=True, wrap=True)
                r += 1
        r += 1


# ══════════════════════════════════════════════════════════════════════════════
# TAB 4 – INCOME STATEMENT
# ══════════════════════════════════════════════════════════════════════════════
def build_income(wb):
    ws = wb.create_sheet("4. Income Statement")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, {1: 4, 2: 36, 3: 16, 4: 16, 5: 16, 6: 16, 7: 16, 8: 20})
    freeze(ws, "C4")

    r = 2
    ws.merge_cells("B2:H2")
    wc(ws, r, 2, "INCOME STATEMENT — THE TRADE DESK (USD Millions)", bold=True, fg=HEADER_FG,
       bg=TTD_DARK, align="center", size=18)

    r = 3
    header_row(ws, r, [2, 3, 4, 5, 6, 7, 8],
               ["Metric", "FY 2021", "FY 2022", "FY 2023", "FY 2024", "FY 2025", "CAGR 21-25"])

    income_data = [
        # (label, values for 2021-2025, bold, num_fmt, cagr)
        ("REVENUE", [1196, 1578, 1946, 2445, 2896], True, '#,##0', None),
        ("  Revenue Growth YoY", [None, 0.319, 0.233, 0.256, 0.185], False, '0.0%', None),
        ("", [], False, None, None),
        ("Gross Profit", [975, 1297, 1581, 1973, 2277], False, '#,##0', None),
        ("  Gross Margin %", [0.815, 0.822, 0.812, 0.807, 0.786], False, '0.0%', None),
        ("", [], False, None, None),
        ("Operating Expenses (ex-COGS)", [], False, None, None),
        ("  Platform Operations (COGS)", [221, 281, 365, 472, 619], False, '#,##0', None),
        ("  Technology & Development", [323, 432, 556, 611, 680], False, '#,##0', None),
        ("  Sales & Marketing", [269, 359, 411, 498, 556], False, '#,##0', None),
        ("  General & Administrative", [158, 191, 213, 256, 302], False, '#,##0', None),
        ("  Total OpEx (incl. COGS)", [971, 1464, 1746, 2018, 2307], False, '#,##0', None),
        ("", [], False, None, None),
        ("OPERATING INCOME (GAAP)", [125, 114, 200, 427, 589], True, '#,##0', None),
        ("  Operating Margin %", [0.104, 0.072, 0.103, 0.175, 0.204], False, '0.0%', None),
        ("", [], False, None, None),
        ("Interest & Other Income", [19, 22, 38, 64, 50], False, '#,##0', None),
        ("Pre-tax Income", [144, 136, 238, 491, 639], False, '#,##0', None),
        ("Income Tax", [6, 83, 59, 98, 196], False, '#,##0', None),
        ("", [], False, None, None),
        ("NET INCOME (GAAP)", [138, 53, 179, 393, 443], True, '#,##0', None),
        ("  Net Margin %", [0.115, 0.034, 0.092, 0.161, 0.153], False, '0.0%', None),
        ("  EPS (Diluted)", [0.28, 0.11, 0.36, 0.78, 0.90], False, '$0.00', None),
        ("", [], False, None, None),
        ("EBITDA (GAAP)", [167, 168, 281, 515, 705], True, '#,##0', None),
        ("  EBITDA Margin %", [0.140, 0.107, 0.144, 0.211, 0.243], False, '0.0%', None),
        ("", [], False, None, None),
        ("Adjusted EBITDA (Non-GAAP)", [None, None, None, 939, 1200], True, '#,##0', None),
        ("  Adj. EBITDA Margin %", [None, None, None, 0.384, 0.414], False, '0.0%', None),
        ("  Stock-Based Compensation", [None, None, None, 424, 495], False, '#,##0', None),
        ("  SBC as % of Revenue", [None, None, None, 0.173, 0.171], False, '0.0%', None),
    ]

    # Calculate CAGRs for revenue, net income, etc.
    cagr_map = {
        "REVENUE": (1196, 2896, 4),
        "NET INCOME (GAAP)": (138, 443, 4),
        "EBITDA (GAAP)": (167, 705, 4),
        "Gross Profit": (975, 2277, 4),
    }

    for i, (label, values, bold, num_fmt, _) in enumerate(income_data):
        r += 1
        ws.row_dimensions[r].height = 22
        alt = i % 2 == 0
        bg = ALT_ROW if alt else WHITE

        if not label:
            for c in range(2, 9):
                ws.cell(row=r, column=c).value = None
            continue

        is_section = label and not label.startswith("  ") and bold
        lbg = SUBHDR_BG if is_section else bg

        wc(ws, r, 2, label, bold=bold, fg=TTD_DARK if is_section else "222222",
           bg=lbg, border=True, wrap=False)

        for j, v in enumerate(values):
            cell = ws.cell(row=r, column=3 + j, value=v)
            cell.font = mf(bold=bold, color="000000")
            cell.fill = mfill(lbg if is_section else bg)
            cell.alignment = cal("right")
            cell.border = mborder()
            if num_fmt and v is not None:
                cell.number_format = num_fmt

        # Fill missing columns
        for j in range(len(values), 5):
            cell = ws.cell(row=r, column=3 + j)
            cell.fill = mfill(lbg if is_section else bg)
            cell.border = mborder()

        # CAGR column
        cagr_col = ws.cell(row=r, column=8)
        cagr_col.fill = mfill(ALT_ROW)
        cagr_col.border = mborder()
        if label in cagr_map:
            v0, v1, n = cagr_map[label]
            cagr = (v1 / v0) ** (1 / n) - 1
            cagr_col.value = cagr
            cagr_col.number_format = "0.0%"
            cagr_col.font = mf(bold=True, color=GREEN if cagr > 0.15 else "000000")
            cagr_col.alignment = cal("right")

    # Rule of 50 note
    r += 2
    ws.merge_cells(f"B{r}:H{r}")
    wc(ws, r, 2,
       "RULE OF 50 (SaaS Benchmark): Revenue Growth % + FCF Margin % ≥ 50 is considered elite-tier SaaS "
       "performance. TTD scores: 2023: 51.7 | 2024: 51.8 | 2025: 46.0 — consistently near/above threshold.",
       bold=True, fg=WHITE, bg=TTD_MID, wrap=True, size=14)

    r += 1
    rule50_data = [
        ("Year", "FY 2022", "FY 2023", "FY 2024", "FY 2025", ""),
        ("Revenue Growth", "31.9%", "23.3%", "25.6%", "18.5%", ""),
        ("FCF Margin", "29.5%", "28.4%", "26.2%", "27.5%", ""),
        ("Rule of 50 Score", "61.4", "51.7", "51.8", "46.0", "⚠ Below 50 in 2025"),
        ("Status", "✓ Excellent", "✓ Passes", "✓ Passes", "⚠ Near Miss", "Watch trend"),
    ]
    for row_data in rule50_data:
        r += 1
        ws.row_dimensions[r].height = 22
        for j, val in enumerate(row_data[:6]):
            bg = SUBHDR_BG if j == 0 else (ALT_ROW if r % 2 == 0 else WHITE)
            bold = j == 0
            wc(ws, r, 2 + j, val, bold=bold, fg=TTD_DARK, bg=bg, border=True, align="center")


# ══════════════════════════════════════════════════════════════════════════════
# TAB 5 – BALANCE SHEET
# ══════════════════════════════════════════════════════════════════════════════
def build_balance(wb):
    ws = wb.create_sheet("5. Balance Sheet")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, {1: 4, 2: 38, 3: 16, 4: 16, 5: 16, 6: 16, 7: 16})
    freeze(ws, "C4")

    r = 2
    ws.merge_cells("B2:G2")
    wc(ws, r, 2, "BALANCE SHEET — THE TRADE DESK (USD Millions)", bold=True, fg=HEADER_FG,
       bg=TTD_DARK, align="center", size=18)

    r = 3
    header_row(ws, r, [2, 3, 4, 5, 6, 7],
               ["Metric", "FY 2021", "FY 2022", "FY 2023", "FY 2024", "FY 2025"])

    bs_data = [
        # ASSETS
        ("ASSETS", [], True, None),
        ("Cash & Cash Equivalents", [754, 1031, 895, 1369, 658], False, '#,##0'),
        ("Investments (Short-term)", [236, 185, 335, 396, 452], False, '#,##0'),
        ("Total Liquid Assets", [990, 1216, 1230, 1765, 1110], True, '#,##0'),
        ("Accounts Receivable", [1231, 1556, 1853, 2161, 2668], False, '#,##0'),
        ("Other Current Assets", [85, 102, 126, 148, 178], False, '#,##0'),
        ("Total Current Assets", [2306, 2874, 3209, 4074, 3956], True, '#,##0'),
        ("", [], False, None),
        ("Property & Equipment (net)", [87, 116, 148, 182, 268], False, '#,##0'),
        ("Operating Lease Right-of-Use", [212, 235, 285, 325, 400], False, '#,##0'),
        ("Intangibles & Other", [152, 156, 147, 131, 329], False, '#,##0'),
        ("Goodwill", [820, 1000, 1100, 1400, 1200], False, '#,##0'),
        ("TOTAL ASSETS", [3577, 4381, 4889, 6112, 6153], True, '#,##0'),
        ("", [], False, None),
        # LIABILITIES
        ("LIABILITIES", [], True, None),
        ("Accounts Payable (Customer Funds)", [1678, 1938, 2279, 2596, 3100], False, '#,##0'),
        ("Accrued Liabilities", [122, 147, 178, 198, 234], False, '#,##0'),
        ("Operating Lease Liabilities (Current)", [58, 68, 77, 92, 108], False, '#,##0'),
        ("Total Current Liabilities", [1858, 2153, 2534, 2886, 3442], True, '#,##0'),
        ("", [], False, None),
        ("Long-Term Debt", [0, 0, 0, 0, 0], False, '#,##0'),
        ("Finance & Operating Leases (LT)", [285, 261, 236, 312, 436], False, '#,##0'),
        ("Other Long-Term Liabilities", [55, 50, 45, 52, 59], False, '#,##0'),
        ("TOTAL LIABILITIES", [2050, 2265, 2724, 3163, 3669], True, '#,##0'),
        ("", [], False, None),
        # EQUITY
        ("SHAREHOLDERS' EQUITY", [], True, None),
        ("Common Stock & APIC", [1456, 1805, 2108, 2491, 2904], False, '#,##0'),
        ("Retained Earnings (Deficit)", [163, 218, 298, 621, 872], False, '#,##0'),
        ("Treasury Stock / Buybacks", [-92, -308, -568, -788, -1292], False, '#,##0'),
        ("TOTAL EQUITY", [1527, 2115, 2164, 2949, 2484], True, '#,##0'),
        ("", [], False, None),
        # KEY RATIOS
        ("KEY BALANCE SHEET RATIOS", [], True, None),
        ("Net Cash (Cash - All Debt)", [469, 770, 659, 1057, 222], False, '#,##0'),
        ("Current Ratio", [1.24, 1.33, 1.27, 1.41, 1.15], False, '0.00'),
        ("Debt-to-Equity Ratio", [0.19, 0.12, 0.11, 0.11, 0.18], False, '0.00'),
        ("Tangible Book Value / Share ($)", [1.50, 2.35, 2.26, 3.27, 2.74], False, '$0.00'),
        ("Total Liabilities / Total Assets", [0.57, 0.52, 0.56, 0.52, 0.60], False, '0.0%'),
    ]

    for i, row_item in enumerate(bs_data):
        r += 1
        ws.row_dimensions[r].height = 22
        label, values, bold, num_fmt = row_item
        alt = i % 2 == 0

        if not label:
            continue

        is_header = not label.startswith(" ") and bold and not values
        is_total = bold and values
        bg = SUBHDR_BG if is_header else (ALT_ROW if alt else WHITE)

        wc(ws, r, 2, label, bold=bold, fg=TTD_DARK if (is_header or is_total) else "333333",
           bg=bg, border=True)

        for j, v in enumerate(values):
            cell = ws.cell(row=r, column=3 + j, value=v)
            cell.font = mf(bold=bold, color="000000")
            cell.fill = mfill(bg)
            cell.alignment = cal("right")
            cell.border = mborder()
            if num_fmt and v is not None:
                cell.number_format = num_fmt

        for j in range(len(values), 5):
            cell = ws.cell(row=r, column=3 + j)
            cell.fill = mfill(bg)
            cell.border = mborder()

    # Notes
    r += 2
    ws.merge_cells(f"B{r}:G{r}")
    wc(ws, r, 2,
       "NOTES: The Trade Desk has ZERO long-term financial debt — a major strength. 'Debt' shown above is "
       "operating/finance leases and right-of-use obligations. Accounts Payable is largely customer funds held "
       "in transit (not traditional AP — these represent media spend not yet disbursed to publishers). "
       "Net cash fell from $1.06B (2024) to $222M (2025) due to $1.4B in share repurchases.",
       bold=False, fg=TTD_DARK, bg=ALT_ROW, wrap=True, size=13)


# ══════════════════════════════════════════════════════════════════════════════
# TAB 6 – CASH FLOW ANALYSIS
# ══════════════════════════════════════════════════════════════════════════════
def build_cashflow(wb):
    ws = wb.create_sheet("6. Cash Flow Analysis")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, {1: 4, 2: 38, 3: 16, 4: 16, 5: 16, 6: 16, 7: 16})
    freeze(ws, "C4")

    r = 2
    ws.merge_cells("B2:G2")
    wc(ws, r, 2, "CASH FLOW ANALYSIS — THE TRADE DESK (USD Millions)", bold=True, fg=HEADER_FG,
       bg=TTD_DARK, align="center", size=18)

    r = 3
    header_row(ws, r, [2, 3, 4, 5, 6, 7],
               ["Metric", "FY 2021", "FY 2022", "FY 2023", "FY 2024", "FY 2025"])

    cf_data = [
        ("OPERATING CASH FLOWS", [], True, None),
        ("Net Income", [138, 53, 179, 393, 443], False, '#,##0'),
        ("  + Depreciation & Amortization", [42, 54, 81, 88, 116], False, '#,##0'),
        ("  + Stock-Based Compensation", [165, 262, 336, 424, 495], False, '#,##0'),
        ("  +/- Working Capital Changes", [34, 180, 2, -166, -61], False, '#,##0'),
        ("OPERATING CASH FLOW (OCF)", [379, 549, 598, 739, 993], True, '#,##0'),
        ("  OCF Growth YoY", [None, 0.449, 0.089, 0.236, 0.343], False, '0.0%'),
        ("  OCF Margin", [0.317, 0.348, 0.307, 0.302, 0.343], False, '0.0%'),
        ("", [], False, None),
        ("INVESTING CASH FLOWS", [], True, None),
        ("  Capital Expenditures (CapEx)", [-55, -84, -47, -98, -197], False, '#,##0'),
        ("  Capitalized Software Development", [-8, -12, -15, -18, -22], False, '#,##0'),
        ("  Acquisitions / Investments", [-31, -208, -46, -42, -74], False, '#,##0'),
        ("INVESTING CASH FLOW", [-94, -304, -108, -158, -293], True, '#,##0'),
        ("", [], False, None),
        ("FINANCING CASH FLOWS", [], True, None),
        ("  Share Repurchases", [0, -24, -673, -148, -1411], False, '#,##0'),
        ("  Stock Option Exercises", [32, 56, 47, 40, None], False, '#,##0'),
        ("FINANCING CASH FLOW", [32, 32, -626, -108, -1411], True, '#,##0'),
        ("", [], False, None),
        ("NET CHANGE IN CASH", [317, 277, -136, 473, -711], True, '#,##0'),
        ("", [], False, None),
        ("FREE CASH FLOW ANALYSIS", [], True, None),
        ("FREE CASH FLOW (OCF - CapEx)", [324, 465, 551, 641, 796], True, '#,##0'),
        ("  FCF Growth YoY", [None, 0.435, 0.185, 0.163, 0.241], False, '0.0%'),
        ("  FCF Margin", [0.271, 0.295, 0.283, 0.262, 0.275], False, '0.0%'),
        ("  FCF Conversion (FCF / Net Inc)", [2.35, 8.77, 3.08, 1.63, 1.80], False, '0.00x'),
        ("  CapEx as % of Revenue", [0.046, 0.053, 0.024, 0.040, 0.068], False, '0.0%'),
        ("", [], False, None),
        ("CASH QUALITY & OWNER EARNINGS", [], True, None),
        ("  Stock-Based Comp (SBC) Dilution", [165, 262, 336, 424, 495], False, '#,##0'),
        ("  SBC as % of Revenue", [0.138, 0.166, 0.173, 0.173, 0.171], False, '0.0%'),
        ("  Owner Earnings (FCF - SBC)", [159, 203, 215, 217, 301], True, '#,##0'),
        ("  Owner Earnings Margin", [0.133, 0.129, 0.110, 0.089, 0.104], False, '0.0%'),
        ("", [], False, None),
        ("BUYBACK PROGRAM", [], True, None),
        ("  Shares Repurchased ($M)", [0, 24, 673, 148, 1411], False, '#,##0'),
        ("  Cumulative Buybacks ($M)", [0, 24, 697, 845, 2256], False, '#,##0'),
        ("  Shares Outstanding (M)", [495, 490, 487, 493, 470], False, '#,##0'),
    ]

    for i, row_item in enumerate(cf_data):
        r += 1
        ws.row_dimensions[r].height = 22
        label, values, bold, num_fmt = row_item
        alt = i % 2 == 0

        if not label:
            continue

        is_header = bold and not values
        is_total = bold and values
        bg = SUBHDR_BG if is_header else (ALT_ROW if alt else WHITE)
        val_color = GREEN if is_total and any((v or 0) > 0 for v in values) else "000000"

        wc(ws, r, 2, label, bold=bold, fg=TTD_DARK if (is_header or is_total) else "333333",
           bg=bg, border=True)

        for j, v in enumerate(values):
            cell = ws.cell(row=r, column=3 + j, value=v)
            cell.font = mf(bold=bold, color=val_color if is_total else "000000")
            cell.fill = mfill(bg)
            cell.alignment = cal("right")
            cell.border = mborder()
            if num_fmt and v is not None:
                cell.number_format = num_fmt

        for j in range(len(values), 5):
            cell = ws.cell(row=r, column=3 + j)
            cell.fill = mfill(bg)
            cell.border = mborder()

    # Note on SBC
    r += 2
    ws.merge_cells(f"B{r}:G{r}")
    wc(ws, r, 2,
       "OWNER EARNINGS NOTE: TTD generates exceptional GAAP FCF ($796M in 2025), but ~$495M in stock-based "
       "compensation (17% of revenue) is a real economic cost. 'Owner Earnings' = FCF minus SBC is the more "
       "conservative measure of true cash generation ($301M in 2025 = 10.4% owner earnings margin). "
       "The $1.4B 2025 buyback was funded partly by debt-like lease obligations and existing cash — watch leverage.",
       bold=False, fg=TTD_DARK, bg=ALT_ROW, wrap=True, size=13)


# ══════════════════════════════════════════════════════════════════════════════
# TAB 7 – RETURN ON CAPITAL
# ══════════════════════════════════════════════════════════════════════════════
def build_roic(wb):
    ws = wb.create_sheet("7. Return on Capital")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, {1: 4, 2: 40, 3: 16, 4: 16, 5: 16, 6: 16, 7: 16})
    freeze(ws, "C4")

    r = 2
    ws.merge_cells("B2:G2")
    wc(ws, r, 2, "RETURN ON CAPITAL — THE TRADE DESK", bold=True, fg=HEADER_FG,
       bg=TTD_DARK, align="center", size=18)

    r = 3
    header_row(ws, r, [2, 3, 4, 5, 6, 7],
               ["Metric", "FY 2021", "FY 2022", "FY 2023", "FY 2024", "FY 2025"])

    roic_data = [
        ("PROFITABILITY RETURNS", [], True, None),
        ("Return on Equity (ROE)", [0.090, 0.025, 0.083, 0.133, 0.178], False, '0.0%'),
        ("Return on Assets (ROA)", [0.039, 0.012, 0.037, 0.064, 0.072], False, '0.0%'),
        ("Return on Invested Capital (ROIC)", [0.087, 0.035, 0.083, 0.192, 0.251], False, '0.0%'),
        ("WACC (Estimated)", [0.090, 0.095, 0.100, 0.100, 0.100], False, '0.0%'),
        ("Economic Spread (ROIC - WACC)", [-0.003, -0.060, -0.017, 0.092, 0.151], True, '0.0%'),
        ("", [], False, None),
        ("INCREMENTAL CAPITAL RETURNS", [], True, None),
        ("Revenue per Dollar of Assets", [0.335, 0.360, 0.398, 0.400, 0.471], False, '$0.000'),
        ("FCF per Diluted Share ($)", [0.65, 0.95, 1.13, 1.30, 1.69], False, '$0.00'),
        ("FCF per Share Growth YoY", [None, 0.462, 0.189, 0.150, 0.300], False, '0.0%'),
        ("Revenue per Employee (est. $K)", [625, 680, 740, 810, 870], False, '#,##0'),
        ("", [], False, None),
        ("CAPITAL EFFICIENCY", [], True, None),
        ("Gross Profit / Total Assets", [0.273, 0.296, 0.323, 0.323, 0.370], False, '0.0%'),
        ("Asset Turnover", [0.335, 0.360, 0.398, 0.400, 0.471], False, '0.00x'),
        ("CapEx Intensity (CapEx / Revenue)", [0.046, 0.053, 0.024, 0.040, 0.068], False, '0.0%'),
        ("Maintenance CapEx (est.)", [25, 30, 25, 40, 60], False, '#,##0'),
        ("Growth CapEx (est.)", [30, 54, 22, 58, 137], False, '#,##0'),
        ("", [], False, None),
        ("INCREMENTAL ROIC", [], True, None),
        ("Δ Revenue (vs prior year)", [None, 382, 368, 499, 451], False, '#,##0'),
        ("Δ Invested Capital (vs prior year)", [None, 563, 424, 1212, 41], False, '#,##0'),
        ("Incremental ROIC (Δ Op. Income / Δ IC)", [None, -0.016, 0.204, 0.188, None], False, '0.0%'),
        ("", [], False, None),
        ("CAPITAL ALLOCATION SCORECARD", [], True, None),
        ("Organic R&D Investment", ["High", "High", "High", "High", "High"], False, None),
        ("Share Repurchases ($M)", [0, 24, 673, 148, 1411], False, '#,##0'),
        ("Acquisitions ($M)", [31, 208, 46, 42, 74], False, '#,##0'),
        ("Dividends Paid", ["None", "None", "None", "None", "None"], False, None),
        ("Debt Raised / Repaid ($M)", [0, 0, 0, 0, 0], False, '#,##0'),
    ]

    for i, row_item in enumerate(roic_data):
        r += 1
        ws.row_dimensions[r].height = 22
        label, values, bold, num_fmt = row_item
        alt = i % 2 == 0

        if not label:
            continue

        is_header = bold and not values
        is_total = bold and values
        bg = SUBHDR_BG if is_header else (ALT_ROW if alt else WHITE)

        wc(ws, r, 2, label, bold=bold, fg=TTD_DARK if (is_header or is_total) else "333333",
           bg=bg, border=True)

        for j, v in enumerate(values):
            cell = ws.cell(row=r, column=3 + j, value=v)
            if isinstance(v, float) and label == "Economic Spread (ROIC - WACC)":
                fg_color = GREEN if v > 0 else RED
                cell.font = mf(bold=True, color=fg_color)
            else:
                cell.font = mf(bold=bold)
            cell.fill = mfill(bg)
            cell.alignment = cal("right" if not isinstance(v, str) else "center")
            cell.border = mborder()
            if num_fmt and v is not None and not isinstance(v, str):
                cell.number_format = num_fmt

        for j in range(len(values), 5):
            cell = ws.cell(row=r, column=3 + j)
            cell.fill = mfill(bg)
            cell.border = mborder()

    # Commentary
    r += 2
    ws.merge_cells(f"B{r}:G{r}")
    wc(ws, r, 2,
       "ROIC COMMENTARY: TTD's ROIC of 25.1% (2025) substantially exceeds our estimated WACC of 10%, "
       "confirming it creates economic value. The $1.4B 2025 buyback at $21-40/share prices represents "
       "attractive capital allocation IF fair value is above $30+ (our base case). High SBC ($495M) reduces "
       "true owner earnings, but the core business generates significant returns on incremental investment.",
       bold=False, fg=TTD_DARK, bg=ALT_ROW, wrap=True, size=13)


# ══════════════════════════════════════════════════════════════════════════════
# TAB 8 – MANAGEMENT ANALYSIS
# ══════════════════════════════════════════════════════════════════════════════
def build_management(wb):
    ws = wb.create_sheet("8. Management")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, {1: 4, 2: 32, 3: 14, 4: 58})

    r = 2
    ws.merge_cells("B2:D2")
    wc(ws, r, 2, "MANAGEMENT ANALYSIS — THE TRADE DESK", bold=True, fg=HEADER_FG,
       bg=TTD_DARK, align="center", size=18)

    mgmt_sections = [
        ("KEY LEADERSHIP", [
            ("Jeff Green — CEO & Co-Founder", "Founder-led since 2009. Holds ~11.1% economic interest (Class B super-voting "
             "shares give ~49% voting control). Previously sold AdECN to Microsoft. Widely regarded as the most visionary "
             "leader in ad-tech. Known for bold long-term bets: UID2, Kokai, OpenPath. Communication style: highly transparent "
             "and intellectually aggressive in earnings calls — willing to take controversial public stances (e.g., "
             "versus Google, versus agency opacity)."),
            ("Dave Pickles — CTO & Co-Founder", "Co-founder and technical architect of the platform. Led all major "
             "platform evolutions including the Kokai AI transition. Deeply technical and long-tenured — key person risk."),
            ("CFO — Vacant (2025)", "Blake Grayson departed as CFO in late 2025 — a negative signal. No permanent "
             "replacement announced as of April 2026. This is an execution risk and governance concern."),
            ("Laura Schenkein — Chief Legal Officer", "Long-tenured legal leadership. Oversees SEC compliance, IP protection."),
        ]),
        ("CEO OWNER-OPERATOR ASSESSMENT", [
            ("Skin in the Game", "Jeff Green owns ~11.1% of shares ($1.1B at $21 stock price). In March 2026, he "
             "purchased 6 million shares for ~$148 million — the largest insider purchase in company history. "
             "This is a strong signal that he believes the stock is materially undervalued at current levels."),
            ("Long-term Orientation", "TTD has consistently invested in long-cycle bets: UID2 (identity), Kokai (AI), "
             "OpenPath (direct publisher access), CTV (streaming). These are 3-7 year initiatives, not short-term revenue "
             "maximizers. This is hallmark behavior of owner-operators planting seeds vs. borrowing from the future."),
            ("Capital Allocation Philosophy", "Priority order: 1) R&D and platform investment; 2) Share repurchases when "
             "undervalued; 3) Targeted acquisitions; 4) No dividends. The $1.4B repurchase at $20-40 demonstrates conviction "
             "at low prices. Total buybacks since 2021 = $2.26B, reducing diluted shares from ~495M to ~470M."),
            ("Transparency & Culture", "Green is unusually forthcoming in earnings calls — acknowledging mistakes, "
             "explaining strategy in detail. However, the Publicis fee dispute suggests some operational practices "
             "(undisclosed fees) were not aligned with the transparency brand promise."),
        ]),
        ("INCENTIVE STRUCTURE (PROXY ANALYSIS)", [
            ("CEO Base Salary (2025)", "$1,200,000 — modest relative to equity value; aligned with owner-operator model."),
            ("CEO Total Compensation", "Primarily equity-linked RSUs and options. FY2025: ~$38M total comp (incl. equity awards) "
             "per proxy. New March 2026 grant: 398,089 restricted shares + 737,028 stock options."),
            ("Performance Metrics (STI)", "Annual cash bonus tied to revenue growth, Adjusted EBITDA, and platform adoption "
             "metrics. 2025 metrics included Kokai client transition and CTV growth."),
            ("Long-Term Equity (LTI)", "RSU vesting over 3-4 years, options with 4-year cliff/vest. Structure aligns "
             "long-term with shareholders. But total SBC at $495M/year (17% of revenue) is dilutive."),
            ("Class B Voting Structure", "Jeff Green holds Class B shares with 10x voting power. Board proposed extending "
             "the Class B sunset from 2025 to 2035 — a controversial governance move that entrenches founder control. "
             "Reduces minority shareholder power."),
            ("Say-on-Pay", "Moved from triennial to annual starting 2026 — positive governance change."),
        ]),
        ("CAPITAL ALLOCATION DECISIONS", [
            ("Share Repurchases", "2025: $1.4B repurchased (~14% of market cap at time of purchase). At $21-35/share, "
             "this appears value-accretive IF fair value is above $30. CEO's personal $148M purchase in March 2026 "
             "at ~$25/share is a strong endorsement of the buyback decision."),
            ("R&D Investment", "Technology & development expense of $680M in 2025 (23.5% of revenue). Heavy investment "
             "in Kokai AI, UID2, and new product lines. This is the right spending priority for platform durability."),
            ("M&A Track Record", "Acquisitions have been small and targeted ($74M in 2025). No dilutive mega-mergers. "
             "Consistent with disciplined capital allocation philosophy."),
            ("Leverage Policy", "Zero long-term financial debt — unusual restraint in a low-rate environment. "
             "Management has preferred organic investment and buybacks over debt-funded growth."),
        ]),
        ("PLANTING SEEDS VS. BORROWING FROM THE FUTURE", [
            ("SEEDS PLANTED", "UID2 (future identity infrastructure), Kokai AI (platform differentiation), "
             "OpenSincera (supply chain transparency tool), Retail Media business unit, international expansion, "
             "audio/CTV/DOOH channel buildout. All require multi-year investment before full ROI."),
            ("POTENTIAL SHORT-TERMISM", "The $1.4B buyback consumed most of the company's cash position — "
             "net cash fell from $1.06B to $222M. This reduces financial flexibility for opportunistic M&A or "
             "recession defense. Also, fee structure concerns (Publicis audit) suggest some revenue may have "
             "been extracted at the expense of long-term agency relationships."),
            ("Overall Verdict", "TTD management is primarily 'seed planting.' Jeff Green consistently takes "
             "long-term bets and communicates them clearly. The CFO departure and Publicis dispute are the main "
             "blemishes. Overall: above-average management quality, with founder-owner alignment."),
        ]),
    ]

    for sec_title, items in mgmt_sections:
        r += 1
        ws.row_dimensions[r].height = 26
        ws.merge_cells(f"B{r}:D{r}")
        wc(ws, r, 2, sec_title, bold=True, fg=HEADER_FG, bg=TTD_MID, align="left", size=15)
        r += 1
        for i, (label, detail) in enumerate(items):
            ws.row_dimensions[r].height = 65
            alt = i % 2 == 0
            bg = ALT_ROW if alt else WHITE
            wc(ws, r, 2, label, bold=True, fg=TTD_DARK, bg=bg, border=True, wrap=True)
            wc(ws, r, 3, "→", bold=False, fg=TTD_BLUE, bg=bg, border=True, align="center")
            wc(ws, r, 4, detail, bold=False, fg="222222", bg=bg, border=True, wrap=True)
            r += 1
        r += 1


# ══════════════════════════════════════════════════════════════════════════════
# TAB 9 – RISKS
# ══════════════════════════════════════════════════════════════════════════════
def build_risks(wb):
    ws = wb.create_sheet("9. Risks")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, {1: 4, 2: 30, 3: 14, 4: 14, 5: 52})

    r = 2
    ws.merge_cells("B2:E2")
    wc(ws, r, 2, "RISK ANALYSIS — THE TRADE DESK (TTD)", bold=True, fg=HEADER_FG,
       bg=TTD_DARK, align="center", size=18)

    r = 3
    header_row(ws, r, [2, 3, 4, 5], ["Risk Factor", "Probability", "Impact", "Analysis & Mitigation"])

    risks = [
        # (category, risk, probability, impact, detail)
        ("COMPETITIVE", "Walled Garden Dominance", "HIGH", "HIGH",
         "Google, Meta, Amazon control ~70%+ of digital ad spend. Advertisers find it easier to buy within "
         "these closed ecosystems. TTD cannot access inventory on these platforms. "
         "MITIGATION: CTV/streaming is NOT walled — it's TTD's opportunity. Regulatory pressure (DOJ antitrust) "
         "may force inventory opening. Open internet projected to grow from ~$200B to $400B+ by 2030."),
        ("COMPETITIVE", "Amazon DSP Market Share Gains", "HIGH", "MEDIUM",
         "Amazon's retail data is uniquely powerful for CPG, grocery, and e-commerce advertisers. "
         "Amazon is actively investing in DSP capabilities and offering bundled pricing. "
         "MITIGATION: TTD's strength is cross-channel breadth; Amazon DSP is narrower. "
         "TTD is building its own retail media unit with non-Amazon retailer data."),
        ("CLIENT", "Publicis / Agency Fee Dispute", "HIGH", "HIGH",
         "Publicis Groupe (representing ~10-15% of TTD gross spend) stopped recommending TTD after a FirmDecisions audit "
         "found fee irregularities. WPP and Dentsu have raised similar concerns. "
         "MITIGATION: TTD disputed findings; some clients may stay direct. BUT if 30% of agency spend is at risk, "
         "this could reduce 2026 revenue growth to single digits. Q1/Q2 2026 results will be the key test."),
        ("CLIENT", "Customer Concentration / Agency Dependence", "MEDIUM", "HIGH",
         "Holding companies represent ~30% of gross spend. Loss of one major holding company relationship could "
         "reduce revenue by 5-10%. "
         "MITIGATION: TTD is actively growing direct brand relationships (35%+ of spend) and independent agencies."),
        ("EXECUTION", "CFO Departure & Governance Gap", "MEDIUM", "MEDIUM",
         "Blake Grayson's CFO departure in 2025 was unexpected and unexplained. No permanent replacement as of April 2026. "
         "CFO is critical for investor relations, capital allocation, and financial controls. "
         "MITIGATION: Jeff Green has strong operational control; existing finance team is experienced."),
        ("EXECUTION", "Kokai Transition Risk", "LOW", "MEDIUM",
         "The transition from legacy platform to Kokai encountered temporary client performance variability. "
         "~75% of spend on Kokai by Q3 2025 — near complete. Full migration by year-end 2025. "
         "MITIGATION: Transition substantially complete; early adopters showing strong results."),
        ("TECHNOLOGY", "AI / Identity Solution Commoditization", "MEDIUM", "MEDIUM",
         "Kokai's AI advantage may erode as Google, Amazon, and startups build comparable optimization layers. "
         "UID2 could become less relevant if alternative identity standards emerge or Google finds a new cookie replacement. "
         "MITIGATION: TTD has 15+ years of bidding data to train models. UID2 has 200+ publisher partners — ecosystem lock-in."),
        ("MACRO", "Digital Ad Spend Recession Sensitivity", "MEDIUM", "HIGH",
         "Digital advertising is cyclical. In a recession, brand/discovery advertising is cut first. "
         "CTV and performance advertising are more resilient than display. "
         "MITIGATION: TTD's channel mix has shifted toward more resilient CTV/performance formats. "
         "Managed programmatic spend fell ~15-20% in 2022 macro tightening; TTD still grew revenue 32%."),
        ("REGULATORY", "Privacy Regulations (GDPR, CCPA, AI Acts)", "MEDIUM", "LOW",
         "Cookie deprecation, data privacy rules, and AI regulation could restrict audience targeting capabilities. "
         "MITIGATION: UID2 is specifically designed as a privacy-safe alternative. TTD is well-positioned relative to "
         "competitors for the post-cookie world."),
        ("GOVERNANCE", "Founder Control (Class B Shares to 2035)", "MEDIUM", "LOW",
         "Extension of Jeff Green's super-voting rights to 2035 limits minority shareholder governance. "
         "This is a structural risk: CEO could resist board oversight, strategic pivots, or M&A. "
         "MITIGATION: Green's track record and personal $148M investment suggest alignment. But governance risk exists."),
        ("FINANCIAL", "High SBC Dilution ($495M/Year)", "HIGH", "MEDIUM",
         "Stock-based compensation of $495M (17% of revenue) creates significant dilution risk and "
         "inflates GAAP earnings quality. Owner earnings are far lower than GAAP FCF. "
         "MITIGATION: Ongoing buybacks offset some dilution ($1.4B in 2025). But net shares are barely declining."),
        ("FINANCIAL", "Net Cash Reduction from Buybacks", "MEDIUM", "LOW",
         "2025 buyback consumed cash — net cash fell from $1.06B to $222M. "
         "Reduces buffer for economic downturns or opportunistic M&A. "
         "MITIGATION: Strong ongoing FCF generation ($796M/year) rebuilds cash rapidly. True financial debt remains zero."),
    ]

    prob_colors = {"HIGH": RED, "MEDIUM": ORANGE, "LOW": GREEN}

    current_cat = None
    for i, (cat, risk, prob, impact, detail) in enumerate(risks):
        if cat != current_cat:
            r += 1
            ws.row_dimensions[r].height = 24
            ws.merge_cells(f"B{r}:E{r}")
            wc(ws, r, 2, f"── {cat} RISKS ──", bold=True, fg=HEADER_FG, bg=TTD_DARK,
               align="left", size=14)
            current_cat = cat

        r += 1
        ws.row_dimensions[r].height = 65
        alt = i % 2 == 0
        bg = ALT_ROW if alt else WHITE

        wc(ws, r, 2, risk, bold=True, fg=TTD_DARK, bg=bg, border=True, wrap=True)
        wc(ws, r, 3, prob, bold=True, fg=WHITE, bg=prob_colors.get(prob, ORANGE),
           align="center", border=True)
        wc(ws, r, 4, impact, bold=True, fg=WHITE, bg=prob_colors.get(impact, ORANGE),
           align="center", border=True)
        wc(ws, r, 5, detail, bold=False, fg="222222", bg=bg, border=True, wrap=True)


# ══════════════════════════════════════════════════════════════════════════════
# TAB 10 – VALUATION
# ══════════════════════════════════════════════════════════════════════════════
def build_valuation(wb):
    ws = wb.create_sheet("10. Valuation")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, {1: 4, 2: 36, 3: 18, 4: 18, 5: 18, 6: 18, 7: 18})

    r = 2
    ws.merge_cells("B2:G2")
    wc(ws, r, 2, "VALUATION ANALYSIS — THE TRADE DESK (TTD)", bold=True, fg=HEADER_FG,
       bg=TTD_DARK, align="center", size=18)

    # ── Section 1: Current Market Snapshot ──────────────────────────────────
    r += 1
    ws.merge_cells(f"B{r}:G{r}")
    wc(ws, r, 2, "SECTION 1: CURRENT MARKET SNAPSHOT (April 13, 2026)", bold=True, fg=HEADER_FG,
       bg=TTD_MID, size=15)

    mkt_data = [
        ("Stock Price", "$21.00", "52-Week High", "$91.45", "52-Week Low", "$19.74"),
        ("Shares Outstanding", "470.1M", "Market Capitalization", "$9.87B", "Short Interest", "14.2%"),
        ("Cash & Investments", "$1,110M", "Total Debt (leases)", "$436M", "Enterprise Value", "~$9.2B"),
        ("P/E Ratio (TTM)", "23.6x", "Forward P/E (FY26E)", "10.3x", "EV/Revenue", "3.2x"),
        ("EV/EBITDA (GAAP)", "13.0x", "EV/Adj. EBITDA", "7.7x", "EV/FCF", "11.6x"),
        ("Price/FCF", "12.4x", "Price/Sales", "3.4x", "Beta", "1.17"),
        ("52-Week Price Decline", "-77%", "Insider Ownership", "11.1%", "Institutional Own.", "76.3%"),
    ]

    r += 1
    for row_data in mkt_data:
        ws.row_dimensions[r].height = 22
        bg = ALT_ROW if r % 2 == 0 else WHITE
        for j in range(3):
            wc(ws, r, 2 + j*2, row_data[j*2], bold=True, fg=TTD_DARK, bg=bg, border=True)
            wc(ws, r, 3 + j*2, row_data[j*2+1], bold=False, fg="000000", bg=bg, border=True,
               align="right")
        r += 1

    # ── Section 2: Relative Valuation ──────────────────────────────────────
    r += 1
    ws.merge_cells(f"B{r}:G{r}")
    wc(ws, r, 2, "SECTION 2: RELATIVE VALUATION vs. PEERS", bold=True, fg=HEADER_FG,
       bg=TTD_MID, size=15)
    r += 1
    header_row(ws, r, [2, 3, 4, 5, 6, 7],
               ["Company", "Revenue ($B)", "Rev Growth", "EV/Revenue", "EV/EBITDA", "FCF Margin"])
    peers = [
        ("The Trade Desk (TTD)", 2.90, 0.185, 3.2, 7.7, 0.275),
        ("Alphabet / Google", 360.0, 0.145, 5.8, 18.2, 0.220),
        ("Meta Platforms", 165.0, 0.215, 8.3, 21.4, 0.350),
        ("Amazon (Ads)", 69.0, 0.190, None, None, None),
        ("Magnite (MGNI)", 0.84, 0.120, 2.1, 8.5, 0.185),
        ("PubMatic (PUBM)", 0.30, 0.115, 1.8, 7.2, 0.220),
        ("Integral Ad Science", 0.48, 0.080, 2.5, 9.1, 0.180),
        ("DoubleVerify (DV)", 0.63, 0.140, 3.4, 11.2, 0.215),
    ]
    r += 1
    for i, (co, rev, growth, ev_rev, ev_ebitda, fcf_margin) in enumerate(peers):
        ws.row_dimensions[r].height = 22
        alt = i % 2 == 0
        bg = ALT_ROW if alt else WHITE
        is_ttd = co.startswith("The Trade Desk")
        b = is_ttd
        fg = TTD_DARK if is_ttd else "333333"
        row_bg = SUBHDR_BG if is_ttd else bg

        wc(ws, r, 2, co, bold=b, fg=fg, bg=row_bg, border=True)
        wc(ws, r, 3, rev, bold=b, fg=fg, bg=row_bg, border=True, align="right", num_fmt='$#,##0.00')
        wc(ws, r, 4, growth, bold=b, fg=fg, bg=row_bg, border=True, align="right", num_fmt='0.0%')
        wc(ws, r, 5, ev_rev, bold=b, fg=fg, bg=row_bg, border=True, align="right", num_fmt='0.0x')
        wc(ws, r, 6, ev_ebitda, bold=b, fg=fg, bg=row_bg, border=True, align="right", num_fmt='0.0x')
        wc(ws, r, 7, fcf_margin, bold=b, fg=fg, bg=row_bg, border=True, align="right",
           num_fmt='0.0%' if fcf_margin is not None else None)
        r += 1

    # ── Section 3: DCF Valuation ────────────────────────────────────────────
    r += 1
    ws.merge_cells(f"B{r}:G{r}")
    wc(ws, r, 2, "SECTION 3: DCF VALUATION — 3-SCENARIO ANALYSIS", bold=True, fg=HEADER_FG,
       bg=TTD_MID, size=15)
    r += 1

    dcf_assumptions = [
        ("Discount Rate (WACC)", "10.0%", "10.0%", "10.0%", "", ""),
        ("Terminal Growth Rate", "2.5%", "3.0%", "4.0%", "", ""),
        ("Revenue CAGR (2026-2030)", "10%", "20%", "27%", "", ""),
        ("Terminal FCF Margin", "24%", "28%", "32%", "", ""),
    ]

    header_row(ws, r, [2, 3, 4, 5, 6, 7],
               ["Assumption", "Bear Case", "Base Case", "Bull Case", "", ""])
    r += 1
    for label, bear, base, bull, _, __ in dcf_assumptions:
        ws.row_dimensions[r].height = 22
        wc(ws, r, 2, label, bold=True, fg=TTD_DARK, bg=SUBHDR_BG, border=True)
        wc(ws, r, 3, bear, bold=False, bg=ALT_ROW, border=True, align="center")
        wc(ws, r, 4, base, bold=False, bg=ALT_ROW, border=True, align="center")
        wc(ws, r, 5, bull, bold=False, bg=ALT_ROW, border=True, align="center")
        for col in [6, 7]:
            wc(ws, r, col, "", bg=WHITE, border=True)
        r += 1

    # DCF Projections — Base Case
    r += 1
    ws.merge_cells(f"B{r}:G{r}")
    wc(ws, r, 2, "BASE CASE DCF PROJECTION (20% Rev CAGR | 28% FCF Margin | 10% WACC | 3% TGR)",
       bold=True, fg=HEADER_FG, bg=TTD_DARK, size=14)
    r += 1
    header_row(ws, r, [2, 3, 4, 5, 6, 7],
               ["Year", "FY 2026E", "FY 2027E", "FY 2028E", "FY 2029E", "FY 2030E"])

    base_proj = [
        ("Revenue ($M)", [3475, 4170, 4921, 5757, 6621]),
        ("Revenue Growth", [0.200, 0.200, 0.180, 0.170, 0.150]),
        ("FCF Margin", [0.270, 0.280, 0.290, 0.290, 0.300]),
        ("Free Cash Flow ($M)", [938, 1168, 1427, 1670, 1986]),
        ("Discount Factor (10%)", [0.909, 0.826, 0.751, 0.683, 0.621]),
        ("PV of FCF ($M)", [853, 965, 1072, 1140, 1233]),
    ]
    r += 1
    for i, (label, values) in enumerate(base_proj):
        ws.row_dimensions[r].height = 22
        alt = i % 2 == 0
        bg = ALT_ROW if alt else WHITE
        is_key = "Revenue ($M)" in label or "Free Cash Flow" in label
        wc(ws, r, 2, label, bold=is_key, fg=TTD_DARK, bg=bg, border=True)
        for j, v in enumerate(values):
            fmt = '#,##0' if isinstance(v, int) or (isinstance(v, float) and v > 1) else '0.0%'
            if "Factor" in label or "Discount" in label:
                fmt = '0.000'
            wc(ws, r, 3 + j, v, bold=is_key, bg=bg, border=True, align="right", num_fmt=fmt)
        r += 1

    # DCF Summary
    r += 1
    dcf_summary = [
        ("Sum of PV (FCFs)", "$5,263M", "$5,263M", "$4,180M", "$7,200M", ""),
        ("Terminal Value (@ Year 5 FCF)", "$28,446M", "$29,263M", "$15,000M", "$55,000M", ""),
        ("PV of Terminal Value", "$17,661M", "$18,170M", "$9,313M", "$34,154M", ""),
        ("Enterprise Value (DCF)", "$22,924M", "$23,433M", "$13,493M", "$41,354M", ""),
        ("+ Net Cash", "$222M", "$222M", "$222M", "$222M", ""),
        ("Equity Value", "$23,146M", "$23,655M", "$13,715M", "$41,576M", ""),
        ("Shares Outstanding", "470M", "470M", "470M", "470M", ""),
        ("INTRINSIC VALUE PER SHARE", "$49.20", "$50.30", "$29.20", "$88.50", ""),
        ("Current Price", "$21.00", "$21.00", "$21.00", "$21.00", ""),
        ("Upside / (Downside)", "+134%", "+140%", "+39%", "+321%", ""),
        ("Margin of Safety", "HIGH", "HIGH", "MODERATE", "VERY HIGH", ""),
    ]

    header_row(ws, r, [2, 3, 4, 5, 6],
               ["DCF Summary", "Peer Base Case", "Base Case", "Bear Case", "Bull Case"])
    r += 1
    for i, row_data in enumerate(dcf_summary):
        ws.row_dimensions[r].height = 22
        label, v1, v2, v3, v4 = row_data[0], row_data[1], row_data[2], row_data[3], row_data[4]
        alt = i % 2 == 0
        bg = ALT_ROW if alt else WHITE
        is_key = "INTRINSIC" in label or "Margin of Safety" in label or "Upside" in label
        row_bg = SUBHDR_BG if is_key else bg

        wc(ws, r, 2, label, bold=is_key, fg=TTD_DARK if is_key else "333333", bg=row_bg, border=True)
        for col, val in zip([3, 4, 5, 6], [v1, v2, v3, v4]):
            fg_color = "000000"
            if is_key and "INTRINSIC" in label:
                fg_color = GREEN
            elif "Upside" in label:
                fg_color = GREEN if "+" in str(val) else RED
            elif "Margin of Safety" in label:
                fg_color = GREEN if val in ["HIGH", "VERY HIGH"] else ORANGE
            wc(ws, r, col, val, bold=is_key, fg=fg_color, bg=row_bg,
               border=True, align="center" if is_key else "right")
        r += 1

    # ── Section 4: Historical Multiples ────────────────────────────────────
    r += 1
    ws.merge_cells(f"B{r}:G{r}")
    wc(ws, r, 2, "SECTION 4: HISTORICAL VALUATION MULTIPLES", bold=True, fg=HEADER_FG,
       bg=TTD_MID, size=15)
    r += 1
    header_row(ws, r, [2, 3, 4, 5, 6, 7],
               ["Multiple", "2021", "2022", "2023", "2024", "2025 (Apr)"])
    hist_mult = [
        ("EV/Revenue (NTM)", "25x", "8x", "15x", "18x", "3.2x"),
        ("EV/EBITDA (NTM)", "175x", "75x", "100x", "85x", "7.7x (adj)"),
        ("Price/FCF", "290x", "80x", "140x", "125x", "12.4x"),
        ("Revenue Growth Rate", "43%", "32%", "23%", "26%", "18%"),
        ("Commentary", "Peak bubble valuation", "Macro correction", "Recovery", "Growth premium", "Crisis-level cheapness"),
    ]
    r += 1
    for i, row_data in enumerate(hist_mult):
        ws.row_dimensions[r].height = 22
        bg = ALT_ROW if i % 2 == 0 else WHITE
        for j, val in enumerate(row_data):
            wc(ws, r, 2 + j, val, bold=(j == 0), fg=TTD_DARK if j == 0 else "333333",
               bg=bg, border=True, align="left" if j == 0 else "center")
        r += 1


# ══════════════════════════════════════════════════════════════════════════════
# TAB 11 – MARKET SENTIMENT
# ══════════════════════════════════════════════════════════════════════════════
def build_sentiment(wb):
    ws = wb.create_sheet("11. Market Sentiment")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, {1: 4, 2: 34, 3: 14, 4: 52})

    r = 2
    ws.merge_cells("B2:D2")
    wc(ws, r, 2, "MARKET SENTIMENT — THE TRADE DESK (April 2026)", bold=True, fg=HEADER_FG,
       bg=TTD_DARK, align="center", size=18)

    sentiment_sections = [
        ("ANALYST CONSENSUS", [
            ("Consensus Rating", "BUY (32 of 36 covering analysts)", "Analyst community remains largely bullish despite stock's "
             "75% decline from highs. Most view the Publicis dispute as temporary and maintain long-term structural thesis."),
            ("Average Price Target", "$46.63 (range: $23 - $100)", "Implies 122% upside from $21 current price. "
             "The wide range ($23-$100) reflects genuine uncertainty about moat durability and agency relationship outcome."),
            ("Recent Target Cuts", "Multiple analysts cut targets in late 2025 / early 2026", "Following Q3 2025 miss on "
             "forward guidance and Publicis dispute. Average target came down from ~$80 to ~$47. "
             "Morningstar cut fair value to $29 and removed moat rating."),
            ("Q1 2026 Preview", "May 7, 2026 earnings date — critical catalyst", "Will be first full quarter showing "
             "Publicis impact. Revenue guidance of $575-600M would show minimal impact; below $550M would confirm thesis."),
        ]),
        ("INSTITUTIONAL & INSIDER ACTIVITY", [
            ("Institutional Ownership", "76.3% of shares (417M of 470M float)", "Strong institutional backing from ARK Invest, "
             "Baillie Gifford, and multiple growth-oriented funds. Potential selling pressure if growth story breaks."),
            ("CEO Insider Buy (March 2026)", "Jeff Green purchased 6M shares for ~$148M", "Record largest insider purchase in "
             "company history. CEO paid ~$25/share. Largest insider buy signals strong conviction of undervaluation."),
            ("Short Interest", "14.2% of shares outstanding (66.8M shares)", "Elevated short interest — highest in company history. "
             "Bears are betting on agency relationship deterioration and ad spend slowdown. This also creates potential "
             "short squeeze dynamics if Q1 2026 results come in above expectations."),
            ("Recent SEC Form 4s", "Trust transfers (non-economic) in Aug/Nov 2025", "Jeff Green transferred shares to trusts "
             "for estate planning — not economic sales. Net: no selling, significant buying."),
        ]),
        ("CURRENT MARKET NARRATIVE", [
            ("Bull Narrative", "Structural winner in programmatic", "CTV + retail media + identity (UID2) = three multi-year "
             "secular growth vectors. At $21, priced for zero growth. 95%+ retention proves platform stickiness. "
             "CEO buying $148M at $25 is the signal. Margin of safety is exceptional."),
            ("Bear Narrative", "Moat is eroding", "Publicis dispute exposes that TTD's 'transparency' brand is inconsistent "
             "with undisclosed fees. If WPP and Dentsu follow Publicis, 30% of gross spend is at risk. Meanwhile "
             "Amazon DSP and Google DV360 are becoming more capable. Revenue growth decelerating from 43% to 18%."),
            ("Macro Headwinds", "Ad spend tied to economic cycle", "In tariff/trade uncertainty environment (April 2026), "
             "brand advertisers are cutting budgets. Performance advertising is more resilient. TTD has limited CTV "
             "direct response, which is more budget-stable than brand awareness campaigns."),
            ("AI Disruption Narrative", "Generative AI threatening the ad ecosystem", "AI-generated content flooding the "
             "internet reduces publisher quality. AI-driven targeting may shift power away from DSPs toward AI agents "
             "that buy media directly. But TTD is investing in AI (Kokai) and this may be opportunity not threat."),
        ]),
        ("COMPETITIVE TRENDS (2026)", [
            ("CTV / Streaming", "FAVORABLE", "Streaming advertising is growing rapidly as cord-cutting continues. "
             "Disney+, Peacock, Max all launched/expanded ad tiers. TTD is well-positioned as the dominant independent "
             "CTV DSP. Netflix is exploring TTD partnership for ad inventory."),
            ("Retail Media", "FAVORABLE", "Retail media is the fastest-growing digital ad format. TTD created a dedicated "
             "Retail Media business unit in December 2025. Partners with Walmart, Target, Kroger. "
             "Counter-positioning against Amazon's retail data advantage."),
            ("Audio/Podcast", "FAVORABLE", "Programmatic audio growing 27% YoY. Spotify, podcast networks, and "
             "streaming radio increasingly moving to programmatic. TTD is investing in audio capabilities."),
            ("Open Web Display", "CHALLENGED", "Traditional display advertising on the open web is commoditizing. "
             "Brands prefer walled gardens for performance marketing. This is the weakest part of TTD's business."),
        ]),
    ]

    for sec_title, items in sentiment_sections:
        r += 1
        ws.row_dimensions[r].height = 26
        ws.merge_cells(f"B{r}:D{r}")
        wc(ws, r, 2, sec_title, bold=True, fg=HEADER_FG, bg=TTD_MID, align="left", size=15)
        r += 1
        if len(items[0]) == 3:
            header_row(ws, r, [2, 3, 4], ["Topic / Signal", "Data Point", "Analysis"])
        else:
            header_row(ws, r, [2, 3, 4], ["Topic", "Assessment", "Detail"])
        r += 1
        for i, row_item in enumerate(items):
            ws.row_dimensions[r].height = 60
            alt = i % 2 == 0
            bg = ALT_ROW if alt else WHITE
            label, mid, detail = row_item[0], row_item[1], row_item[2]
            if "FAVORABLE" in mid:
                mid_bg = GREEN
                mid_fg = WHITE
            elif "CHALLENGED" in mid:
                mid_bg = ORANGE
                mid_fg = WHITE
            elif "BULL" in label:
                mid_bg = GREEN
                mid_fg = WHITE
            elif "BEAR" in label:
                mid_bg = RED
                mid_fg = WHITE
            else:
                mid_bg = bg
                mid_fg = TTD_DARK
            wc(ws, r, 2, label, bold=True, fg=TTD_DARK, bg=bg, border=True, wrap=True)
            wc(ws, r, 3, mid, bold=True, fg=mid_fg, bg=mid_bg, border=True, wrap=True, align="center")
            wc(ws, r, 4, detail, bold=False, fg="222222", bg=bg, border=True, wrap=True)
            r += 1
        r += 1


# ══════════════════════════════════════════════════════════════════════════════
# TAB 12 – KEY INDICATORS
# ══════════════════════════════════════════════════════════════════════════════
def build_key_indicators(wb):
    ws = wb.create_sheet("12. Key Indicators")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, {1: 4, 2: 38, 3: 16, 4: 16, 5: 16, 6: 16, 7: 16, 8: 20})
    freeze(ws, "C4")

    r = 2
    ws.merge_cells("B2:H2")
    wc(ws, r, 2, "KEY PERFORMANCE INDICATORS — THE TRADE DESK (TTD)", bold=True, fg=HEADER_FG,
       bg=TTD_DARK, align="center", size=18)

    r = 3
    header_row(ws, r, [2, 3, 4, 5, 6, 7, 8],
               ["KPI", "FY 2021", "FY 2022", "FY 2023", "FY 2024", "FY 2025", "Trend"])

    kpi_sections = [
        ("GROWTH METRICS", [
            ("Revenue ($M)", [1196, 1578, 1946, 2445, 2896], '#,##0', "↗ Strong"),
            ("Revenue Growth YoY", [0.431, 0.319, 0.233, 0.256, 0.185], '0.0%', "↘ Decelerating"),
            ("3-Year Revenue CAGR", [None, None, None, 0.271, 0.234], '0.0%', "↘ Moderating"),
            ("Platform Spend Growth (Gross)", [0.45, 0.30, 0.24, 0.27, 0.20], '0.0%', "↘ Normalizing"),
            ("Customer Count Growth (est.)", [0.12, 0.10, 0.09, 0.08, 0.07], '0.0%', "→ Stable"),
        ]),
        ("PROFITABILITY METRICS", [
            ("Gross Margin %", [0.815, 0.822, 0.812, 0.807, 0.786], '0.0%', "↘ Slight compression"),
            ("Operating Margin % (GAAP)", [0.104, 0.072, 0.103, 0.175, 0.204], '0.0%', "↗ Improving"),
            ("Adj. EBITDA Margin %", [None, None, None, 0.384, 0.414], '0.0%', "↗ Expanding"),
            ("Net Income Margin %", [0.115, 0.034, 0.092, 0.161, 0.153], '0.0%', "→ Stable"),
            ("FCF Margin %", [0.271, 0.295, 0.283, 0.262, 0.275], '0.0%', "→ Resilient"),
        ]),
        ("CASH & CAPITAL", [
            ("Operating Cash Flow ($M)", [379, 549, 598, 739, 993], '#,##0', "↗ Strong growth"),
            ("Free Cash Flow ($M)", [324, 465, 551, 641, 796], '#,##0', "↗ Consistent growth"),
            ("FCF per Share ($)", [0.65, 0.95, 1.13, 1.30, 1.69], '$0.00', "↗ Growing"),
            ("Net Cash Position ($M)", [469, 770, 659, 1057, 222], '#,##0', "↘ Buyback impact"),
            ("Share Buybacks ($M)", [0, 24, 673, 148, 1411], '#,##0', "↗ Accelerating"),
        ]),
        ("RETURNS & EFFICIENCY", [
            ("ROIC %", [0.087, 0.035, 0.083, 0.192, 0.251], '0.0%', "↗ Improving rapidly"),
            ("ROE %", [0.090, 0.025, 0.083, 0.133, 0.178], '0.0%', "↗ Strong"),
            ("ROA %", [0.039, 0.012, 0.037, 0.064, 0.072], '0.0%', "↗ Improving"),
            ("Asset Turnover", [0.335, 0.360, 0.398, 0.400, 0.471], '0.000x', "↗ Efficient"),
            ("Revenue / Employee (est. $K)", [625, 680, 740, 810, 870], '#,##0', "↗ Productivity gains"),
        ]),
        ("QUALITY METRICS", [
            ("Customer Retention Rate", [0.95, 0.95, 0.95, 0.95, 0.95], '0.0%', "→ Rock-solid"),
            ("SBC as % of Revenue", [0.138, 0.166, 0.173, 0.173, 0.171], '0.0%', "→ Stubbornly high"),
            ("SBC ($M)", [165, 262, 336, 424, 495], '#,##0', "↗ Growing (dilutive)"),
            ("FCF Conversion (FCF/NI)", [2.35, 8.77, 3.08, 1.63, 1.80], '0.00x', "↘ Normalizing"),
            ("Owner Earnings ($M)", [159, 203, 215, 217, 301], '#,##0', "↗ Growing slowly"),
        ]),
        ("RULE OF 50 (SAAS BENCHMARK)", [
            ("Revenue Growth %", [0.431, 0.319, 0.233, 0.256, 0.185], '0.0%', "↘ Declining"),
            ("FCF Margin %", [0.271, 0.295, 0.283, 0.262, 0.275], '0.0%', "→ Stable"),
            ("Rule of 50 Score", [70.2, 61.4, 51.7, 51.8, 46.0], '0.0', "↘ Fell below 50"),
            ("Rule of 50 Pass/Fail", ["PASS", "PASS", "PASS", "PASS", "FAIL"], None, "⚠ Watch in 2026"),
        ]),
        ("VALUATION TRACKING", [
            ("Stock Price (year-end)", [None, None, None, 90.0, 21.0], '$0.00', "↘ Massive de-rating"),
            ("Market Cap ($B)", [None, None, None, 44.4, 9.9], '$0.0', "↘ -78% from peak"),
            ("EV/Revenue Multiple", [25.0, 8.0, 15.0, 18.0, 3.2], '0.0x', "↘ Near trough"),
            ("EV/FCF Multiple", [None, None, 140.0, 125.0, 11.6], '0.0x', "↘ Historically cheap"),
            ("P/E Ratio (TTM)", [None, None, None, 115.0, 23.6], '0.0x', "↘ Normalized"),
        ]),
    ]

    trend_colors = {
        "↗": GREEN, "↘": RED, "→": TTD_BLUE,
        "⚠": ORANGE, "✓": GREEN,
    }

    for sec_title, items in kpi_sections:
        r += 1
        ws.row_dimensions[r].height = 26
        ws.merge_cells(f"B{r}:H{r}")
        wc(ws, r, 2, sec_title, bold=True, fg=HEADER_FG, bg=TTD_MID, align="left", size=15)
        r += 1
        for i, (label, values, num_fmt, trend) in enumerate(items):
            ws.row_dimensions[r].height = 22
            alt = i % 2 == 0
            bg = ALT_ROW if alt else WHITE
            wc(ws, r, 2, label, bold=False, fg="333333", bg=bg, border=True)
            for j, v in enumerate(values):
                cell = ws.cell(row=r, column=3 + j, value=v)
                cell.font = mf()
                cell.fill = mfill(bg)
                cell.border = mborder()
                cell.alignment = cal("right" if not isinstance(v, str) else "center")
                if num_fmt and v is not None and not isinstance(v, str):
                    cell.number_format = num_fmt
                if isinstance(v, str) and v in ["PASS", "FAIL"]:
                    cell.fill = mfill(GREEN if v == "PASS" else RED)
                    cell.font = mf(bold=True, color=WHITE)
                    cell.alignment = cal("center")

            # Trend column
            trend_cell = ws.cell(row=r, column=8, value=trend)
            tcolor = "000000"
            for arrow, col in trend_colors.items():
                if trend.startswith(arrow):
                    tcolor = col
                    break
            trend_cell.font = mf(bold=True, color=tcolor)
            trend_cell.fill = mfill(bg)
            trend_cell.border = mborder()
            trend_cell.alignment = cal("left")
            r += 1
        r += 1

    # Final summary
    r += 1
    ws.merge_cells(f"B{r}:H{r}")
    wc(ws, r, 2,
       "SUMMARY: TTD is a high-quality programmatic advertising business with 95%+ customer retention, 78%+ gross margins, "
       "and growing ROIC (25%+). The key risks are agency fee disputes and growth deceleration. At $21/share "
       "($9.9B market cap), the stock appears to price in significant deterioration. Our DCF suggests $50+ fair value "
       "in a base case. Rule of 50 fell below threshold in 2025 (46.0) — watch for recovery in 2026 guidance.",
       bold=True, fg=WHITE, bg=TTD_DARK, wrap=True, size=14)


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════
def main():
    output_dir = "/Users/naji/WORK/github.com/AI/claude/Agent/MarketResearch/output"
    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(output_dir, "TTD_Financial_Analysis.xlsx")

    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    print("Building tabs...")
    build_cover(wb)
    print("  ✓ Cover")
    build_business(wb)
    print("  ✓ Business Overview")
    build_moat(wb)
    print("  ✓ Moat Analysis")
    build_income(wb)
    print("  ✓ Income Statement")
    build_balance(wb)
    print("  ✓ Balance Sheet")
    build_cashflow(wb)
    print("  ✓ Cash Flow Analysis")
    build_roic(wb)
    print("  ✓ Return on Capital")
    build_management(wb)
    print("  ✓ Management")
    build_risks(wb)
    print("  ✓ Risks")
    build_valuation(wb)
    print("  ✓ Valuation")
    build_sentiment(wb)
    print("  ✓ Market Sentiment")
    build_key_indicators(wb)
    print("  ✓ Key Indicators")

    wb.save(output_path)
    print(f"\n✅ Saved: {output_path}")
    return output_path


if __name__ == "__main__":
    main()
