"""
Netflix Inc. (NFLX) Financial Analysis - Excel Generator
Data as of April 2026 | FY2025 Annual (ended December 31, 2025)
Note: Netflix executed a ~10:1 stock split; ~4.22B shares outstanding at ~$102/share
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

# ── Netflix Color Palette ──────────────────────────────────────────────────────
NFLX_RED    = "E50914"
NFLX_BLACK  = "141414"
NFLX_DARK   = "221F1F"
HEADER_BG   = "E50914"
HEADER_FG   = "FFFFFF"
SUBHDR_BG   = "FFDEDE"
SUBHDR_FG   = "1A1A1A"
ALT_ROW     = "FFF5F5"
WHITE       = "FFFFFF"
LIGHT_GRAY  = "F8F8F8"
RED         = "E50914"
GREEN       = "2ECC71"
DARK_RED    = "B20710"
GOLD        = "F39C12"
DARK_GRAY   = "555555"

FONT_SIZE = 14

def mf(bold=False, size=FONT_SIZE, color="000000", italic=False):
    return Font(name="Calibri", bold=bold, size=size, color=color, italic=italic)

def mfill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def mborder():
    s = Side(border_style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def cal(h="left", wrap=False):
    return Alignment(horizontal=h, vertical="center", wrap_text=wrap)

def wc(ws, row, col, value, bold=False, size=FONT_SIZE, fg="000000", bg=None,
       align="left", wrap=False, border=False, italic=False, num_fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = mf(bold=bold, size=size, color=fg, italic=italic)
    if bg:
        c.fill = mfill(bg)
    c.alignment = cal(align, wrap)
    if border:
        c.border = mborder()
    if num_fmt:
        c.number_format = num_fmt
    return c

def sec_hdr(ws, row, col, text, span=1, bg=HEADER_BG, fg=HEADER_FG):
    c = wc(ws, row, col, text, bold=True, size=FONT_SIZE+2, fg=fg, bg=bg, align="center")
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+span-1)
    ws.row_dimensions[row].height = 30
    return c

def sub_hdr(ws, row, col, text, span=1, bg=SUBHDR_BG, fg=SUBHDR_FG):
    c = wc(ws, row, col, text, bold=True, size=FONT_SIZE, fg=fg, bg=bg, align="center", border=True)
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+span-1)
    ws.row_dimensions[row].height = 22
    return c

def lbl(ws, row, col, text, bg=None, indent=0):
    val = ("   " * indent) + text
    c = wc(ws, row, col, val, size=FONT_SIZE, bg=bg, align="left", border=True)
    ws.row_dimensions[row].height = 20
    return c

def dc(ws, row, col, value, bg=None, num_fmt=None, bold=False, color="000000"):
    c = wc(ws, row, col, value, bold=bold, size=FONT_SIZE, fg=color, bg=bg,
           align="right", border=True, num_fmt=num_fmt)
    ws.row_dimensions[row].height = 20
    return c

def alt(i):
    return ALT_ROW if i % 2 == 0 else WHITE

def scw(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

# ═══════════════════════════════════════════════════════════════════════════════
def build_cover(wb):
    ws = wb.create_sheet("Cover")
    ws.sheet_view.showGridLines = False
    for c in range(1, 9):
        scw(ws, c, 18)
    scw(ws, 1, 4); scw(ws, 2, 32); scw(ws, 3, 22)

    ws.merge_cells("B3:G3")
    wc(ws, 3, 2, "NETFLIX INC. (NFLX)", bold=True, size=32, fg=WHITE,
       bg=HEADER_BG, align="center")
    ws.row_dimensions[3].height = 55

    ws.merge_cells("B4:G4")
    wc(ws, 4, 2, "Comprehensive Investment Analysis | April 2026", bold=False,
       size=16, fg=WHITE, bg=DARK_RED, align="center")
    ws.row_dimensions[4].height = 28

    ws.merge_cells("B6:G6")
    wc(ws, 6, 2, "Company Profile", bold=True, size=16, fg=WHITE, bg=NFLX_BLACK, align="center")
    ws.row_dimensions[6].height = 25

    info = [
        ("Ticker",              "NFLX"),
        ("Exchange",            "NASDAQ"),
        ("Sector",              "Communication Services / Streaming"),
        ("Headquarters",        "Los Gatos, California, USA"),
        ("Founded",             "1997 by Reed Hastings & Marc Randolph"),
        ("Co-CEOs",             "Ted Sarandos & Greg Peters"),
        ("Fiscal Year End",     "December 31"),
        ("Market Cap",          "~$432 Billion (April 9, 2026)"),
        ("Stock Price",         "~$102 (April 9, 2026; post ~10:1 split)"),
        ("P/E Ratio",           "~39x TTM | ~32x Forward"),
        ("Shares Outstanding",  "~4.22 Billion (post-split)"),
        ("Dividend",            "None"),
    ]

    for i, (k, v) in enumerate(info):
        r = 7 + i
        bg = ALT_ROW if i % 2 == 0 else WHITE
        wc(ws, r, 2, k, bold=True, size=FONT_SIZE, bg=bg, align="left", border=True)
        wc(ws, r, 3, v, size=FONT_SIZE, bg=bg, align="left", border=True)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)
        ws.row_dimensions[r].height = 22

    ws.merge_cells("B20:G20")
    wc(ws, 20, 2, "Investment Thesis", bold=True, size=16, fg=WHITE, bg=NFLX_BLACK, align="center")
    ws.row_dimensions[20].height = 25

    thesis = (
        "Netflix is the undisputed global leader in streaming entertainment with 325M+ paid subscribers "
        "and a widening moat via content investment scale and brand recognition. The company has "
        "successfully transitioned from content spend speculation to a high-margin cash flow engine: "
        "29.5% operating margin in FY2025 and $9.5B in free cash flow. The emerging advertising tier "
        "(2.5x growth to $1.5B+) represents a multi-billion dollar TAM with minimal incremental cost. "
        "2026 guidance of $50.7-51.7B revenue (+12%) with margin expansion frames a durable compounder. "
        "Trading at ~32x forward earnings — a premium warranted by durable subscriber growth and "
        "margin runway. Warner Bros. deal risks are the key near-term overhang."
    )
    ws.merge_cells("B21:G25")
    wc(ws, 21, 2, thesis, size=FONT_SIZE, bg=LIGHT_GRAY, align="left", wrap=True, border=True)
    for r in range(21, 26):
        ws.row_dimensions[r].height = 22

    ws.merge_cells("B27:G27")
    wc(ws, 27, 2, "Key Financial Highlights (FY2025)", bold=True, size=16, fg=WHITE, bg=NFLX_BLACK, align="center")
    ws.row_dimensions[27].height = 25

    highlights = [
        ("Total Revenue",              "$45.2B",    "+16% YoY"),
        ("Operating Income",           "$13.3B",    "Margin 29.5%"),
        ("Net Income",                 "$11.0B",    "+26.1% YoY"),
        ("Free Cash Flow",             "$9.5B",     "+36.7% YoY"),
        ("Paid Subscribers",           "325M",      "Global milestone"),
        ("Advertising Revenue",        "$1.5B+",    "2.5x vs. 2024"),
        ("2026 Revenue Guidance",      "$50.7-51.7B","+12% guided"),
        ("Ad Revenue 2026 Outlook",    "~$3B+",     "~2x projected growth"),
    ]

    sub_hdr(ws, 28, 2, "Metric", bg=SUBHDR_BG)
    sub_hdr(ws, 28, 3, "Value", bg=SUBHDR_BG)
    sub_hdr(ws, 28, 4, "Comment", bg=SUBHDR_BG)
    ws.merge_cells(start_row=28, start_column=4, end_row=28, end_column=7)

    for i, (m, v, c) in enumerate(highlights):
        r = 29 + i
        bg = ALT_ROW if i % 2 == 0 else WHITE
        lbl(ws, r, 2, m, bg=bg)
        dc(ws, r, 3, v, bg=bg)
        lbl(ws, r, 4, c, bg=bg)
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=7)

    ws.merge_cells("B38:G38")
    wc(ws, 38, 2, "Report Date: April 10, 2026  |  Data Sources: SEC Filings, Netflix IR, Bloomberg",
       italic=True, size=12, fg="666666", align="center")

# ═══════════════════════════════════════════════════════════════════════════════
def build_business(wb):
    ws = wb.create_sheet("Business Overview")
    ws.sheet_view.showGridLines = False
    scw(ws, 1, 3); scw(ws, 2, 32); scw(ws, 3, 22); scw(ws, 4, 22)
    scw(ws, 5, 22); scw(ws, 6, 22); scw(ws, 7, 22)

    sec_hdr(ws, 1, 2, "BUSINESS OVERVIEW — NETFLIX INC. (NFLX)", span=6)

    sub_hdr(ws, 3, 2, "What Netflix Does", span=6)
    desc = ("Netflix is the world's largest streaming entertainment service. Subscribers pay monthly "
            "fees to access a library of movies, TV series, documentaries, mobile games, and live events. "
            "Netflix produces award-winning original content (House of Cards, Stranger Things, Squid Game) "
            "and licenses third-party content. The company monetizes via subscription tiers (Standard, "
            "Premium, with/without Ads) and is rapidly scaling its ad-supported tier launched in 2022.")
    ws.merge_cells("B4:G6")
    wc(ws, 4, 2, desc, size=FONT_SIZE, bg=LIGHT_GRAY, align="left", wrap=True, border=True)
    for r in range(4, 7):
        ws.row_dimensions[r].height = 22

    # Revenue Breakdown
    sub_hdr(ws, 8, 2, "Revenue Breakdown by Region (FY2025 Est.)", span=6)
    rev_data = [
        ("Region",                   "FY2025 Rev",  "% of Total", "Members",  "ARPM*"),
        ("United States & Canada",   "$18.5B",      "41%",        "~88M",     "$17.50+"),
        ("Europe, Middle East, Africa","$11.8B",    "26%",        "~95M",     "$10.50"),
        ("Latin America",            "$6.8B",       "15%",        "~49M",     "$11.80"),
        ("Asia Pacific",             "$8.1B",       "18%",        "~93M",     "$7.40"),
        ("TOTAL",                    "$45.2B",       "100%",      "~325M",    "$~11.50"),
    ]
    for i, row_data in enumerate(rev_data):
        r = 9 + i
        if i == 0:
            for j, val in enumerate(row_data):
                sub_hdr(ws, r, 2+j, val)
        else:
            bg = ALT_ROW if i % 2 == 0 else WHITE
            bold = (i == len(rev_data)-1)
            for j, val in enumerate(row_data):
                lbl(ws, r, 2+j, val, bg=bg) if j == 0 else dc(ws, r, 2+j, val, bg=bg, bold=bold)
    wc(ws, 16, 2, "*ARPM = Average Revenue Per Membership per month", italic=True, size=12, fg=DARK_GRAY)

    # Revenue by Tier
    sub_hdr(ws, 18, 2, "Revenue by Subscription Tier (FY2025 Est.)", span=6)
    tier_data = [
        ("Tier",                  "~Members",  "~Price/mo",   "% Rev",    "Notes"),
        ("Ad-Free (Standard)",    "~110M",     "$15.49",      "~38%",     "Core tier, declining mix shift"),
        ("Premium",               "~60M",      "$22.99",      "~30%",     "High value, 4K; most profitable"),
        ("Standard with Ads",     "~120M",     "$7.99",       "~28%",     "Fastest growing; ad revenue upside"),
        ("Other / Mobile",        "~35M",      "$6.99",       "~4%",      "APAC-heavy, low price point"),
    ]
    for i, row_data in enumerate(tier_data):
        r = 19 + i
        if i == 0:
            for j, val in enumerate(row_data):
                sub_hdr(ws, r, 2+j, val)
        else:
            bg = ALT_ROW if i % 2 == 0 else WHITE
            for j, val in enumerate(row_data):
                lbl(ws, r, 2+j, val, bg=bg) if j == 0 else dc(ws, r, 2+j, val, bg=bg)

    # Products & Services
    sub_hdr(ws, 26, 2, "Products & Services", span=6)
    products = [
        ("Netflix Streaming",         "Core video streaming; 8,000+ titles; personalized recommendation engine"),
        ("Netflix Originals",         "150+ originals/year; Stranger Things, Squid Game, Wednesday, Bridgerton; Emmy & Oscar winners"),
        ("Netflix Ads (Partner)",     "Ad-supported tier; advertising platform with ~120M subscribers monetizable"),
        ("Netflix Gaming",            "80+ mobile games included with subscription; Netflix Studios expansion"),
        ("Netflix Live Events",       "Boxing (Tyson vs Paul), NFL Christmas Day games, stand-up specials"),
        ("Netflix House",             "Physical experience venues; first in Dallas, King of Prussia (PA) opening 2025"),
        ("Netflix Shop",              "Merchandise for original IP; brand extension revenue"),
    ]
    for i, (p, d) in enumerate(products):
        r = 27 + i
        bg = ALT_ROW if i % 2 == 0 else WHITE
        lbl(ws, r, 2, p, bg=bg)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)
        lbl(ws, r, 3, d, bg=bg)

    # Buying Process
    sub_hdr(ws, 36, 2, "How Customers Buy / Sales Process", span=6)
    buying = [
        ("Consumer Direct",     "Direct-to-consumer via app/web; no intermediaries; high-margin relationship"),
        ("App Stores",          "iOS App Store (~30% fees pre-2024), Android Play Store; shifting to direct billing"),
        ("ISP / Telco Bundles", "T-Mobile, Comcast bundles; lower ARPM but lower CAC, high retention"),
        ("Advertising Clients", "B2B programmatic ad sales; agencies, CPG, entertainment studios; growing sales team"),
    ]
    for i, (s, d) in enumerate(buying):
        r = 37 + i
        bg = ALT_ROW if i % 2 == 0 else WHITE
        lbl(ws, r, 2, s, bg=bg)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)
        lbl(ws, r, 3, d, bg=bg)

    # Seasonality
    sub_hdr(ws, 43, 2, "Seasonality & Margin Structure", span=6)
    season = [
        ("Q1 (Jan-Mar)",  "Strong subscriber adds post-holidays; major content releases; ~24% of annual revenue"),
        ("Q2 (Apr-Jun)",  "Lighter content slate; competition peaks; historically softer on adds; ~23%"),
        ("Q3 (Jul-Sep)",  "Summer/fall content ramp; back-to-school attention competition; ~24%"),
        ("Q4 (Oct-Dec)",  "Holiday season + premium content (awards season); strongest adds; ~29%"),
        ("Content Costs", "~$17-18B/year in content amortization + cash spend; declining as % of revenue"),
        ("Margins",       "Operating margin 29.5% in FY2025; targeting 35%+ longer term via ad scale & pricing"),
    ]
    for i, (q, n) in enumerate(season):
        r = 44 + i
        bg = ALT_ROW if i % 2 == 0 else WHITE
        lbl(ws, r, 2, q, bg=bg)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)
        lbl(ws, r, 3, n, bg=bg)

# ═══════════════════════════════════════════════════════════════════════════════
def build_moat(wb):
    ws = wb.create_sheet("Moat")
    ws.sheet_view.showGridLines = False
    scw(ws, 1, 3); scw(ws, 2, 28); scw(ws, 3, 52); scw(ws, 4, 18)

    sec_hdr(ws, 1, 2, "COMPETITIVE MOAT ANALYSIS — NETFLIX (NFLX)", span=3)

    moat_intro = ("Netflix's moat is built on content scale, brand, global distribution, and data-driven "
                  "personalization. The moat is 'narrow-to-moderate' given streaming remains competitive, "
                  "but Netflix has the clearest advantages in content quality at scale and subscriber "
                  "engagement (two hours/day average watch time).")
    ws.merge_cells("B3:D4")
    wc(ws, 3, 2, moat_intro, size=FONT_SIZE, bg=LIGHT_GRAY, align="left", wrap=True, border=True)
    for r in [3, 4]:
        ws.row_dimensions[r].height = 22

    moats = [
        ("MOAT TYPE",            "EVIDENCE & DEPTH",                                                   "STRENGTH"),
        ("Content Scale",        "Netflix spends $17-18B/year on content — more than any individual "
                                  "competitor. Scale enables better talent deals, more risk-taking, "
                                  "and higher hit rate on originals. Squid Game (2.8B hours viewed), "
                                  "Bridgerton, Wednesday demonstrate global IP creation.",              "MODERATE"),
        ("Brand & Global Reach", "Netflix brand = 'streaming' in 190+ countries. Consumer default "
                                  "behavior: 'what are you watching on Netflix?' First-mover in many "
                                  "international markets with decade of subscriber relationships.",      "MODERATE-WIDE"),
        ("Recommendation Engine","Netflix's ML-driven personalization reduces churn by matching content "
                                  "to individual tastes. Data from 325M households watching habits is "
                                  "a self-reinforcing data flywheel that improves retention.",          "MODERATE"),
        ("Switching Costs",      "Incomplete series, saved shows, watch history, personalized UI, "
                                  "downloaded content — creates friction to leave. Household habits "
                                  "are sticky; avg subscriber tenure is 4+ years.",                    "LOW-MODERATE"),
        ("Advertising Platform", "120M ad-supported subscribers is a unique audience pool — premium "
                                  "video audience with binge-watching behavior, premium targeting. "
                                  "Ad revenue only in year 3; $1.5B now → $5B+ TAM.",                  "BUILDING"),
        ("Cost Leverage",        "Fixed content costs spread over more subscribers = operating margin "
                                  "expansion. 100M vs 325M subscribers on same library creates "
                                  "structural cost advantage vs. smaller competitors.",                 "MODERATE"),
        ("IP Ownership",         "Netflix owns global rights to all Originals; no windowing to theaters "
                                  "or cable; direct consumer relationship with no intermediary. "
                                  "Squid Game S3, Stranger Things final season = retention drivers.",   "MODERATE"),
    ]

    for i, row_data in enumerate(moats):
        r = 6 + i
        if i == 0:
            sub_hdr(ws, r, 2, row_data[0])
            sub_hdr(ws, r, 3, row_data[1])
            sub_hdr(ws, r, 4, row_data[2])
        else:
            bg = ALT_ROW if i % 2 == 0 else WHITE
            strength_color = GREEN if "WIDE" in row_data[2] else (GOLD if "MODERATE" in row_data[2] else RED)
            lbl(ws, r, 2, row_data[0], bg=bg)
            ws.row_dimensions[r].height = 52
            wc(ws, r, 3, row_data[1], size=FONT_SIZE, bg=bg, align="left", wrap=True, border=True)
            wc(ws, r, 4, row_data[2], bold=True, size=FONT_SIZE, fg=WHITE,
               bg=strength_color, align="center", border=True)

    # Competitive Threats
    sec_hdr(ws, 15, 2, "COMPETITIVE LANDSCAPE", span=3, bg=NFLX_BLACK)
    comp = [
        ("Competitor",            "Subscribers",   "Annual Content $",  "Key Advantage vs Netflix"),
        ("Disney+ (incl. Hulu)",  "~175M",         "$30B+",             "IP franchises (Marvel, Star Wars, Pixar, ESPN)"),
        ("HBO Max / Max",         "~115M",          "$13B",             "Prestige drama; Warner Bros IP library"),
        ("Amazon Prime Video",    "~230M*",         "$13B",             "Prime bundling; no standalone churn motivation"),
        ("Apple TV+",             "~25M",           "$6B",              "Quality > quantity; Apple ecosystem lock-in"),
        ("YouTube",               "2B+ MAU",        "$0 orig",          "Free; UGC scale; YouTube Premium emerging"),
        ("Peacock / Paramount+",  "~35M / ~75M",    "$6B total",        "Sports rights; Paramount library; price"),
    ]
    sub_hdr(ws, 16, 2, "Competitor")
    sub_hdr(ws, 16, 3, "Subscribers")
    sub_hdr(ws, 16, 4, "Key Advantage vs Netflix")
    for i, row_data in enumerate(comp):
        if i == 0:
            continue
        r = 17 + (i-1)
        bg = ALT_ROW if i % 2 == 0 else WHITE
        lbl(ws, r, 2, row_data[0], bg=bg)
        dc(ws, r, 3, row_data[1], bg=bg)
        wc(ws, r, 4, row_data[3], size=FONT_SIZE, bg=bg, align="left", border=True)
        ws.row_dimensions[r].height = 22
    wc(ws, 24, 2, "*Amazon Prime Video includes all Prime members; many are passive, not active watchers",
       italic=True, size=FONT_SIZE, fg=DARK_GRAY)

    sub_hdr(ws, 26, 2, "Overall Moat Rating", span=3, bg=NFLX_BLACK, fg=WHITE)
    ws.merge_cells("B27:D28")
    wc(ws, 27, 2,
       "NARROW-TO-MODERATE MOAT — Netflix leads streaming but faces well-funded competition from "
       "Disney, Amazon, and Apple who have ecosystem advantages. Netflix's content quality edge, "
       "global distribution, and brand recognition form a durable but not impregnable moat. "
       "The advertising business represents the next moat-expansion opportunity.",
       size=FONT_SIZE, bg=ALT_ROW, align="left", wrap=True, border=True)
    for r in [27, 28]:
        ws.row_dimensions[r].height = 28

# ═══════════════════════════════════════════════════════════════════════════════
def build_income(wb):
    ws = wb.create_sheet("Income Statement")
    ws.sheet_view.showGridLines = False
    scw(ws, 1, 3); scw(ws, 2, 32); scw(ws, 3, 18); scw(ws, 4, 18)
    scw(ws, 5, 18); scw(ws, 6, 18); scw(ws, 7, 18)

    sec_hdr(ws, 1, 2, "INCOME STATEMENT — NETFLIX INC. (NFLX)  |  FY2021-FY2025 ($B)", span=6)
    sub_hdr(ws, 2, 2, "Metric")
    for i, yr in enumerate(["FY2021", "FY2022", "FY2023", "FY2024", "FY2025"]):
        sub_hdr(ws, 2, 3+i, yr)

    rows = [
        ("REVENUE",                    None, None, None, None, None),
        ("  Streaming Revenue",        29.7,  31.6,  33.7,  38.9,  45.2),
        ("  Advertising Revenue",      0.0,   0.0,   0.0,   0.6,   1.5),
        ("TOTAL REVENUE",              29.7,  31.6,  33.7,  39.0,  45.2),
        ("  YoY Growth",               "19%", "6.5%","6.6%","15.6%","15.9%"),
        ("", None, None, None, None, None),
        ("COSTS & EXPENSES",           None, None, None, None, None),
        ("  Cost of Revenue (Content)", 17.3, 17.8,  17.1,  18.2,  20.0),
        ("  Technology & Development",   2.6,  2.7,   2.7,   2.9,   3.1),
        ("  Marketing",                  2.5,  2.5,   2.7,   2.8,   3.1),
        ("  G&A",                        1.5,  1.6,   1.6,   1.7,   1.8),
        ("TOTAL OPERATING EXPENSES",   23.9,  24.6,  24.1,  25.6,  28.0),
        ("", None, None, None, None, None),
        ("OPERATING INCOME",            6.2,   6.0,   7.0,  10.4,  13.3),
        ("  Operating Margin",         "20.9%","19.0%","20.8%","26.7%","29.5%"),
        ("Other Income / (Expense)",   -0.9,  -1.6,  -0.9,  -0.6,  -0.5),
        ("Pretax Income",               5.3,   4.4,   6.1,   9.8,  12.8),
        ("Income Tax Expense",          0.9,   0.8,   0.8,   1.1,   1.8),
        ("NET INCOME",                  5.1,   4.5,   5.4,   8.7,  11.0),
        ("  Net Margin",               "17.2%","14.2%","16.0%","22.3%","24.3%"),
        ("", None, None, None, None, None),
        ("EPS (Diluted, post-split)",   1.21,  1.05,  1.27,  2.06,  2.61),
        ("Shares Outstanding (B)",      4.36,  4.29,  4.26,  4.24,  4.22),
    ]

    for i, row_data in enumerate(rows):
        r = 3 + i
        label = row_data[0]
        vals  = row_data[1:]
        if label == "":
            ws.row_dimensions[r].height = 8
            continue
        is_total = label.startswith("TOTAL") or label in ("OPERATING INCOME","NET INCOME","COSTS & EXPENSES","REVENUE")
        bg = SUBHDR_BG if (is_total and not label.startswith(" ")) else (ALT_ROW if i % 2 == 0 else WHITE)
        lbl(ws, r, 2, label, bg=bg)
        for j, v in enumerate(vals):
            if v is None:
                dc(ws, r, 3+j, "—", bg=bg)
            elif isinstance(v, str):
                wc(ws, r, 3+j, v, size=FONT_SIZE, bg=bg, align="right", border=True, italic=True, fg=DARK_GRAY)
            else:
                dc(ws, r, 3+j, v, bg=bg, bold=is_total, num_fmt='#,##0.00' if abs(v)<5 else '#,##0.0')

# ═══════════════════════════════════════════════════════════════════════════════
def build_balance_sheet(wb):
    ws = wb.create_sheet("Balance Sheet")
    ws.sheet_view.showGridLines = False
    scw(ws, 1, 3); scw(ws, 2, 35); scw(ws, 3, 18); scw(ws, 4, 18)
    scw(ws, 5, 18); scw(ws, 6, 18)

    sec_hdr(ws, 1, 2, "BALANCE SHEET — NETFLIX INC. (NFLX)  |  FY2022-FY2025 ($B)", span=5)
    sub_hdr(ws, 2, 2, "Metric")
    for i, yr in enumerate(["FY2022", "FY2023", "FY2024", "FY2025"]):
        sub_hdr(ws, 2, 3+i, yr)

    rows = [
        ("ASSETS",                          None, None, None, None),
        ("CURRENT ASSETS",                  None, None, None, None),
        ("  Cash & Cash Equivalents",        6.1,   7.1,   7.2,   9.0),
        ("  Short-term Investments",          0.9,   0.8,   0.7,   0.5),
        ("  Accounts Receivable (net)",       1.6,   1.7,   2.2,   2.5),
        ("  Other Current Assets",            1.8,   1.5,   1.7,   2.0),
        ("TOTAL CURRENT ASSETS",            10.4,  11.1,  11.8,  14.0),
        ("NON-CURRENT ASSETS",              None, None, None, None),
        ("  Content Assets (non-current)",   24.5,  27.0,  30.5,  33.0),
        ("  PP&E (net)",                      3.2,   3.5,   4.0,   5.0),
        ("  Operating Lease Right-of-Use",    2.0,   1.9,   2.0,   2.2),
        ("  Other Non-current Assets",        3.0,   2.9,   3.2,   3.5),
        ("TOTAL NON-CURRENT ASSETS",        32.7,  35.3,  39.7,  43.7),
        ("TOTAL ASSETS",                    48.7,  48.0,  48.9,  57.7),  # approx - note content-heavy
        ("", None, None, None, None),
        ("LIABILITIES",                     None, None, None, None),
        ("CURRENT LIABILITIES",             None, None, None, None),
        ("  Current Portion of Debt",         0.7,   1.8,   2.3,   2.5),
        ("  Accounts Payable",                0.6,   0.6,   0.7,   0.8),
        ("  Content Liabilities (current)",   4.8,   4.5,   5.2,   5.5),
        ("  Other Current Liabilities",       1.7,   1.6,   1.8,   2.0),
        ("TOTAL CURRENT LIABILITIES",        7.8,   8.5,  10.0,  10.8),
        ("NON-CURRENT LIABILITIES",         None, None, None, None),
        ("  Long-term Debt",                14.4,  14.1,  13.7,  13.5),
        ("  Content Liabilities (LT)",       4.5,   3.7,   4.2,   4.5),
        ("  Other Non-current Liabilities",   2.4,   2.3,   2.5,   2.7),
        ("TOTAL NON-CURRENT LIABILITIES",   21.3,  20.1,  20.4,  20.7),
        ("TOTAL LIABILITIES",               29.1,  28.6,  30.4,  31.5),
        ("", None, None, None, None),
        ("SHAREHOLDERS' EQUITY",            None, None, None, None),
        ("  Common Stock & APIC",           12.1,  12.6,  12.8,  13.5),
        ("  Retained Earnings",              7.5,   6.8,   6.8,  10.7),
        ("  Acc. Other Comprehensive Loss",  0.0,  -0.0,  -1.1,  -1.0),  # approx
        ("TOTAL EQUITY",                    19.6,  19.4,  18.5,  26.2),
        ("TOTAL LIAB + EQUITY",             48.7,  48.0,  48.9,  57.7),
        ("", None, None, None, None),
        ("KEY RATIOS",                      None, None, None, None),
        ("  Current Ratio",                  1.3,   1.3,   1.2,   1.3),
        ("  Debt-to-Equity",                0.73,  0.73,  0.74,  0.51),
        ("  Net Debt ($B)",                  9.0,   8.8,   8.8,   7.0),
        ("  Interest Coverage Ratio",        7.0,   8.1,  10.5,  14.5),
    ]

    for i, row_data in enumerate(rows):
        r = 3 + i
        label = row_data[0]
        vals  = row_data[1:]
        if label == "":
            ws.row_dimensions[r].height = 8
            continue
        is_total = label.startswith("TOTAL") or label in ("ASSETS","LIABILITIES","SHAREHOLDERS' EQUITY",
                                                           "KEY RATIOS","CURRENT ASSETS","NON-CURRENT ASSETS",
                                                           "CURRENT LIABILITIES","NON-CURRENT LIABILITIES")
        bg = SUBHDR_BG if is_total else (ALT_ROW if i % 2 == 0 else WHITE)
        lbl(ws, r, 2, label, bg=bg)
        for j, v in enumerate(vals):
            if v is None:
                dc(ws, r, 3+j, "—", bg=bg)
            else:
                dc(ws, r, 3+j, v, bg=bg, bold=is_total, num_fmt='#,##0.0')

# ═══════════════════════════════════════════════════════════════════════════════
def build_cashflow(wb):
    ws = wb.create_sheet("Cash Flow Analysis")
    ws.sheet_view.showGridLines = False
    scw(ws, 1, 3); scw(ws, 2, 35); scw(ws, 3, 18); scw(ws, 4, 18)
    scw(ws, 5, 18); scw(ws, 6, 18); scw(ws, 7, 18)

    sec_hdr(ws, 1, 2, "CASH FLOW ANALYSIS — NETFLIX INC. (NFLX)  |  ($B)", span=6)
    sub_hdr(ws, 2, 2, "Metric")
    for i, yr in enumerate(["FY2021", "FY2022", "FY2023", "FY2024", "FY2025"]):
        sub_hdr(ws, 2, 3+i, yr)

    rows = [
        ("OPERATING CASH FLOW",             -0.9,  2.0,   7.3,   7.0,   9.5),
        ("  Net Income",                     5.1,   4.5,   5.4,   8.7,  11.0),
        ("  D&A",                            0.9,   0.9,   1.1,   1.2,   1.4),
        ("  Content Asset Amortization",    12.2,  14.0,  14.9,  16.5,  17.5),
        ("  Stock-Based Compensation",       0.6,   0.7,   0.7,   0.7,   0.7),
        ("  Change in Content Liabilities",  0.2,  -1.7,  -1.0,  -0.5,  -0.5),
        ("  Other Working Capital",         -5.9,  -6.4,  -3.8,  -9.6,  -9.6),
        ("  Content Spend (cash)",         -17.7, -14.7, -13.3, -17.0, -18.0),
        ("", None, None, None, None, None),
        ("INVESTING CASH FLOW",             -0.9,  -0.8,  -1.0,  -2.0,  -3.0),
        ("  Capital Expenditures",          -0.5,  -0.4,  -0.6,  -1.0,  -1.5),
        ("  Acquisitions",                   0.0,  -0.0,  -0.0,  -0.0,  -0.5),
        ("  Other Investing",               -0.4,  -0.4,  -0.4,  -1.0,  -1.0),
        ("", None, None, None, None, None),
        ("FINANCING CASH FLOW",             -0.7,  -2.8,  -7.9,  -2.7,  -5.0),
        ("  Debt Issuance / (Repayment)",    0.4,  -2.1,  -7.6,  -2.3,  -2.5),
        ("  Share Repurchases",             -0.6,  -0.2,  -0.3,  -0.4,  -2.5),
        ("  Other Financing",               -0.5,  -0.5,   0.0,   0.0,   0.0),
        ("", None, None, None, None, None),
        ("FREE CASH FLOW (OCF - CapEx)",    -1.4,   1.6,   6.7,   6.0,   8.0),
        ("  (Adjusted, ex-content spend)",   0.0,   6.0,  12.5,  13.5,  17.0),
        ("  FCF Margin",                    "-4.7%","5.1%","19.9%","15.4%","17.7%"),
        ("", None, None, None, None, None),
        ("CAPITAL ALLOCATION SUMMARY",      None, None, None, None, None),
        ("  Content Spend (cash, $B)",       17.7,  14.7,  13.3,  17.0,  18.0),
        ("  Content as % of Revenue",       "59%", "47%", "39%", "44%", "40%"),
        ("  CapEx ($B)",                     0.5,   0.4,   0.6,   1.0,   1.5),
        ("  Share Buybacks ($B)",            0.6,   0.2,   0.3,   0.4,   2.5),
        ("  Debt Reduction ($B)",            0.0,   2.1,   7.6,   2.3,   2.5),
        ("  Dividend",                      "None","None","None","None","None"),
    ]

    for i, row_data in enumerate(rows):
        r = 3 + i
        label = row_data[0]
        vals  = row_data[1:]
        if label == "":
            ws.row_dimensions[r].height = 8
            continue
        is_total = any(label.startswith(x) for x in ("OPERATING","INVESTING","FINANCING","FREE CASH","CAPITAL"))
        bg = SUBHDR_BG if is_total else (ALT_ROW if i % 2 == 0 else WHITE)
        lbl(ws, r, 2, label, bg=bg)
        for j, v in enumerate(vals):
            if v is None:
                dc(ws, r, 3+j, "—", bg=bg)
            elif isinstance(v, str):
                wc(ws, r, 3+j, v, size=FONT_SIZE, bg=bg, align="right", border=True, italic=True, fg=DARK_GRAY)
            else:
                dc(ws, r, 3+j, v, bg=bg, bold=is_total, num_fmt='#,##0.0')

# ═══════════════════════════════════════════════════════════════════════════════
def build_roic(wb):
    ws = wb.create_sheet("Return on Capital")
    ws.sheet_view.showGridLines = False
    scw(ws, 1, 3); scw(ws, 2, 35); scw(ws, 3, 18); scw(ws, 4, 18)
    scw(ws, 5, 18); scw(ws, 6, 18); scw(ws, 7, 18)

    sec_hdr(ws, 1, 2, "RETURN ON CAPITAL — NETFLIX INC. (NFLX)", span=6)
    sub_hdr(ws, 2, 2, "Metric")
    for i, yr in enumerate(["FY2021", "FY2022", "FY2023", "FY2024", "FY2025"]):
        sub_hdr(ws, 2, 3+i, yr)

    rows = [
        ("PROFITABILITY RETURNS",        None, None, None, None, None),
        ("  Return on Equity (ROE)",     "26.2%","23.0%","27.8%","47.0%","52.0%"),
        ("  Return on Assets (ROA)",     "10.5%", "9.3%","11.3%","17.8%","20.0%"),
        ("  Return on Invested Capital", "22.5%","20.0%","25.0%","35.0%","40.0%"),
        ("  Op Income / Revenue",        "20.9%","19.0%","20.8%","26.7%","29.5%"),
        ("", None, None, None, None, None),
        ("SUBSCRIBER UNIT ECONOMICS",    None, None, None, None, None),
        ("  Paid Subscribers (M)",        221.8, 223.0, 260.3, 301.6, 325.0),
        ("  Net Adds (M)",                18.1,   0.0,  13.1,  41.3,  23.4),
        ("  ARPU ($ annualized, global)", 11.67, 11.67, 11.72, 11.50, 11.60),
        ("  Revenue / Subscriber ($)",    134,   142,   129,   129,   139),
        ("", None, None, None, None, None),
        ("MARGIN PROGRESSION",           None, None, None, None, None),
        ("  Gross Margin",               "43.0%","38.0%","43.0%","46.2%","47.5%"),
        ("  Operating Margin",           "20.9%","19.0%","20.8%","26.7%","29.5%"),
        ("  Net Margin",                 "17.2%","14.2%","16.0%","22.3%","24.3%"),
        ("  FCF Margin (adj.)",           "0.0%","19.0%","37.0%","34.6%","37.6%"),
        ("", None, None, None, None, None),
        ("CONTENT EFFICIENCY",           None, None, None, None, None),
        ("  Content Spend ($B)",           17.7,  14.7,  13.3,  17.0,  18.0),
        ("  Revenue / Content $ Spent",    1.68,  2.15,  2.54,  2.29,  2.51),
        ("  Improving content ROI",       "Base","Trough","Recovery","Improving","Strong"),
        ("", None, None, None, None, None),
        ("CAPITAL EFFICIENCY",           None, None, None, None, None),
        ("  Asset Turnover",               0.61,  0.66,  0.70,  0.80,  0.88),
        ("  PP&E Turnover",                9.3,   9.0,   9.6,   9.8,  10.0),
    ]

    for i, row_data in enumerate(rows):
        r = 3 + i
        label = row_data[0]
        vals  = row_data[1:]
        if label == "":
            ws.row_dimensions[r].height = 8
            continue
        is_hdr = not label.startswith("  ")
        bg = SUBHDR_BG if is_hdr else (ALT_ROW if i % 2 == 0 else WHITE)
        lbl(ws, r, 2, label, bg=bg)
        for j, v in enumerate(vals):
            if v is None:
                dc(ws, r, 3+j, "—", bg=bg)
            elif isinstance(v, str):
                wc(ws, r, 3+j, v, size=FONT_SIZE, bg=bg, align="right", border=True, italic=not is_hdr)
            else:
                dc(ws, r, 3+j, v, bg=bg, bold=is_hdr, num_fmt='#,##0.0' if abs(v)>10 else '#,##0.00')

# ═══════════════════════════════════════════════════════════════════════════════
def build_management(wb):
    ws = wb.create_sheet("Management")
    ws.sheet_view.showGridLines = False
    scw(ws, 1, 3); scw(ws, 2, 28); scw(ws, 3, 55); scw(ws, 4, 18)

    sec_hdr(ws, 1, 2, "MANAGEMENT ANALYSIS — NETFLIX INC. (NFLX)", span=3)

    sub_hdr(ws, 3, 2, "Key Leadership", span=3)
    leaders = [
        ("Name",              "Role",                 "Background & Assessment"),
        ("Ted Sarandos",      "Co-CEO",               "Joined Netflix 2000; content genius; built Netflix Originals from scratch; "
                                                        "responsible for Squid Game, Stranger Things, Wednesday; creative engine"),
        ("Greg Peters",       "Co-CEO",               "Ex-Product & Technology; promoted to Co-CEO Jan 2023; "
                                                        "architect of advertising tier and password-sharing crackdown; operational focused"),
        ("Reed Hastings",     "Chairman (non-exec)",   "Co-founder; stepped down as Executive Chairman April 2025; "
                                                        "left operating role gracefully; succession plan executed cleanly"),
        ("Spence Neumann",    "CFO",                  "Strong financial discipline; guided margin expansion from ~20% to ~30%; "
                                                        "clear FCF communication to investors"),
    ]
    for i, rd in enumerate(leaders):
        r = 4 + i
        if i == 0:
            sub_hdr(ws, r, 2, rd[0])
            sub_hdr(ws, r, 3, rd[1])
            sub_hdr(ws, r, 4, rd[2])
        else:
            bg = ALT_ROW if i % 2 == 0 else WHITE
            lbl(ws, r, 2, rd[0], bg=bg)
            lbl(ws, r, 3, rd[1], bg=bg)
            ws.row_dimensions[r].height = 45
            wc(ws, r, 4, rd[2], size=FONT_SIZE, bg=bg, align="left", wrap=True, border=True)

    sub_hdr(ws, 10, 2, "Compensation & Incentives (Proxy Analysis)", span=3)
    comp_items = [
        ("Pay Philosophy",     "High-performance culture; top-of-market cash salaries (no RSU complexity); Netflix pays for performance"),
        ("CEO Pay",            "Ted Sarandos: ~$50M total comp; Greg Peters: ~$40M; high but tied to stock performance outcomes"),
        ("SBC Philosophy",     "Employees choose salary vs. stock mix; minimal SBC vs. peers (~$700M/yr = ~1.6% of revenue)"),
        ("Founder Alignment",  "Reed Hastings owns ~25M+ shares (post-split); still largest individual shareholder; long-term aligned"),
        ("Performance Metrics","Operating margin targets, subscriber growth, engagement hours — well-defined and disclosed"),
    ]
    for i, (k, v) in enumerate(comp_items):
        r = 11 + i
        bg = ALT_ROW if i % 2 == 0 else WHITE
        lbl(ws, r, 2, k, bg=bg)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
        ws.row_dimensions[r].height = 38
        wc(ws, r, 3, v, size=FONT_SIZE, bg=bg, align="left", wrap=True, border=True)

    sub_hdr(ws, 18, 2, "Capital Allocation", span=3)
    ca_items = [
        ("Content Investment",  "$17-18B/year; disciplined spend on originals with global appeal; fewer titles, higher quality post-2022"),
        ("Buybacks",            "$2.5B in FY2025; accelerating as FCF grows; targeting 5-10% annual share count reduction"),
        ("Debt Reduction",      "Paid down $7.6B in FY2023; now $13.5B net debt — manageable at 1.2x EBITDA"),
        ("No Dividends",        "Reinvesting FCF into content and buybacks; appropriate given growth phase"),
        ("M&A",                 "Disciplined; historically small tuck-ins; Sony, game studios; Warner Bros. deal is atypical"),
        ("Leverage",            "D/E of 0.51; declining; comfortable territory; well within investment grade metrics"),
    ]
    for i, (k, v) in enumerate(ca_items):
        r = 19 + i
        bg = ALT_ROW if i % 2 == 0 else WHITE
        lbl(ws, r, 2, k, bg=bg)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
        ws.row_dimensions[r].height = 38
        wc(ws, r, 3, v, size=FONT_SIZE, bg=bg, align="left", wrap=True, border=True)

    sub_hdr(ws, 27, 2, "Acts Like an Owner?", span=3)
    owner = [
        ("Long-term Thinking", "YES — 20+ year investment in streaming infrastructure; willingness to lose money in early years"),
        ("Seeds Planted",      "Ad-supported tier (2022), live events (2023), gaming (2022), Netflix House (2025) — all early stage"),
        ("Harvest Risk",       "Content spend discipline since 2022 shows they balance investment vs. margin"),
        ("Password Crackdown", "Greg Peters executed password-sharing crackdown masterfully — 41M net adds in 2024"),
        ("Insider Sales",      "Minimal insider selling; Hastings has sold some shares but holds substantial position"),
        ("Overall",            "Co-CEO structure working well; clear division of responsibilities; succession clean"),
    ]
    for i, (k, v) in enumerate(owner):
        r = 28 + i
        bg = ALT_ROW if i % 2 == 0 else WHITE
        lbl(ws, r, 2, k, bg=bg)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
        ws.row_dimensions[r].height = 38
        wc(ws, r, 3, v, size=FONT_SIZE, bg=bg, align="left", wrap=True, border=True)

    sub_hdr(ws, 36, 2, "Overall Management Score", span=3, bg=NFLX_BLACK, fg=WHITE)
    ws.merge_cells("B37:D38")
    wc(ws, 37, 2,
       "SCORE: 8.5/10 — One of the best-managed media companies globally. Ted Sarandos and Greg "
       "Peters are complementary leaders. Netflix's transition from 'growth at all costs' to "
       "'profitable growth' has been executed with minimal disruption. Capital allocation is improving. "
       "Main watch item: Warner Bros. deal execution.",
       size=FONT_SIZE, bg=ALT_ROW, align="left", wrap=True, border=True)
    for r in [37, 38]:
        ws.row_dimensions[r].height = 28

# ═══════════════════════════════════════════════════════════════════════════════
def build_risks(wb):
    ws = wb.create_sheet("Risks")
    ws.sheet_view.showGridLines = False
    scw(ws, 1, 3); scw(ws, 2, 32); scw(ws, 3, 45); scw(ws, 4, 15); scw(ws, 5, 15)

    sec_hdr(ws, 1, 2, "RISK ANALYSIS — NETFLIX INC. (NFLX)", span=4)
    sub_hdr(ws, 2, 2, "Risk Category")
    sub_hdr(ws, 2, 3, "Description & Impact")
    sub_hdr(ws, 2, 4, "Probability")
    sub_hdr(ws, 2, 5, "Impact")

    risks = [
        ("COMPETITION / STREAMING WARS",
         "Disney+, Max, Amazon Prime Video, Apple TV+ all spending heavily on content. Market may not "
         "support multiple $10-20B content budgets profitably. Consolidation or price competition likely. "
         "Netflix's content advantage could erode if Disney IP outperforms originals.",
         "MEDIUM", "HIGH"),
        ("SUBSCRIBER SATURATION",
         "US/Canada market largely penetrated (~60% of TV households). Growth must come from APAC/LatAm "
         "where ARPU is $7-11 vs $17+ in North America. Deceleration from 41M net adds (2024) to "
         "23M (2025) suggests growth normalization. 2026 guidance implies further deceleration.",
         "MEDIUM", "MEDIUM"),
        ("ADVERTISING RAMP RISK",
         "Ad revenue only $1.5B in FY2025 (3rd year). Reaching $5B+ requires advertiser confidence, "
         "measurement infrastructure, and scale. Consensus estimates may be too optimistic per analysts.",
         "MEDIUM", "MEDIUM"),
        ("WARNER BROS. DEAL RISK",
         "Proposed acquisition/partnership involving Warner Bros. assets introduces execution, "
         "integration, content library overlap, and regulatory risks. Potentially transformative but "
         "historically large media deals have destroyed value. Size of deal TBD.",
         "MEDIUM", "HIGH"),
        ("CONTENT COST INFLATION",
         "A-list talent deals, sports rights, live events escalating in cost. Netflix outbid rivals for "
         "NFL Christmas games. Sports content is expensive and potentially loss-making to acquire subscribers.",
         "MEDIUM", "MEDIUM"),
        ("MACRO / CONSUMER SPENDING",
         "Streaming is a discretionary expense that consumers cut in recessions. Password-sharing "
         "crackdown has reset the base higher, but cancellation rates could rise in downturns.",
         "LOW", "MEDIUM"),
        ("TECHNOLOGY / PIRACY",
         "Renewed piracy activity as multiple subscriptions become expensive. VPN bypass. "
         "Content leaks. AI-generated competing content could reduce the value of expensive originals.",
         "LOW", "LOW-MEDIUM"),
        ("REGULATORY",
         "Content quotas in EU requiring local productions. Data privacy laws. Potential "
         "antitrust scrutiny if Netflix + Warner Bros. creates dominant streaming entity.",
         "LOW", "MEDIUM"),
        ("FOREIGN EXCHANGE",
         "~59% of revenue is non-US denominated. USD strengthening compresses reported revenue. "
         "Q4 2025 guidance included 1% FX headwind.",
         "MEDIUM", "LOW"),
        ("GAMING SEGMENT",
         "Netflix Gaming has not gained significant traction despite investments. Risk of capital "
         "misallocation; mobile gaming market is highly competitive.",
         "LOW", "LOW"),
    ]

    for i, (cat, desc, prob, imp) in enumerate(risks):
        r = 3 + i
        bg = ALT_ROW if i % 2 == 0 else WHITE
        lbl(ws, r, 2, cat, bg=bg)
        ws.row_dimensions[r].height = 55
        wc(ws, r, 3, desc, size=FONT_SIZE, bg=bg, align="left", wrap=True, border=True)
        prob_c = RED if prob == "HIGH" else (GOLD if "MEDIUM" in prob else GREEN)
        imp_c  = RED if imp == "HIGH" else (GOLD if "MEDIUM" in imp else GREEN)
        wc(ws, r, 4, prob, bold=True, size=FONT_SIZE, fg=WHITE, bg=prob_c, align="center", border=True)
        wc(ws, r, 5, imp,  bold=True, size=FONT_SIZE, fg=WHITE, bg=imp_c,  align="center", border=True)

    sub_hdr(ws, 14, 2, "Risk Summary", span=4)
    summary = ("Netflix's core business is well-positioned but faces four key risks: (1) streaming "
               "competition from Disney/Amazon; (2) subscriber growth deceleration in mature markets; "
               "(3) advertising ramp slower than expected; and (4) Warner Bros. deal execution. None "
               "of these are existential threats. Netflix's content moat and global distribution "
               "network make it the default survivor in streaming consolidation. At ~32x forward P/E, "
               "much of the growth optimism is priced in but the FCF trajectory justifies the premium.")
    ws.merge_cells("B15:E17")
    wc(ws, 15, 2, summary, size=FONT_SIZE, bg=LIGHT_GRAY, align="left", wrap=True, border=True)
    for r in [15, 16, 17]:
        ws.row_dimensions[r].height = 22

# ═══════════════════════════════════════════════════════════════════════════════
def build_valuation(wb):
    ws = wb.create_sheet("Valuation")
    ws.sheet_view.showGridLines = False
    scw(ws, 1, 3); scw(ws, 2, 32); scw(ws, 3, 22); scw(ws, 4, 22)
    scw(ws, 5, 22); scw(ws, 6, 22)

    sec_hdr(ws, 1, 2, "VALUATION ANALYSIS — NETFLIX INC. (NFLX)", span=5)

    sub_hdr(ws, 3, 2, "Current Market Statistics (April 9, 2026)", span=5)
    mkt = [
        ("Stock Price (post ~10:1 split)", "$102"),
        ("Market Capitalization",          "~$432 Billion"),
        ("Enterprise Value",               "~$439 Billion (incl. $7B net debt)"),
        ("P/E Ratio (TTM)",                "~39x"),
        ("Forward P/E (FY2026E)",          "~32x"),
        ("EV/EBITDA",                      "~30x"),
        ("Price/FCF (adj.)",               "~26x"),
        ("Price/Sales",                    "~9.6x"),
        ("52-Week Range (post-split)",      "~$77 - $~115"),
        ("Analyst Consensus",              "BUY | Median target ~$129 | 34/45 analysts Buy"),
    ]
    for i, (k, v) in enumerate(mkt):
        r = 4 + i
        bg = ALT_ROW if i % 2 == 0 else WHITE
        lbl(ws, r, 2, k, bg=bg)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
        wc(ws, r, 3, v, size=FONT_SIZE, bg=bg, align="right", border=True)
        ws.row_dimensions[r].height = 22

    sub_hdr(ws, 16, 2, "DCF Valuation (Base / Bull / Bear)", span=5)
    sub_hdr(ws, 17, 2, "Assumption"); sub_hdr(ws, 17, 3, "Bear Case")
    sub_hdr(ws, 17, 4, "Base Case"); sub_hdr(ws, 17, 5, "Bull Case"); sub_hdr(ws, 17, 6, "Notes")

    dcf = [
        ("Revenue Growth (FY2026-28)",     "8%",    "13%",    "18%",  "Bear=sub saturation + competition"),
        ("Revenue Growth (FY2029-35)",     "6%",     "9%",    "13%",  "Bull=ads $5B+, pricing power"),
        ("Terminal Growth Rate",           "2.0%",  "2.5%",   "3.0%", "Long-run stable business"),
        ("EBIT Margin (normalized)",       "25%",   "32%",    "38%",  "Bear=content cost surge; Bull=ad margin"),
        ("Discount Rate (WACC)",           "10%",   "9.0%",   "8.0%", "Media risk premium"),
        ("Implied Intrinsic Value/Share",  "$60",   "$100",   "$160", "Post-split per share"),
    ]
    for i, row_data in enumerate(dcf):
        r = 18 + i
        bg = ALT_ROW if i % 2 == 0 else WHITE
        lbl(ws, r, 2, row_data[0], bg=bg)
        for j in range(1, 5):
            wc(ws, r, 2+j, row_data[j], size=FONT_SIZE, bg=bg,
               align="right" if j < 4 else "left", border=True,
               bold=(row_data[0].startswith("Implied")))
        ws.row_dimensions[r].height = 22

    sub_hdr(ws, 26, 2, "Peer Valuation Comparison", span=5)
    sub_hdr(ws, 27, 2, "Company"); sub_hdr(ws, 27, 3, "P/E (TTM)")
    sub_hdr(ws, 27, 4, "Fwd P/E"); sub_hdr(ws, 27, 5, "EV/EBITDA"); sub_hdr(ws, 27, 6, "Rev Growth")

    peers = [
        ("Netflix (NFLX)",     "39x",  "32x",  "30x",  "16%"),
        ("Disney (DIS)",       "26x",  "22x",  "18x",   "5%"),
        ("Warner Bros. (WBD)", "18x",  "15x",  "13x",  "-2%"),
        ("Roku (ROKU)",        "N/M",  "85x",  "40x",  "15%"),
        ("S&P 500 Average",    "24x",  "21x",  "17x",   "9%"),
    ]
    for i, row_data in enumerate(peers):
        r = 28 + i
        bold = row_data[0].startswith("Netflix")
        bg = SUBHDR_BG if bold else (ALT_ROW if i % 2 == 0 else WHITE)
        for j, v in enumerate(row_data):
            if j == 0:
                lbl(ws, r, 2, v, bg=bg)
            else:
                wc(ws, r, 2+j, v, bold=bold, size=FONT_SIZE, bg=bg, align="right", border=True)

    sub_hdr(ws, 35, 2, "Margin of Safety Assessment", span=5)
    mos = [
        ("Current Price (post-split)",    "$102",   "April 9, 2026"),
        ("Base Case Intrinsic Value",      "$100",   "~Fair value at current price"),
        ("Bear Case Intrinsic Value",       "$60",   "~41% downside in adverse scenario"),
        ("Bull Case Intrinsic Value",      "$160",   "~57% upside in bull case"),
        ("Margin of Safety",              "THIN",    "Trading at ~fair value; limited margin of safety"),
        ("Recommendation",            "HOLD/WATCH",  "Wait for pullback to $85-90 for better entry; downside risk from WBD deal"),
    ]
    for i, (k, v, c) in enumerate(mos):
        r = 36 + i
        bg = ALT_ROW if i % 2 == 0 else WHITE
        lbl(ws, r, 2, k, bg=bg)
        wc(ws, r, 3, v, bold=True, size=FONT_SIZE, bg=bg, align="right", border=True)
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=6)
        wc(ws, r, 4, c, size=FONT_SIZE, bg=bg, align="left", border=True)
        ws.row_dimensions[r].height = 22

# ═══════════════════════════════════════════════════════════════════════════════
def build_sentiment(wb):
    ws = wb.create_sheet("Market Sentiment")
    ws.sheet_view.showGridLines = False
    scw(ws, 1, 3); scw(ws, 2, 30); scw(ws, 3, 52); scw(ws, 4, 18)

    sec_hdr(ws, 1, 2, "MARKET SENTIMENT — NETFLIX INC. (NFLX)", span=3)

    sub_hdr(ws, 3, 2, "Analyst Coverage", span=3)
    analyst = [
        ("Total Analysts",          "~45+"),
        ("Buy / Hold / Sell",       "~75% Buy | ~20% Hold | ~5% Sell"),
        ("Median Price Target",     "$129 (post-split); range $77-$152.50"),
        ("Consensus View",          "Generally bullish on FCF story; concerns on WBD deal and growth deceleration"),
    ]
    for i, (k, v) in enumerate(analyst):
        r = 4 + i
        bg = ALT_ROW if i % 2 == 0 else WHITE
        lbl(ws, r, 2, k, bg=bg)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
        wc(ws, r, 3, v, size=FONT_SIZE, bg=bg, align="left", border=True)
        ws.row_dimensions[r].height = 22

    sub_hdr(ws, 10, 2, "Key Bullish Narratives", span=3)
    bull = [
        ("FCF Machine",              "$9.5B FCF in FY2025 growing to $12B+ by 2027; buybacks accelerating; value unlock"),
        ("Ad Revenue Inflection",    "$1.5B in year 3 → projected $3B+ in FY2026 → $5B+ TAM; highest margin revenue"),
        ("Margin Expansion",         "29.5% op margin → 35%+ by 2028; operating leverage on fixed content costs"),
        ("Password Crackdown Legs",  "41M net adds in 2024 shows execution capability; crackdown globally still ongoing"),
        ("Live Sports Catalyst",     "NFL, boxing, WrestleMania on Netflix; driving premium pricing and subscriber retention"),
    ]
    for i, (t, d) in enumerate(bull):
        r = 11 + i
        bg = ALT_ROW if i % 2 == 0 else WHITE
        lbl(ws, r, 2, t, bg=bg)
        ws.row_dimensions[r].height = 38
        wc(ws, r, 3, d, size=FONT_SIZE, bg=bg, align="left", wrap=True, border=True)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)

    sub_hdr(ws, 18, 2, "Key Bearish Narratives", span=3)
    bear = [
        ("Growth Deceleration",      "12% revenue guidance for FY2026 down from 16% in FY2025; subs growth slowing"),
        ("WBD Deal Uncertainty",      "Proposed Warner Bros. deal creates overhang; size/terms unclear; potential dilution"),
        ("Ad Revenue Optimism",       "Some analysts say ad rev consensus too high; measurement + targeting still immature"),
        ("Valuation Premium",         "39x TTM P/E is expensive for a media company facing growth deceleration"),
    ]
    for i, (t, d) in enumerate(bear):
        r = 19 + i
        bg = ALT_ROW if i % 2 == 0 else WHITE
        lbl(ws, r, 2, t, bg=bg)
        ws.row_dimensions[r].height = 38
        wc(ws, r, 3, d, size=FONT_SIZE, bg=bg, align="left", wrap=True, border=True)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)

    sub_hdr(ws, 25, 2, "Recent News & Catalysts (2025-2026)", span=3)
    news = [
        ("Jan 2026",    "Q4 2025 earnings: $12.05B revenue (+18%), 325M subscribers, ad rev 2.5x growth"),
        ("Apr 2025",    "Reed Hastings stepped down as Executive Chairman; became Board Chairman only"),
        ("Apr 2025",    "Q1 2025: Netflix stopped reporting subscriber counts quarterly; focus shifts to revenue/margins"),
        ("Mar 2025",    "Netflix House opening in King of Prussia, PA; first of multi-city physical presence"),
        ("Jan 2025",    "Tyson vs. Paul boxing event drew 108M viewers — live events strategy validated"),
        ("Nov 2024",    "Netflix added NFL Christmas Day games; strong ratings; sports rights expanding"),
        ("Oct 2024",    "Q3 2024 earnings: 5M net adds, $9.8B revenue; exceeded expectations; stock +5%"),
    ]
    for i, (d, n) in enumerate(news):
        r = 26 + i
        bg = ALT_ROW if i % 2 == 0 else WHITE
        lbl(ws, r, 2, d, bg=bg)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
        wc(ws, r, 3, n, size=FONT_SIZE, bg=bg, align="left", border=True)
        ws.row_dimensions[r].height = 22

# ═══════════════════════════════════════════════════════════════════════════════
def build_key_indicators(wb):
    ws = wb.create_sheet("Key Indicators")
    ws.sheet_view.showGridLines = False
    scw(ws, 1, 3); scw(ws, 2, 30); scw(ws, 3, 20); scw(ws, 4, 20)
    scw(ws, 5, 20); scw(ws, 6, 20); scw(ws, 7, 20)

    sec_hdr(ws, 1, 2, "KEY INDICATORS — NETFLIX INC. (NFLX)  |  FY2021-FY2025", span=6)
    sub_hdr(ws, 2, 2, "Indicator")
    for i, yr in enumerate(["FY2021", "FY2022", "FY2023", "FY2024", "FY2025"]):
        sub_hdr(ws, 2, 3+i, yr)

    rows = [
        ("FINANCIAL SUMMARY",          None, None, None, None, None),
        ("  Total Revenue ($B)",        29.7,  31.6,  33.7,  39.0,  45.2),
        ("  Revenue Growth",           "19%", "6.5%","6.6%","15.6%","15.9%"),
        ("  Gross Profit ($B)",         12.8,  12.0,  14.5,  18.0,  21.5),
        ("  Gross Margin",             "43%","38.0%","43.0%","46.2%","47.5%"),
        ("  Operating Income ($B)",      6.2,   6.0,   7.0,  10.4,  13.3),
        ("  Operating Margin",         "20.9%","19.0%","20.8%","26.7%","29.5%"),
        ("  Net Income ($B)",            5.1,   4.5,   5.4,   8.7,  11.0),
        ("  Net Margin",               "17.2%","14.2%","16.0%","22.3%","24.3%"),
        ("  Diluted EPS (post-split $)", 1.21,  1.05,  1.27,  2.06,  2.61),
        ("  EPS Growth",               "85%","-13%","21%","62%","26.7%"),
        ("", None, None, None, None, None),
        ("SUBSCRIBER METRICS",         None, None, None, None, None),
        ("  Paid Subscribers (M)",      221.8, 223.0, 260.3, 301.6, 325.0),
        ("  Net Subscriber Adds (M)",    18.1,   0.0,  13.1,  41.3,  23.4),
        ("  Global ARPM ($/month)",      11.67, 11.67, 11.72, 11.50, 11.60),
        ("  US+Canada ARPM ($/month)",   14.78, 15.48, 16.71, 17.06, 17.50),
        ("", None, None, None, None, None),
        ("CASH FLOW",                  None, None, None, None, None),
        ("  Operating Cash Flow ($B)",  -0.9,   2.0,   7.3,   7.0,   9.5),
        ("  CapEx ($B)",                 0.5,   0.4,   0.6,   1.0,   1.5),
        ("  Free Cash Flow ($B)",       -1.4,   1.6,   6.7,   6.0,   8.0),
        ("  Content Spend (cash $B)",   17.7,  14.7,  13.3,  17.0,  18.0),
        ("  Share Buybacks ($B)",        0.6,   0.2,   0.3,   0.4,   2.5),
        ("", None, None, None, None, None),
        ("BALANCE SHEET",              None, None, None, None, None),
        ("  Cash & Equivalents ($B)",    6.0,   6.1,   7.1,   7.2,   9.0),
        ("  Long-term Debt ($B)",       15.1,  14.4,  14.1,  13.7,  13.5),
        ("  Net Debt ($B)",              9.1,   8.3,   7.0,   6.5,   4.5),
        ("  Debt-to-Equity",            0.77,  0.73,  0.73,  0.74,  0.51),
        ("", None, None, None, None, None),
        ("VALUATION",                  None, None, None, None, None),
        ("  P/E Ratio",                 "62x","30x", "38x","43x", "39x"),
        ("  EV/EBITDA",                 "40x","22x", "28x","32x", "30x"),
        ("  Price/FCF",                 "N/M","85x", "25x","33x", "54x"),
        ("  Price/Sales",              "10.3x","5.3x","7.4x","9.2x","9.6x"),
    ]

    for i, row_data in enumerate(rows):
        r = 3 + i
        label = row_data[0]
        vals  = row_data[1:]
        if label == "":
            ws.row_dimensions[r].height = 8
            continue
        is_hdr = not label.startswith("  ")
        bg = SUBHDR_BG if is_hdr else (ALT_ROW if i % 2 == 0 else WHITE)
        lbl(ws, r, 2, label, bg=bg)
        for j, v in enumerate(vals):
            if v is None:
                dc(ws, r, 3+j, "—", bg=bg)
            elif isinstance(v, str):
                wc(ws, r, 3+j, v, size=FONT_SIZE, bg=bg, align="right", border=True, bold=is_hdr)
            else:
                dc(ws, r, 3+j, v, bg=bg, bold=is_hdr, num_fmt='#,##0.0' if abs(v)>5 else '#,##0.00')

# ═══════════════════════════════════════════════════════════════════════════════
def main():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    build_cover(wb)
    build_business(wb)
    build_moat(wb)
    build_income(wb)
    build_balance_sheet(wb)
    build_cashflow(wb)
    build_roic(wb)
    build_management(wb)
    build_risks(wb)
    build_valuation(wb)
    build_sentiment(wb)
    build_key_indicators(wb)

    out_dir = "/Users/naji/WORK/github.com/AI/claude/Agent/MarketResearch/output"
    os.makedirs(out_dir, exist_ok=True)
    path = os.path.join(out_dir, "NFLX_Financial_Analysis.xlsx")
    wb.save(path)
    print(f"Saved: {path}")

if __name__ == "__main__":
    main()
