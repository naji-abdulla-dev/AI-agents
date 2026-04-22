"""
ESCO Technologies Inc. (NYSE: ESE) - Comprehensive Financial Analysis
Generated: April 20, 2026
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT_PATH = "/Users/naji/WORK/github.com/AI/claude/Agent/MarketResearch/output/ESE_Financial_Analysis.xlsx"

# ── colour palette ──────────────────────────────────────────────────────────
DARK_NAVY   = "0D1B2A"
MID_NAVY    = "1B3A5C"
LIGHT_NAVY  = "C5D8EC"
ACCENT_GOLD = "C9A84C"
DARK_GREEN  = "1E5631"
LIGHT_GREEN = "D9EAD3"
RED         = "C00000"
LIGHT_RED   = "FFE0E0"
GREY        = "F2F2F2"
MID_GREY    = "D9D9D9"
WHITE       = "FFFFFF"
ORANGE      = "E06C00"
LIGHT_ORANGE= "FFF0DC"

FONT_SIZE = 14

def font(bold=False, size=FONT_SIZE, color=None, italic=False):
    kw = {"name": "Calibri", "bold": bold, "size": size, "italic": italic}
    if color:
        kw["color"] = color
    return Font(**kw)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def thin_border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def center(wrap=True):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def left(wrap=True):
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)

def right_al():
    return Alignment(horizontal="right", vertical="center", wrap_text=False)

def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

def merge_write(ws, rng, value, bold=False, size=FONT_SIZE,
                bg=None, fg=WHITE, align=None, italic=False, num_fmt=None):
    ws.merge_cells(rng)
    cell = ws[rng.split(":")[0]]
    cell.value = value
    cell.font = font(bold=bold, size=size, color=fg, italic=italic)
    if bg:
        cell.fill = fill(bg)
    cell.alignment = align or center()
    if num_fmt:
        cell.number_format = num_fmt
    return cell

def wc(ws, row, col, value, bold=False, size=FONT_SIZE,
        bg=None, fg=None, align=None, num_fmt=None, italic=False, border=False):
    cell = ws.cell(row=row, column=col, value=value)
    default_fg = "000000" if not bg else WHITE
    cell.font = font(bold=bold, size=size, color=fg or default_fg, italic=italic)
    if bg:
        cell.fill = fill(bg)
    cell.alignment = align or left()
    if num_fmt:
        cell.number_format = num_fmt
    if border:
        cell.border = thin_border()
    return cell

def header_row(ws, row, cols_values, bg=DARK_NAVY, fg=WHITE, height=30):
    ws.row_dimensions[row].height = height
    for col, val in enumerate(cols_values, 1):
        wc(ws, row, col, val, bold=True, bg=bg, fg=fg, align=center(), size=FONT_SIZE)

def sec_hdr(ws, row, c1, c2, title, bg=MID_NAVY, fg=WHITE, height=28):
    rng = f"{get_column_letter(c1)}{row}:{get_column_letter(c2)}{row}"
    merge_write(ws, rng, title, bold=True, bg=bg, fg=fg, size=FONT_SIZE+1, align=center())
    ws.row_dimensions[row].height = height

def data_row(ws, row, c_start, values, even=True, num_fmts=None, bold_first=True):
    bg = GREY if even else WHITE
    for i, val in enumerate(values):
        col = c_start + i
        fmt = num_fmts[i] if num_fmts and i < len(num_fmts) else None
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = font(size=FONT_SIZE, bold=(i == 0 and bold_first))
        cell.fill = fill(bg)
        cell.alignment = left() if i == 0 else center()
        if fmt:
            cell.number_format = fmt
        cell.border = thin_border()
    ws.row_dimensions[row].height = 22


# ════════════════════════════════════════════════════════════════════════════
# TAB 1 – COVER
# ════════════════════════════════════════════════════════════════════════════
def build_cover(wb):
    ws = wb.create_sheet("1. Cover")
    ws.sheet_view.showGridLines = False
    for col in range(1, 7):
        set_col_width(ws, col, 26)
    for r in range(1, 70):
        ws.row_dimensions[r].height = 18

    for r in range(1, 8):
        for c in range(1, 7):
            ws.cell(r, c).fill = fill(DARK_NAVY)

    merge_write(ws, "A1:F7",
                "ESCO TECHNOLOGIES INC. (NYSE: ESE)\nComprehensive Financial Analysis",
                bold=True, size=28, bg=DARK_NAVY, fg=WHITE, align=center())

    for c in range(1, 7):
        ws.cell(8, c).fill = fill(ACCENT_GOLD)
    merge_write(ws, "A8:F8", "Prepared: April 20, 2026  |  Research Mode  |  Think Like an Owner",
                bold=True, size=FONT_SIZE, bg=ACCENT_GOLD, fg=WHITE, align=center())

    rows = [
        (10, "COMPANY",          "ESCO Technologies Inc."),
        (11, "TICKER",           "ESE (NYSE)"),
        (12, "SECTOR",           "Industrials — Highly Engineered Products"),
        (13, "HEADQUARTERS",     "St. Louis, Missouri, USA"),
        (14, "FOUNDED",          "1990 (spun off from Emerson Electric)"),
        (15, "CEO",              "Bryan H. Sayler (President & CEO)"),
        (16, "CFO",              "Christopher L. Tucker (SVP & CFO)"),
        (17, "FISCAL YEAR END",  "September 30"),
        (19, "CURRENT PRICE",    "~$215 (est., Apr 2026)"),
        (20, "MARKET CAP",       "~$5.6 Billion"),
        (21, "52-WEEK RANGE",    "~$130 – $240"),
        (22, "YTD PERFORMANCE",  "+50% Year-to-Date (as of Apr 2026)"),
        (23, "ANALYST RATING",   "Strong Buy | Median PT: $300 | Range: $270–$350"),
        (25, "FY2025 REVENUE",   "$1,095M (+19.2% YoY)"),
        (26, "FY2025 ADJ EBIT",  "$222.4M (20.3% margin)"),
        (27, "FY2025 NET INCOME","$116.3M (continuing ops)"),
        (28, "FY2025 FCF",       "~$164M (normalized, after $36M CapEx)"),
        (29, "FY2026 GUIDANCE",  "Revenue $1.29–1.33B | Adj EPS $7.90–$8.15"),
        (30, "Q1 FY2026 ORDERS", "$557M (+143% YoY) | Backlog: $1.4B (record)"),
        (31, "MEGGER DEAL",      "Acquiring Megger Group for $2.35B (Apr 15, 2026)"),
    ]

    for (r, label, val) in rows:
        wc(ws, r, 1, label, bold=True, bg=MID_NAVY, fg=WHITE, align=center(), size=FONT_SIZE)
        ws.merge_cells(f"B{r}:F{r}")
        wc(ws, r, 2, val, bold=False, bg=LIGHT_NAVY, fg="000000", align=left(), size=FONT_SIZE)
        ws.row_dimensions[r].height = 22

    sec_hdr(ws, 33, 1, 6, "INVESTMENT THESIS", bg=DARK_GREEN)
    thesis = (
        "ESCO Technologies is a disciplined compounder of highly engineered, mission-critical products "
        "serving three defensible end-markets: (1) Aerospace & Defense – ~44% of revenue, driven by "
        "long-cycle U.S. Navy submarine & surface ship programs with a record $1.4B backlog providing "
        "multi-year revenue visibility; (2) Utility Solutions Group (USG) – ~35% of revenue via Doble "
        "Engineering's embedded grid-testing tools used by virtually every major utility in North America; "
        "and (3) RF Test & Measurement – ~21% of revenue through ETS-Lindgren's EMC chambers. The company "
        "has deployed capital well: the $472M SM&P acquisition (April 2025) added Navy Maritime revenue, "
        "and the pending $2.35B Megger acquisition (April 2026) will more than double USG and create a "
        "global utility-testing powerhouse. Free cash flow is healthy ($164M normalized FY2025) and "
        "management's incentives are aligned with EPS + operating cash flow. Key risks: post-Megger "
        "leverage (~$1.5B new debt), integration execution, defense budget uncertainty, and stock "
        "re-rating risk after 50% YTD gain."
    )
    ws.row_dimensions[34].height = 160
    merge_write(ws, "A34:F41", thesis, bold=False, size=FONT_SIZE,
                bg=LIGHT_GREEN, fg="000000",
                align=Alignment(horizontal="left", vertical="center", wrap_text=True))

    for c in range(1, 7):
        ws.cell(43, c).fill = fill(DARK_GREEN)
    merge_write(ws, "A43:F43",
                "RATING:  HOLD/BUY  |  12-Month Target: $250  |  Fair-Value DCF: $178–$225  |  Megger Optionality Upside",
                bold=True, size=FONT_SIZE+2, bg=DARK_GREEN, fg=WHITE, align=center())


# ════════════════════════════════════════════════════════════════════════════
# TAB 2 – BUSINESS OVERVIEW
# ════════════════════════════════════════════════════════════════════════════
def build_business(wb):
    ws = wb.create_sheet("2. Business Overview")
    ws.sheet_view.showGridLines = False
    widths = [32, 22, 18, 18, 18, 18, 18]
    for i, w in enumerate(widths, 1):
        set_col_width(ws, i, w)

    merge_write(ws, "A1:G1", "ESCO TECHNOLOGIES — BUSINESS OVERVIEW", bold=True,
                size=18, bg=DARK_NAVY, fg=WHITE, align=center())
    ws.row_dimensions[1].height = 35

    # ── Company Description ──
    sec_hdr(ws, 3, 1, 7, "COMPANY DESCRIPTION", bg=MID_NAVY)
    desc = (
        "ESCO Technologies Inc. (NYSE: ESE) designs and manufactures highly engineered products and solutions "
        "for niche industrial and defense markets. Founded in 1990 as a spin-off from Emerson Electric, the "
        "company has transformed from a diversified manufacturer into a focused industrial compounder with three "
        "core segments. Revenue has compounded at ~8% annually over the past decade, with accelerating growth "
        "following strategic acquisitions in the Navy/defense space and now the $2.35B Megger deal. The company "
        "serves the U.S. Department of Defense (primarily Navy), global utilities, and commercial aerospace clients."
    )
    ws.row_dimensions[4].height = 100
    merge_write(ws, "A4:G7", desc, bold=False, size=FONT_SIZE, bg=LIGHT_NAVY, fg="000000",
                align=Alignment(horizontal="left", vertical="center", wrap_text=True))

    # ── Segment Revenue ──
    sec_hdr(ws, 9, 1, 7, "SEGMENT REVENUE BREAKDOWN (FY2025 vs FY2024, $M)", bg=MID_NAVY)
    header_row(ws, 10, ["Segment", "FY2024 Rev ($M)", "FY2025 Rev ($M)", "YoY Growth",
                         "% of FY2025", "Adj EBIT Margin", "Key Brands"], height=28)
    seg_data = [
        ["Aerospace & Defense (A&D)", "$341M", "$478M", "+40.4%", "43.7%", "~26-29%",
         "ESCO Maritime Solutions, Doble Lemke, Westland"],
        ["Utility Solutions Group (USG)", "$372M", "$383M", "+2.9%", "35.0%", "~28-30%",
         "Doble Engineering, NRG Systems, Trayer"],
        ["RF Test & Measurement (Test)", "$206M", "$234M", "+13.6%", "21.3%", "~14-16%",
         "ETS-Lindgren, ESCO Japan"],
        ["TOTAL", "$919M", "$1,095M", "+19.2%", "100%", "20.3% (overall)", ""],
    ]
    for i, row in enumerate(seg_data):
        is_total = i == len(seg_data) - 1
        bg = DARK_NAVY if is_total else (GREY if i % 2 == 0 else WHITE)
        fg_c = WHITE if is_total else "000000"
        for j, val in enumerate(row):
            cell = ws.cell(row=11+i, column=j+1, value=val)
            cell.font = font(size=FONT_SIZE, bold=is_total, color=fg_c)
            cell.fill = fill(bg)
            cell.alignment = center() if j > 0 else left()
            cell.border = thin_border()
        ws.row_dimensions[11+i].height = 22

    # ── Products & Value Propositions ──
    sec_hdr(ws, 16, 1, 7, "PRODUCTS, CLIENTS & VALUE PROPOSITIONS", bg=MID_NAVY)
    prod_hdr = ["Segment", "Core Products / Services", "Key Clients", "Value Proposition",
                "Buying Process", "Seasonality", "Revenue Model"]
    header_row(ws, 17, prod_hdr, height=28)
    prod_data = [
        ["A&D", "Submarine power systems, EMI filters, naval electronics, filtration, maritime command solutions",
         "U.S. Navy (primary), U.S. Army, Commercial Aerospace OEMs",
         "Mission-critical; no substitute; designed into multi-decade platform programs",
         "Long-cycle DoD procurement; 5-10yr contracts; low price sensitivity",
         "Relatively even; multi-year contracts smooth lumpy orders",
         "Product + services; gov't cost-plus and fixed-price"],
        ["USG (Doble)", "Transformer testing, asset management diagnostics, grid condition monitoring, wind sensors",
         "~80% of U.S. electric utilities; global utilities; renewable energy operators",
         "Reduces grid failure risk; embedded in utility maintenance schedules; sticky SaaS-like recurring revenue",
         "Utility capital budget cycle; long vendor approval; high switching costs post-integration",
         "Slight bias toward H2 (utility CapEx often back-half loaded)",
         "Hardware + recurring service/software subscriptions"],
        ["Test (ETS-Lindgren)", "Anechoic chambers, RF shielding, EMC test systems, MRI suite shielding",
         "Aerospace OEMs, telecom, medical device manufacturers, U.S. government labs",
         "FAA/FCC/MIL-STD compliance required; certified chambers are expensive to replace",
         "Project-based bids; often 1-2yr lead times for large chamber installations",
         "Uneven by project; not strongly seasonal",
         "Large capital project + service contracts"],
    ]
    for i, row in enumerate(prod_data):
        bg = GREY if i % 2 == 0 else WHITE
        for j, val in enumerate(row):
            cell = ws.cell(row=18+i, column=j+1, value=val)
            cell.font = font(size=FONT_SIZE)
            cell.fill = fill(bg)
            cell.alignment = left()
            cell.border = thin_border()
        ws.row_dimensions[18+i].height = 60

    # ── Geographic Mix ──
    sec_hdr(ws, 22, 1, 7, "REVENUE GEOGRAPHY (FY2025 Estimates)", bg=MID_NAVY)
    header_row(ws, 23, ["Geography", "Est. Revenue", "% of Total", "Key Drivers", "", "", ""], height=24)
    geo_data = [
        ["United States",    "$820M",  "~75%", "DoD/Navy, U.S. utilities, domestic EMC testing", "", "", ""],
        ["Europe",           "$175M",  "~16%", "Doble Europe, ETS-Lindgren EU, export defense", "", "", ""],
        ["Asia-Pacific",     "$55M",   "~5%",  "ESCO Japan (Test), Asia utility growth", "", "", ""],
        ["Rest of World",    "$45M",   "~4%",  "Middle East defense, Latin America utilities", "", "", ""],
        ["POST-MEGGER (est)","~$1.9B", "–",    "Megger adds $590M; significant UK/Europe expansion", "", "", ""],
    ]
    for i, row in enumerate(geo_data):
        bg = LIGHT_ORANGE if i == 4 else (GREY if i % 2 == 0 else WHITE)
        for j, val in enumerate(row):
            cell = ws.cell(row=24+i, column=j+1, value=val)
            cell.font = font(size=FONT_SIZE, bold=(i == 4))
            cell.fill = fill(bg)
            cell.alignment = center() if j in [1, 2] else left()
            cell.border = thin_border()
        ws.row_dimensions[24+i].height = 22

    # ── Megger Acquisition Note ──
    sec_hdr(ws, 30, 1, 7, "MEGGER ACQUISITION — STRATEGIC CONTEXT (Apr 15, 2026)", bg=ORANGE)
    meg_text = (
        "ESCO agreed to acquire Megger Group Limited (from TBG AG) for $2.35B = $922M cash + 5.1M ESCO shares. "
        "Megger is a leading global provider of electrical testing & monitoring solutions for utilities and critical "
        "infrastructure, with projected 2026 revenue of ~$590M. The deal values Megger at ~14x 2026 EBITDA including "
        "$60M in synergies (3-year target). Financing: $1.5B JPMorgan committed credit facility + existing cash. "
        "Pro-forma leverage rises to ~4x Net Debt/EBITDA, a significant step-up from <0.3x pre-deal. "
        "Post-close, ESCO becomes a ~$1.9B revenue global industrial technology company with utility testing dominance."
    )
    ws.row_dimensions[31].height = 110
    merge_write(ws, "A31:G35", meg_text, bold=False, size=FONT_SIZE, bg=LIGHT_ORANGE, fg="000000",
                align=Alignment(horizontal="left", vertical="center", wrap_text=True))


# ════════════════════════════════════════════════════════════════════════════
# TAB 3 – MOAT
# ════════════════════════════════════════════════════════════════════════════
def build_moat(wb):
    ws = wb.create_sheet("3. Moat")
    ws.sheet_view.showGridLines = False
    for i, w in enumerate([30, 20, 15, 15, 40], 1):
        set_col_width(ws, i, w)

    merge_write(ws, "A1:E1", "ESCO TECHNOLOGIES — COMPETITIVE MOAT ANALYSIS", bold=True,
                size=18, bg=DARK_NAVY, fg=WHITE, align=center())
    ws.row_dimensions[1].height = 35

    sec_hdr(ws, 3, 1, 5, "MOAT SCORECARD", bg=MID_NAVY)
    header_row(ws, 4, ["Moat Source", "Segment", "Strength (1-5)", "Durability", "Evidence / Commentary"], height=28)
    moat_data = [
        ["Long-Cycle Defense Contracts", "A&D", "★★★★★", "10-20 years",
         "ESCO systems are designed into U.S. Navy submarine programs (Columbia, Virginia class) — switching is contractually & technically prohibitive. Record $1.4B backlog provides multi-year visibility."],
        ["Switching Costs — Utility Testing", "USG", "★★★★☆", "7-15 years",
         "Doble's diagnostics are embedded in utility asset-management workflows. Utilities train engineers on Doble equipment; historical test data resides in Doble platforms. Very high cost to re-certify with a competitor."],
        ["Regulatory & Certification Barriers", "Test", "★★★★☆", "5-10 years",
         "ETS-Lindgren chambers must meet FAA, FCC, MIL-STD specs. Competitors require extensive re-certification to displace a certified chamber. Specialized installation teams create further lock-in."],
        ["Engineering Depth & IP", "All", "★★★★☆", "Ongoing",
         "Deep application engineering (e.g., submarine EMI filtering tolerances, MRI shielding). Proprietary designs not easily reverse-engineered. Ongoing R&D: $23M/year."],
        ["Customer Relationships & Sole-Source Status", "A&D", "★★★★★", "Program life",
         "Many Navy programs have ESCO as sole-qualified supplier. Qualification is multi-year; no viable alternative exists mid-program. Similar dynamic in some Doble utility contracts."],
        ["Scale & Global Installed Base", "USG + Test", "★★★☆☆", "5-10 years",
         "Doble serves ~80% of U.S. utilities; ETS-Lindgren chambers in most major aerospace/defense test facilities. Creates reference-sale network effects."],
        ["Cost Advantages (Niche Focus)", "All", "★★★☆☆", "Moderate",
         "ESCO is not the lowest-cost producer but competes on value/reliability. Margins are protected by mission-criticality, not volume scale. Smaller niche players lack breadth."],
    ]
    for i, row in enumerate(moat_data):
        bg = GREY if i % 2 == 0 else WHITE
        for j, val in enumerate(row):
            cell = ws.cell(row=5+i, column=j+1, value=val)
            cell.font = font(size=FONT_SIZE, bold=(j == 0))
            cell.fill = fill(bg)
            cell.alignment = left()
            cell.border = thin_border()
        ws.row_dimensions[5+i].height = 55

    sec_hdr(ws, 13, 1, 5, "MOAT vs. PEERS COMPARISON", bg=MID_NAVY)
    header_row(ws, 14, ["Company", "Segment Overlap", "Est. EBIT Margin", "Revenue Model", "Key Differentiator vs ESE"], height=28)
    peer_data = [
        ["ESCO Technologies (ESE)", "All three", "~20% adj", "Product + Service", "Sole-source DoD; Doble utility dominance; post-Megger scale"],
        ["VIAVI Solutions (VIAV)", "Test (partial)", "~14%", "Product + SaaS", "Fiber/network test; less defense; no utility exposure"],
        ["Keysight Technologies (KEYS)", "Test", "~25%", "Product + SW", "Electronic test breadth; less niche defense; 10x larger"],
        ["Roper Technologies (ROP)", "USG (adjacent)", "~35%", "Software/SaaS", "Higher margins but much higher valuation; software-heavy"],
        ["Watts Water (WTS)", "USG (partial)", "~18%", "Product", "Water/flow control; no defense; commoditized"],
        ["Teledyne Technologies (TDY)", "A&D (overlap)", "~23%", "Product + Service", "Broader defense sensors; more overlap but less Navy-specific"],
        ["Megger Group (post-acquisition)", "USG (complement)", "~18-20%", "Product + Service", "Utility testing globally; becomes core after acquisition"],
    ]
    for i, row in enumerate(peer_data):
        is_ese = i == 0
        bg = LIGHT_NAVY if is_ese else (GREY if i % 2 == 1 else WHITE)
        for j, val in enumerate(row):
            cell = ws.cell(row=15+i, column=j+1, value=val)
            cell.font = font(size=FONT_SIZE, bold=is_ese)
            cell.fill = fill(bg)
            cell.alignment = center() if j in [1,2,3] else left()
            cell.border = thin_border()
        ws.row_dimensions[15+i].height = 28

    sec_hdr(ws, 23, 1, 5, "MOAT VERDICT", bg=DARK_GREEN)
    verdict = (
        "ESCO has a NARROW-TO-WIDE economic moat driven primarily by switching costs and sole-source "
        "defense positioning. The A&D segment has the strongest moat — no rational Navy program manager "
        "risks submarine readiness to save cost on a niche supplier. USG/Doble has sticky recurring revenue "
        "from embedded diagnostics. Test is the weakest moat but benefits from certification barriers. "
        "Post-Megger, the USG moat widens considerably: combined Doble + Megger would create a global "
        "utility-testing duopoly with significant pricing power and cross-selling opportunities. "
        "Management's record backlog ($1.4B) and 1.92x book-to-bill are the clearest signals of moat strength."
    )
    ws.row_dimensions[24].height = 140
    merge_write(ws, "A24:E29", verdict, bold=False, size=FONT_SIZE, bg=LIGHT_GREEN, fg="000000",
                align=Alignment(horizontal="left", vertical="center", wrap_text=True))


# ════════════════════════════════════════════════════════════════════════════
# TAB 4 – INCOME STATEMENTS
# ════════════════════════════════════════════════════════════════════════════
def build_income(wb):
    ws = wb.create_sheet("4. Income Statements")
    ws.sheet_view.showGridLines = False
    widths = [30, 16, 16, 16, 16, 16, 16]
    for i, w in enumerate(widths, 1):
        set_col_width(ws, i, w)

    merge_write(ws, "A1:G1", "ESCO TECHNOLOGIES — INCOME STATEMENT ANALYSIS", bold=True,
                size=18, bg=DARK_NAVY, fg=WHITE, align=center())
    ws.row_dimensions[1].height = 35

    years = ["FY2021E", "FY2022E", "FY2023E", "FY2024A", "FY2025A", "FY2026E"]
    sec_hdr(ws, 3, 1, 7, "CONSOLIDATED INCOME STATEMENT ($M, FY ends Sep 30)", bg=MID_NAVY)
    header_row(ws, 4, ["Metric"] + years, height=28)

    usd = '#,##0.0'
    pct = '0.0%'
    inc_data = [
        ("Revenue",                     742.0, 795.0, 858.0,  919.0,  1095.0, 1310.0),
        ("YoY Growth",                  None,  "+7.1%","+7.9%","+7.1%","+19.2%","+19.6%"),
        ("Cost of Goods Sold",          -452,  -484,  -516,   -542,   -634,   -762),
        ("Gross Profit",                290,   311,   342,    377,    461,    548),
        ("Gross Margin %",              "39.1%","39.1%","39.9%","41.0%","42.1%","41.8%"),
        ("SG&A Expenses",               -92,   -99,   -103,   -112,   -134,   -158),
        ("R&D Expenses",                -19,   -20,   -21,    -22,    -23,    -27),
        ("Adj. EBIT",                   147,   162,   181,    170,    222,    305),
        ("Adj. EBIT Margin",            "19.8%","20.4%","21.1%","18.5%","20.3%","23.3%"),
        ("Amortization (acq.)",         -15,   -16,   -17,    -18,    -47,    -85),
        ("GAAP EBIT (est.)",            132,   146,   164,    152,    175,    220),
        ("Interest Expense",            -8,    -8,    -9,     -10,    -12,    -65),
        ("Pre-Tax Income",              124,   138,   155,    142,    163,    155),
        ("Income Tax (28%)",            -35,   -39,   -43,    -40,    -47,    -43),
        ("Net Income (cont. ops)",      89,    99,    112,    103,    116,    112),
        ("Net Margin %",                "12.0%","12.5%","13.1%","11.2%","10.6%","8.6%"),
        ("Diluted Shares (M)",          25.6,  25.7,  25.8,   25.9,   25.9,   31.0),
        ("GAAP Diluted EPS",            "$3.48","$3.85","$4.34","$3.97","$4.49","$3.61"),
        ("Adjusted Diluted EPS",        "$4.20","$4.75","$5.30","$5.20","$6.30","$8.00"),
        ("EPS Growth (adj.)",           None,  "+13%","+12%", "-2%",  "+21%", "+27%"),
    ]
    for i, row_data in enumerate(inc_data):
        label = row_data[0]
        vals = row_data[1:]
        is_pct = "%" in label or label in ("YoY Growth", "EPS Growth (adj.)")
        is_header = label in ("Gross Profit", "Adj. EBIT", "Net Income (cont. ops)", "Adjusted Diluted EPS")
        is_subtotal = label in ("Gross Profit", "Adj. EBIT", "Net Income (cont. ops)")
        bg = LIGHT_NAVY if is_subtotal else (GREY if i % 2 == 0 else WHITE)
        fg_c = "000000"

        cell = ws.cell(row=5+i, column=1, value=label)
        cell.font = font(size=FONT_SIZE, bold=is_header)
        cell.fill = fill(bg)
        cell.alignment = left()
        cell.border = thin_border()

        for j, val in enumerate(vals):
            col = j + 2
            cell2 = ws.cell(row=5+i, column=col, value=val)
            cell2.font = font(size=FONT_SIZE, bold=is_header, color=fg_c)
            cell2.fill = fill(bg)
            cell2.alignment = center()
            cell2.border = thin_border()
            if isinstance(val, (int, float)) and not is_pct:
                cell2.number_format = usd
        ws.row_dimensions[5+i].height = 22

    # Segment Revenue Breakdown
    row_start = 5 + len(inc_data) + 2
    sec_hdr(ws, row_start, 1, 7, "SEGMENT REVENUE ($M)", bg=MID_NAVY)
    row_start += 1
    header_row(ws, row_start, ["Segment"] + years, height=28)
    seg_inc = [
        ("Aerospace & Defense",    295, 315, 330, 341,  478,  570),
        ("% of Total",             "39.8%","39.6%","38.5%","37.1%","43.7%","43.5%"),
        ("Utility Solutions (USG)", 278, 300, 330, 372,  383,  440),
        ("% of Total",             "37.5%","37.7%","38.5%","40.5%","35.0%","33.6%"),
        ("RF Test & Measurement",  169, 180, 198, 206,  234,  300),
        ("% of Total",             "22.8%","22.6%","23.1%","22.4%","21.4%","22.9%"),
        ("Total Revenue",          742, 795, 858, 919, 1095, 1310),
    ]
    for i, row_data in enumerate(seg_inc):
        label = row_data[0]
        vals = row_data[1:]
        is_total = label == "Total Revenue"
        is_pct = "%" in label
        bg = DARK_NAVY if is_total else (LIGHT_NAVY if not is_pct else (GREY if i % 2 == 0 else WHITE))
        fg_c = WHITE if is_total else "000000"
        ws.row_dimensions[row_start+1+i].height = 22

        cell = ws.cell(row=row_start+1+i, column=1, value=label)
        cell.font = font(size=FONT_SIZE, bold=is_total, color=fg_c)
        cell.fill = fill(bg)
        cell.alignment = left()
        cell.border = thin_border()
        for j, val in enumerate(vals):
            col = j + 2
            c2 = ws.cell(row=row_start+1+i, column=col, value=val)
            c2.font = font(size=FONT_SIZE, bold=is_total, color=fg_c)
            c2.fill = fill(bg)
            c2.alignment = center()
            c2.border = thin_border()
            if isinstance(val, (int, float)):
                c2.number_format = usd

    # Segment EBIT Margins
    row2 = row_start + len(seg_inc) + 3
    sec_hdr(ws, row2, 1, 7, "SEGMENT ADJ. EBIT MARGINS", bg=MID_NAVY)
    row2 += 1
    header_row(ws, row2, ["Segment"] + years, height=28)
    seg_margin = [
        ("Aerospace & Defense",    "23%","24%","24%","22%","27%","28%"),
        ("Utility Solutions (USG)", "28%","28%","29%","29%","29%","30%"),
        ("RF Test & Measurement",  "14%","14%","15%","15%","15%","16%"),
        ("Consolidated Adj. EBIT", "19.8%","20.4%","21.1%","18.5%","20.3%","23.3%"),
    ]
    for i, row_data in enumerate(seg_margin):
        label = row_data[0]
        vals = row_data[1:]
        is_con = label == "Consolidated Adj. EBIT"
        bg = DARK_NAVY if is_con else (GREY if i % 2 == 0 else WHITE)
        fg_c = WHITE if is_con else "000000"
        cell = ws.cell(row=row2+1+i, column=1, value=label)
        cell.font = font(size=FONT_SIZE, bold=is_con, color=fg_c)
        cell.fill = fill(bg)
        cell.alignment = left()
        cell.border = thin_border()
        for j, val in enumerate(vals):
            c2 = ws.cell(row=row2+1+i, column=j+2, value=val)
            c2.font = font(size=FONT_SIZE, bold=is_con, color=fg_c)
            c2.fill = fill(bg)
            c2.alignment = center()
            c2.border = thin_border()
        ws.row_dimensions[row2+1+i].height = 22


# ════════════════════════════════════════════════════════════════════════════
# TAB 5 – BALANCE SHEET
# ════════════════════════════════════════════════════════════════════════════
def build_balance(wb):
    ws = wb.create_sheet("5. Balance Sheet")
    ws.sheet_view.showGridLines = False
    widths = [32, 16, 16, 16, 16, 16]
    for i, w in enumerate(widths, 1):
        set_col_width(ws, i, w)

    merge_write(ws, "A1:F1", "ESCO TECHNOLOGIES — BALANCE SHEET", bold=True,
                size=18, bg=DARK_NAVY, fg=WHITE, align=center())
    ws.row_dimensions[1].height = 35

    years_bs = ["FY2022A", "FY2023A", "FY2024A", "FY2025A", "Pro-Forma\n(Post-Megger)"]
    sec_hdr(ws, 3, 1, 6, "BALANCE SHEET ($M, as of September 30 unless noted)", bg=MID_NAVY)
    header_row(ws, 4, ["Line Item"] + years_bs, height=35)

    usd = '#,##0.0'
    bs_data = [
        # ASSETS
        ("── ASSETS ──",              None, None, None, None, None),
        ("Cash & Equivalents",         65,   72,   88,  101,   50),
        ("Accounts Receivable",        170,  183,  195,  220,  310),
        ("Inventories",                120,  128,  140,  160,  210),
        ("Other Current Assets",        35,   38,   40,   45,   60),
        ("TOTAL CURRENT ASSETS",       390,  421,  463,  526,  630),
        ("PP&E (net)",                 185,  196,  208,  235,  290),
        ("Goodwill",                   685,  690,  695,  918,  2400),
        ("Intangible Assets (net)",    120,  108,  96,   280,   900),
        ("Other Long-term Assets",      30,   32,   34,   52,   80),
        ("TOTAL ASSETS",              1410, 1447, 1496, 2011,  4300),
        # LIABILITIES
        ("── LIABILITIES ──",          None, None, None, None, None),
        ("Accounts Payable",            65,   70,   75,   90,  125),
        ("Accrued Liabilities",         95,  100,  110,  130,  175),
        ("Current Debt",                 5,    5,   10,   15,   50),
        ("Other Current Liabilities",   40,   45,   50,   55,   80),
        ("TOTAL CURRENT LIABILITIES",  205,  220,  245,  290,  430),
        ("Long-term Debt",              80,   75,  120,  166, 1550),
        ("Deferred Tax & Other",        80,   85,   90,  110,  180),
        ("TOTAL LIABILITIES",          365,  380,  455,  566, 2160),
        # EQUITY
        ("── EQUITY ──",               None, None, None, None, None),
        ("Common Stock & APIC",        520,  535,  545,  560, 1710),
        ("Retained Earnings",          570,  587,  555,  960,  490),
        ("Treasury Stock & Other",      -45,  -55,  -59,  -75,  -60),
        ("TOTAL EQUITY",              1045, 1067, 1041, 1445, 2140),
        ("TOTAL LIAB. + EQUITY",      1410, 1447, 1496, 2011, 4300),
        # RATIOS
        ("── KEY RATIOS ──",           None, None, None, None, None),
        ("Net Debt ($M)",               20,    8,   42,   64, 1550),
        ("Net Debt / EBITDA",         "0.1x","0.0x","0.2x","0.3x","~4.0x"),
        ("Current Ratio",             "1.9x","1.9x","1.9x","1.8x","1.5x"),
        ("Debt / Equity",             "0.08x","0.07x","0.12x","0.11x","0.72x"),
        ("Book Value / Share",        "$40.7","$41.5","$40.2","$55.8","$69.1"),
    ]

    for i, row_data in enumerate(bs_data):
        label = row_data[0]
        vals = row_data[1:]
        is_total = label.startswith("TOTAL")
        is_section = label.startswith("──")
        is_ratio_hdr = label == "── KEY RATIOS ──"
        bg = (MID_NAVY if is_section else
              (DARK_NAVY if is_total else
               (GREY if i % 2 == 0 else WHITE)))
        fg_c = WHITE if (is_section or is_total) else "000000"

        cell = ws.cell(row=5+i, column=1, value=label)
        cell.font = font(size=FONT_SIZE, bold=(is_total or is_section))
        cell.fill = fill(bg)
        cell.alignment = left()
        cell.border = thin_border()
        ws.row_dimensions[5+i].height = 22

        for j, val in enumerate(vals):
            c2 = ws.cell(row=5+i, column=j+2, value=val)
            c2.font = font(size=FONT_SIZE, bold=(is_total or is_section), color=fg_c)
            c2.fill = fill(bg)
            c2.alignment = center()
            c2.border = thin_border()
            if isinstance(val, (int, float)) and not is_section:
                c2.number_format = usd

    ws.row_dimensions[5 + bs_data.index(("── ASSETS ──", None, None, None, None, None))].height = 22


# ════════════════════════════════════════════════════════════════════════════
# TAB 6 – CASH FLOW ANALYSIS
# ════════════════════════════════════════════════════════════════════════════
def build_cashflow(wb):
    ws = wb.create_sheet("6. Cash Flow Analysis")
    ws.sheet_view.showGridLines = False
    widths = [32, 16, 16, 16, 16, 16, 16]
    for i, w in enumerate(widths, 1):
        set_col_width(ws, i, w)

    merge_write(ws, "A1:G1", "ESCO TECHNOLOGIES — CASH FLOW ANALYSIS", bold=True,
                size=18, bg=DARK_NAVY, fg=WHITE, align=center())
    ws.row_dimensions[1].height = 35

    years = ["FY2021E", "FY2022E", "FY2023E", "FY2024A", "FY2025A", "FY2026E"]
    sec_hdr(ws, 3, 1, 7, "CASH FLOW STATEMENT ($M, FY ends Sep 30)", bg=MID_NAVY)
    header_row(ws, 4, ["Line Item"] + years, height=28)

    usd = '#,##0.0'
    cf_data = [
        ("── OPERATING ──",              None, None, None, None, None, None),
        ("Net Income",                    89,   99,  112,  103,  116,  112),
        ("D&A (excl. acq. amort.)",       28,   30,   32,   34,   42,   55),
        ("Acq. Amortization",             15,   16,   17,   18,   47,   85),
        ("Working Capital Changes",       -5,   -8,  -10,  -12,  -10,  -15),
        ("Other Operating Items",          3,    4,    5,    4,    5,    8),
        ("OPERATING CASH FLOW",          130,  141,  156,  147,  200,  245),
        ("CapEx",                        -22,  -26,  -28,  -28,  -36,  -48),
        ("FREE CASH FLOW (normalized)",  108,  115,  128,  119,  164,  197),
        ("FCF Margin %",               "14.6%","14.5%","14.9%","12.9%","15.0%","15.0%"),
        ("FCF Conversion (FCF/NI)",     "121%","116%","114%","116%","141%","176%"),
        ("── INVESTING ──",              None, None, None, None, None, None),
        ("CapEx",                        -22,  -26,  -28,  -28,  -36,  -48),
        ("Acquisitions (net)",            -30,  -25,  -22,  -18, -472, -922),
        ("Divestitures",                   0,    0,    0,    0,   42,    0),
        ("Other Investing",               -5,   -5,   -5,   -5,   -5,   -5),
        ("NET INVESTING CF",             -57,  -56,  -55,  -51, -471, -975),
        ("── FINANCING ──",              None, None, None, None, None, None),
        ("Net Debt Issuance / (Repay.)",  -5,   -5,   44,   46,   46, 1384),
        ("Dividends Paid",                -9,   -9,  -10,  -10,  -10,  -10),
        ("Share Buybacks",               -10,  -10,   -8,   -5,   -2,   -2),
        ("Other Financing",               -2,   -2,   -2,   -2,   -2,   -2),
        ("NET FINANCING CF",             -26,  -26,   24,   29,   32, 1370),
        ("── SUMMARY ──",                None, None, None, None, None, None),
        ("Net Change in Cash",            47,   59,  125,  125,  -239,  640),
        ("Cash, End of Period",           65,   72,   88,  101,   50,  100),
        ("── FCF QUALITY CHECK ──",      None, None, None, None, None, None),
        ("OCF / Net Income",            "146%","142%","139%","143%","172%","219%"),
        ("CapEx / Revenue",             "3.0%","3.3%","3.3%","3.0%","3.3%","3.7%"),
        ("CapEx / Depreciation",        "79%", "87%", "88%", "82%", "86%", "87%"),
        ("R&D Spend ($M)",               19,   20,   21,   22,   23,   27),
    ]

    for i, row_data in enumerate(cf_data):
        label = row_data[0]
        vals = row_data[1:]
        is_total = label.startswith(("OPERATING", "FREE CASH", "NET INVEST", "NET FINANC"))
        is_section = label.startswith("──")
        bg = (MID_NAVY if is_section else
              (DARK_NAVY if is_total else
               (LIGHT_GREEN if label.startswith("FREE CASH") else
                (GREY if i % 2 == 0 else WHITE))))
        fg_c = WHITE if (is_section or is_total) else "000000"
        if label.startswith("FREE CASH"):
            fg_c = "000000"

        cell = ws.cell(row=5+i, column=1, value=label)
        cell.font = font(size=FONT_SIZE, bold=(is_total or is_section))
        cell.fill = fill(bg)
        cell.alignment = left()
        cell.border = thin_border()
        ws.row_dimensions[5+i].height = 22

        for j, val in enumerate(vals):
            c2 = ws.cell(row=5+i, column=j+2, value=val)
            c2.font = font(size=FONT_SIZE, bold=(is_total or is_section), color=fg_c)
            c2.fill = fill(bg)
            c2.alignment = center()
            c2.border = thin_border()
            if isinstance(val, (int, float)) and not is_section:
                c2.number_format = usd

    # FCF Commentary
    note_row = 5 + len(cf_data) + 2
    sec_hdr(ws, note_row, 1, 7, "FCF QUALITY COMMENTARY", bg=DARK_GREEN)
    note_text = (
        "ESCO generates high-quality free cash flow with OCF consistently above net income (non-cash charges: "
        "D&A + amortization). FCF margin of ~15% is solid for an industrial company. FY2025 FCF was boosted by "
        "working capital improvements and the VACCO divestiture proceeds. FY2026E FCF of ~$197M reflects the SM&P "
        "acquisition contribution but not Megger (deal not yet closed). Post-Megger close, FCF will initially "
        "compress due to interest costs (~$65M/year on $1.5B debt) but Megger itself generates meaningful FCF "
        "(est. $90-100M/yr). Management targets synergies of $60M within 3 years, which should restore FCF "
        "conversion. Capital allocation priority: debt paydown post-Megger, then opportunistic M&A."
    )
    ws.row_dimensions[note_row+1].height = 130
    merge_write(ws, f"A{note_row+1}:G{note_row+5}", note_text, bold=False, size=FONT_SIZE,
                bg=LIGHT_GREEN, fg="000000",
                align=Alignment(horizontal="left", vertical="center", wrap_text=True))


# ════════════════════════════════════════════════════════════════════════════
# TAB 7 – RETURN ON CAPITAL
# ════════════════════════════════════════════════════════════════════════════
def build_roc(wb):
    ws = wb.create_sheet("7. Return on Capital")
    ws.sheet_view.showGridLines = False
    widths = [32, 16, 16, 16, 16, 16, 16]
    for i, w in enumerate(widths, 1):
        set_col_width(ws, i, w)

    merge_write(ws, "A1:G1", "ESCO TECHNOLOGIES — RETURN ON CAPITAL", bold=True,
                size=18, bg=DARK_NAVY, fg=WHITE, align=center())
    ws.row_dimensions[1].height = 35

    years = ["FY2021E", "FY2022E", "FY2023E", "FY2024A", "FY2025A", "FY2026E"]
    sec_hdr(ws, 3, 1, 7, "RETURNS ANALYSIS ($M unless stated)", bg=MID_NAVY)
    header_row(ws, 4, ["Metric"] + years, height=28)

    roc_data = [
        ("Net Income (cont. ops)",      89,   99,  112,  103,  116,  112),
        ("Adj. EBIT",                  147,  162,  181,  170,  222,  305),
        ("NOPAT (Adj. EBIT × 72%)",    106,  117,  130,  122,  160,  220),
        ("Total Assets",              1410, 1447, 1496, 2011, 2011, 4300),
        ("Total Equity",              1045, 1067, 1041, 1445, 1445, 2140),
        ("Invested Capital (est.)",    980, 1000, 1030, 1150, 1580, 3550),
        ("── RETURNS ──",             None, None, None, None, None, None),
        ("ROIC (NOPAT / Inv. Cap.)",  "10.8%","11.7%","12.6%","10.6%","10.1%","6.2%"),
        ("ROE (NI / Equity)",          "8.5%","9.3%","10.8%","7.1%","8.0%","5.2%"),
        ("ROA (NI / Assets)",          "6.3%","6.8%","7.5%","5.1%","5.8%","2.6%"),
        ("Adj. EBIT / Assets",        "10.4%","11.2%","12.1%","8.5%","11.0%","7.1%"),
        ("── INCREMENTAL RETURNS ──",  None, None, None, None, None, None),
        ("Incremental Revenue ($M)",   None,   53,   63,   61,  176,  215),
        ("Incremental Adj. EBIT",      None,   15,   19,  -11,   52,   83),
        ("ROIIC (Inc EBIT/Inc Rev.)",  None, "28%","30%","-18%","30%","39%"),
        ("── CAPITAL EFFICIENCY ──",   None, None, None, None, None, None),
        ("Asset Turnover",            "0.53x","0.55x","0.57x","0.46x","0.54x","0.30x"),
        ("Revenue / Invested Cap.",   "0.76x","0.80x","0.83x","0.80x","0.69x","0.37x"),
        ("FCF / Invested Capital",    "11.0%","11.5%","12.4%","10.3%","10.4%","5.5%"),
        ("── WACC COMPARISON ──",      None, None, None, None, None, None),
        ("WACC (est.)",               "9.0%","9.0%","9.0%","9.0%","9.0%","9.0%"),
        ("ROIC - WACC (spread)",      "+1.8%","+2.7%","+3.6%","+1.6%","+1.1%","-2.8%"),
        ("Value Creation Status",     "Positive","Positive","Positive","Positive","Slim","Negative"),
    ]

    for i, row_data in enumerate(roc_data):
        label = row_data[0]
        vals = row_data[1:]
        is_section = label.startswith("──")
        is_wacc = label in ("ROIC - WACC (spread)", "Value Creation Status")
        bg = (MID_NAVY if is_section else
              (LIGHT_GREEN if label == "ROIC (NOPAT / Inv. Cap.)" else
               (LIGHT_RED if is_wacc and vals[-1] == "Negative" else
                (GREY if i % 2 == 0 else WHITE))))
        fg_c = WHITE if is_section else "000000"

        cell = ws.cell(row=5+i, column=1, value=label)
        cell.font = font(size=FONT_SIZE, bold=is_section)
        cell.fill = fill(bg)
        cell.alignment = left()
        cell.border = thin_border()
        ws.row_dimensions[5+i].height = 22

        for j, val in enumerate(vals):
            c2 = ws.cell(row=5+i, column=j+2, value=val)
            c2.font = font(size=FONT_SIZE, bold=is_section, color=fg_c)
            c2.fill = fill(bg)
            c2.alignment = center()
            c2.border = thin_border()
            if isinstance(val, (int, float)) and not is_section:
                c2.number_format = '#,##0.0'

    note_row = 5 + len(roc_data) + 2
    sec_hdr(ws, note_row, 1, 7, "ROIC COMMENTARY — THINKING LIKE AN OWNER", bg=DARK_GREEN)
    note_text = (
        "ESCO's ROIC has hovered around 10-13% over the past 5 years — decent but not exceptional for an "
        "industrial compounder. The core business (ex-acquisitions) earns 12-15% ROIC, but goodwill from "
        "serial acquisitions dilutes the consolidated figure. Critically, the Megger deal at $2.35B will "
        "dramatically expand the invested capital base: pro-forma ROIC (FY2026 post-close) falls to ~6%, "
        "below WACC. Management needs to execute on $60M synergies and grow Megger revenue to earn its way "
        "back above 9-10% ROIC within 3-4 years. The SM&P (Maritime) acquisition has been additive: "
        "A&D segment EBIT margins held ~27% with Maritime included. ROIIC on the SM&P deal was ~15% in "
        "year 1 — reasonable but not exceptional. The key question for the Megger thesis: can management "
        "replicate its integration track record at a much larger acquisition ($2.35B vs. prior deals averaging $100-200M)?"
    )
    ws.row_dimensions[note_row+1].height = 150
    merge_write(ws, f"A{note_row+1}:G{note_row+6}", note_text, bold=False, size=FONT_SIZE,
                bg=LIGHT_GREEN, fg="000000",
                align=Alignment(horizontal="left", vertical="center", wrap_text=True))


# ════════════════════════════════════════════════════════════════════════════
# TAB 8 – MANAGEMENT
# ════════════════════════════════════════════════════════════════════════════
def build_management(wb):
    ws = wb.create_sheet("8. Management")
    ws.sheet_view.showGridLines = False
    for i, w in enumerate([28, 18, 15, 15, 30], 1):
        set_col_width(ws, i, w)

    merge_write(ws, "A1:E1", "ESCO TECHNOLOGIES — MANAGEMENT QUALITY ANALYSIS", bold=True,
                size=18, bg=DARK_NAVY, fg=WHITE, align=center())
    ws.row_dimensions[1].height = 35

    # Leadership
    sec_hdr(ws, 3, 1, 5, "EXECUTIVE LEADERSHIP", bg=MID_NAVY)
    header_row(ws, 4, ["Name", "Title", "Tenure", "Comp (FY2025)", "Background & Notes"], height=28)
    leaders = [
        ["Bryan H. Sayler", "President & CEO", "~3 years as CEO", "$5.59M (+44% YoY)",
         "Long ESCO insider; previously ran segments before CEO role. Deep operational background. Has executed SM&P deal and now Megger — acting boldly on M&A. Strong communicator."],
        ["Christopher L. Tucker", "SVP & CFO", "Multiple years", "~$2.5M (est.)",
         "Finance veteran; manages capital structure. Arranged $1.5B JPMorgan facility for Megger. Conservative on leverage historically — Megger is a departure."],
        ["David M. Schatz", "SVP, General Counsel & Secretary", "Long tenure", "N/A",
         "Governance & legal. Leads compliance and SEC filings. Critical for Megger regulatory approvals."],
    ]
    for i, row in enumerate(leaders):
        bg = GREY if i % 2 == 0 else WHITE
        for j, val in enumerate(row):
            cell = ws.cell(row=5+i, column=j+1, value=val)
            cell.font = font(size=FONT_SIZE)
            cell.fill = fill(bg)
            cell.alignment = left()
            cell.border = thin_border()
        ws.row_dimensions[5+i].height = 60

    # Compensation Structure
    sec_hdr(ws, 9, 1, 5, "COMPENSATION STRUCTURE (PROXY ANALYSIS)", bg=MID_NAVY)
    header_row(ws, 10, ["Component", "Weight", "Metric Used", "Payout Range", "Assessment"], height=28)
    comp_data = [
        ["Base Salary", "~15% of total", "N/A (fixed)", "$750K–900K est.", "Reasonable base; not excessive relative to peers"],
        ["Annual Cash Incentive", "~35% of total", "Adj. EPS + Adj. Operating CF", "0%–200% of target", "Well-aligned with shareholder value; both metrics are key owner metrics"],
        ["Long-term Equity (RSUs)", "~50% of total", "Adj. EPS growth (3-yr) + TSR vs. peers", "Cliff vesting", "Majority in equity = skin in the game; 3-yr vesting period appropriate"],
        ["Total CEO Comp (FY2025)", "$5.59M", "+44% YoY", "ISS Score: 1 (best)", "44% jump flagged by some; driven by strong EPS outperformance"],
    ]
    for i, row in enumerate(comp_data):
        is_total = i == 3
        bg = LIGHT_NAVY if is_total else (GREY if i % 2 == 0 else WHITE)
        for j, val in enumerate(row):
            cell = ws.cell(row=11+i, column=j+1, value=val)
            cell.font = font(size=FONT_SIZE, bold=is_total)
            cell.fill = fill(bg)
            cell.alignment = center() if j in [1,2,3] else left()
            cell.border = thin_border()
        ws.row_dimensions[11+i].height = 40

    # Capital Allocation Track Record
    sec_hdr(ws, 16, 1, 5, "CAPITAL ALLOCATION DECISIONS (OWNER'S LENS)", bg=MID_NAVY)
    header_row(ws, 17, ["Decision", "Year", "Size", "Outcome", "Owner Assessment"], height=28)
    capalloc = [
        ["SM&P (Maritime) Acquisition", "Apr 2025", "$472M", "Added $95M revenue in H2 FY2025; EBIT margin held ~27% in A&D; backlog exploded",
         "POSITIVE — Disciplined price, strong strategic fit, improved Navy positioning"],
        ["VACCO Divestiture", "FY2025", "~$42M proceeds", "Sold non-core filtration unit to focus portfolio",
         "POSITIVE — Pruned the tree; redeployed proceeds toward higher-quality assets"],
        ["Megger Acquisition", "Apr 2026", "$2.35B", "Pending — transforms USG into global utility-testing leader",
         "BOLD — Largest deal in company history; leverage risk is real but strategic logic is compelling"],
        ["Share Buybacks", "Ongoing", "~$2-10M/yr", "Modest; not the primary capital return tool",
         "NEUTRAL — Appropriately deprioritized vs. growth M&A"],
        ["Dividends", "Ongoing", "~$10M/yr ($0.38/share)", "Symbolic; very low yield (~0.2%)",
         "NEUTRAL — Not an income stock; growth focus appropriate"],
        ["CapEx / R&D", "Annual", "$36M CapEx + $23M R&D", "Maintained productive asset base; R&D growing modestly",
         "POSITIVE — CapEx below D&A suggests asset-light; R&D investment sustains competitive position"],
    ]
    for i, row in enumerate(capalloc):
        bg = GREY if i % 2 == 0 else WHITE
        for j, val in enumerate(row):
            cell = ws.cell(row=18+i, column=j+1, value=val)
            cell.font = font(size=FONT_SIZE)
            cell.fill = fill(bg)
            cell.alignment = left()
            cell.border = thin_border()
        ws.row_dimensions[18+i].height = 50

    # CEO as Owner Assessment
    sec_hdr(ws, 25, 1, 5, "DOES THE CEO ACT LIKE AN OWNER?", bg=DARK_GREEN)
    owner_data = [
        ("Incentive Alignment",  "HIGH",    "85% of compensation is at-risk equity/bonus tied to EPS + CF"),
        ("Capital Discipline",   "MEDIUM",  "Good track record except Megger — largest bet ever at ~42% of market cap"),
        ("Long-term Thinking",   "HIGH",    "SM&P deal positions for 10-20yr Navy programs; Megger for global utility cycle"),
        ("Shareholder Communication","HIGH", "Clear segment reporting, raised guidance, transparent about Megger risk"),
        ("Insider Ownership",    "MEDIUM",  "RSU grants vest over 3 years; limited open-market purchases observed"),
        ("Seeds for the Future", "HIGH",    "Backlog at $1.4B and growing; investing in capacity for Navy surge demand"),
        ("Borrowing from Future", "LOW",    "GAAP accounting is clean; no aggressive revenue recognition"),
    ]
    for i, (factor, rating, comment) in enumerate(owner_data):
        bg_r = (LIGHT_GREEN if rating == "HIGH" else
                (LIGHT_ORANGE if rating == "MEDIUM" else
                 LIGHT_RED))
        cell = ws.cell(row=26+i, column=1, value=factor)
        cell.font = font(size=FONT_SIZE, bold=True)
        cell.fill = fill(bg_r)
        cell.alignment = left()
        cell.border = thin_border()
        c2 = ws.cell(row=26+i, column=2, value=rating)
        c2.font = font(size=FONT_SIZE, bold=True)
        c2.fill = fill(bg_r)
        c2.alignment = center()
        c2.border = thin_border()
        ws.merge_cells(f"C{26+i}:E{26+i}")
        c3 = ws.cell(row=26+i, column=3, value=comment)
        c3.font = font(size=FONT_SIZE)
        c3.fill = fill(bg_r)
        c3.alignment = left()
        c3.border = thin_border()
        ws.row_dimensions[26+i].height = 24

    # Insider Transactions
    sec_hdr(ws, 34, 1, 5, "INSIDER TRANSACTIONS (Recent SEC Form 4 Filings)", bg=MID_NAVY)
    header_row(ws, 35, ["Date", "Insider", "Transaction", "Shares", "Assessment"], height=24)
    insider = [
        ["Oct 2025",  "Director", "RSU grant (dividend equiv.)", "3.3 RSUs",   "Routine; not indicative of conviction"],
        ["Nov 2025",  "Director", "RSU conversion to shares",    "705 shares", "Vest event; no cash purchase signal"],
        ["Feb 2026",  "Multiple", "Form 4 filings (various RSU)","Routine",    "No meaningful open-market purchases noted"],
        ["ASSESSMENT","–",        "No significant insider buying observed; primarily RSU grants and conversions. "
         "CEO has substantial RSU exposure from grants (aligning long-term). Absence of open-market buying "
         "at current valuation (~34x trailing adj. P/E) is notable but not alarming given 50% YTD run.",
         "–",         "NEUTRAL"],
    ]
    for i, row in enumerate(insider):
        is_assess = i == 3
        bg = LIGHT_ORANGE if is_assess else (GREY if i % 2 == 0 else WHITE)
        for j, val in enumerate(row):
            cell = ws.cell(row=36+i, column=j+1, value=val)
            cell.font = font(size=FONT_SIZE, bold=is_assess)
            cell.fill = fill(bg)
            cell.alignment = left()
            cell.border = thin_border()
        ws.row_dimensions[36+i].height = 55 if is_assess else 24


# ════════════════════════════════════════════════════════════════════════════
# TAB 9 – RISKS
# ════════════════════════════════════════════════════════════════════════════
def build_risks(wb):
    ws = wb.create_sheet("9. Risks")
    ws.sheet_view.showGridLines = False
    for i, w in enumerate([35, 14, 14, 14, 38], 1):
        set_col_width(ws, i, w)

    merge_write(ws, "A1:E1", "ESCO TECHNOLOGIES — RISK ANALYSIS", bold=True,
                size=18, bg=DARK_NAVY, fg=WHITE, align=center())
    ws.row_dimensions[1].height = 35

    sec_hdr(ws, 3, 1, 5, "RISK MATRIX (Probability × Impact → Overall)", bg=MID_NAVY)
    header_row(ws, 4, ["Risk Factor", "Probability", "Impact", "Overall", "Mitigation / Commentary"], height=28)

    risks = [
        ("Megger Integration Failure / Leverage Overhang",
         "MEDIUM", "VERY HIGH", "HIGH",
         "At $2.35B, this is ESCO's largest deal (42% of current mkt cap). Leverage rises to ~4x EBITDA. "
         "Any integration miss, demand slowdown, or interest rate increase compounds the risk. "
         "Mitigant: $60M synergies are achievable; JPMorgan committed financing; Megger is profitable."),
        ("Defense Budget Cuts / Navy Program Risk",
         "LOW-MED", "HIGH", "MEDIUM",
         "U.S. defense spending is a political variable. Columbia-class submarine program delays or budget "
         "sequesters would hit A&D backlog conversion. Mitigant: Navy programs are top DoD priority; backlog "
         "of $1.4B provides 1+ year visibility even if new orders slow."),
        ("Execution Risk — Post-SM&P Integration",
         "LOW", "MEDIUM", "LOW-MED",
         "SM&P was closed April 2025 and is still being integrated. Any margin compression or synergy miss "
         "while managing the larger Megger deal simultaneously is a real operational risk."),
        ("Valuation / Multiple Compression",
         "MEDIUM", "HIGH", "MEDIUM",
         "At 34x trailing adj P/E and 27x forward, ESE is priced for execution perfection. Any earnings miss "
         "or guidance cut could cause sharp de-rating, especially given 50% YTD run. "
         "Mitigant: record backlog and strong book-to-bill provide earnings visibility."),
        ("Key Customer Concentration (U.S. Navy)",
         "LOW", "HIGH", "MEDIUM",
         "Navy-related revenue is ~35%+ of total. Sole-source positions protect pricing but create "
         "concentration risk. Diversification via USG/Megger helps over time."),
        ("M&A Integration Capability at Scale",
         "MEDIUM", "HIGH", "MEDIUM-HIGH",
         "ESCO has historically done smaller tuck-ins ($50-200M). Megger ($2.35B) tests management's "
         "integration playbook at a new scale. Cultural integration of a UK-headquartered global company adds complexity."),
        ("Interest Rate Sensitivity",
         "LOW-MED", "MEDIUM", "MEDIUM",
         "Post-Megger debt of ~$1.55B at floating rates could add $20-30M in additional interest per "
         "100bps rate increase. Net income will be pressured until debt is paid down."),
        ("Competition in RF Test Segment",
         "MEDIUM", "MEDIUM", "MEDIUM",
         "Keysight, VIAVI, and smaller chamber builders compete with ETS-Lindgren. Margin compression "
         "possible if large defense/aerospace clients consolidate vendors. Test segment is the weakest moat."),
        ("Currency / FX Risk",
         "MEDIUM", "MEDIUM", "MEDIUM",
         "Megger is UK-based with significant European revenue. GBP/EUR exposure adds FX risk not "
         "currently in ESCO's model. No current hedging program disclosed."),
        ("ESG / Supply Chain Disruptions",
         "LOW", "LOW-MED", "LOW",
         "Aerospace/defense supply chain has been tight post-COVID. Semiconductor shortages, raw material "
         "inflation, and labor constraints could delay shipments. Mitigant: backlog-driven demand smooths impact."),
    ]

    for i, (risk, prob, impact, overall, comment) in enumerate(risks):
        overall_bg = (RED if overall in ("HIGH", "VERY HIGH") else
                      (ORANGE if "MEDIUM" in overall else
                       DARK_GREEN))
        row_bg = (LIGHT_RED if overall in ("HIGH", "VERY HIGH") else
                  (LIGHT_ORANGE if "MEDIUM" in overall else
                   LIGHT_GREEN))
        row = 5 + i
        cell = ws.cell(row=row, column=1, value=risk)
        cell.font = font(size=FONT_SIZE, bold=True)
        cell.fill = fill(row_bg)
        cell.alignment = left()
        cell.border = thin_border()

        for j, (val, bg) in enumerate([(prob, row_bg), (impact, row_bg), (overall, overall_bg)]):
            c2 = ws.cell(row=row, column=j+2, value=val)
            c2.font = font(size=FONT_SIZE, bold=(j == 2), color=(WHITE if j == 2 else "000000"))
            c2.fill = fill(bg)
            c2.alignment = center()
            c2.border = thin_border()

        c5 = ws.cell(row=row, column=5, value=comment)
        c5.font = font(size=FONT_SIZE)
        c5.fill = fill(row_bg)
        c5.alignment = left()
        c5.border = thin_border()
        ws.row_dimensions[row].height = 65

    # Overall risk verdict
    verd_row = 5 + len(risks) + 1
    sec_hdr(ws, verd_row, 1, 5, "OVERALL RISK VERDICT", bg=ORANGE)
    verdict_text = (
        "ESCO's risk profile has ELEVATED MATERIALLY in 2026 due to the Megger acquisition. Pre-Megger, "
        "ESCO was a low-leverage, high-quality industrial with a steady risk profile. Post-Megger announcement, "
        "the company is taking on ~4x net debt/EBITDA — a significant departure from its conservative history. "
        "The strategic logic is sound (utility testing leadership, cross-sell Doble + Megger), but the "
        "execution bar is high. Investors should treat this as a 'show me the synergies' story over the "
        "next 2-3 years. The 50% YTD stock run already prices in considerable optimism. Risk/reward "
        "is less compelling at current levels versus 12 months ago."
    )
    ws.row_dimensions[verd_row+1].height = 140
    merge_write(ws, f"A{verd_row+1}:E{verd_row+5}", verdict_text, bold=False, size=FONT_SIZE,
                bg=LIGHT_ORANGE, fg="000000",
                align=Alignment(horizontal="left", vertical="center", wrap_text=True))


# ════════════════════════════════════════════════════════════════════════════
# TAB 10 – VALUATION
# ════════════════════════════════════════════════════════════════════════════
def build_valuation(wb):
    ws = wb.create_sheet("10. Valuation")
    ws.sheet_view.showGridLines = False
    for i, w in enumerate([32, 18, 18, 18, 18, 18], 1):
        set_col_width(ws, i, w)

    merge_write(ws, "A1:F1", "ESCO TECHNOLOGIES — STOCK VALUATION", bold=True,
                size=18, bg=DARK_NAVY, fg=WHITE, align=center())
    ws.row_dimensions[1].height = 35

    # Current Pricing
    sec_hdr(ws, 3, 1, 6, "CURRENT MARKET SNAPSHOT", bg=MID_NAVY)
    snap = [
        ("Current Stock Price (est.)",    "~$215"),
        ("Shares Outstanding (diluted)",  "~25.9M"),
        ("Market Capitalization",         "~$5.57B"),
        ("Net Debt (FY2025)",             "$64.7M"),
        ("Enterprise Value (EV)",         "~$5.63B"),
        ("FY2025 Revenue (TTM)",          "$1,095M"),
        ("FY2025 Adj. EBIT",              "$222.4M"),
        ("FY2025 Adj. EBITDA (est.)",     "~$267M"),
        ("FY2025 Adj. EPS",               "~$6.30"),
        ("FY2026E Adj. EPS (guidance)",   "$7.90–$8.15"),
        ("FY2026E Adj. EPS (mid)",        "$8.00"),
        ("52-Week Range",                 "~$130–$240"),
        ("YTD Performance",               "+50%"),
    ]
    for i, (label, val) in enumerate(snap):
        r = 4 + i
        wc(ws, r, 1, label, bold=True, bg=MID_NAVY, fg=WHITE, align=center(), size=FONT_SIZE)
        ws.merge_cells(f"B{r}:F{r}")
        wc(ws, r, 2, val, bold=False, bg=LIGHT_NAVY, fg="000000", align=center(), size=FONT_SIZE)
        ws.row_dimensions[r].height = 22

    # Multiples Valuation
    mv_row = 4 + len(snap) + 2
    sec_hdr(ws, mv_row, 1, 6, "MULTIPLES VALUATION", bg=MID_NAVY)
    header_row(ws, mv_row+1,
               ["Method", "Multiple Used", "Base Metric", "Implied Value", "Per Share", "vs. Current ($215)"],
               height=28)
    mult = [
        ["Trailing Adj. P/E",           "34.0x",  "FY2025 Adj. EPS $6.30",  "N/A",     "$214",  "-0.5%"],
        ["Forward Adj. P/E",            "27.0x",  "FY2026E Adj. EPS $8.00", "N/A",     "$216",  "+0.5%"],
        ["EV / EBITDA (trailing)",       "21.1x",  "FY2025 EBITDA $267M",    "$5.63B",  "$216",  "+0.5%"],
        ["EV / EBITDA (forward FY2026)", "17.1x",  "FY2026E EBITDA $330M",   "$5.64B",  "$217",  "+1.0%"],
        ["EV / Revenue",                 "5.1x",   "FY2025 Revenue $1,095M", "$5.58B",  "$214",  "-0.5%"],
        ["P/FCF (normalized)",          "34.0x",  "FY2025 FCF $164M",       "N/A",     "$215",  "0.0%"],
        ["Analyst Median Target",        "–",      "Wall Street consensus",   "N/A",     "$300",  "+39.5%"],
    ]
    for i, row in enumerate(mult):
        is_analyst = i == len(mult) - 1
        bg = LIGHT_GREEN if is_analyst else (GREY if i % 2 == 0 else WHITE)
        for j, val in enumerate(row):
            cell = ws.cell(row=mv_row+2+i, column=j+1, value=val)
            cell.font = font(size=FONT_SIZE, bold=is_analyst)
            cell.fill = fill(bg)
            cell.alignment = center() if j > 0 else left()
            cell.border = thin_border()
        ws.row_dimensions[mv_row+2+i].height = 22

    # DCF Valuation
    dcf_row = mv_row + len(mult) + 4
    sec_hdr(ws, dcf_row, 1, 6, "DCF VALUATION (Base Case)", bg=MID_NAVY)
    header_row(ws, dcf_row+1, ["Year", "FCF ($M)", "Growth Rate", "Discount Factor (9%)", "PV of FCF ($M)", ""], height=28)
    dcf_years = [
        ("FY2026E", 197,  "–",      0.917,  181,  ""),
        ("FY2027E", 236, "+20%",    0.842,  199,  ""),
        ("FY2028E", 278, "+18%",    0.772,  215,  ""),
        ("FY2029E", 320, "+15%",    0.708,  227,  ""),
        ("FY2030E", 362, "+13%",    0.650,  235,  ""),
    ]
    for i, row in enumerate(dcf_years):
        for j, val in enumerate(row):
            cell = ws.cell(row=dcf_row+2+i, column=j+1, value=val)
            cell.font = font(size=FONT_SIZE)
            cell.fill = fill(GREY if i % 2 == 0 else WHITE)
            cell.alignment = center() if j > 0 else left()
            cell.border = thin_border()
        ws.row_dimensions[dcf_row+2+i].height = 22

    dcf_sum_row = dcf_row + 8
    dcf_sum = [
        ("Sum of PV (FCFs FY2026-2030)",  "$1,057M"),
        ("Terminal Growth Rate",           "3.0%"),
        ("Terminal Year FCF (FY2030)",     "$362M"),
        ("Terminal Value (TV)",            "$6,216M (TV = FCF×1.03/(9%-3%))"),
        ("PV of Terminal Value",           "$4,044M"),
        ("Enterprise Value (DCF)",         "$5,101M"),
        ("Less: Net Debt",                 "$65M"),
        ("Equity Value",                   "$5,036M"),
        ("Shares Outstanding",             "25.9M"),
        ("Intrinsic Value (Base)",         "~$194/share"),
        ("Bear Case (12% WACC, 2% TGR)",   "~$155/share"),
        ("Bull Case (8% WACC, 3.5% TGR)", "~$240/share"),
        ("Current Price",                  "~$215/share"),
        ("Margin of Safety (Base)",        "≈ NONE (stock near/above fair value)"),
    ]
    for i, (label, val) in enumerate(dcf_sum):
        is_key = label in ("Intrinsic Value (Base)", "Margin of Safety (Base.)", "Current Price", "Margin of Safety (Base)")
        bg = (LIGHT_GREEN if "Bull" in label else
              (LIGHT_RED if "Bear" in label or "NONE" in val else
               (LIGHT_NAVY if is_key else
                (GREY if i % 2 == 0 else WHITE))))
        r = dcf_sum_row + i
        wc(ws, r, 1, label, bold=is_key, bg=bg, fg="000000", align=left(), size=FONT_SIZE, border=True)
        ws.merge_cells(f"B{r}:F{r}")
        wc(ws, r, 2, val, bold=is_key, bg=bg, fg="000000", align=center(), size=FONT_SIZE, border=True)
        ws.row_dimensions[r].height = 22

    # Valuation Conclusion
    conc_row = dcf_sum_row + len(dcf_sum) + 2
    sec_hdr(ws, conc_row, 1, 6, "VALUATION VERDICT — IS THERE A MARGIN OF SAFETY?", bg=ORANGE)
    conc_text = (
        "At ~$215/share, ESCO Technologies is trading very close to our base-case DCF intrinsic value of ~$194. "
        "The stock has run 50% YTD and is pricing in significant earnings growth AND a successful Megger "
        "integration. On a pure multiples basis, the stock looks fair (27x forward P/E, ~17x forward EV/EBITDA) "
        "for a high-quality industrial compounder with a record backlog. THERE IS LIMITED MARGIN OF SAFETY at "
        "current levels.\n\n"
        "Bull Case ($240+): Megger synergies ($60M) flow through by FY2028; ROIC recovers to 10%+; Navy "
        "funding remains robust. Post-Megger EPS power could reach $12-14 by FY2028 → $250-300 stock.\n\n"
        "Bear Case ($150-170): Megger integration challenges; defense budget pressure; leverage constrains "
        "flexibility; multiple compression from 27x to 20x → $160/share.\n\n"
        "Analyst consensus ($300 target) implies full execution of Megger + sustained 25%+ growth — that's "
        "the optimistic scenario. A disciplined owner would wait for a pullback to $170-185 for a clear "
        "margin of safety, or size the position small acknowledging the embedded optionality in Megger."
    )
    ws.row_dimensions[conc_row+1].height = 200
    merge_write(ws, f"A{conc_row+1}:F{conc_row+9}", conc_text, bold=False, size=FONT_SIZE,
                bg=LIGHT_ORANGE, fg="000000",
                align=Alignment(horizontal="left", vertical="center", wrap_text=True))


# ════════════════════════════════════════════════════════════════════════════
# TAB 11 – MARKET SENTIMENT
# ════════════════════════════════════════════════════════════════════════════
def build_sentiment(wb):
    ws = wb.create_sheet("11. Market Sentiment")
    ws.sheet_view.showGridLines = False
    for i, w in enumerate([30, 18, 18, 18, 28], 1):
        set_col_width(ws, i, w)

    merge_write(ws, "A1:E1", "ESCO TECHNOLOGIES — MARKET SENTIMENT", bold=True,
                size=18, bg=DARK_NAVY, fg=WHITE, align=center())
    ws.row_dimensions[1].height = 35

    # Analyst Coverage
    sec_hdr(ws, 3, 1, 5, "ANALYST COVERAGE & PRICE TARGETS", bg=MID_NAVY)
    header_row(ws, 4, ["Firm (est.)", "Rating", "Price Target", "Implied Upside", "Key Thesis"], height=28)
    analysts = [
        ["Analyst Consensus (5 firms)", "Strong Buy", "$300 (median)", "+39.5%", "Record backlog; A&D surge; Megger transforms USG"],
        ["Low Target",                  "Buy",        "$270",          "+25.6%", "Conservative; leverage risk from Megger"],
        ["High Target",                 "Strong Buy", "$350",          "+62.8%", "Bull case: Megger synergies + Navy supercycle"],
        ["ISS Governance Score",        "N/A",        "QualityScore: 3", "N/A", "Compensation: 1 (best), Board: 2, S/H Rights: 7, Audit: 7"],
    ]
    for i, row in enumerate(analysts):
        is_cons = i == 0
        bg = LIGHT_GREEN if is_cons else (GREY if i % 2 == 0 else WHITE)
        for j, val in enumerate(row):
            cell = ws.cell(row=5+i, column=j+1, value=val)
            cell.font = font(size=FONT_SIZE, bold=is_cons)
            cell.fill = fill(bg)
            cell.alignment = center() if j in [1,2,3] else left()
            cell.border = thin_border()
        ws.row_dimensions[5+i].height = 28

    # Recent Catalysts & News
    sec_hdr(ws, 10, 1, 5, "RECENT CATALYSTS & KEY EVENTS", bg=MID_NAVY)
    header_row(ws, 11, ["Date", "Event", "Impact", "Market Reaction", "Commentary"], height=28)
    events = [
        ["Q1 FY2026\n(Feb 2026)", "Q1 Results: Rev +35%, Orders +143%, EPS $1.64", "VERY POSITIVE",
         "Stock rallied strongly; new record backlog $1.4B", "Blowout quarter; book-to-bill of 1.92x was the standout — implied >$1B of revenue locked in over next 12 months"],
        ["Apr 15, 2026", "Megger Acquisition Announced ($2.35B)", "MIXED",
         "Stock dropped ~2.7% on day of announcement", "Market concerned about leverage (4x ND/EBITDA) and dilution (5.1M shares); strategic logic praised but execution risk highlighted"],
        ["FY2025 Q4 (Nov 2025)", "Record Annual Results: $1.1B revenue, FCF $164M", "POSITIVE",
         "Stock added to YTD gain", "Strong close to the year; VACCO divestiture completed; SM&P contributing ahead of schedule"],
        ["Apr 2025", "SM&P (Maritime) Acquisition Closed ($472M)", "POSITIVE",
         "Well received; perceived as disciplined bet on Navy", "ROIC-accretive; demonstrated management's ability to close and begin integrating acquisitions quickly"],
        ["YTD 2026", "Stock +50% YTD (as of Apr 2026)", "CONTEXT",
         "Outperforming industrials sector", "Driven by record backlog, Q1 blowout, and Megger announcement. Valuation has expanded significantly."],
    ]
    for i, row in enumerate(events):
        bg = (LIGHT_GREEN if "POSITIVE" in row[2] else
              (LIGHT_RED if "NEGATIVE" in row[2] else
               (LIGHT_ORANGE if "MIXED" in row[2] else GREY)))
        for j, val in enumerate(row):
            cell = ws.cell(row=12+i, column=j+1, value=val)
            cell.font = font(size=FONT_SIZE)
            cell.fill = fill(bg)
            cell.alignment = center() if j in [0,2,3] else left()
            cell.border = thin_border()
        ws.row_dimensions[12+i].height = 65

    # Competitive Trends
    sec_hdr(ws, 18, 1, 5, "COMPETITIVE LANDSCAPE & MARKET TRENDS", bg=MID_NAVY)
    trends = [
        ("U.S. Navy / DoD Spending",    "TAILWIND",  "Columbia-class submarine program, AUKUS alliance, Navy shipbuilding expansion = sustained demand for 10-20 years. "
         "ESCO is sole-qualified on multiple programs. Congressional support remains strong."),
        ("Utility Grid Modernization",  "TAILWIND",  "Global grid investment driven by electrification, EV charging, renewable integration. "
         "Doble and (post-close) Megger are direct beneficiaries. IRA (Inflation Reduction Act) spending adds to U.S. tailwind."),
        ("EMC Test Demand",             "NEUTRAL+",  "5G rollout, EV testing, medical device proliferation drive EMC testing demand. "
         "ETS-Lindgren well positioned but faces competition from Keysight and specialized chamber builders."),
        ("Defense Consolidation",       "NEUTRAL",   "Large primes (Raytheon, L3Harris, Northrop) have been consolidating smaller defense sub-suppliers. "
         "ESCO's niche position partially protects it but could face pricing pressure if primes insource."),
        ("Interest Rate Environment",   "HEADWIND",  "Post-Megger leverage at ~4x EBITDA creates sensitivity to higher-for-longer rates. "
         "$65M annual interest cost will compress near-term GAAP EPS and FCF conversion."),
    ]
    header_row(ws, 19, ["Trend", "Direction", "Detail", "", ""], height=28)
    for i, (trend, direction, detail) in enumerate(trends):
        bg = (LIGHT_GREEN if direction == "TAILWIND" else
              (LIGHT_RED if direction == "HEADWIND" else GREY))
        dir_col = (DARK_GREEN if direction == "TAILWIND" else
                   (RED if direction == "HEADWIND" else MID_NAVY))
        row = 20 + i
        cell = ws.cell(row=row, column=1, value=trend)
        cell.font = font(size=FONT_SIZE, bold=True)
        cell.fill = fill(bg)
        cell.alignment = left()
        cell.border = thin_border()
        c2 = ws.cell(row=row, column=2, value=direction)
        c2.font = font(size=FONT_SIZE, bold=True, color=WHITE)
        c2.fill = fill(dir_col)
        c2.alignment = center()
        c2.border = thin_border()
        ws.merge_cells(f"C{row}:E{row}")
        c3 = ws.cell(row=row, column=3, value=detail)
        c3.font = font(size=FONT_SIZE)
        c3.fill = fill(bg)
        c3.alignment = left()
        c3.border = thin_border()
        ws.row_dimensions[row].height = 65


# ════════════════════════════════════════════════════════════════════════════
# TAB 12 – KEY INDICATORS
# ════════════════════════════════════════════════════════════════════════════
def build_indicators(wb):
    ws = wb.create_sheet("12. Key Indicators")
    ws.sheet_view.showGridLines = False
    for i, w in enumerate([32, 20, 20, 20, 24], 1):
        set_col_width(ws, i, w)

    merge_write(ws, "A1:E1", "ESCO TECHNOLOGIES — KEY INDICATOR DASHBOARD", bold=True,
                size=18, bg=DARK_NAVY, fg=WHITE, align=center())
    ws.row_dimensions[1].height = 35
    merge_write(ws, "A2:E2", "Snapshot as of April 20, 2026  |  FY ends September 30",
                bold=False, size=FONT_SIZE, bg=MID_NAVY, fg=WHITE, align=center())
    ws.row_dimensions[2].height = 24

    sections = [
        ("GROWTH METRICS", [
            ("Revenue CAGR (FY2021-FY2025)",     "~10.2% p.a.",       "+", "Strong organic + M&A-driven growth"),
            ("Revenue Growth FY2025 YoY",         "+19.2%",            "+", "Accelerated via SM&P/Maritime contribution"),
            ("Adj. EPS Growth FY2025 YoY",        "+21%",              "+", "Ahead of revenue growth; margin expansion"),
            ("Order Growth Q1 FY2026",            "+143% YoY",         "++","Record orders; 1.92x book-to-bill"),
            ("Backlog Growth (FY2025 YE vs PY)",  "+70.7%",            "+", "From $664M to $1,134M (continuing ops)"),
            ("FY2026E Revenue Growth Guidance",   "+19.6% (mid)",      "+", "Organic + Maritime; Megger not in guidance yet"),
        ]),
        ("PROFITABILITY METRICS", [
            ("Gross Margin FY2025",               "42.1%",             "+", "Expanded ~110bps vs FY2024; A&D mix shift helps"),
            ("Adj. EBIT Margin FY2025",           "20.3%",             "+", "Expanded 180bps vs FY2024 (~18.5%)"),
            ("A&D Adj. EBIT Margin (FY2025)",     "~27%",              "+", "High-quality Navy segment margin"),
            ("USG Adj. EBIT Margin (FY2025)",     "~29%",              "++","Best-in-class utility testing margins"),
            ("Test Adj. EBIT Margin (FY2025)",    "~15%",              "=", "Lowest segment; competitive market"),
            ("Net Income Margin (cont. ops)",     "10.6%",             "=", "GAAP compressed by M&A amortization"),
        ]),
        ("CASH FLOW METRICS", [
            ("Operating Cash Flow FY2025",        "$200M",             "+", "+$79M vs FY2024; strong conversion"),
            ("Normalized FCF FY2025",             "~$164M",            "+", "FCF margin ~15%; high quality"),
            ("FCF Conversion (FCF/Net Income)",   "~141%",             "+", "Non-cash charges boost conversion"),
            ("CapEx / Revenue",                   "3.3%",              "+", "Asset-light relative to revenue base"),
            ("Q1 FY2026 Operating CF",            "$68.9M (+101% YoY)","++","Very strong start to FY2026"),
        ]),
        ("BALANCE SHEET & LEVERAGE", [
            ("Cash (FY2025 YE)",                  "$101.3M",           "+", "Adequate pre-Megger; partially used for deal"),
            ("Long-term Debt (FY2025 YE)",        "$166M",             "+", "Low leverage pre-Megger"),
            ("Net Debt / EBITDA (pre-Megger)",    "~0.3x",             "+", "Conservative; about to change dramatically"),
            ("Pro-Forma Net Debt / EBITDA",       "~4.0x",             "–", "Post-Megger: highest leverage in company history"),
            ("Current Ratio",                     "~1.8x",             "+", "Comfortable liquidity"),
            ("Goodwill / Total Assets",           "~46% (FY2025)",     "=", "Elevated; will rise post-Megger to ~56%"),
        ]),
        ("VALUATION & RETURNS", [
            ("Market Cap (est.)",                 "~$5.57B",           "=", "Based on ~$215/share × 25.9M shares"),
            ("Enterprise Value",                  "~$5.63B",           "=", "EV = Mkt cap + net debt"),
            ("Trailing Adj. P/E",                 "~34x",              "–", "Rich; priced for continued strong execution"),
            ("Forward P/E (FY2026E)",              "~27x",              "=", "More reasonable if guidance is met"),
            ("EV / EBITDA (trailing)",            "~21x",              "–", "Premium to industrials peers"),
            ("EV / EBITDA (forward)",             "~17x",              "=", "More in-line if EBITDA grows ~25%"),
            ("ROIC (Q1 FY2026 annualized)",       "10.1%",             "=", "Slim spread over WACC; Megger will compress"),
            ("ROE FY2025",                        "~8%",               "=", "Low-ish; goodwill drag from acquisitions"),
            ("Dividend Yield",                    "~0.2%",             "=", "Not an income stock; growth focus"),
            ("Analyst Median Target",             "$300 (+39.5%)",     "+", "Strong Buy consensus; priced for upside"),
        ]),
        ("STRATEGIC SCORECARD", [
            ("Moat Strength",                     "NARROW-TO-WIDE",    "+", "A&D sole-source + Doble switching costs"),
            ("Management Quality",                "B+",                "+", "Owner-like incentives; bold M&A; proven integrators"),
            ("Capital Allocation",                "B+",                "+", "SM&P excellent; Megger the big test"),
            ("Revenue Visibility",                "HIGH",              "+", "Record backlog; 1.92x book-to-bill"),
            ("Earnings Quality",                  "HIGH",              "+", "Clean accounting; FCF > net income"),
            ("Growth Reinvestment",               "HIGH",              "+", "Megger = massive reinvestment; synergy upside"),
            ("Margin of Safety (current price)",  "LIMITED",           "–", "Stock near DCF intrinsic value; 50% YTD run"),
            ("Overall Rating",                    "HOLD / ACCUMULATE", "=", "Great business; fair price. Buy on pullback to $180-190"),
        ]),
    ]

    row = 4
    for section_name, items in sections:
        sec_hdr(ws, row, 1, 5, section_name, bg=MID_NAVY)
        row += 1
        header_row(ws, row, ["Indicator", "Value", "Signal", "Notes"], height=24)
        row += 1
        for i, item in enumerate(items):
            label, val, signal, note = item
            sig_bg = (DARK_GREEN if signal == "++" else
                      (DARK_GREEN if signal == "+" else
                       (RED if signal == "--" else
                        (RED if signal == "–" else
                         MID_NAVY))))
            row_bg = (LIGHT_GREEN if signal in ("++", "+") else
                      (LIGHT_RED if signal in ("--", "–") else GREY))
            wc(ws, row, 1, label, bold=True, bg=row_bg, fg="000000", align=left(), size=FONT_SIZE, border=True)
            wc(ws, row, 2, val, bold=True, bg=row_bg, fg="000000", align=center(), size=FONT_SIZE, border=True)
            c3 = ws.cell(row=row, column=3, value=signal)
            c3.font = font(size=FONT_SIZE, bold=True, color=WHITE)
            c3.fill = fill(sig_bg)
            c3.alignment = center()
            c3.border = thin_border()
            ws.merge_cells(f"D{row}:E{row}")
            wc(ws, row, 4, note, bold=False, bg=row_bg, fg="000000", align=left(), size=FONT_SIZE, border=True)
            ws.row_dimensions[row].height = 24
            row += 1
        row += 1


# ════════════════════════════════════════════════════════════════════════════
# MAIN
# ════════════════════════════════════════════════════════════════════════════
def main():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    build_cover(wb)
    build_business(wb)
    build_moat(wb)
    build_income(wb)
    build_balance(wb)
    build_cashflow(wb)
    build_roc(wb)
    build_management(wb)
    build_risks(wb)
    build_valuation(wb)
    build_sentiment(wb)
    build_indicators(wb)

    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"Saved: {OUTPUT_PATH}")

if __name__ == "__main__":
    import os
    main()
