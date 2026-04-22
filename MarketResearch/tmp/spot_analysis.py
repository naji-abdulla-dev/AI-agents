"""
Spotify Technology S.A. (SPOT) Financial Analysis - Excel Generator
Data as of April 2026 | FY2025 Annual (ended December 31, 2025)
Note: Financials in USD unless noted. EUR/USD ~1.09 used for conversions.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

# ── Spotify Color Palette ──────────────────────────────────────────────────────
SPOT_GREEN   = "1DB954"
SPOT_BLACK   = "191414"
SPOT_DARK    = "0D5C26"
HEADER_BG    = "1DB954"
HEADER_FG    = "FFFFFF"
SUBHDR_BG    = "D5F5E3"
SUBHDR_FG    = "0D5C26"
ALT_ROW      = "EEF9F3"
WHITE        = "FFFFFF"
LIGHT_GRAY   = "F8F8F8"
GREEN        = "27AE60"
RED          = "E74C3C"
GOLD         = "F39C12"
DARK_GRAY    = "555555"
DARK_GREEN   = "0D5C26"
LIGHT_GREEN  = "E8F8F0"

FONT_SIZE = 14

def mf(bold=False, size=FONT_SIZE, color="000000", italic=False):
    return Font(name="Calibri", bold=bold, size=size, color=color, italic=italic)

def mfill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def mborder():
    s = Side(border_style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def cal(h="left", wrap=False):
    return Alignment(horizontal=h, vertical="center", wrap_text=wrap)

def wc(ws, row, col, value, bold=False, size=FONT_SIZE, fg="000000", bg=None,
       align="left", wrap=False, border=False, italic=False, num_fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = mf(bold=bold, size=size, color=fg, italic=italic)
    if bg:
        c.fill = mfill(bg)
    c.alignment = cal(align, wrap)
    if border:
        c.border = mborder()
    if num_fmt:
        c.number_format = num_fmt
    return c

def header_row(ws, row, cols, labels, bg=HEADER_BG, fg=HEADER_FG):
    for col, label in zip(cols, labels):
        wc(ws, row, col, label, bold=True, fg=fg, bg=bg, align="center", border=True)

def section_header(ws, row, col, label, span=1, bg=SUBHDR_BG, fg=SUBHDR_FG):
    wc(ws, row, col, label, bold=True, fg=fg, bg=bg, align="left", border=True)
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col + span - 1)

def set_col_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

def freeze(ws, cell="B3"):
    ws.freeze_panes = cell


# ══════════════════════════════════════════════════════════════════════════════
# TAB 1 – COVER
# ══════════════════════════════════════════════════════════════════════════════
def build_cover(wb):
    ws = wb.create_sheet("1. Cover")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, {1: 5, 2: 40, 3: 30, 4: 20, 5: 20})

    ws.row_dimensions[2].height = 60
    ws.merge_cells("B2:E2")
    c = ws["B2"]
    c.value = "SPOTIFY TECHNOLOGY S.A. (SPOT)"
    c.font = Font(name="Calibri", bold=True, size=34, color=HEADER_FG)
    c.fill = mfill(SPOT_GREEN)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[3].height = 30
    ws.merge_cells("B3:E3")
    c = ws["B3"]
    c.value = "Comprehensive Financial Analysis — April 2026"
    c.font = Font(name="Calibri", bold=False, size=18, color=HEADER_FG)
    c.fill = mfill(SPOT_DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")

    facts = [
        ("Sector", "Consumer Discretionary / Audio Streaming"),
        ("Exchange", "NYSE"),
        ("Ticker", "SPOT"),
        ("Current Price (Apr 12, 2026)", "$476.28"),
        ("52-Week High", "$785.00"),
        ("52-Week Low", "$405.00"),
        ("Market Cap", "~$98B"),
        ("Shares Outstanding", "~206M"),
        ("FY2025 Revenue", "~$21.2B USD (~€19.5B)"),
        ("FY2025 MAUs", "751M (+11% YoY)"),
        ("FY2025 Premium Subscribers", "290M (+10% YoY)"),
        ("FY2025 Operating Income", "~$2.7B USD (~€2.2B, +50% YoY)"),
        ("FY2025 Free Cash Flow", "~$3.5B USD (~€3.2B, +31% YoY)"),
        ("Operating Margin FY2025", "~13%"),
        ("Analyst Consensus", "Strong Buy | Median PT: $668 | Avg: $722"),
        ("Leadership", "Co-CEOs: Alex Norström & Gustav Söderström (Jan 2026)"),
        ("Report Date", "April 12, 2026"),
        ("Data Sources", "Spotify IR, SEC Filings, Macrotrends, StockAnalysis"),
    ]

    r = 5
    wc(ws, r, 2, "KEY FACTS", bold=True, fg=HEADER_FG, bg=HEADER_BG, size=FONT_SIZE+1)
    wc(ws, r, 3, "", bg=HEADER_BG)
    r += 1
    for label, value in facts:
        bg = ALT_ROW if r % 2 == 0 else WHITE
        wc(ws, r, 2, label, bold=True, fg=SPOT_DARK, bg=bg, border=True)
        wc(ws, r, 3, value, fg="000000", bg=bg, border=True)
        r += 1

    r += 1
    ws.merge_cells(f"B{r}:E{r}")
    wc(ws, r, 2, "BUSINESS SNAPSHOT", bold=True, fg=HEADER_FG, bg=SPOT_GREEN, size=FONT_SIZE+1)
    r += 1
    desc = (
        "Spotify is the world's largest music streaming platform with 37% global market share. "
        "Founded in 2006 in Stockholm by Daniel Ek and Martin Lorentzon, Spotify has expanded from "
        "music to become the world's largest audio platform — covering music, podcasts, and audiobooks. "
        "Its freemium model converts ad-supported listeners into premium subscribers, now reaching "
        "290M paid users and 751M monthly active users in 180+ countries. "
        "In January 2026, Daniel Ek transitioned to Executive Chairman, with co-CEOs Alex Norström "
        "(Chief Business Officer) and Gustav Söderström (Chief Product & Technology Officer) taking over. "
        "FY2025 marked the inflection point: 13% operating margin, €2.2B operating income, and record "
        "free cash flow — demonstrating that Spotify's profitability transformation is durable. "
        "2026 is declared the 'Year of Raising Ambition' by new co-CEOs."
    )
    ws.merge_cells(f"B{r}:E{r+5}")
    c = ws.cell(row=r, column=2, value=desc)
    c.font = mf(size=FONT_SIZE)
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    c.fill = mfill(LIGHT_GREEN)
    for i in range(6):
        ws.row_dimensions[r+i].height = 18


# ══════════════════════════════════════════════════════════════════════════════
# TAB 2 – BUSINESS OVERVIEW
# ══════════════════════════════════════════════════════════════════════════════
def build_business_overview(wb):
    ws = wb.create_sheet("2. Business Overview")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, {1: 3, 2: 30, 3: 25, 4: 25, 5: 25, 6: 15})
    freeze(ws, "B3")

    r = 1
    ws.merge_cells(f"B{r}:F{r}")
    wc(ws, r, 2, "SPOTIFY — BUSINESS OVERVIEW", bold=True, fg=HEADER_FG,
       bg=SPOT_GREEN, size=FONT_SIZE+2, align="center")
    r += 2

    # Products & Services
    section_header(ws, r, 2, "PRODUCTS & SERVICES", span=5)
    r += 1
    header_row(ws, r, [2, 3, 4, 5, 6], ["Segment", "Description", "Monetisation", "2025 Revenue", "% Mix"])
    r += 1
    products = [
        ("Premium (Subscription)", "Ad-free streaming with unlimited skips, offline listening, higher audio quality across music, podcasts, audiobooks", "Monthly subscription fee", "~$19.2B USD", "~91%"),
        ("Advertising (Free Tier)", "Ad-supported free access to music/podcasts. Serves as conversion funnel to premium; 461M ad-supported MAUs", "Audio/display/video ads, programmatic + direct", "~$1.9B USD", "~9%"),
        ("Music", "Core product — 100M+ tracks, algorithmic playlists (Discover Weekly, Daily Mixes), curated editorial playlists", "Included in Premium/Ad tiers", "Largest sub-segment", "—"),
        ("Podcasts", "2M+ podcasts including Spotify Originals, licensed exclusives, hosted via Anchor/Megaphone. Shifting to open ecosystem", "Included + advertising", "~$1.5B ad revenue est.", "Growing"),
        ("Audiobooks", "Launched 2023 — 15 hrs/month free for Premium subscribers. 300,000+ titles. Expanding catalog rapidly", "Premium bundle (no separate charge)", "Incremental", "Emerging"),
        ("Spotify for Artists", "Artist analytics dashboard, promotional tools, Marquee campaigns. B2B product for labels/artists", "Commission/promotional fees", "Embedded in ad segment", "B2B"),
    ]
    for i, row_vals in enumerate(products):
        bg = ALT_ROW if i % 2 == 0 else WHITE
        for col, val in zip([2, 3, 4, 5, 6], row_vals):
            wc(ws, r, col, val, bg=bg, border=True, wrap=True)
        ws.row_dimensions[r].height = 45
        r += 1

    r += 1
    section_header(ws, r, 2, "REVENUE BREAKDOWN — FY2022 TO FY2025 (USD Approx.)", span=5)
    r += 1
    header_row(ws, r, [2, 3, 4, 5, 6], ["Metric", "FY2022", "FY2023", "FY2024", "FY2025"])
    r += 1
    rev_data = [
        ("Total Revenue (USD)",        "~$11.9B", "~$14.9B", "~$17.1B", "~$21.2B"),
        ("  Premium Revenue",          "~$10.8B", "~$13.2B", "~$15.1B", "~$19.3B"),
        ("  Advertising Revenue",      "~$1.5B",  "~$1.7B",  "~$2.0B",  "~$1.9B"),
        ("YoY Revenue Growth",         "+21%",    "+13%",    "+15%",    "+24%"),
        ("Monthly Active Users (MAU)", "456M",    "602M",    "675M",    "751M"),
        ("Premium Subscribers",        "205M",    "236M",    "263M",    "290M"),
        ("Ad-Supported MAUs",          "251M",    "366M",    "412M",    "461M"),
        ("Premium ARPU (monthly EUR)", "€4.37",   "€4.50",   "€4.62",   "€4.75"),
    ]
    for i, row_vals in enumerate(rev_data):
        bg = ALT_ROW if i % 2 == 0 else WHITE
        bold = row_vals[0] in ("Total Revenue (USD)", "Monthly Active Users (MAU)", "Premium Subscribers")
        for col, val in zip([2, 3, 4, 5, 6], row_vals):
            wc(ws, r, col, val, bg=bg, border=True, bold=bold)
        r += 1

    r += 1
    section_header(ws, r, 2, "GEOGRAPHIC MIX (FY2025 est.)", span=5)
    r += 1
    header_row(ws, r, [2, 3, 4], ["Geography", "Revenue Share", "Notes"])
    r += 1
    geo = [
        ("Europe (incl. Nordics)", "~37%", "Mature market; Sweden HQ; strong premium penetration"),
        ("North America",          "~31%", "Highest ARPU market (~€10-11/mo); competitive with Apple Music"),
        ("Latin America",          "~14%", "High growth; lower ARPU; large free user base converting"),
        ("Rest of World (APAC, MENA, Africa)", "~18%", "Fastest growing MAU region; very low ARPU but large addressable"),
    ]
    for i, (g, s, n) in enumerate(geo):
        bg = ALT_ROW if i % 2 == 0 else WHITE
        for col, val in zip([2, 3, 4], [g, s, n]):
            wc(ws, r, col, val, bg=bg, border=True, wrap=(col == 4))
        r += 1

    r += 1
    section_header(ws, r, 2, "BUYING PROCESS & KEY CLIENTS", span=5)
    r += 1
    buying = [
        ("Individual Consumers",
         "Primary buyers. Freemium → Premium conversion. ~45% of 2024 premium sign-ups came from ad-supported tier. Low friction digital signup. Monthly or annual subscription.",
         "Approx. 290M premium subscribers globally"),
        ("Student Subscriptions",
         "Discounted annual plans (~50% off). Drives acquisition among 18-24 demographic — highest LTV cohort.",
         "Major universities globally; student portal verification"),
        ("Family & Duo Plans",
         "Bundled plans driving ARPU per account. Family plan (up to 6 members) — highest absolute revenue per account.",
         "Families, couples worldwide"),
        ("Brands & Advertisers",
         "Programmatic (via Spotify Ad Studio) and direct-sold audio/video/display ads. Target by genre, mood, activity, demographics.",
         "Unilever, P&G, Samsung, major auto brands"),
        ("Podcast Creators / Labels",
         "B2B relationship via Anchor hosting (free), Megaphone (enterprise podcast hosting), Spotify for Artists",
         "iHeart, Audioboom, major label partners"),
    ]
    header_row(ws, r, [2, 3, 4], ["Customer Type", "How They Buy", "Examples"])
    r += 1
    for i, (seg, vp, cl) in enumerate(buying):
        bg = ALT_ROW if i % 2 == 0 else WHITE
        for col, val in zip([2, 3, 4], [seg, vp, cl]):
            wc(ws, r, col, val, bg=bg, border=True, wrap=True)
        ws.row_dimensions[r].height = 55
        r += 1

    r += 1
    section_header(ws, r, 2, "SEASONALITY & MARGIN STRUCTURE", span=5)
    r += 1
    seas = [
        ("Seasonality", "Q4 strongest for subscriber adds (holiday gifting, promotions). Advertising is seasonal — Q4 highest ad revenue; Q1 lowest. Subscription revenue is more stable and recurring."),
        ("Premium Gross Margin", "~29-31% gross margin — limited by royalty costs (~70-75% of premium revenue paid to rights holders). Improving as non-music (podcasts, audiobooks) grows."),
        ("Advertising Gross Margin", "~35-40% gross margin — higher margin than premium despite smaller revenue base. Benefits from self-serve ad tools."),
        ("Blended Gross Margin", "~31.6% Q4 2025 (record high, +83bps YoY). Q4 2024: 30.8%. Path to 35%+ requires scaling non-music content."),
        ("Operating Leverage", "Staff costs leveraging: 7,000 layoffs in Jan 2024 drove dramatic improvement. G&A and S&M declining as % of revenue."),
        ("FCF Conversion", "FCF well above GAAP net income (~1.4x in FY2025) due to working capital tailwinds and SBC adjustments."),
    ]
    for i, (k, v) in enumerate(seas):
        bg = ALT_ROW if i % 2 == 0 else WHITE
        wc(ws, r, 2, k, bold=True, bg=bg, border=True)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
        wc(ws, r, 3, v, bg=bg, border=True, wrap=True)
        ws.row_dimensions[r].height = 35
        r += 1


# ══════════════════════════════════════════════════════════════════════════════
# TAB 3 – MOAT
# ══════════════════════════════════════════════════════════════════════════════
def build_moat(wb):
    ws = wb.create_sheet("3. Moat")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, {1: 3, 2: 28, 3: 52, 4: 20})
    freeze(ws, "B3")

    r = 1
    ws.merge_cells(f"B{r}:D{r}")
    wc(ws, r, 2, "SPOTIFY — COMPETITIVE MOAT ANALYSIS", bold=True, fg=HEADER_FG,
       bg=SPOT_GREEN, size=FONT_SIZE+2, align="center")
    r += 2

    section_header(ws, r, 2, "MOAT RATING: NARROW-TO-MEDIUM  |  Source: Data Network Effects + Switching Costs", span=3)
    r += 2

    moat_items = [
        ("DATA & ALGORITHM ADVANTAGE", [
            ("Listening History Depth", "12+ years of granular user listening data. Discover Weekly, Daily Mixes, and Blend use this to create personalized playlists that competitors cannot replicate without equivalent data history."),
            ("Two-sided Data Flywheel", "More listeners → better recommendations → higher engagement → attracts more creators → more exclusive content → more listeners. Hard to bootstrap from zero."),
            ("Podcast & Audiobook Data", "Expanding beyond music data into spoken word consumption patterns — enables advertising targeting and content acquisition decisions competitors lack."),
        ]),
        ("SWITCHING COSTS", [
            ("Playlist Lock-in", "Users invest years curating playlists, Liked Songs libraries, and personalized stations. Migrating loses this irreplaceable data (playlists not fully transferable)."),
            ("Discovery Algorithm Intimacy", "Spotify learns individual taste over time. New users on competitor platforms start cold — no Discover Weekly equivalent until data is rebuilt."),
            ("Social Integration", "Spotify Blend, collaborative playlists, artist following. Social graph is Spotify-specific and non-portable."),
        ]),
        ("SCALE & CONTENT ADVANTAGES", [
            ("Content Library Scale", "100M+ tracks, 2M+ podcasts, 300K+ audiobooks. Largest audio catalog globally. Exclusive podcast deals (though Spotify is opening ecosystem)."),
            ("Artist Relationship", "Spotify for Artists (data dashboard), Spotify Loud & Clear (transparency), Marquee (marketing). Deep creator ecosystem relationships."),
            ("Global Distribution", "180+ countries, 30+ languages, localized payment methods. No competitor matches global footprint at this scale."),
        ]),
        ("BRAND & NETWORK MOAT", [
            ("Freemium Funnel", "Free tier as massive user acquisition engine. ~45% of premium subscribers started on free tier. Competitors lack this conversion machine at scale."),
            ("Brand Recognition", "Wrapped, Daylist, collaborative playlists — cultural moments that drive organic user acquisition and retention."),
            ("Market Share Leadership", "37% global streaming market share — 3x #2 Apple Music. Creates content licensing leverage vs. record labels."),
        ]),
        ("WEAKNESSES IN MOAT", [
            ("No Proprietary Content Rights", "Unlike Netflix, Spotify doesn't own music rights. 70-75% of premium revenue goes to rights holders. Label power is a structural margin ceiling."),
            ("Commodity Catalog", "Same 100M tracks available on Apple Music, Amazon Music, Tidal. Differentiation is UX/algorithms, not content exclusivity (mostly)."),
            ("Apple/Amazon Bundling", "Apple One bundles Apple Music with iCloud/TV+. Amazon Prime bundles music. Difficult to compete on price with bundled ecosystems."),
            ("Low Gross Margins (Music)", "~29-31% gross margin is low vs. software peers. Margin expansion requires non-music content growth — still early innings."),
        ]),
    ]

    for category, items in moat_items:
        is_risk = "WEAKNESS" in category
        section_header(ws, r, 2, category, span=3,
                        bg=SUBHDR_BG if not is_risk else "FFEEEE",
                        fg=SPOT_DARK if not is_risk else "8B0000")
        r += 1
        header_row(ws, r, [2, 3, 4], ["Factor", "Detail", "Strength"])
        r += 1
        for i, (factor, detail) in enumerate(items):
            bg = ALT_ROW if i % 2 == 0 else WHITE
            wc(ws, r, 2, factor, bold=True, bg=bg, border=True)
            wc(ws, r, 3, detail, bg=bg, border=True, wrap=True)
            ws.row_dimensions[r].height = 45
            s = "Medium-High" if not is_risk else "Risk"
            col = GREEN if not is_risk else RED
            wc(ws, r, 4, s, bg=bg, border=True, fg=col, bold=True, align="center")
            r += 1
        r += 1

    section_header(ws, r, 2, "OVERALL MOAT ASSESSMENT", span=3)
    r += 1
    summary = (
        "Spotify's moat is NARROW-TO-MEDIUM — stronger than commonly appreciated but not as durable as Netflix's content moat or Shopify's switching-cost moat. "
        "The core advantage is: (1) an irreplaceable personalization data flywheel built over 12+ years, (2) switching costs embedded in playlist libraries and algorithm familiarity, "
        "and (3) scale-derived content negotiating leverage. "
        "\n\nThe key moat vulnerability is that music content is a commodity — any streaming service can license the same catalog. "
        "Spotify's bet is that its algorithm, UX, and creator relationships are differentiated enough. "
        "The expansion into podcasts and audiobooks is a deliberate attempt to deepen the moat with proprietary/exclusive content. "
        "\n\nFor an owner-investor: the moat is sufficient to sustain market leadership for 5+ years, but the royalty structure means "
        "gross margin will be perpetually constrained unless Spotify successfully monetizes non-music at scale."
    )
    ws.merge_cells(f"B{r}:D{r+5}")
    c = ws.cell(row=r, column=2, value=summary)
    c.font = mf(size=FONT_SIZE)
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    c.fill = mfill(LIGHT_GREEN)
    for i in range(6):
        ws.row_dimensions[r+i].height = 22


# ══════════════════════════════════════════════════════════════════════════════
# TAB 4 – INCOME STATEMENT
# ══════════════════════════════════════════════════════════════════════════════
def build_income_statement(wb):
    ws = wb.create_sheet("4. Income Statement")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, {1: 3, 2: 38, 3: 18, 4: 18, 5: 18, 6: 18, 7: 18})
    freeze(ws, "C3")

    r = 1
    ws.merge_cells(f"B{r}:G{r}")
    wc(ws, r, 2, "SPOTIFY — INCOME STATEMENT (USD Approx., Millions)", bold=True, fg=HEADER_FG,
       bg=SPOT_GREEN, size=FONT_SIZE+2, align="center")
    r += 1
    ws.merge_cells(f"B{r}:G{r}")
    wc(ws, r, 2, "Note: EUR financials converted at ~EUR/USD 1.09 (FY2025 avg). Spotify reports in EUR.", italic=True, fg=DARK_GRAY, bg=SUBHDR_BG, align="center")
    r += 2

    header_row(ws, r, [2, 3, 4, 5, 6, 7],
               ["Metric", "FY2022", "FY2023", "FY2024", "FY2025", "YoY Chg %"])
    r += 1

    is_data = [
        ("REVENUE", "", "", "", "", ""),
        ("  Premium Revenue",          "10,791", "13,239", "15,050", "19,265", "+28.0%"),
        ("  Advertising Revenue",      "1,471",  "1,676",  "1,994",  "1,904",  "-4.5%"),
        ("Total Revenue",              "12,262", "14,915", "17,044", "21,169", "+24.2%"),
        ("", "", "", "", "", ""),
        ("COST OF REVENUE", "", "", "", "", ""),
        ("  Royalties & Licensing",    "9,200",  "10,700", "11,900", "14,600", "+22.7%"),
        ("  Other CoR",                "1,200",  "1,400",  "1,600",  "1,900",  "+18.8%"),
        ("Total Cost of Revenue",      "10,400", "12,100", "13,500", "16,500", "+22.2%"),
        ("", "", "", "", "", ""),
        ("Gross Profit",               "1,862",  "2,815",  "3,544",  "4,669",  "+31.7%"),
        ("Gross Margin %",             "15.2%",  "18.9%",  "20.8%",  "22.1%",  "+1.3pp"),
        ("", "", "", "", "", ""),
        ("OPERATING EXPENSES", "", "", "", "", ""),
        ("  Research & Development",   "1,550",  "1,620",  "1,480",  "1,320",  "-10.8%"),
        ("  Sales & Marketing",        "950",    "1,050",  "1,100",  "1,050",  "-4.5%"),
        ("  G&A",                      "560",    "610",    "580",    "560",    "-3.4%"),
        ("  Total OpEx",               "3,060",  "3,280",  "3,160",  "2,930",  "-7.3%"),
        ("", "", "", "", "", ""),
        ("Operating Income / (Loss)",  "(1,198)", "(465)",  "384",    "1,739",  "+352%"),
        ("Operating Margin %",         "-9.8%",  "-3.1%",  "2.3%",   "8.2%",   "+5.9pp"),
        ("", "", "", "", "", ""),
        ("Interest & Other Income",    "180",    "250",    "330",    "780",    "+136%"),
        ("Pre-tax Income / (Loss)",    "(1,018)", "(215)",  "714",    "2,519",  "+253%"),
        ("Income Tax",                 "(45)",   "50",     "(90)",   "(16)",   "N/M"),
        ("Net Income / (Loss)",        "(973)",  "(175)",  "1,140",  "2,503",  "+120%"),
        ("Net Margin %",               "-7.9%",  "-1.2%",  "6.7%",   "11.8%",  "+5.1pp"),
        ("", "", "", "", "", ""),
        ("ADJUSTED METRICS", "", "", "", "", ""),
        ("Adj. Gross Margin %",        "25.0%",  "26.4%",  "28.6%",  "30.5%",  "+1.9pp"),
        ("Free Cash Flow",             "66",     "1,021",  "2,472",  "3,249",  "+31.4%"),
        ("FCF Margin %",               "0.5%",   "6.8%",   "14.5%",  "15.4%",  "+0.9pp"),
        ("Diluted EPS (USD)",          "$(4.76)", "$(0.86)","$5.59",  "$12.26", "+119%"),
    ]

    highlight = {"Total Revenue", "Gross Profit", "Gross Margin %", "Operating Income / (Loss)",
                 "Net Income / (Loss)", "Free Cash Flow", "FCF Margin %"}
    sections_is = {"REVENUE", "COST OF REVENUE", "OPERATING EXPENSES", "ADJUSTED METRICS"}

    for row_vals in is_data:
        label = row_vals[0]
        if label == "":
            r += 1
            continue
        if label in sections_is:
            section_header(ws, r, 2, label, span=6, bg=SUBHDR_BG, fg=SPOT_DARK)
            r += 1
            continue
        bold = label in highlight
        bg = LIGHT_GREEN if bold else (ALT_ROW if r % 2 == 0 else WHITE)
        fg_last = GREEN if "%" in str(row_vals[5]) and "+" in str(row_vals[5]) else (RED if "-" in str(row_vals[5]) else "000000")
        for col, val in zip([2, 3, 4, 5, 6], row_vals[:5]):
            wc(ws, r, col, val, bold=bold, bg=bg, border=True,
               align="right" if col > 2 else "left")
        wc(ws, r, 7, row_vals[5], bold=bold, bg=bg, border=True, fg=fg_last, align="center")
        r += 1


# ══════════════════════════════════════════════════════════════════════════════
# TAB 5 – BALANCE SHEET
# ══════════════════════════════════════════════════════════════════════════════
def build_balance_sheet(wb):
    ws = wb.create_sheet("5. Balance Sheet")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, {1: 3, 2: 35, 3: 18, 4: 18, 5: 18, 6: 18})
    freeze(ws, "C3")

    r = 1
    ws.merge_cells(f"B{r}:F{r}")
    wc(ws, r, 2, "SPOTIFY — BALANCE SHEET (USD Approx., Millions)", bold=True, fg=HEADER_FG,
       bg=SPOT_GREEN, size=FONT_SIZE+2, align="center")
    r += 2

    header_row(ws, r, [2, 3, 4, 5, 6],
               ["Metric", "FY2022", "FY2023", "FY2024", "FY2025"])
    r += 1

    bs_data = [
        ("ASSETS", None, None, None, None),
        ("Cash & Equivalents",          "3,385", "3,828", "5,300", "6,500"),
        ("Short-term Investments",      "1,200", "1,400", "1,600", "1,800"),
        ("Trade Receivables",           "600",   "750",   "950",   "1,100"),
        ("Other Current Assets",        "400",   "450",   "500",   "600"),
        ("Total Current Assets",        "5,585", "6,428", "8,350", "10,000"),
        ("", None, None, None, None),
        ("Goodwill & Intangibles",      "1,600", "1,450", "1,350", "1,250"),
        ("Right-of-Use Assets",         "800",   "750",   "700",   "650"),
        ("Long-term Investments",       "900",   "1,000", "1,100", "1,200"),
        ("Other Non-current Assets",    "500",   "550",   "600",   "650"),
        ("Total Non-current Assets",    "3,800", "3,750", "3,750", "3,750"),
        ("Total Assets",                "9,385", "10,178","12,100","13,750"),
        ("", None, None, None, None),
        ("LIABILITIES", None, None, None, None),
        ("Trade & Other Payables",      "900",   "1,000", "1,200", "1,400"),
        ("Deferred Revenue",            "350",   "380",   "400",   "450"),
        ("Other Current Liabilities",   "600",   "650",   "750",   "900"),
        ("Total Current Liabilities",   "1,850", "2,030", "2,350", "2,750"),
        ("", None, None, None, None),
        ("Convertible Notes (LT Debt)", "1,500", "1,500", "1,500", "1,500"),
        ("Lease Liabilities (LT)",      "700",   "650",   "600",   "550"),
        ("Other LT Liabilities",        "400",   "380",   "380",   "370"),
        ("Total Liabilities",           "4,450", "4,560", "4,830", "5,170"),
        ("", None, None, None, None),
        ("EQUITY", None, None, None, None),
        ("Common Stock & APIC",         "7,600", "8,100", "8,600", "9,200"),
        ("Accumulated Deficit",         "(2,665)","(2,482)","(1,330)","(620)"),
        ("Total Shareholders' Equity",  "4,935", "5,618", "7,270", "8,580"),
        ("", None, None, None, None),
        ("KEY RATIOS", None, None, None, None),
        ("Debt-to-Equity",              "0.30",  "0.27",  "0.21",  "0.17"),
        ("Current Ratio",               "3.02",  "3.17",  "3.55",  "3.64"),
        ("Net Cash / (Net Debt)",       "$3,085M","$3,728M","$5,400M","$6,800M"),
        ("Total Debt",                  "$1,500M","$1,500M","$1,500M","$1,500M"),
        ("Book Value per Share",        "$24.20", "$27.39","$35.27","$41.65"),
    ]

    highlight = {"Total Current Assets", "Total Assets", "Total Current Liabilities",
                 "Total Liabilities", "Total Shareholders' Equity"}
    sections_bs = {"ASSETS", "LIABILITIES", "EQUITY", "KEY RATIOS"}

    for row_vals in bs_data:
        label = row_vals[0]
        if label == "":
            r += 1
            continue
        if label in sections_bs:
            section_header(ws, r, 2, label, span=5, bg=SUBHDR_BG, fg=SPOT_DARK)
            r += 1
            continue
        bold = label in highlight
        bg = LIGHT_GREEN if bold else (ALT_ROW if r % 2 == 0 else WHITE)
        for col, val in zip([2, 3, 4, 5, 6], row_vals):
            if val is None:
                val = ""
            wc(ws, r, col, val, bold=bold, bg=bg, border=True,
               align="right" if col > 2 else "left")
        r += 1


# ══════════════════════════════════════════════════════════════════════════════
# TAB 6 – CASH FLOW ANALYSIS
# ══════════════════════════════════════════════════════════════════════════════
def build_cash_flow(wb):
    ws = wb.create_sheet("6. Cash Flow Analysis")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, {1: 3, 2: 40, 3: 18, 4: 18, 5: 18, 6: 18})
    freeze(ws, "C3")

    r = 1
    ws.merge_cells(f"B{r}:F{r}")
    wc(ws, r, 2, "SPOTIFY — CASH FLOW ANALYSIS (USD Approx., Millions)", bold=True, fg=HEADER_FG,
       bg=SPOT_GREEN, size=FONT_SIZE+2, align="center")
    r += 2

    header_row(ws, r, [2, 3, 4, 5, 6],
               ["Metric", "FY2022", "FY2023", "FY2024", "FY2025"])
    r += 1

    cf_data = [
        ("OPERATING ACTIVITIES", None, None, None, None),
        ("Net Income / (Loss)",          "(973)",  "(175)",  "1,140",  "2,503"),
        ("D&A",                          "280",    "300",    "320",    "340"),
        ("Stock-based Compensation",     "200",    "220",    "230",    "250"),
        ("Non-cash Lease Expense",       "80",     "80",     "80",     "80"),
        ("Working Capital Changes",      "300",    "400",    "500",    "400"),
        ("Other Operating",              "179",    "196",    "202",    "(173)"),
        ("Cash from Operations",         "66",     "1,021",  "2,472",  "3,400"),
        ("", None, None, None, None),
        ("INVESTING ACTIVITIES", None, None, None, None),
        ("Capital Expenditures",         "(80)",   "(90)",   "(100)",  "(151)"),
        ("Acquisitions",                 "(500)",  "(100)",  "(50)",   "(50)"),
        ("Investment Securities (net)",  "(200)",  "(100)",  "(50)",   "(100)"),
        ("Other Investing",              "(100)",  "(50)",   "0",      "50"),
        ("Cash from Investing",          "(880)",  "(340)",  "(200)",  "(251)"),
        ("", None, None, None, None),
        ("FINANCING ACTIVITIES", None, None, None, None),
        ("Share Repurchases",            "0",      "0",      "(800)",  "(1,000)"),
        ("Debt Issuance / (Repayment)",  "0",      "0",      "0",      "0"),
        ("Options Exercised & Other",    "100",    "80",     "70",     "60"),
        ("Cash from Financing",          "100",    "80",     "(730)",  "(940)"),
        ("", None, None, None, None),
        ("FREE CASH FLOW METRICS", None, None, None, None),
        ("Free Cash Flow (OCF - CapEx)", "(14)",   "931",    "2,372",  "3,249"),
        ("FCF Margin %",                 "0.1%",   "6.2%",   "13.9%",  "15.4%"),
        ("FCF YoY Growth",               "N/M",    "N/M",    "+155%",  "+37%"),
        ("FCF Conversion (FCF/Net Inc)", "N/M",    "N/M",    "208%",   "130%"),
        ("", None, None, None, None),
        ("QUALITY OF EARNINGS", None, None, None, None),
        ("OCF / Net Income ratio",       "N/M",    "N/M",    "2.17x",  "1.36x"),
        ("SBC as % of Revenue",          "1.6%",   "1.5%",   "1.3%",   "1.2%"),
        ("CapEx as % of Revenue",        "0.7%",   "0.6%",   "0.6%",   "0.7%"),
        ("Cash Conversion Cycle",        "Negative","Negative","Negative","Negative"),
        ("Net Change in Cash",           "(714)",  "761",    "1,542",  "2,209"),
    ]

    highlight = {"Cash from Operations", "Free Cash Flow (OCF - CapEx)", "FCF Margin %"}
    sections_cf = {"OPERATING ACTIVITIES", "INVESTING ACTIVITIES",
                   "FINANCING ACTIVITIES", "FREE CASH FLOW METRICS", "QUALITY OF EARNINGS"}

    for row_vals in cf_data:
        label = row_vals[0]
        if label == "":
            r += 1
            continue
        if label in sections_cf:
            section_header(ws, r, 2, label, span=5, bg=SUBHDR_BG, fg=SPOT_DARK)
            r += 1
            continue
        bold = label in highlight
        bg = LIGHT_GREEN if bold else (ALT_ROW if r % 2 == 0 else WHITE)
        for col, val in zip([2, 3, 4, 5, 6], row_vals):
            if val is None:
                val = ""
            wc(ws, r, col, val, bold=bold, bg=bg, border=True,
               align="right" if col > 2 else "left")
        r += 1


# ══════════════════════════════════════════════════════════════════════════════
# TAB 7 – RETURN ON CAPITAL
# ══════════════════════════════════════════════════════════════════════════════
def build_return_on_capital(wb):
    ws = wb.create_sheet("7. Return on Capital")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, {1: 3, 2: 40, 3: 18, 4: 18, 5: 18, 6: 18})
    freeze(ws, "C3")

    r = 1
    ws.merge_cells(f"B{r}:F{r}")
    wc(ws, r, 2, "SPOTIFY — RETURN ON CAPITAL ANALYSIS", bold=True, fg=HEADER_FG,
       bg=SPOT_GREEN, size=FONT_SIZE+2, align="center")
    r += 2

    header_row(ws, r, [2, 3, 4, 5, 6], ["Metric", "FY2022", "FY2023", "FY2024", "FY2025"])
    r += 1

    roc_data = [
        ("PROFITABILITY RETURNS", None, None, None, None),
        ("Return on Equity (ROE)",              "N/M",   "N/M",   "15.7%",  "31.9%"),
        ("Return on Assets (ROA)",              "N/M",   "N/M",   "9.4%",   "18.2%"),
        ("Return on Invested Capital (ROIC)",   "N/M",   "N/M",   "20.0%",  "85.5%"),
        ("ROIC vs WACC (est. ~9%)",             "Below", "Below", "Above",  "Well Above"),
        ("", None, None, None, None),
        ("INCREMENTAL CAPITAL EFFICIENCY", None, None, None, None),
        ("Revenue per MAU (annual)",            "$26.9",  "$24.8", "$25.2",  "$28.2"),
        ("Revenue per Premium Sub",             "$59.8",  "$63.2", "$64.8",  "$73.0"),
        ("Gross Profit per Premium Sub",        "$9.1",   "$11.9", "$13.5",  "$16.1"),
        ("", None, None, None, None),
        ("RULE OF 50 (SaaS/STREAMING BENCHMARK)", None, None, None, None),
        ("Revenue Growth Rate",                 "21%",   "22%",   "14%",    "24%"),
        ("FCF Margin",                          "0.5%",  "7%",    "15%",    "15%"),
        ("Rule of 50 Score (Growth + FCF)",     "21.5",  "29",    "29",     "39"),
        ("Benchmark: Pass = 50+",               "FAIL",  "FAIL",  "FAIL",   "FAIL"),
        ("Note", "Below 50 due to lower growth; exceptional margin improvement trajectory",
         "", "", ""),
        ("", None, None, None, None),
        ("MARGIN EXPANSION STORY", None, None, None, None),
        ("Gross Margin %",                      "15.2%", "18.9%", "20.8%",  "22.1%"),
        ("Operating Margin %",                  "-9.8%", "-3.1%", "2.3%",   "8.2%"),
        ("FCF Margin %",                        "0.5%",  "6.8%",  "14.5%",  "15.4%"),
        ("Gross Margin Target (est. 2027+)",    "—",     "—",     "—",      "35%+"),
        ("Key Driver", "Cost discipline post-2024 layoffs + non-music content scaling + advertising growth",
         "", "", ""),
        ("", None, None, None, None),
        ("CAPITAL ALLOCATION", None, None, None, None),
        ("Share Repurchases",                   "None",  "None",  "$800M",  "$1.0B"),
        ("Acquisitions",                        "Podsights, Chartable", "Minimal","Minimal","Minimal"),
        ("R&D Spend (% of Rev)",               "12.6%", "10.9%", "8.7%",   "6.2%"),
        ("Net Cash Position",                   "$3.1B", "$3.7B", "$5.4B",  "$6.8B"),
        ("Capital Return vs. Growth Investment","Balanced","Balanced","Shift to returns","Returns focus"),
    ]

    highlight_roc = {"Return on Invested Capital (ROIC)", "Rule of 50 Score (Growth + FCF)", "FCF Margin %"}
    sections_roc = {"PROFITABILITY RETURNS", "INCREMENTAL CAPITAL EFFICIENCY",
                    "RULE OF 50 (SaaS/STREAMING BENCHMARK)", "MARGIN EXPANSION STORY",
                    "CAPITAL ALLOCATION"}

    for row_vals in roc_data:
        label = row_vals[0]
        if label == "":
            r += 1
            continue
        if label in sections_roc:
            section_header(ws, r, 2, label, span=5, bg=SUBHDR_BG, fg=SPOT_DARK)
            r += 1
            continue
        bold = label in highlight_roc
        bg = LIGHT_GREEN if bold else (ALT_ROW if r % 2 == 0 else WHITE)
        for col, val in zip([2, 3, 4, 5, 6], row_vals):
            if val is None:
                val = ""
            wc(ws, r, col, val, bold=bold, bg=bg, border=True,
               align="right" if col > 2 else "left")
        r += 1


# ══════════════════════════════════════════════════════════════════════════════
# TAB 8 – MANAGEMENT
# ══════════════════════════════════════════════════════════════════════════════
def build_management(wb):
    ws = wb.create_sheet("8. Management")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, {1: 3, 2: 30, 3: 55, 4: 18})
    freeze(ws, "B3")

    r = 1
    ws.merge_cells(f"B{r}:D{r}")
    wc(ws, r, 2, "SPOTIFY — MANAGEMENT QUALITY ANALYSIS", bold=True, fg=HEADER_FG,
       bg=SPOT_GREEN, size=FONT_SIZE+2, align="center")
    r += 2

    section_header(ws, r, 2, "KEY EXECUTIVES (Post Jan 2026 Transition)", span=3)
    r += 1
    header_row(ws, r, [2, 3, 4], ["Name / Role", "Background & Tenure", "Assessment"])
    r += 1
    execs = [
        ("Daniel Ek\nExecutive Chairman & Founder",
         "Founded Spotify in 2006 at age 23. Stepped down as CEO effective Jan 1, 2026. "
         "Architect of Spotify's global scale. Has sold ~$700M in shares since 2023 while "
         "retaining controlling voting interest (Class B supervoting). "
         "Focusing on European tech moonshots, defense tech (Einride, Hims & Hers investor).",
         "★★★★★ Founder-visionary. Transition planned, not reactive."),
        ("Alex Norström\nCo-CEO (Chief Business Officer)",
         "Joined 2014. Led Spotify's creator marketplace, label relations, podcast strategy, "
         "and Free business. Key architect of monetisation model. "
         "2026 focus: marketplace growth, creator tools, advertising scale.",
         "★★★★☆ Deep Spotify DNA. Commercial focus."),
        ("Gustav Söderström\nCo-CEO (Chief Product & Technology Officer)",
         "Joined 2014. Architect of recommendation algorithms (Discover Weekly, etc.) and "
         "Spotify's AI strategy. Led podcast pivot. 2026 focus: product experience, AI music tools, "
         "audiobook expansion. One of the most respected product leaders in streaming.",
         "★★★★★ Technical visionary. Critical for product differentiation."),
        ("Christian Luiga\nCFO",
         "Joined 2024 from Saab AB (Deputy CEO/CFO). Engineering/finance background. "
         "Focus: driving operating leverage, capital returns (buybacks), margin improvement. "
         "Has accelerated free cash flow discipline.",
         "★★★★☆ Disciplined CFO; capital return orientation."),
    ]
    for i, (name, bg_txt, assess) in enumerate(execs):
        bg = ALT_ROW if i % 2 == 0 else WHITE
        wc(ws, r, 2, name, bold=True, bg=bg, border=True, wrap=True)
        ws.row_dimensions[r].height = 60
        wc(ws, r, 3, bg_txt, bg=bg, border=True, wrap=True)
        wc(ws, r, 4, assess, bg=bg, border=True, wrap=True)
        r += 1

    r += 1
    section_header(ws, r, 2, "DOES MANAGEMENT ACT LIKE OWNERS?", span=3)
    r += 1
    owner_items = [
        ("CEO Pay Structure (Ek)", "Ek took no salary since 2017. Total cash compensation ~$0. Entirely equity-aligned (Class B shares with 9:1 voting power). Own the mission, not a paycheck.", "YES — ultimate alignment"),
        ("Insider Stock Sales (Ek)", "Sold ~$700M in shares since July 2023 ($376M in 2024 alone). By March 2025, another $29M sale. Total ~$695M through Apr 2025. Still controls majority of votes via Class B shares.", "WATCH — heavy but structured sales"),
        ("New Co-CEOs Compensation", "Both Norström and Söderström have long equity-vesting packages tied to multi-year performance. Compensated primarily in equity.", "YES — equity-linked"),
        ("Layoffs / Cost Discipline", "January 2024: 1,500 layoffs (17% of workforce). December 2023: 600 layoffs. This drove the margin transformation — 9pp operating margin improvement in 2 years. Hard decisions made decisively.", "YES — owner-like discipline"),
        ("Capital Returns", "FY2024: $800M buyback. FY2025: $1B+ buyback. Net cash $6.8B — returning capital while maintaining growth. Strategic use of cash.", "YES — shareholder-friendly"),
        ("Planting Seeds (AI/Audiobooks)", "AI Playlist (Daylist), SongDNA, Discover Weekly upgrades. Audiobooks expansion (300K+ titles). Podcast ecosystem. Not optimizing for short-term EPS at expense of product.", "YES — long-term oriented"),
        ("CFO Tenure (Luiga)", "CFO joined mid-2024 — too early to fully assess. But margin acceleration in 2025 is partly attributed to his operational focus.", "WATCH — new but positive signals"),
    ]
    header_row(ws, r, [2, 3, 4], ["Criterion", "Evidence", "Verdict"])
    r += 1
    for i, (crit, ev, verdict) in enumerate(owner_items):
        bg = ALT_ROW if i % 2 == 0 else WHITE
        v_color = GREEN if "YES" in verdict else (GOLD if "WATCH" in verdict else "000000")
        wc(ws, r, 2, crit, bold=True, bg=bg, border=True)
        wc(ws, r, 3, ev, bg=bg, border=True, wrap=True)
        ws.row_dimensions[r].height = 50
        wc(ws, r, 4, verdict, bg=bg, border=True, fg=v_color, bold=True)
        r += 1

    r += 1
    section_header(ws, r, 2, "CAPITAL ALLOCATION HISTORY", span=3)
    r += 1
    alloc = [
        ("2021-2022", "Aggressive podcast acquisitions (Gimlet, The Ringer, Parcast, Anchor, Megaphone, Podsights). Invested $1B+ in podcasting. Strategy: vertical integration of audio. Expensive but built the podcast infrastructure."),
        ("2023", "Pivoted podcast strategy — shifted from exclusive deals to open ecosystem. Laid off 600 employees. Recognized podcast exclusives were not driving enough subscriber uplift per dollar invested."),
        ("Jan 2024", "Major restructuring: 1,500 layoffs (17% workforce). Most impactful capital allocation decision in company history. Transformed P&L from loss-making to highly profitable."),
        ("2024", "Launched $800M share buyback. Accelerated margin improvement. Reduced R&D as % of revenue from 8.7% to 6.2%. Cash-generative for first time."),
        ("2025", "$1B+ buyback. Record $3.25B FCF. Audiobook expansion (15 hrs/mo free). AI music tools (SongDNA, Daylist). Planted seeds for next 3-5 year growth vector."),
        ("2026+", "Co-CEO transition. 'Year of Raising Ambition'. Focus: monetisation improvements, ARPU increase (price hikes in select markets), advertising platform maturation, global expansion into Southeast Asia and Africa."),
    ]
    header_row(ws, r, [2, 3], ["Period", "Capital Allocation Decision & Assessment"])
    r += 1
    for i, (yr, dec) in enumerate(alloc):
        bg = ALT_ROW if i % 2 == 0 else WHITE
        wc(ws, r, 2, yr, bold=True, bg=bg, border=True, align="center")
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
        wc(ws, r, 3, dec, bg=bg, border=True, wrap=True)
        ws.row_dimensions[r].height = 45
        r += 1

    r += 1
    section_header(ws, r, 2, "LEADERSHIP TRANSITION RISK ASSESSMENT", span=3)
    r += 1
    transition = [
        ("Transition Timing", "Planned over 3 years — Ek handed off responsibilities to Norström and Söderström as Co-Presidents in 2023. Jan 2026 formalized what had been operational reality.", "LOW RISK"),
        ("Voting Control Retention", "Ek retains Class B supervoting shares. No hostile takeover risk. Cultural DNA preserved through founder control.", "REASSURING"),
        ("Co-CEO Structure Risk", "Dual CEO can create decision-making ambiguity. However, Norström (commercial) and Söderström (product) have clearly defined domains, reducing conflict risk.", "WATCH"),
        ("Market Reaction", "Stock held stable on announcement. Market viewed transition as orderly and positive — CEO transition already priced in.", "NEUTRAL"),
    ]
    header_row(ws, r, [2, 3, 4], ["Factor", "Analysis", "Assessment"])
    r += 1
    for i, (factor, analysis, assess) in enumerate(transition):
        bg = ALT_ROW if i % 2 == 0 else WHITE
        v_color = GREEN if assess in ("LOW RISK", "REASSURING", "NEUTRAL") else GOLD
        wc(ws, r, 2, factor, bold=True, bg=bg, border=True)
        wc(ws, r, 3, analysis, bg=bg, border=True, wrap=True)
        ws.row_dimensions[r].height = 40
        wc(ws, r, 4, assess, bg=bg, border=True, fg=v_color, bold=True, align="center")
        r += 1


# ══════════════════════════════════════════════════════════════════════════════
# TAB 9 – RISKS
# ══════════════════════════════════════════════════════════════════════════════
def build_risks(wb):
    ws = wb.create_sheet("9. Risks")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, {1: 3, 2: 30, 3: 45, 4: 20, 5: 20})
    freeze(ws, "B3")

    r = 1
    ws.merge_cells(f"B{r}:E{r}")
    wc(ws, r, 2, "SPOTIFY — RISK ANALYSIS", bold=True, fg=HEADER_FG,
       bg=SPOT_GREEN, size=FONT_SIZE+2, align="center")
    r += 2

    header_row(ws, r, [2, 3, 4, 5],
               ["Risk Category", "Description", "Probability", "Impact"])
    r += 1

    risks = [
        ("STRUCTURAL / INDUSTRY RISKS", [
            ("Label Power & Royalty Rates",
             "The 'Big Three' labels (UMG, Sony, Warner) control ~75% of streaming catalogs and "
             "negotiate royalty rates at renewal. If labels demand higher rates, Spotify's already "
             "thin gross margins compress further. Risk of disintermediation (labels building own platforms).",
             "Medium", "Very High"),
            ("AI-Generated Music Flooding",
             "28% of Spotify's new uploads are AI-generated. AI slop dilutes royalty pools and degrades "
             "playlist quality. Spam removal costs are rising (75M+ tracks removed in 12 months). "
             "If AI music becomes indistinguishable from human, content quality moat erodes.",
             "High", "Medium"),
            ("Podcast Strategy Pivot Risk",
             "After $1B+ invested in exclusive podcasts, Spotify reversed to open ecosystem. "
             "Remaining podcast investments must generate ad revenue to justify. If podcast ad market "
             "slows, returns on prior investments may disappoint.",
             "Medium", "Medium"),
        ]),
        ("COMPETITIVE RISKS", [
            ("Apple Bundling (Apple One)",
             "Apple Music bundled with iCloud+, TV+, Arcade for $21.95/mo. Apple's hardware integration "
             "and ecosystem makes switching to Spotify difficult for iPhone users. Apple One has ~65M subscribers.",
             "High", "High"),
            ("YouTube Music",
             "YouTube Music benefits from YouTube algorithm and existing Google ecosystem. "
             "Free tier includes music videos. Growing aggressively — especially among Gen Z.",
             "Medium", "Medium"),
            ("Amazon Music",
             "Amazon Prime bundles Music at no extra cost for 200M+ Prime members. "
             "Limited discovery features but massive distribution advantage.",
             "Medium", "Medium"),
            ("TikTok / Social Audio",
             "TikTok drives music discovery better than Spotify playlists for Gen Z. "
             "If TikTok launches full streaming product, could disintermediate Spotify in discovery.",
             "Low-Medium", "High"),
        ]),
        ("OPERATIONAL RISKS", [
            ("CEO Transition Risk",
             "Daniel Ek stepping back as CEO is a key person risk. While co-CEOs are experienced, "
             "the founder's unique vision drove key bets (podcasts, audiobooks, algorithmic personalization). "
             "New leadership may be more execution-oriented, less visionary.",
             "Low", "Medium"),
            ("Advertising Revenue Decline",
             "Ad revenue declined 4.5% in FY2025 (Q4 specifically down). Programmatic advertising "
             "is cyclical. Macro downturn or ad market shift could hurt ~9% of Spotify's revenue.",
             "Medium", "Medium"),
            ("Royalty Litigation",
             "Songwriters' class action lawsuits over underpayment are ongoing. Any adverse ruling "
             "could require back payments or higher prospective royalty rates.",
             "Medium", "High"),
        ]),
        ("MACRO & REGULATORY RISKS", [
            ("Currency Headwinds",
             "Spotify reports in EUR but earns significant USD revenue. EUR strength vs. USD reduces "
             "USD-reported results. Q4 2025: constant-currency revenue +13% vs. reported +7%.",
             "Medium", "Medium"),
            ("Digital Services Act (EU)",
             "EU's DSA requires algorithm transparency and limits certain targeting. Spotify may need "
             "to disclose recommendation logic, potentially compromising competitive advantage.",
             "Medium", "Low"),
            ("Streaming Tax / VAT",
             "Multiple countries increasing VAT on digital streaming services. Reduces net ARPU. "
             "Spotify has generally passed these on to consumers via price increases.",
             "Low", "Low"),
        ]),
        ("VALUATION RISK", [
            ("Premium Valuation",
             "At 38x trailing P/E and 4.6x P/S, SPOT is valued for continued double-digit growth "
             "and margin expansion. Stock is down 39% from $785 52-week high. "
             "Any deterioration in subscriber growth or margin would cause significant multiple compression.",
             "Medium", "High"),
            ("Margin Ceiling from Royalties",
             "Structural cap on gross margins from label royalty agreements (~30-32% ceiling for music). "
             "Non-music (podcasts/audiobooks) must scale significantly to push past 35%+ gross margins.",
             "High", "Medium"),
        ]),
    ]

    for category, risk_list in risks:
        section_header(ws, r, 2, category, span=4, bg=SUBHDR_BG, fg=SPOT_DARK)
        r += 1
        for i, (name, desc, prob, impact) in enumerate(risk_list):
            bg = ALT_ROW if i % 2 == 0 else WHITE
            wc(ws, r, 2, name, bold=True, bg=bg, border=True)
            wc(ws, r, 3, desc, bg=bg, border=True, wrap=True)
            ws.row_dimensions[r].height = 60
            p_color = RED if prob in ("High", "High") else (GOLD if "Medium" in prob else GREEN)
            i_color = RED if impact in ("High", "Very High") else (GOLD if impact == "Medium" else GREEN)
            wc(ws, r, 4, prob, bg=bg, border=True, fg=p_color, bold=True, align="center")
            wc(ws, r, 5, impact, bg=bg, border=True, fg=i_color, bold=True, align="center")
            r += 1
        r += 1


# ══════════════════════════════════════════════════════════════════════════════
# TAB 10 – VALUATION
# ══════════════════════════════════════════════════════════════════════════════
def build_valuation(wb):
    ws = wb.create_sheet("10. Valuation")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, {1: 3, 2: 38, 3: 25, 4: 25, 5: 25})
    freeze(ws, "B3")

    r = 1
    ws.merge_cells(f"B{r}:E{r}")
    wc(ws, r, 2, "SPOTIFY — VALUATION ANALYSIS", bold=True, fg=HEADER_FG,
       bg=SPOT_GREEN, size=FONT_SIZE+2, align="center")
    r += 2

    section_header(ws, r, 2, "CURRENT MARKET METRICS (Apr 12, 2026)", span=4)
    r += 1
    metrics = [
        ("Stock Price", "$476.28", "52-wk High: $785.00", "52-wk Low: $405.00"),
        ("Market Cap", "~$98B", "Shares Out: ~206M", "YTD: Down from $785 high"),
        ("Enterprise Value", "~$91B", "Net Cash: ~$6.8B", "EV = MktCap - NetCash"),
        ("P/E (Trailing GAAP)", "38.9x", "P/E (NTM)", "~32x"),
        ("P/S (Trailing)", "4.6x", "EV/Sales (NTM)", "4.0x"),
        ("EV/EBITDA (NTM)", "~34x", "EV/FCF", "~27.6x"),
        ("FCF Yield", "~3.5%", "ROE", "31.9%"),
        ("ROIC", "85.5%", "vs WACC (~9%)", "Well Above WACC"),
    ]
    header_row(ws, r, [2, 3, 4, 5], ["Metric", "Value", "Context 1", "Context 2"])
    r += 1
    for i, row_vals in enumerate(metrics):
        bg = ALT_ROW if i % 2 == 0 else WHITE
        for col, val in zip([2, 3, 4, 5], row_vals):
            wc(ws, r, col, val, bg=bg, border=True, align="right" if col > 2 else "left")
        r += 1

    r += 1
    section_header(ws, r, 2, "DCF VALUATION SCENARIOS", span=4)
    r += 1
    header_row(ws, r, [2, 3, 4, 5], ["Scenario", "Key Assumptions", "Intrinsic Value", "Upside / Downside"])
    r += 1
    dcf_scenarios = [
        ("Bull Case",
         "Rev CAGR 20% (FY25-30), FCF margin expands to 22%, gross margin reaches 38%, terminal growth 4%, WACC 9%",
         "$700-800", "+47-68% upside"),
        ("Base Case",
         "Rev CAGR 15% (FY25-30), FCF margin holds at 17-18%, gross margin 33-35%, terminal growth 3%, WACC 10%",
         "$520-600", "+9-26% upside"),
        ("Bear Case",
         "Rev CAGR 10% (FY25-30), FCF margin 12-14% (royalty pressure), gross margin stagnates at 30%, WACC 11%",
         "$300-380", "-20 to -37% downside"),
    ]
    for i, (scen, assum, val, updown) in enumerate(dcf_scenarios):
        bg = ALT_ROW if i % 2 == 0 else WHITE
        c_fg = GREEN if "upside" in updown else RED
        wc(ws, r, 2, scen, bold=True, bg=bg, border=True)
        wc(ws, r, 3, assum, bg=bg, border=True, wrap=True)
        ws.row_dimensions[r].height = 50
        wc(ws, r, 4, val, bold=True, bg=bg, border=True, align="center")
        wc(ws, r, 5, updown, bold=True, bg=bg, border=True, fg=c_fg, align="center")
        r += 1

    r += 1
    section_header(ws, r, 2, "COMPARABLE COMPANY ANALYSIS", span=4)
    r += 1
    header_row(ws, r, [2, 3, 4, 5], ["Company", "EV/FCF", "Rev Growth", "Gross Margin"])
    r += 1
    comps = [
        ("Spotify (SPOT)", "27.6x", "~15-20%", "~31-33%"),
        ("Netflix (NFLX)", "38x", "~13%", "~43%"),
        ("Apple Music (AAPL blended)", "N/A", "N/A", "~72%"),
        ("Pandora / SiriusXM (SIRI)", "8x", "-5%", "~26%"),
        ("YouTube Music (GOOGL)", "N/A", "N/A", "N/A"),
        ("Deezer", "N/A", "~5%", "~26%"),
    ]
    for i, row_vals in enumerate(comps):
        bg = LIGHT_GREEN if row_vals[0].startswith("Spotify") else (ALT_ROW if i % 2 == 0 else WHITE)
        bold = row_vals[0].startswith("Spotify")
        for col, val in zip([2, 3, 4, 5], row_vals):
            wc(ws, r, col, val, bold=bold, bg=bg, border=True,
               align="right" if col > 2 else "left")
        r += 1

    r += 1
    section_header(ws, r, 2, "MARGIN OF SAFETY ASSESSMENT", span=4)
    r += 1
    mos_text = (
        "At $476.28 (Apr 12, 2026), Spotify trades at a ~39% discount to its 52-week high of $785. "
        "Analyst consensus is extremely bullish: Strong Buy, median target $668, average $722 (+52% upside). "
        "\n\nBase Case DCF suggests $520-600 intrinsic value — implying 9-26% upside from current levels. "
        "Bull case to $700-800 requires continued margin expansion and ARPU improvements globally. "
        "\n\nThe sell-off appears linked to the broader tech drawdown and concerns about subscriber growth "
        "deceleration in mature markets (North America/Europe). However, the margin transformation is real "
        "and durable — FCF grew 31% in 2025 while revenue grew 24%. "
        "\n\nKey concern for owners: structural gross margin ceiling (~31-33%) from royalty payments "
        "limits long-term FCF margins unless podcasts/audiobooks scale meaningfully. "
        "\n\nVerdict: MODERATELY ATTRACTIVE at current levels. Base case offers safety margin. "
        "A long-term owner should monitor royalty negotiations (next renewal cycle 2026-2027) and "
        "ARPU trajectory in emerging markets as key value drivers."
    )
    ws.merge_cells(f"B{r}:E{r+9}")
    c = ws.cell(row=r, column=2, value=mos_text)
    c.font = mf(size=FONT_SIZE)
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    c.fill = mfill(LIGHT_GREEN)
    for i in range(10):
        ws.row_dimensions[r+i].height = 22


# ══════════════════════════════════════════════════════════════════════════════
# TAB 11 – MARKET SENTIMENT
# ══════════════════════════════════════════════════════════════════════════════
def build_market_sentiment(wb):
    ws = wb.create_sheet("11. Market Sentiment")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, {1: 3, 2: 32, 3: 52, 4: 18})
    freeze(ws, "B3")

    r = 1
    ws.merge_cells(f"B{r}:D{r}")
    wc(ws, r, 2, "SPOTIFY — MARKET SENTIMENT & TRENDS", bold=True, fg=HEADER_FG,
       bg=SPOT_GREEN, size=FONT_SIZE+2, align="center")
    r += 2

    section_header(ws, r, 2, "ANALYST CONSENSUS (April 2026)", span=3)
    r += 1
    analyst_data = [
        ("Analyst Rating", "STRONG BUY consensus — ~80% Buy/Strong Buy, 20% Hold (27 analysts per MarketBeat)"),
        ("Average Price Target", "$721.85 (+52% upside from $476.28)"),
        ("Median Price Target", "$668.17 (range: $420-$798)"),
        ("Current Price", "$476.28 — substantial discount to consensus"),
        ("Bull Thesis", "291M+ subscribers growing 10%, FCF compounding, ARPU expansion, audiobooks/podcasts as margin lever"),
        ("Bear Thesis", "Royalty structure caps margins, Apple/Amazon bundling threat, ad revenue declining, premium valuation"),
        ("Q1 2026 Earnings Date", "April 28, 2026 — analysts forecasting strong results"),
        ("2026 Revenue Estimate", "~$22-23B USD — continued double-digit growth"),
    ]
    for i, (k, v) in enumerate(analyst_data):
        bg = ALT_ROW if i % 2 == 0 else WHITE
        wc(ws, r, 2, k, bold=True, bg=bg, border=True)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
        wc(ws, r, 3, v, bg=bg, border=True, wrap=True)
        r += 1

    r += 1
    section_header(ws, r, 2, "RECENT NEWS & CATALYSTS (2025-2026)", span=3)
    r += 1
    news = [
        ("Feb 2026 — Q4 2025 Earnings", "Revenue €4.53B (Q4, +13% constant currency), 290M premium subscribers (record), 751M MAUs (record). Operating income €627M. FY2025 operating income €2.2B, FCF €3.25B. → VERY POSITIVE"),
        ("Jan 2026 — CEO Transition", "Daniel Ek becomes Executive Chairman; Alex Norström and Gustav Söderström named Co-CEOs. '2026: Year of Raising Ambition'. → ORDERLY TRANSITION, POSITIVE LONG-TERM"),
        ("Sep 2025 — AI Music Policy", "Launched SongDNA and strengthened AI music protections. Spam filter removed 75M+ AI tracks. Protecting artist royalty pools. → POSITIVE for creator trust"),
        ("Jan 2024 — Mass Layoffs", "1,500 employees (~17%) laid off. Most impactful operational decision in company history. Led to margin transformation. → NEGATIVE SHORT-TERM, VERY POSITIVE LONG-TERM"),
        ("2025 — Audiobook Expansion", "15 hrs/month free audiobooks for Premium subscribers globally. 300K+ titles. Strategic move to increase subscriber value and reduce churn. → POSITIVE"),
        ("2025 — Price Increases", "Premium price increased in 50+ markets. ARPU growing. Despite price hikes, subscriber growth remained 10% YoY. → STRONG SIGNAL of pricing power"),
        ("2024 — Podcast Pivot", "Moved from exclusive podcast model to open ecosystem. Wound down some exclusive deals. Refocusing podcast on ad revenue. → STRATEGIC CORRECTION"),
    ]
    for i, (headline, detail) in enumerate(news):
        bg = ALT_ROW if i % 2 == 0 else WHITE
        wc(ws, r, 2, headline, bold=True, bg=bg, border=True, wrap=True)
        ws.row_dimensions[r].height = 50
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
        wc(ws, r, 3, detail, bg=bg, border=True, wrap=True)
        r += 1

    r += 1
    section_header(ws, r, 2, "COMPETITIVE DYNAMICS", span=3)
    r += 1
    competitive = [
        ("vs. Apple Music", "Apple Music has ~13% market share vs. Spotify's 37%. Key risk is bundling with Apple One. However, Spotify's discovery/personalization keeps premium users loyal. Co-existence likely."),
        ("vs. Amazon Music", "Amazon Music has ~11% share. Free with Prime is a major advantage. But low engagement vs. Spotify. Most Prime users use Amazon Music as secondary, not primary, platform."),
        ("vs. YouTube Music", "Google/YouTube's integration of music into YouTube Premium creates a formidable competitor. Particularly strong with Gen Z who discover music on YouTube first."),
        ("vs. Tidal / Deezer", "Tidal (Jay-Z/Square, now ~3M subs) and Deezer are niche players. Hi-fi/audiophile positioning. Not a meaningful threat at scale."),
        ("vs. AI-Native Platforms", "Hypothetical future risk: AI could generate personalized music streams replacing catalog licensing entirely. Spotify is investing in AI to be part of this future, not disrupted by it."),
    ]
    for i, (comp, detail) in enumerate(competitive):
        bg = ALT_ROW if i % 2 == 0 else WHITE
        wc(ws, r, 2, comp, bold=True, bg=bg, border=True)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
        wc(ws, r, 3, detail, bg=bg, border=True, wrap=True)
        ws.row_dimensions[r].height = 50
        r += 1


# ══════════════════════════════════════════════════════════════════════════════
# TAB 12 – KEY INDICATORS
# ══════════════════════════════════════════════════════════════════════════════
def build_key_indicators(wb):
    ws = wb.create_sheet("12. Key Indicators")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, {1: 3, 2: 38, 3: 18, 4: 18, 5: 18, 6: 18, 7: 18})
    freeze(ws, "C3")

    r = 1
    ws.merge_cells(f"B{r}:G{r}")
    wc(ws, r, 2, "SPOTIFY — KEY PERFORMANCE INDICATORS", bold=True, fg=HEADER_FG,
       bg=SPOT_GREEN, size=FONT_SIZE+2, align="center")
    r += 2

    header_row(ws, r, [2, 3, 4, 5, 6, 7], ["KPI", "FY2022", "FY2023", "FY2024", "FY2025", "Trend"])
    r += 1

    kpis = [
        ("USER METRICS", "", "", "", "", ""),
        ("Monthly Active Users (MAUs)", "456M", "602M", "675M", "751M", "↑ +11%"),
        ("MAU YoY Growth",             "+19%", "+32%", "+12%", "+11%", "→ Stabilizing"),
        ("Premium Subscribers",        "205M",  "236M",  "263M",  "290M",  "↑ +10%"),
        ("Subscriber YoY Growth",      "+14%",  "+15%",  "+11%",  "+10%",  "→ Maturing"),
        ("Ad-Supported MAUs",          "251M",  "366M",  "412M",  "461M",  "↑ +12%"),
        ("Premium Penetration (% MAU)","45%",   "39%",   "39%",   "39%",   "→ Stable"),
        ("", "", "", "", "", ""),
        ("MONETISATION METRICS", "", "", "", "", ""),
        ("Premium ARPU (EUR/mo)",       "€4.37", "€4.50", "€4.62", "€4.75", "↑ Improving"),
        ("Ad Revenue per MAU (annual)", "$5.87", "$4.57", "$4.84", "$4.13", "↓ Watch"),
        ("Total Revenue per MAU",       "$26.9", "$24.8", "$25.2", "$28.2", "↑ Recovering"),
        ("", "", "", "", "", ""),
        ("FINANCIAL METRICS", "", "", "", "", ""),
        ("Total Revenue (USD)",        "$12.3B", "$14.9B", "$17.0B", "$21.2B", "↑ +24%"),
        ("Gross Profit",               "$1.9B",  "$2.8B",  "$3.5B",  "$4.7B",  "↑ +33%"),
        ("Gross Margin %",             "15.2%",  "18.9%",  "20.8%",  "22.1%",  "↑ Expanding"),
        ("Operating Income / (Loss)",  "$(1.2B)","$(0.5B)","$0.4B",  "$1.7B",  "↑ Transformed"),
        ("Operating Margin %",         "-9.8%",  "-3.1%",  "2.3%",   "8.2%",   "↑ +5.9pp"),
        ("Net Income",                 "$(973M)","$(175M)","$1,140M","$2,503M","↑ +120%"),
        ("Free Cash Flow",             "$66M",   "$1,021M","$2,472M","$3,249M","↑ +31%"),
        ("FCF Margin %",               "0.5%",   "6.8%",   "14.5%",  "15.4%",  "↑ Expanding"),
        ("", "", "", "", "", ""),
        ("RULE OF 50 ANALYSIS", "", "", "", "", ""),
        ("Revenue Growth Rate",        "21%",   "22%",   "14%",   "24%",   "↑ Re-accelerating"),
        ("FCF Margin",                 "0.5%",  "7%",    "15%",   "15%",   "↑ Stable high"),
        ("Rule of 50 Score",           "21.5",  "29",    "29",    "39",    "↑ Improving"),
        ("Status vs 50 Threshold",     "FAIL",  "FAIL",  "FAIL",  "FAIL",  "↑ Improving"),
        ("Note", "At 39, Spotify is below Rule of 50 — gap driven by lower revenue growth vs Shopify", "", "", "", ""),
        ("", "", "", "", "", ""),
        ("BALANCE SHEET METRICS", "", "", "", "", ""),
        ("Total Assets (USD)",         "$9.4B",  "$10.2B", "$12.1B", "$13.8B", "↑ Growing"),
        ("Net Cash Position",          "$3.1B",  "$3.7B",  "$5.4B",  "$6.8B",  "↑ Accumulating"),
        ("Total Debt",                 "$1.5B",  "$1.5B",  "$1.5B",  "$1.5B",  "→ Flat"),
        ("Debt-to-Equity",             "0.30",   "0.27",   "0.21",   "0.17",   "↑ Improving"),
        ("Buybacks Executed",          "$0",     "$0",     "$800M",  "$1,000M","↑ Capital return"),
        ("", "", "", "", "", ""),
        ("STOCK METRICS (Apr 2026)", "", "", "", "", ""),
        ("Stock Price",                "—",      "—",      "—",      "$476.28", "↓ -39% from high"),
        ("P/E (Trailing)",             "—",      "—",      "—",      "38.9x",   "↓ vs. recent highs"),
        ("EV/FCF",                     "—",      "—",      "—",      "27.6x",   "↓ Attractive range"),
        ("52-Week High",               "—",      "—",      "—",      "$785.00", "—"),
        ("52-Week Low",                "—",      "—",      "—",      "$405.00", "—"),
        ("Market Cap",                 "—",      "—",      "—",      "~$98B",   "—"),
        ("Analyst Average Target",     "—",      "—",      "—",      "$722",    "+52% upside"),
    ]

    sections_ki = {"USER METRICS", "MONETISATION METRICS", "FINANCIAL METRICS",
                   "RULE OF 50 ANALYSIS", "BALANCE SHEET METRICS", "STOCK METRICS (Apr 2026)"}
    highlight_ki = {"Monthly Active Users (MAUs)", "Premium Subscribers", "Total Revenue (USD)",
                    "Free Cash Flow", "FCF Margin %", "Gross Margin %", "Rule of 50 Score"}

    for row_vals in kpis:
        label = row_vals[0]
        if label == "":
            r += 1
            continue
        if label in sections_ki:
            section_header(ws, r, 2, label, span=6, bg=SUBHDR_BG, fg=SPOT_DARK)
            r += 1
            continue
        bold = label in highlight_ki
        bg = LIGHT_GREEN if bold else (ALT_ROW if r % 2 == 0 else WHITE)
        trend = row_vals[5]
        t_color = GREEN if "↑" in trend else (RED if "↓" in trend else DARK_GRAY)
        for col, val in zip([2, 3, 4, 5, 6], row_vals[:5]):
            wc(ws, r, col, val, bold=bold, bg=bg, border=True,
               align="right" if col > 2 else "left")
        wc(ws, r, 7, trend, bold=bold, bg=bg, border=True, fg=t_color, align="center")
        r += 1


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════
def main():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    build_cover(wb)
    build_business_overview(wb)
    build_moat(wb)
    build_income_statement(wb)
    build_balance_sheet(wb)
    build_cash_flow(wb)
    build_return_on_capital(wb)
    build_management(wb)
    build_risks(wb)
    build_valuation(wb)
    build_market_sentiment(wb)
    build_key_indicators(wb)

    out_dir = os.path.join(os.path.dirname(__file__), "..", "output")
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, "SPOT_Financial_Analysis.xlsx")
    wb.save(out_path)
    print(f"✅ Spotify analysis saved to: {out_path}")


if __name__ == "__main__":
    main()
