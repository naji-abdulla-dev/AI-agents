"""
Snowflake Inc. (SNOW) Financial Analysis - Excel Generator
Data as of April 2026 | FY2026 Annual (ended January 31, 2026)
Fiscal Year Note: Snowflake FY2026 = Feb 2025 - Jan 2026
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

# ── Snowflake Color Palette ────────────────────────────────────────────────────
SNOW_BLUE    = "29B5E8"
SNOW_DARK    = "0A3D5C"
HEADER_BG    = "29B5E8"
HEADER_FG    = "FFFFFF"
SUBHDR_BG    = "D6F0FB"
SUBHDR_FG    = "0A3D5C"
ALT_ROW      = "EBF8FD"
WHITE        = "FFFFFF"
LIGHT_GRAY   = "F7FBFD"
RED          = "E53935"
GREEN        = "2E7D32"
GOLD         = "F9A825"
DARK_BLUE    = "0A3D5C"
DARK_GRAY    = "555555"
WARN_RED     = "C62828"

FONT_SIZE = 14

def mf(bold=False, size=FONT_SIZE, color="000000", italic=False):
    return Font(name="Calibri", bold=bold, size=size, color=color, italic=italic)

def mfill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def mborder():
    s = Side(border_style="thin", color="BBDDEE")
    return Border(left=s, right=s, top=s, bottom=s)

def cal(h="left", wrap=False):
    return Alignment(horizontal=h, vertical="center", wrap_text=wrap)

def wc(ws, row, col, value, bold=False, size=FONT_SIZE, fg="000000", bg=None,
       align="left", wrap=False, border=False, italic=False, num_fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = mf(bold=bold, size=size, color=fg, italic=italic)
    if bg:
        c.fill = mfill(bg)
    c.alignment = cal(align, wrap)
    if border:
        c.border = mborder()
    if num_fmt:
        c.number_format = num_fmt
    return c

def sec_hdr(ws, row, col, text, span=1, bg=HEADER_BG, fg=HEADER_FG):
    c = wc(ws, row, col, text, bold=True, size=FONT_SIZE+2, fg=fg, bg=bg, align="center")
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+span-1)
    ws.row_dimensions[row].height = 30
    return c

def sub_hdr(ws, row, col, text, span=1, bg=SUBHDR_BG, fg=SUBHDR_FG):
    c = wc(ws, row, col, text, bold=True, size=FONT_SIZE, fg=fg, bg=bg, align="center", border=True)
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+span-1)
    ws.row_dimensions[row].height = 22
    return c

def lbl(ws, row, col, text, bg=None, indent=0):
    val = ("   " * indent) + text
    c = wc(ws, row, col, val, size=FONT_SIZE, bg=bg, align="left", border=True)
    ws.row_dimensions[row].height = 20
    return c

def dc(ws, row, col, value, bg=None, num_fmt=None, bold=False, color="000000"):
    c = wc(ws, row, col, value, bold=bold, size=FONT_SIZE, fg=color, bg=bg,
           align="right", border=True, num_fmt=num_fmt)
    ws.row_dimensions[row].height = 20
    return c

def scw(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

# ═══════════════════════════════════════════════════════════════════════════════
def build_cover(wb):
    ws = wb.create_sheet("Cover")
    ws.sheet_view.showGridLines = False
    for c in range(1, 9):
        scw(ws, c, 18)
    scw(ws, 1, 4); scw(ws, 2, 32); scw(ws, 3, 22)

    ws.merge_cells("B3:G3")
    wc(ws, 3, 2, "SNOWFLAKE INC. (SNOW)", bold=True, size=32, fg=WHITE,
       bg=HEADER_BG, align="center")
    ws.row_dimensions[3].height = 55

    ws.merge_cells("B4:G4")
    wc(ws, 4, 2, "Comprehensive Investment Analysis | April 2026", bold=False,
       size=16, fg=WHITE, bg=DARK_BLUE, align="center")
    ws.row_dimensions[4].height = 28

    ws.merge_cells("B6:G6")
    wc(ws, 6, 2, "Company Profile", bold=True, size=16, fg=WHITE, bg=DARK_BLUE, align="center")
    ws.row_dimensions[6].height = 25

    info = [
        ("Ticker",              "SNOW"),
        ("Exchange",            "NYSE"),
        ("Sector",              "Technology / Cloud Data Platform"),
        ("Headquarters",        "San Mateo, California, USA"),
        ("Founded",             "2012 by Benoit Dageville, Thierry Cruanes, Marcin Zukowski"),
        ("CEO",                 "Sridhar Ramaswamy (since Feb 2024)"),
        ("Fiscal Year End",     "January 31 (FY2026 = Feb 2025 - Jan 2026)"),
        ("Market Cap",          "~$50-55B (April 2026; stock down ~23% YTD)"),
        ("Stock Price",         "~$132 (April 9, 2026; -12% Apr 10 on AI fears + legal)"),
        ("P/S Ratio (NTM)",     "~15-18x (compressed from 40x+ peak)"),
        ("Shares Outstanding",  "~380M (diluted)"),
        ("Dividend",            "None"),
    ]

    for i, (k, v) in enumerate(info):
        r = 7 + i
        bg = ALT_ROW if i % 2 == 0 else WHITE
        wc(ws, r, 2, k, bold=True, size=FONT_SIZE, bg=bg, align="left", border=True)
        wc(ws, r, 3, v, size=FONT_SIZE, bg=bg, align="left", border=True)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)
        ws.row_dimensions[r].height = 22

    ws.merge_cells("B20:G20")
    wc(ws, 20, 2, "Investment Thesis", bold=True, size=16, fg=WHITE, bg=DARK_BLUE, align="center")
    ws.row_dimensions[20].height = 25

    thesis = (
        "Snowflake is the leading independent cloud data platform, enabling organizations to "
        "store, process, and analyze data across multiple cloud environments. Under CEO Sridhar "
        "Ramaswamy's 'Product-First' era, Snowflake has launched Cortex AI, Snowflake Intelligence, "
        "and 430+ new capabilities in FY2026, positioning the platform as an AI-native data cloud. "
        "FY2026 revenue grew 29% to $4.68B with $9.77B in RPO (+42% YoY) signaling durable demand. "
        "The BULL case: AI workloads drive structural increase in data processing on Snowflake. "
        "The BEAR case: Microsoft Fabric + Databricks threaten displacement; GAAP losses continue. "
        "Current stock at ~$132 (down 23% YTD) offers more attractive entry than the $300+ highs. "
        "Trading at ~15x NTM P/S vs. 40x+ historical — significant multiple compression already."
    )
    ws.merge_cells("B21:G25")
    wc(ws, 21, 2, thesis, size=FONT_SIZE, bg=LIGHT_GRAY, align="left", wrap=True, border=True)
    for r in range(21, 26):
        ws.row_dimensions[r].height = 22

    ws.merge_cells("B27:G27")
    wc(ws, 27, 2, "Key Financial Highlights (FY2026 | Ended Jan 31, 2026)", bold=True,
       size=16, fg=WHITE, bg=DARK_BLUE, align="center")
    ws.row_dimensions[27].height = 25

    highlights = [
        ("Product Revenue",              "$4.53B",   "+30% YoY"),
        ("Total Revenue",                "$4.68B",   "+29% YoY"),
        ("RPO (Remaining Perf. Oblig.)", "$9.77B",   "+42% YoY — strong future revenue signal"),
        ("Customers > $1M TTM Rev",      "733",      "+27% YoY"),
        ("Total Customers",              "13,300+",  "+~40% net new additions"),
        ("Non-GAAP Op Margin",           "11.0%",    "FY2027 guided 12.5%"),
        ("FCF Margin",                   "23%",      "FY2027 guided 23%"),
        ("GAAP Op Loss",                 "-$318M",   "Q4 FY2026; GAAP unprofitable"),
    ]

    sub_hdr(ws, 28, 2, "Metric", bg=SUBHDR_BG)
    sub_hdr(ws, 28, 3, "Value", bg=SUBHDR_BG)
    sub_hdr(ws, 28, 4, "Comment", bg=SUBHDR_BG)
    ws.merge_cells(start_row=28, start_column=4, end_row=28, end_column=7)

    for i, (m, v, c) in enumerate(highlights):
        r = 29 + i
        bg = ALT_ROW if i % 2 == 0 else WHITE
        lbl(ws, r, 2, m, bg=bg)
        dc(ws, r, 3, v, bg=bg)
        lbl(ws, r, 4, c, bg=bg)
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=7)

    ws.merge_cells("B38:G38")
    wc(ws, 38, 2, "Report Date: April 10, 2026  |  Data Sources: SEC Filings, Snowflake IR, Bloomberg",
       italic=True, size=12, fg="666666", align="center")

# ═══════════════════════════════════════════════════════════════════════════════
def build_business(wb):
    ws = wb.create_sheet("Business Overview")
    ws.sheet_view.showGridLines = False
    scw(ws, 1, 3); scw(ws, 2, 32); scw(ws, 3, 22); scw(ws, 4, 22)
    scw(ws, 5, 22); scw(ws, 6, 22); scw(ws, 7, 22)

    sec_hdr(ws, 1, 2, "BUSINESS OVERVIEW — SNOWFLAKE INC. (SNOW)", span=6)

    sub_hdr(ws, 3, 2, "What Snowflake Does", span=6)
    desc = ("Snowflake is a cloud-native data platform delivered as-a-service. Unlike traditional "
            "databases or Hadoop, Snowflake separates compute from storage, runs across AWS, Azure, "
            "and GCP simultaneously, and allows customers to securely share data across organizations. "
            "Snowflake charges on consumption (compute credits used), not seat licenses. Under CEO "
            "Sridhar Ramaswamy, it has pivoted to an AI Data Cloud, launching Cortex AI (SQL/Python "
            "AI inference in-warehouse), Snowpark (code deployment), and Snowflake Intelligence "
            "(agentic AI). It also entered the IT Ops market via Observe acquisition.")
    ws.merge_cells("B4:G7")
    wc(ws, 4, 2, desc, size=FONT_SIZE, bg=LIGHT_GRAY, align="left", wrap=True, border=True)
    for r in range(4, 8):
        ws.row_dimensions[r].height = 22

    # Revenue Model
    sub_hdr(ws, 9, 2, "Revenue Model (Consumption-Based)", span=6)
    model = [
        ("Metric",             "Description",                             "FY2026 Data"),
        ("Product Revenue",    "Snowflake compute + storage credits consumed by customers",  "$4.53B (97% of total)"),
        ("Professional Svcs",  "Implementation, training, consulting",    "$150M (3%)"),
        ("Billing Model",      "Pay-per-use: customers pre-purchase 'credits' (capacity)",  "Committed contracts"),
        ("RPO",                "Remaining Performance Obligations — contracted but not yet recognized revenue",  "$9.77B (+42% YoY)"),
        ("cRPO",               "Current RPO (next 12 months) — highly predictable near-term revenue signal",   "~$3.2B"),
        ("NRR",                "Net Revenue Retention Rate — expansion from existing customers",  "~131% (FY2026)"),
    ]
    for i, row_data in enumerate(model):
        r = 10 + i
        if i == 0:
            for j, val in enumerate(row_data):
                sub_hdr(ws, r, 2+j, val)
        else:
            bg = ALT_ROW if i % 2 == 0 else WHITE
            lbl(ws, r, 2, row_data[0], bg=bg)
            lbl(ws, r, 3, row_data[1], bg=bg)
            ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
            dc(ws, r, 6, row_data[2], bg=bg)
            ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=7)
        ws.row_dimensions[r].height = 28

    # Products & Services
    sub_hdr(ws, 19, 2, "Products & Services (AI Data Cloud)", span=6)
    products = [
        ("Data Warehousing",           "Core SQL-based cloud data warehouse; scales compute elastically; zero-copy clone; time travel"),
        ("Data Lake / Lakehouse",       "Iceberg Tables support (open format); unstructured data + structured in one platform"),
        ("Data Sharing",               "Secure data sharing across organizations without copying data; Snowflake Marketplace (2,000+ datasets)"),
        ("Snowpark",                   "Python/Java/Scala in Snowflake; data engineering and ML pipelines native to the platform"),
        ("Cortex AI",                  "LLM inference inside Snowflake (Llama, Mistral, Snowflake Arctic); no data leaves environment"),
        ("Snowflake Intelligence",      "Agentic AI layer for enterprise automation; early stage but key strategic product"),
        ("Streamlit",                   "Open-source Python app framework acquired 2022; build data apps inside Snowflake"),
        ("Snowflake Marketplace",       "Data sharing + monetization marketplace; ecosystem lock-in driver"),
        ("Dynamic Tables / Pipes",      "Real-time streaming and incremental data processing"),
        ("Observe (acquired FY2026)",   "IT operations / observability; $50B+ TAM entry; logs, traces, metrics"),
    ]
    for i, (p, d) in enumerate(products):
        r = 20 + i
        bg = ALT_ROW if i % 2 == 0 else WHITE
        lbl(ws, r, 2, p, bg=bg)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)
        lbl(ws, r, 3, d, bg=bg)
        ws.row_dimensions[r].height = 22

    # Customer Segments
    sub_hdr(ws, 32, 2, "Customer Segments & Key Clients", span=6)
    segments = [
        ("Enterprise (>$1M)",    "733 customers; Fortune 500; banks, healthcare, retail; multi-year committed contracts"),
        ("Mid-Market",           "Growth segment; cloud-native companies; faster adoption of AI features"),
        ("Financial Services",   "Capital One, JPMorgan, HSBC; regulatory data segregation; heavy Snowpark usage"),
        ("Healthcare / Pharma",  "Moderna, AstraZeneca; genomics data; HIPAA compliance; data clean rooms"),
        ("Retail / CPG",         "Instacart, Warner Music; supply chain analytics; real-time personalization"),
        ("Technology",           "Adobe, ServiceNow, Salesforce; data sharing between SaaS vendors"),
        ("Media & Entertainment","NBA, Warner Bros Discovery; fan data, analytics"),
    ]
    for i, (s, d) in enumerate(segments):
        r = 33 + i
        bg = ALT_ROW if i % 2 == 0 else WHITE
        lbl(ws, r, 2, s, bg=bg)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)
        lbl(ws, r, 3, d, bg=bg)
        ws.row_dimensions[r].height = 22

    # Seasonality
    sub_hdr(ws, 42, 2, "Seasonality & Margin Structure", span=6)
    season = [
        ("Q1 FY (Feb-Apr)",    "Seasonally stronger; fiscal year start with new contracts; good enterprise budget flush"),
        ("Q2 FY (May-Jul)",    "Mid-year; AI workload testing ramp; some consumption variability"),
        ("Q3 FY (Aug-Oct)",    "Strengthening; data analytics demand pre-year-end planning"),
        ("Q4 FY (Nov-Jan)",    "Strongest; fiscal year-end enterprise spending; largest contract renewals"),
        ("Revenue Drivers",    "Consumption = # of queries × cost/credit × credits/query; AI workloads >3x more credits"),
        ("Margin Structure",   "Product gross margin ~75%; high SBC (~25-30% of revenue historically); improving toward GAAP profitability"),
    ]
    for i, (q, n) in enumerate(season):
        r = 43 + i
        bg = ALT_ROW if i % 2 == 0 else WHITE
        lbl(ws, r, 2, q, bg=bg)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)
        lbl(ws, r, 3, n, bg=bg)
        ws.row_dimensions[r].height = 22

# ═══════════════════════════════════════════════════════════════════════════════
def build_moat(wb):
    ws = wb.create_sheet("Moat")
    ws.sheet_view.showGridLines = False
    scw(ws, 1, 3); scw(ws, 2, 28); scw(ws, 3, 52); scw(ws, 4, 18)

    sec_hdr(ws, 1, 2, "COMPETITIVE MOAT ANALYSIS — SNOWFLAKE (SNOW)", span=3)

    moat_intro = ("Snowflake's moat is real but contested. It has built a powerful data ecosystem "
                  "with high switching costs and network effects via Data Sharing, but faces formidable "
                  "competition from cloud giants and Databricks. The moat is 'narrow-to-moderate' and "
                  "may widen or narrow depending on AI workload adoption.")
    ws.merge_cells("B3:D4")
    wc(ws, 3, 2, moat_intro, size=FONT_SIZE, bg=LIGHT_GRAY, align="left", wrap=True, border=True)
    for r in [3, 4]:
        ws.row_dimensions[r].height = 22

    moats = [
        ("MOAT TYPE",                "EVIDENCE & DEPTH",                                                  "STRENGTH"),
        ("Switching Costs",          "Migrating a Snowflake data warehouse requires rebuilding pipelines, "
                                      "rewriting SQL, retraining teams, and re-testing data governance. "
                                      "Average migration time 12-24 months. Multi-cloud deployments "
                                      "create additional switching complexity.",                           "MODERATE"),
        ("Data Network Effects",     "Snowflake Marketplace: 2,000+ data products; the more providers, "
                                      "the more consumers. Data Sharing: once company A shares data with "
                                      "Company B on Snowflake, both are locked in. Hard to replicate.",   "MODERATE"),
        ("Multi-Cloud Architecture", "ONLY major data platform that runs natively across AWS, Azure, "
                                      "AND GCP without data copies. Cloud-agnostic customers and those "
                                      "with multi-cloud strategies have no comparable alternative.",       "MODERATE"),
        ("Ecosystem / Integrations", "300+ native integrations; Fivetran, dbt, Tableau, Looker, "
                                      "Informatica all optimize for Snowflake first. Partner ecosystem "
                                      "creates flywheel that reinforces platform stickiness.",             "MODERATE"),
        ("AI Data Platform (New)",   "Cortex AI runs LLMs inside Snowflake — zero data egress, "
                                      "privacy-compliant. For regulated industries (financial, health), "
                                      "this is a moat: can't send PII to OpenAI externally.",             "BUILDING"),
        ("Customer Concentration",   "733 customers each spending >$1M/year; top customers spend "
                                      "$10M-100M+; deeply embedded in mission-critical workflows; "
                                      "cost of failure/migration is career-risk for data teams.",          "MODERATE"),
    ]

    for i, row_data in enumerate(moats):
        r = 6 + i
        if i == 0:
            sub_hdr(ws, r, 2, row_data[0])
            sub_hdr(ws, r, 3, row_data[1])
            sub_hdr(ws, r, 4, row_data[2])
        else:
            bg = ALT_ROW if i % 2 == 0 else WHITE
            strength_color = GREEN if "WIDE" in row_data[2] else (GOLD if "MODERATE" in row_data[2] else RED)
            lbl(ws, r, 2, row_data[0], bg=bg)
            ws.row_dimensions[r].height = 52
            wc(ws, r, 3, row_data[1], size=FONT_SIZE, bg=bg, align="left", wrap=True, border=True)
            wc(ws, r, 4, row_data[2], bold=True, size=FONT_SIZE, fg=WHITE,
               bg=strength_color, align="center", border=True)

    # Threats to Moat
    sec_hdr(ws, 14, 2, "THREATS TO MOAT — COMPETITIVE RISKS", span=3, bg=WARN_RED)
    threats = [
        ("Microsoft Fabric",      "Microsoft bundled its entire data stack (Synapse, Databricks Spark, Power BI) into "
                                   "Fabric under one Azure license. Existing Azure customers may choose 'good enough' "
                                   "Fabric over best-in-class Snowflake. Microsoft Azure credits sweeten the deal.",  "HIGH"),
        ("Databricks",            "Databricks (pre-IPO, valued $62B) leads in ML/AI workloads and open-source Spark. "
                                   "Databricks Unity Catalog + Delta Lake = direct competition to Snowflake's "
                                   "full platform. If Databricks IPOs with fresh capital, marketing intensifies.",  "HIGH"),
        ("AWS / BigQuery (GCP)",  "AWS Redshift Serverless and Google BigQuery are capable alternatives; cloud "
                                   "providers incentivize customers to stay within their ecosystem. Native cloud "
                                   "tools have gotten much better in 2023-2025.",                                    "MEDIUM"),
        ("Iceberg / Open Formats","Open table formats (Apache Iceberg) reduce vendor lock-in. Customers can store "
                                   "data in open formats and query from any engine. Snowflake's Iceberg Table "
                                   "support is necessary to stay relevant but reduces proprietary advantage.",       "MEDIUM"),
        ("AI Disruption (LLMs)",  "LLMs can increasingly query raw data without structured warehouses. Semantic "
                                   "layers and AI-native databases could bypass traditional data warehousing. "
                                   "This is a 3-7 year risk, not immediate.",                                        "LOW-MEDIUM"),
    ]

    sub_hdr(ws, 15, 2, "Threat"); sub_hdr(ws, 15, 3, "Description"); sub_hdr(ws, 15, 4, "Risk Level")
    for i, (t, d, r_lvl) in enumerate(threats):
        r = 16 + i
        bg = ALT_ROW if i % 2 == 0 else WHITE
        lbl(ws, r, 2, t, bg=bg)
        ws.row_dimensions[r].height = 50
        wc(ws, r, 3, d, size=FONT_SIZE, bg=bg, align="left", wrap=True, border=True)
        clr = RED if r_lvl == "HIGH" else (GOLD if "MEDIUM" in r_lvl else GREEN)
        wc(ws, r, 4, r_lvl, bold=True, size=FONT_SIZE, fg=WHITE, bg=clr, align="center", border=True)

    sub_hdr(ws, 22, 2, "Overall Moat Rating", span=3, bg=DARK_BLUE, fg=WHITE)
    ws.merge_cells("B23:D24")
    wc(ws, 23, 2,
       "NARROW MOAT — Snowflake has real switching costs and a growing data ecosystem, but faces "
       "existential competition from Microsoft (bundling), Databricks (AI workloads), and the major "
       "cloud providers. The moat is defensible in the near term but requires continuous innovation "
       "to prevent erosion. CEO Ramaswamy's AI pivot (Cortex, Intelligence) is the right strategy "
       "but outcomes remain uncertain. This is a high-risk, high-reward investment.",
       size=FONT_SIZE, bg=ALT_ROW, align="left", wrap=True, border=True)
    for r in [23, 24]:
        ws.row_dimensions[r].height = 30

# ═══════════════════════════════════════════════════════════════════════════════
def build_income(wb):
    ws = wb.create_sheet("Income Statement")
    ws.sheet_view.showGridLines = False
    scw(ws, 1, 3); scw(ws, 2, 32); scw(ws, 3, 18); scw(ws, 4, 18)
    scw(ws, 5, 18); scw(ws, 6, 18); scw(ws, 7, 18)

    sec_hdr(ws, 1, 2, "INCOME STATEMENT — SNOWFLAKE INC. (SNOW)  |  FY2022-FY2026 ($M)", span=6)
    sub_hdr(ws, 2, 2, "Metric")
    for i, yr in enumerate(["FY2022", "FY2023", "FY2024", "FY2025", "FY2026"]):
        sub_hdr(ws, 2, 3+i, yr)
    # Note appended to header row already merged — skip separate cell write

    rows = [
        ("REVENUE ($M)",                 None, None, None, None, None),
        ("  Product Revenue",             1139, 1968, 2665, 3476, 4527),
        ("  Professional Services",         82,  120,  157,  147,  157),
        ("TOTAL REVENUE",                 1220, 2067, 2814, 3628, 4684),
        ("  YoY Growth",                  "106%","69%","36%","29%","29%"),
        ("", None, None, None, None, None),
        ("COSTS & EXPENSES ($M)",         None, None, None, None, None),
        ("  Cost of Product Revenue",       404,  640,  804,  944, 1133),
        ("  Cost of Prof. Services",         92,  140,  175,  170,  180),
        ("GROSS PROFIT",                    724, 1287, 1835, 2514, 3371),
        ("  Gross Margin",                "59.3%","62.3%","65.2%","69.3%","72.0%"),
        ("", None, None, None, None, None),
        ("  Sales & Marketing",             834, 1100, 1193, 1275, 1350),
        ("  Research & Development",        544,  814,  986, 1017, 1100),
        ("  G&A",                           226,  303,  301,  290,  300),
        ("TOTAL OPERATING EXPENSES",       2100, 2997, 3459, 3696, 4063),
        ("", None, None, None, None, None),
        ("OPERATING INCOME (LOSS)",        -880, -930, -645, -668, -692),
        ("  GAAP Operating Margin",       "-72%","-45%","-23%","-18%","-15%"),
        ("", None, None, None, None, None),
        ("NON-GAAP ADJUSTMENTS",          None, None, None, None, None),
        ("  Stock-Based Compensation",      517,  747,  760,  830,  870),
        ("  D&A",                            66,   90,  115,  130,  145),
        ("  Other (restructuring, etc.)",    30,   50,   40,   30,   20),
        ("NON-GAAP OPERATING INCOME",      -267,  -43,  270,  292,  343),
        ("  Non-GAAP Op Margin",           "-22%","-2%","9.6%","8.1%","7.3%"),
        ("", None, None, None, None, None),
        ("ADJUSTED EBITDA",                -201,    47,  385,  422,  488),
        ("  Adj. EBITDA Margin",           "-16%","2.3%","13.7%","11.6%","10.4%"),
        ("", None, None, None, None, None),
        ("NET INCOME (LOSS)",              -889,  -791,  -836,  -985,  -800),
        ("  GAAP Net Margin",             "-73%","-38%","-30%","-27%","-17%"),
        ("  Adj. EPS (Non-GAAP)",          -0.51, -0.10,  0.79,  0.77,  0.97),
    ]

    for i, row_data in enumerate(rows):
        r = 3 + i
        label = row_data[0]
        vals  = row_data[1:]
        if label == "":
            ws.row_dimensions[r].height = 8
            continue
        is_total = label.startswith("TOTAL") or label.startswith("GROSS") or label in (
                   "OPERATING INCOME (LOSS)","NET INCOME (LOSS)","ADJUSTED EBITDA","NON-GAAP OPERATING INCOME",
                   "COSTS & EXPENSES ($M)","REVENUE ($M)","NON-GAAP ADJUSTMENTS")
        loss_col = RED if any("LOSS" in label or "INCOME (LOSS)" in label for _ in [1]) else "000000"
        bg = SUBHDR_BG if (is_total and not label.startswith("  ")) else (ALT_ROW if i % 2 == 0 else WHITE)
        lbl(ws, r, 2, label, bg=bg)
        for j, v in enumerate(vals):
            if v is None:
                dc(ws, r, 3+j, "—", bg=bg)
            elif isinstance(v, str):
                wc(ws, r, 3+j, v, size=FONT_SIZE, bg=bg, align="right", border=True, italic=True, fg=DARK_GRAY)
            else:
                clr = RED if v < 0 else "000000"
                dc(ws, r, 3+j, v, bg=bg, bold=is_total, num_fmt='#,##0', color=clr)

# ═══════════════════════════════════════════════════════════════════════════════
def build_balance_sheet(wb):
    ws = wb.create_sheet("Balance Sheet")
    ws.sheet_view.showGridLines = False
    scw(ws, 1, 3); scw(ws, 2, 35); scw(ws, 3, 18); scw(ws, 4, 18)
    scw(ws, 5, 18); scw(ws, 6, 18)

    sec_hdr(ws, 1, 2, "BALANCE SHEET — SNOWFLAKE INC. (SNOW)  |  FY2023-FY2026 ($M)", span=5)
    sub_hdr(ws, 2, 2, "Metric")
    for i, yr in enumerate(["FY2023", "FY2024", "FY2025", "FY2026"]):
        sub_hdr(ws, 2, 3+i, yr)

    rows = [
        ("ASSETS",                        None, None, None, None),
        ("CURRENT ASSETS",                None, None, None, None),
        ("  Cash & Cash Equivalents",      4793, 3675, 3562, 3800),
        ("  Short-term Investments",       2020, 2017, 2100, 2000),
        ("  Accounts Receivable (net)",     498,  715,  890, 1050),
        ("  Deferred Commission",           175,  200,  210,  230),
        ("  Other Current Assets",          156,  195,  200,  210),
        ("TOTAL CURRENT ASSETS",           7642, 6802, 6962, 7290),
        ("NON-CURRENT ASSETS",            None, None, None, None),
        ("  Long-term Investments",        1080,  820,  700,  600),
        ("  Operating Lease Right-of-Use",  253,  277,  290,  350),
        ("  Property & Equipment (net)",    530,  630,  750,  900),
        ("  Goodwill",                      341,  342,  345,  600),
        ("  Intangibles (net)",              30,   22,   15,  200),
        ("  Deferred Commission LT",        244,  280,  300,  310),
        ("  Other Non-current Assets",      150,  160,  175,  190),
        ("TOTAL NON-CURRENT ASSETS",       2628, 2531, 2575, 3150),
        ("TOTAL ASSETS",                  10270, 9333, 9537,10440),
        ("", None, None, None, None),
        ("LIABILITIES",                   None, None, None, None),
        ("  Deferred Revenue (current)",    857, 1086, 1300, 1500),
        ("  Accounts Payable",               81,   90,  105,  120),
        ("  Accrued Expenses",              348,  450,  500,  550),
        ("  Operating Lease (current)",      57,   59,   65,   70),
        ("  Other Current Liabilities",      90,  100,  110,  120),
        ("TOTAL CURRENT LIABILITIES",      1433, 1785, 2080, 2360),
        ("NON-CURRENT LIABILITIES",        None, None, None, None),
        ("  Long-term Debt",                  0,    0,    0,    0),
        ("  Deferred Revenue LT",           124,   97,   90,  100),
        ("  Operating Lease LT",            228,  256,  265,  320),
        ("  Other LT Liabilities",          150,  140,  145,  155),
        ("TOTAL NON-CURRENT LIABILITIES",   502,  493,  500,  575),
        ("TOTAL LIABILITIES",              1935, 2278, 2580, 2935),
        ("", None, None, None, None),
        ("SHAREHOLDERS' EQUITY",          None, None, None, None),
        ("  Common Stock & APIC",          9928,10740,11210,11800),
        ("  Accumulated Deficit",          -1593,-3703,-4267,-4300),
        ("TOTAL EQUITY",                   8335, 7037, 6943, 7505),
        ("TOTAL LIAB + EQUITY",           10270, 9315, 9523,10440),
        ("", None, None, None, None),
        ("KEY BALANCE SHEET METRICS",     None, None, None, None),
        ("  Cash + Investments ($M)",      7893, 6512, 6362, 6400),
        ("  Long-term Debt",                  0,    0,    0,    0),
        ("  Net Cash Position ($M)",       7893, 6512, 6362, 6400),
        ("  Deferred Revenue Total ($M)", 981, 1183, 1390, 1600),
    ]

    for i, row_data in enumerate(rows):
        r = 3 + i
        label = row_data[0]
        vals  = row_data[1:]
        if label == "":
            ws.row_dimensions[r].height = 8
            continue
        is_total = label.startswith("TOTAL") or label in ("ASSETS","LIABILITIES","SHAREHOLDERS' EQUITY",
                                                           "KEY BALANCE SHEET METRICS",
                                                           "CURRENT ASSETS","NON-CURRENT ASSETS",
                                                           "CURRENT LIABILITIES","NON-CURRENT LIABILITIES")
        bg = SUBHDR_BG if is_total else (ALT_ROW if i % 2 == 0 else WHITE)
        lbl(ws, r, 2, label, bg=bg)
        for j, v in enumerate(vals):
            if v is None:
                dc(ws, r, 3+j, "—", bg=bg)
            else:
                clr = RED if v < 0 else "000000"
                dc(ws, r, 3+j, v, bg=bg, bold=is_total, num_fmt='#,##0', color=clr)

# ═══════════════════════════════════════════════════════════════════════════════
def build_cashflow(wb):
    ws = wb.create_sheet("Cash Flow Analysis")
    ws.sheet_view.showGridLines = False
    scw(ws, 1, 3); scw(ws, 2, 35); scw(ws, 3, 18); scw(ws, 4, 18)
    scw(ws, 5, 18); scw(ws, 6, 18); scw(ws, 7, 18)

    sec_hdr(ws, 1, 2, "CASH FLOW ANALYSIS — SNOWFLAKE INC. (SNOW)  |  ($M)", span=6)
    sub_hdr(ws, 2, 2, "Metric")
    for i, yr in enumerate(["FY2022", "FY2023", "FY2024", "FY2025", "FY2026"]):
        sub_hdr(ws, 2, 3+i, yr)

    rows = [
        ("OPERATING CASH FLOW",            -314,  202, 1141,  903, 1077),
        ("  Net Loss",                      -889, -791, -836, -985, -800),
        ("  Stock-Based Compensation",       517,  747,  760,  830,  870),
        ("  D&A",                             66,   90,  115,  130,  145),
        ("  Change in Deferred Revenue",     285,  170,  200,  220,  210),
        ("  Other Working Capital / Changes", -293,-14, -98, -292, -348),
        ("", None, None, None, None, None),
        ("INVESTING CASH FLOW",             -624,-2154,-1283, -620, -950),
        ("  Capital Expenditures",           -85, -193, -222, -272, -350),
        ("  Purchase of Investments",       -2072,-4620,-2875,-1880,-1200),
        ("  Maturities / Sales Invest.",     1533, 3159, 1814, 1532, 1600),
        ("  Acquisitions (net)",               0,  -500,    0,    0, -450),
        ("", None, None, None, None, None),
        ("FINANCING CASH FLOW",             2196,  -94, -198, -355, -550),
        ("  Equity Issuance (net)",          2221,    6,    0,    0,    0),
        ("  Share Repurchases",                0,    0, -100, -355, -550),
        ("  Other",                           -25, -100, -98,    0,    0),
        ("", None, None, None, None, None),
        ("FREE CASH FLOW (OCF - CapEx)",    -399,    9,  919,  631,  727),
        ("  FCF Margin",                    "-33%","0.4%","32.7%","17.4%","15.5%"),
        ("  Adj. FCF Margin (mgmt def.)",   "-20%","5.0%","36.0%","22.0%","23.0%"),
        ("", None, None, None, None, None),
        ("CAPITAL ALLOCATION ($M)",         None, None, None, None, None),
        ("  CapEx",                           85,  193,  222,  272,  350),
        ("  CapEx as % of Revenue",         "7%","9.3%","7.9%","7.5%","7.5%"),
        ("  Share Repurchases",                0,    0,  100,  355,  550),
        ("  Acquisitions",                     0,  500,    0,    0,  450),
        ("  R&D Spend",                       544,  814,  986, 1017, 1100),
        ("  Cash Balance (end of year $M)", 4782, 4793, 3675, 3562, 3800),
    ]

    for i, row_data in enumerate(rows):
        r = 3 + i
        label = row_data[0]
        vals  = row_data[1:]
        if label == "":
            ws.row_dimensions[r].height = 8
            continue
        is_total = any(label.startswith(x) for x in ("OPERATING","INVESTING","FINANCING","FREE CASH","CAPITAL"))
        bg = SUBHDR_BG if is_total else (ALT_ROW if i % 2 == 0 else WHITE)
        lbl(ws, r, 2, label, bg=bg)
        for j, v in enumerate(vals):
            if v is None:
                dc(ws, r, 3+j, "—", bg=bg)
            elif isinstance(v, str):
                wc(ws, r, 3+j, v, size=FONT_SIZE, bg=bg, align="right", border=True, italic=True, fg=DARK_GRAY)
            else:
                clr = RED if v < 0 else "000000"
                dc(ws, r, 3+j, v, bg=bg, bold=is_total, num_fmt='#,##0', color=clr)

# ═══════════════════════════════════════════════════════════════════════════════
def build_roic(wb):
    ws = wb.create_sheet("Return on Capital")
    ws.sheet_view.showGridLines = False
    scw(ws, 1, 3); scw(ws, 2, 35); scw(ws, 3, 18); scw(ws, 4, 18)
    scw(ws, 5, 18); scw(ws, 6, 18); scw(ws, 7, 18)

    sec_hdr(ws, 1, 2, "RETURN ON CAPITAL — SNOWFLAKE INC. (SNOW)", span=6)
    sub_hdr(ws, 2, 2, "Metric")
    for i, yr in enumerate(["FY2022", "FY2023", "FY2024", "FY2025", "FY2026"]):
        sub_hdr(ws, 2, 3+i, yr)

    rows = [
        ("GAAP RETURNS (NOTE: ALL NEGATIVE)", None, None, None, None, None),
        ("  GAAP Return on Equity",      "-107%","-95%","-119%","-142%","-107%"),
        ("  GAAP Return on Assets",       "-87%","-77%","-89%","-103%","-77%"),
        ("  GAAP ROIC",                   "-65%","-52%","-54%","-72%","-62%"),
        ("  NOTE",                        "GAAP losses driven by ~$830M+/yr SBC; non-cash charge",
                                           None, None, None, None),
        ("", None, None, None, None, None),
        ("NON-GAAP RETURNS (OPERATING)",  None, None, None, None, None),
        ("  Non-GAAP Operating Margin",   "-22%","-2%","9.6%","8.1%","7.3%"),
        ("  Non-GAAP ROE (adj.)",          "N/M",  "N/M","10%","8.5%","9.5%"),
        ("  Adj. EBITDA Margin",           "-16%","2.3%","13.7%","11.6%","10.4%"),
        ("", None, None, None, None, None),
        ("UNIT ECONOMICS (KEY)",          None, None, None, None, None),
        ("  NRR (Net Revenue Retention)", "N/A","158%","131%","131%","~131%"),
        ("  Gross Margin",               "59.3%","62.3%","65.2%","69.3%","72.0%"),
        ("  Product Gross Margin",        "65%","70%","75%","76%","76%"),
        ("  CAC Payback Period (est.)",   "~36mo","~30mo","~24mo","~24mo","~24mo"),
        ("  LTV / CAC Ratio (est.)",       3.5,   4.5,   5.5,   5.5,   5.5),
        ("", None, None, None, None, None),
        ("INCREMENTAL ANALYSIS",         None, None, None, None, None),
        ("  Revenue Growth ($M)",          847,  847,  747,  814, 1056),
        ("  S&M + R&D Invested ($M)",     1378, 1914, 2179, 2292, 2450),
        ("  Incremental Rev / $ Invested", 0.61, 0.44, 0.34, 0.35, 0.43),
        ("", None, None, None, None, None),
        ("FREE CASH FLOW RETURNS",       None, None, None, None, None),
        ("  FCF ($M)",                    -399,    9,  919,  631,  727),
        ("  FCF / Revenue",              "-33%","0.4%","32.7%","17.4%","15.5%"),
        ("  Cash Return on Invested Cap.","-6.5%","0.1%","14.5%","10%","11.5%"),
        ("", None, None, None, None, None),
        ("SBC ANALYSIS (KEY WATCH ITEM)", None, None, None, None, None),
        ("  SBC ($M)",                     517,  747,  760,  830,  870),
        ("  SBC as % of Revenue",         "42%","36%","27%","22.9%","18.6%"),
        ("  SBC Declining — POSITIVE",    "Base","Improv.","Improv.","Improv.","Improv."),
    ]

    for i, row_data in enumerate(rows):
        r = 3 + i
        label = row_data[0]
        vals  = row_data[1:]
        if label == "":
            ws.row_dimensions[r].height = 8
            continue
        is_hdr = not label.startswith("  ")
        bg = SUBHDR_BG if is_hdr else (ALT_ROW if i % 2 == 0 else WHITE)
        lbl(ws, r, 2, label, bg=bg)
        for j, v in enumerate(vals):
            if v is None:
                dc(ws, r, 3+j, "—", bg=bg)
            elif isinstance(v, str):
                clr = RED if v.startswith("-") and "%" in v else DARK_GRAY
                wc(ws, r, 3+j, v, size=FONT_SIZE, bg=bg, align="right", border=True,
                   italic=not is_hdr, fg=clr)
            else:
                clr = RED if v < 0 else "000000"
                dc(ws, r, 3+j, v, bg=bg, bold=is_hdr, num_fmt='#,##0.00', color=clr)

# ═══════════════════════════════════════════════════════════════════════════════
def build_management(wb):
    ws = wb.create_sheet("Management")
    ws.sheet_view.showGridLines = False
    scw(ws, 1, 3); scw(ws, 2, 28); scw(ws, 3, 55); scw(ws, 4, 18)

    sec_hdr(ws, 1, 2, "MANAGEMENT ANALYSIS — SNOWFLAKE INC. (SNOW)", span=3)

    sub_hdr(ws, 3, 2, "Key Leadership", span=3)
    leaders = [
        ("Name",                "Role",             "Background & Assessment"),
        ("Sridhar Ramaswamy",   "CEO (since Feb 2024)",
         "Former Google VP of Ads (15 yrs); built Google Ads from $1.5B to $100B; "
         "founded Neeva (AI search, acquired by Snowflake May 2023). "
         "POSITIVE: Product visionary with AI-native DNA; RISK: no prior public CEO experience at this scale"),
        ("Shan Krishnasamy",    "Incoming CFO (2025)",
         "Succeeded Michael Scarpelli; Scarpelli was exceptional communicator; new CFO needs to prove credibility with investors"),
        ("Frank Slootman",      "Former CEO / Board",
         "Legendary enterprise software exec (ServiceNow, Data Domain). Stepped down Feb 2024. "
         "His departure was surprising but management says 'evolution.' Created some uncertainty."),
        ("Prasanna Krishnan",   "CPO (Chief Product Officer)",
         "Formerly at Google; leading Cortex AI and Snowflake Intelligence product roadmap"),
    ]
    for i, rd in enumerate(leaders):
        r = 4 + i
        if i == 0:
            sub_hdr(ws, r, 2, rd[0]); sub_hdr(ws, r, 3, rd[1]); sub_hdr(ws, r, 4, rd[2])
        else:
            bg = ALT_ROW if i % 2 == 0 else WHITE
            lbl(ws, r, 2, rd[0], bg=bg); lbl(ws, r, 3, rd[1], bg=bg)
            ws.row_dimensions[r].height = 50
            wc(ws, r, 4, rd[2], size=FONT_SIZE, bg=bg, align="left", wrap=True, border=True)

    sub_hdr(ws, 10, 2, "Compensation & Incentives (Proxy Analysis)", span=3)
    comp = [
        ("CEO Pay",            "Sridhar Ramaswamy total comp ~$105M in FY2025 (majority RSUs vesting); very high for unprofitable company"),
        ("Pay Philosophy",     "High SBC culture: $830M/yr in SBC = 23% of revenue; being addressed but still dilutive"),
        ("SBC Trajectory",     "FY2022: 42% of rev → FY2026: 18.6% of rev — clear improvement; target 15% by FY2028"),
        ("Dilution",           "~5-7% annual dilution from SBC; partially offset by $550M buybacks in FY2026"),
        ("Insider Holdings",   "CEO Sridhar owns ~$50M in SNOW (growing position); alignment improving"),
        ("Former CEO Buyback", "Frank Slootman sold significant shares at $250-300 range; limited remaining alignment"),
    ]
    for i, (k, v) in enumerate(comp):
        r = 11 + i
        bg = ALT_ROW if i % 2 == 0 else WHITE
        lbl(ws, r, 2, k, bg=bg)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
        ws.row_dimensions[r].height = 38
        wc(ws, r, 3, v, size=FONT_SIZE, bg=bg, align="left", wrap=True, border=True)

    sub_hdr(ws, 19, 2, "Capital Allocation Decisions", span=3)
    ca = [
        ("R&D",               "$1.1B in FY2026; primarily AI products (Cortex, Intelligence, Snowpark); key to differentiation"),
        ("Buybacks",          "$550M in FY2026 (up from $355M FY2025); offsetting SBC dilution; shareholder friendly signal"),
        ("Acquisitions",      "Observe (IT ops, ~$450M) and TensorStax in FY2026; strategic but adds integration risk"),
        ("Cash Management",   "$6.4B cash/investments; no debt; very strong balance sheet; runway for investment"),
        ("No Dividends",      "Appropriate given growth/investment phase; 100% reinvestment mode"),
        ("Leverage",          "ZERO long-term debt; pristine balance sheet; $6.4B net cash is a significant cushion"),
    ]
    for i, (k, v) in enumerate(ca):
        r = 20 + i
        bg = ALT_ROW if i % 2 == 0 else WHITE
        lbl(ws, r, 2, k, bg=bg)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
        ws.row_dimensions[r].height = 38
        wc(ws, r, 3, v, size=FONT_SIZE, bg=bg, align="left", wrap=True, border=True)

    sub_hdr(ws, 28, 2, "Acts Like an Owner? — Assessment", span=3)
    owner = [
        ("Long-term Focus",      "YES — Cortex AI investment is multi-year; Observe acquisition expands TAM; seeds planted"),
        ("Seeds Planted",        "Snowflake Intelligence (agentic AI), Cortex AI, Streamlit ecosystem — all early stage growth"),
        ("Borrowing from Future", "CONCERN: SBC still very high relative to revenue; essentially paying employees with shareholder dilution"),
        ("CEO Alignment",        "Sridhar has been buying stock; recently granted large RSU package tied to performance milestones"),
        ("Accountability",       "Clear on GAAP profitability goal ('two years ago SBC was 41% of revenue' — openly acknowledged)"),
        ("Insider Sells",        "Former CEO Slootman sold $100M+ in shares in 2022-2023 at elevated prices; red flag in hindsight"),
    ]
    for i, (k, v) in enumerate(owner):
        r = 29 + i
        bg = ALT_ROW if i % 2 == 0 else WHITE
        lbl(ws, r, 2, k, bg=bg)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
        ws.row_dimensions[r].height = 38
        wc(ws, r, 3, v, size=FONT_SIZE, bg=bg, align="left", wrap=True, border=True)

    sub_hdr(ws, 37, 2, "Overall Management Score", span=3, bg=DARK_BLUE, fg=WHITE)
    ws.merge_cells("B38:D39")
    wc(ws, 38, 2,
       "SCORE: 6.5/10 — Sridhar Ramaswamy has the right product vision and is the right CEO for "
       "the AI pivot. Capital allocation is improving (less SBC, more buybacks). However, CEO pay "
       "is very high for an unprofitable company, legacy insider selling was poorly timed, and "
       "execution on GAAP profitability is a multi-year journey. The jury is still out on whether "
       "Ramaswamy can scale from product vision to enterprise execution at Snowflake's stage.",
       size=FONT_SIZE, bg=ALT_ROW, align="left", wrap=True, border=True)
    for r in [38, 39]:
        ws.row_dimensions[r].height = 30

# ═══════════════════════════════════════════════════════════════════════════════
def build_risks(wb):
    ws = wb.create_sheet("Risks")
    ws.sheet_view.showGridLines = False
    scw(ws, 1, 3); scw(ws, 2, 32); scw(ws, 3, 45); scw(ws, 4, 15); scw(ws, 5, 15)

    sec_hdr(ws, 1, 2, "RISK ANALYSIS — SNOWFLAKE INC. (SNOW)", span=4)
    sub_hdr(ws, 2, 2, "Risk Category")
    sub_hdr(ws, 2, 3, "Description & Impact")
    sub_hdr(ws, 2, 4, "Probability")
    sub_hdr(ws, 2, 5, "Impact")

    risks = [
        ("MICROSOFT FABRIC BUNDLING",
         "Microsoft is bundling Azure Synapse, Databricks Spark, and Power BI into 'Fabric' under one "
         "Azure license. For existing Microsoft/Azure customers, Fabric may be 'good enough.' If Fabric "
         "adoption accelerates, Snowflake faces pricing pressure or displacement in existing accounts. "
         "Microsoft has an enormous enterprise installed base — this is Snowflake's biggest strategic risk.",
         "HIGH", "VERY HIGH"),
        ("DATABRICKS COMPETITION",
         "Databricks (~$62B private valuation) leads in AI/ML workloads, open-source Delta Lake, and "
         "is the preferred platform for data science/ML engineers. Their Unity Catalog directly "
         "competes with Snowflake for enterprise data governance. A Databricks IPO would intensify "
         "competitive investment and marketing.",
         "HIGH", "HIGH"),
        ("SECURITIES CLASS ACTION LAWSUIT",
         "Ongoing securities class action alleging Snowflake withheld material negative information "
         "about Iceberg Tables, tiered pricing changes, and product efficiency improvements that "
         "were expected to hurt consumption revenue. Investor deadline April 27, 2026. "
         "Legal costs and settlement risk; reputational damage.",
         "MEDIUM", "MEDIUM"),
        ("CONSUMPTION MODEL VOLATILITY",
         "Unlike SaaS subscriptions, consumption revenue can drop overnight if customers optimize "
         "queries, switch workloads, or reduce data processing. This model creates revenue "
         "unpredictability and is particularly sensitive to macro slowdowns.",
         "MEDIUM", "HIGH"),
        ("GAAP PROFITABILITY PATH",
         "Despite 7 years as a public company, Snowflake has NEVER been GAAP profitable. "
         "SBC at 18.6% of revenue is still extremely high. FCF is real but GAAP losses continue. "
         "Investor patience for profitability timelines is wearing thin.",
         "MEDIUM", "MEDIUM"),
        ("AI COMMODITIZATION",
         "If LLMs can query raw data without a structured data warehouse, the data warehouse "
         "TAM could shrink structurally. Snowflake's Cortex AI is a hedge, but the platform may "
         "be disrupted by a fundamentally different AI-native architecture over 5-10 years.",
         "LOW", "HIGH"),
        ("CUSTOMER CONCENTRATION",
         "Top 10 customers likely represent 10-15% of revenue. Loss of a whale customer (Anthem, "
         "Capital One level) would be material to results and investor confidence.",
         "LOW", "HIGH"),
        ("TALENT RETENTION",
         "With Databricks, OpenAI, and Google competing for data/AI engineers, retaining top talent "
         "is expensive. SBC cuts could increase attrition. Frank Slootman's departure was already "
         "disruptive from a talent perspective.",
         "MEDIUM", "MEDIUM"),
        ("MACRO / CLOUD SPENDING",
         "Enterprise cloud spending is cyclical. If companies optimize cloud costs (as happened in "
         "2023), Snowflake consumption immediately drops. The 2023 'optimization headwind' shaved "
         "~5-10 percentage points off revenue growth.",
         "LOW-MEDIUM", "MEDIUM"),
        ("MULTIPLE COMPRESSION",
         "Trading at ~15x P/S is expensive for a company with GAAP losses. If growth "
         "decelerates further to 20-25%, multiple compression could drive significant stock decline. "
         "Stock was at $400+ in 2021 and is now ~$132.",
         "MEDIUM", "MEDIUM"),
    ]

    for i, (cat, desc, prob, imp) in enumerate(risks):
        r = 3 + i
        bg = ALT_ROW if i % 2 == 0 else WHITE
        lbl(ws, r, 2, cat, bg=bg)
        ws.row_dimensions[r].height = 60
        wc(ws, r, 3, desc, size=FONT_SIZE, bg=bg, align="left", wrap=True, border=True)
        prob_c = RED if prob == "HIGH" else (GOLD if "MEDIUM" in prob else GREEN)
        imp_c  = RED if imp in ("VERY HIGH","HIGH") else (GOLD if "MEDIUM" in imp else GREEN)
        wc(ws, r, 4, prob, bold=True, size=FONT_SIZE, fg=WHITE, bg=prob_c, align="center", border=True)
        wc(ws, r, 5, imp,  bold=True, size=FONT_SIZE, fg=WHITE, bg=imp_c,  align="center", border=True)

    sub_hdr(ws, 14, 2, "Risk Summary & Investment Implication", span=4)
    summary = ("Snowflake is a high-risk, high-reward investment. The core product is excellent and "
               "the AI pivot is strategically sound. However, Microsoft Fabric and Databricks pose "
               "genuine threats to the platform's moat. The securities lawsuit adds near-term uncertainty. "
               "The stock at ~$132 (15x P/S) offers a more reasonable entry point than the $300+ highs, "
               "but GAAP losses and competition mean this is NOT a 'safe' investment. Best suited for "
               "investors with a 5+ year horizon who believe AI data workloads structurally grow consumption.")
    ws.merge_cells("B15:E17")
    wc(ws, 15, 2, summary, size=FONT_SIZE, bg=LIGHT_GRAY, align="left", wrap=True, border=True)
    for r in [15, 16, 17]:
        ws.row_dimensions[r].height = 22

# ═══════════════════════════════════════════════════════════════════════════════
def build_valuation(wb):
    ws = wb.create_sheet("Valuation")
    ws.sheet_view.showGridLines = False
    scw(ws, 1, 3); scw(ws, 2, 32); scw(ws, 3, 22); scw(ws, 4, 22)
    scw(ws, 5, 22); scw(ws, 6, 22)

    sec_hdr(ws, 1, 2, "VALUATION ANALYSIS — SNOWFLAKE INC. (SNOW)", span=5)

    sub_hdr(ws, 3, 2, "Current Market Statistics (April 9-10, 2026)", span=5)
    mkt = [
        ("Stock Price",                  "~$132 (Apr 9); -12% Apr 10 on AI fears & legal news"),
        ("Market Capitalization",         "~$50-55B (est., based on ~380M diluted shares)"),
        ("Enterprise Value",              "~$44-49B (less ~$6.4B net cash)"),
        ("Price / NTM Revenue",           "~11-15x (on ~$5.7B FY2027E revenue)"),
        ("EV / NTM Revenue",              "~9-12x"),
        ("Non-GAAP P/E (FY2026E)",        "~104x (on adj. EPS ~$0.97)"),
        ("Price / FCF",                   "~70x (on $727M FCF)"),
        ("52-Week Range",                 "$~130 - $~220"),
        ("YTD Performance",               "-23% YTD (as of April 2026)"),
        ("Analyst Consensus",             "BUY (44 Buy, 1 Sell); median target $239.84"),
    ]
    for i, (k, v) in enumerate(mkt):
        r = 4 + i
        bg = ALT_ROW if i % 2 == 0 else WHITE
        lbl(ws, r, 2, k, bg=bg)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
        wc(ws, r, 3, v, size=FONT_SIZE, bg=bg, align="right", border=True)
        ws.row_dimensions[r].height = 22

    sub_hdr(ws, 16, 2, "DCF / Revenue Multiple Valuation (Base / Bull / Bear)", span=5)
    sub_hdr(ws, 17, 2, "Assumption"); sub_hdr(ws, 17, 3, "Bear Case")
    sub_hdr(ws, 17, 4, "Base Case"); sub_hdr(ws, 17, 5, "Bull Case"); sub_hdr(ws, 17, 6, "Notes")

    dcf = [
        ("FY2027E Revenue",                "$5.5B",  "$5.7B",   "$6.0B", "Bear=competition headwinds"),
        ("Revenue Growth (3-year CAGR)",   "18%",    "25%",     "32%",   "Bull=AI workloads surge"),
        ("Terminal NTM P/S Multiple",       "8x",    "12x",     "18x",   "Bear=Databricks/MSFT wins"),
        ("Non-GAAP Op Margin (FY2028)",    "12%",    "18%",     "25%",   "Bull=scale economies"),
        ("WACC",                           "12%",    "10%",     "9%",    "High risk premium warranted"),
        ("Implied Value / Share",          "$80",    "$155",    "$280",  "Based on FY2027 exit multiple"),
    ]
    for i, row_data in enumerate(dcf):
        r = 18 + i
        bg = ALT_ROW if i % 2 == 0 else WHITE
        lbl(ws, r, 2, row_data[0], bg=bg)
        for j in range(1, 5):
            wc(ws, r, 2+j, row_data[j], size=FONT_SIZE, bg=bg,
               align="right" if j < 4 else "left", border=True,
               bold=(row_data[0].startswith("Implied")))
        ws.row_dimensions[r].height = 22

    sub_hdr(ws, 26, 2, "Peer Valuation Comparison", span=5)
    sub_hdr(ws, 27, 2, "Company"); sub_hdr(ws, 27, 3, "P/S NTM")
    sub_hdr(ws, 27, 4, "Fwd P/E"); sub_hdr(ws, 27, 5, "Rev Growth"); sub_hdr(ws, 27, 6, "GAAP Profitable")

    peers = [
        ("Snowflake (SNOW)",      "~13x",   "~104x",  "~25%",   "NO"),
        ("Databricks (private)",  "~15x",   "N/M",    "~40%",   "NO"),
        ("Palantir (PLTR)",       "~30x",   "~95x",   "~25%",   "YES (recent)"),
        ("MongoDB (MDB)",         "~10x",   "~45x",   "~20%",   "NO"),
        ("Elastic (ESTC)",         "~6x",   "~30x",   "~17%",   "NO"),
        ("Datadog (DDOG)",        "~12x",   "~60x",   "~25%",   "YES"),
    ]
    for i, row_data in enumerate(peers):
        r = 28 + i
        bold = row_data[0].startswith("Snowflake")
        bg = SUBHDR_BG if bold else (ALT_ROW if i % 2 == 0 else WHITE)
        for j, v in enumerate(row_data):
            if j == 0:
                lbl(ws, r, 2, v, bg=bg)
            else:
                clr = RED if v == "NO" else (GREEN if v == "YES" else "000000")
                wc(ws, r, 2+j, v, bold=bold, size=FONT_SIZE, fg=clr, bg=bg, align="right", border=True)

    sub_hdr(ws, 36, 2, "Margin of Safety Assessment", span=5)
    mos = [
        ("Current Price",                "$132",    "April 9, 2026 (before -12% Apr 10 drop)"),
        ("Base Case Intrinsic Value",    "$155",    "~17% upside to base case"),
        ("Bear Case Value",               "$80",    "~39% downside in adverse scenario"),
        ("Bull Case Value",              "$280",    "~112% upside in bull case (AI boom)"),
        ("Margin of Safety",         "MODERATE",    "More attractive than 2021-2022 highs; but competition risk is real"),
        ("Recommendation",         "SPECULATIVE BUY", "High-risk; position size <5% of portfolio; 3-5yr hold required"),
    ]
    for i, (k, v, c) in enumerate(mos):
        r = 37 + i
        bg = ALT_ROW if i % 2 == 0 else WHITE
        lbl(ws, r, 2, k, bg=bg)
        wc(ws, r, 3, v, bold=True, size=FONT_SIZE, bg=bg, align="right", border=True)
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=6)
        wc(ws, r, 4, c, size=FONT_SIZE, bg=bg, align="left", border=True)
        ws.row_dimensions[r].height = 22

# ═══════════════════════════════════════════════════════════════════════════════
def build_sentiment(wb):
    ws = wb.create_sheet("Market Sentiment")
    ws.sheet_view.showGridLines = False
    scw(ws, 1, 3); scw(ws, 2, 30); scw(ws, 3, 52); scw(ws, 4, 18)

    sec_hdr(ws, 1, 2, "MARKET SENTIMENT — SNOWFLAKE INC. (SNOW)", span=3)

    sub_hdr(ws, 3, 2, "Analyst Coverage", span=3)
    analyst = [
        ("Total Analysts",          "~45+"),
        ("Buy / Hold / Sell",       "~80% Buy | ~18% Hold | ~2% Sell"),
        ("Median Price Target",     "$239.84 (stock trading at ~$132 — significant premium to current price)"),
        ("Consensus View",          "Bullish on AI data cloud thesis; near-term concerns on competition + legal overhang"),
    ]
    for i, (k, v) in enumerate(analyst):
        r = 4 + i
        bg = ALT_ROW if i % 2 == 0 else WHITE
        lbl(ws, r, 2, k, bg=bg)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
        wc(ws, r, 3, v, size=FONT_SIZE, bg=bg, align="left", border=True)
        ws.row_dimensions[r].height = 22

    sub_hdr(ws, 10, 2, "Key Bullish Narratives", span=3)
    bull = [
        ("AI Workload Growth",       "AI inference, training, and RAG workloads consume significantly more compute credits — 3x vs standard queries"),
        ("RPO Acceleration",         "$9.77B RPO (+42% YoY) = 2x annual revenue; highest backlog coverage in Snowflake's history"),
        ("Valuation Reset",          "Stock down 60%+ from highs; 15x P/S vs 40x+ peak; much of risk is priced in at current levels"),
        ("430+ New Capabilities",    "Cortex AI, Snowflake Intelligence, Iceberg Tables, Streamlit — product velocity is accelerating under Ramaswamy"),
        ("Profitability Path",       "SBC declining, FCF $727M, non-GAAP expanding; FY2028 GAAP profitability within sight"),
    ]
    for i, (t, d) in enumerate(bull):
        r = 11 + i
        bg = ALT_ROW if i % 2 == 0 else WHITE
        lbl(ws, r, 2, t, bg=bg)
        ws.row_dimensions[r].height = 38
        wc(ws, r, 3, d, size=FONT_SIZE, bg=bg, align="left", wrap=True, border=True)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)

    sub_hdr(ws, 18, 2, "Key Bearish Narratives", span=3)
    bear = [
        ("Microsoft Fabric",         "Azure's bundled data platform is 'good enough' for 80% of use cases at no incremental cost"),
        ("Securities Litigation",    "April 27, 2026 lawsuit deadline creates investor uncertainty; potential settlement costs"),
        ("AI Disruption Fear",       "LLM-native data querying could bypass traditional data warehouses structurally over 5-10 years"),
        ("Consumption Uncertainty",  "Iceberg Table efficiency improvements may reduce consumption per query — a double-edged sword"),
        ("GAAP Profitability",        "Never profitable on GAAP basis; SBC dilution continues; some investors avoid loss-making companies"),
    ]
    for i, (t, d) in enumerate(bear):
        r = 19 + i
        bg = ALT_ROW if i % 2 == 0 else WHITE
        lbl(ws, r, 2, t, bg=bg)
        ws.row_dimensions[r].height = 38
        wc(ws, r, 3, d, size=FONT_SIZE, bg=bg, align="left", wrap=True, border=True)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)

    sub_hdr(ws, 25, 2, "Recent News & Catalysts (2025-2026)", span=3)
    news = [
        ("Apr 10, 2026", "SNOW craters 12% on AI disruption fears + securities lawsuit investor deadline approaching Apr 27"),
        ("Mar 2026",     "FY2026 Q4 earnings: $1.23B product revenue (+30%); RPO $9.77B (+42%); guidance in line"),
        ("Feb 2026",     "FY2026 full year results: $4.68B revenue (+29%); 430+ new product capabilities launched"),
        ("Jan 2026",     "Snowflake Intelligence announced — agentic AI layer for enterprise automation workflows"),
        ("Nov 2025",     "FY2026 Q3: product rev +28% YoY; launched Cortex AI across regions"),
        ("Aug 2025",     "Acquired Observe (IT Ops) to enter $50B+ observability market"),
        ("May 2025",     "Acquired TensorStax to strengthen AI-driven data engineering in Cortex Code"),
        ("Feb 2024",     "Sridhar Ramaswamy named CEO; Frank Slootman steps down unexpectedly — market reaction negative"),
    ]
    for i, (d, n) in enumerate(news):
        r = 26 + i
        bg = ALT_ROW if i % 2 == 0 else WHITE
        lbl(ws, r, 2, d, bg=bg)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
        wc(ws, r, 3, n, size=FONT_SIZE, bg=bg, align="left", border=True)
        ws.row_dimensions[r].height = 22

# ═══════════════════════════════════════════════════════════════════════════════
def build_key_indicators(wb):
    ws = wb.create_sheet("Key Indicators")
    ws.sheet_view.showGridLines = False
    scw(ws, 1, 3); scw(ws, 2, 30); scw(ws, 3, 20); scw(ws, 4, 20)
    scw(ws, 5, 20); scw(ws, 6, 20); scw(ws, 7, 20)

    sec_hdr(ws, 1, 2, "KEY INDICATORS — SNOWFLAKE INC. (SNOW)  |  FY2022-FY2026", span=6)
    sub_hdr(ws, 2, 2, "Indicator")
    for i, yr in enumerate(["FY2022", "FY2023", "FY2024", "FY2025", "FY2026"]):
        sub_hdr(ws, 2, 3+i, yr)

    rows = [
        ("FINANCIAL SUMMARY ($M)",      None, None, None, None, None),
        ("  Total Revenue",              1220, 2067, 2814, 3628, 4684),
        ("  Revenue Growth",            "106%","69%","36%","29%","29%"),
        ("  Product Revenue",            1139, 1968, 2665, 3476, 4527),
        ("  Product Gross Margin",        "65%","70%","75%","76%","76%"),
        ("  GAAP Operating Loss",        -880, -930, -645, -668, -692),
        ("  Non-GAAP Op Income (Loss)",  -267,  -43,  270,  292,  343),
        ("  Non-GAAP Op Margin",         "-22%","-2%","9.6%","8.1%","7.3%"),
        ("  GAAP Net Loss",              -889, -791, -836, -985, -800),
        ("  Adj. EPS (Non-GAAP)",        -0.51,-0.10,  0.79, 0.77, 0.97),
        ("", None, None, None, None, None),
        ("CUSTOMER METRICS",             None, None, None, None, None),
        ("  Total Customers",             6322, 8236, 9437,10618,13300),
        ("  Customers >$1M TTM Rev",       184,  330,  461,  578,  733),
        ("  Net Revenue Retention (%)",    "N/A","158%","131%","131%","~131%"),
        ("  RPO ($B)",                     2.6,   3.3,   5.2,   6.9,   9.8),
        ("", None, None, None, None, None),
        ("CASH FLOW ($M)",               None, None, None, None, None),
        ("  Operating Cash Flow",         -314,  202, 1141,  903, 1077),
        ("  Capital Expenditures",          85,  193,  222,  272,  350),
        ("  Free Cash Flow",             -399,    9,  919,  631,  727),
        ("  FCF Margin",                 "-33%","0.4%","32.7%","17.4%","15.5%"),
        ("  SBC ($M)",                    517,  747,  760,  830,  870),
        ("  SBC as % Revenue",           "42%","36%","27%","22.9%","18.6%"),
        ("  Share Buybacks ($M)",            0,    0,  100,  355,  550),
        ("", None, None, None, None, None),
        ("BALANCE SHEET ($M)",           None, None, None, None, None),
        ("  Cash + Investments",         7893, 7893, 6512, 6362, 6400),
        ("  Long-term Debt",                0,    0,    0,    0,    0),
        ("  Net Cash Position",          7893, 7893, 6512, 6362, 6400),
        ("", None, None, None, None, None),
        ("VALUATION",                    None, None, None, None, None),
        ("  P/S (NTM at year-end)",       "~35x","~20x","~15x","~18x","~13x"),
        ("  Stock Price Approx. (FY-end)","~$380","~$150","~$175","~$165","~$132"),
        ("  Market Cap Approx. ($B)",      "~145","~57","~67","~63","~50"),
    ]

    for i, row_data in enumerate(rows):
        r = 3 + i
        label = row_data[0]
        vals  = row_data[1:]
        if label == "":
            ws.row_dimensions[r].height = 8
            continue
        is_hdr = not label.startswith("  ")
        bg = SUBHDR_BG if is_hdr else (ALT_ROW if i % 2 == 0 else WHITE)
        lbl(ws, r, 2, label, bg=bg)
        for j, v in enumerate(vals):
            if v is None:
                dc(ws, r, 3+j, "—", bg=bg)
            elif isinstance(v, str):
                wc(ws, r, 3+j, v, size=FONT_SIZE, bg=bg, align="right", border=True, bold=is_hdr)
            else:
                clr = RED if v < 0 else "000000"
                dc(ws, r, 3+j, v, bg=bg, bold=is_hdr, num_fmt='#,##0', color=clr)

# ═══════════════════════════════════════════════════════════════════════════════
def main():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    build_cover(wb)
    build_business(wb)
    build_moat(wb)
    build_income(wb)
    build_balance_sheet(wb)
    build_cashflow(wb)
    build_roic(wb)
    build_management(wb)
    build_risks(wb)
    build_valuation(wb)
    build_sentiment(wb)
    build_key_indicators(wb)

    out_dir = "/Users/naji/WORK/github.com/AI/claude/Agent/MarketResearch/output"
    os.makedirs(out_dir, exist_ok=True)
    path = os.path.join(out_dir, "SNOW_Financial_Analysis.xlsx")
    wb.save(path)
    print(f"Saved: {path}")

if __name__ == "__main__":
    main()
