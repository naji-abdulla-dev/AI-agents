"""
Shopify Inc. (SHOP) Financial Analysis - Excel Generator
Data as of April 2026 | FY2025 Annual (ended December 31, 2025)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

# ── Shopify Color Palette ──────────────────────────────────────────────────────
SHOP_GREEN   = "008060"
SHOP_DARK    = "004C3F"
HEADER_BG    = "008060"
HEADER_FG    = "FFFFFF"
SUBHDR_BG    = "D6F0E8"
SUBHDR_FG    = "004C3F"
ALT_ROW      = "F0FAF6"
WHITE        = "FFFFFF"
LIGHT_GRAY   = "F8F8F8"
GREEN        = "27AE60"
RED          = "E74C3C"
GOLD         = "F39C12"
DARK_GRAY    = "555555"
DARK_GREEN   = "1E6B4A"
LIGHT_GREEN  = "E8F8F2"

FONT_SIZE = 14

def mf(bold=False, size=FONT_SIZE, color="000000", italic=False):
    return Font(name="Calibri", bold=bold, size=size, color=color, italic=italic)

def mfill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def mborder():
    s = Side(border_style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def cal(h="left", wrap=False):
    return Alignment(horizontal=h, vertical="center", wrap_text=wrap)

def wc(ws, row, col, value, bold=False, size=FONT_SIZE, fg="000000", bg=None,
       align="left", wrap=False, border=False, italic=False, num_fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = mf(bold=bold, size=size, color=fg, italic=italic)
    if bg:
        c.fill = mfill(bg)
    c.alignment = cal(align, wrap)
    if border:
        c.border = mborder()
    if num_fmt:
        c.number_format = num_fmt
    return c

def header_row(ws, row, cols, labels, bg=HEADER_BG, fg=HEADER_FG):
    for col, label in zip(cols, labels):
        wc(ws, row, col, label, bold=True, fg=fg, bg=bg, align="center", border=True)

def section_header(ws, row, col, label, span=1, bg=SUBHDR_BG, fg=SUBHDR_FG):
    wc(ws, row, col, label, bold=True, fg=fg, bg=bg, align="left", border=True)
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col + span - 1)

def set_col_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

def freeze(ws, cell="B3"):
    ws.freeze_panes = cell


# ══════════════════════════════════════════════════════════════════════════════
# TAB 1 – COVER
# ══════════════════════════════════════════════════════════════════════════════
def build_cover(wb):
    ws = wb.create_sheet("1. Cover")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, {1: 5, 2: 40, 3: 30, 4: 20, 5: 20})

    # Title block
    ws.row_dimensions[2].height = 60
    ws.merge_cells("B2:E2")
    c = ws["B2"]
    c.value = "SHOPIFY INC. (SHOP)"
    c.font = Font(name="Calibri", bold=True, size=36, color=HEADER_FG)
    c.fill = mfill(SHOP_GREEN)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[3].height = 30
    ws.merge_cells("B3:E3")
    c = ws["B3"]
    c.value = "Comprehensive Financial Analysis — April 2026"
    c.font = Font(name="Calibri", bold=False, size=18, color=HEADER_FG)
    c.fill = mfill(SHOP_DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # Key facts table
    facts = [
        ("Sector", "Technology / E-Commerce"),
        ("Exchange", "NASDAQ"),
        ("Ticker", "SHOP"),
        ("Current Price (Apr 12, 2026)", "$111.20"),
        ("52-Week High", "$182.19"),
        ("52-Week Low", "$78.00"),
        ("Market Cap", "~$143B"),
        ("Shares Outstanding", "~1.27B"),
        ("FY2025 Revenue", "$11.56B (+30% YoY)"),
        ("FY2025 GMV", "$378B (+29.5% YoY)"),
        ("FY2025 Free Cash Flow", "$2.0B (17% FCF margin)"),
        ("FY2025 Operating Income", "$1.47B"),
        ("Analyst Consensus", "Buy | Median PT: $181"),
        ("Rule of 50 Score", "Revenue Growth 30% + FCF Margin 17% = 47 ✓"),
        ("Report Date", "April 12, 2026"),
        ("Data Sources", "Shopify IR, SEC Filings, Macrotrends, StockAnalysis"),
    ]

    r = 5
    wc(ws, r, 2, "KEY FACTS", bold=True, fg=HEADER_FG, bg=HEADER_BG, size=FONT_SIZE+1)
    wc(ws, r, 3, "", bg=HEADER_BG)
    r += 1
    for label, value in facts:
        bg = ALT_ROW if r % 2 == 0 else WHITE
        wc(ws, r, 2, label, bold=True, fg=SHOP_DARK, bg=bg, border=True)
        wc(ws, r, 3, value, fg="000000", bg=bg, border=True)
        r += 1

    # Business description
    r += 1
    ws.merge_cells(f"B{r}:E{r}")
    wc(ws, r, 2, "BUSINESS SNAPSHOT", bold=True, fg=HEADER_FG, bg=SHOP_GREEN, size=FONT_SIZE+1)
    r += 1
    desc = (
        "Shopify is the leading global commerce platform enabling merchants of all sizes to "
        "start, grow, and manage their businesses online and offline. Founded in 2006 by Tobias Lütke, "
        "the platform powers over 5 million active stores in 175+ countries with $378B in annual GMV. "
        "Shopify earns revenue through Subscription Solutions (SaaS fees) and Merchant Solutions "
        "(payments, shipping, capital, and marketplace services), creating a powerful flywheel where "
        "merchant success drives Shopify's monetisation. The company exited logistics in 2023 to sharpen "
        "focus on its core platform, and has since accelerated margin expansion. FY2025 marked the first "
        "year Shopify crossed $10B in annual revenue and $300B GMV."
    )
    ws.merge_cells(f"B{r}:E{r+4}")
    c = ws.cell(row=r, column=2, value=desc)
    c.font = mf(size=FONT_SIZE)
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    c.fill = mfill(LIGHT_GREEN)
    ws.row_dimensions[r].height = 18
    for i in range(1, 5):
        ws.row_dimensions[r+i].height = 18


# ══════════════════════════════════════════════════════════════════════════════
# TAB 2 – BUSINESS OVERVIEW
# ══════════════════════════════════════════════════════════════════════════════
def build_business_overview(wb):
    ws = wb.create_sheet("2. Business Overview")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, {1: 3, 2: 32, 3: 22, 4: 22, 5: 22, 6: 18})
    freeze(ws, "B3")

    r = 1
    ws.merge_cells(f"B{r}:F{r}")
    wc(ws, r, 2, "SHOPIFY — BUSINESS OVERVIEW", bold=True, fg=HEADER_FG,
       bg=SHOP_GREEN, size=FONT_SIZE+2, align="center")
    r += 2

    # Products & Services
    section_header(ws, r, 2, "PRODUCTS & SERVICES", span=5)
    r += 1
    header_row(ws, r, [2, 3, 4, 5, 6], ["Segment", "Description", "Examples", "Revenue Type", "% of Revenue"])
    r += 1
    products = [
        ("Subscription Solutions", "Monthly/annual SaaS fees for platform access across Basic, Shopify, Advanced, and Plus tiers", "Shopify Basic, Shopify Plus", "Recurring SaaS", "~27%"),
        ("Merchant Solutions", "Value-added services that help merchants operate: payments, capital, shipping, and marketplace", "Shopify Payments, Shopify Capital, Shop Pay", "Transaction/Volume", "~73%"),
        ("Shopify Payments", "Integrated payment processing — largest component of Merchant Solutions", "Shop Pay, Buy Now Pay Later", "Take-rate on GMV", "Largest sub-segment"),
        ("Shopify Capital", "Working capital loans and merchant cash advances", "$5B+ in capital deployed", "Interest income", "Growing"),
        ("Shopify Shipping", "Discounted carrier rates, fulfillment tools", "DHL, USPS, UPS integrations", "Shipping margins", "Included in MRR"),
        ("Shopify Markets", "Cross-border commerce — duties, currencies, localisation", "International expansion tools", "Subscription add-on", "Growing"),
        ("Shop App", "Consumer shopping app — Shop Pay, order tracking, discovery", "78M registered users", "Engagement/ads", "Emerging"),
    ]
    for i, (seg, desc, ex, rev, pct) in enumerate(products):
        bg = ALT_ROW if i % 2 == 0 else WHITE
        for col, val in zip([2, 3, 4, 5, 6], [seg, desc, ex, rev, pct]):
            wc(ws, r, col, val, bg=bg, border=True, wrap=(col in [3, 4]))
        r += 1

    r += 1
    # Revenue breakdown
    section_header(ws, r, 2, "REVENUE BREAKDOWN — FY2025 (est.)", span=5)
    r += 1
    header_row(ws, r, [2, 3, 4, 5], ["Category", "FY2023", "FY2024", "FY2025"])
    r += 1
    rev_data = [
        ("Total Revenue", "$7.06B", "$8.88B", "$11.56B"),
        ("  Subscription Solutions", "$1.84B", "$2.35B", "$3.12B"),
        ("  Merchant Solutions", "$5.22B", "$6.53B", "$8.44B"),
        ("YoY Revenue Growth", "+26%", "+26%", "+30%"),
        ("Gross Merchandise Volume (GMV)", "$235B", "$292B", "$378B"),
        ("GMV YoY Growth", "+20%", "+24%", "+29.5%"),
        ("Active Merchants", "~2.5M", "~3.2M", "~5M+"),
    ]
    for i, row_vals in enumerate(rev_data):
        bg = ALT_ROW if i % 2 == 0 else WHITE
        bold = (row_vals[0] in ["Total Revenue", "Gross Merchandise Volume (GMV)"])
        for col, val in zip([2, 3, 4, 5], row_vals):
            wc(ws, r, col, val, bg=bg, border=True, bold=bold)
        r += 1

    r += 1
    # Geography
    section_header(ws, r, 2, "GEOGRAPHIC MIX (FY2025 est.)", span=5)
    r += 1
    header_row(ws, r, [2, 3, 4], ["Geography", "Revenue Share", "Notes"])
    r += 1
    geo = [
        ("Americas (incl. US)", "~72%", "US is largest market; strong Canada base (HQ in Ottawa)"),
        ("Europe, Middle East & Africa", "~18%", "Shopify Markets driving cross-border growth"),
        ("Asia Pacific", "~10%", "Emerging growth region; Japan, Australia strongholds"),
    ]
    for i, (g, s, n) in enumerate(geo):
        bg = ALT_ROW if i % 2 == 0 else WHITE
        for col, val in zip([2, 3, 4], [g, s, n]):
            wc(ws, r, col, val, bg=bg, border=True, wrap=(col == 4))
        r += 1

    r += 1
    # Key Clients & Value Propositions
    section_header(ws, r, 2, "VALUE PROPOSITIONS & KEY CLIENTS", span=5)
    r += 1
    vals = [
        ("Enterprise (Plus)", "Shopify Plus — customizable, 99.99% uptime, advanced APIs", "Heinz, Kylie Cosmetics, Netflix Merch, Supreme, Gymshark"),
        ("Mid-Market", "All-in-one platform with 16,000+ apps, Shopify Markets", "Allbirds, MVMT, Bombas"),
        ("SMBs & Solopreneurs", "Easy setup, unified commerce, low total cost of ownership", "Millions of DTC brands globally"),
        ("B2B Buyers (Shopify B2B)", "Wholesale ordering portals, net terms, custom pricing", "Large wholesale merchants"),
    ]
    header_row(ws, r, [2, 3, 4], ["Segment", "Value Proposition", "Example Clients"])
    r += 1
    for i, (seg, vp, cl) in enumerate(vals):
        bg = ALT_ROW if i % 2 == 0 else WHITE
        for col, val in zip([2, 3, 4], [seg, vp, cl]):
            wc(ws, r, col, val, bg=bg, border=True, wrap=True)
        r += 1

    r += 1
    # Seasonality
    section_header(ws, r, 2, "SEASONALITY & MARGIN STRUCTURE", span=5)
    r += 1
    seas = [
        ("Seasonality", "Q4 (Oct-Dec) is peak: Holiday/BFCM drives ~30-33% of annual GMV. Q1 weakest quarter."),
        ("Subscription Margin", "~55-58% gross margin — high-quality recurring SaaS revenue"),
        ("Merchant Solutions Margin", "~40-42% gross margin — lower due to payment processing costs"),
        ("Blended Gross Margin", "~50-51% (FY2025 est.) — improving as higher-margin segments scale"),
        ("FCF Margin", "17% FY2025; targeting 20%+ longer-term"),
        ("Operating Leverage", "Fixed costs spread over growing GMV base; S&M declining as % of revenue"),
    ]
    for i, (k, v) in enumerate(seas):
        bg = ALT_ROW if i % 2 == 0 else WHITE
        wc(ws, r, 2, k, bold=True, bg=bg, border=True)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
        wc(ws, r, 3, v, bg=bg, border=True, wrap=True)
        ws.row_dimensions[r].height = 30
        r += 1


# ══════════════════════════════════════════════════════════════════════════════
# TAB 3 – MOAT
# ══════════════════════════════════════════════════════════════════════════════
def build_moat(wb):
    ws = wb.create_sheet("3. Moat")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, {1: 3, 2: 28, 3: 50, 4: 20})
    freeze(ws, "B3")

    r = 1
    ws.merge_cells(f"B{r}:D{r}")
    wc(ws, r, 2, "SHOPIFY — COMPETITIVE MOAT ANALYSIS", bold=True, fg=HEADER_FG,
       bg=SHOP_GREEN, size=FONT_SIZE+2, align="center")
    r += 2

    section_header(ws, r, 2, "MOAT RATING: WIDE  |  Source: Network Effects + Switching Costs + Scale Economies", span=3)
    r += 2

    moat_items = [
        ("SWITCHING COSTS", [
            ("Platform Stickiness", "Merchants integrate Shopify deeply into operations (ERP, CRM, POS). Migration costs $3K-$25K and 6-12 weeks. Churn rate historically <5% annually."),
            ("App Ecosystem Lock-in", "16,000+ apps in the Shopify App Store create a web of integrations specific to Shopify APIs. Migrating means losing app compatibility."),
            ("Merchant Data", "Years of customer data, order history, analytics dashboards — all locked in Shopify's infrastructure."),
        ]),
        ("NETWORK EFFECTS", [
            ("Partner Ecosystem", "Over $1B partner economy; 40,000+ Shopify Partners (developers, agencies, theme makers) who only build for Shopify."),
            ("Shop App Discovery", "78M+ registered Shop App users create a consumer discovery loop feeding merchant revenue."),
            ("Two-sided Marketplace", "More merchants → more apps/tools built → better platform → attracts more merchants (self-reinforcing)."),
        ]),
        ("SCALE ECONOMIES", [
            ("Payment Volume", "Shopify Payments processes $200B+ per year, enabling better interchange rates than any single merchant could negotiate."),
            ("Shopify Capital", "Proprietary underwriting using GMV/sales data gives Shopify lower default rates than traditional lenders → cheaper capital."),
            ("Logistics Intelligence", "Even after exiting Flexport JV, Shopify has carrier data from billions of shipments improving Shopify Shipping rates."),
        ]),
        ("BRAND & DISTRIBUTION", [
            ("Merchant Trust", "Shopify's brand is synonymous with DTC e-commerce. It powers 73% of top 800 DTC brands."),
            ("Developer Mindshare", "Shopify Liquid, Hydrogen, Oxygen — full-stack commerce developer ecosystem. First choice for new e-com projects."),
            ("Market Share Leadership", "29% U.S. ecommerce platform market share — 2x the #2 competitor WooCommerce in top-tier merchants."),
        ]),
        ("RISKS TO MOAT", [
            ("Amazon Competition", "Amazon's own seller programs (FBA, Buy with Prime) target the same merchants. Risk of disintermediation."),
            ("WooCommerce/Headless", "Open-source alternatives (WooCommerce, Medusa) appeal to cost-sensitive or technically sophisticated merchants."),
            ("TikTok Shop / Social Commerce", "Emerging social-first platforms could shift merchant acquisition away from Shopify's traditional channels."),
            ("Macro Risk", "If merchant failures rise in recession, Shopify Capital credit losses increase and GMV shrinks."),
        ]),
    ]

    for category, items in moat_items:
        section_header(ws, r, 2, category, span=3,
                        bg=SUBHDR_BG if "RISK" not in category else "FFEEEE",
                        fg=SHOP_DARK if "RISK" not in category else "8B0000")
        r += 1
        header_row(ws, r, [2, 3, 4], ["Factor", "Detail", "Strength"])
        r += 1
        strengths = {"SWITCHING COSTS": "High", "NETWORK EFFECTS": "High",
                     "SCALE ECONOMIES": "Medium-High", "BRAND & DISTRIBUTION": "High",
                     "RISKS TO MOAT": "Monitor"}
        for i, (factor, detail) in enumerate(items):
            bg = ALT_ROW if i % 2 == 0 else WHITE
            wc(ws, r, 2, factor, bold=True, bg=bg, border=True)
            wc(ws, r, 3, detail, bg=bg, border=True, wrap=True)
            ws.row_dimensions[r].height = 40
            s = strengths[category]
            col = GREEN if "RISK" not in category else RED
            wc(ws, r, 4, s, bg=bg, border=True, fg=col, bold=True, align="center")
            r += 1
        r += 1

    # Moat summary
    section_header(ws, r, 2, "OVERALL MOAT ASSESSMENT", span=3)
    r += 1
    summary = (
        "Shopify has a WIDE MOAT driven by three reinforcing pillars: (1) High switching costs from deep platform integration, "
        "(2) Network effects from its massive partner/app ecosystem, and (3) Scale economies in payments and capital. "
        "The moat is defended but not impregnable — Amazon and social commerce platforms are credible threats. "
        "Shopify's strategic response (Shopify Magic AI, Shop App, B2B expansion) suggests management is aware and investing "
        "to extend the moat. For a long-term owner, the switching costs alone are worth significant premium valuation."
    )
    ws.merge_cells(f"B{r}:D{r+3}")
    c = ws.cell(row=r, column=2, value=summary)
    c.font = mf(size=FONT_SIZE)
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    c.fill = mfill(LIGHT_GREEN)
    for i in range(4):
        ws.row_dimensions[r+i].height = 20


# ══════════════════════════════════════════════════════════════════════════════
# TAB 4 – INCOME STATEMENT
# ══════════════════════════════════════════════════════════════════════════════
def build_income_statement(wb):
    ws = wb.create_sheet("4. Income Statement")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, {1: 3, 2: 35, 3: 18, 4: 18, 5: 18, 6: 18, 7: 18})
    freeze(ws, "C3")

    r = 1
    ws.merge_cells(f"B{r}:G{r}")
    wc(ws, r, 2, "SHOPIFY — INCOME STATEMENT (USD Millions)", bold=True, fg=HEADER_FG,
       bg=SHOP_GREEN, size=FONT_SIZE+2, align="center")
    r += 2

    header_row(ws, r, [2, 3, 4, 5, 6, 7],
               ["Metric", "FY2022", "FY2023", "FY2024", "FY2025", "YoY Chg %"])
    r += 1

    is_data = [
        ("REVENUE", "", "", "", "", ""),
        ("  Subscription Solutions", "1,489", "1,841", "2,350", "3,120", "+32.8%"),
        ("  Merchant Solutions",     "4,355", "5,222", "6,530", "8,440", "+29.2%"),
        ("Total Revenue",            "5,844", "7,061", "8,880", "11,556", "+30.1%"),
        ("", "", "", "", "", ""),
        ("COST OF REVENUE", "", "", "", "", ""),
        ("  Subscription CoR",       "366",   "406",   "470",   "560",   "+19.1%"),
        ("  Merchant Solutions CoR", "2,721", "3,197", "3,894", "5,065", "+30.1%"),
        ("Total Cost of Revenue",    "3,087", "3,603", "4,364", "5,625", "+28.9%"),
        ("", "", "", "", "", ""),
        ("Gross Profit",             "2,757", "3,458", "4,516", "5,931", "+31.3%"),
        ("Gross Margin %",           "47.2%", "49.0%", "50.9%", "51.3%", "+0.4pp"),
        ("", "", "", "", "", ""),
        ("OPERATING EXPENSES", "", "", "", "", ""),
        ("  Sales & Marketing",      "1,023", "986",   "1,150", "1,380", "+20.0%"),
        ("  R&D",                    "1,268", "1,243", "1,370", "1,580", "+15.3%"),
        ("  G&A",                    "376",   "325",   "485",   "580",   "+19.6%"),
        ("  Total OpEx",             "2,667", "2,554", "3,005", "3,540", "+17.8%"),
        ("", "", "", "", "", ""),
        ("Operating Income (GAAP)",  "90",    "904",   "1,080", "1,470", "+36.1%"),
        ("Operating Margin %",       "1.5%",  "12.8%", "12.2%", "12.7%", "+0.5pp"),
        ("", "", "", "", "", ""),
        ("Interest & Other Income",  "(19)",  "143",   "210",   "280",   "+33.3%"),
        ("Pre-tax Income",           "71",    "1,047", "1,290", "1,750", "+35.7%"),
        ("Income Tax",               "13",    "(73)",  "(95)",  "(120)", "+26.3%"),
        ("Net Income (GAAP)",        "58",    "974",   "1,195", "1,630", "+36.4%"),
        ("Net Margin %",             "1.0%",  "13.8%", "13.5%", "14.1%", "+0.6pp"),
        ("", "", "", "", "", ""),
        ("ADJUSTED METRICS", "", "", "", "", ""),
        ("Adjusted Operating Income","(269)", "912",   "1,370", "1,970", "+43.8%"),
        ("Adj. Operating Margin %",  "-4.6%", "12.9%", "15.4%", "17.0%", "+1.6pp"),
        ("Free Cash Flow",           "(175)", "905",   "1,361", "2,000", "+47.0%"),
        ("FCF Margin %",             "-3.0%", "12.8%", "15.3%", "17.3%", "+2.0pp"),
        ("Diluted EPS (GAAP)",       "$0.05", "$0.77", "$0.94", "$1.27", "+35.1%"),
    ]

    highlight = {"Total Revenue", "Gross Profit", "Gross Margin %", "Operating Income (GAAP)",
                 "Net Income (GAAP)", "Free Cash Flow", "FCF Margin %"}
    for row_vals in is_data:
        label = row_vals[0]
        if label == "":
            r += 1
            continue
        if label in ("REVENUE", "COST OF REVENUE", "OPERATING EXPENSES", "ADJUSTED METRICS"):
            section_header(ws, r, 2, label, span=6, bg=SUBHDR_BG, fg=SHOP_DARK)
            r += 1
            continue
        bold = label in highlight
        bg = LIGHT_GREEN if bold else (ALT_ROW if r % 2 == 0 else WHITE)
        fg_last = GREEN if "%" in str(row_vals[5]) and "+" in str(row_vals[5]) else (RED if "-" in str(row_vals[5]) else "000000")
        for col, val in zip([2, 3, 4, 5, 6], row_vals[:5]):
            wc(ws, r, col, val, bold=bold, bg=bg, border=True,
               align="right" if col > 2 else "left")
        wc(ws, r, 7, row_vals[5], bold=bold, bg=bg, border=True, fg=fg_last, align="center")
        r += 1


# ══════════════════════════════════════════════════════════════════════════════
# TAB 5 – BALANCE SHEET
# ══════════════════════════════════════════════════════════════════════════════
def build_balance_sheet(wb):
    ws = wb.create_sheet("5. Balance Sheet")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, {1: 3, 2: 35, 3: 18, 4: 18, 5: 18, 6: 18})
    freeze(ws, "C3")

    r = 1
    ws.merge_cells(f"B{r}:F{r}")
    wc(ws, r, 2, "SHOPIFY — BALANCE SHEET (USD Millions)", bold=True, fg=HEADER_FG,
       bg=SHOP_GREEN, size=FONT_SIZE+2, align="center")
    r += 2

    header_row(ws, r, [2, 3, 4, 5, 6],
               ["Metric", "FY2022", "FY2023", "FY2024", "FY2025"])
    r += 1

    bs_data = [
        ("ASSETS", None, None, None, None),
        ("Cash & Equivalents",        "4,882", "5,024", "5,100", "5,200"),
        ("Short-term Investments",    "0",     "0",     "0",     "0"),
        ("Accounts Receivable",       "310",   "484",   "540",   "680"),
        ("Merchant Cash Advances",    "740",   "1,100", "1,500", "1,900"),
        ("Other Current Assets",      "200",   "260",   "310",   "380"),
        ("Total Current Assets",      "6,132", "6,868", "7,450", "8,160"),
        ("", None, None, None, None),
        ("PP&E (net)",                "280",   "300",   "310",   "320"),
        ("Intangibles & Goodwill",    "2,400", "2,100", "1,900", "1,750"),
        ("Other Long-term Assets",    "1,200", "1,400", "1,600", "1,780"),
        ("Total Assets",              "10,012","10,668","11,260","12,010"),
        ("", None, None, None, None),
        ("LIABILITIES", None, None, None, None),
        ("Accounts Payable",          "130",   "170",   "200",   "240"),
        ("Deferred Revenue",          "120",   "145",   "175",   "210"),
        ("Other Current Liabilities", "450",   "520",   "600",   "700"),
        ("Total Current Liabilities", "700",   "835",   "975",   "1,150"),
        ("", None, None, None, None),
        ("Long-term Debt",            "912",   "912",   "920",   "920"),
        ("Other LT Liabilities",      "400",   "360",   "360",   "380"),
        ("Total Liabilities",         "2,012", "2,107", "2,255", "2,450"),
        ("", None, None, None, None),
        ("EQUITY", None, None, None, None),
        ("Common Stock & APIC",       "8,600", "9,200", "9,800", "10,350"),
        ("Retained Earnings (Deficit)", "(600)","(639)", "205",   "1,210"),
        ("Total Shareholders' Equity","8,000", "8,561", "9,005", "11,560"),
        ("", None, None, None, None),
        ("KEY RATIOS", None, None, None, None),
        ("Debt-to-Equity",            "0.11",  "0.11",  "0.10",  "0.08"),
        ("Current Ratio",             "8.76",  "8.22",  "7.64",  "7.10"),
        ("Cash as % of Market Cap",   "3.4%",  "3.5%",  "3.6%",  "3.6%"),
        ("Net Debt / (Net Cash)",     "$(3,970)","$(4,112)","$(4,180)","$(4,280)"),
    ]

    highlight = {"Total Current Assets", "Total Assets", "Total Current Liabilities",
                 "Total Liabilities", "Total Shareholders' Equity"}
    sections = {"ASSETS", "LIABILITIES", "EQUITY", "KEY RATIOS"}

    for row_vals in bs_data:
        label = row_vals[0]
        if label == "":
            r += 1
            continue
        if label in sections:
            section_header(ws, r, 2, label, span=5, bg=SUBHDR_BG, fg=SHOP_DARK)
            r += 1
            continue
        bold = label in highlight
        bg = LIGHT_GREEN if bold else (ALT_ROW if r % 2 == 0 else WHITE)
        for col, val in zip([2, 3, 4, 5, 6], row_vals):
            if val is None:
                val = ""
            wc(ws, r, col, val, bold=bold, bg=bg, border=True,
               align="right" if col > 2 else "left")
        r += 1


# ══════════════════════════════════════════════════════════════════════════════
# TAB 6 – CASH FLOW ANALYSIS
# ══════════════════════════════════════════════════════════════════════════════
def build_cash_flow(wb):
    ws = wb.create_sheet("6. Cash Flow Analysis")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, {1: 3, 2: 38, 3: 18, 4: 18, 5: 18, 6: 18})
    freeze(ws, "C3")

    r = 1
    ws.merge_cells(f"B{r}:F{r}")
    wc(ws, r, 2, "SHOPIFY — CASH FLOW ANALYSIS (USD Millions)", bold=True, fg=HEADER_FG,
       bg=SHOP_GREEN, size=FONT_SIZE+2, align="center")
    r += 2

    header_row(ws, r, [2, 3, 4, 5, 6],
               ["Metric", "FY2022", "FY2023", "FY2024", "FY2025"])
    r += 1

    cf_data = [
        ("OPERATING ACTIVITIES", None, None, None, None),
        ("Net Income",                    "58",    "974",    "1,195", "1,630"),
        ("D&A",                           "180",   "190",    "200",   "210"),
        ("Stock-based Compensation",      "600",   "370",    "390",   "420"),
        ("Working Capital Changes",       "(200)", "100",    "(180)", "(80)"),
        ("Merchant Cash Adv. (net)",      "(300)", "(360)",  "(400)", "(400)"),
        ("Other Operating",               "62",    "(28)",   "(45)",  "61"),
        ("Cash from Operations",          "400",   "1,246",  "1,600", "1,841"),
        ("", None, None, None, None),
        ("INVESTING ACTIVITIES", None, None, None, None),
        ("Capital Expenditures",          "(160)", "(150)",  "(180)", "(200)"),
        ("Acquisitions / Investments",    "(3,400)","(180)", "(200)", "(250)"),
        ("Other Investing",               "200",   "(80)",   "100",   "50"),
        ("Cash from Investing",           "(3,360)","(410)", "(280)", "(400)"),
        ("", None, None, None, None),
        ("FINANCING ACTIVITIES", None, None, None, None),
        ("Share Repurchases",             "0",     "0",      "0",     "(500)"),
        ("Debt Issuance / (Repayment)",   "0",     "0",      "0",     "0"),
        ("Stock Options Exercised",       "200",   "150",    "120",   "100"),
        ("Other Financing",               "(50)",  "(50)",   "(50)",  "(735)"),
        ("Cash from Financing",           "150",   "100",    "70",    "(1,135)"),
        ("", None, None, None, None),
        ("FREE CASH FLOW METRICS", None, None, None, None),
        ("Free Cash Flow (OCF - CapEx)",  "240",   "1,096",  "1,420", "1,641"),
        ("Adj. FCF (excl. SBC leases)",  "(175)", "905",    "1,361", "2,000"),
        ("FCF Margin %",                  "-3.0%", "12.8%",  "15.3%", "17.3%"),
        ("FCF Conversion (FCF/Net Inc)",  "N/M",   "93%",    "114%",  "123%"),
        ("", None, None, None, None),
        ("QUALITY OF EARNINGS", None, None, None, None),
        ("OCF / Net Income ratio",        "6.9x",  "1.28x",  "1.34x", "1.13x"),
        ("SBC as % of Revenue",          "10.3%", "5.2%",   "4.4%",  "3.6%"),
        ("CapEx as % of Revenue",        "2.7%",  "2.1%",   "2.0%",  "1.7%"),
        ("Net Change in Cash",            "(2,810)","936",   "1,390", "306"),
    ]

    highlight = {"Cash from Operations", "Free Cash Flow (OCF - CapEx)",
                 "Adj. FCF (excl. SBC leases)", "FCF Margin %"}
    sections = {"OPERATING ACTIVITIES", "INVESTING ACTIVITIES",
                "FINANCING ACTIVITIES", "FREE CASH FLOW METRICS", "QUALITY OF EARNINGS"}

    for row_vals in cf_data:
        label = row_vals[0]
        if label == "":
            r += 1
            continue
        if label in sections:
            section_header(ws, r, 2, label, span=5, bg=SUBHDR_BG, fg=SHOP_DARK)
            r += 1
            continue
        bold = label in highlight
        bg = LIGHT_GREEN if bold else (ALT_ROW if r % 2 == 0 else WHITE)
        for col, val in zip([2, 3, 4, 5, 6], row_vals):
            if val is None:
                val = ""
            wc(ws, r, col, val, bold=bold, bg=bg, border=True,
               align="right" if col > 2 else "left")
        r += 1


# ══════════════════════════════════════════════════════════════════════════════
# TAB 7 – RETURN ON CAPITAL
# ══════════════════════════════════════════════════════════════════════════════
def build_return_on_capital(wb):
    ws = wb.create_sheet("7. Return on Capital")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, {1: 3, 2: 38, 3: 18, 4: 18, 5: 18, 6: 18})
    freeze(ws, "C3")

    r = 1
    ws.merge_cells(f"B{r}:F{r}")
    wc(ws, r, 2, "SHOPIFY — RETURN ON CAPITAL ANALYSIS", bold=True, fg=HEADER_FG,
       bg=SHOP_GREEN, size=FONT_SIZE+2, align="center")
    r += 2

    header_row(ws, r, [2, 3, 4, 5, 6], ["Metric", "FY2022", "FY2023", "FY2024", "FY2025"])
    r += 1

    roc_data = [
        ("PROFITABILITY RETURNS", None, None, None, None),
        ("Return on Equity (ROE)",              "0.7%",  "11.4%",  "13.3%",  "14.1%"),
        ("Return on Assets (ROA)",              "0.6%",  "9.1%",   "10.6%",  "13.6%"),
        ("Return on Invested Capital (ROIC)",   "N/M",   "18.2%",  "20.5%",  "30.6%"),
        ("ROIC vs WACC (est. ~10%)",            "Below", "Above",  "Above",  "Well Above"),
        ("", None, None, None, None),
        ("INCREMENTAL CAPITAL EFFICIENCY", None, None, None, None),
        ("Revenue per Employee",                "~$600K","~$680K", "$780K",  "$900K+"),
        ("GMV per Active Merchant",             "$90K",  "$94K",   "$91K",   "$76K"),
        ("Gross Profit per Merchant",           "$1,100","$1,384", "$1,411", "$1,186"),
        ("", None, None, None, None),
        ("RULE OF 50 (SaaS Benchmark)", None, None, None, None),
        ("Revenue Growth Rate",                 "21.4%", "26.1%",  "25.8%",  "30.1%"),
        ("FCF Margin",                          "-3.0%", "12.8%",  "15.3%",  "17.3%"),
        ("Rule of 50 Score (Growth + FCF)",     "18.4",  "38.9",   "41.1",   "47.4"),
        ("Benchmark: Pass = 50+",               "FAIL",  "FAIL",   "FAIL",   "NEAR PASS"),
        ("Note",                                "Improving rapidly", "Strong trajectory",
                                                "Above average SaaS", "Approaching Rule of 50"),
        ("", None, None, None, None),
        ("CAPITAL ALLOCATION SCORECARD", None, None, None, None),
        ("R&D Investment",                      "High",  "Moderate","Moderate","Growing"),
        ("Acquisitions",                        "Deliverr", "Minimal","Minimal","Minimal"),
        ("Share Repurchases",                   "None",  "None",   "None",   "$500M authorized"),
        ("Dividends",                           "None",  "None",   "None",   "None"),
        ("Net Cash Position",                   "$3,970M","$4,112M","$4,180M","$4,280M"),
        ("Capital Efficiency Trend",            "Poor",  "Good",   "Very Good","Excellent"),
    ]

    highlight = {"Return on Invested Capital (ROIC)", "Rule of 50 Score (Growth + FCF)", "Free Cash Flow"}
    sections = {"PROFITABILITY RETURNS", "INCREMENTAL CAPITAL EFFICIENCY",
                "RULE OF 50 (SaaS Benchmark)", "CAPITAL ALLOCATION SCORECARD"}

    for row_vals in roc_data:
        label = row_vals[0]
        if label == "":
            r += 1
            continue
        if label in sections:
            section_header(ws, r, 2, label, span=5, bg=SUBHDR_BG, fg=SHOP_DARK)
            r += 1
            continue
        bold = label in highlight
        bg = LIGHT_GREEN if bold else (ALT_ROW if r % 2 == 0 else WHITE)
        for col, val in zip([2, 3, 4, 5, 6], row_vals):
            if val is None:
                val = ""
            wc(ws, r, col, val, bold=bold, bg=bg, border=True,
               align="right" if col > 2 else "left")
        r += 1


# ══════════════════════════════════════════════════════════════════════════════
# TAB 8 – MANAGEMENT
# ══════════════════════════════════════════════════════════════════════════════
def build_management(wb):
    ws = wb.create_sheet("8. Management")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, {1: 3, 2: 30, 3: 55, 4: 18})
    freeze(ws, "B3")

    r = 1
    ws.merge_cells(f"B{r}:D{r}")
    wc(ws, r, 2, "SHOPIFY — MANAGEMENT QUALITY ANALYSIS", bold=True, fg=HEADER_FG,
       bg=SHOP_GREEN, size=FONT_SIZE+2, align="center")
    r += 2

    # Key executives
    section_header(ws, r, 2, "KEY EXECUTIVES", span=3)
    r += 1
    header_row(ws, r, [2, 3, 4], ["Name / Role", "Background & Tenure", "Assessment"])
    r += 1
    execs = [
        ("Tobias (Tobi) Lütke\nCEO & Co-Founder",
         "Founded Shopify in 2006 from his own e-commerce frustration. Computer programmer by training. "
         "Holds supervoting shares (Class B). Compensation mostly equity-aligned ($1 salary). "
         "Named one of Glassdoor's top CEOs. Deep technical background.",
         "★★★★★ Owner-operator mindset. Long-term oriented."),
        ("Harley Finkelstein\nPresident",
         "Joined 2010. Lawyer by training. Runs merchant relationships, policy, partnerships. "
         "Highly visible public face of Shopify. Critical to ecosystem building.",
         "★★★★☆ Excellent operator and brand ambassador."),
        ("Jeff Hoffmeister\nCFO",
         "Joined 2022 from Goldman Sachs (16 years). Brings institutional capital discipline. "
         "Led Shopify's exit from logistics/Flexport and margin improvement story.",
         "★★★★☆ Disciplined capital allocator; margin focus."),
        ("Mikhail Parakhin\nCTO",
         "Joined 2024 from Microsoft Bing AI division. Leading Shopify's AI-first commerce strategy "
         "(Shopify Magic, Sidekick AI assistant). Critical hire for next growth phase.",
         "★★★★☆ AI-era leadership hire."),
    ]
    for i, (name, bg_txt, assess) in enumerate(execs):
        bg = ALT_ROW if i % 2 == 0 else WHITE
        wc(ws, r, 2, name, bold=True, bg=bg, border=True, wrap=True)
        ws.row_dimensions[r].height = 50
        wc(ws, r, 3, bg_txt, bg=bg, border=True, wrap=True)
        wc(ws, r, 4, assess, bg=bg, border=True, wrap=True)
        r += 1

    r += 1
    # CEO owner behavior
    section_header(ws, r, 2, "DOES CEO ACT LIKE AN OWNER?", span=3)
    r += 1
    owner_items = [
        ("$1 Base Salary", "Tobi Lütke takes only $1 base salary; compensation is 100% performance equity.", "YES — fully aligned"),
        ("Equity Ownership", "Lütke owns ~7% of Shopify through Class B supervoting shares (~$10B+ stake). He is the largest individual shareholder.", "YES — strong skin in game"),
        ("Stock Sales (ASDP)", "Filed to sell up to 2.56M shares via pre-arranged ASDP through Dec 2025. This is diversification — not a bearish signal. Total sales modest vs. total holdings.", "NEUTRAL — planned diversification"),
        ("Compensation Package 2024", "Awarded ~$200M in options/shares in Feb 2024; ISS and Glass Lewis flagged as excessive. Shareholders approved.", "WATCH — magnitude is large"),
        ("Long-Term Thinking", "Exited logistics (Flexport JV) to refocus core platform. Chose profitability over growth theater. Invests in AI infrastructure for 5-year payoff.", "YES — long-term orientation"),
        ("Operator Track Record", "Grew Shopify from $105M (2015 IPO) to $11.5B revenue in 10 years. Compounding at ~40% CAGR.", "YES — exceptional track record"),
        ("Capital Buybacks", "Authorized $2B buyback in Feb 2026 — first time. Signals confidence and balance sheet strength.", "YES — shareholder-friendly shift"),
    ]
    header_row(ws, r, [2, 3, 4], ["Criterion", "Evidence", "Verdict"])
    r += 1
    for i, (crit, ev, verdict) in enumerate(owner_items):
        bg = ALT_ROW if i % 2 == 0 else WHITE
        v_color = GREEN if "YES" in verdict else (GOLD if "WATCH" in verdict or "NEUTRAL" in verdict else "000000")
        wc(ws, r, 2, crit, bold=True, bg=bg, border=True)
        wc(ws, r, 3, ev, bg=bg, border=True, wrap=True)
        ws.row_dimensions[r].height = 40
        wc(ws, r, 4, verdict, bg=bg, border=True, fg=v_color, bold=True)
        r += 1

    r += 1
    # Capital allocation
    section_header(ws, r, 2, "CAPITAL ALLOCATION HISTORY", span=3)
    r += 1
    alloc = [
        ("2021", "Acquired Deliverr for $2.1B (logistics). Later exited. Expensive lesson in non-core M&A."),
        ("2023", "Sold Deliverr/logistics to Flexport for Flexport equity + $550M. Disciplined retreat to core platform."),
        ("2024", "Minimal M&A; invested in R&D (AI, Hydrogen storefront, B2B). FCF used for platform. SBC declining as % of revenue."),
        ("2025", "Launched $2B share buyback. Net cash position ~$4.3B. No debt concerns. R&D investment in Shopify Magic AI."),
        ("2026+", "B2B expansion, international markets (Shopify Markets Pro), AI-commerce tools. Disciplined organic investment."),
    ]
    header_row(ws, r, [2, 3], ["Year", "Capital Allocation Decision & Assessment"])
    r += 1
    for i, (yr, dec) in enumerate(alloc):
        bg = ALT_ROW if i % 2 == 0 else WHITE
        wc(ws, r, 2, yr, bold=True, bg=bg, border=True, align="center")
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
        wc(ws, r, 3, dec, bg=bg, border=True, wrap=True)
        ws.row_dimensions[r].height = 35
        r += 1

    r += 1
    # Planting seeds vs borrowing
    section_header(ws, r, 2, "PLANTING SEEDS vs. BORROWING FROM THE FUTURE", span=3)
    r += 1
    seeds = [
        ("Shopify Magic (AI)", "Generative AI built into every merchant workflow. Launched 2023; iterating fast. Designed for 2025-2030 commerce.", "SEED — multi-year investment"),
        ("Shopify B2B", "Enterprise wholesale platform. Attacking $9T B2B commerce market. Early innings.", "SEED — 5-year horizon"),
        ("Shop App / Shop Pay Network", "Consumer side of commerce flywheel. 78M+ registered users. Loyalty/ads potential.", "SEED — building network"),
        ("SBC Expense", "SBC at 3.6% of revenue in FY2025 (down from 10% in 2022). Not borrowing from future via excessive dilution.", "DISCIPLINED — improving"),
        ("Shopify Capital", "Merchant lending from internally generated cash. High-return use of capital with data advantage.", "SEED — compounding asset"),
    ]
    header_row(ws, r, [2, 3, 4], ["Initiative", "Description", "Assessment"])
    r += 1
    for i, (init, desc, assess) in enumerate(seeds):
        bg = ALT_ROW if i % 2 == 0 else WHITE
        wc(ws, r, 2, init, bold=True, bg=bg, border=True)
        wc(ws, r, 3, desc, bg=bg, border=True, wrap=True)
        ws.row_dimensions[r].height = 40
        v_color = GREEN if "SEED" in assess or "DISCIPLINED" in assess else GOLD
        wc(ws, r, 4, assess, bg=bg, border=True, fg=v_color, bold=True)
        r += 1


# ══════════════════════════════════════════════════════════════════════════════
# TAB 9 – RISKS
# ══════════════════════════════════════════════════════════════════════════════
def build_risks(wb):
    ws = wb.create_sheet("9. Risks")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, {1: 3, 2: 30, 3: 45, 4: 20, 5: 20})
    freeze(ws, "B3")

    r = 1
    ws.merge_cells(f"B{r}:E{r}")
    wc(ws, r, 2, "SHOPIFY — RISK ANALYSIS", bold=True, fg=HEADER_FG,
       bg=SHOP_GREEN, size=FONT_SIZE+2, align="center")
    r += 2

    header_row(ws, r, [2, 3, 4, 5],
               ["Risk Category", "Description", "Probability", "Impact"])
    r += 1

    risks = [
        ("MACRO & TRADE RISKS", [
            ("US-China Trade War / Tariffs",
             "Tariff escalation hurts cross-border GMV. De minimis rule changes (suspended Aug 2025) "
             "reduce competitive moat for Shopify merchants importing from Asia. "
             "Merchants selling on Amazon/Shopify with China-sourced goods face margin compression.",
             "High", "High"),
            ("Consumer Spending Slowdown",
             "Shopify's GMV is directly tied to consumer spending. A US recession reduces merchant sales "
             "→ less GMV → less Merchant Solutions revenue. Historical precedent: COVID-era boost reversed.",
             "Medium", "High"),
            ("Inflation & SMB Stress",
             "Rising input costs squeeze SMB margins, leading to merchant churn or reduced plan tiers. "
             "Shopify Capital credit losses increase if merchants default.",
             "Medium", "Medium"),
        ]),
        ("COMPETITIVE RISKS", [
            ("Amazon Competition",
             "Amazon's Buy with Prime, FBA, and own marketplace compete directly for merchants. "
             "Risk: Amazon re-platforms DTC brands onto its own ecosystem.",
             "Medium", "High"),
            ("TikTok Shop & Social Commerce",
             "Social-first platforms (TikTok Shop, Instagram Checkout) bypass traditional merchant websites. "
             "Shopify has integrated these but relies on third-party platforms.",
             "Medium", "Medium"),
            ("WooCommerce / Open-Source",
             "Enterprise/technical merchants may prefer headless/open-source solutions for cost savings. "
             "Medusa.js, commerce.js are emerging alternatives.",
             "Low", "Medium"),
        ]),
        ("OPERATIONAL RISKS", [
            ("Platform Outages",
             "Shopify's infrastructure serves millions of merchants. Any BFCM downtime is catastrophic "
             "for merchant trust. Major outage in 2023 demonstrated risk.",
             "Low", "Very High"),
            ("Shopify Capital Credit Risk",
             "Merchant cash advance defaults increase in economic downturns. $1.9B+ in outstanding capital. "
             "Underwriting model uses GMV data but not fully tested in severe downturn.",
             "Medium", "High"),
            ("Key Person Risk (Tobi Lütke)",
             "Tobi is the founder-CEO with supervoting shares. His departure or shift in focus would be "
             "materially negative for the company's direction and culture.",
             "Low", "Very High"),
        ]),
        ("REGULATORY & LEGAL RISKS", [
            ("Payments Regulation",
             "Shopify Payments is subject to banking and payment regulations globally. Changes in "
             "interchange regulation (like Durbin Amendment expansion) could compress payment margins.",
             "Medium", "Medium"),
            ("Data Privacy (GDPR, CCPA)",
             "Shopify stores merchant customer data across 175+ countries. Privacy law enforcement is "
             "intensifying globally. Compliance costs are rising.",
             "Medium", "Low"),
            ("Antitrust / Platform Regulation",
             "As market share grows to 30%+, regulatory scrutiny on App Store fees and exclusivity "
             "practices may increase.",
             "Low", "Medium"),
        ]),
        ("VALUATION RISK", [
            ("Premium Valuation Compression",
             "At 9-10x P/S and 120x trailing P/E, SHOP is priced for continued 25-30% growth. "
             "Any revenue deceleration below 20% risks significant multiple compression.",
             "Medium", "High"),
            ("Rising Interest Rates",
             "High-duration growth stocks are sensitive to discount rate increases. Rate hikes in "
             "2022-2023 caused SHOP to fall 80%+ from all-time highs.",
             "Low-Medium", "High"),
        ]),
    ]

    for category, risk_list in risks:
        section_header(ws, r, 2, category, span=4, bg=SUBHDR_BG, fg=SHOP_DARK)
        r += 1
        for i, (name, desc, prob, impact) in enumerate(risk_list):
            bg = ALT_ROW if i % 2 == 0 else WHITE
            wc(ws, r, 2, name, bold=True, bg=bg, border=True)
            wc(ws, r, 3, desc, bg=bg, border=True, wrap=True)
            ws.row_dimensions[r].height = 55
            p_color = RED if prob == "High" else (GOLD if prob == "Medium" else GREEN)
            i_color = RED if impact in ("High", "Very High") else (GOLD if impact == "Medium" else GREEN)
            wc(ws, r, 4, prob, bg=bg, border=True, fg=p_color, bold=True, align="center")
            wc(ws, r, 5, impact, bg=bg, border=True, fg=i_color, bold=True, align="center")
            r += 1
        r += 1


# ══════════════════════════════════════════════════════════════════════════════
# TAB 10 – VALUATION
# ══════════════════════════════════════════════════════════════════════════════
def build_valuation(wb):
    ws = wb.create_sheet("10. Valuation")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, {1: 3, 2: 35, 3: 25, 4: 25, 5: 25})
    freeze(ws, "B3")

    r = 1
    ws.merge_cells(f"B{r}:E{r}")
    wc(ws, r, 2, "SHOPIFY — VALUATION ANALYSIS", bold=True, fg=HEADER_FG,
       bg=SHOP_GREEN, size=FONT_SIZE+2, align="center")
    r += 2

    # Current metrics
    section_header(ws, r, 2, "CURRENT MARKET METRICS (Apr 12, 2026)", span=4)
    r += 1
    metrics = [
        ("Stock Price", "$111.20", "52-wk High: $182.19", "52-wk Low: $78.00"),
        ("Market Cap", "~$143B", "Shares Out: ~1.27B", "YTD Change: -30%"),
        ("Enterprise Value", "~$139B", "Net Cash: $4.3B", "EV = MktCap - NetCash"),
        ("P/E (Trailing GAAP)", "88x", "P/E (NTM)", "~70x"),
        ("P/S (Trailing)", "12.4x", "EV/Sales (NTM)", "9.4x"),
        ("EV/EBITDA (NTM)", "~45x", "EV/FCF", "~70x"),
        ("PEG Ratio", "~2.3x", "FCF Yield", "~1.4%"),
    ]
    header_row(ws, r, [2, 3, 4, 5], ["Metric", "Value", "Context 1", "Context 2"])
    r += 1
    for i, row_vals in enumerate(metrics):
        bg = ALT_ROW if i % 2 == 0 else WHITE
        for col, val in zip([2, 3, 4, 5], row_vals):
            wc(ws, r, col, val, bg=bg, border=True, align="right" if col > 2 else "left")
        r += 1

    r += 1
    # DCF Analysis
    section_header(ws, r, 2, "DISCOUNTED CASH FLOW (DCF) ANALYSIS", span=4)
    r += 1
    header_row(ws, r, [2, 3, 4, 5], ["Scenario", "Assumptions", "Intrinsic Value", "Upside / Downside"])
    r += 1
    dcf_scenarios = [
        ("Bull Case",
         "Rev CAGR 28% (FY25-30), FCF margin expands to 25%, terminal growth 4%, WACC 10%",
         "$185-200", "+66-80% upside"),
        ("Base Case",
         "Rev CAGR 22% (FY25-30), FCF margin expands to 20%, terminal growth 3%, WACC 11%",
         "$135-155", "+21-39% upside"),
        ("Bear Case",
         "Rev CAGR 15% (FY25-30), FCF margin stays 17%, terminal growth 2%, WACC 12%",
         "$80-95", "-15 to -28% downside"),
    ]
    for i, (scen, assum, val, updown) in enumerate(dcf_scenarios):
        bg = ALT_ROW if i % 2 == 0 else WHITE
        c_fg = GREEN if "upside" in updown else RED
        wc(ws, r, 2, scen, bold=True, bg=bg, border=True)
        wc(ws, r, 3, assum, bg=bg, border=True, wrap=True)
        ws.row_dimensions[r].height = 45
        wc(ws, r, 4, val, bold=True, bg=bg, border=True, align="center")
        wc(ws, r, 5, updown, bold=True, bg=bg, border=True, fg=c_fg, align="center")
        r += 1

    r += 1
    # Comparable companies
    section_header(ws, r, 2, "COMPARABLE COMPANY ANALYSIS", span=4)
    r += 1
    header_row(ws, r, [2, 3, 4, 5], ["Company", "P/S (NTM)", "Rev Growth", "FCF Margin"])
    r += 1
    comps = [
        ("Shopify (SHOP)", "9.4x", "~25-28%", "17%"),
        ("HubSpot (HUBS)", "8.5x", "~17%", "~15%"),
        ("Salesforce (CRM)", "6.2x", "~9%", "~33%"),
        ("Wix (WIX)", "3.8x", "~12%", "~18%"),
        ("BigCommerce (BIGC)", "2.1x", "~7%", "Negative"),
        ("Amazon (eComm segment)", "N/A", "~10%", "N/A"),
    ]
    for i, row_vals in enumerate(comps):
        bg = LIGHT_GREEN if row_vals[0].startswith("Shopify") else (ALT_ROW if i % 2 == 0 else WHITE)
        bold = row_vals[0].startswith("Shopify")
        for col, val in zip([2, 3, 4, 5], row_vals):
            wc(ws, r, col, val, bold=bold, bg=bg, border=True,
               align="right" if col > 2 else "left")
        r += 1

    r += 1
    # Safety of margin assessment
    section_header(ws, r, 2, "MARGIN OF SAFETY ASSESSMENT", span=4)
    r += 1
    mos_text = (
        "At $111.20 (Apr 12, 2026), Shopify trades at a ~30% discount to its 52-week high of $182.19 "
        "and near the low end of analyst consensus ($130-$220 range, median $181). "
        "\n\nBase Case DCF suggests $135-155 intrinsic value, implying 21-39% upside from current levels. "
        "The selloff appears driven by macro fears (tariffs, consumer slowdown) rather than "
        "fundamental business deterioration — revenue and FCF continue to grow at 30%+. "
        "\n\nFor a long-term owner: current price offers a meaningful margin of safety vs. "
        "3-5 year growth trajectory. The business continues to compound. Key risk: multiple "
        "compression if growth decelerates below 20% or macro worsens materially. "
        "\n\nVerdict: ATTRACTIVE at current levels for patient investors with 3-5 year horizon."
    )
    ws.merge_cells(f"B{r}:E{r+7}")
    c = ws.cell(row=r, column=2, value=mos_text)
    c.font = mf(size=FONT_SIZE)
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    c.fill = mfill(LIGHT_GREEN)
    for i in range(8):
        ws.row_dimensions[r+i].height = 22


# ══════════════════════════════════════════════════════════════════════════════
# TAB 11 – MARKET SENTIMENT
# ══════════════════════════════════════════════════════════════════════════════
def build_market_sentiment(wb):
    ws = wb.create_sheet("11. Market Sentiment")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, {1: 3, 2: 32, 3: 52, 4: 18})
    freeze(ws, "B3")

    r = 1
    ws.merge_cells(f"B{r}:D{r}")
    wc(ws, r, 2, "SHOPIFY — MARKET SENTIMENT & TRENDS", bold=True, fg=HEADER_FG,
       bg=SHOP_GREEN, size=FONT_SIZE+2, align="center")
    r += 2

    section_header(ws, r, 2, "ANALYST CONSENSUS (April 2026)", span=3)
    r += 1
    analyst_data = [
        ("Analyst Rating", "BUY consensus — 75% Buy/Strong Buy, 25% Hold, 0% Sell (60 analysts)"),
        ("Median Price Target", "$181.00 (range: $130-$220)"),
        ("Average Price Target", "$161.05"),
        ("Current Price", "$111.20 — trading at significant discount to consensus"),
        ("Bull Thesis", "30% revenue growth + FCF expansion + $300B+ GMV engine with wide moat"),
        ("Bear Thesis", "Premium valuation, tariff headwinds, Amazon/social commerce threats"),
        ("2026 Revenue Estimate", "$14.9B (avg) — +29% YoY"),
        ("2026 EBITDA Estimate", "$2.7B (avg) — +84% YoY"),
    ]
    for i, (k, v) in enumerate(analyst_data):
        bg = ALT_ROW if i % 2 == 0 else WHITE
        wc(ws, r, 2, k, bold=True, bg=bg, border=True)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
        wc(ws, r, 3, v, bg=bg, border=True, wrap=True)
        r += 1

    r += 1
    section_header(ws, r, 2, "RECENT NEWS & CATALYSTS (2025-2026)", span=3)
    r += 1
    news = [
        ("Feb 2026 — Q4 2025 Earnings Beat", "Revenue $3.67B (+30.6% YoY), FCF $700M+. $2B buyback announced. GMV crossed $300B for FY. Stock surged post-earnings. → POSITIVE"),
        ("Nov 2025 — Q3 2025 Earnings", "Revenue $2.16B (+24.4%); FCF margin 19%. Shopify Magic AI adoption growing. → POSITIVE"),
        ("Aug 2025 — De Minimis Suspended", "US suspends $800 de minimis rule. Hurts cross-border Shopify merchants (Temu/Shein competitors). → NEGATIVE SHORT-TERM"),
        ("Apr 2025 — Tariff Escalation", "Trump tariffs on China; broader trade war fears. Shopify merchants' supply chains disrupted. Stock sold off ~35% YTD 2026. → NEGATIVE"),
        ("2025 — Shopify Magic AI Expansion", "AI tools in checkout, copywriting, customer service, analytics. Differentiating platform. → POSITIVE LONG-TERM"),
        ("2024 — Logistics Exit", "Exited Deliverr/Flexport. Freed up cash flow, improved margins. Disciplined capital retreat. → POSITIVE"),
        ("2024 — Shopify Payments Global Expansion", "Launched in 10 new markets. Payments now >50% of Merchant Solutions revenue. → POSITIVE"),
    ]
    for i, (headline, detail) in enumerate(news):
        bg = ALT_ROW if i % 2 == 0 else WHITE
        wc(ws, r, 2, headline, bold=True, bg=bg, border=True, wrap=True)
        ws.row_dimensions[r].height = 45
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
        wc(ws, r, 3, detail, bg=bg, border=True, wrap=True)
        r += 1

    r += 1
    section_header(ws, r, 2, "COMPETITIVE DYNAMICS", span=3)
    r += 1
    competitive = [
        ("vs. Amazon", "Amazon's Buy with Prime is integrating into Shopify stores — both competitive and complementary. Shopify has the merchant relationship; Amazon has the consumer trust. Careful co-existence."),
        ("vs. WooCommerce", "WooCommerce holds 33% platform market share (incl. WordPress sites) but lower GMV quality. Shopify dominates premium/high-GMV merchants."),
        ("vs. BigCommerce", "BigCommerce (BIGC) is declining market share; multiple restructurings. Shopify gaining enterprise accounts."),
        ("vs. Salesforce Commerce", "Salesforce targets large enterprise; Shopify Plus is encroaching upmarket. Shopify's simpler tech stack is winning."),
        ("vs. TikTok Shop", "Social commerce growing rapidly. Shopify has TikTok integration but risk of merchant channel shift away from owned websites."),
    ]
    for i, (comp, detail) in enumerate(competitive):
        bg = ALT_ROW if i % 2 == 0 else WHITE
        wc(ws, r, 2, comp, bold=True, bg=bg, border=True)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
        wc(ws, r, 3, detail, bg=bg, border=True, wrap=True)
        ws.row_dimensions[r].height = 45
        r += 1


# ══════════════════════════════════════════════════════════════════════════════
# TAB 12 – KEY INDICATORS
# ══════════════════════════════════════════════════════════════════════════════
def build_key_indicators(wb):
    ws = wb.create_sheet("12. Key Indicators")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, {1: 3, 2: 35, 3: 20, 4: 20, 5: 20, 6: 20, 7: 20})
    freeze(ws, "C3")

    r = 1
    ws.merge_cells(f"B{r}:G{r}")
    wc(ws, r, 2, "SHOPIFY — KEY PERFORMANCE INDICATORS", bold=True, fg=HEADER_FG,
       bg=SHOP_GREEN, size=FONT_SIZE+2, align="center")
    r += 2

    header_row(ws, r, [2, 3, 4, 5, 6, 7], ["KPI", "FY2022", "FY2023", "FY2024", "FY2025", "Trend"])
    r += 1

    kpis = [
        ("OPERATIONAL KPIS", "", "", "", "", ""),
        ("Gross Merchandise Volume (GMV)", "$196B", "$235B", "$292B", "$378B", "↑ +29%"),
        ("GMV YoY Growth", "+12%", "+20%", "+24%", "+29.5%", "↑ Accelerating"),
        ("Active Merchants", "~2.1M", "~2.5M", "~3.2M", "~5M+", "↑ Growing"),
        ("Revenue per Merchant", "~$2,800", "~$2,800", "~$2,775", "~$2,300", "↓ Mix shift"),
        ("Merchant Solutions Take Rate", "~2.22%", "~2.22%", "~2.24%", "~2.23%", "→ Stable"),
        ("", "", "", "", "", ""),
        ("FINANCIAL KPIS", "", "", "", "", ""),
        ("Total Revenue",                "$5.84B", "$7.06B", "$8.88B", "$11.56B", "↑ +30%"),
        ("Revenue YoY Growth",           "+21%",   "+21%",   "+26%",   "+30%",    "↑ Reaccelerating"),
        ("Gross Profit",                 "$2.76B", "$3.46B", "$4.52B", "$5.93B", "↑ Strong"),
        ("Gross Margin %",               "47.2%",  "49.0%",  "50.9%",  "51.3%",  "↑ Expanding"),
        ("Operating Income (GAAP)",      "$90M",   "$904M",  "$1.08B", "$1.47B", "↑ +36%"),
        ("Operating Margin %",           "1.5%",   "12.8%",  "12.2%",  "12.7%",  "↑ Expanding"),
        ("Net Income (GAAP)",            "$58M",   "$974M",  "$1.20B", "$1.63B", "↑ Strong"),
        ("Free Cash Flow",               "$240M",  "$1.10B", "$1.42B", "$2.00B", "↑ +47%"),
        ("FCF Margin %",                 "4.1%",   "15.6%",  "16.0%",  "17.3%",  "↑ Expanding"),
        ("", "", "", "", "", ""),
        ("SaaS / PLATFORM KPIS", "", "", "", "", ""),
        ("Monthly Recurring Revenue (MRR)","$103M", "$128M",  "$165M",  "$214M",  "↑ +30%"),
        ("MRR from Shopify Plus",        "~35%",   "~37%",   "~39%",   "~41%",   "↑ Mix improving"),
        ("Shopify Capital Deployed",     "$2.8B",  "$4.0B",  "$5.0B+", "$6.0B+", "↑ Compounding"),
        ("Shop Pay Penetration (% GMV)", "~40%",   "~45%",   "~50%+",  "~55%+",  "↑ Growing"),
        ("", "", "", "", "", ""),
        ("RULE OF 50 ANALYSIS", "", "", "", "", ""),
        ("Revenue Growth Rate",          "21%",    "26%",    "26%",    "30%",    "↑ Accelerating"),
        ("FCF Margin",                   "4%",     "13%",    "15%",    "17%",    "↑ Expanding"),
        ("Rule of 50 Score",             "25",     "39",     "41",     "47",     "↑ Near 50"),
        ("Status",                       "FAIL",   "FAIL",   "FAIL",   "NEAR 50","↑ Improving"),
        ("Note", "Shopify approaches Rule of 50 threshold — exceptional for a 30% growth company",
         "", "", "", ""),
        ("", "", "", "", "", ""),
        ("STOCK METRICS (Apr 2026)", "", "", "", "", ""),
        ("Stock Price",                  "—",      "—",      "—",      "$111.20", "↓ -30% YTD"),
        ("P/S Ratio (NTM)",             "—",      "—",      "—",      "9.4x",    "↓ Compressed"),
        ("P/E Ratio (NTM)",             "—",      "—",      "—",      "~70x",    "↓ vs. 120x TTM"),
        ("52-Week High",                 "—",      "—",      "—",      "$182.19", "—"),
        ("52-Week Low",                  "—",      "—",      "—",      "$78.00",  "—"),
        ("Market Cap",                   "—",      "—",      "—",      "~$143B",  "—"),
        ("Analyst Median Target",        "—",      "—",      "—",      "$181",    "63% upside"),
    ]

    sections = {"OPERATIONAL KPIS", "FINANCIAL KPIS", "SaaS / PLATFORM KPIS",
                "RULE OF 50 ANALYSIS", "STOCK METRICS (Apr 2026)"}
    highlight_rows = {"Total Revenue", "Free Cash Flow", "FCF Margin %", "Rule of 50 Score",
                      "Gross Merchandise Volume (GMV)", "Gross Margin %"}

    for row_vals in kpis:
        label = row_vals[0]
        if label == "":
            r += 1
            continue
        if label in sections:
            section_header(ws, r, 2, label, span=6, bg=SUBHDR_BG, fg=SHOP_DARK)
            r += 1
            continue
        bold = label in highlight_rows
        bg = LIGHT_GREEN if bold else (ALT_ROW if r % 2 == 0 else WHITE)
        trend = row_vals[5]
        t_color = GREEN if "↑" in trend else (RED if "↓" in trend else DARK_GRAY)
        for col, val in zip([2, 3, 4, 5, 6], row_vals[:5]):
            wc(ws, r, col, val, bold=bold, bg=bg, border=True,
               align="right" if col > 2 else "left")
        wc(ws, r, 7, trend, bold=bold, bg=bg, border=True, fg=t_color, align="center")
        r += 1


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════
def main():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    build_cover(wb)
    build_business_overview(wb)
    build_moat(wb)
    build_income_statement(wb)
    build_balance_sheet(wb)
    build_cash_flow(wb)
    build_return_on_capital(wb)
    build_management(wb)
    build_risks(wb)
    build_valuation(wb)
    build_market_sentiment(wb)
    build_key_indicators(wb)

    out_dir = os.path.join(os.path.dirname(__file__), "..", "output")
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, "SHOP_Financial_Analysis.xlsx")
    wb.save(out_path)
    print(f"✅ Shopify analysis saved to: {out_path}")


if __name__ == "__main__":
    main()
