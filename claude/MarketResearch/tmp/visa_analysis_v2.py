"""
Visa Inc. (NYSE: V) — Comprehensive Financial Analysis
Generated: April 15, 2026
Data Sources: StockAnalysis, MacroTrends, SEC EDGAR, Visa IR,
              MarketBeat, GuruFocus, AlphaSpread, Investing.com
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.utils import get_column_letter
import os

# ─────────────────────────────────────────────────────────────
# CONSTANTS & THEME
# ─────────────────────────────────────────────────────────────
FONT_SIZE = 14
FONT_NAME = "Calibri"

VISA_DARK_BLUE = "1A1F71"
VISA_GOLD      = "F7B600"
VISA_MID_BLUE  = "2962B8"
ACCENT_TEAL    = "17A589"
ACCENT_RED     = "C0392B"
ACCENT_GREEN   = "1E8449"
LIGHT_BLUE_BG  = "D6E4F0"
LIGHT_GOLD_BG  = "FEF9E7"
LIGHT_GREEN_BG = "D5F5E3"
LIGHT_RED_BG   = "FADBD8"
VERY_LIGHT_BG  = "F2F3F4"
WHITE          = "FFFFFF"
DARK_TEXT      = "1C2833"
HEADER_TEXT    = "FFFFFF"
GREY_BG        = "E8E8E8"

OUTPUT_DIR  = "/Users/naji/WORK/github.com/AI/claude/Agent/MarketResearch/output"
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "Visa_Financial_Analysis.xlsx")

# ─────────────────────────────────────────────────────────────
# STYLE HELPERS
# ─────────────────────────────────────────────────────────────
def mfont(size=FONT_SIZE, bold=False, color=DARK_TEXT, italic=False):
    return Font(name=FONT_NAME, size=size, bold=bold, color=color, italic=italic)

def mfill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def mborder(style="thin"):
    s = Side(border_style=style, color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def thick_border():
    t = Side(border_style="medium", color="7F8C8D")
    return Border(left=t, right=t, top=t, bottom=t)

def sc(ws, row, col, value, bold=False, fill_hex=None, size=FONT_SIZE,
       align_h="left", align_v="center", num_fmt=None, italic=False,
       fcolor=DARK_TEXT, border=True, wrap=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = mfont(size, bold, fcolor, italic)
    c.alignment = Alignment(horizontal=align_h, vertical=align_v, wrap_text=wrap)
    if fill_hex:
        c.fill = mfill(fill_hex)
    if num_fmt:
        c.number_format = num_fmt
    if border:
        c.border = mborder()
    return c

def header_row(ws, row, labels, col_start=1, fill=VISA_DARK_BLUE, fc=HEADER_TEXT, h=28):
    for i, lbl in enumerate(labels):
        sc(ws, row, col_start + i, lbl, bold=True, fill_hex=fill,
           fcolor=fc, align_h="center")
    ws.row_dimensions[row].height = h

def sec_hdr(ws, row, label, span=10, fill=VISA_MID_BLUE, fc=HEADER_TEXT):
    c = ws.cell(row=row, column=1, value=label)
    c.font = mfont(FONT_SIZE + 1, bold=True, color=fc)
    c.fill = mfill(fill)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = mborder()
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    ws.row_dimensions[row].height = 26

def sub_hdr(ws, row, label, span=10, fill=LIGHT_BLUE_BG, fc=DARK_TEXT):
    c = ws.cell(row=row, column=1, value=label)
    c.font = mfont(FONT_SIZE, bold=True, color=fc)
    c.fill = mfill(fill)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = mborder()
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    ws.row_dimensions[row].height = 22

def cw(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

def rh(ws, row, h):
    ws.row_dimensions[row].height = h

def blank(ws, row, h=6):
    ws.row_dimensions[row].height = h

def merge_val(ws, row, col1, col2, value, bold=False, fill_hex=None,
              align_h="left", size=FONT_SIZE, italic=False, fcolor=DARK_TEXT, wrap=False):
    ws.merge_cells(start_row=row, start_column=col1, end_row=row, end_column=col2)
    c = ws.cell(row=row, column=col1, value=value)
    c.font = mfont(size, bold, fcolor, italic)
    c.alignment = Alignment(horizontal=align_h, vertical="center", wrap_text=wrap)
    if fill_hex:
        c.fill = mfill(fill_hex)
    c.border = mborder()
    return c

def alt_row(row_num):
    return LIGHT_BLUE_BG if row_num % 2 == 0 else WHITE

# ─────────────────────────────────────────────────────────────
# WORKBOOK SETUP
# ─────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
wb.remove(wb.active)

TABS = [
    "Cover", "Business Overview", "Moat",
    "Income Statement", "Balance Sheet", "Cash Flow Analysis",
    "Return on Capital", "Management", "Risks",
    "Valuation", "Market Sentiment", "Key Indicators",
]

sheets = {}
for name in TABS:
    ws = wb.create_sheet(title=name)
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = VISA_DARK_BLUE
    sheets[name] = ws

sheets["Cover"].sheet_properties.tabColor = VISA_GOLD

# ═══════════════════════════════════════════════════════════════
# 1. COVER
# ═══════════════════════════════════════════════════════════════
ws = sheets["Cover"]
cw(ws, {1:3, 2:40, 3:25, 4:25, 5:3})

for r in range(1, 9):
    for c in range(1, 6):
        ws.cell(r, c).fill = mfill(VISA_DARK_BLUE)
    ws.row_dimensions[r].height = 20

ws.merge_cells("B2:D7")
c = ws.cell(2, 2, "VISA INC.  (NYSE: V)")
c.font = Font(name=FONT_NAME, size=38, bold=True, color=VISA_GOLD)
c.alignment = Alignment(horizontal="center", vertical="center")
c.fill = mfill(VISA_DARK_BLUE)

ws.merge_cells("B8:D8")
c = ws.cell(8, 2, "Comprehensive Financial Analysis  ·  April 15, 2026")
c.font = Font(name=FONT_NAME, size=FONT_SIZE, color="CCCCCC", italic=True)
c.alignment = Alignment(horizontal="center", vertical="center")
c.fill = mfill(VISA_DARK_BLUE)

blank(ws, 9, 8)

# Key stats
kv = [
    ("Stock Price (Apr 15, 2026)",  "$312.68"),
    ("Market Capitalisation",       "$595.8 Billion"),
    ("52-Week Range",               "$293.89 – $375.51"),
    ("FY2025 Net Revenue",          "$40.0 Billion"),
    ("FY2025 Net Income",           "$20.1 Billion"),
    ("Free Cash Flow (FY2025)",     "$21.6 Billion"),
    ("ROIC (FY2025 TTM)",           "~28%"),
    ("FCF Margin",                  "~54%"),
    ("Forward P/E",                 "23.5×"),
    ("Dividend Yield",              "0.86%  ($2.68/share)"),
    ("Next Earnings",               "April 28, 2026  (Q2 FY2026)"),
    ("Analyst Consensus",           "STRONG BUY (21 analysts)"),
    ("Avg. Price Target",           "~$397  (+27% upside)"),
]

row = 10
sec_hdr(ws, row, "  KEY SNAPSHOT", 4)
row += 1
for label, val in kv:
    f = LIGHT_BLUE_BG if row % 2 == 0 else WHITE
    sc(ws, row, 2, label, bold=True, fill_hex=f, align_h="left")
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
    sc(ws, row, 3, val, fill_hex=f, align_h="right")
    rh(ws, row, 22)
    row += 1

blank(ws, row, 8); row += 1

# Investment thesis
sec_hdr(ws, row, "  INVESTMENT THESIS", 4)
row += 1
thesis_points = [
    "Visa is the world's largest payment network — a digital tollbooth on global commerce.",
    "Four-sided network effects: 4.5B+ cards · 130M+ merchants · 200+ countries · 160+ currencies.",
    "Asset-light model: ~81% gross margin, ~60% operating margin, ~50% net margin — elite economics.",
    "Capital return machine: $30B buyback authorisation (2025); $22.9B returned to shareholders in FY2025.",
    "Value-Added Services growing 26% YoY — diversifying beyond core network revenue.",
    "Stablecoin/AI integration positions Visa as payments infrastructure for next-gen commerce.",
    "Current price (~$313) trades at ~21% discount to median analyst target of $397.",
    "Q1 FY2026 beat: Revenue $10.9B (+15%), EPS $3.17 vs $3.14 est.; Q2 FY2026 earnings: Apr 28.",
]
for pt in thesis_points:
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    c = ws.cell(row, 2, f"  •  {pt}")
    c.font = mfont(FONT_SIZE)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c.fill = mfill(LIGHT_GOLD_BG)
    c.border = mborder()
    rh(ws, row, 24)
    row += 1

blank(ws, row, 8); row += 1

# TOC
sec_hdr(ws, row, "  TABLE OF CONTENTS", 4)
row += 1
toc = [
    ("1.  Business Overview",   "Products, Revenue Mix, Value Props, Clients, Seasonality"),
    ("2.  Moat",                "Network Effects, Scale, Switching Costs, Brand"),
    ("3.  Income Statement",    "Revenue, Margins, EPS — 5-Year History"),
    ("4.  Balance Sheet",       "Assets, Liabilities, Equity — 5-Year History"),
    ("5.  Cash Flow Analysis",  "OCF, FCF, CapEx, Capital Returns — 5-Year History"),
    ("6.  Return on Capital",   "ROIC, ROE, ROA, Incremental Returns"),
    ("7.  Management",          "CEO Profile, Incentives, Capital Allocation, Insider Activity"),
    ("8.  Risks",               "Regulatory, Competitive, Macro, Technology"),
    ("9.  Valuation",           "DCF, Multiples, Bull / Base / Bear Scenarios"),
    ("10. Market Sentiment",    "Analyst Ratings, Price Targets, Institutional Ownership"),
    ("11. Key Indicators",      "Operational KPIs, Volumes, Margins Dashboard"),
]
for tab_name, desc in toc:
    f = VERY_LIGHT_BG if row % 2 == 0 else WHITE
    sc(ws, row, 2, tab_name, bold=True, fill_hex=f)
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
    sc(ws, row, 3, desc, fill_hex=f, italic=True)
    rh(ws, row, 22)
    row += 1

blank(ws, row, 6); row += 1
ws.merge_cells(f"B{row}:D{row}")
c = ws.cell(row, 2, "Sources: StockAnalysis · MacroTrends · SEC EDGAR · Visa IR · MarketBeat · GuruFocus · AlphaSpread")
c.font = Font(name=FONT_NAME, size=FONT_SIZE - 1, italic=True, color="888888")
c.alignment = Alignment(horizontal="center")

# ═══════════════════════════════════════════════════════════════
# 2. BUSINESS OVERVIEW
# ═══════════════════════════════════════════════════════════════
ws = sheets["Business Overview"]
cw(ws, {1:3, 2:28, 3:20, 4:20, 5:20, 6:20, 7:3})

row = 1
sec_hdr(ws, row, "  BUSINESS OVERVIEW  —  Visa Inc. (NYSE: V)", 6); row += 1

# Company description
sub_hdr(ws, row, "  What Visa Does", 6); row += 1
desc = (
    "Visa Inc. operates the world's largest retail electronic payments network. "
    "Founded in 1958 and publicly listed in 2008, Visa does NOT issue credit/debit cards or extend credit. "
    "Instead, it provides the technology, brand, and rules that connect 15,000+ financial institution issuers, "
    "130M+ merchants, and 4.5B+ cardholders across 200+ countries. Visa earns fees on every transaction "
    "processed over its VisaNet network — a true 'tollbooth' business model on global commerce."
)
ws.merge_cells(f"B{row}:F{row}")
c = ws.cell(row, 2, desc)
c.font = mfont(FONT_SIZE)
c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
c.fill = mfill(LIGHT_GOLD_BG)
c.border = mborder()
rh(ws, row, 70); row += 1
blank(ws, row); row += 1

# Products / Services
sub_hdr(ws, row, "  Products & Services", 6); row += 1
products = [
    ("CORE NETWORK", "VisaNet — processes $15.7T in payment volume annually; authorization, clearing, settlement"),
    ("VISA CREDIT",  "Premium credit products (Visa Infinite, Signature, Platinum) issued by bank partners"),
    ("VISA DEBIT",   "Debit/prepaid products (Visa Classic, Electron) — growing in emerging markets"),
    ("VISA DIRECT",  "Real-time push payments to cards/accounts; P2P, payroll, gig economy, B2B"),
    ("VISA B2B CONNECT", "Multilateral B2B cross-border payments network — bypasses correspondent banking"),
    ("VALUE-ADDED SERVICES", "Fraud prevention (CyberSource), data analytics, consulting, issuer processing, tokenisation"),
    ("VISA TOKENISATION", "EMV token service: converts card numbers into secure tokens; 11B+ tokens issued"),
    ("STABLECOIN / CRYPTO", "Crypto-linked card programs; stablecoin settlement pilots with USDC on Solana / Ethereum"),
]
header_row(ws, row, ["Product / Service", "Description", "", "", "", ""], col_start=2, h=24)
row += 1
for i, (prod, desc2) in enumerate(products):
    f = alt_row(i)
    sc(ws, row, 2, prod, bold=True, fill_hex=f)
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)
    c = ws.cell(row, 3, desc2)
    c.font = mfont(FONT_SIZE)
    c.fill = mfill(f)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c.border = mborder()
    rh(ws, row, 30)
    row += 1

blank(ws, row); row += 1

# Revenue breakdown by segment
sub_hdr(ws, row, "  Revenue Breakdown by Segment  (FY2025, Net Revenue $40.0B)", 6); row += 1
header_row(ws, row, ["Revenue Stream", "FY2025 Gross ($M)", "FY2025 Net ($M)", "% of Net Rev", "YoY Growth", "Notes"], col_start=2, h=24)
row += 1
seg_data = [
    ("Service Revenues",              15_900, 15_900, "39.8%", "+10%",  "Fees based on prior-qtr payment volumes; most stable segment"),
    ("Data Processing Revenues",      14_400, 14_400, "36.0%", "+13%",  "Authorization, clearing, settlement per-transaction fees"),
    ("International Transaction Rev.", 14_200, 14_200, "35.5%", "+14%",  "Currency conversion + cross-border fees; high-margin"),
    ("Other Revenues (VAS)",           4_000,  4_000, "10.0%", "+26%",  "CyberSource, data analytics, consulting, licensing"),
    ("Client Incentives (deducted)",  (8_500), (8_500),"-21.3%", "+12%", "Contractual incentives to issuers & acquirers"),
    ("NET REVENUE",                   40_000, 40_000, "100%",  "+11%",  "FY2025 total net revenue"),
]
for i, row_data in enumerate(seg_data):
    f = LIGHT_GOLD_BG if row_data[0] == "NET REVENUE" else alt_row(i)
    bold = row_data[0] == "NET REVENUE"
    sc(ws, row, 2, row_data[0], bold=bold, fill_hex=f)
    sc(ws, row, 3, row_data[1], bold=bold, fill_hex=f, align_h="right",
       num_fmt='#,##0' if row_data[1] > 0 else '#,##0;(#,##0)')
    sc(ws, row, 4, row_data[2], bold=bold, fill_hex=f, align_h="right",
       num_fmt='#,##0' if row_data[2] > 0 else '#,##0;(#,##0)')
    sc(ws, row, 5, row_data[3], bold=bold, fill_hex=f, align_h="center")
    sc(ws, row, 6, row_data[4], bold=bold, fill_hex=f, align_h="center")
    sc(ws, row, 7, row_data[5] if len(row_data) > 5 else "", fill_hex=f, wrap=True)
    rh(ws, row, 28)
    row += 1

blank(ws, row); row += 1

# Geography
sub_hdr(ws, row, "  Revenue by Geography  (FY2025)", 6); row += 1
header_row(ws, row, ["Region", "FY2025 Revenue ($M)", "% of Total", "YoY Growth", "Key Markets", ""], col_start=2, h=24)
row += 1
geo_data = [
    ("United States",    15_630, "39.1%", "+7%",  "Largest market; mature but still growing debit/VAS"),
    ("Non-U.S. / International", 24_370, "60.9%", "+15%", "Europe, Asia-Pacific, CEMEA, LAC — faster growing"),
    ("  Europe",          7_500, "18.8%", "+12%", "UK, Germany, France — regulatory risk from interchange caps"),
    ("  Asia-Pacific",    7_200, "18.0%", "+18%", "China (UnionPay competition), India, SE Asia — high growth"),
    ("  Canada",          2_800, "7.0%",  "+8%",  "Mature, duopoly with MC"),
    ("  CEMEA",           3_870, "9.7%",  "+19%", "High-growth; mobile-first consumers"),
    ("  Latin America",   3_000, "7.5%",  "+14%", "Brazil (Pix competition), Mexico, others"),
    ("TOTAL",            40_000, "100%",  "+11%", "FY2025 net revenue"),
]
for i, gd in enumerate(geo_data):
    f = LIGHT_GOLD_BG if gd[0] == "TOTAL" else alt_row(i)
    bold = gd[0] == "TOTAL"
    sc(ws, row, 2, gd[0], bold=bold, fill_hex=f)
    sc(ws, row, 3, gd[1], bold=bold, fill_hex=f, align_h="right", num_fmt='#,##0')
    sc(ws, row, 4, gd[2], bold=bold, fill_hex=f, align_h="center")
    sc(ws, row, 5, gd[3], bold=bold, fill_hex=f, align_h="center")
    ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=7)
    sc(ws, row, 6, gd[4], fill_hex=f, wrap=True)
    rh(ws, row, 24)
    row += 1

blank(ws, row); row += 1

# Value proposition & clients
sub_hdr(ws, row, "  Value Proposition & Key Clients", 6); row += 1
vp = [
    ("For Cardholders",    "Accepted everywhere, fraud protection, global usability, rewards programs"),
    ("For Merchants",      "Guaranteed payment, lower fraud loss vs. cash, global acceptance infrastructure"),
    ("For Issuers",        "Revenue share on interchange, fraud tools, data insights, co-brand programs"),
    ("For Acquirers",      "Access to Visa network, processing infrastructure, dispute resolution"),
    ("Key Issuer Clients", "JPMorgan Chase, Bank of America, Citibank, Wells Fargo, HDFC, Barclays, BBVA"),
    ("Key Merchant Clients","Amazon, Walmart, Apple (Apple Pay), Alibaba, McDonald's — global merchant base"),
]
for i, (k, v) in enumerate(vp):
    f = alt_row(i)
    sc(ws, row, 2, k, bold=True, fill_hex=f)
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=7)
    sc(ws, row, 3, v, fill_hex=f, wrap=True)
    rh(ws, row, 26)
    row += 1

blank(ws, row); row += 1

# Seasonality & Margin
sub_hdr(ws, row, "  Seasonality & Margin Structure", 6); row += 1
seas = [
    ("Q1 (Oct–Dec)",  "Strongest quarter — holiday spending, Black Friday, Christmas; ~28% of annual revenue"),
    ("Q2 (Jan–Mar)",  "Moderate; post-holiday dip but VAS and B2B contribute"),
    ("Q3 (Apr–Jun)",  "Picks up; travel season begins, cross-border fees ramp"),
    ("Q4 (Jul–Sep)",  "Summer travel peak; cross-border strongest; fiscal year-end"),
    ("Gross Margin",  "~81% — Visa has minimal COGS; cost base is primarily network technology & personnel"),
    ("Op. Margin",    "~60% — one of the highest operating margins in global finance / technology"),
    ("Net Margin",    "~50% — after interest on $21B debt and low effective tax rate (~15.9%)"),
    ("FCF Margin",    "~54% — capex-light; CapEx ~$1.5B vs. $23B operating cash flow"),
]
for i, (k, v) in enumerate(seas):
    f = alt_row(i)
    sc(ws, row, 2, k, bold=True, fill_hex=f)
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=7)
    sc(ws, row, 3, v, fill_hex=f, wrap=True)
    rh(ws, row, 26)
    row += 1

# ═══════════════════════════════════════════════════════════════
# 3. MOAT
# ═══════════════════════════════════════════════════════════════
ws = sheets["Moat"]
cw(ws, {1:3, 2:30, 3:60, 4:3})

row = 1
sec_hdr(ws, row, "  ECONOMIC MOAT ANALYSIS  —  Visa Inc.", 3); row += 1

moat_sections = [
    ("1. NETWORK EFFECTS  (★★★★★)", [
        ("Global Two-Sided Network",
         "Visa connects ~15,000 financial institutions (issuers), 130M+ merchants, and 4.5B+ cardholders "
         "in 200+ countries. Each new participant makes the network MORE valuable for all others. This "
         "self-reinforcing flywheel has taken 60+ years to build and is virtually impossible to replicate."),
        ("Four-Sided Platform",
         "Cardholders → Merchants → Issuers → Acquirers. Visa sits at the intersection of all four, "
         "charging a small fee on each transaction. The more each group uses Visa, the more valuable "
         "it becomes for everyone else — a 'winner-takes-most' dynamic."),
        ("Scale Moat in Fraud",
         "Processing >264 billion transactions/year generates unparalleled fraud data. Visa's AI-powered "
         "fraud engine (sub-0.1% fraud rate) is impossible to replicate at scale — better data → fewer "
         "false declines → happier merchants → more Visa usage."),
    ]),
    ("2. SWITCHING COSTS  (★★★★☆)", [
        ("Issuer Switching Costs",
         "Banks sign long-term contracts (5–10 years) to issue Visa-branded cards. Switching to a competitor "
         "requires re-issuing millions of cards, renegotiating merchant contracts, retaining staff, and "
         "risking customer churn. Cost of switching is astronomical."),
        ("Merchant Acceptance",
         "Merchants accept Visa because consumers carry Visa cards. They accept all networks to avoid "
         "losing sales. This means Visa benefits from acceptance ubiquity regardless of any individual "
         "merchant's preference."),
        ("Consumer Card Loyalty",
         "Co-branded cards (Chase Sapphire, United MileagePlus, etc.) create loyalty through rewards "
         "programs. Consumers rarely change their primary card — sticky, recurring transaction volumes."),
    ]),
    ("3. BRAND & TRUST  (★★★★☆)", [
        ("60+ Year Brand Heritage",
         "The Visa brand represents security, reliability, and global acceptance. This is especially "
         "powerful in cross-border transactions where local payment rails are unavailable or untrusted."),
        ("Regulatory Relationships",
         "Visa has spent decades building relationships with central banks, regulators, and governments. "
         "New entrants face enormous compliance hurdles. Visa's existing licenses and regulatory approvals "
         "represent significant barriers to entry."),
    ]),
    ("4. SCALE ECONOMICS  (★★★★★)", [
        ("Incremental Cost Near Zero",
         "Processing the 264 billionth transaction costs Visa nearly nothing incremental. Fixed cost "
         "infrastructure (VisaNet data centers) supports exponential volume growth with minimal marginal cost. "
         "This is why operating margins have expanded from ~50% in 2010 to ~60% today."),
        ("R&D Investment Advantage",
         "Visa spends ~$2.5B/year on technology. This maintains its edge in security, processing speed "
         "(millisecond authorization), uptime (99.999%), and innovation (tokenisation, AI fraud, Visa Direct)."),
    ]),
    ("5. MOAT THREATS (Monitoring Required)", [
        ("CCCA Regulatory Risk",
         "The Credit Card Competition Act (reintroduced Jan 2026) would require large banks to enable "
         "alternative networks on credit cards. If passed, could reduce Visa's volumes by 10–20% over time. "
         "Probability: ~20–30% over 3 years. Impact: Material but manageable given VAS diversification."),
        ("Central Bank Real-Time Rails",
         "UPI (India), Pix (Brazil), FedNow (USA) are government-backed instant payment systems that "
         "bypass card networks for domestic transactions. Long-term, these could reduce Visa's addressable "
         "market in specific geographies, though Visa has partnered with several (e.g., Visa on UPI)."),
        ("Crypto / Stablecoin Disruption",
         "Stablecoins and CBDCs could theoretically enable peer-to-peer value transfer without a network "
         "intermediary. Visa is actively positioning as infrastructure (USDC settlement, crypto card programs) "
         "rather than being disrupted — a wise defensive move."),
        ("Big Tech Wallets",
         "Apple Pay, Google Pay route through Visa rails — net positive today. Risk is if Apple builds its "
         "own payment rails (unlikely given regulatory scrutiny). Currently, Big Tech AMPLIFIES Visa's reach."),
    ]),
    ("6. MOAT VERDICT", [
        ("Overall Moat Rating",
         "WIDE MOAT — Visa possesses one of the strongest economic moats in global business. The combination "
         "of network effects, switching costs, brand, and scale creates a near-impenetrable competitive position. "
         "ROIC of ~28% vs. WACC of ~8% = economic profit spread of ~20 percentage points. This moat has "
         "been durable for 60+ years and is self-reinforcing. The primary risk is regulatory intervention, "
         "not competitive disruption from the private sector."),
    ]),
]

for section_title, items in moat_sections:
    sec_hdr(ws, row, f"  {section_title}", 3,
            fill=VISA_MID_BLUE if "THREAT" not in section_title else ACCENT_RED,
            fc=HEADER_TEXT)
    row += 1
    for title, body in items:
        sc(ws, row, 2, title, bold=True, fill_hex=LIGHT_BLUE_BG)
        rh(ws, row, 22)
        row += 1
        ws.merge_cells(f"B{row}:C{row}")
        c = ws.cell(row, 2, body)
        c.font = mfont(FONT_SIZE)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c.fill = mfill(VERY_LIGHT_BG)
        c.border = mborder()
        rh(ws, row, 72)
        row += 1
    blank(ws, row); row += 1

# ═══════════════════════════════════════════════════════════════
# 4. INCOME STATEMENT
# ═══════════════════════════════════════════════════════════════
ws = sheets["Income Statement"]
cw(ws, {1:3, 2:36, 3:18, 4:18, 5:18, 6:18, 7:18, 8:18, 9:3})

row = 1
sec_hdr(ws, row, "  INCOME STATEMENT  —  Visa Inc.  (FY2021 – FY2025, $M)", 8); row += 1

# Annual income statement
years = ["FY2021", "FY2022", "FY2023", "FY2024", "FY2025"]
header_row(ws, row, [""] + years + ["TTM"], col_start=2, h=28); row += 1

is_data = [
    # (label, [FY21, FY22, FY23, FY24, FY25, TTM], bold, fill)
    ("NET REVENUE ($M)",           [24105, 29310, 32653, 35926, 40000, 41391], True, LIGHT_BLUE_BG),
    ("  YoY Growth",               ["—", "21.6%", "11.4%", "10.0%", "11.3%", "12.5%"], False, WHITE),
    ("Cost of Revenue ($M)",       [4580, 5568, 6204, 6826, 7568, 7864], False, WHITE),
    ("GROSS PROFIT ($M)",          [19525, 23742, 26449, 29100, 32432, 33527], True, LIGHT_BLUE_BG),
    ("  Gross Margin",             ["81.0%", "81.0%", "81.0%", "81.0%", "81.1%", "81.0%"], False, WHITE),
    ("Operating Expenses ($M)",    [3721, 4929, 5449, 5505, 8438, 9030], False, WHITE),
    ("OPERATING INCOME ($M)",      [15804, 18813, 21000, 23595, 23994, 24497], True, LIGHT_BLUE_BG),
    ("  Operating Margin",         ["65.6%", "64.2%", "64.3%", "65.7%", "60.0%", "59.2%"], False, WHITE),
    ("Interest & Other (net)",     [-626, -706, -811, -817, -883, -895], False, WHITE),
    ("Pre-Tax Income ($M)",        [15178, 18107, 20189, 22778, 23111, 23602], False, WHITE),
    ("Income Tax ($M)",            [2867, 3150, 2916, 3035, 3053, 3090], False, WHITE),
    ("  Effective Tax Rate",       ["18.9%", "17.4%", "14.4%", "13.3%", "13.2%", "13.1%"], False, WHITE),
    ("NET INCOME ($M)",            [12311, 14957, 17273, 19743, 20058, 20792], True, LIGHT_GOLD_BG),
    ("  Net Margin",               ["51.1%", "51.0%", "52.9%", "54.9%", "50.1%", "50.2%"], False, WHITE),
    ("Diluted Shares (M)",         [2185, 2136, 2087, 2029, 1966, 1950], False, WHITE),
    ("EPS (Diluted)",              [5.63, 7.00, 8.28, 9.73, 10.20, 10.66], True, LIGHT_BLUE_BG),
    ("  EPS YoY Growth",           ["—", "24.3%", "18.3%", "17.5%", "4.8%", "7.5%"], False, WHITE),
    ("EBITDA ($M)",                [16840, 20135, 22421, 25125, 25700, 25760], True, LIGHT_BLUE_BG),
    ("  EBITDA Margin",            ["69.8%", "68.7%", "68.7%", "69.9%", "64.3%", "62.2%"], False, WHITE),
]

for lbl, vals, bold, fill in is_data:
    sc(ws, row, 2, lbl, bold=bold, fill_hex=fill)
    for j, v in enumerate(vals):
        col = 3 + j
        if isinstance(v, float) and v > 100:
            sc(ws, row, col, v, bold=bold, fill_hex=fill, align_h="right", num_fmt='#,##0.00')
        elif isinstance(v, int):
            fmt = '#,##0' if v >= 0 else '#,##0;(#,##0)'
            sc(ws, row, col, v, bold=bold, fill_hex=fill, align_h="right", num_fmt=fmt)
        else:
            sc(ws, row, col, v, bold=bold, fill_hex=fill, align_h="center")
    rh(ws, row, 22)
    row += 1

blank(ws, row); row += 1

# Q1 FY2026
sec_hdr(ws, row, "  Q1 FY2026 QUARTERLY RESULTS (Jan 29, 2026)", 8); row += 1
header_row(ws, row, ["Metric", "Q1 FY2025", "Q1 FY2026", "YoY Change", "vs. Estimate", "Beat/Miss", "", ""], col_start=2, h=24); row += 1
q1_data = [
    ("Net Revenue ($M)", "9,488", "10,900", "+14.9%", "$10,680M est.", "BEAT +2.1%"),
    ("Operating Income ($M)", "5,850", "6,700", "+14.5%", "—", "—"),
    ("Net Income ($M)", "5,100", "5,900", "+15.7%", "—", "—"),
    ("EPS (Diluted)", "$2.75", "$3.17", "+15.3%", "$3.14 est.", "BEAT +$0.03"),
    ("Payments Volume ($B)", "$3,700", "$3,990", "+7.8%", "—", "—"),
    ("Processed Transactions (B)", "63.1", "69.0", "+9.3%", "—", "—"),
    ("Cross-Border Volume Growth", "+16%", "+16%", "Flat", "—", "—"),
    ("VAS Revenue Growth", "+22%", "+28%", "+6pp", "—", "Beat"),
]
for i, (m, q1_25, q1_26, chg, est, beat) in enumerate(q1_data):
    f = alt_row(i)
    sc(ws, row, 2, m, bold=True, fill_hex=f)
    sc(ws, row, 3, q1_25, fill_hex=f, align_h="center")
    sc(ws, row, 4, q1_26, fill_hex=f, align_h="center")
    sc(ws, row, 5, chg, fill_hex=LIGHT_GREEN_BG, align_h="center")
    sc(ws, row, 6, est, fill_hex=f, align_h="center")
    fc2 = ACCENT_GREEN if "BEAT" in beat else (ACCENT_RED if "MISS" in beat else DARK_TEXT)
    sc(ws, row, 7, beat, fill_hex=f, align_h="center", fcolor=fc2, bold="BEAT" in beat or "MISS" in beat)
    rh(ws, row, 22)
    row += 1

blank(ws, row); row += 1

# Segment revenue table
sec_hdr(ws, row, "  ANNUAL REVENUE BY SEGMENT ($M)", 8); row += 1
header_row(ws, row, ["Segment"] + years, col_start=2, h=24); row += 1
seg_annual = [
    ("Service Revenues",            [10418, 14024, 14981, 15985, 15900]),
    ("Data Processing Revenues",    [7791,  9715,  11405, 12709, 14400]),
    ("Intl Transaction Revenues",   [8056,  9924,  10886, 11988, 14200]),
    ("Other Revenues",              [1545,  1874,  2128,  2765,  4000]),
    ("Client Incentives",           [-3705, -6227, -6747, -9521, -8500]),
    ("NET REVENUE",                 [24105, 29310, 32653, 35926, 40000]),
]
for i, (seg, vals) in enumerate(seg_annual):
    f = LIGHT_GOLD_BG if seg == "NET REVENUE" else alt_row(i)
    bold = seg == "NET REVENUE"
    sc(ws, row, 2, seg, bold=bold, fill_hex=f)
    for j, v in enumerate(vals):
        fmt = '#,##0' if v >= 0 else '#,##0;(#,##0)'
        sc(ws, row, 3+j, v, bold=bold, fill_hex=f, align_h="right", num_fmt=fmt)
    rh(ws, row, 22)
    row += 1

# ═══════════════════════════════════════════════════════════════
# 5. BALANCE SHEET
# ═══════════════════════════════════════════════════════════════
ws = sheets["Balance Sheet"]
cw(ws, {1:3, 2:36, 3:18, 4:18, 5:18, 6:18, 7:18, 8:3})

row = 1
sec_hdr(ws, row, "  BALANCE SHEET  —  Visa Inc.  (FY2021 – FY2025, $M)", 7); row += 1

header_row(ws, row, [""] + years, col_start=2, h=28); row += 1

bs_data = [
    # ASSETS
    ("─── ASSETS ───", [], True, VISA_MID_BLUE, HEADER_TEXT),
    ("Cash & Cash Equivalents",          [16487, 15689, 16286, 11975, 14756], False, WHITE, DARK_TEXT),
    ("Short-Term Investments",           [2250,  2800,  2500,  1800,  2100], False, WHITE, DARK_TEXT),
    ("Accounts Receivable",              [2080,  2580,  3020,  3180,  3500], False, WHITE, DARK_TEXT),
    ("Other Current Assets",             [1960,  2200,  2680,  2850,  3200], False, WHITE, DARK_TEXT),
    ("TOTAL CURRENT ASSETS",             [22777, 23269, 24486, 19805, 23556], True, LIGHT_BLUE_BG, DARK_TEXT),
    ("Property, Plant & Equip (net)",    [2820,  3210,  3580,  4020,  4680], False, WHITE, DARK_TEXT),
    ("Goodwill",                         [13484, 13484, 17783, 17810, 17820], False, WHITE, DARK_TEXT),
    ("Intangible Assets (net)",          [27520, 25420, 24500, 23200, 21800], False, WHITE, DARK_TEXT),
    ("Deferred Tax & Other Assets",      [16295, 20118, 20150, 29676, 28958], False, WHITE, DARK_TEXT),
    ("TOTAL ASSETS",                     [82896, 85501, 90499, 94511, 96814], True, LIGHT_GOLD_BG, DARK_TEXT),
    ("─── LIABILITIES ───", [], True, VISA_MID_BLUE, HEADER_TEXT),
    ("Accounts Payable & Accruals",      [3890,  4200,  4890,  5190,  5450], False, WHITE, DARK_TEXT),
    ("Current Debt",                     [999,   999,   0,     2000,  0], False, WHITE, DARK_TEXT),
    ("Other Current Liabilities",        [7038,  7500,  8500,  9200,  9800], False, WHITE, DARK_TEXT),
    ("TOTAL CURRENT LIABILITIES",        [11927, 12699, 13390, 16390, 15250], True, LIGHT_BLUE_BG, DARK_TEXT),
    ("Long-Term Debt",                   [19978, 21451, 20463, 18836, 21177], False, WHITE, DARK_TEXT),
    ("Other Long-Term Liabilities",      [13402, 15770, 17913, 20148, 21610], False, WHITE, DARK_TEXT),
    ("TOTAL LIABILITIES",                [45307, 49920, 51766, 55374, 58037], True, LIGHT_RED_BG, DARK_TEXT),
    ("─── SHAREHOLDERS' EQUITY ───", [], True, VISA_MID_BLUE, HEADER_TEXT),
    ("Common Stock (Class A/B/C)",       [3600,  3600,  3600,  3600,  3600], False, WHITE, DARK_TEXT),
    ("Retained Earnings",                [19875, 22600, 25900, 26340, 26130], False, WHITE, DARK_TEXT),
    ("Other Comprehensive Inc. (Loss)",  [14114, 9381,  9233,  9197,  9047], False, WHITE, DARK_TEXT),
    ("TOTAL SHAREHOLDERS' EQUITY",       [37589, 35581, 38733, 39137, 38777], True, LIGHT_GREEN_BG, DARK_TEXT),
    ("─── KEY RATIOS ───", [], True, VISA_MID_BLUE, HEADER_TEXT),
    ("Debt / Equity",                    ["0.56x","0.63x","0.53x","0.53x","0.55x"], False, WHITE, DARK_TEXT),
    ("Current Ratio",                    ["1.91x","1.83x","1.83x","1.21x","1.54x"], False, WHITE, DARK_TEXT),
    ("Net Cash / (Net Debt) ($M)",       [-2465, -3928, -335, -8861, -5937], False, WHITE, DARK_TEXT),
    ("Book Value per Share",             ["$17.22","$16.64","$18.56","$19.29","$20.29"], False, WHITE, DARK_TEXT),
]

for lbl, vals, bold, fill, fc in bs_data:
    if not vals:  # section header
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
        c = ws.cell(row, 2, lbl)
        c.font = mfont(FONT_SIZE, bold=True, color=fc)
        c.fill = mfill(fill)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = mborder()
        rh(ws, row, 22); row += 1
        continue
    sc(ws, row, 2, lbl, bold=bold, fill_hex=fill, fcolor=fc)
    for j, v in enumerate(vals):
        col = 3 + j
        if isinstance(v, int):
            fmt = '#,##0' if v >= 0 else '#,##0;(#,##0)'
            sc(ws, row, col, v, bold=bold, fill_hex=fill, align_h="right", num_fmt=fmt, fcolor=fc)
        else:
            sc(ws, row, col, v, bold=bold, fill_hex=fill, align_h="center", fcolor=fc)
    rh(ws, row, 22); row += 1

# ═══════════════════════════════════════════════════════════════
# 6. CASH FLOW ANALYSIS
# ═══════════════════════════════════════════════════════════════
ws = sheets["Cash Flow Analysis"]
cw(ws, {1:3, 2:36, 3:18, 4:18, 5:18, 6:18, 7:18, 8:3})

row = 1
sec_hdr(ws, row, "  CASH FLOW ANALYSIS  —  Visa Inc.  (FY2021 – FY2025, $M)", 7); row += 1
header_row(ws, row, [""] + years, col_start=2, h=28); row += 1

cf_data = [
    ("─── OPERATING CASH FLOW ───", [], True, VISA_MID_BLUE, HEADER_TEXT),
    ("Net Income",                   [12311, 14957, 17273, 19743, 20058], False, WHITE, DARK_TEXT),
    ("D&A",                          [1036,  1322,  1421,  1530,  1706], False, WHITE, DARK_TEXT),
    ("Stock-Based Compensation",     [535,   542,   578,   620,   680], False, WHITE, DARK_TEXT),
    ("Changes in Working Capital",   [1345,  2028,  1483, -1943,  615], False, WHITE, DARK_TEXT),
    ("OPERATING CASH FLOW",          [15227, 18849, 20755, 19950, 23059], True, LIGHT_GREEN_BG, DARK_TEXT),
    ("  OCF Margin",                 ["63.2%","64.3%","63.6%","55.5%","57.6%"], False, WHITE, DARK_TEXT),
    ("─── INVESTING ACTIVITIES ───", [], True, VISA_MID_BLUE, HEADER_TEXT),
    ("Capital Expenditures",         [-705,  -970, -1059, -1257, -1482], False, WHITE, DARK_TEXT),
    ("Acquisitions & Investments",   [-450,  -890,  -650,  -520,  -780], False, WHITE, DARK_TEXT),
    ("Other Investing",              [280,   -120,   450,   640,   480], False, WHITE, DARK_TEXT),
    ("INVESTING CASH FLOW",          [-875, -1980, -1259, -1137, -1782], True, LIGHT_RED_BG, DARK_TEXT),
    ("─── FREE CASH FLOW ───", [], True, VISA_MID_BLUE, HEADER_TEXT),
    ("FREE CASH FLOW",               [14522, 17879, 19696, 18693, 21577], True, LIGHT_GOLD_BG, DARK_TEXT),
    ("  FCF YoY Growth",             ["—","23.1%","10.2%","-5.1%","15.4%"], False, WHITE, DARK_TEXT),
    ("  FCF Margin",                 ["60.2%","61.0%","60.3%","52.0%","53.9%"], False, WHITE, DARK_TEXT),
    ("  FCF per Share",              ["$6.64","$8.37","$9.44","$9.21","$10.98"], False, WHITE, DARK_TEXT),
    ("─── FINANCING ACTIVITIES ───", [], True, VISA_MID_BLUE, HEADER_TEXT),
    ("Dividends Paid",               [-2798, -3203, -3751, -4217, -4634], False, WHITE, DARK_TEXT),
    ("Share Repurchases",            [-8676,-11589,-12101,-16713,-18316], False, WHITE, DARK_TEXT),
    ("Net Debt Issuance / (Repay.)", [1560,  1980,  -970,  -860,   510], False, WHITE, DARK_TEXT),
    ("FINANCING CASH FLOW",          [-9914,-12812,-16822,-21790,-22440], True, LIGHT_RED_BG, DARK_TEXT),
    ("─── CAPITAL RETURNS ───", [], True, VISA_MID_BLUE, HEADER_TEXT),
    ("Total Capital Returned ($M)",  [11474, 14792, 15852, 20930, 22950], True, LIGHT_GREEN_BG, DARK_TEXT),
    ("  % of FCF Returned",          ["79.0%","82.7%","80.5%","111.9%","106.4%"], False, WHITE, DARK_TEXT),
    ("  Dividends per Share",        ["$1.28","$1.50","$1.80","$2.08","$2.52"], False, WHITE, DARK_TEXT),
    ("  Div. Payout Ratio (GAAP)",   ["22.7%","21.4%","21.7%","21.4%","23.1%"], False, WHITE, DARK_TEXT),
    ("  Shares Repurchased (M)",     [315,   260,   244,   225,   245], False, WHITE, DARK_TEXT),
    ("  Avg. Repurchase Price",      ["$27.5","$44.6","$49.6","$74.3","$74.8"], False, WHITE, DARK_TEXT),
]

for lbl, vals, bold, fill, fc in cf_data:
    if not vals:
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
        c = ws.cell(row, 2, lbl)
        c.font = mfont(FONT_SIZE, bold=True, color=fc)
        c.fill = mfill(fill)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = mborder()
        rh(ws, row, 22); row += 1
        continue
    sc(ws, row, 2, lbl, bold=bold, fill_hex=fill, fcolor=fc)
    for j, v in enumerate(vals):
        col = 3 + j
        if isinstance(v, (int, float)) and not isinstance(v, bool):
            if isinstance(v, float) and v < 1:  # small float = percentage
                sc(ws, row, col, v, bold=bold, fill_hex=fill, align_h="center", fcolor=fc)
            else:
                fmt = '#,##0' if v >= 0 else '#,##0;(#,##0)'
                sc(ws, row, col, v, bold=bold, fill_hex=fill, align_h="right", num_fmt=fmt, fcolor=fc)
        else:
            sc(ws, row, col, v, bold=bold, fill_hex=fill, align_h="center", fcolor=fc)
    rh(ws, row, 22); row += 1

blank(ws, row); row += 1

# Cash flow quality commentary
sec_hdr(ws, row, "  CASH FLOW QUALITY ASSESSMENT", 7); row += 1
comments = [
    ("OCF vs. Net Income Conversion", "OCF consistently exceeds net income — hallmark of high-quality earnings. In FY2025, OCF of $23.1B vs. Net Income of $20.1B = 115% conversion ratio. No concern over accounting-driven inflation of earnings."),
    ("CapEx Intensity", "CapEx is just $1.5B on $40B revenue — a 3.7% intensity rate. This is the defining feature of Visa's asset-light business model. The company processes trillions in payments on infrastructure requiring minimal ongoing capital investment."),
    ("FCF Growth Trajectory", "FCF has grown from $14.5B (FY2021) to $21.6B (FY2025) — a CAGR of ~10.4%. This compares favorably to revenue CAGR of ~13.5%, indicating slight margin compression from increasing client incentives and operating expenses."),
    ("Capital Return Discipline", "Visa returned $22.9B to shareholders in FY2025 against $21.6B FCF — implying mild balance sheet usage to fund buybacks. With $30B buyback authorization remaining, this is a deliberate, value-enhancing strategy given share count reduction of ~12% since FY2021."),
    ("Dividend Sustainability", "Dividend has grown at ~18% CAGR since 2021. Payout ratio ~23% of GAAP earnings — very conservative, with significant room to grow. Visa is a reliable dividend grower; 15+ consecutive years of increases."),
]
for i, (k, v) in enumerate(comments):
    f = alt_row(i)
    sc(ws, row, 2, k, bold=True, fill_hex=f)
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=7)
    c = ws.cell(row, 3, v)
    c.font = mfont(FONT_SIZE)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c.fill = mfill(f)
    c.border = mborder()
    rh(ws, row, 52)
    row += 1

# ═══════════════════════════════════════════════════════════════
# 7. RETURN ON CAPITAL
# ═══════════════════════════════════════════════════════════════
ws = sheets["Return on Capital"]
cw(ws, {1:3, 2:36, 3:18, 4:18, 5:18, 6:18, 7:18, 8:3})

row = 1
sec_hdr(ws, row, "  RETURN ON CAPITAL  —  Visa Inc.  (FY2021 – FY2025)", 7); row += 1
header_row(ws, row, ["Metric"] + years, col_start=2, h=28); row += 1

roc_data = [
    ("─── CORE RETURN METRICS ───", [], True, VISA_MID_BLUE, HEADER_TEXT),
    ("Return on Equity (ROE)",          ["32.8%","40.4%","44.8%","51.0%","51.5%"], False, LIGHT_BLUE_BG, DARK_TEXT),
    ("Return on Assets (ROA)",          ["14.9%","17.6%","19.3%","21.1%","20.9%"], False, WHITE, DARK_TEXT),
    ("Return on Invested Capital (ROIC)",["20.5%","24.2%","26.1%","27.5%","27.9%"], False, LIGHT_GREEN_BG, DARK_TEXT),
    ("WACC (est.)",                     ["7.5%", "7.8%", "8.0%", "8.0%", "8.0%"], False, WHITE, DARK_TEXT),
    ("Economic Profit Spread (ROIC-WACC)",["13.0%","16.4%","18.1%","19.5%","19.9%"], True, LIGHT_GOLD_BG, DARK_TEXT),
    ("─── FCF RETURN METRICS ───", [], True, VISA_MID_BLUE, HEADER_TEXT),
    ("FCF Return on Equity",            ["38.6%","50.3%","50.8%","47.8%","55.6%"], False, WHITE, DARK_TEXT),
    ("FCF Yield (on Market Cap)",        ["2.3%","2.8%","3.1%","2.9%","3.5%"], False, LIGHT_BLUE_BG, DARK_TEXT),
    ("FCF per Share",                    ["$6.64","$8.37","$9.44","$9.21","$10.98"], False, WHITE, DARK_TEXT),
    ("─── INCREMENTAL RETURNS ───", [], True, VISA_MID_BLUE, HEADER_TEXT),
    ("Incremental Revenue ($M)",         ["—", "5205", "3343", "3273", "4074"], False, WHITE, DARK_TEXT),
    ("Incremental CapEx ($M)",           ["—", "265", "89", "198", "225"], False, WHITE, DARK_TEXT),
    ("Return on Incremental Capital",    ["—", "19.6x","37.6x","16.5x","18.1x"], False, LIGHT_GOLD_BG, DARK_TEXT),
    ("─── ASSET EFFICIENCY ───", [], True, VISA_MID_BLUE, HEADER_TEXT),
    ("Asset Turnover",                   ["0.29x","0.34x","0.36x","0.38x","0.41x"], False, WHITE, DARK_TEXT),
    ("Revenue / Employee ($M)",          ["0.95","1.18","1.32","1.44","1.60"], False, WHITE, DARK_TEXT),
    ("Net Income / Employee ($M)",       ["0.49","0.60","0.70","0.79","0.80"], False, WHITE, DARK_TEXT),
    ("Employees (approx.)",              ["25,300","25,000","26,500","24,800","25,000"], False, WHITE, DARK_TEXT),
]

for lbl, vals, bold, fill, fc in roc_data:
    if not vals:
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
        c = ws.cell(row, 2, lbl)
        c.font = mfont(FONT_SIZE, bold=True, color=fc)
        c.fill = mfill(fill)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = mborder()
        rh(ws, row, 22); row += 1
        continue
    sc(ws, row, 2, lbl, bold=bold, fill_hex=fill, fcolor=fc)
    for j, v in enumerate(vals):
        sc(ws, row, 3+j, v, bold=bold, fill_hex=fill, align_h="center", fcolor=fc)
    rh(ws, row, 22); row += 1

blank(ws, row); row += 1

# Commentary
sec_hdr(ws, row, "  RETURN ON CAPITAL ASSESSMENT", 7); row += 1
roc_commentary = [
    ("ROIC vs. WACC Spread",
     "Visa earns an economic profit spread of ~20 percentage points (ROIC ~28% vs. WACC ~8%). "
     "This is exceptional among large-cap global companies. It reflects the pricing power and "
     "capital efficiency of the network model — each dollar reinvested generates outsized returns."),
    ("ROE Expansion",
     "ROE has expanded from 33% (FY2021) to 52% (FY2025) driven by: (1) margin expansion, "
     "(2) aggressive share buybacks reducing equity base, and (3) operating leverage. Even with "
     "$20B+ in goodwill/intangibles, ROE remains exceptional."),
    ("Return on Incremental Capital",
     "The return on each incremental dollar of CapEx is extraordinary — Visa's business does not "
     "require large capital investments to grow revenue. A $225M incremental CapEx supports $4B+ "
     "in incremental revenue. This is the hallmark of a capital-light, high-quality business."),
    ("FCF / Net Income > 1x",
     "FCF consistently exceeds net income, which is rare. This confirms that earnings quality is "
     "high, working capital is neutral-to-positive, and the business is not consuming cash to "
     "support operations. Every dollar earned can be returned to shareholders."),
]
for i, (k, v) in enumerate(roc_commentary):
    f = alt_row(i)
    sc(ws, row, 2, k, bold=True, fill_hex=f)
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=7)
    c = ws.cell(row, 3, v)
    c.font = mfont(FONT_SIZE)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c.fill = mfill(f)
    c.border = mborder()
    rh(ws, row, 60)
    row += 1

# ═══════════════════════════════════════════════════════════════
# 8. MANAGEMENT
# ═══════════════════════════════════════════════════════════════
ws = sheets["Management"]
cw(ws, {1:3, 2:32, 3:55, 4:3})

row = 1
sec_hdr(ws, row, "  MANAGEMENT ANALYSIS  —  Visa Inc.", 3); row += 1

# CEO Profile
sub_hdr(ws, row, "  CEO: Ryan McInerney — Profile", 3); row += 1
ceo_facts = [
    ("Name / Title",        "Ryan McInerney — President & CEO"),
    ("Tenure as CEO",       "February 2023 – present (Visa employee since June 2013; 13+ years at Visa)"),
    ("Previous Role",       "President of Visa Inc. (2013–2023); CFO at JPMorgan Chase retail bank"),
    ("Education",           "University of Notre Dame (Finance); Notre Dame Law School (J.D.)"),
    ("Compensation (FY2025)","$31.56M total; ~5% salary ($1.5M), ~95% variable (stock options, RSUs, cash bonus)"),
    ("Shareholding",         "Owns ~68,000 shares after recent sales; value ~$21M at current prices"),
    ("Insider Activity",     "37 SELLS, 0 BUYS over past 5 years — all under Rule 10b5-1 plan (adopted Apr 2024). "
                             "CEO McInerney sold 10,485 shares at ~$349 (Jan 2026) and 8,620 shares at ~$363 (Mar 2025)."),
]
for i, (k, v) in enumerate(ceo_facts):
    f = alt_row(i)
    sc(ws, row, 2, k, bold=True, fill_hex=f)
    merge_val(ws, row, 3, 3, v, fill_hex=f, wrap=True, align_h="left")
    rh(ws, row, 36)
    row += 1
blank(ws, row); row += 1

# Leadership team
sub_hdr(ws, row, "  Senior Leadership Team", 3); row += 1
leaders = [
    ("Ryan McInerney",    "President & CEO — Strategy, P&L accountability, board liaison"),
    ("Chris Suh",         "EVP & CFO — Capital allocation, investor relations, financial planning"),
    ("Rajat Taneja",      "President, Technology — VisaNet platform, AI/ML, cybersecurity, Visa Direct"),
    ("Kim Lawrence",      "President, North America — US and Canada consumer/commercial card strategy"),
    ("Oliver Jenkyn",     "Group President, North America (former) — now advisor"),
    ("Alan Gallo",        "EVP & General Counsel — Regulatory, legal, compliance"),
    ("Lynne Biggar",      "EVP, Chief Marketing Officer — Visa brand, sponsorships (FIFA, Olympics)"),
]
header_row(ws, row, ["Executive", "Role & Responsibility", ""], col_start=2, h=22); row += 1
for i, (name, role) in enumerate(leaders):
    f = alt_row(i)
    sc(ws, row, 2, name, bold=True, fill_hex=f)
    merge_val(ws, row, 3, 3, role, fill_hex=f, wrap=True)
    rh(ws, row, 26)
    row += 1
blank(ws, row); row += 1

# Proxy / Incentives
sub_hdr(ws, row, "  CEO Incentive Structure (Proxy Statement Analysis)", 3); row += 1
incentive_rows = [
    ("Base Salary",             "~$1.5M (4.8% of total comp) — demonstrates emphasis on variable pay"),
    ("Annual Cash Bonus",       "Based on: Revenue growth (25%), EPS growth (25%), FCF (25%), ESG/strategic (25%)"),
    ("Long-Term Equity (RSUs)", "3-year cliff vest; tied to 3-year relative TSR vs. peer group (S&P 500 Financials)"),
    ("Stock Options",           "5-year vest; underwater if stock doesn't appreciate — aligned with shareholders"),
    ("Peer Group",              "AmEx, Mastercard, PayPal, Fiserv, FIS, JPMorgan, S&P 500 index"),
    ("Our Assessment",          "Incentive structure is WELL-ALIGNED: short-term bonuses tied to financial KPIs, "
                                "long-term equity tied to TSR. High variable comp ratio (95%) ensures CEO is "
                                "incentivised as an owner. Rule 10b5-1 sales are pre-planned and not opportunistic."),
]
for i, (k, v) in enumerate(incentive_rows):
    f = alt_row(i)
    sc(ws, row, 2, k, bold=True, fill_hex=f)
    merge_val(ws, row, 3, 3, v, fill_hex=f, wrap=True)
    rh(ws, row, 38)
    row += 1
blank(ws, row); row += 1

# Capital allocation
sub_hdr(ws, row, "  Capital Allocation — Does Management Act Like Owners?", 3); row += 1
cap_alloc = [
    ("Share Buybacks",
     "$18.3B repurchased in FY2025; $30B authorization announced Q2 2025. Shares outstanding "
     "reduced from 2,185M (FY2021) to 1,966M (FY2025) — ~10% reduction. Buybacks accretive "
     "given FCF yield of ~3.5% vs. cost of capital of ~8%... borderline but acceptable at these "
     "multiples. Management has consistently bought back shares — a commitment to capital return."),
    ("Dividends",
     "Dividend grew from $1.28/share (FY2021) to $2.52/share (FY2025) — 18% CAGR. Payout ratio "
     "~23% of GAAP EPS leaves substantial room for continued growth. Visa has raised dividends "
     "every year for 15+ consecutive years."),
    ("M&A Strategy",
     "Visa attempted to acquire Plaid ($5.3B) in 2019 but abandoned due to DOJ antitrust concerns. "
     "Primarily bolt-on acquisitions (CyberSource, Tink, Currencycloud) to build VAS capabilities. "
     "No transformative M&A since 2016 Visa Europe acquisition ($21B) — management cautious on "
     "overpaying or inviting regulatory scrutiny."),
    ("CapEx Discipline",
     "CapEx of $1.5B on $40B revenue = 3.7% intensity — very disciplined. Investment in VisaNet "
     "infrastructure, tokenisation, and AI fraud systems. Management invests in durable, "
     "productivity-enhancing technologies, not vanity projects."),
    ("Use of Leverage",
     "Long-term debt of $21.2B ($10.66 EPS coverage). Net debt position of ~$5.9B is manageable. "
     "Debt is used to fund buybacks at attractive spreads — a deliberate, sensible strategy. "
     "Interest coverage ratio of ~27x provides substantial safety margin."),
    ("Seeds for Future Growth",
     "VAS (Value-Added Services) +26% YoY — management clearly investing for future diversification. "
     "Visa Direct (real-time payments) growing rapidly. Stablecoin/crypto integration positioning "
     "for next decade. Management is planting seeds without sacrificing current profitability."),
    ("Overall Assessment",
     "★★★★☆  Excellent management team. McInerney has continued the disciplined capital allocation "
     "philosophy of his predecessors. The $30B buyback authorization while maintaining dividend "
     "growth and investing in VAS demonstrates owner-operator mentality. Primary concern: systematic "
     "insider selling (37 sells, 0 buys) under 10b5-1 plan — though scheduled/pre-planned, it signals "
     "insiders do not view current price as cheap."),
]
for i, (k, v) in enumerate(cap_alloc):
    bold_label = True
    f = alt_row(i)
    if k == "Overall Assessment":
        f = LIGHT_GOLD_BG
    sc(ws, row, 2, k, bold=bold_label, fill_hex=f)
    merge_val(ws, row, 3, 3, v, fill_hex=f, wrap=True)
    rh(ws, row, 70)
    row += 1

# ═══════════════════════════════════════════════════════════════
# 9. RISKS
# ═══════════════════════════════════════════════════════════════
ws = sheets["Risks"]
cw(ws, {1:3, 2:28, 3:18, 4:18, 5:48, 6:3})

row = 1
sec_hdr(ws, row, "  RISK ANALYSIS  —  Visa Inc.", 5); row += 1
header_row(ws, row, ["Risk Category", "Severity", "Probability", "Description & Mitigation"], col_start=2, h=24); row += 1

risks = [
    # (category, severity, prob, description, fill)
    ("REGULATORY RISKS", "", "", "", VISA_MID_BLUE),
    ("Credit Card Competition Act (CCCA)",
     "HIGH", "20–30% (3yr)",
     "Reintroduced Jan 2026. Requires large banks (>$100B assets) to enable at least one non-Visa/MC "
     "routing network on credit cards. If passed, could reduce Visa's credit card volume by 10–20% "
     "over time as merchants route cheaper alternatives. MITIGATION: Visa lobbying aggressively; "
     "historical precedent (Durbin Amendment on debit) shows market adaptation is possible; "
     "VAS revenues partially offset volume-based risk.",
     LIGHT_RED_BG),
    ("EU / UK Interchange Caps",
     "MEDIUM", "Ongoing",
     "EU caps interchange at 0.3% (credit) and 0.2% (debit) since 2015. UK reviewing post-Brexit. "
     "Continued pressure on take-rates in Europe (~19% of non-US revenue). MITIGATION: Visa has "
     "adapted through increasing service fees and VAS revenue in Europe.",
     LIGHT_RED_BG),
    ("US DOJ / FTC Antitrust",
     "MEDIUM", "15–20%",
     "DOJ sued Visa in Sept 2024 over debit network monopoly practices (blocking Visa competitors). "
     "Case ongoing. If Visa loses, could face forced behavioral remedies or divestiture. "
     "MITIGATION: Visa has strong legal team and precedent of settlements; full divestiture is unlikely.",
     LIGHT_RED_BG),
    ("COMPETITIVE RISKS", "", "", "", VISA_MID_BLUE),
    ("Real-Time Payment Systems",
     "MEDIUM", "Structural/Long-term",
     "UPI (India, 13B+ monthly transactions), Pix (Brazil), FedNow (USA) are government-backed "
     "instant payment rails that bypass card networks for domestic P2P and P2B payments. "
     "MITIGATION: Visa is partnering with UPI (Visa on UPI program); Visa Direct competes directly "
     "in real-time push payments. Cross-border and cross-currency transactions remain Visa's stronghold.",
     LIGHT_GOLD_BG),
    ("Stablecoin / CBDC Disruption",
     "LOW-MEDIUM", "5–10% (5yr)",
     "If stablecoins (USDC, USDT) or Central Bank Digital Currencies achieve mass consumer adoption, "
     "peer-to-peer value transfer could theoretically bypass payment networks. "
     "MITIGATION: Visa is actively piloting USDC settlement on Solana; enabling crypto-linked card "
     "programs; positioning as infrastructure layer, not disrupted incumbent.",
     LIGHT_GOLD_BG),
    ("Big Tech Wallet Disintermediation",
     "LOW", "5–10%",
     "Apple Pay, Google Pay, and Samsung Pay currently ROUTE THROUGH Visa rails — increasing "
     "Visa volume. Risk is if Apple/Google build proprietary payment rails and bypass Visa. "
     "MITIGATION: Regulatory scrutiny of Apple's NFC monopoly; Apple Pay already accounts for "
     "significant Visa volumes (net positive). Apple Card is issued by Goldman on Mastercard, not Visa.",
     LIGHT_GOLD_BG),
    ("Mastercard Competition",
     "LOW", "Ongoing",
     "Mastercard is Visa's primary duopoly competitor. Well-matched in technology, global coverage, "
     "and customer relationships. MITIGATION: Duopoly structure means competition is rational; "
     "neither company competes on price. Both benefit from same tailwinds.",
     WHITE),
    ("MACRO / CYCLICAL RISKS", "", "", "", VISA_MID_BLUE),
    ("Consumer Spending Slowdown",
     "MEDIUM", "Cyclical",
     "Visa's revenue is correlated with consumer spending and GDP. A deep recession reduces "
     "payment volumes. Historical: FY2020 revenue declined ~5% (COVID). "
     "MITIGATION: Visa has very low fixed cost structure; asset-light model means it can weather "
     "volume declines; cross-border travel recovery is a long-term secular driver.",
     LIGHT_GOLD_BG),
    ("Currency / FX Risk",
     "LOW-MEDIUM", "Ongoing",
     "~61% of revenue is non-US. Strong USD headwinds reduce reported revenue from international "
     "operations. MITIGATION: Natural hedge (revenue and costs in same currencies in many markets); "
     "constant-currency growth outpaces reported growth in dollar-strength environments.",
     WHITE),
    ("Credit / Counterparty Risk",
     "LOW", "Low",
     "Visa does NOT extend credit — it processes payments. Credit risk sits with card-issuing banks. "
     "MITIGATION: Visa has settlement risk exposure (~$4-5B daily) but manages this through "
     "settlement finality windows and issuer guarantees.",
     WHITE),
    ("TECHNOLOGY RISKS", "", "", "", VISA_MID_BLUE),
    ("Cybersecurity / Data Breach",
     "MEDIUM", "Ongoing",
     "As the world's largest payment network, Visa is a high-value target for cybercriminals. "
     "A major breach would be catastrophic for brand trust. "
     "MITIGATION: Visa invests ~$10B+ cumulatively in cybersecurity; 99.999% network uptime; "
     "tokenisation removes actual card numbers from transactions; no breach of VisaNet at scale to date.",
     LIGHT_GOLD_BG),
    ("AI Disruption to Fraud / Analytics",
     "LOW", "5yr horizon",
     "AI-powered fraud attacks are increasingly sophisticated. Visa's own AI is industry-leading "
     "but the arms race continues. "
     "MITIGATION: Visa spends heavily on AI/ML; ROIC from fraud prevention is extremely high; "
     "this is a strength, not a weakness, for Visa vs. smaller competitors.",
     WHITE),
]

for r_data in risks:
    cat, sev, prob, desc, fill = r_data
    if sev == "" and prob == "" and desc == "":
        # Section header
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
        c = ws.cell(row, 2, f"  {cat}")
        c.font = mfont(FONT_SIZE, bold=True, color=HEADER_TEXT)
        c.fill = mfill(fill)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = mborder()
        rh(ws, row, 22)
        row += 1
        continue
    sc(ws, row, 2, cat, bold=True, fill_hex=fill)
    fcolor = ACCENT_RED if sev == "HIGH" else (VISA_DARK_BLUE if sev == "MEDIUM" else ACCENT_GREEN)
    sc(ws, row, 3, sev, bold=True, fill_hex=fill, align_h="center", fcolor=fcolor)
    sc(ws, row, 4, prob, fill_hex=fill, align_h="center")
    ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=5)
    c = ws.cell(row, 5, desc)
    c.font = mfont(FONT_SIZE)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c.fill = mfill(fill)
    c.border = mborder()
    rh(ws, row, 68)
    row += 1

# ═══════════════════════════════════════════════════════════════
# 10. VALUATION
# ═══════════════════════════════════════════════════════════════
ws = sheets["Valuation"]
cw(ws, {1:3, 2:36, 3:22, 4:22, 5:22, 6:22, 7:3})

row = 1
sec_hdr(ws, row, "  VALUATION  —  Visa Inc. (NYSE: V)  |  As of April 15, 2026", 6); row += 1

# Current multiples
sub_hdr(ws, row, "  Current Trading Multiples", 6); row += 1
header_row(ws, row, ["Metric", "Current (TTM)", "FY2026E", "5-Year Avg", "Peer Median", "Premium/(Discount)"], col_start=2, h=24); row += 1
multiples = [
    ("Stock Price",          "$312.68",    "$312.68",  "—",      "—",      "—"),
    ("Market Cap",           "$595.8B",    "$595.8B",  "—",      "—",      "—"),
    ("Enterprise Value",     "$602.3B",    "$602.3B",  "—",      "—",      "—"),
    ("P/E Ratio",            "29.3×",      "23.5×",    "27.5×",  "22.0×",  "+33% premium"),
    ("EV / EBITDA",          "23.4×",      "20.1×",    "22.0×",  "17.5×",  "+34% premium"),
    ("EV / Revenue (P/S)",   "14.5×",      "13.0×",    "13.5×",  "10.5×",  "+38% premium"),
    ("P / FCF",              "27.5×",      "23.8×",    "26.0×",  "20.0×",  "+38% premium"),
    ("Dividend Yield",       "0.86%",      "0.86%",    "0.62%",  "0.75%",  "Premium yield"),
    ("PEG Ratio",            "2.0×",       "—",        "—",      "—",      "Fairly valued for growth"),
]
for i, row_data in enumerate(multiples):
    f = alt_row(i)
    for j, v in enumerate(row_data):
        sc(ws, row, 2+j, v, fill_hex=f, align_h="center" if j > 0 else "left",
           bold=(j == 0))
    rh(ws, row, 22)
    row += 1
blank(ws, row); row += 1

# DCF Analysis
sub_hdr(ws, row, "  DCF Valuation  (5-Year FCF Projection + Terminal Value)", 6); row += 1
header_row(ws, row, ["Year", "Revenue Est. ($M)", "FCF Margin", "FCF ($M)", "Discount Factor (8%)", "PV of FCF ($M)"], col_start=2, h=24); row += 1

dcf_rows = [
    ("FY2026E",  43500, "54.5%", 23698, 0.926,  21944),
    ("FY2027E",  47200, "55.0%", 25960, 0.857,  22248),
    ("FY2028E",  51100, "55.5%", 28361, 0.794,  22519),
    ("FY2029E",  55100, "56.0%", 30856, 0.735,  22679),
    ("FY2030E",  59400, "56.5%", 33561, 0.681,  22853),
    ("Terminal Value (3% grow)", "—", "—", "—", "—", 322000),
    ("Total PV of Cash Flows",    "—", "—", "—", "—", 434243),
]
for i, (yr, rev, fcfm, fcf, df, pv) in enumerate(dcf_rows):
    f = LIGHT_GOLD_BG if "Total" in str(yr) or "Terminal" in str(yr) else alt_row(i)
    bold = "Total" in str(yr) or "Terminal" in str(yr)
    sc(ws, row, 2, yr, bold=bold, fill_hex=f)
    sc(ws, row, 3, rev, bold=bold, fill_hex=f, align_h="right",
       num_fmt='#,##0' if isinstance(rev, int) else '@')
    sc(ws, row, 4, fcfm, bold=bold, fill_hex=f, align_h="center")
    sc(ws, row, 5, fcf, bold=bold, fill_hex=f, align_h="right",
       num_fmt='#,##0' if isinstance(fcf, int) else '@')
    sc(ws, row, 6, df, bold=bold, fill_hex=f, align_h="center")
    sc(ws, row, 7, pv, bold=bold, fill_hex=f, align_h="right",
       num_fmt='#,##0' if isinstance(pv, int) else '@')
    rh(ws, row, 22)
    row += 1

# DCF summary
dcf_summary = [
    ("Total PV of FCFs ($M)", "112,243"),
    ("Terminal Value PV ($M)", "322,000"),
    ("Enterprise Value ($M)",  "434,243"),
    ("Less: Net Debt ($M)",    "(5,937)"),
    ("Equity Value ($M)",      "428,306"),
    ("Shares Outstanding (M)", "1,910"),
    ("DCF Intrinsic Value / Share", "$224.24"),
    ("Current Price", "$312.68"),
    ("Premium to DCF", "+39.4% (stock trades above base-case DCF)"),
    ("Note", "DCF is sensitive to WACC and terminal growth. At 7% WACC + 4% terminal growth → $380/share"),
]
blank(ws, row); row += 1
for i, (k, v) in enumerate(dcf_summary):
    f = LIGHT_GOLD_BG if "DCF Intrinsic" in k or "Note" in k else alt_row(i)
    sc(ws, row, 2, k, bold=True, fill_hex=f)
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)
    sc(ws, row, 3, v, fill_hex=f, align_h="left")
    rh(ws, row, 22)
    row += 1
blank(ws, row); row += 1

# Scenarios
sub_hdr(ws, row, "  Bull / Base / Bear Valuation Scenarios", 6); row += 1
header_row(ws, row, ["Scenario", "Key Assumptions", "Revenue CAGR", "FCF CAGR", "Target P/FCF", "Price Target"], col_start=2, h=24); row += 1
scenarios = [
    ("BULL CASE",
     "VAS accelerates to >30% growth; CCCA fails; stablecoin integration drives new revenue; "
     "buybacks continue aggressively; no major regulatory action.",
     "14–15%", "13–14%", "30×", "$430 – $450"),
    ("BASE CASE",
     "Revenue grows ~11–12%; VAS continues at 20–25%; moderate regulatory headwinds; "
     "buybacks sustain 10–12% EPS growth; interest rates stable.",
     "11–12%", "10–12%", "26×", "$370 – $400"),
    ("BEAR CASE",
     "CCCA passes, forcing alternative routing; EU/US interchange caps; macro recession reduces "
     "payment volumes; Big Tech builds competing rails; FCF margin contracts.",
     "6–7%",  "5–6%",  "20×", "$220 – $260"),
    ("CURRENT PRICE", "→  $312.68 (April 15, 2026)",
     "—", "—", "~28×", "$312.68"),
]
fills = [LIGHT_GREEN_BG, LIGHT_BLUE_BG, LIGHT_RED_BG, LIGHT_GOLD_BG]
for i, (scen, assum, rev, fcf, pe, pt) in enumerate(scenarios):
    f = fills[i]
    bold = i == 3
    sc(ws, row, 2, scen, bold=True, fill_hex=f)
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=3)
    c = ws.cell(row, 3, assum)
    c.font = mfont(FONT_SIZE)
    c.fill = mfill(f)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c.border = mborder()
    sc(ws, row, 4, rev, bold=bold, fill_hex=f, align_h="center")
    sc(ws, row, 5, fcf, bold=bold, fill_hex=f, align_h="center")
    sc(ws, row, 6, pe, bold=bold, fill_hex=f, align_h="center")
    sc(ws, row, 7, pt, bold=True, fill_hex=f, align_h="center")
    rh(ws, row, 56)
    row += 1
blank(ws, row); row += 1

# Safety of margin
sub_hdr(ws, row, "  Is There a Margin of Safety at $312.68?", 6); row += 1
mos_text = (
    "ASSESSMENT:  The current price of $312.68 is BELOW the analyst consensus target of $397 (+27% upside) "
    "and the upper end of the base-case scenario. However, it is ABOVE the strict base-case DCF of ~$224 "
    "(using 8% WACC, 3% terminal growth). The margin of safety depends on your assumptions:\n\n"
    "•  DISCOUNT RATE VIEW: At a 7% WACC (justified given Visa's stable cash flows and wide moat), "
    "   intrinsic value rises to ~$340–380, making the current price slightly undervalued.\n\n"
    "•  GROWTH INVESTOR VIEW: Paying 23.5× forward earnings for a business with 28% ROIC, ~11% revenue "
    "   CAGR, ~15% EPS CAGR, and minimal capital requirements is reasonable to attractive.\n\n"
    "•  RISK-ADJUSTED VIEW: CCCA passage probability (20–30%) creates headline risk. DOJ debit case "
    "   creates overhang. These regulatory risks explain the current ~21% discount to analyst targets.\n\n"
    "VERDICT:  Visa represents a FAIR to SLIGHTLY UNDERVALUED opportunity relative to intrinsic value "
    "at current prices. There is ~21–27% upside to analyst consensus. A regulatory resolution or "
    "continued VAS growth acceleration would be a catalyst. Position sizing should reflect regulatory risk."
)
ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
c = ws.cell(row, 2, mos_text)
c.font = mfont(FONT_SIZE)
c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
c.fill = mfill(LIGHT_GOLD_BG)
c.border = mborder()
rh(ws, row, 200)

# ═══════════════════════════════════════════════════════════════
# 11. MARKET SENTIMENT
# ═══════════════════════════════════════════════════════════════
ws = sheets["Market Sentiment"]
cw(ws, {1:3, 2:30, 3:20, 4:20, 5:20, 6:20, 7:3})

row = 1
sec_hdr(ws, row, "  MARKET SENTIMENT  —  Visa Inc. (V)  |  April 15, 2026", 6); row += 1

# Analyst consensus
sub_hdr(ws, row, "  Analyst Ratings & Price Targets", 6); row += 1
header_row(ws, row, ["Metric", "Value", "Detail", "", "", ""], col_start=2, h=24); row += 1
analyst_data = [
    ("Consensus Rating",      "STRONG BUY",    "21 analysts covering Visa as of April 2026"),
    ("Buy / Hold / Sell",     "19 / 2 / 0",    "No SELL ratings; 2 HOLD ratings"),
    ("Average Price Target",  "$397.43",        "Implies +27.1% upside from $312.68"),
    ("Median Price Target",   "$400.60",        "Median of 54 analyst targets"),
    ("Highest Price Target",  "$450.00",        "Implies +43.9% upside"),
    ("Lowest Price Target",   "$323.00",        "Implies +3.3% upside — most conservative"),
    ("Current Price",         "$312.68",        "As of April 15, 2026 — near 52-week low"),
    ("52-Week Range",         "$293.89 – $375.51", "Currently in lower third of range"),
    ("Stock Performance YTD", "-8.3%",          "Underperforming S&P 500 YTD 2026"),
    ("Next Earnings",         "April 28, 2026", "Q2 FY2026 — potential catalyst"),
    ("Recent Upgrade",        "Freedom Capital: Hold → Buy (Feb 17, 2026)", "Post-Q1 beat"),
    ("TD Cowen",              "MAINTAIN BUY — $380 target", "After Q1 FY2026 results"),
]
for i, (metric, val, detail) in enumerate(analyst_data):
    f = alt_row(i)
    sc(ws, row, 2, metric, bold=True, fill_hex=f)
    green_metrics = {"STRONG BUY", "$397.43", "$400.60"}
    red_metrics = {"-8.3%"}
    fc = ACCENT_GREEN if val in green_metrics else (ACCENT_RED if val in red_metrics else DARK_TEXT)
    sc(ws, row, 3, val, bold=val in green_metrics, fill_hex=f, align_h="center", fcolor=fc)
    ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=6)
    sc(ws, row, 4, detail, fill_hex=f, italic=True, wrap=True)
    rh(ws, row, 26)
    row += 1
blank(ws, row); row += 1

# Institutional
sub_hdr(ws, row, "  Institutional Ownership & Insider Activity", 6); row += 1
inst_data = [
    ("Institutional Ownership",  "~87%", "Vanguard, BlackRock, State Street are top holders"),
    ("Float",                    "~99%",  "Very high float; highly liquid large-cap"),
    ("Short Interest",           "~0.8%","Low; limited bearish positioning despite regulatory overhang"),
    ("Short Interest Trend",     "Increasing", "Short interest has grown modestly in early 2026"),
    ("Insider Net Activity",     "SELLING","CEO McInerney: 37 sells, 0 buys over 5 years (Rule 10b5-1)"),
    ("Largest Shareholder",      "Vanguard Group", "~8.2% ownership; passive index holder"),
    ("2nd Largest",              "BlackRock",       "~7.4% ownership; passive index holder"),
    ("3rd Largest",              "State Street",    "~4.1% ownership; passive index holder"),
    ("Notable Activity (Q1 2026)","Mixed Institutional","V Square Quant (-13.5%); Richardson Fin. (-62.7%)"),
]
header_row(ws, row, ["Metric", "Value", "Detail", "", "", ""], col_start=2, h=22); row += 1
for i, (m, v, d) in enumerate(inst_data):
    f = alt_row(i)
    sc(ws, row, 2, m, bold=True, fill_hex=f)
    sc(ws, row, 3, v, fill_hex=f, align_h="center")
    ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=6)
    sc(ws, row, 4, d, fill_hex=f, italic=True, wrap=True)
    rh(ws, row, 26)
    row += 1
blank(ws, row); row += 1

# Recent news catalysts
sub_hdr(ws, row, "  Recent Catalysts & News  (2025–2026)", 6); row += 1
news = [
    ("Jan 29, 2026", "Q1 FY2026 Beat",          "+", "Revenue $10.9B (+15%), EPS $3.17 beat $3.14 est.; stock rose +1.5% AH"),
    ("Jan 2026",     "CCCA Reintroduced",        "-", "Credit Card Competition Act reintroduced in Congress; regulatory overhang on stock"),
    ("Sep 2025",     "DOJ Debit Case Filed",     "-", "DOJ sued Visa alleging debit card monopoly; pending court proceedings"),
    ("May 2025",     "$30B Buyback Authorization","+" ,"Board approved $30B share repurchase program — largest in company history"),
    ("Oct 2025",     "FY2025 Full-Year Results",  "+", "Revenue $40B (+11%), EPS $10.20 (+5%); VAS +26%; $22.9B returned to shareholders"),
    ("Feb 2026",     "Freedom Capital Upgrade",   "+", "Upgraded from Hold to Buy; sees regulatory overhang as excessive"),
    ("Apr 2026",     "Q2 FY2026 Upcoming",        "→", "Earnings April 28, 2026; street expects ~$11.0B revenue, $3.28 EPS"),
]
header_row(ws, row, ["Date", "Event", "+/-", "Description", "", ""], col_start=2, h=22); row += 1
for i, (date, event, sign, desc) in enumerate(news):
    f = LIGHT_GREEN_BG if sign == "+" else (LIGHT_RED_BG if sign == "-" else LIGHT_GOLD_BG)
    sc(ws, row, 2, date, fill_hex=f, align_h="center")
    sc(ws, row, 3, event, bold=True, fill_hex=f)
    sc(ws, row, 4, sign, bold=True, fill_hex=f, align_h="center",
       fcolor=ACCENT_GREEN if sign == "+" else (ACCENT_RED if sign == "-" else DARK_TEXT))
    ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)
    c = ws.cell(row, 5, desc)
    c.font = mfont(FONT_SIZE)
    c.fill = mfill(f)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c.border = mborder()
    rh(ws, row, 36)
    row += 1

# ═══════════════════════════════════════════════════════════════
# 12. KEY INDICATORS
# ═══════════════════════════════════════════════════════════════
ws = sheets["Key Indicators"]
cw(ws, {1:3, 2:36, 3:18, 4:18, 5:18, 6:18, 7:18, 8:3})

row = 1
sec_hdr(ws, row, "  KEY PERFORMANCE INDICATORS  —  Visa Inc.  (FY2021 – FY2025)", 7); row += 1

# Volume KPIs
sub_hdr(ws, row, "  Operational Volume Metrics", 7); row += 1
header_row(ws, row, ["KPI"] + years + ["Q1 FY2026"], col_start=2, h=24); row += 1
kpi_data = [
    ("Payments Volume ($T)",          ["7.1", "11.6", "12.3", "13.2", "15.7", "~$4.0 (qtr)"],  False, WHITE),
    ("  YoY Growth",                  ["—",  "+63%",  "+6%",  "+7%",  "+19%",  "+8%"],           False, VERY_LIGHT_BG),
    ("Processed Transactions (B)",    ["138", "192",  "212",  "233",  "264",   "69 (qtr)"],       False, WHITE),
    ("  YoY Growth",                  ["—",  "+39%",  "+10%", "+10%", "+13%",  "+9%"],            False, VERY_LIGHT_BG),
    ("Cross-Border Volume Growth",    ["-29%", "+45%", "+31%", "+21%", "+16%", "+16%"],           False, WHITE),
    ("Cards in Circulation (B)",      ["3.6", "3.9",  "4.1",  "4.4",  "4.5+",  "—"],             False, VERY_LIGHT_BG),
    ("Merchant Locations (M)",        ["80",  "100",  "110",  "120",  "130+",  "—"],              False, WHITE),
    ("Countries / Territories",       ["200+","200+", "200+", "200+", "200+",  "200+"],           False, VERY_LIGHT_BG),
    ("VAS Revenue Growth",            ["—",  "+21%",  "+25%", "+23%", "+26%",  "+28%"],           False, WHITE),
]
for lbl, vals, bold, fill in kpi_data:
    sc(ws, row, 2, lbl, bold=bold, fill_hex=fill)
    for j, v in enumerate(vals):
        sc(ws, row, 3+j, v, bold=bold, fill_hex=fill, align_h="center")
    rh(ws, row, 22)
    row += 1
blank(ws, row); row += 1

# Margins dashboard
sub_hdr(ws, row, "  Profitability Margins Dashboard", 7); row += 1
header_row(ws, row, ["Margin"] + years + ["TTM"], col_start=2, h=24); row += 1
margin_data = [
    ("Gross Margin",     ["81.0%","81.0%","81.0%","81.0%","81.1%","81.0%"], True, LIGHT_GREEN_BG),
    ("Operating Margin", ["65.6%","64.2%","64.3%","65.7%","60.0%","59.2%"], True, LIGHT_BLUE_BG),
    ("Net Margin",       ["51.1%","51.0%","52.9%","54.9%","50.1%","50.2%"], True, LIGHT_GOLD_BG),
    ("EBITDA Margin",    ["69.8%","68.7%","68.7%","69.9%","64.3%","62.2%"], True, LIGHT_GREEN_BG),
    ("FCF Margin",       ["60.2%","61.0%","60.3%","52.0%","53.9%","55.4%"], True, LIGHT_BLUE_BG),
]
for lbl, vals, bold, fill in margin_data:
    sc(ws, row, 2, lbl, bold=bold, fill_hex=fill)
    for j, v in enumerate(vals):
        sc(ws, row, 3+j, v, bold=bold, fill_hex=fill, align_h="center")
    rh(ws, row, 24)
    row += 1
blank(ws, row); row += 1

# Returns dashboard
sub_hdr(ws, row, "  Returns & Capital Efficiency Dashboard", 7); row += 1
header_row(ws, row, ["Metric"] + years + ["TTM/Est."], col_start=2, h=24); row += 1
returns_data = [
    ("ROIC",              ["20.5%","24.2%","26.1%","27.5%","27.9%","~28%"],    True, LIGHT_GREEN_BG),
    ("ROE",               ["32.8%","40.4%","44.8%","51.0%","51.5%","~52%"],    True, LIGHT_BLUE_BG),
    ("ROA",               ["14.9%","17.6%","19.3%","21.1%","20.9%","~21%"],    True, WHITE),
    ("EPS (Diluted)",     ["$5.63","$7.00","$8.28","$9.73","$10.20","$10.66"], True, LIGHT_GOLD_BG),
    ("EPS YoY Growth",    ["—",   "+24.3%","+18.3%","+17.5%","+4.8%","~7.5%"],False, WHITE),
    ("FCF per Share",     ["$6.64","$8.37","$9.44","$9.21","$10.98","$10.52"], True, LIGHT_BLUE_BG),
    ("Dividend/Share",    ["$1.28","$1.50","$1.80","$2.08","$2.52","$2.68"],   False, WHITE),
    ("Shares Out. (M)",   ["2,185","2,136","2,087","2,029","1,966","~1,910"],  False, VERY_LIGHT_BG),
    ("Buybacks ($B)",     ["$8.7","$11.6","$12.1","$16.7","$18.3","Ongoing"],  False, WHITE),
]
for lbl, vals, bold, fill in returns_data:
    sc(ws, row, 2, lbl, bold=bold, fill_hex=fill)
    for j, v in enumerate(vals):
        sc(ws, row, 3+j, v, bold=bold, fill_hex=fill, align_h="center")
    rh(ws, row, 22)
    row += 1
blank(ws, row); row += 1

# Q1 FY2026 vs prior quarters
sub_hdr(ws, row, "  Quarterly Revenue Trend ($M)", 7); row += 1
header_row(ws, row, ["Quarter", "Revenue", "YoY Growth", "EPS", "EPS Growth", "Key Highlight", ""], col_start=2, h=24); row += 1
quarterly_trend = [
    ("Q1 FY2024", "8,634",  "+9.0%",  "$2.41", "+22%", "Cross-border recovery accelerates"),
    ("Q2 FY2024", "8,775",  "+10.0%", "$2.51", "+20%", "Q2 seasonality; debit strong"),
    ("Q3 FY2024", "8,900",  "+9.5%",  "$2.42", "+17%", "Summer cross-border peak"),
    ("Q4 FY2024", "9,617",  "+12.0%", "$2.71", "+16%", "Fiscal year-end; VAS milestone"),
    ("Q1 FY2025", "9,488",  "+9.9%",  "$2.75", "+14%", "Holiday season; beat estimates"),
    ("Q2 FY2025", "9,909",  "+12.9%", "$2.88", "+15%", "VAS acceleration noted"),
    ("Q3 FY2025", "10,014", "+12.5%", "$2.68", "+11%", "$30B buyback authorization"),
    ("Q4 FY2025", "10,649", "+10.7%", "$2.71", "0%",   "DOJ debit case announced; drag"),
    ("Q1 FY2026", "10,900", "+14.9%", "$3.17", "+15%", "BEAT: Rev +2.1%, EPS +$0.03 est."),
    ("Q2 FY2026E","~11,000","~+11%",  "~$3.28","~+14%", "Est; earnings April 28, 2026"),
]
for i, (q, rev, yoy, eps, eg, hl) in enumerate(quarterly_trend):
    f = LIGHT_GOLD_BG if "Q1 FY2026" in q else (LIGHT_BLUE_BG if "E" in q else alt_row(i))
    bold = "Q1 FY2026" in q
    sc(ws, row, 2, q, bold=bold, fill_hex=f, align_h="center")
    sc(ws, row, 3, rev, bold=bold, fill_hex=f, align_h="right", num_fmt='@')
    sc(ws, row, 4, yoy, bold=bold, fill_hex=f, align_h="center")
    sc(ws, row, 5, eps, bold=bold, fill_hex=f, align_h="center")
    sc(ws, row, 6, eg, bold=bold, fill_hex=f, align_h="center")
    ws.merge_cells(start_row=row, start_column=7, end_row=row, end_column=8)
    c = ws.cell(row, 7, hl)
    c.font = mfont(FONT_SIZE, italic=True)
    c.fill = mfill(f)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c.border = mborder()
    rh(ws, row, 24)
    row += 1

blank(ws, row); row += 1

# Sources
sec_hdr(ws, row, "  DATA SOURCES & METHODOLOGY", 7); row += 1
sources = [
    "StockAnalysis.com — Income Statement, Balance Sheet, Cash Flow (FY2021–TTM)",
    "Visa Investor Relations (investor.visa.com) — Q1 FY2026 Earnings Release, Q4/FY2025 Results",
    "SEC EDGAR — 10-K Annual Report (FY2025), Form 4 Insider Filings, Proxy Statement",
    "MacroTrends.net — Historical Financial Ratios, ROIC, ROE, Margins",
    "GuruFocus.com — ROIC, Intrinsic Value (DCF), Insider Transactions",
    "AlphaSpread.com — DCF Valuation, WACC Analysis",
    "MarketBeat.com — Analyst Consensus, Price Targets, Short Interest",
    "Investing.com — Earnings Call Transcripts (Q1 FY2026, Q4 FY2025)",
    "Yahoo Finance — Earnings Estimates, Forward Metrics",
    "Benzinga / FinViz — Market Sentiment, News Catalysts",
    "Yale TAP Paper 2025 — Academic Analysis of Visa's Competitive Position",
    "Payments Dive — Industry Analysis of Competitive Dynamics",
]
for i, src in enumerate(sources):
    f = alt_row(i)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
    c = ws.cell(row, 2, f"  •  {src}")
    c.font = mfont(FONT_SIZE - 1, italic=True)
    c.fill = mfill(f)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = mborder()
    rh(ws, row, 20)
    row += 1

# ─────────────────────────────────────────────────────────────
# SAVE
# ─────────────────────────────────────────────────────────────
os.makedirs(OUTPUT_DIR, exist_ok=True)
wb.save(OUTPUT_FILE)
print(f"✅  Visa Financial Analysis saved → {OUTPUT_FILE}")
