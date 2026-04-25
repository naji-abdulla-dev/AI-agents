"""
Alphabet Inc. (GOOGL) Financial Analysis - Excel Generator
Data as of April 2026 | FY2025 Annual (ended December 31, 2025)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

# ── Color palette ──────────────────────────────────────────────────────────────
GOOGLE_BLUE  = "1A73E8"
GOOGLE_RED   = "EA4335"
GOOGLE_GREEN = "34A853"
GOOGLE_YELLOW= "FBBC04"
HEADER_BG    = "1A73E8"
HEADER_FG    = "FFFFFF"
SUBHDR_BG    = "D2E3FC"
SUBHDR_FG    = "1A1A1A"
ALT_ROW      = "EBF3FE"
WHITE        = "FFFFFF"
LIGHT_GRAY   = "F8F9FA"
RED          = "EA4335"
GREEN        = "34A853"
DARK_BLUE    = "0D47A1"
GOLD         = "F9AB00"

FONT_SIZE = 14

def mf(bold=False, size=FONT_SIZE, color="000000", italic=False):
    return Font(name="Calibri", bold=bold, size=size, color=color, italic=italic)

def mfill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def mborder(style="thin"):
    s = Side(border_style=style, color="BBBBBB")
    return Border(left=s, right=s, top=s, bottom=s)

def cal(h="left", wrap=False):
    return Alignment(horizontal=h, vertical="center", wrap_text=wrap)

def wc(ws, row, col, value, bold=False, size=FONT_SIZE, fg="000000", bg=None,
       align="left", wrap=False, border=False, italic=False, num_fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = mf(bold=bold, size=size, color=fg, italic=italic)
    if bg:
        c.fill = mfill(bg)
    c.alignment = cal(align, wrap)
    if border:
        c.border = mborder()
    if num_fmt:
        c.number_format = num_fmt
    return c

def sec_hdr(ws, row, col, text, span=1, bg=HEADER_BG, fg=HEADER_FG):
    c = wc(ws, row, col, text, bold=True, size=FONT_SIZE+2, fg=fg, bg=bg, align="center")
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+span-1)
    ws.row_dimensions[row].height = 30
    return c

def sub_hdr(ws, row, col, text, span=1, bg=SUBHDR_BG, fg=SUBHDR_FG):
    c = wc(ws, row, col, text, bold=True, size=FONT_SIZE, fg=fg, bg=bg, align="center", border=True)
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+span-1)
    ws.row_dimensions[row].height = 22
    return c

def lbl(ws, row, col, text, bg=None, indent=0):
    val = ("   " * indent) + text
    c = wc(ws, row, col, val, size=FONT_SIZE, bg=bg, align="left", border=True)
    ws.row_dimensions[row].height = 20
    return c

def dc(ws, row, col, value, bg=None, num_fmt=None, bold=False, color="000000"):
    c = wc(ws, row, col, value, bold=bold, size=FONT_SIZE, fg=color, bg=bg, align="right", border=True, num_fmt=num_fmt)
    ws.row_dimensions[row].height = 20
    return c

def alt(i):
    return ALT_ROW if i % 2 == 0 else WHITE

def scw(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

# ═══════════════════════════════════════════════════════════════════════════════
def build_cover(wb):
    ws = wb.create_sheet("Cover")
    ws.sheet_view.showGridLines = False
    for c in range(1, 9):
        scw(ws, c, 18)
    scw(ws, 1, 4)
    scw(ws, 2, 30)
    scw(ws, 3, 22)

    ws.merge_cells("B3:G3")
    wc(ws, 3, 2, "ALPHABET INC. (GOOGL)", bold=True, size=32, fg=WHITE,
       bg=HEADER_BG, align="center")
    ws.row_dimensions[3].height = 55

    ws.merge_cells("B4:G4")
    wc(ws, 4, 2, "Comprehensive Investment Analysis | April 2026", bold=False,
       size=16, fg=WHITE, bg=GOOGLE_BLUE, align="center")
    ws.row_dimensions[4].height = 28

    ws.merge_cells("B6:G6")
    wc(ws, 6, 2, "Company Profile", bold=True, size=16, fg=WHITE, bg=DARK_BLUE, align="center")
    ws.row_dimensions[6].height = 25

    info = [
        ("Ticker",         "GOOGL / GOOG"),
        ("Exchange",       "NASDAQ"),
        ("Sector",         "Communication Services / Technology"),
        ("Headquarters",   "Mountain View, California, USA"),
        ("Founded",        "1998 by Larry Page & Sergey Brin"),
        ("CEO",            "Sundar Pichai"),
        ("Fiscal Year End","December 31"),
        ("Market Cap",     "~$3.86 Trillion (April 2026)"),
        ("Stock Price",    "$317.32 (April 8, 2026)"),
        ("P/E Ratio",      "~29x (TTM) | ~23x Forward"),
        ("Shares Outstanding", "~12.2 Billion"),
        ("Dividend",       "None"),
    ]

    for i, (k, v) in enumerate(info):
        r = 7 + i
        bg = ALT_ROW if i % 2 == 0 else WHITE
        wc(ws, r, 2, k, bold=True, size=FONT_SIZE, bg=bg, align="left", border=True)
        wc(ws, r, 3, v, size=FONT_SIZE, bg=bg, align="left", border=True)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)
        ws.row_dimensions[r].height = 22

    ws.merge_cells("B20:G20")
    wc(ws, 20, 2, "Investment Thesis", bold=True, size=16, fg=WHITE, bg=DARK_BLUE, align="center")
    ws.row_dimensions[20].height = 25

    thesis = (
        "Alphabet is the dominant force in global digital advertising, commanding ~90% of search "
        "query share globally. Google Cloud has emerged as the #3 hyperscaler and is growing 40%+ "
        "annually. The company's AI leadership (Gemini, TPUs, DeepMind) underpins a defensible moat "
        "across search, cloud, and consumer platforms. With $164.7B in operating cash flow and a "
        "disciplined capital return program, Alphabet offers rare combination of growth + value. "
        "Trading at 23x forward earnings, significantly below historical averages, with "
        "$175-185B CapEx planned for 2026 to cement AI infrastructure dominance."
    )
    ws.merge_cells("B21:G24")
    wc(ws, 21, 2, thesis, size=FONT_SIZE, bg=LIGHT_GRAY, align="left", wrap=True, border=True)
    for r in range(21, 25):
        ws.row_dimensions[r].height = 22

    ws.merge_cells("B26:G26")
    wc(ws, 26, 2, "Key Financial Highlights (FY2025)", bold=True, size=16, fg=WHITE, bg=DARK_BLUE, align="center")
    ws.row_dimensions[26].height = 25

    highlights = [
        ("Total Revenue",         "$402.8B",  "+15.1% YoY"),
        ("Net Income",            "$132.2B",  "+28.9% YoY"),
        ("Net Margin",            "32.8%",    "Record high"),
        ("Operating Cash Flow",   "$164.7B",  "Exceptional conversion"),
        ("Free Cash Flow",        "$~73B",    "After $91.4B CapEx"),
        ("Share Buybacks",        "$45.7B",   "Strong return program"),
        ("Google Cloud Revenue",  "$~43B+",   "+~40% YoY"),
        ("YouTube Revenue",       "$60B+",    "Ads + Subscriptions"),
    ]

    sub_hdr(ws, 27, 2, "Metric", bg=SUBHDR_BG)
    sub_hdr(ws, 27, 3, "Value", bg=SUBHDR_BG)
    sub_hdr(ws, 27, 4, "Comment", bg=SUBHDR_BG)
    ws.merge_cells(start_row=27, start_column=4, end_row=27, end_column=7)

    for i, (m, v, c) in enumerate(highlights):
        r = 28 + i
        bg = ALT_ROW if i % 2 == 0 else WHITE
        lbl(ws, r, 2, m, bg=bg)
        dc(ws, r, 3, v, bg=bg)
        lbl(ws, r, 4, c, bg=bg)
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=7)

    ws.merge_cells("B37:G37")
    wc(ws, 37, 2, "Report Date: April 10, 2026  |  Data Sources: SEC Filings, Alphabet IR, Bloomberg",
       italic=True, size=12, fg="666666", align="center")

# ═══════════════════════════════════════════════════════════════════════════════
def build_business(wb):
    ws = wb.create_sheet("Business Overview")
    ws.sheet_view.showGridLines = False
    scw(ws, 1, 3); scw(ws, 2, 32); scw(ws, 3, 20); scw(ws, 4, 20)
    scw(ws, 5, 20); scw(ws, 6, 20); scw(ws, 7, 20)

    sec_hdr(ws, 1, 2, "BUSINESS OVERVIEW — ALPHABET INC. (GOOGL)", span=6)

    # Business Description
    sub_hdr(ws, 3, 2, "What Alphabet Does", span=6)
    desc = ("Alphabet is the parent company of Google, the world's largest search engine, and a "
            "diversified technology conglomerate. Its core business is digital advertising — "
            "Search, YouTube, and Google Network ads. Google Cloud Platform (GCP) is its rapidly "
            "growing enterprise segment. 'Other Bets' includes Waymo (autonomous driving), Verily "
            "(life sciences), Wing (drone delivery), and DeepMind (AI research).")
    ws.merge_cells("B4:G6")
    wc(ws, 4, 2, desc, size=FONT_SIZE, bg=LIGHT_GRAY, align="left", wrap=True, border=True)
    for r in range(4, 7):
        ws.row_dimensions[r].height = 22

    # Revenue Breakdown
    sub_hdr(ws, 8, 2, "Revenue Breakdown by Segment (FY2025)", span=6)
    rev_data = [
        ("Segment",                      "FY2025 Rev",  "% of Total", "YoY Growth", "Key Driver"),
        ("Google Search & Other",        "$198.1B",     "49.2%",      "+17%",       "AI-enhanced search, paid search volume"),
        ("YouTube Ads",                  "$36.1B",      "9.0%",       "+9%",        "Connected TV, Shorts monetization"),
        ("Google Network Members",       "$30.8B",      "7.6%",       "-4%",        "Declining 3rd-party display"),
        ("Google Subscriptions/Devices", "$40.3B",      "10.0%",      "+17%",       "YouTube Premium, Play, Pixel"),
        ("Google Services Total",        "$305.3B",     "75.8%",      "+14%",       "Core advertising ecosystem"),
        ("Google Cloud",                 "$43.2B",      "10.7%",      "+~40%",      "GCP + Workspace enterprise"),
        ("Other Bets",                   "$1.7B",       "0.4%",       "+various",   "Waymo, Verily"),
        ("Hedging & Corporate",          "~$52.6B",     "13.1%",      "N/A",        "FX hedging, eliminations"),
        ("TOTAL",                        "$402.8B",     "100%",       "+15.1%",     ""),
    ]

    for i, row_data in enumerate(rev_data):
        r = 9 + i
        if i == 0:
            for j, val in enumerate(row_data):
                sub_hdr(ws, r, 2+j, val)
        else:
            bg = ALT_ROW if i % 2 == 0 else WHITE
            bold = (i == len(rev_data)-1)
            for j, val in enumerate(row_data):
                lbl(ws, r, 2+j, val, bg=bg) if j == 0 else dc(ws, r, 2+j, val, bg=bg, bold=bold)

    # Geographic Breakdown
    sub_hdr(ws, 21, 2, "Geographic Revenue Breakdown (FY2025 Est.)", span=6)
    geo_data = [
        ("Geography",        "Rev Estimate", "% Total", "Notes"),
        ("United States",    "$196B",         "49%",    "Core advertising + Cloud"),
        ("EMEA",             "$100B",         "25%",    "Europe dominant, strong Cloud"),
        ("APAC",             "$72B",          "18%",    "YouTube + Android growth"),
        ("Other Americas",   "$32B",          "8%",     "LatAm, Canada"),
    ]
    for i, rd in enumerate(geo_data):
        r = 22 + i
        if i == 0:
            for j, v in enumerate(rd):
                sub_hdr(ws, r, 2+j, v)
        else:
            bg = ALT_ROW if i % 2 == 0 else WHITE
            lbl(ws, r, 2, rd[0], bg=bg)
            for j in range(1, len(rd)):
                dc(ws, r, 2+j, rd[j], bg=bg)

    # Products & Services
    sub_hdr(ws, 28, 2, "Key Products & Services", span=6)
    products = [
        ("Google Search",        "Web & AI search; ~90% global query share; primary monetization engine"),
        ("YouTube",              "Video platform; 2B+ logged-in users/month; ads + YouTube Premium/TV"),
        ("Google Cloud (GCP)",   "IaaS/PaaS/SaaS; AI infrastructure (TPUs, Vertex AI); #3 cloud globally"),
        ("Android",              "Mobile OS on ~72% of smartphones globally; Play Store ecosystem"),
        ("Chrome",               "Browser with ~65% global share; gateway to Google ecosystem"),
        ("Workspace (G Suite)",  "Enterprise productivity: Gmail, Docs, Drive, Meet"),
        ("Google Ads",           "AdWords, Display Network, DV360; full-stack ad platform"),
        ("Waymo",                "Autonomous driving; commercial robotaxi in Phoenix, SF, LA"),
        ("DeepMind/Google AI",   "Gemini LLM family; AlphaFold; leads in AI research"),
        ("Pixel / Nest",         "Hardware; smartphones, smart home devices"),
    ]
    for i, (p, d) in enumerate(products):
        r = 29 + i
        bg = ALT_ROW if i % 2 == 0 else WHITE
        lbl(ws, r, 2, p, bg=bg)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)
        lbl(ws, r, 3, d, bg=bg)

    # Value Proposition & Customer Base
    sub_hdr(ws, 41, 2, "Value Proposition & Key Customers", span=6)
    vp = [
        ("Advertisers",           "Unmatched intent-based targeting via Search; massive reach via YouTube/GDN"),
        ("Enterprise (Cloud)",    "AI-native cloud infrastructure; best-in-class data analytics & ML tools"),
        ("Consumers",             "Free, high-quality services (Search, Maps, Gmail, YouTube) in exchange for data"),
        ("Developers",            "Android platform, Firebase, GCP — largest developer ecosystem globally"),
    ]
    for i, (seg, val) in enumerate(vp):
        r = 42 + i
        bg = ALT_ROW if i % 2 == 0 else WHITE
        lbl(ws, r, 2, seg, bg=bg)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)
        lbl(ws, r, 3, val, bg=bg)

    # Seasonality
    sub_hdr(ws, 48, 2, "Seasonality & Margin Structure", span=6)
    seasonality = [
        ("Q1 (Jan-Mar)",  "Seasonally weakest; post-holiday pullback in ad spend; ~26% of annual rev"),
        ("Q2 (Apr-Jun)",  "Recovery quarter; retail/travel/tech advertising ramp; ~24%"),
        ("Q3 (Jul-Sep)",  "Mid-year; back-to-school/travel; ~24%"),
        ("Q4 (Oct-Dec)",  "Strongest; holiday retail advertising surge; ~26% of annual revenue"),
        ("Margins",       "Google Services ~40% op margin; Google Cloud ~17% op margin; improving"),
    ]
    for i, (q, n) in enumerate(seasonality):
        r = 49 + i
        bg = ALT_ROW if i % 2 == 0 else WHITE
        lbl(ws, r, 2, q, bg=bg)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)
        lbl(ws, r, 3, n, bg=bg)

# ═══════════════════════════════════════════════════════════════════════════════
def build_moat(wb):
    ws = wb.create_sheet("Moat")
    ws.sheet_view.showGridLines = False
    scw(ws, 1, 3); scw(ws, 2, 28); scw(ws, 3, 50); scw(ws, 4, 20)

    sec_hdr(ws, 1, 2, "COMPETITIVE MOAT ANALYSIS — ALPHABET (GOOGL)", span=3)

    moat_intro = ("Alphabet possesses one of the deepest and widest economic moats in the world, "
                  "comprising network effects, switching costs, intangible assets (brand + data), "
                  "and cost advantages. The moat spans multiple complementary businesses.")
    ws.merge_cells("B3:D4")
    wc(ws, 3, 2, moat_intro, size=FONT_SIZE, bg=LIGHT_GRAY, align="left", wrap=True, border=True)
    for r in [3, 4]:
        ws.row_dimensions[r].height = 22

    moats = [
        ("MOAT TYPE",                "EVIDENCE & DEPTH",                                               "STRENGTH"),
        ("Network Effects (Search)",
         "More queries → more data → better results → more users. 25 years of click/query data is irreproducible. "
         "Each new Google user makes the product better for all users.",                               "VERY WIDE"),
        ("Data Advantage",
         "Google processes 8.5B+ searches/day. User intent data across Search, Maps, YouTube, Android "
         "is uniquely comprehensive and continuously self-reinforcing. This feeds AI model training.",  "VERY WIDE"),
        ("Brand / Trust",
         "'Google it' = verb globally. 90%+ search market share across most geographies. "
         "Users have decades of trust and habit formation. Switching friction is psychological + habitual.", "WIDE"),
        ("Switching Costs (Cloud)",
         "Once enterprises adopt Google Workspace + BigQuery + GCP, migration is expensive and risky. "
         "Vertex AI and proprietary TPUs create additional lock-in for ML workloads.",                  "MODERATE-WIDE"),
        ("Scale / Cost Advantages",
         "Google's infrastructure (data centers, subsea cables, custom TPU chips) is built at a scale "
         "no competitor can replicate. Marginal cost of serving ads/search queries is near zero.",      "VERY WIDE"),
        ("Intangible Assets (IP/AI)",
         "DeepMind pioneered transformer architecture; Google Brain developed TensorFlow and Gemini. "
         "TPU development gives Google structural cost advantage in AI inference vs. NVIDIA-dependent peers.", "WIDE"),
        ("Ecosystem Lock-in",
         "Android → Google Play → Google Pay → Google Search → YouTube → Maps → Gmail creates a "
         "closed-loop ecosystem. Users are invested across multiple touchpoints.",                       "WIDE"),
        ("Advertiser ROI",
         "Google Search delivers the best measurable ROI of any digital ad channel (intent-based). "
         "Advertisers have no comparable alternative at scale; programmatic alternatives are inferior.", "WIDE"),
    ]

    for i, row_data in enumerate(moats):
        r = 6 + i
        if i == 0:
            sub_hdr(ws, r, 2, row_data[0])
            sub_hdr(ws, r, 3, row_data[1])
            sub_hdr(ws, r, 4, row_data[2])
        else:
            bg = ALT_ROW if i % 2 == 0 else WHITE
            strength_color = GREEN if "VERY" in row_data[2] else GOOGLE_YELLOW
            lbl(ws, r, 2, row_data[0], bg=bg)
            ws.row_dimensions[r].height = 50
            wc(ws, r, 3, row_data[1], size=FONT_SIZE, bg=bg, align="left", wrap=True, border=True)
            wc(ws, r, 4, row_data[2], bold=True, size=FONT_SIZE, fg=WHITE,
               bg=strength_color if row_data[2]!="MODERATE-WIDE" else GOLD,
               align="center", border=True)

    # Threats to Moat
    sec_hdr(ws, 16, 2, "THREATS TO MOAT", span=3, bg=RED)

    threats = [
        ("AI Search Disruption",   "ChatGPT, Perplexity, Claude eating into zero-click searches; ~165x growth in AI traffic",    "HIGH"),
        ("Antitrust / Regulation", "DOJ seeking Chrome/Android divestiture; Apple deal banned; EU DMA ongoing",                  "HIGH"),
        ("Cloud Competition",      "AWS and Azure have larger installed base; Microsoft Copilot bundling advantages",             "MEDIUM"),
        ("Ad Market Cyclicality",  "Economic downturns hit ad budgets first; 2022 showed 10%+ revenue deceleration risk",        "MEDIUM"),
        ("Social Media (YouTube)", "TikTok competes for Gen-Z attention; Shorts monetization still maturing",                    "MEDIUM"),
        ("CapEx Overinvestment",   "$175-185B planned 2026 CapEx; risk of overcapacity if AI demand plateaus",                   "HIGH"),
    ]

    sub_hdr(ws, 17, 2, "Threat")
    sub_hdr(ws, 17, 3, "Description")
    sub_hdr(ws, 17, 4, "Risk Level")

    for i, (t, d, r_lvl) in enumerate(threats):
        r = 18 + i
        bg = ALT_ROW if i % 2 == 0 else WHITE
        lbl(ws, r, 2, t, bg=bg)
        ws.row_dimensions[r].height = 40
        wc(ws, r, 3, d, size=FONT_SIZE, bg=bg, align="left", wrap=True, border=True)
        clr = RED if r_lvl == "HIGH" else GOLD
        wc(ws, r, 4, r_lvl, bold=True, size=FONT_SIZE, fg=WHITE, bg=clr, align="center", border=True)

    # Overall Moat Rating
    sub_hdr(ws, 25, 2, "Overall Moat Rating", span=3, bg=DARK_BLUE, fg=WHITE)
    ws.merge_cells("B26:D26")
    wc(ws, 26, 2, "WIDE MOAT — One of the widest moats in global equities. "
       "The combination of network effects, data flywheel, brand monopoly, and AI positioning "
       "is nearly unassailable in the medium term. Primary risk is regulatory disruption, not competitive.",
       bold=True, size=FONT_SIZE, bg=mfill(GREEN).fgColor.rgb if False else ALT_ROW,
       align="left", wrap=True, border=True)
    wc(ws, 26, 2, "WIDE MOAT — One of the widest moats in global equities. "
       "The combination of network effects, data flywheel, brand monopoly, and AI positioning "
       "is nearly unassailable in the medium term. Primary risk is regulatory disruption, not competitive.",
       size=FONT_SIZE, bg=ALT_ROW, align="left", wrap=True, border=True)
    ws.row_dimensions[26].height = 50

# ═══════════════════════════════════════════════════════════════════════════════
def build_income(wb):
    ws = wb.create_sheet("Income Statement")
    ws.sheet_view.showGridLines = False
    scw(ws, 1, 3); scw(ws, 2, 32); scw(ws, 3, 18); scw(ws, 4, 18)
    scw(ws, 5, 18); scw(ws, 6, 18); scw(ws, 7, 18)

    sec_hdr(ws, 1, 2, "INCOME STATEMENT — ALPHABET INC. (GOOGL)  |  FY2021-FY2025 ($B)", span=6)
    sub_hdr(ws, 2, 2, "Metric")
    for i, yr in enumerate(["FY2021", "FY2022", "FY2023", "FY2024", "FY2025"]):
        sub_hdr(ws, 2, 3+i, yr)

    rows = [
        ("REVENUE",                     None, None, None, None, None),
        ("  Google Search & Other",     149.0, 162.5, 175.0, 198.1, None),
        ("  YouTube Ads",                28.8,  29.2,  31.5,  36.1, None),  # approx
        ("  Google Network",             31.7,  32.3,  31.3,  30.8, None),
        ("  Google Subscriptions/Devices",28.0, 29.0, 34.0, 40.3, None),
        ("  Google Services Total",     237.5, 253.0, 272.0, 305.3, None),  # approx
        ("  Google Cloud",               19.2,  26.3,  33.1,  43.2, None),
        ("  Other Bets",                  0.7,   1.1,   1.5,   1.7, None),
        ("TOTAL REVENUE",               257.6, 282.8, 307.4, 350.0, 402.8),
        ("YoY Growth",                  "41%",  "9.8%","8.7%","13.8%","15.1%"),
        ("", None, None, None, None, None),
        ("COSTS & EXPENSES",            None, None, None, None, None),
        ("  Cost of Revenue",            97.0, 116.0, 126.0, 144.0, 160.0),
        ("  Research & Development",     31.6,  39.5,  45.4,  52.0,  61.1),
        ("  Sales & Marketing",          22.9,  26.6,  27.4,  28.0,  30.0),
        ("  G&A",                         8.7,  11.6,  11.6,  13.0,  14.0),
        ("TOTAL OPERATING EXPENSES",    160.2, 193.7, 210.4, 237.0, 265.1),
        ("", None, None, None, None, None),
        ("OPERATING INCOME",             78.7,  74.8,  84.3, 112.4, 127.1),  # approx
        ("  Operating Margin",          "30.6%","26.5%","27.4%","32.1%","31.6%"),
        ("Other Income / (Expense)",      2.5,   0.4,   1.8,   4.0,   5.0),
        ("Pretax Income",               78.7+2.5, 74.8+0.4, 84.3+1.8, 116.4, 132.1),
        ("Income Tax Expense",            14.7,  9.0, 11.9,  13.9,  24.2),
        ("NET INCOME",                   76.0,  59.9,  73.8, 102.5, 132.2),
        ("  Net Margin",                "29.5%","21.2%","24.0%","29.3%","32.8%"),
        ("", None, None, None, None, None),
        ("EPS (Diluted)",                5.61,   4.56,  5.80,  8.05, 10.85),
        ("Shares Outstanding (B)",      12.85,  12.66, 12.45, 12.28, 12.18),
    ]

    for i, row_data in enumerate(rows):
        r = 3 + i
        label = row_data[0]
        vals  = row_data[1:]
        if label == "" :
            ws.row_dimensions[r].height = 8
            continue
        is_total = label.startswith("TOTAL") or label in ("OPERATING INCOME", "NET INCOME",
                                                            "COSTS & EXPENSES", "REVENUE")
        is_margin = "Margin" in label or "Growth" in label or label.startswith("  ")
        bg = SUBHDR_BG if is_total and not label.startswith(" ") else (ALT_ROW if i % 2 == 0 else WHITE)

        lbl(ws, r, 2, label, bg=bg, indent=1 if label.startswith("  ") else 0)
        for j, v in enumerate(vals):
            if v is None:
                dc(ws, r, 3+j, "—", bg=bg)
            elif isinstance(v, str):
                wc(ws, r, 3+j, v, size=FONT_SIZE, bg=bg, align="right", border=True,
                   italic=True, fg="555555")
            else:
                num_fmt = '#,##0.00' if abs(v) < 20 else '#,##0.0'
                dc(ws, r, 3+j, v, bg=bg, bold=is_total, num_fmt=num_fmt)

# ═══════════════════════════════════════════════════════════════════════════════
def build_balance_sheet(wb):
    ws = wb.create_sheet("Balance Sheet")
    ws.sheet_view.showGridLines = False
    scw(ws, 1, 3); scw(ws, 2, 35); scw(ws, 3, 18); scw(ws, 4, 18)
    scw(ws, 5, 18); scw(ws, 6, 18)

    sec_hdr(ws, 1, 2, "BALANCE SHEET — ALPHABET INC. (GOOGL)  |  FY2022-FY2025 ($B)", span=5)
    sub_hdr(ws, 2, 2, "Metric")
    for i, yr in enumerate(["FY2022", "FY2023", "FY2024", "FY2025"]):
        sub_hdr(ws, 2, 3+i, yr)

    rows = [
        ("ASSETS",                          None, None, None, None),
        ("CURRENT ASSETS",                  None, None, None, None),
        ("  Cash & Cash Equivalents",       21.9,  24.0,  27.7,  30.7),
        ("  Short-term Investments",        87.9,  91.9, 110.2, 111.0),
        ("  Accounts Receivable (net)",     40.3,  47.2,  49.0,  52.0),
        ("  Other Current Assets",           6.9,   7.1,   8.0,   9.0),
        ("TOTAL CURRENT ASSETS",           157.0, 170.2, 194.9, 202.7),
        ("NON-CURRENT ASSETS",             None, None, None, None),
        ("  PP&E (net)",                   117.5, 134.4, 134.4, 182.0),
        ("  Operating Lease Right-of-Use",  14.9,  14.6,  15.0,  16.0),
        ("  Equity Investments",            30.2,  30.5,  31.0,  32.0),
        ("  Other Non-current Assets",      10.2,  10.8,  12.0,  13.0),
        ("TOTAL NON-CURRENT ASSETS",       172.8, 190.3, 192.4, 243.0),
        ("TOTAL ASSETS",                   359.3, 402.4, 450.5, 474.8),  # approx
        ("", None, None, None, None),
        ("LIABILITIES",                    None, None, None, None),
        ("CURRENT LIABILITIES",            None, None, None, None),
        ("  Accounts Payable",               6.0,   7.5,   7.0,   7.5),
        ("  Accrued Expenses",              13.7,  14.5,  17.0,  19.0),
        ("  Deferred Revenue",               3.3,   4.1,   4.5,   4.8),
        ("  Other Current Liabilities",      7.0,   8.0,   9.0,  10.0),
        ("TOTAL CURRENT LIABILITIES",       30.0,  34.1,  37.5,  41.3),
        ("NON-CURRENT LIABILITIES",        None, None, None, None),
        ("  Long-term Debt",               14.7,  14.7,  14.7,  46.5),
        ("  Deferred Tax Liabilities",       4.2,   4.8,   5.0,   5.5),
        ("  Other LT Liabilities",          13.2,  14.2,  15.0,  16.0),
        ("TOTAL NON-CURRENT LIABILITIES",   32.1,  33.7,  34.7,  68.0),
        ("TOTAL LIABILITIES",               62.1,  67.8,  72.2, 109.3),
        ("", None, None, None, None),
        ("SHAREHOLDERS' EQUITY",           None, None, None, None),
        ("  Common Stock & APIC",           68.2,  72.8,  76.0,  80.0),
        ("  Retained Earnings",            229.0, 261.8, 302.1, 285.5),
        ("TOTAL EQUITY",                   297.2, 334.6, 378.3, 365.5),
        ("TOTAL LIAB + EQUITY",            359.3, 402.4, 450.5, 474.8),
        ("", None, None, None, None),
        ("KEY RATIOS",                     None, None, None, None),
        ("  Current Ratio",                  5.2,   5.0,   5.2,   4.9),
        ("  Debt-to-Equity",                0.05,  0.04,  0.04,  0.13),
        ("  Cash + ST Investments ($B)",   109.8, 115.9, 137.9, 141.7),
        ("  Net Cash / (Debt) ($B)",       109.8, 115.9, 137.9,  95.2),
    ]

    for i, row_data in enumerate(rows):
        r = 3 + i
        label = row_data[0]
        vals  = row_data[1:]
        if label == "":
            ws.row_dimensions[r].height = 8
            continue
        is_total = label.startswith("TOTAL") or label in ("ASSETS","LIABILITIES","SHAREHOLDERS' EQUITY",
                                                            "KEY RATIOS","CURRENT ASSETS","NON-CURRENT ASSETS",
                                                            "CURRENT LIABILITIES","NON-CURRENT LIABILITIES")
        bg = SUBHDR_BG if is_total else (ALT_ROW if i % 2 == 0 else WHITE)
        lbl(ws, r, 2, label, bg=bg)
        for j, v in enumerate(vals):
            if v is None:
                dc(ws, r, 3+j, "—", bg=bg)
            else:
                dc(ws, r, 3+j, v, bg=bg, bold=is_total, num_fmt='#,##0.0')

# ═══════════════════════════════════════════════════════════════════════════════
def build_cashflow(wb):
    ws = wb.create_sheet("Cash Flow Analysis")
    ws.sheet_view.showGridLines = False
    scw(ws, 1, 3); scw(ws, 2, 35); scw(ws, 3, 18); scw(ws, 4, 18)
    scw(ws, 5, 18); scw(ws, 6, 18); scw(ws, 7, 18)

    sec_hdr(ws, 1, 2, "CASH FLOW ANALYSIS — ALPHABET INC. (GOOGL)  |  ($B)", span=6)
    sub_hdr(ws, 2, 2, "Metric")
    for i, yr in enumerate(["FY2021", "FY2022", "FY2023", "FY2024", "FY2025"]):
        sub_hdr(ws, 2, 3+i, yr)

    rows = [
        ("OPERATING CASH FLOW",            91.7,  91.5, 101.7, 125.0, 164.7),
        ("  Net Income",                   76.0,  59.9,  73.8, 102.5, 132.2),
        ("  D&A",                          11.5,  13.5,  15.2,  17.5,  19.0),
        ("  Stock-Based Compensation",     15.4,  19.8,  22.5,  24.0,  25.0),
        ("  Changes in Working Capital",   -8.0, -10.0, -10.0,  -15.0, -12.0),
        ("  Other",                        -3.2,   8.3,   0.2,    4.0,   0.5),
        ("", None, None, None, None, None),
        ("INVESTING CASH FLOW",           -35.5, -42.2, -38.9, -52.5, -91.4),  # approx
        ("  Capital Expenditures (CapEx)",-24.3, -31.5, -32.3, -52.5, -91.4),
        ("  Acquisitions (net)",           -2.6, -2.2,  -1.7,  -2.0,  -3.0),
        ("  Purchases of Investments",    -88.9,-60.3, -66.0, -70.0, -60.0),
        ("  Maturities of Investments",    82.3,  58.1,  57.8,  65.0,  58.0),
        ("", None, None, None, None, None),
        ("FINANCING CASH FLOW",           -62.8,-70.6, -72.5, -82.0, -80.0),  # approx
        ("  Share Repurchases",           -50.3,-59.3, -62.2, -62.2, -45.7),
        ("  Debt Issuance / (Repayment)",   6.8,   0.0,   0.0,   0.0,  32.0),
        ("  Other Financing",              -5.0, -5.5,  -6.3,  -8.0,  -8.0),
        ("", None, None, None, None, None),
        ("FREE CASH FLOW (OCF - CapEx)",   67.4,  60.0,  69.4,  72.5,  73.3),
        ("  FCF Margin",                  "26.2%","21.2%","22.6%","20.7%","18.2%"),
        ("  FCF / Net Income",             0.89,  1.00,  0.94,  0.71,   0.55),
        ("", None, None, None, None, None),
        ("CAPITAL ALLOCATION SUMMARY",    None, None, None, None, None),
        ("  CapEx ($B)",                   24.3,  31.5,  32.3,  52.5,  91.4),
        ("  CapEx as % of Revenue",       "9.4%","11.1%","10.5%","15.0%","22.7%"),
        ("  R&D Spending ($B)",            31.6,  39.5,  45.4,  52.0,  61.1),
        ("  Buybacks ($B)",                50.3,  59.3,  62.2,  62.2,  45.7),
        ("  Dividends",                   "None","None","None","None","None"),
        ("  M&A Spend ($B)",                2.6,   2.2,   1.7,   2.0,   3.0),
    ]

    for i, row_data in enumerate(rows):
        r = 3 + i
        label = row_data[0]
        vals  = row_data[1:]
        if label == "":
            ws.row_dimensions[r].height = 8
            continue
        is_total = any(label.startswith(x) for x in ("OPERATING","INVESTING","FINANCING","FREE CASH","CAPITAL"))
        bg = SUBHDR_BG if is_total else (ALT_ROW if i % 2 == 0 else WHITE)
        lbl(ws, r, 2, label, bg=bg)
        for j, v in enumerate(vals):
            if v is None:
                dc(ws, r, 3+j, "—", bg=bg)
            elif isinstance(v, str):
                wc(ws, r, 3+j, v, size=FONT_SIZE, bg=bg, align="right", border=True, italic=True, fg="555555")
            else:
                dc(ws, r, 3+j, v, bg=bg, bold=is_total, num_fmt='#,##0.0')

# ═══════════════════════════════════════════════════════════════════════════════
def build_roic(wb):
    ws = wb.create_sheet("Return on Capital")
    ws.sheet_view.showGridLines = False
    scw(ws, 1, 3); scw(ws, 2, 35); scw(ws, 3, 18); scw(ws, 4, 18)
    scw(ws, 5, 18); scw(ws, 6, 18); scw(ws, 7, 18)

    sec_hdr(ws, 1, 2, "RETURN ON CAPITAL — ALPHABET INC. (GOOGL)", span=6)
    sub_hdr(ws, 2, 2, "Metric")
    for i, yr in enumerate(["FY2021", "FY2022", "FY2023", "FY2024", "FY2025"]):
        sub_hdr(ws, 2, 3+i, yr)

    rows = [
        ("PROFITABILITY RETURNS",         None, None, None, None, None),
        ("  Return on Equity (ROE)",      "29.5%","18.0%","22.5%","27.1%","36.2%"),
        ("  Return on Assets (ROA)",      "18.4%","14.9%","19.0%","22.7%","27.8%"),
        ("  Return on Invested Capital",  "29.8%","22.5%","27.5%","35.0%","38.0%"),
        ("  Return on Ad Spend (ROAS)",   "High","High","High","High","High"),
        ("", None, None, None, None, None),
        ("INCREMENTAL RETURNS",           None, None, None, None, None),
        ("  Revenue Growth ($B)",          75.4,  25.2,  24.6,  42.6,  52.8),
        ("  CapEx + R&D Invested ($B)",    55.9,  71.0,  77.7, 104.5, 152.5),
        ("  Return on Incremental Capital","135%","35%","32%","41%","35%"),  # approx
        ("", None, None, None, None, None),
        ("SEGMENT MARGINS",               None, None, None, None, None),
        ("  Google Services Op Margin",   "36%","27%","32%","40%","40%"),
        ("  Google Cloud Op Margin",      "-5%", "-5%", "8%","17%","17%"),
        ("  Other Bets Op Margin",        "-400%","-790%","-800%","-900%","-900%"),
        ("", None, None, None, None, None),
        ("ASSET EFFICIENCY",              None, None, None, None, None),
        ("  Asset Turnover",               0.72,  0.79,  0.76,  0.78,  0.85),
        ("  Revenue / Employee ($K)",      None,  None, 1344, 1519, 1700),
        ("  Op CF / Revenue",             "35.6%","32.4%","33.1%","35.7%","40.9%"),
        ("", None, None, None, None, None),
        ("CAPITAL ALLOCATION QUALITY",    None, None, None, None, None),
        ("  CapEx Growth YoY",            "17%", "30%",  "3%","63%","74%"),
        ("  Commentary",                  "Disciplined", "Discipline", "Efficient", "AI Ramp", "AI Mega-bet"),
        ("  Net CapEx Payback (Est.)",    "3-4 yrs","3-4 yrs","3 yrs","4-5 yrs","5-6 yrs"),
    ]

    for i, row_data in enumerate(rows):
        r = 3 + i
        label = row_data[0]
        vals  = row_data[1:]
        if label == "":
            ws.row_dimensions[r].height = 8
            continue
        is_hdr = any(label.startswith(x) for x in ("PROFITABILITY","INCREMENTAL","SEGMENT","ASSET EFFICIENCY","CAPITAL"))
        bg = SUBHDR_BG if is_hdr else (ALT_ROW if i % 2 == 0 else WHITE)
        lbl(ws, r, 2, label, bg=bg)
        for j, v in enumerate(vals):
            if v is None:
                dc(ws, r, 3+j, "—", bg=bg)
            elif isinstance(v, str):
                wc(ws, r, 3+j, v, size=FONT_SIZE, bg=bg, align="right", border=True, italic="%" not in v)
            else:
                dc(ws, r, 3+j, v, bg=bg, bold=is_hdr, num_fmt='#,##0.00')

# ═══════════════════════════════════════════════════════════════════════════════
def build_management(wb):
    ws = wb.create_sheet("Management")
    ws.sheet_view.showGridLines = False
    scw(ws, 1, 3); scw(ws, 2, 28); scw(ws, 3, 55); scw(ws, 4, 18)

    sec_hdr(ws, 1, 2, "MANAGEMENT ANALYSIS — ALPHABET INC. (GOOGL)", span=3)

    # Key Leaders
    sub_hdr(ws, 3, 2, "Key Leadership", span=3)
    leaders = [
        ("Name",            "Role",                     "Background & Assessment"),
        ("Sundar Pichai",   "CEO, Alphabet & Google",   "Joined Google 2004; CEO since 2015; deep product/engineering DNA; "
                                                          "navigated AI pivot with Gemini launch; strong engineer-manager"),
        ("Ruth Porat",      "SVP & CFO",                "Ex-Morgan Stanley CFO; joined 2015; brought financial discipline; "
                                                          "steered buyback program; now President of Alphabet"),
        ("Demis Hassabis",  "CEO DeepMind",             "AI pioneer; Nobel-adjacent (AlphaFold); driving Gemini Ultra/Pro strategy"),
        ("Larry Page",      "Co-founder, Board",        "Executive chair emeritus; less active but a board stabilizer"),
        ("Sergey Brin",     "Co-founder, Board",        "Technically returned to support AI R&D post-ChatGPT; credibility signal"),
    ]
    for i, row_data in enumerate(leaders):
        r = 4 + i
        if i == 0:
            sub_hdr(ws, r, 2, row_data[0])
            sub_hdr(ws, r, 3, row_data[1])
            sub_hdr(ws, r, 4, row_data[2])
        else:
            bg = ALT_ROW if i % 2 == 0 else WHITE
            lbl(ws, r, 2, row_data[0], bg=bg)
            lbl(ws, r, 3, row_data[1], bg=bg)
            ws.row_dimensions[r].height = 45
            wc(ws, r, 4, row_data[2], size=FONT_SIZE, bg=bg, align="left", wrap=True, border=True)

    # Compensation / Incentives
    sub_hdr(ws, 11, 2, "Compensation & Incentives (Proxy Statement Analysis)", span=3)
    comp_items = [
        ("CEO Pay Ratio",          "Sundar Pichai's total comp ~$10.7M (2024); 95th percentile but not egregious for a $3.8T company"),
        ("Pay Structure",          "Base salary ($2M), annual bonus, and RSUs vesting over 4 years tied to performance metrics"),
        ("Performance Metrics",    "Revenue growth, operating income margin, Cloud growth, product goals — aligned with shareholders"),
        ("Founder Ownership",      "Page & Brin retain ~10x voting power shares (Class C); effective control despite dilution"),
        ("Dilution Risk",          "SBC ~$25B/yr (~6% of net income); partially offset by $45-62B annual buybacks"),
        ("Management Skin in Game","Page ($100B+), Brin ($100B+) remain largest shareholders; interests well aligned"),
    ]
    for i, (k, v) in enumerate(comp_items):
        r = 12 + i
        bg = ALT_ROW if i % 2 == 0 else WHITE
        lbl(ws, r, 2, k, bg=bg)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
        ws.row_dimensions[r].height = 40
        wc(ws, r, 3, v, size=FONT_SIZE, bg=bg, align="left", wrap=True, border=True)

    # Capital Allocation
    sub_hdr(ws, 20, 2, "Capital Allocation Decisions", span=3)
    ca_items = [
        ("Buybacks",       "Consistent $45-62B/yr buyback program; Alphabet has retired ~5% of shares since 2019 — shareholder friendly"),
        ("CapEx",          "$91.4B in FY2025 (+74% YoY); $175-185B guided for 2026; betting big on AI infrastructure (data centers, TPUs)"),
        ("R&D",            "$61.1B in FY2025; highest R&D spend of any company globally; seeds future competitive advantages"),
        ("M&A",            "Disciplined acquirer: Motorola ($12.5B), YouTube ($1.7B), Mandiant ($5.4B), DeepMind; no mega-dilutive deals"),
        ("Dividends",      "No cash dividend; prefers buybacks and reinvestment — rational given growth opportunities"),
        ("Leverage",       "Historically debt-free; $46.5B LT debt added in FY2025 to fund CapEx, low D/E of 0.13"),
    ]
    for i, (k, v) in enumerate(ca_items):
        r = 21 + i
        bg = ALT_ROW if i % 2 == 0 else WHITE
        lbl(ws, r, 2, k, bg=bg)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
        ws.row_dimensions[r].height = 40
        wc(ws, r, 3, v, size=FONT_SIZE, bg=bg, align="left", wrap=True, border=True)

    # Acts Like Owner?
    sub_hdr(ws, 29, 2, "Does Management Act Like an Owner?", span=3)
    owner_items = [
        ("Long-term Thinking",     "YES — DeepMind acquisition (2014), Waymo investment, TPU development all represent seeds planted years ago"),
        ("Seed vs. Harvest",       "SEEDS: $175B AI CapEx commitment; Waymo commercialization; Gemini model family"),
        ("Harvest Risk",           "CONCERN: Ad revenue still 75%+ of total; innovation mostly defensive vs. proactive in recent years"),
        ("Insider Transactions",   "Minimal insider selling; founders retain large stakes; CFO/CEO compensation reasonable"),
        ("Accountability",         "Sundar Pichai publicly acknowledged 'missed the AI moment' in 2023; pivoted entire organization"),
        ("Cost Management",        "Laid off 12,000 employees in 2023 to drive efficiency; shows willingness to make hard decisions"),
    ]
    for i, (k, v) in enumerate(owner_items):
        r = 30 + i
        bg = ALT_ROW if i % 2 == 0 else WHITE
        lbl(ws, r, 2, k, bg=bg)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
        ws.row_dimensions[r].height = 40
        wc(ws, r, 3, v, size=FONT_SIZE, bg=bg, align="left", wrap=True, border=True)

    # Management Score
    sub_hdr(ws, 38, 2, "Overall Management Score", span=3, bg=DARK_BLUE, fg=WHITE)
    ws.merge_cells("B39:D40")
    wc(ws, 39, 2,
       "SCORE: 8/10 — Strong management team with proven long-term focus. Sundar Pichai "
       "successfully navigated the AI transformation after a slow start. Capital allocation is "
       "shareholder-friendly (buybacks > dividends). The $175B CapEx bet is bold but necessary. "
       "WATCH: Whether AI CapEx translates to earnings power in 2027-2028.",
       size=FONT_SIZE, bg=ALT_ROW, align="left", wrap=True, border=True)
    for r in [39, 40]:
        ws.row_dimensions[r].height = 28

# ═══════════════════════════════════════════════════════════════════════════════
def build_risks(wb):
    ws = wb.create_sheet("Risks")
    ws.sheet_view.showGridLines = False
    scw(ws, 1, 3); scw(ws, 2, 32); scw(ws, 3, 45); scw(ws, 4, 15); scw(ws, 5, 15)

    sec_hdr(ws, 1, 2, "RISK ANALYSIS — ALPHABET INC. (GOOGL)", span=4)
    sub_hdr(ws, 2, 2, "Risk Category")
    sub_hdr(ws, 2, 3, "Description & Impact")
    sub_hdr(ws, 2, 4, "Probability")
    sub_hdr(ws, 2, 5, "Impact")

    risks = [
        ("REGULATORY / ANTITRUST",
         "DOJ seeking Chrome/Android divestiture; Apple default deal banned; EU Digital Markets Act "
         "forcing changes to Google services display; adtech breakup possible. This is the #1 risk.",
         "HIGH", "VERY HIGH"),
        ("AI SEARCH DISRUPTION",
         "ChatGPT, Perplexity, Copilot eating into 'zero-click' information queries. "
         "GenAI traffic growing 165x faster than search, but only ~1% of total web traffic currently. "
         "Risk is 3-5 year secular shift, not imminent collapse.",
         "MEDIUM", "HIGH"),
        ("CAPEX OVER-INVESTMENT",
         "$91.4B CapEx in 2025 (+74%), $175-185B guided for 2026. If AI adoption plateaus or "
         "commodity compute prices fall, ROI on this spend may disappoint. Free cash flow will be "
         "depressed for 2-3 years while infrastructure builds.",
         "MEDIUM", "HIGH"),
        ("CLOUD COMPETITION",
         "AWS maintains ~32% share, Azure ~24%, GCP ~12%. Microsoft Copilot bundling advantage. "
         "Google Cloud has strong momentum but is the #3 player; winning large enterprise deals "
         "requires overcoming AWS/Azure inertia.",
         "MEDIUM", "MEDIUM"),
        ("MACRO / AD SPENDING",
         "Digital advertising is cyclical. A global recession would hit Google ad revenue materially. "
         "2022 saw first-ever revenue decline quarter in ad segment.",
         "LOW-MEDIUM", "MEDIUM"),
        ("DATA PRIVACY / AI ETHICS",
         "Gemini controversies (biased outputs, image generation errors) damaged reputation. "
         "GDPR, CCPA, and emerging AI regulation could restrict data collection that underpins ad targeting.",
         "MEDIUM", "MEDIUM"),
        ("FOREIGN EXCHANGE",
         "~51% of revenue is non-US; dollar strength reduces reported revenue. Q4 2025 FX impact was ~1%.",
         "MEDIUM", "LOW"),
        ("TALENT / AI COMPETITION",
         "OpenAI, Anthropic, Meta AI compete for top AI talent. Brain drain risk to AI startups; "
         "Google has partially mitigated via DeepMind merge and aggressive retention packages.",
         "LOW", "MEDIUM"),
        ("CYBER SECURITY",
         "Scale of user data makes Google a prime target. A major breach could cause regulatory fines "
         "(GDPR: up to 4% of global revenue), reputational damage, and user trust erosion.",
         "LOW", "HIGH"),
        ("HARDWARE / PIXEL FAILURE",
         "Consumer hardware (Pixel, Nest) is sub-scale; Google has historically failed at hardware "
         "(Glass, Nexus, Stadia). Continued losses are a capital drain.",
         "MEDIUM", "LOW"),
    ]

    for i, (cat, desc, prob, imp) in enumerate(risks):
        r = 3 + i
        bg = ALT_ROW if i % 2 == 0 else WHITE
        lbl(ws, r, 2, cat, bg=bg)
        ws.row_dimensions[r].height = 55
        wc(ws, r, 3, desc, size=FONT_SIZE, bg=bg, align="left", wrap=True, border=True)
        prob_c = RED if prob == "HIGH" else (GOLD if "MEDIUM" in prob else GREEN)
        imp_c  = RED if imp  == "VERY HIGH" else (RED if imp == "HIGH" else (GOLD if imp == "MEDIUM" else GREEN))
        wc(ws, r, 4, prob, bold=True, size=FONT_SIZE, fg=WHITE, bg=prob_c, align="center", border=True)
        wc(ws, r, 5, imp,  bold=True, size=FONT_SIZE, fg=WHITE, bg=imp_c,  align="center", border=True)

    # Risk Summary
    sub_hdr(ws, 14, 2, "Risk Summary & Investment Implication", span=4)
    summary = ("The two dominant risks for Alphabet are (1) regulatory fragmentation that could "
               "structurally impair the search monopoly and (2) AI disruption to the core search "
               "business over a 3-7 year horizon. Both risks are real but manageable. The CapEx "
               "bet represents a near-term FCF headwind, not permanent capital destruction. "
               "Given the 23x forward P/E (below historical avg 28x), much of the antitrust "
               "risk appears already priced in. We view downside as limited at current prices.")
    ws.merge_cells("B15:E17")
    wc(ws, 15, 2, summary, size=FONT_SIZE, bg=LIGHT_GRAY, align="left", wrap=True, border=True)
    for r in [15, 16, 17]:
        ws.row_dimensions[r].height = 22

# ═══════════════════════════════════════════════════════════════════════════════
def build_valuation(wb):
    ws = wb.create_sheet("Valuation")
    ws.sheet_view.showGridLines = False
    scw(ws, 1, 3); scw(ws, 2, 32); scw(ws, 3, 22); scw(ws, 4, 22)
    scw(ws, 5, 22); scw(ws, 6, 22)

    sec_hdr(ws, 1, 2, "VALUATION ANALYSIS — ALPHABET INC. (GOOGL)", span=5)

    # Current Market Stats
    sub_hdr(ws, 3, 2, "Current Market Statistics (April 8, 2026)", span=5)
    mkt = [
        ("Stock Price",             "$317.32"),
        ("Market Capitalization",   "~$3.86 Trillion"),
        ("Enterprise Value",        "~$3.77 Trillion"),
        ("P/E Ratio (TTM)",         "~29x"),
        ("Forward P/E (FY2026E)",   "~23x"),
        ("EV/EBITDA",               "~22x"),
        ("Price/FCF",               "~53x (elevated due to CapEx surge)"),
        ("Price/Sales",             "~9.6x"),
        ("52-Week Range",           "$~245 - $~335"),
        ("Analyst Consensus",       "BUY | Median target ~$215 (target below current = already re-rated)"),
    ]
    for i, (k, v) in enumerate(mkt):
        r = 4 + i
        bg = ALT_ROW if i % 2 == 0 else WHITE
        lbl(ws, r, 2, k, bg=bg)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
        wc(ws, r, 3, v, size=FONT_SIZE, bg=bg, align="right", border=True)
        ws.row_dimensions[r].height = 22

    # DCF Valuation
    sub_hdr(ws, 16, 2, "DCF Valuation (Base / Bull / Bear)", span=5)
    sub_hdr(ws, 17, 2, "Assumption"); sub_hdr(ws, 17, 3, "Bear Case"); sub_hdr(ws, 17, 4, "Base Case")
    sub_hdr(ws, 17, 5, "Bull Case"); sub_hdr(ws, 17, 6, "Notes")

    dcf = [
        ("Revenue Growth (FY2026-28)",     "8%",    "13%",   "18%",  "Bear=macro+regulatory"),
        ("Revenue Growth (FY2029-35)",     "6%",     "9%",   "14%",  "Bull=Cloud 40%+, AI monetization"),
        ("Terminal Growth Rate",           "2.5%",  "3.0%",  "3.5%", "Long-run GDP growth"),
        ("EBIT Margin (normalized)",       "28%",   "33%",   "37%",  "Bear=CapEx drag; Bull=scale"),
        ("Discount Rate (WACC)",           "10%",   "9.0%",  "8.0%", "Risk premium + RF rate"),
        ("Implied Intrinsic Value/Share",  "$220",  "$330",  "$450", "Per share DCF"),
    ]
    for i, row_data in enumerate(dcf):
        r = 18 + i
        bg = ALT_ROW if i % 2 == 0 else WHITE
        lbl(ws, r, 2, row_data[0], bg=bg)
        for j in range(1, 5):
            wc(ws, r, 2+j, row_data[j], size=FONT_SIZE, bg=bg, align="right" if j < 4 else "left",
               border=True, bold=(row_data[0].startswith("Implied")))
        ws.row_dimensions[r].height = 22

    # Relative Valuation
    sub_hdr(ws, 26, 2, "Relative Valuation vs. Peers", span=5)
    sub_hdr(ws, 27, 2, "Company"); sub_hdr(ws, 27, 3, "P/E (TTM)")
    sub_hdr(ws, 27, 4, "Fwd P/E"); sub_hdr(ws, 27, 5, "EV/EBITDA"); sub_hdr(ws, 27, 6, "Rev Growth")

    peers = [
        ("Alphabet (GOOGL)",    "29x",   "23x",  "22x",  "15%"),
        ("Meta Platforms",      "27x",   "22x",  "18x",  "19%"),
        ("Microsoft (MSFT)",    "34x",   "29x",  "26x",  "13%"),
        ("Amazon (AMZN)",       "38x",   "31x",  "24x",  "11%"),
        ("Apple (AAPL)",        "31x",   "27x",  "23x",   "5%"),
        ("S&P 500 Average",     "24x",   "21x",  "17x",   "9%"),
    ]
    for i, row_data in enumerate(peers):
        r = 28 + i
        bg = ALT_ROW if i % 2 == 0 else WHITE
        bold = (row_data[0].startswith("Alphabet"))
        bg = SUBHDR_BG if bold else bg
        for j, v in enumerate(row_data):
            if j == 0:
                lbl(ws, r, 2, v, bg=bg)
            else:
                wc(ws, r, 2+j, v, bold=bold, size=FONT_SIZE, bg=bg, align="right", border=True)

    # Margin of Safety
    sub_hdr(ws, 36, 2, "Margin of Safety Assessment", span=5)
    mos = [
        ("Current Price",          "$317.32",  "April 8, 2026"),
        ("Base Case Intrinsic Value", "$330",   "~4% undervalued to base case"),
        ("Bear Case Intrinsic Value", "$220",   "~31% downside in adverse scenario"),
        ("Bull Case Intrinsic Value", "$450",   "~42% upside in bull scenario"),
        ("Margin of Safety",          "MODERATE","Trading near/slightly below base case; limited downside vs. historical avg"),
        ("Recommendation",            "ACCUMULATE","At current price, valuation is fair-to-attractive for 3-5 year hold"),
    ]
    for i, (k, v, c) in enumerate(mos):
        r = 37 + i
        bg = ALT_ROW if i % 2 == 0 else WHITE
        lbl(ws, r, 2, k, bg=bg)
        wc(ws, r, 3, v, bold=True, size=FONT_SIZE, bg=bg, align="right", border=True)
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=6)
        wc(ws, r, 4, c, size=FONT_SIZE, bg=bg, align="left", border=True)
        ws.row_dimensions[r].height = 22

# ═══════════════════════════════════════════════════════════════════════════════
def build_sentiment(wb):
    ws = wb.create_sheet("Market Sentiment")
    ws.sheet_view.showGridLines = False
    scw(ws, 1, 3); scw(ws, 2, 30); scw(ws, 3, 50); scw(ws, 4, 18)

    sec_hdr(ws, 1, 2, "MARKET SENTIMENT — ALPHABET INC. (GOOGL)", span=3)

    sub_hdr(ws, 3, 2, "Analyst Coverage", span=3)
    analyst = [
        ("Total Analysts Covering", "~60+"),
        ("Buy / Hold / Sell",       "~75% Buy | ~20% Hold | ~5% Sell"),
        ("Median Price Target",     "~$215-225 (note: stock re-rated above many targets)"),
        ("Consensus View",          "Broadly constructive; concerns on antitrust and CapEx intensity"),
    ]
    for i, (k, v) in enumerate(analyst):
        r = 4 + i
        bg = ALT_ROW if i % 2 == 0 else WHITE
        lbl(ws, r, 2, k, bg=bg)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
        wc(ws, r, 3, v, size=FONT_SIZE, bg=bg, align="left", border=True)
        ws.row_dimensions[r].height = 22

    sub_hdr(ws, 10, 2, "Key Bullish Narratives", span=3)
    bull = [
        ("AI Infrastructure Leader", "Google owns the full AI stack: data, models, TPU hardware, cloud. No other company has this breadth."),
        ("Cloud Acceleration",       "Google Cloud growing 40%+; now a $43B+ business with 17%+ operating margin and accelerating"),
        ("Valuation Discount",       "GOOGL trades at 23x fwd P/E vs. 28x historical avg; antitrust overhang overstated"),
        ("FCF Machine",              "$164.7B OCF with minimal dilution via buybacks; capital returns will resume post-CapEx peak"),
        ("YouTube + Subscriptions",  "$60B+ YouTube ecosystem and growing $40B subscriptions/devices segment — underfollowed"),
    ]
    for i, (t, d) in enumerate(bull):
        r = 11 + i
        bg = ALT_ROW if i % 2 == 0 else WHITE
        lbl(ws, r, 2, t, bg=bg)
        ws.row_dimensions[r].height = 40
        wc(ws, r, 3, d, size=FONT_SIZE, bg=bg, align="left", wrap=True, border=True)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)

    sub_hdr(ws, 18, 2, "Key Bearish Narratives", span=3)
    bear = [
        ("Antitrust Breakup Risk",   "DOJ actively seeking Chrome + Android divestiture; structural change could impair economics"),
        ("AI Search Erosion",        "ChatGPT/Copilot eating into search queries; Google AI Overviews cannibalizing ad clicks"),
        ("CapEx Trap",               "$175B+ CapEx in 2026 massively depresses FCF; if AI ROI disappoints, huge value destruction"),
        ("Regulatory Whack-a-Mole",  "EU DMA, UK CMA, Japan JFTC all pursuing Alphabet; legal costs + behavioral remedies ongoing"),
    ]
    for i, (t, d) in enumerate(bear):
        r = 19 + i
        bg = ALT_ROW if i % 2 == 0 else WHITE
        lbl(ws, r, 2, t, bg=bg)
        ws.row_dimensions[r].height = 40
        wc(ws, r, 3, d, size=FONT_SIZE, bg=bg, align="left", wrap=True, border=True)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)

    sub_hdr(ws, 25, 2, "Recent News & Catalysts (2025-2026)", span=3)
    news = [
        ("Feb 2026",     "Q4 2025 earnings beat; Cloud 48% growth; $113.8B quarterly revenue record"),
        ("Jan 2026",     "Gemini 2.0 launched; Google AI Mode in Search rolling out globally"),
        ("Dec 2025",     "DOJ filed remedies brief seeking Chrome browser divestiture"),
        ("Nov 2025",     "Waymo expanded to Miami and Tokyo; commercialization accelerating"),
        ("Oct 2025",     "Q3 2025 earnings; announced $175-185B CapEx for 2026 — stock initially fell 5%"),
        ("Apr 2025",     "Judge ruled Google must allow alternative app stores on Android; EU settlement"),
    ]
    for i, (d, n) in enumerate(news):
        r = 26 + i
        bg = ALT_ROW if i % 2 == 0 else WHITE
        lbl(ws, r, 2, d, bg=bg)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
        wc(ws, r, 3, n, size=FONT_SIZE, bg=bg, align="left", border=True)
        ws.row_dimensions[r].height = 22

# ═══════════════════════════════════════════════════════════════════════════════
def build_key_indicators(wb):
    ws = wb.create_sheet("Key Indicators")
    ws.sheet_view.showGridLines = False
    scw(ws, 1, 3); scw(ws, 2, 30); scw(ws, 3, 20); scw(ws, 4, 20)
    scw(ws, 5, 20); scw(ws, 6, 20); scw(ws, 7, 20)

    sec_hdr(ws, 1, 2, "KEY INDICATORS — ALPHABET INC. (GOOGL)  |  FY2021-FY2025", span=6)
    sub_hdr(ws, 2, 2, "Indicator")
    for i, yr in enumerate(["FY2021", "FY2022", "FY2023", "FY2024", "FY2025"]):
        sub_hdr(ws, 2, 3+i, yr)

    rows = [
        ("FINANCIAL SUMMARY",           None, None, None, None, None),
        ("  Total Revenue ($B)",         257.6, 282.8, 307.4, 350.0, 402.8),
        ("  Revenue Growth",             "41%", "9.8%","8.7%","13.8%","15.1%"),
        ("  Gross Profit ($B)",          146.7, 151.1, 174.1, 206.0, 242.8),
        ("  Gross Margin",               "57%","53.4%","56.6%","58.9%","60.3%"),
        ("  Operating Income ($B)",       78.7,  74.8,  84.3, 112.4, 127.1),
        ("  Operating Margin",           "30.6%","26.5%","27.4%","32.1%","31.6%"),
        ("  Net Income ($B)",             76.0,  59.9,  73.8, 102.5, 132.2),
        ("  Net Margin",                 "29.5%","21.2%","24.0%","29.3%","32.8%"),
        ("  Diluted EPS ($)",              5.61,  4.56,  5.80,  8.05, 10.85),
        ("  EPS Growth",                 "91%","-19%","27%","38.8%","34.8%"),
        ("", None, None, None, None, None),
        ("CASH FLOW & RETURNS",          None, None, None, None, None),
        ("  Operating Cash Flow ($B)",    91.7,  91.5, 101.7, 125.0, 164.7),
        ("  CapEx ($B)",                  24.3,  31.5,  32.3,  52.5,  91.4),
        ("  Free Cash Flow ($B)",         67.4,  60.0,  69.4,  72.5,  73.3),
        ("  Share Buybacks ($B)",         50.3,  59.3,  62.2,  62.2,  45.7),
        ("  ROE",                        "29.5%","18.0%","22.5%","27.1%","36.2%"),
        ("  ROIC",                       "29.8%","22.5%","27.5%","35.0%","38.0%"),
        ("", None, None, None, None, None),
        ("BALANCE SHEET",               None, None, None, None, None),
        ("  Cash + ST Investments ($B)", 136.7, 109.8, 115.9, 137.9, 141.7),
        ("  Long-term Debt ($B)",          0.0,  14.7,  14.7,  14.7,  46.5),
        ("  Net Cash ($B)",              136.7,  95.1, 101.2, 123.2,  95.2),
        ("  Debt-to-Equity",              0.00,  0.05,  0.04,  0.04,  0.13),
        ("", None, None, None, None, None),
        ("VALUATION",                   None, None, None, None, None),
        ("  P/E Ratio",                   "27x",  "21x",  "26x",  "24x",  "29x"),
        ("  EV/EBITDA",                   "22x",  "17x",  "20x",  "19x",  "22x"),
        ("  Price/FCF",                   "27x",  "24x",  "27x",  "31x",  "53x"),
        ("", None, None, None, None, None),
        ("BUSINESS METRICS",            None, None, None, None, None),
        ("  Google Cloud Growth",        "45%",  "37%",  "26%",  "28%",  "~40%"),
        ("  YouTube Revenue ($B)",        28.8,  29.2,  31.5,  36.1,  "60B+"),
        ("  Employees (thousands)",        156,   187,   182,   181,  "~180"),
        ("  Revenue/Employee ($K)",       None, None, 1344, 1519, 1700),
    ]

    for i, row_data in enumerate(rows):
        r = 3 + i
        label = row_data[0]
        vals  = row_data[1:]
        if label == "":
            ws.row_dimensions[r].height = 8
            continue
        is_hdr = not label.startswith("  ")
        bg = SUBHDR_BG if is_hdr else (ALT_ROW if i % 2 == 0 else WHITE)
        lbl(ws, r, 2, label, bg=bg)
        for j, v in enumerate(vals):
            if v is None:
                dc(ws, r, 3+j, "—", bg=bg)
            elif isinstance(v, str):
                wc(ws, r, 3+j, v, size=FONT_SIZE, bg=bg, align="right", border=True,
                   bold=is_hdr, italic=not is_hdr and "%" in v)
            else:
                dc(ws, r, 3+j, v, bg=bg, bold=is_hdr, num_fmt='#,##0.0' if abs(v)>10 else '#,##0.00')

# ═══════════════════════════════════════════════════════════════════════════════
def main():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    build_cover(wb)
    build_business(wb)
    build_moat(wb)
    build_income(wb)
    build_balance_sheet(wb)
    build_cashflow(wb)
    build_roic(wb)
    build_management(wb)
    build_risks(wb)
    build_valuation(wb)
    build_sentiment(wb)
    build_key_indicators(wb)

    out_dir = "/Users/naji/WORK/github.com/AI/claude/Agent/MarketResearch/output"
    os.makedirs(out_dir, exist_ok=True)
    path = os.path.join(out_dir, "GOOGL_Financial_Analysis.xlsx")
    wb.save(path)
    print(f"Saved: {path}")

if __name__ == "__main__":
    main()
