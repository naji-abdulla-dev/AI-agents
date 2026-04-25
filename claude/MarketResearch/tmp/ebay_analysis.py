"""
eBay Inc. (EBAY) Financial Analysis - Excel Generator
Data as of April 2026 | FY2025 Annual (ended December 31, 2025)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
import os

# ── Color palette ──────────────────────────────────────────────────────────────
EBAY_BLUE   = "3665F3"
EBAY_RED    = "E53238"
EBAY_YELLOW = "F5AF02"
EBAY_GREEN  = "86B817"
DARK_BG     = "1A1A2E"
HEADER_BG   = "3665F3"
HEADER_FG   = "FFFFFF"
SUBHDR_BG   = "D6E0FD"
SUBHDR_FG   = "1A1A2E"
ALT_ROW     = "EEF2FF"
WHITE       = "FFFFFF"
LIGHT_GRAY  = "F5F6FA"
RED         = "C62828"
GREEN       = "2E7D32"
WARN_YELLOW = "FFF9C4"
POSITIVE    = "E8F5E9"
NEGATIVE    = "FFEBEE"

FONT_SIZE = 14

def mf(bold=False, size=FONT_SIZE, color="000000", italic=False):
    return Font(name="Calibri", bold=bold, size=size, color=color, italic=italic)

def mfill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def mborder(style="thin"):
    s = Side(border_style=style, color="BBBBBB")
    return Border(left=s, right=s, top=s, bottom=s)

def cal(h="left", wrap=False):
    return Alignment(horizontal=h, vertical="center", wrap_text=wrap)

def wc(ws, row, col, value, bold=False, size=FONT_SIZE, fg="000000", bg=None,
       align="left", wrap=False, border=False, italic=False, num_fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = mf(bold=bold, size=size, color=fg, italic=italic)
    if bg:
        c.fill = mfill(bg)
    c.alignment = cal(align, wrap)
    if border:
        c.border = mborder()
    if num_fmt:
        c.number_format = num_fmt
    return c

def sec_hdr(ws, row, col, text, span=1, bg=HEADER_BG, fg=HEADER_FG):
    c = wc(ws, row, col, text, bold=True, size=FONT_SIZE+2, fg=fg, bg=bg, align="center")
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+span-1)
    ws.row_dimensions[row].height = 30
    return c

def sub_hdr(ws, row, col, text, span=1, bg=SUBHDR_BG, fg=SUBHDR_FG):
    c = wc(ws, row, col, text, bold=True, size=FONT_SIZE, fg=fg, bg=bg, align="center", border=True)
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+span-1)
    ws.row_dimensions[row].height = 22
    return c

def lbl(ws, row, col, text, bg=None, indent=0):
    val = ("   " * indent) + text
    c = wc(ws, row, col, val, size=FONT_SIZE, bg=bg, align="left", border=True)
    ws.row_dimensions[row].height = 20
    return c

def dc(ws, row, col, value, bg=None, num_fmt=None, bold=False, color="000000"):
    c = wc(ws, row, col, value, bold=bold, size=FONT_SIZE, fg=color, bg=bg, align="right", border=True, num_fmt=num_fmt)
    ws.row_dimensions[row].height = 20
    return c

def alt(i):
    return ALT_ROW if i % 2 == 0 else WHITE

def scw(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

def pct_color(val):
    if val is None: return WHITE
    if isinstance(val, str): return WHITE
    return POSITIVE if val >= 0 else NEGATIVE

# ═══════════════════════════════════════════════════════════════════════════════
def build_cover(wb):
    ws = wb.create_sheet("Cover")
    ws.sheet_view.showGridLines = False
    for c in range(1, 9):
        scw(ws, c, 18)
    scw(ws, 1, 4)
    scw(ws, 2, 34)
    scw(ws, 3, 24)

    # Title banner
    ws.merge_cells("B3:G3")
    wc(ws, 3, 2, "eBay Inc. (EBAY)", bold=True, size=36, fg=WHITE, bg=EBAY_BLUE, align="center")
    ws.row_dimensions[3].height = 60

    ws.merge_cells("B4:G4")
    wc(ws, 4, 2, "Comprehensive Financial Analysis", bold=True, size=20, fg=WHITE, bg=DARK_BG, align="center")
    ws.row_dimensions[4].height = 36

    ws.merge_cells("B5:G5")
    wc(ws, 5, 2, "April 2026  |  FY2025 Full-Year Results", bold=False, size=FONT_SIZE, fg=HEADER_FG, bg=EBAY_BLUE, align="center", italic=True)
    ws.row_dimensions[5].height = 28

    # Key snapshot table
    row = 7
    sec_hdr(ws, row, 2, "Company Snapshot", span=5)
    row += 1

    snap = [
        ("Ticker / Exchange",      "EBAY / NASDAQ"),
        ("Sector / Industry",      "Consumer Discretionary / E-Commerce Marketplace"),
        ("Founded",                "1995, San Jose, California"),
        ("CEO",                    "Jamie Iannone (since April 2020)"),
        ("FY2025 Revenue",         "$11.1 Billion  (+7.9% YoY)"),
        ("FY2025 GMV",             "$79.6 Billion  (+7% YoY)"),
        ("FY2025 Net Income",      "$2.03 Billion  (18.3% margin)"),
        ("FY2025 Free Cash Flow",  "$1.47 Billion"),
        ("Active Buyers",          "~135 Million"),
        ("Live Listings",          "~2.5 Billion"),
        ("Stock Price (Apr 2026)", "~$96.01"),
        ("Market Cap",             "~$45.7 Billion"),
        ("Dividend Yield",         "~1.5%"),
        ("P/E (Trailing GAAP)",    "~22.5x"),
        ("Forward P/E",            "~15.5x"),
        ("Analyst Consensus",      "Hold  |  Avg. Target $97–98"),
    ]
    for i, (label, val) in enumerate(snap):
        bg = alt(i)
        lbl(ws, row, 2, label, bg=bg)
        c = wc(ws, row, 3, val, size=FONT_SIZE, bg=bg, align="left", border=True)
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)
        row += 1

    # Analyst notes
    row += 1
    sec_hdr(ws, row, 2, "Analyst Summary", span=5)
    row += 1
    notes = [
        "eBay is the world's largest pure-play marketplace, connecting ~135M buyers with ~18M sellers across ~2.5B listings.",
        "Unlike Amazon, eBay owns NO inventory or logistics — a capital-light, high-margin model with ~71% gross margins.",
        "Recommerce (pre-owned goods) represents >40% of GMV and is eBay's fastest-growing segment — structurally advantaged.",
        "Advertising revenue reached ~$2B in FY2025 (18% of revenue), growing 17%+ and expanding the take rate.",
        "The Depop acquisition (pending, Q2 2026) extends eBay's C2C and youth fashion exposure.",
        "Management returned $3.0B to shareholders in 2025 via buybacks ($2.5B) + dividends ($0.53B).",
        "Key risks: Temu/Shein competition in low-price segments, EU de minimis rule changes, compliance leadership gap.",
        "DCF analysis suggests intrinsic value of $62–$95/share depending on FCF assumptions and growth rates.",
    ]
    for i, note in enumerate(notes):
        bg = alt(i)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        wc(ws, row, 2, f"• {note}", size=FONT_SIZE, bg=bg, align="left", wrap=True, border=True)
        ws.row_dimensions[row].height = 28
        row += 1

    # Color legend row
    row += 1
    wc(ws, row, 2, "eBay Brand Colors:", bold=True, size=FONT_SIZE-1, fg=DARK_BG)
    for col, (label, clr) in enumerate([("Blue","3665F3"),("Red","E53238"),("Yellow","F5AF02"),("Green","86B817")], start=3):
        c = wc(ws, row, col, label, bold=True, size=FONT_SIZE-1, fg=WHITE, bg=clr, align="center")
    ws.row_dimensions[row].height = 18


# ═══════════════════════════════════════════════════════════════════════════════
def build_business_overview(wb):
    ws = wb.create_sheet("Business Overview")
    ws.sheet_view.showGridLines = False
    scw(ws, 1, 3); scw(ws, 2, 32); scw(ws, 3, 20); scw(ws, 4, 20)
    scw(ws, 5, 20); scw(ws, 6, 20); scw(ws, 7, 20)

    row = 2
    sec_hdr(ws, row, 2, "eBay Inc. — Business Overview", span=5)
    row += 2

    # ── What eBay does ──
    sec_hdr(ws, row, 2, "Products & Services", span=5, bg=EBAY_BLUE)
    row += 1
    products = [
        ("Marketplace Platform",
         "Core two-sided marketplace connecting buyers and sellers for new, used, refurbished, "
         "and collectible goods. Revenue via transaction take rate (~14% of GMV). "
         "2.5B live listings across virtually every product category."),
        ("First-Party Advertising",
         "Promoted Listings (standard + advanced) and display ads. Sellers pay to boost "
         "visibility. FY2025 first-party ad revenue ~$1.89B (+17% YoY). "
         "Take rate growing: ads are now ~2.5% of GMV."),
        ("Managed Payments",
         "Fully in-house payments processing (since 2021). Replaced PayPal. "
         "Generates incremental revenue and improved unit economics. "
         "Processes all transactions across 190+ countries."),
        ("Authentication Services",
         "Authenticity Guarantee for luxury watches, sneakers, trading cards, "
         "handbags, jewelry. Third-party verification builds trust in high-value categories. "
         "Directly combats StockX, Chrono24, and other niche competitors."),
        ("Depop (pending acquisition)",
         "Social fashion resale app popular with Gen Z. Acquisition expected Q2 2026. "
         "Will add ~1–2 pp to FX-neutral GMV growth but create a low-single-digit "
         "headwind to near-term non-GAAP operating income."),
    ]
    for i, (name, desc) in enumerate(products):
        bg = alt(i)
        lbl(ws, row, 2, name, bg=bg)
        c = wc(ws, row, 3, desc, size=FONT_SIZE, bg=bg, align="left", wrap=True, border=True)
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)
        ws.row_dimensions[row].height = 52
        row += 1

    row += 1
    # ── Revenue breakdown ──
    sec_hdr(ws, row, 2, "Revenue Breakdown — FY2025 ($11.1B)", span=5, bg=EBAY_RED)
    row += 1
    sub_hdr(ws, row, 2, "Category"); sub_hdr(ws, row, 3, "Revenue ($M)"); sub_hdr(ws, row, 4, "% of Total"); sub_hdr(ws, row, 5, "YoY Growth"); sub_hdr(ws, row, 6, "Notes")
    row += 1
    rev_data = [
        ("Marketplace (Take Rate)",  "~$9,100",  "~82%",  "+5–6%",  "Transaction fees on ~$79.6B GMV"),
        ("Advertising (1P + 3P)",    "~$2,000",  "~18%",  "+17%",   "Fastest-growing segment; ~2.5% GMV penetration"),
        ("Total Net Revenue",        "$11,100",  "100%",  "+7.9%",  "Outpaced GMV growth by ~1 point"),
    ]
    for i, row_data in enumerate(rev_data):
        bg = alt(i) if i < 2 else SUBHDR_BG
        bold = (i == 2)
        for col, val in enumerate(row_data, start=2):
            wc(ws, row, col, val, size=FONT_SIZE, bg=bg, align="right" if col > 2 else "left", border=True, bold=bold)
        row += 1

    row += 1
    sec_hdr(ws, row, 2, "Revenue by Geography — FY2025 (Estimated)", span=5, bg=EBAY_GREEN)
    row += 1
    sub_hdr(ws, row, 2, "Geography"); sub_hdr(ws, row, 3, "Revenue ($M)"); sub_hdr(ws, row, 4, "% of Total"); sub_hdr(ws, row, 5, "YoY Growth"); sub_hdr(ws, row, 6, "Key Notes")
    row += 1
    geo_data = [
        ("United States",     "~$6,100",  "~55%",  "+6%",   "Largest market; strong focus category growth"),
        ("United Kingdom",    "~$1,550",  "~14%",  "+9%",   "2nd largest; recommerce leadership"),
        ("Germany",           "~$1,330",  "~12%",  "+8%",   "Strong auto parts; Kleinanzeigen integration"),
        ("Rest of Europe",    "~$1,220",  "~11%",  "+8%",   "Fragmented but growing"),
        ("Rest of World",     "~$900",    "~8%",   "+12%",  "APAC, Canada, Australia, MENA"),
        ("Total",             "$11,100",  "100%",  "+7.9%", ""),
    ]
    for i, row_data in enumerate(geo_data):
        bg = alt(i) if i < 5 else SUBHDR_BG
        bold = (i == 5)
        for col, val in enumerate(row_data, start=2):
            wc(ws, row, col, val, size=FONT_SIZE, bg=bg, align="right" if col > 2 else "left", border=True, bold=bold)
        row += 1

    row += 1
    # ── Value proposition & clients ──
    sec_hdr(ws, row, 2, "Value Proposition & Key Customer Segments", span=5, bg=EBAY_YELLOW, fg=DARK_BG)
    row += 1
    vp_data = [
        ("Casual Sellers (C2C)",
         "~25% of GMV. Individuals selling unwanted items — clothing, electronics, collectibles. "
         "eBay's simplest seller journey vs. Amazon FBA complexity."),
        ("Small Business Sellers",
         "Largest seller cohort by volume. Niche retailers, vintage dealers, refurbishers. "
         "Low platform fees relative to Amazon, no required fulfillment."),
        ("Large / Professional Sellers",
         "eBay Stores subscribers (~600K stores). Auto dealers (via eBay Motors). "
         "Managed Payments provides consolidated settlement."),
        ("Collectors & Enthusiasts",
         "Sports cards, vintage watches, rare sneakers, vinyl records. "
         "Authenticity Guarantee builds premium trust. eBay dominates this niche."),
        ("Buyers Seeking Value",
         "Price-sensitive shoppers seeking used, refurbished, and overstock. "
         "Recommerce tailwind as consumers downsize and sustainability awareness grows."),
    ]
    for i, (seg, desc) in enumerate(vp_data):
        bg = alt(i)
        lbl(ws, row, 2, seg, bg=bg)
        c = wc(ws, row, 3, desc, size=FONT_SIZE, bg=bg, align="left", wrap=True, border=True)
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)
        ws.row_dimensions[row].height = 44
        row += 1

    row += 1
    # ── GMV focus categories ──
    sec_hdr(ws, row, 2, "GMV Breakdown by Focus Category (FY2025)", span=5, bg=EBAY_BLUE)
    row += 1
    sub_hdr(ws, row, 2, "Category"); sub_hdr(ws, row, 3, "Est. GMV"); sub_hdr(ws, row, 4, "YoY Growth"); sub_hdr(ws, row, 5, "Differentiation"); sub_hdr(ws, row, 6, "Moat?")
    row += 1
    cat_data = [
        ("Fashion / Apparel",        ">$10B",    "+12%+",  "Authenticity Guarantee; recommerce",     "Strong"),
        ("Collectibles",             ">$10B",    "+12%+",  "Trading cards, coins, stamps; grading",  "Very Strong"),
        ("Automotive Parts & Acc.",  ">$10B",    "+12%+",  "eBay Motors; B2B and DIY buyers",        "Strong"),
        ("Consumer Electronics",     ">$10B",    "+8%",    "Refurb & used; competitive vs Amazon",   "Moderate"),
        ("Home & Garden",            ">$10B",    "+8%",    "Estate sales, furniture, décor",         "Moderate"),
        ("Recommerce (total)",        ">$32B",    "+15%+",  "40%+ of GMV; pre-owned everything",      "Growing"),
        ("C2C",                       "~$20B",    "+10%",   "25% of GMV; individuals selling",        "Moderate"),
    ]
    for i, row_data in enumerate(cat_data):
        bg = alt(i)
        for col, val in enumerate(row_data, start=2):
            wc(ws, row, col, val, size=FONT_SIZE, bg=bg, align="right" if col == 3 else "left", border=True)
        row += 1

    row += 1
    # ── Seasonality ──
    sec_hdr(ws, row, 2, "Seasonality & Margin Structure", span=5, bg=EBAY_RED)
    row += 1
    seas = [
        ("Strongest Quarter", "Q4 (Oct–Dec)", "Holiday shopping; GMV $21.2B in Q4 2025 (+10%)"),
        ("Weakest Quarter",   "Q1 (Jan–Mar)", "Post-holiday lull; GMV $18.8B in Q1 2025 (+1%)"),
        ("Gross Margin",      "~71.5%",       "Capital-light; no inventory, no logistics costs"),
        ("Operating Margin",  "~20.5%",       "R&D + S&M heavy; room for expansion via advertising"),
        ("Non-GAAP Op Margin","~26.1%",       "Adjusted for SBC and D&A"),
        ("FCF Margin",        "~13.2%",       "FCF $1.47B / Revenue $11.1B = 13.2%"),
    ]
    for i, (item, val, note) in enumerate(seas):
        bg = alt(i)
        lbl(ws, row, 2, item, bg=bg)
        wc(ws, row, 3, val, size=FONT_SIZE, bg=bg, align="center", border=True, bold=True)
        c = wc(ws, row, 4, note, size=FONT_SIZE, bg=bg, align="left", wrap=True, border=True)
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=6)
        ws.row_dimensions[row].height = 22
        row += 1


# ═══════════════════════════════════════════════════════════════════════════════
def build_moat(wb):
    ws = wb.create_sheet("Moat")
    ws.sheet_view.showGridLines = False
    scw(ws, 1, 3); scw(ws, 2, 28); scw(ws, 3, 18); scw(ws, 4, 40); scw(ws, 5, 18); scw(ws, 6, 18)

    row = 2
    sec_hdr(ws, row, 2, "eBay — Competitive Moat Analysis", span=5, bg=DARK_BG)
    row += 2

    moat_items = [
        (EBAY_BLUE, "1. Network Effects (STRONG)",
         "eBay's #1 moat. 135M buyers + 18M sellers create a self-reinforcing loop: "
         "more buyers attract more sellers and vice versa. 2.5B live listings mean near-infinite "
         "selection that no new entrant can replicate. Liquidity in niche categories (e.g., 1970s "
         "baseball cards, rare vinyl) is unmatched globally.",
         "Rating: 8/10", "Durability: High"),

        (EBAY_RED, "2. Category Leadership in Recommerce (STRONG)",
         "40%+ of GMV is used/refurbished goods. As consumers embrace sustainability and value, "
         "recommerce is structurally growing. eBay is the dominant global platform. "
         "Competitors (ThredUp, Poshmark, Depop) address narrow sub-categories; "
         "eBay spans everything. The Depop acquisition deepens this advantage.",
         "Rating: 8/10", "Durability: High"),

        (EBAY_YELLOW, "3. Authentication Infrastructure (STRONG)",
         "Authenticity Guarantee for luxury watches, sneakers, trading cards, handbags, and jewelry. "
         "Physical inspection by third-party experts. This is expensive to replicate and creates "
         "high buyer trust — directly countering StockX, GOAT, and Chrono24 in their home categories.",
         "Rating: 7/10", "Durability: Medium-High"),

        (EBAY_GREEN, "4. Capital-Light Business Model (STRONG)",
         "eBay holds zero inventory and operates no logistics infrastructure. "
         "This yields 71.5% gross margins — far above Amazon (~44%) or Walmart (~25%). "
         "The capital-light model enables consistent FCF generation and massive buybacks "
         "even during revenue downturns.",
         "Rating: 8/10", "Durability: High"),

        (EBAY_BLUE, "5. Managed Payments + Proprietary Data (MODERATE)",
         "Full in-house payment processing (since 2021) creates a transaction data layer: "
         "pricing intelligence, buyer behavior, and fraud signals. "
         "AI tools (Magical Listing, dynamic pricing) lower seller friction. "
         "Payment margins improve take rate over time.",
         "Rating: 6/10", "Durability: Medium"),

        (EBAY_RED, "6. Brand Trust & Global Reach (MODERATE)",
         "eBay is a 30-year-old global brand with operations in 190+ countries. "
         "Buyers/sellers worldwide recognize and trust the brand — especially important "
         "for cross-border trade (estimated ~20% of GMV is cross-border).",
         "Rating: 6/10", "Durability: Medium"),

        (EBAY_YELLOW, "7. Switching Costs (LOW-MODERATE)",
         "Seller switching costs are moderate: eBay store subscribers, seller ratings, "
         "listing history, and customer base all create inertia. "
         "Buyer switching costs are low — many buyers use multiple platforms. "
         "Mitigated by loyalty programs and exclusive authentication services.",
         "Rating: 5/10", "Durability: Medium"),
    ]

    for bg, title, desc, rating, durability in moat_items:
        sub_hdr(ws, row, 2, title, span=5, bg=bg, fg=WHITE)
        row += 1
        c = wc(ws, row, 2, desc, size=FONT_SIZE, bg=WHITE, align="left", wrap=True, border=True)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        wc(ws, row, 5, rating, size=FONT_SIZE, bg=WHITE, align="center", border=True, bold=True)
        wc(ws, row, 6, durability, size=FONT_SIZE, bg=WHITE, align="center", border=True)
        ws.row_dimensions[row].height = 60
        row += 2

    # Competitive landscape
    sec_hdr(ws, row, 2, "Competitive Landscape", span=5, bg=DARK_BG)
    row += 1
    sub_hdr(ws, row, 2, "Competitor"); sub_hdr(ws, row, 3, "Est. Rev (2025)"); sub_hdr(ws, row, 4, "Overlap"); sub_hdr(ws, row, 5, "Threat Level"); sub_hdr(ws, row, 6, "eBay Advantage")
    row += 1
    comps = [
        ("Amazon Marketplace", "~$700B+ GMV",  "General merchandise, electronics", "High",        "Lower fees, auction format, used goods"),
        ("Temu / Shein",       "~$40B+ GMV",   "Low-price fashion & goods",        "High",         "Trust, authenticity, recommerce focus"),
        ("Etsy",               "~$13B GMV",    "Handmade, vintage, crafts",        "Low-Moderate", "Scale, brand diversity, global reach"),
        ("MercadoLibre",       "~$55B GMV",    "LATAM (minimal US overlap)",       "Low",          "US/EU dominance, payments integration"),
        ("Facebook Marketplace","~$30B+ GMV",  "Local C2C, furniture, cars",       "Moderate",     "Shipping, trust, global reach"),
        ("StockX / GOAT",      "Niche",        "Sneakers, collectibles",           "Moderate",     "Breadth, authentication across categories"),
        ("Poshmark / Depop",   "~$2–3B GMV",   "Fashion C2C",                      "Moderate",     "Scale; Depop being acquired by eBay"),
        ("Walmart.com",        "~$75B+ GMV",   "General merchandise",              "Moderate",     "Used/recommerce, auction, niche categories"),
    ]
    for i, row_data in enumerate(comps):
        bg = alt(i)
        for col, val in enumerate(row_data, start=2):
            threat = row_data[3]
            cell_bg = bg
            if col == 5:
                cell_bg = NEGATIVE if "High" in threat else (WARN_YELLOW if "Moderate" in threat else POSITIVE)
            wc(ws, row, col, val, size=FONT_SIZE, bg=cell_bg, align="left", border=True)
        row += 1

    row += 1
    # Moat summary score
    sec_hdr(ws, row, 2, "Overall Moat Assessment", span=5, bg=EBAY_BLUE)
    row += 1
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    wc(ws, row, 2,
       "eBay possesses a NARROW-TO-WIDE MOAT driven by strong network effects, recommerce leadership, "
       "and authentication infrastructure. The moat is NOT as wide as Visa or MSFT because buyer "
       "switching costs are low and Amazon/Temu compete aggressively. However, eBay's structural "
       "position in recommerce, collectibles, and C2C provides durable competitive advantages "
       "that are difficult and expensive to replicate. Overall Moat Score: 7/10.",
       size=FONT_SIZE, bg=ALT_ROW, align="left", wrap=True, border=True, bold=True)
    ws.row_dimensions[row].height = 60


# ═══════════════════════════════════════════════════════════════════════════════
def build_income_statement(wb):
    ws = wb.create_sheet("Income Statement")
    ws.sheet_view.showGridLines = False
    scw(ws, 1, 3); scw(ws, 2, 36); scw(ws, 3, 18); scw(ws, 4, 18); scw(ws, 5, 18); scw(ws, 6, 18); scw(ws, 7, 18); scw(ws, 8, 18)

    row = 2
    sec_hdr(ws, row, 2, "Income Statement — eBay Inc. ($ Millions)", span=7, bg=EBAY_BLUE)
    row += 1
    sub_hdr(ws, row, 2, "Metric")
    for col, yr in enumerate(["FY2021","FY2022","FY2023","FY2024","FY2025","Q4 2025"], start=3):
        sub_hdr(ws, row, col, yr)
    row += 1

    IS = [
        # (label, indent, [2021, 2022, 2023, 2024, 2025, Q4-2025], bold, fmt)
        ("NET REVENUE",                  0, [10421, 9795,  10112, 10283, 11100, 2970],  True,  "#,##0"),
        ("Cost of Net Revenue",          1, [3006,  2815,  2930,  2979,  3167,  848],   False, "#,##0"),
        ("GROSS PROFIT",                 0, [7415,  6980,  7182,  7304,  7933,  2122],  True,  "#,##0"),
        ("Gross Margin %",               1, ["71.2%","71.3%","71.0%","71.0%","71.5%","71.4%"], False, None),
        ("",                             0, [None]*6, False, None),
        ("Research & Development",       1, [1459,  1478,  1408,  1373,  1400,  352],   False, "#,##0"),
        ("Sales & Marketing",            1, [1583,  1523,  1467,  1422,  1505,  401],   False, "#,##0"),
        ("General & Administrative",     1, [815,   798,   778,   774,   750,   189],   False, "#,##0"),
        ("Provision / Restructuring",    1, [0,     31,    93,    72,    0,     0],     False, "#,##0"),
        ("Total Operating Expenses",     0, [3857,  3830,  3746,  3641,  3655,  942],   True,  "#,##0"),
        ("",                             0, [None]*6, False, None),
        ("OPERATING INCOME",             0, [2923,  2353,  1941,  2318,  2278,  600],   True,  "#,##0"),
        ("Operating Margin %",           1, ["28.1%","24.0%","19.2%","22.5%","20.5%","20.2%"], False, None),
        ("",                             0, [None]*6, False, None),
        ("Interest & Other Income(Exp)", 1, [-187,  -153,  -175,  -176,  -165,  -40],   False, "#,##0"),
        ("PRE-TAX INCOME",               0, [2736,  2200,  1766,  2142,  2113,  560],   True,  "#,##0"),
        ("Income Tax Expense",           1, [395,   819,   97,    215,   83,    35],    False, "#,##0"),
        ("Effective Tax Rate %",         1, ["14.4%","37.2%","5.5%","10.0%","3.9%","6.3%"], False, None),
        ("",                             0, [None]*6, False, None),
        ("NET INCOME (cont. ops)",       0, [2341,  1381,  1669,  1927,  2030,  525],   True,  "#,##0"),
        ("Net Margin %",                 1, ["22.5%","14.1%","16.5%","18.7%","18.3%","17.7%"], False, None),
        ("",                             0, [None]*6, False, None),
        ("EPS Diluted (GAAP)",           0, [3.84,  2.56,  3.37,  4.05,  4.26,  1.14], True,  "#,##0.00"),
        ("EPS Diluted (Non-GAAP)",       0, [4.09,  3.05,  4.10,  4.88,  5.52,  1.41], False, "#,##0.00"),
        ("Non-GAAP EPS Growth %",        1, ["N/A","-25.4%","+34.4%","+19.0%","+13.1%","N/A"], False, None),
        ("",                             0, [None]*6, False, None),
        ("EBITDA",                       0, [3486,  2901,  2472,  2877,  2700,  720],   True,  "#,##0"),
        ("EBITDA Margin %",              1, ["33.5%","29.6%","24.4%","28.0%","24.3%","24.2%"], False, None),
        ("",                             0, [None]*6, False, None),
        ("Diluted Shares (M)",           0, [609,   540,   495,   476,   476,   462],   False, "#,##0"),
    ]

    for i, (label, indent, vals, bold, fmt) in enumerate(IS):
        bg = LIGHT_GRAY if bold else alt(i)
        if not label:
            for col in range(2, 9):
                ws.cell(row=row, column=col).value = None
            row += 1
            continue
        lbl(ws, row, 2, ("   " * indent) + label, bg=bg)
        ws.cell(row=row, column=2).font = mf(bold=bold, size=FONT_SIZE)
        ws.cell(row=row, column=2).border = mborder()
        ws.cell(row=row, column=2).fill = mfill(bg)
        for col, val in enumerate(vals, start=3):
            c_bg = bg
            if label == "Non-GAAP EPS Growth %" and isinstance(val, str) and val not in ("N/A",""):
                try:
                    num = float(val.replace("%","").replace("+",""))
                    c_bg = POSITIVE if num >= 0 else NEGATIVE
                except: pass
            dc(ws, row, col, val, bg=c_bg, num_fmt=fmt, bold=bold)
        row += 1

    # Quarterly breakdown 2025
    row += 2
    sec_hdr(ws, row, 2, "Quarterly Revenue & GMV — FY2025 ($ Billions)", span=7, bg=EBAY_RED)
    row += 1
    sub_hdr(ws, row, 2, "Metric")
    for col, q in enumerate(["Q1 2025","Q2 2025","Q3 2025","Q4 2025","FY2025","Q1 2026E"], start=3):
        sub_hdr(ws, row, col, q)
    row += 1
    qtrs = [
        ("Revenue ($B)",           [2.59, 2.73, 2.82, 2.97, 11.10, "3.00–3.05"], True),
        ("YoY Revenue Growth",     ["+1%", "+6%", "+9%", "+15%", "+7.9%", "+13–15%"], False),
        ("GMV ($B)",               [18.8, 19.7, 19.9, 21.2, 79.6, "21.5–21.9"], True),
        ("YoY GMV Growth",         ["+1%", "+5%", "+8%", "+10%", "+7%", "+10–12%"], False),
        ("Gross Margin %",         ["71.5%","71.5%","71.4%","71.4%","71.5%", "~71%"], False),
        ("Op. Margin (Non-GAAP)",  ["24.5%","25.8%","25.8%","26.1%","~25.6%","~25–26%"], False),
        ("Non-GAAP EPS",           [1.25, 1.44, 1.42, 1.41, 5.52, "1.53–1.59"], True),
    ]
    for i, (label, vals, bold) in enumerate(qtrs):
        bg = LIGHT_GRAY if bold else alt(i)
        lbl(ws, row, 2, label, bg=bg)
        ws.cell(row=row, column=2).font = mf(bold=bold, size=FONT_SIZE)
        ws.cell(row=row, column=2).fill = mfill(bg)
        ws.cell(row=row, column=2).border = mborder()
        for col, val in enumerate(vals, start=3):
            dc(ws, row, col, val, bg=bg, bold=bold)
        row += 1


# ═══════════════════════════════════════════════════════════════════════════════
def build_balance_sheet(wb):
    ws = wb.create_sheet("Balance Sheet")
    ws.sheet_view.showGridLines = False
    scw(ws, 1, 3); scw(ws, 2, 38); scw(ws, 3, 18); scw(ws, 4, 18); scw(ws, 5, 18); scw(ws, 6, 18); scw(ws, 7, 18)

    row = 2
    sec_hdr(ws, row, 2, "Balance Sheet — eBay Inc. ($ Millions, Dec 31)", span=6, bg=EBAY_BLUE)
    row += 1
    sub_hdr(ws, row, 2, "Line Item")
    for col, yr in enumerate(["FY2021","FY2022","FY2023","FY2024","FY2025"], start=3):
        sub_hdr(ws, row, col, yr)
    row += 1

    BS = [
        # ASSETS
        ("ASSETS",                         True,  [None]*5),
        ("Current Assets",                 False, [None]*5),
        ("  Cash & Equivalents",           False, [3495, 2158, 2146, 2012, 1900]),
        ("  Short-Term Investments",       False, [1666, 1302, 1144, 1242, 1330]),
        ("  Accounts Receivable",          False, [521,  498,  496,  492,  488]),
        ("  Other Current Assets",         False, [356,  320,  310,  305,  295]),
        ("TOTAL CURRENT ASSETS",           True,  [6038, 4278, 4096, 4051, 4013]),
        ("",                               False, [None]*5),
        ("Non-Current Assets",             False, [None]*5),
        ("  Long-Term Investments",        False, [3480, 2501, 2478, 1810, 1600]),
        ("  Property, Plant & Equipment",  False, [1842, 1650, 1523, 1415, 1400]),
        ("  Goodwill",                     False, [6508, 6508, 6508, 6508, 6508]),
        ("  Intangible Assets",            False, [3020, 2740, 2520, 2320, 2180]),
        ("  Other Non-Current Assets",     False, [3312, 2823, 2675, 2596, 1899]),
        ("TOTAL NON-CURRENT ASSETS",       True,  [18162, 16222, 15704, 14649, 13587]),
        ("",                               False, [None]*5),
        ("TOTAL ASSETS",                   True,  [24200, 20500, 19800, 18700, 17600]),
        ("",                               False, [None]*5),
        # LIABILITIES
        ("LIABILITIES & STOCKHOLDERS' EQUITY", True, [None]*5),
        ("Current Liabilities",            False, [None]*5),
        ("  Current Portion LT Debt",      False, [752,  506,  1506, 752,  750]),
        ("  Accounts Payable",             False, [412,  395,  388,  401,  398]),
        ("  Accrued Expenses",             False, [1124, 1052, 1001, 985,  1010]),
        ("  Customer Accounts Payable",    False, [1874, 2100, 2250, 2421, 2500]),
        ("  Other Current Liabilities",    False, [312,  298,  285,  278,  272]),
        ("TOTAL CURRENT LIABILITIES",      True,  [4474, 4351, 5430, 4837, 4930]),
        ("",                               False, [None]*5),
        ("  Long-Term Debt",               False, [7720, 7728, 7186, 6477, 5998]),
        ("  Deferred Tax Liabilities",     False, [3510, 2987, 2741, 2844, 3260]),
        ("  Operating Lease Liabilities",  False, [548,  476,  423,  387,  350]),
        ("  Other Long-Term Liabilities",  False, [2248, 1958, 2020, 1755, 1662]),
        ("TOTAL NON-CURRENT LIABILITIES",  True,  [14026, 13149, 12370, 11463, 11270]),
        ("",                               False, [None]*5),
        ("TOTAL LIABILITIES",              True,  [18500, 17500, 17800, 16300, 16200]),
        ("",                               False, [None]*5),
        ("Stockholders' Equity (Deficit)", False, [None]*5),
        ("  Common Stock & APIC",          False, [16862, 17210, 17580, 17950, 18400]),
        ("  Treasury Stock",               False, [-14800,-17400,-19600,-21000,-23300]),
        ("  Retained Earnings",            False, [4280,  3640,  4620,  5810,  6820]),
        ("  Accumulated OCI",              False, [-642,  -450,  -600,  -360,  -320]),
        ("TOTAL STOCKHOLDERS' EQUITY",     True,  [5700,  3000,  2000,  2400,  1600]),
        ("",                               False, [None]*5),
        ("TOTAL LIABILITIES + EQUITY",     True,  [24200, 20500, 19800, 18700, 17800]),
    ]

    note_row = row
    for i, (label, bold, vals) in enumerate(BS):
        if not label:
            row += 1
            continue
        bg = LIGHT_GRAY if bold else alt(i)
        lbl(ws, row, 2, label, bg=bg)
        ws.cell(row=row, column=2).font = mf(bold=bold, size=FONT_SIZE)
        ws.cell(row=row, column=2).fill = mfill(bg)
        ws.cell(row=row, column=2).border = mborder()
        for col, val in enumerate(vals, start=3):
            dc(ws, row, col, val, bg=bg, num_fmt="#,##0" if isinstance(val, (int, float)) else None, bold=bold)
        row += 1

    row += 1
    # Key ratios from balance sheet
    sec_hdr(ws, row, 2, "Balance Sheet Ratios", span=6, bg=EBAY_GREEN, fg=WHITE)
    row += 1
    sub_hdr(ws, row, 2, "Ratio")
    for col, yr in enumerate(["FY2021","FY2022","FY2023","FY2024","FY2025"], start=3):
        sub_hdr(ws, row, col, yr)
    row += 1
    ratios = [
        ("Current Ratio",           ["1.35x", "0.98x", "0.75x", "0.84x", "0.81x"]),
        ("Debt-to-Total-Assets",    ["47.5%", "54.1%", "55.7%", "55.7%", "57.9%"]),
        ("Net Debt ($M)",           ["4,557", "6,576", "6,548", "5,467", "5,448"]),
        ("Goodwill / Total Assets", ["26.9%", "31.7%", "32.9%", "34.8%", "37.0%"]),
        ("Book Value per Share",    ["$9.36", "$5.56", "$4.04", "$5.04", "$3.36"]),
    ]
    for i, (lbl_txt, vals) in enumerate(ratios):
        bg = alt(i)
        lbl(ws, row, 2, lbl_txt, bg=bg)
        for col, val in enumerate(vals, start=3):
            dc(ws, row, col, val, bg=bg)
        row += 1

    row += 1
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
    wc(ws, row, 2,
       "Note: Balance sheet figures for FY2021–FY2024 are estimated from public filings and "
       "may differ slightly from audited figures. FY2025 figures are estimated based on Q4 2025 "
       "earnings release and analyst data. Negative stockholders' equity in prior years reflects "
       "eBay's aggressive share buyback program ($25B+ cumulative). Consult the 10-K for precise data.",
       size=FONT_SIZE-1, fg="666666", align="left", wrap=True, italic=True)
    ws.row_dimensions[row].height = 36


# ═══════════════════════════════════════════════════════════════════════════════
def build_cash_flow(wb):
    ws = wb.create_sheet("Cash Flow Analysis")
    ws.sheet_view.showGridLines = False
    scw(ws, 1, 3); scw(ws, 2, 38); scw(ws, 3, 18); scw(ws, 4, 18); scw(ws, 5, 18); scw(ws, 6, 18); scw(ws, 7, 18)

    row = 2
    sec_hdr(ws, row, 2, "Cash Flow Statement — eBay Inc. ($ Millions)", span=6, bg=EBAY_BLUE)
    row += 1
    sub_hdr(ws, row, 2, "Line Item")
    for col, yr in enumerate(["FY2021","FY2022","FY2023","FY2024","FY2025"], start=3):
        sub_hdr(ws, row, col, yr)
    row += 1

    CF = [
        ("OPERATING ACTIVITIES",              True,  [None]*5),
        ("  Net Income",                      False, [2341, 1381, 1669, 1927, 2030]),
        ("  Depreciation & Amortization",     False, [563,  548,  531,  559,  542]),
        ("  Stock-Based Compensation",        False, [517,  538,  497,  489,  500]),
        ("  Changes in Working Capital",      False, [-80,  -248, -312, -296, -310]),
        ("  Other Operating Items",           False, [-341, 161,  -235, -569, -762]),
        ("OPERATING CASH FLOW",               True,  [3000, 2380, 2150, 2110, 2000]),
        ("  % of Revenue",                    False, ["28.8%","24.3%","21.3%","20.5%","18.0%"]),
        ("",                                  False, [None]*5),
        ("INVESTING ACTIVITIES",              True,  [None]*5),
        ("  Capital Expenditures",            False, [-473, -520, -540, -458, -525]),
        ("  Purchases of Investments",        False, [-2891,-2315,-2128,-1943,-1812]),
        ("  Proceeds from Investments",       False, [5210, 3510, 2821, 2445, 1972]),
        ("  Acquisitions (net)",              False, [-312, -89,  -142, -210, -185]),
        ("  Other Investing",                 False, [43,   28,   19,   11,   15]),
        ("NET INVESTING CASH FLOW",           True,  [1577, 614,  30,   -155, -535]),
        ("",                                  False, [None]*5),
        ("FINANCING ACTIVITIES",              True,  [None]*5),
        ("  Share Repurchases",               False, [-2281,-2992,-3400,-2560,-2500]),
        ("  Dividends Paid",                  False, [-332, -432, -468, -519, -531]),
        ("  Debt Issued / (Repaid)",          False, [247,  -125, 498,  -752, -271]),
        ("  Other Financing",                 False, [45,   38,   22,   18,   12]),
        ("NET FINANCING CASH FLOW",           True,  [-2321,-3511,-3348,-3813,-3290]),
        ("",                                  False, [None]*5),
        ("NET CHANGE IN CASH",                True,  [2256, -517, -1168,-1858,-1825]),
        ("BEGINNING CASH",                    False, [1239, 3495, 2978, 1810, 1952]),
        ("ENDING CASH",                       False, [3495, 2978, 1810, 1952,  127]),
        ("",                                  False, [None]*5),
        ("FREE CASH FLOW",                    True,  [2527, 1860, 1610, 1652, 1475]),
        ("  FCF Margin %",                    False, ["24.3%","19.0%","15.9%","16.1%","13.3%"]),
        ("  YoY FCF Growth",                  False, ["N/A", "-26.4%","-13.4%","+2.6%","-10.7%"]),
        ("",                                  False, [None]*5),
        ("SHAREHOLDER RETURNS",               True,  [None]*5),
        ("  Buybacks",                        False, [2281, 2992, 3400, 2560, 2500]),
        ("  Dividends",                       False, [332,  432,  468,  519,  531]),
        ("  Total Returned to Shareholders",  True,  [2613, 3424, 3868, 3079, 3031]),
        ("  As % of FCF",                     False, ["103%","184%","240%","186%","206%"]),
        ("  Cumulative Buybacks Since 2018",  False, ["~$19B","~$22B","~$25B","~$28B","~$30B"]),
    ]

    for i, (label, bold, vals) in enumerate(CF):
        if not label:
            row += 1
            continue
        bg = LIGHT_GRAY if bold else alt(i)
        lbl(ws, row, 2, label, bg=bg)
        ws.cell(row=row, column=2).font = mf(bold=bold, size=FONT_SIZE)
        ws.cell(row=row, column=2).fill = mfill(bg)
        ws.cell(row=row, column=2).border = mborder()
        for col, val in enumerate(vals, start=3):
            c_bg = bg
            if label == "  YoY FCF Growth" and isinstance(val, str) and val not in ("N/A",""):
                try:
                    n = float(val.replace("%","").replace("+",""))
                    c_bg = POSITIVE if n >= 0 else NEGATIVE
                except: pass
            dc(ws, row, col, val, bg=c_bg, num_fmt="#,##0" if isinstance(val, (int, float)) else None, bold=bold)
        row += 1

    row += 1
    sec_hdr(ws, row, 2, "Capital Allocation Philosophy", span=6, bg=EBAY_RED)
    row += 1
    points = [
        "eBay has returned $30B+ to shareholders since 2018 via buybacks and dividends — exceeding annual revenues.",
        "FY2025: Returned $3.03B ($2.5B buybacks + $531M dividends) despite FCF of only $1.47B — funded by debt and cash.",
        "Buyback yield: With $2.5B in buybacks on a ~$45B market cap, eBay is retiring ~5.5% of shares annually.",
        "Dividend: $0.27/quarter ($1.08 annualized), yield ~1.1% at $96. Dividend raised 3% in Q4 2025.",
        "CapEx intensity is LOW (~4.7% of revenue in 2025) — investment in cloud infrastructure and AI tools.",
        "Depop acquisition signals willingness to acquire for strategic growth (C2C / fashion / Gen Z demographic).",
        "CONCERN: Total returns ($3B) exceed FCF ($1.47B) significantly — eBay is borrowing to fund returns. Sustainable only with debt capacity.",
    ]
    for i, pt in enumerate(points):
        bg = alt(i)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
        wc(ws, row, 2, f"{'⚠️ ' if i == 6 else '• '}{pt}", size=FONT_SIZE, bg=bg, align="left", wrap=True, border=True,
           bold=(i == 6))
        ws.row_dimensions[row].height = 28
        row += 1


# ═══════════════════════════════════════════════════════════════════════════════
def build_return_on_capital(wb):
    ws = wb.create_sheet("Return on Capital")
    ws.sheet_view.showGridLines = False
    scw(ws, 1, 3); scw(ws, 2, 36); scw(ws, 3, 18); scw(ws, 4, 18); scw(ws, 5, 18); scw(ws, 6, 18); scw(ws, 7, 18)

    row = 2
    sec_hdr(ws, row, 2, "Return on Capital — eBay Inc.", span=6, bg=EBAY_BLUE)
    row += 1
    sub_hdr(ws, row, 2, "Metric")
    for col, yr in enumerate(["FY2021","FY2022","FY2023","FY2024","FY2025"], start=3):
        sub_hdr(ws, row, col, yr)
    row += 1

    roc = [
        ("PROFITABILITY RETURNS",        True,  [None]*5),
        ("Return on Equity (ROE)",        False, ["41.1%","46.0%","83.5%","80.3%","126.9%"]),
        ("Return on Assets (ROA)",        False, ["9.7%", "6.7%", "8.4%", "10.3%", "11.5%"]),
        ("Return on Invested Capital",    False, ["15.2%","12.1%","10.1%","13.5%","13.9%"]),
        ("Return on Capital Employed",    False, ["18.5%","15.0%","11.4%","15.8%","16.2%"]),
        ("",                              False, [None]*5),
        ("ROIC DECOMPOSITION",            True,  [None]*5),
        ("NOPAT ($M)",                    False, [2340, 1870, 1560, 1870, 1868]),
        ("  Operating Income",            False, [2923, 2353, 1941, 2318, 2278]),
        ("  Less: Taxes @ 20% rate",      False, [-583, -483, -381, -448, -410]),
        ("Invested Capital ($M)",         False, [15412, 15481, 15444, 13848, 13460]),
        ("  Total Equity",                False, [5700, 3000, 2000, 2400, 1600]),
        ("  Total Debt",                  False, [8472, 8234, 8692, 7229, 6748]),
        ("  Less: Cash & Equiv",          False, [-3495,-2158,-2146,-2012,-1900]),
        ("  Less: Non-Operating Assets",  False, [735,  2405, 2898, 6231, 7012]),
        ("ROIC",                          True,  ["15.2%","12.1%","10.1%","13.5%","13.9%"]),
        ("",                              False, [None]*5),
        ("FCF-BASED RETURNS",             True,  [None]*5),
        ("FCF ($M)",                      False, [2527, 1860, 1610, 1652, 1475]),
        ("FCF / Invested Capital",        False, ["16.4%","12.0%","10.4%","11.9%","11.0%"]),
        ("FCF / Market Cap",              False, ["5.4%", "7.5%", "6.9%", "5.8%", "3.2%"]),
        ("FCF / Enterprise Value",        False, ["5.0%", "6.9%", "6.3%", "5.4%", "3.1%"]),
        ("",                              False, [None]*5),
        ("INCREMENTAL RETURNS",           True,  [None]*5),
        ("Revenue Growth (YoY)",          False, ["+17.3%","-6.0%","+3.2%","+1.7%","+7.9%"]),
        ("Incremental Revenue ($M)",      False, [1546,  -626,  317,   171,   817]),
        ("Incremental Op. Income ($M)",   False, [320,   -570,  -412,  377,   -40]),
        ("ROIC on Incremental Capital",   False, ["Var.", "-91%", "N/M",  "N/M",  "~-5%"]),
        ("",                              False, [None]*5),
        ("PEER COMPARISON (FY2025E)",     True,  [None]*5),
        ("eBay ROIC",                     False, ["—","—","—","—","13.9%"]),
        ("Etsy ROIC (est.)",              False, ["—","—","—","—","~18%"]),
        ("Amazon Marketplace ROIC",       False, ["—","—","—","—","~24%"]),
        ("Visa ROIC (benchmark)",         False, ["—","—","—","—","~50%+"]),
    ]

    for i, (label, bold, vals) in enumerate(roc):
        if not label:
            row += 1
            continue
        bg = LIGHT_GRAY if bold else alt(i)
        lbl(ws, row, 2, label, bg=bg)
        ws.cell(row=row, column=2).font = mf(bold=bold, size=FONT_SIZE)
        ws.cell(row=row, column=2).fill = mfill(bg)
        ws.cell(row=row, column=2).border = mborder()
        for col, val in enumerate(vals, start=3):
            dc(ws, row, col, val, bg=bg, num_fmt="#,##0" if isinstance(val, (int, float)) else None, bold=bold)
        row += 1

    row += 1
    sec_hdr(ws, row, 2, "Owner Earnings Analysis (Buffett Method)", span=6, bg=EBAY_GREEN, fg=WHITE)
    row += 1
    oe = [
        ("Net Income",                    [2341, 1381, 1669, 1927, 2030]),
        ("+ D&A",                         [563,  548,  531,  559,  542]),
        ("+ SBC (add back for owner)",    [517,  538,  497,  489,  500]),
        ("- Maintenance CapEx (est.)",    [-250, -270, -285, -260, -300]),
        ("- Required Working Capital",    [-80,  -100, -90,  -80,  -90]),
        ("OWNER EARNINGS",                [3091, 2097, 2322, 2635, 2682]),
        ("Owner Earnings Yield",          ["6.6%","8.4%","9.7%","9.3%","5.9%"]),
        ("P / Owner Earnings",            ["15.2x","11.9x","10.3x","10.8x","17.0x"]),
    ]
    sub_hdr(ws, row, 2, "Component")
    for col, yr in enumerate(["FY2021","FY2022","FY2023","FY2024","FY2025"], start=3):
        sub_hdr(ws, row, col, yr)
    row += 1
    for i, (label, vals) in enumerate(oe):
        bold = label in ("OWNER EARNINGS","Owner Earnings Yield","P / Owner Earnings")
        bg = LIGHT_GRAY if bold else alt(i)
        lbl(ws, row, 2, label, bg=bg)
        ws.cell(row=row, column=2).font = mf(bold=bold, size=FONT_SIZE)
        ws.cell(row=row, column=2).fill = mfill(bg)
        ws.cell(row=row, column=2).border = mborder()
        for col, val in enumerate(vals, start=3):
            dc(ws, row, col, val, bg=bg, num_fmt="#,##0" if isinstance(val, (int, float)) else None, bold=bold)
        row += 1


# ═══════════════════════════════════════════════════════════════════════════════
def build_management(wb):
    ws = wb.create_sheet("Management")
    ws.sheet_view.showGridLines = False
    scw(ws, 1, 3); scw(ws, 2, 32); scw(ws, 3, 20); scw(ws, 4, 48)

    row = 2
    sec_hdr(ws, row, 2, "Management Analysis — eBay Inc.", span=3, bg=DARK_BG)
    row += 2

    # Key executives
    sec_hdr(ws, row, 2, "Key Executives", span=3, bg=EBAY_BLUE)
    row += 1
    sub_hdr(ws, row, 2, "Name / Title"); sub_hdr(ws, row, 3, "Since"); sub_hdr(ws, row, 4, "Background & Notes")
    row += 1
    execs = [
        ("Jamie Iannone — President & CEO",        "Apr 2020", "Former Walmart eCommerce COO; started at eBay 2001–2009. Deep marketplace DNA. "
                                                               "Led eBay's 'Tech-Led Reimagination' strategy. Tenure: 6 years."),
        ("Steve Priest — CFO",                     "2020",     "Former CFO at Celestica. Oversees capital allocation, buybacks, debt management. "
                                                               "Architect of managed payments financial model."),
        ("Eddie Garcia — Chief Product Officer",   "2021",     "Built eBay's AI/ML product stack. Magical Listing, dynamic pricing, search. "
                                                               "Former Walmart Labs VP. Key to advertising revenue growth."),
        ("Jordan Sweetnam — SVP, Americas",        "2019",     "Leads eBay's largest market. Focus on C2C and recommerce growth in the US."),
        ("Kristin Reinke — SVP, Global Markets",   "2020",     "Oversees UK, Germany, Australia, and Rest of World. Strong international growth execution."),
    ]
    for i, (name, since, note) in enumerate(execs):
        bg = alt(i)
        lbl(ws, row, 2, name, bg=bg)
        ws.cell(row=row, column=2).font = mf(bold=True, size=FONT_SIZE)
        ws.cell(row=row, column=2).fill = mfill(bg)
        ws.cell(row=row, column=2).border = mborder()
        wc(ws, row, 3, since, size=FONT_SIZE, bg=bg, align="center", border=True)
        wc(ws, row, 4, note, size=FONT_SIZE, bg=bg, align="left", wrap=True, border=True)
        ws.row_dimensions[row].height = 48
        row += 1

    row += 1
    # CEO compensation
    sec_hdr(ws, row, 2, "CEO Compensation — Jamie Iannone (FY2024 Proxy)", span=3, bg=EBAY_RED)
    row += 1
    comp = [
        ("Base Salary",                  "$1,000,000",  "Fixed component; <5% of total comp"),
        ("Annual Cash Bonus",            "$2,895,000",  "Performance-based; tied to revenue, GMV, and EBITDA targets"),
        ("Long-Term Stock Awards",       "$16,315,235", "PSUs + RSUs; vesting over 3 years; ~80% of total comp"),
        ("Other Compensation",           "$139,415",    "Benefits, security, and perks"),
        ("TOTAL COMPENSATION",           "$20,349,650", "86% performance-based; aligns with shareholder returns"),
        ("Shares Held",                  "~530,466",    "Worth ~$51M at $96/share — meaningful skin in the game"),
        ("Pay-to-Median-Employee Ratio", "~340:1",      "High ratio; typical for large-cap tech-adjacent companies"),
    ]
    sub_hdr(ws, row, 2, "Component"); sub_hdr(ws, row, 3, "Amount"); sub_hdr(ws, row, 4, "Context")
    row += 1
    for i, (comp_item, amt, note) in enumerate(comp):
        bg = LIGHT_GRAY if "TOTAL" in comp_item else alt(i)
        bold = "TOTAL" in comp_item
        lbl(ws, row, 2, comp_item, bg=bg)
        ws.cell(row=row, column=2).font = mf(bold=bold, size=FONT_SIZE)
        ws.cell(row=row, column=2).fill = mfill(bg)
        ws.cell(row=row, column=2).border = mborder()
        wc(ws, row, 3, amt, bold=bold, size=FONT_SIZE, bg=bg, align="right", border=True)
        wc(ws, row, 4, note, size=FONT_SIZE, bg=bg, align="left", wrap=True, border=True)
        ws.row_dimensions[row].height = 22
        row += 1

    row += 1
    # Capital allocation grades
    sec_hdr(ws, row, 2, "Capital Allocation Scorecard", span=3, bg=EBAY_YELLOW, fg=DARK_BG)
    row += 1
    sub_hdr(ws, row, 2, "Dimension"); sub_hdr(ws, row, 3, "Grade"); sub_hdr(ws, row, 4, "Evidence")
    row += 1
    grades = [
        ("Acts Like an Owner",        "B+", "530K shares owned; 86% performance pay; "
                                            "clear strategic vision on recommerce + focus categories"),
        ("Capital Returns to Shareholders","A", "$30B+ returned since 2018; $3B in 2025; "
                                                "consistent buyback + growing dividend program"),
        ("Organic Investment",        "B",  "R&D ~$1.4B/yr; AI tools (Magical Listing, dynamic pricing); "
                                            "authentication expansion — meaningful but not transformative"),
        ("M&A Discipline",            "B-", "Depop ($1.5B) is strategic but expensive; "
                                            "prior divestiture of Classifieds ($9.2B to Adevinta) was value-accretive"),
        ("Leverage Use",              "C+", "Funded buybacks exceeding FCF with debt; "
                                            "$6.7B long-term debt; manageable but concerning trajectory"),
        ("Planting Seeds for Future", "B",  "AI investment, recommerce positioning, focus categories, "
                                            "advertising scaling — all point to durable future growth"),
        ("Transparency",              "B+", "Clear GMV reporting, non-GAAP bridges explained, "
                                            "annual guidance provided; some complexity in GAAP vs non-GAAP"),
    ]
    for i, (dim, grade, ev) in enumerate(grades):
        bg = alt(i)
        grade_color = POSITIVE if grade.startswith("A") else (WARN_YELLOW if grade.startswith("B") else NEGATIVE)
        lbl(ws, row, 2, dim, bg=bg)
        wc(ws, row, 3, grade, bold=True, size=FONT_SIZE+2, bg=grade_color, align="center", border=True)
        wc(ws, row, 4, ev, size=FONT_SIZE, bg=bg, align="left", wrap=True, border=True)
        ws.row_dimensions[row].height = 44
        row += 1

    row += 1
    # Insider transactions
    sec_hdr(ws, row, 2, "Insider Transactions (12 Months to Apr 2026)", span=3, bg=EBAY_BLUE)
    row += 1
    insider = [
        ("Jamie Iannone (CEO)", "Net Seller", "Sold ~100K shares ($8.5M) via 10b5-1 plan. Retention ~530K shares."),
        ("Steve Priest (CFO)",  "Net Seller", "Sold ~30K shares for tax planning. Normal executive pattern."),
        ("Board Directors",     "Net Buyers", "Several board members made open-market purchases at $60–$75 range (2023–2024)."),
        ("Institutional Holdings","~83% Float","Top holders: Vanguard (9.2%), BlackRock (7.1%), Capital Group (5.8%)"),
        ("Short Interest",      "~2.5%",       "Low short interest — not a heavily shorted stock; consensus is Hold."),
        ("Overall Signal",      "NEUTRAL",     "Insider selling is typical for executives with large equity comp grants. "
                                               "No red-flag cluster selling. Board purchases at lower prices = positive."),
    ]
    sub_hdr(ws, row, 2, "Party"); sub_hdr(ws, row, 3, "Activity"); sub_hdr(ws, row, 4, "Detail")
    row += 1
    for i, (party, act, detail) in enumerate(insider):
        bg = alt(i)
        act_bg = POSITIVE if "Buyer" in act or "NEUTRAL" in act else (WARN_YELLOW if "Seller" in act else ALT_ROW)
        lbl(ws, row, 2, party, bg=bg)
        wc(ws, row, 3, act, size=FONT_SIZE, bg=act_bg, align="center", border=True, bold=True)
        wc(ws, row, 4, detail, size=FONT_SIZE, bg=bg, align="left", wrap=True, border=True)
        ws.row_dimensions[row].height = 36
        row += 1


# ═══════════════════════════════════════════════════════════════════════════════
def build_risks(wb):
    ws = wb.create_sheet("Risks")
    ws.sheet_view.showGridLines = False
    scw(ws, 1, 3); scw(ws, 2, 28); scw(ws, 3, 14); scw(ws, 4, 14); scw(ws, 5, 44)

    row = 2
    sec_hdr(ws, row, 2, "Risk Analysis — eBay Inc.", span=4, bg=EBAY_RED)
    row += 1
    sub_hdr(ws, row, 2, "Risk Factor"); sub_hdr(ws, row, 3, "Severity"); sub_hdr(ws, row, 4, "Likelihood"); sub_hdr(ws, row, 5, "Detail & Mitigation")
    row += 1

    risks = [
        # (name, severity, likelihood, detail)
        ("Competition — Temu & Shein",
         "HIGH", "HIGH",
         "Temu/Shein captured price-sensitive buyers via aggressive pricing and direct-from-manufacturer "
         "shipping. eBay's low-cost commodity volume has eroded. Mitigation: eBay is pivoting away from "
         "commoditized goods toward focus categories, recommerce, and authentication where Temu cannot compete."),

        ("Competition — Amazon Marketplace",
         "HIGH", "HIGH",
         "Amazon dominates new goods marketplace with Prime, FBA logistics, and superior search. "
         "eBay's differentiation is used/vintage/niche — but Amazon is expanding second-hand offerings. "
         "Mitigation: eBay's C2C and collectibles moat is structural; Amazon doesn't do auctions."),

        ("EU De Minimis Exemption Removal",
         "MEDIUM", "HIGH",
         "EU eliminates the <€150 de minimis tax exemption in July 2026 + €3/item handling fee. "
         "This increases cost of cross-border imports and may slow eBay's EU cross-border GMV "
         "(estimated ~10–15% of GMV). Mitigation: May actually help eBay vs. Temu/Shein in EU markets."),

        ("US 1099-K Threshold Change",
         "MEDIUM", "HIGH",
         "IRS now requires 1099-K for sellers receiving >$600/year (down from $20,000). "
         "This may deter casual C2C sellers — eBay's 25% GMV segment. "
         "Mitigation: eBay is investing in seller education and simplified tax tools."),

        ("Advertising Revenue Concentration",
         "MEDIUM", "MEDIUM",
         "Advertising is growing from ~18% of revenue to potentially 25%+. "
         "If ad revenue slows, eBay's outperformance vs. GMV growth reverses. "
         "Risk that sellers resist higher ad take rates. Mitigation: Diversifying ad formats."),

        ("Leverage and Capital Returns vs. FCF",
         "MEDIUM", "MEDIUM",
         "eBay returned $3.03B to shareholders in 2025 vs. FCF of $1.47B — funded partly by debt. "
         "Long-term debt of $6.7B. If FCF declines or rates rise, this strategy becomes strained. "
         "Mitigation: Investment-grade credit rating; manageable interest coverage (~10x EBIT/interest)."),

        ("Compliance Leadership Gap",
         "MEDIUM", "LOW",
         "Chief Risk & Compliance Officer Ryan Jones departed October 2025 with no public successor. "
         "Risk of compliance gaps during transition. Historical context: eBay paid $3M DOJ settlement "
         "in 2023 for cyberstalking incident by executives. Board oversight reforms in place."),

        ("Depop Integration Risk",
         "LOW-MEDIUM", "MEDIUM",
         "Depop acquisition (pending Q2 2026, est. ~$1.5B) adds a younger demographic but also "
         "complexity. Social commerce dynamics differ from eBay's model. Near-term NON-GAAP "
         "operating income headwind of low-single-digits expected."),

        ("AI Disruption to Search & Discovery",
         "MEDIUM", "MEDIUM",
         "LLM-powered shopping assistants (Google Shopping AI, ChatGPT shopping) may reduce "
         "direct eBay site traffic. Buyers may discover items via AI without visiting eBay. "
         "Mitigation: eBay is investing in its own AI tools and SEO optimization."),

        ("Currency & Macro Risk",
         "LOW-MEDIUM", "MEDIUM",
         "45% of revenue is international; USD strength (strong in 2025) creates FX headwind. "
         "As-reported vs. FX-neutral GMV often differ by ~2pp. "
         "Consumer discretionary spending sensitive to recession. "
         "eBay's used-goods model partially hedged — value-seekers increase in downturns."),

        ("Cybersecurity / Data Privacy",
         "MEDIUM", "LOW",
         "Large-scale data breach could damage buyer/seller trust permanently. "
         "GDPR compliance risk in EU. eBay suffered a major breach in 2014; "
         "systems have been substantially upgraded. EU Digital Services Act compliance required."),

        ("Regulatory / Antitrust",
         "LOW", "LOW",
         "eBay is less dominant than Amazon/Google — less antitrust exposure. "
         "Payment regulation (managed payments) could face banking regulation if eBay is deemed a "
         "financial institution. Historically eBay has navigated regulation well."),
    ]

    sev_colors = {"HIGH": NEGATIVE, "MEDIUM": WARN_YELLOW, "LOW-MEDIUM": "FFF3E0", "LOW": POSITIVE}

    for i, (name, sev, like, detail) in enumerate(risks):
        bg = alt(i)
        s_bg = sev_colors.get(sev, WHITE)
        l_bg = sev_colors.get(like, WHITE)
        lbl(ws, row, 2, name, bg=bg)
        ws.cell(row=row, column=2).font = mf(bold=True, size=FONT_SIZE)
        ws.cell(row=row, column=2).fill = mfill(bg)
        ws.cell(row=row, column=2).border = mborder()
        wc(ws, row, 3, sev, bold=True, size=FONT_SIZE, bg=s_bg, align="center", border=True)
        wc(ws, row, 4, like, bold=True, size=FONT_SIZE, bg=l_bg, align="center", border=True)
        wc(ws, row, 5, detail, size=FONT_SIZE, bg=bg, align="left", wrap=True, border=True)
        ws.row_dimensions[row].height = 56
        row += 1

    row += 1
    sec_hdr(ws, row, 2, "Risk Summary Matrix", span=4, bg=DARK_BG)
    row += 1
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
    wc(ws, row, 2,
       "OVERALL RISK LEVEL: MODERATE  |  eBay's capital-light model and dominant recommerce positioning "
       "provide resilience. The key near-term risks are competition (Temu/Shein) and the EU de minimis "
       "change. Leverage is manageable but returning more cash than earned is a structural concern. "
       "Governance/compliance needs monitoring. Net: eBay is a mature platform with identifiable risks, "
       "not a high-growth company with existential threats.",
       size=FONT_SIZE, bg=ALT_ROW, align="left", wrap=True, border=True, bold=True)
    ws.row_dimensions[row].height = 56


# ═══════════════════════════════════════════════════════════════════════════════
def build_valuation(wb):
    ws = wb.create_sheet("Valuation")
    ws.sheet_view.showGridLines = False
    scw(ws, 1, 3); scw(ws, 2, 36); scw(ws, 3, 22); scw(ws, 4, 22); scw(ws, 5, 22); scw(ws, 6, 22); scw(ws, 7, 22)

    row = 2
    sec_hdr(ws, row, 2, "Valuation Analysis — eBay Inc. (EBAY)", span=6, bg=EBAY_BLUE)
    row += 2

    # Current market snapshot
    sec_hdr(ws, row, 2, "Current Market Snapshot (April 2026)", span=6, bg=EBAY_RED)
    row += 1
    snap = [
        ("Stock Price",             "$96.01",    "As of April 7, 2026"),
        ("Diluted Shares",          "~476M",     "FY2025 diluted weighted average"),
        ("Market Capitalization",   "~$45.7B",   "Price × Shares"),
        ("Net Debt",                "~$5.4B",    "LT Debt $6.7B – Cash $4.8B + Operating lease"),
        ("Enterprise Value",        "~$51.1B",   "Market Cap + Net Debt"),
        ("FY2025 Revenue",          "$11.1B",    "+7.9% YoY"),
        ("FY2025 EBITDA",           "$2.70B",    "24.3% margin"),
        ("FY2025 GAAP EPS",         "$4.26",     "+7.8% YoY"),
        ("FY2025 Non-GAAP EPS",     "$5.52",     "+13.1% YoY"),
        ("FY2025 FCF",              "$1.47B",    "13.3% FCF margin"),
    ]
    sub_hdr(ws, row, 2, "Metric"); sub_hdr(ws, row, 3, "Value"); sub_hdr(ws, row, 4, "Notes")
    row += 1
    for i, (m, v, n) in enumerate(snap):
        bg = alt(i)
        lbl(ws, row, 2, m, bg=bg)
        wc(ws, row, 3, v, bold=True, size=FONT_SIZE, bg=bg, align="right", border=True)
        c = wc(ws, row, 4, n, size=FONT_SIZE, bg=bg, align="left", border=True)
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=7)
        row += 1

    row += 1
    # Trading multiples
    sec_hdr(ws, row, 2, "Trading Multiples vs. Peers", span=6, bg=EBAY_GREEN, fg=WHITE)
    row += 1
    sub_hdr(ws, row, 2, "Multiple"); sub_hdr(ws, row, 3, "eBay (TTM)"); sub_hdr(ws, row, 4, "eBay (FWD)"); sub_hdr(ws, row, 5, "Etsy"); sub_hdr(ws, row, 6, "Amazon"); sub_hdr(ws, row, 7, "Sector Avg")
    row += 1
    mults = [
        ("P/E",           "22.5x",  "15.5x",  "~24x",  "~45x",  "~20x"),
        ("EV / EBITDA",   "18.9x",  "16.2x",  "~18x",  "~30x",  "~18x"),
        ("EV / Revenue",  "4.6x",   "4.2x",   "~5.0x", "~3.5x", "~3.8x"),
        ("P / FCF",       "31.1x",  "26.0x",  "~22x",  "~50x",  "~22x"),
        ("FCF Yield",     "3.2%",   "3.8%",   "~4.5%", "~2.0%", "~4.5%"),
        ("Dividend Yield","1.1%",   "1.1%",   "N/A",   "N/A",   "~1.0%"),
        ("P / GMV",       "0.57x",  "0.53x",  "0.22x", "N/M",   "N/M"),
    ]
    for i, row_data in enumerate(mults):
        bg = alt(i)
        for col, val in enumerate(row_data, start=2):
            wc(ws, row, col, val, size=FONT_SIZE, bg=bg, align="right" if col > 2 else "left", border=True)
        row += 1

    row += 1
    # Historical valuation
    sec_hdr(ws, row, 2, "Historical Valuation Bands (P/E)", span=6, bg=EBAY_YELLOW, fg=DARK_BG)
    row += 1
    sub_hdr(ws, row, 2, "Year"); sub_hdr(ws, row, 3, "Avg P/E"); sub_hdr(ws, row, 4, "Low P/E"); sub_hdr(ws, row, 5, "High P/E"); sub_hdr(ws, row, 6, "EPS"); sub_hdr(ws, row, 7, "Avg Price")
    row += 1
    hist_pe = [
        ("2021", "12.1x", "9.5x",  "16.2x", "$3.84",  "~$46.5"),
        ("2022", "10.5x", "8.0x",  "14.0x", "$2.56",  "~$26.9"),
        ("2023", "8.2x",  "6.5x",  "10.5x", "$3.37",  "~$27.6"),
        ("2024", "15.0x", "11.5x", "20.0x", "$4.05",  "~$60.8"),
        ("2025", "19.5x", "14.5x", "24.0x", "$4.26",  "~$83.1"),
        ("2026 YTD","22.5x","19.0x","26.0x", "$4.26E", "~$96.0"),
    ]
    for i, row_data in enumerate(hist_pe):
        bg = alt(i)
        for col, val in enumerate(row_data, start=2):
            wc(ws, row, col, val, size=FONT_SIZE, bg=bg, align="right" if col > 2 else "left", border=True)
        row += 1

    row += 1
    # DCF Model
    sec_hdr(ws, row, 2, "DCF Valuation Model", span=6, bg=DARK_BG)
    row += 1
    sub_hdr(ws, row, 2, "Assumptions")
    row += 1
    dcf_assum = [
        ("Base FCF (FY2025)",            "$1,475M",  "Reported free cash flow from continuing operations"),
        ("WACC",                         "9.0%",     "Risk-free 4.3% + equity risk premium + beta adjustment"),
        ("Terminal Growth Rate",         "3.0%",     "Long-run nominal GDP growth proxy"),
        ("Discount Period",              "10 years", ""),
        ("Net Debt",                     "$5,448M",  "Used to bridge EV → Equity Value"),
        ("Diluted Shares",               "476M",     "FY2025 weighted average"),
    ]
    for i, (a, v, n) in enumerate(dcf_assum):
        bg = alt(i)
        lbl(ws, row, 2, a, bg=bg)
        wc(ws, row, 3, v, bold=True, size=FONT_SIZE, bg=bg, align="right", border=True)
        c = wc(ws, row, 4, n, size=FONT_SIZE, bg=bg, align="left", border=True)
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=7)
        row += 1

    row += 1
    # Year-by-year DCF
    sub_hdr(ws, row, 2, "Year"); sub_hdr(ws, row, 3, "FCF ($M)"); sub_hdr(ws, row, 4, "Growth"); sub_hdr(ws, row, 5, "PV Factor"); sub_hdr(ws, row, 6, "PV of FCF ($M)"); sub_hdr(ws, row, 7, "Cumul. PV ($M)")
    row += 1
    # FCF projections
    fcf_base = 1475
    growth_rates = [0.10, 0.08, 0.07, 0.06, 0.05, 0.05, 0.04, 0.04, 0.04, 0.03]
    wacc = 0.09
    fcfs = []
    cur = fcf_base
    for g in growth_rates:
        cur = cur * (1 + g)
        fcfs.append(cur)

    pv_sum = 0
    dcf_years = []
    for yr_idx, (fcf_val, g) in enumerate(zip(fcfs, growth_rates), start=1):
        pv_factor = 1 / ((1 + wacc) ** yr_idx)
        pv = fcf_val * pv_factor
        pv_sum += pv
        dcf_years.append((f"Year {yr_idx} ({2025+yr_idx})", round(fcf_val), f"{g*100:.0f}%",
                           f"{pv_factor:.4f}", round(pv), round(pv_sum)))

    for i, row_data in enumerate(dcf_years):
        bg = alt(i)
        for col, val in enumerate(row_data, start=2):
            wc(ws, row, col, val, size=FONT_SIZE, bg=bg, align="right" if col > 2 else "left", border=True)
        row += 1

    # Terminal value
    tv_fcf = fcfs[-1]
    tv = tv_fcf * (1 + 0.03) / (wacc - 0.03)
    pv_tv = tv / ((1 + wacc) ** 10)
    total_ev = pv_sum + pv_tv
    equity_val = total_ev - 5448
    per_share = equity_val / 476

    tv_data = [
        ("Terminal FCF (Year 10 × 1.03)",  f"${tv_fcf*1.03/1e3:.2f}B", "", "", "", ""),
        ("Terminal Value",                  f"${tv/1e3:.2f}B",           "", "", "", ""),
        ("PV of Terminal Value",            f"${pv_tv/1e3:.2f}B",        "", "", "", ""),
        ("PV of FCF (Yrs 1–10)",            f"${pv_sum/1e3:.2f}B",       "", "", "", ""),
        ("ENTERPRISE VALUE",                f"${total_ev/1e3:.2f}B",     "", "", "", ""),
        ("Less: Net Debt",                  f"-$5.45B",                  "", "", "", ""),
        ("EQUITY VALUE",                    f"${equity_val/1e3:.2f}B",   "", "", "", ""),
        ("INTRINSIC VALUE / SHARE",         f"${per_share:.2f}",         "", "", "", ""),
        ("Current Price",                   "$96.01",                    "", "", "", ""),
        ("Upside / (Downside)",             f"{((per_share/96.01)-1)*100:.1f}%", "", "", "", ""),
    ]
    row += 1
    for i, row_data in enumerate(tv_data):
        bold = row_data[0] in ("ENTERPRISE VALUE","EQUITY VALUE","INTRINSIC VALUE / SHARE","Upside / (Downside)")
        bg = LIGHT_GRAY if bold else alt(i)
        color = GREEN if bold and row_data[0] == "INTRINSIC VALUE / SHARE" and per_share > 96 else (RED if bold and per_share <= 96 else "000000")
        lbl(ws, row, 2, row_data[0], bg=bg)
        ws.cell(row=row, column=2).font = mf(bold=bold, size=FONT_SIZE)
        ws.cell(row=row, column=2).fill = mfill(bg)
        ws.cell(row=row, column=2).border = mborder()
        wc(ws, row, 3, row_data[1], bold=bold, size=FONT_SIZE, fg=color, bg=bg, align="right", border=True)
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=7)
        row += 1

    row += 1
    # Sensitivity table — FCF-based
    sec_hdr(ws, row, 2, "Sensitivity Analysis — FCF-Based DCF: Intrinsic Value/Share ($)", span=6, bg=EBAY_BLUE)
    row += 1
    sub_hdr(ws, row, 2, "WACC \\ Terminal Growth"); sub_hdr(ws, row, 3, "2.0%"); sub_hdr(ws, row, 4, "2.5%"); sub_hdr(ws, row, 5, "3.0%"); sub_hdr(ws, row, 6, "3.5%"); sub_hdr(ws, row, 7, "4.0%")
    row += 1
    waccs = [0.08, 0.085, 0.09, 0.095, 0.10]
    tgrs = [0.02, 0.025, 0.03, 0.035, 0.04]
    for w in waccs:
        lbl(ws, row, 2, f"{w*100:.1f}%", bg=WHITE)
        ws.cell(row=row, column=2).border = mborder()
        for col_idx, tg in enumerate(tgrs, start=3):
            c_fcfs = []
            c_cur = fcf_base
            for g in growth_rates:
                c_cur = c_cur * (1 + g)
                c_fcfs.append(c_cur)
            c_pv_sum = sum(f / ((1+w)**yr) for yr, f in enumerate(c_fcfs, start=1))
            c_tv = c_fcfs[-1] * (1+tg) / (w - tg) if w > tg else 0
            c_pv_tv = c_tv / ((1+w)**10)
            c_ev = c_pv_sum + c_pv_tv
            c_eq = c_ev - 5448
            c_ps = c_eq / 476
            bg = POSITIVE if c_ps > 96 else (WARN_YELLOW if c_ps > 75 else NEGATIVE)
            wc(ws, row, col_idx, f"${c_ps:.0f}", size=FONT_SIZE, bg=bg, align="center", border=True, bold=(abs(w-0.09)<0.001 and abs(tg-0.03)<0.001))
        row += 1

    row += 1
    # Owner Earnings DCF — higher-conviction scenario
    oe_base = 2682  # Owner Earnings from the Return on Capital sheet FY2025
    sec_hdr(ws, row, 2, "Owner Earnings DCF (Buffett Method) — Base: $2,682M", span=6, bg=EBAY_GREEN, fg=WHITE)
    row += 1
    sub_hdr(ws, row, 2, "WACC \\ Terminal Growth"); sub_hdr(ws, row, 3, "2.0%"); sub_hdr(ws, row, 4, "2.5%"); sub_hdr(ws, row, 5, "3.0%"); sub_hdr(ws, row, 6, "3.5%"); sub_hdr(ws, row, 7, "4.0%")
    row += 1
    for w in waccs:
        lbl(ws, row, 2, f"{w*100:.1f}%", bg=WHITE)
        ws.cell(row=row, column=2).border = mborder()
        for col_idx, tg in enumerate(tgrs, start=3):
            c_fcfs = []
            c_cur = oe_base
            for g in growth_rates:
                c_cur = c_cur * (1 + g)
                c_fcfs.append(c_cur)
            c_pv_sum = sum(f / ((1+w)**yr) for yr, f in enumerate(c_fcfs, start=1))
            c_tv = c_fcfs[-1] * (1+tg) / (w - tg) if w > tg else 0
            c_pv_tv = c_tv / ((1+w)**10)
            c_ev = c_pv_sum + c_pv_tv
            c_eq = c_ev - 5448
            c_ps = c_eq / 476
            bg = POSITIVE if c_ps > 96 else (WARN_YELLOW if c_ps > 75 else NEGATIVE)
            wc(ws, row, col_idx, f"${c_ps:.0f}", size=FONT_SIZE, bg=bg, align="center", border=True, bold=(abs(w-0.09)<0.001 and abs(tg-0.03)<0.001))
        row += 1

    row += 1
    # Owner earnings DCF individual rows
    oe_fcfs = []
    c_cur = oe_base
    for g in growth_rates:
        c_cur = c_cur * (1 + g)
        oe_fcfs.append(c_cur)
    oe_pv_sum = sum(f / ((1+0.09)**yr) for yr, f in enumerate(oe_fcfs, start=1))
    oe_tv = oe_fcfs[-1] * 1.03 / 0.06
    oe_pv_tv = oe_tv / (1.09**10)
    oe_ev = oe_pv_sum + oe_pv_tv
    oe_eq = oe_ev - 5448
    oe_ps = oe_eq / 476

    oe_summary = [
        ("Owner Earnings Base (FY2025)",   f"$2,682M",         "Net Income + D&A + SBC – Maint. CapEx – WC needs"),
        ("PV of Owner Earnings (10 yrs)",  f"${oe_pv_sum/1e3:.2f}B", "Discounted at 9% WACC"),
        ("PV of Terminal Value",           f"${oe_pv_tv/1e3:.2f}B", "Terminal growth 3%, WACC 9%"),
        ("Enterprise Value",               f"${oe_ev/1e3:.2f}B",     "PV FCF + PV Terminal"),
        ("Equity Value",                   f"${oe_eq/1e3:.2f}B",     "EV – Net Debt $5.45B"),
        ("INTRINSIC VALUE / SHARE (OE)",   f"${oe_ps:.2f}",          f"vs. Market Price $96.01  →  {'UNDERVALUED +{:.0f}%'.format((oe_ps/96.01-1)*100) if oe_ps>96 else 'OVERVALUED {:.0f}%'.format((oe_ps/96.01-1)*100)}"),
    ]
    sub_hdr(ws, row, 2, "Component"); sub_hdr(ws, row, 3, "Value");
    c = sub_hdr(ws, row, 4, "Notes"); ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=7)
    row += 1
    for i, (a, v, n) in enumerate(oe_summary):
        bold = "INTRINSIC" in a
        bg = LIGHT_GRAY if bold else alt(i)
        lbl(ws, row, 2, a, bg=bg)
        ws.cell(row=row, column=2).font = mf(bold=bold, size=FONT_SIZE)
        ws.cell(row=row, column=2).fill = mfill(bg)
        ws.cell(row=row, column=2).border = mborder()
        wc(ws, row, 3, v, bold=bold, size=FONT_SIZE, bg=bg, align="right", border=True)
        c = wc(ws, row, 4, n, size=FONT_SIZE, bg=bg, align="left", border=True)
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=7)
        row += 1

    row += 1
    # Valuation conclusion
    sec_hdr(ws, row, 2, "Valuation Conclusion", span=6, bg=EBAY_RED)
    row += 1
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
    wc(ws, row, 2,
       f"FCF-BASED DCF (WACC 9%, TGR 3%): ${per_share:.0f}/share  |  "
       f"OWNER EARNINGS DCF (WACC 9%, TGR 3%): ${oe_ps:.0f}/share  |  "
       f"Analyst Consensus Target: $97–98  |  Citigroup (Buy): $114  |  Simply Wall St DCF: $117  |  GuruFocus GF Value: $71.\n\n"
       f"VERDICT: At $96, eBay is FAIRLY VALUED to MODESTLY OVERVALUED on a conservative FCF basis "
       f"(${per_share:.0f} base DCF). However, the market is pricing eBay on OWNER EARNINGS (~$2.7B), "
       f"where intrinsic value is ${oe_ps:.0f}/share — suggesting the stock is "
       f"{'UNDERVALUED' if oe_ps > 96 else 'FAIRLY VALUED'} on an owner earnings basis. "
       f"The bear case risk (~$60) assumes FCF deterioration from competition. "
       f"LIMITED MARGIN OF SAFETY at current prices for a value investor; "
       f"FAIR ENTRY for a growth/quality investor with a 3–5 year horizon.",
       size=FONT_SIZE, bg=ALT_ROW, align="left", wrap=True, border=True, bold=True)
    ws.row_dimensions[row].height = 100


# ═══════════════════════════════════════════════════════════════════════════════
def build_market_sentiment(wb):
    ws = wb.create_sheet("Market Sentiment")
    ws.sheet_view.showGridLines = False
    scw(ws, 1, 3); scw(ws, 2, 32); scw(ws, 3, 22); scw(ws, 4, 46)

    row = 2
    sec_hdr(ws, row, 2, "Market Sentiment — eBay Inc. (EBAY)", span=3, bg=DARK_BG)
    row += 2

    # Analyst ratings
    sec_hdr(ws, row, 2, "Analyst Coverage & Ratings (April 2026)", span=3, bg=EBAY_BLUE)
    row += 1
    ratings = [
        ("Total Analysts Covering",    "26",           "Broad coverage — large-cap marketplace"),
        ("Consensus Rating",           "HOLD",         "Balanced view; strong execution but limited upside at current valuation"),
        ("Average Price Target",       "$97.27",       "~1.3% upside from $96.01"),
        ("High Price Target",          "$114",         "Citigroup — Buy; cites advertising & recommerce acceleration"),
        ("Low Price Target",           "$72",          "Bearish thesis: competition pressure, FCF decline"),
        ("Buys / Holds / Sells",       "8 / 14 / 4",  "More Holds than Buys — cautious optimism"),
        ("Recent Upgrades",            "Citigroup ↑",  "Raised target to $114; cites Q4 2025 beat and Depop optionality"),
        ("Recent Downgrades",          "None (90 days)","No major downgrades post Q4 2025 earnings"),
    ]
    sub_hdr(ws, row, 2, "Metric"); sub_hdr(ws, row, 3, "Value"); sub_hdr(ws, row, 4, "Context")
    row += 1
    for i, (m, v, n) in enumerate(ratings):
        bg = alt(i)
        lbl(ws, row, 2, m, bg=bg)
        rating_bg = bg
        if v in ("HOLD",): rating_bg = WARN_YELLOW
        elif v in ("BUY","STRONG BUY"): rating_bg = POSITIVE
        elif v in ("SELL",): rating_bg = NEGATIVE
        wc(ws, row, 3, v, bold=True, size=FONT_SIZE, bg=rating_bg, align="center", border=True)
        wc(ws, row, 4, n, size=FONT_SIZE, bg=bg, align="left", wrap=True, border=True)
        ws.row_dimensions[row].height = 24
        row += 1

    row += 1
    # Market sentiment indicators
    sec_hdr(ws, row, 2, "Market Sentiment Indicators", span=3, bg=EBAY_GREEN, fg=WHITE)
    row += 1
    sent = [
        ("Short Interest % Float",    "~2.5%",      "LOW — market not betting against eBay; limited downside speculation"),
        ("Institutional Ownership",   "~83%",       "HIGH — stable institutional base; Vanguard 9.2%, BlackRock 7.1%"),
        ("Retail Sentiment (Reddit/X)","POSITIVE",  "Strong positive reaction to Q4 2025 GMV beat (+10%). "
                                                    "Recommerce narrative resonating with retail investors."),
        ("Options Market (IV)",       "MODERATE",   "~30% implied volatility — pricing moderate uncertainty; "
                                                    "options traders not extreme bulls or bears"),
        ("Relative Strength vs. S&P", "OUTPERFORM", "EBAY +28% YTD 2025 vs. S&P 500 +12%. "
                                                    "Momentum has been positive since Q2 2025 acceleration."),
        ("52-Week Range",             "$63 – $105",  "Trading near upper half of 52-week range at $96."),
        ("Price vs. 200-day MA",      "ABOVE",       "Bullish signal; trading ~15% above 200-day moving average"),
        ("Beta",                      "~1.0",        "Market-neutral volatility; moves roughly with S&P 500"),
    ]
    sub_hdr(ws, row, 2, "Indicator"); sub_hdr(ws, row, 3, "Reading"); sub_hdr(ws, row, 4, "Interpretation")
    row += 1
    for i, (ind, val, interp) in enumerate(sent):
        bg = alt(i)
        sent_bg = POSITIVE if any(x in val.upper() for x in ["LOW","HIGH","OUTPERFORM","ABOVE","POSITIVE"]) else (NEGATIVE if "BEAR" in val.upper() else bg)
        lbl(ws, row, 2, ind, bg=bg)
        wc(ws, row, 3, val, bold=True, size=FONT_SIZE, bg=sent_bg, align="center", border=True)
        wc(ws, row, 4, interp, size=FONT_SIZE, bg=bg, align="left", wrap=True, border=True)
        ws.row_dimensions[row].height = 30
        row += 1

    row += 1
    # Key narratives
    sec_hdr(ws, row, 2, "Prevailing Market Narratives (April 2026)", span=3, bg=EBAY_RED)
    row += 1
    narratives = [
        ("BULL CASE — GMV Re-acceleration",
         "POSITIVE",
         "Q4 2025 GMV of +10% was a significant acceleration from +1% in Q1. "
         "Recommerce structural tailwind is real. Advertising growing >17% annually. "
         "Depop adds Gen Z demographic. Management execution improving."),
        ("BULL CASE — Capital Returns",
         "POSITIVE",
         "$3B returned in 2025 (6.5% yield at $46B cap). "
         "Share count declining ~5% annually. Dividend growing. "
         "FCF yield expanding as buybacks reduce share count."),
        ("BEAR CASE — Temu/Shein Erosion",
         "NEGATIVE",
         "Price-sensitive eBay buyers migrating to Temu. "
         "eBay's cost-leadership in low-end goods is permanently impaired. "
         "Focus category pivot may not offset volume loss."),
        ("BEAR CASE — Leverage Trap",
         "NEGATIVE",
         "Returning $3B/yr vs $1.5B FCF requires debt or cash drawdown. "
         "Long-term debt $6.7B and growing. If growth stalls, "
         "eBay may need to cut returns, spooking income investors."),
        ("NEUTRAL — Mature Platform",
         "NEUTRAL",
         "eBay is not a high-growth story. It's a durable, capital-light platform "
         "growing at 7–10% GMV. The stock re-rated higher in 2025 on execution. "
         "At ~16x forward EPS, it's priced for steady growth, not acceleration."),
    ]
    sub_hdr(ws, row, 2, "Narrative"); sub_hdr(ws, row, 3, "Signal"); sub_hdr(ws, row, 4, "Detail")
    row += 1
    for i, (name, signal, detail) in enumerate(narratives):
        bg = alt(i)
        sig_bg = POSITIVE if "POSITIVE" in signal else (NEGATIVE if "NEGATIVE" in signal else WARN_YELLOW)
        lbl(ws, row, 2, name, bg=bg)
        ws.cell(row=row, column=2).font = mf(bold=True, size=FONT_SIZE)
        ws.cell(row=row, column=2).fill = mfill(bg)
        ws.cell(row=row, column=2).border = mborder()
        wc(ws, row, 3, signal, bold=True, size=FONT_SIZE, bg=sig_bg, align="center", border=True)
        wc(ws, row, 4, detail, size=FONT_SIZE, bg=bg, align="left", wrap=True, border=True)
        ws.row_dimensions[row].height = 56
        row += 1


# ═══════════════════════════════════════════════════════════════════════════════
def build_key_indicators(wb):
    ws = wb.create_sheet("Key Indicators")
    ws.sheet_view.showGridLines = False
    scw(ws, 1, 3); scw(ws, 2, 36); scw(ws, 3, 18); scw(ws, 4, 18); scw(ws, 5, 18); scw(ws, 6, 18); scw(ws, 7, 18)

    row = 2
    sec_hdr(ws, row, 2, "Key Performance Indicators — eBay Inc.", span=6, bg=EBAY_BLUE)
    row += 1
    sub_hdr(ws, row, 2, "KPI")
    for col, yr in enumerate(["FY2021","FY2022","FY2023","FY2024","FY2025"], start=3):
        sub_hdr(ws, row, col, yr)
    row += 1

    kpis = [
        ("MARKETPLACE SCALE",            True,  [None]*5),
        ("Active Buyers (M)",             False, [154,   138,   132,   133,   135]),
        ("YoY Active Buyer Growth",       False, ["+1%", "-10%","-4%", "+1%", "+1.5%"]),
        ("Gross Merchandise Volume ($B)", True,  [87.9,  73.9,  73.2,  74.6,  79.6]),
        ("YoY GMV Growth",                False, ["+21%","-16%","-1%", "+2%", "+7%"]),
        ("Live Listings (B)",             False, ["~1.5","~1.7","~2.0","~2.2","~2.5"]),
        ("Active Sellers (M)",            False, ["~19", "~18", "~18", "~18", "~18"]),
        ("",                              False, [None]*5),
        ("REVENUE & TAKE RATE",          True,  [None]*5),
        ("Total Revenue ($M)",            True,  [10421, 9795,  10112, 10283, 11100]),
        ("Take Rate (Rev/GMV)",           False, ["11.9%","13.3%","13.8%","13.8%","13.9%"]),
        ("Advertising Revenue ($M)",      False, ["~900","~1,200","~1,450","~1,730","~2,000"]),
        ("Advertising % of Revenue",      False, ["~8.6%","~12.2%","~14.3%","~16.8%","~18.0%"]),
        ("1P Ad Revenue Growth",          False, ["N/A", "+30%", "+20%", "+19%", "+17%"]),
        ("Ad Penetration (% GMV)",        False, ["~1.0%","~1.6%","~2.0%","~2.3%","~2.5%"]),
        ("",                              False, [None]*5),
        ("PER-BUYER ECONOMICS",          True,  [None]*5),
        ("Revenue per Active Buyer",      False, ["$67.7","$71.0","$76.6","$77.3","$82.2"]),
        ("GMV per Active Buyer",          False, ["$571","$535","$555","$561","$589"]),
        ("Transactions per Buyer (est.)", False, ["~4.5","~4.6","~4.7","~4.7","~4.8"]),
        ("",                              False, [None]*5),
        ("PROFITABILITY",                True,  [None]*5),
        ("Gross Margin %",                False, ["71.2%","71.3%","71.0%","71.0%","71.5%"]),
        ("Operating Margin % (GAAP)",     False, ["28.1%","24.0%","19.2%","22.5%","20.5%"]),
        ("Operating Margin % (Non-GAAP)", False, ["32.0%","28.1%","24.5%","25.9%","25.6%"]),
        ("Net Margin % (GAAP)",           False, ["22.5%","14.1%","16.5%","18.7%","18.3%"]),
        ("FCF Margin %",                  False, ["24.3%","19.0%","15.9%","16.1%","13.3%"]),
        ("",                              False, [None]*5),
        ("CAPITAL EFFICIENCY",           True,  [None]*5),
        ("Return on Assets (ROA)",        False, ["9.7%", "6.7%", "8.4%","10.3%","11.5%"]),
        ("Return on Equity (ROE)",        False, ["41%",  "46%",  "83%", "80%", "127%"]),
        ("ROIC",                          False, ["15.2%","12.1%","10.1%","13.5%","13.9%"]),
        ("CapEx % of Revenue",            False, ["4.5%", "5.3%", "5.3%", "4.5%", "4.7%"]),
        ("FCF Conversion (FCF/Net Inc.)", False, ["107.9%","134.7%","96.5%","85.7%","72.7%"]),
        ("",                              False, [None]*5),
        ("SHAREHOLDER VALUE",            True,  [None]*5),
        ("Shares Repurchased ($B)",       False, ["$2.28","$2.99","$3.40","$2.56","$2.50"]),
        ("Dividends Paid ($M)",           False, [332,    432,    468,    519,    531]),
        ("Total Shareholder Return ($B)", False, ["$2.61","$3.42","$3.87","$3.08","$3.03"]),
        ("Buyback Yield",                 False, ["~4.8%","~12%", "~14%", "~9%",  "~5.5%"]),
        ("Shares Outstanding (M)",        False, [609,    540,    495,    476,    476]),
        ("YoY Share Count Change",        False, ["N/A", "-11.3%","-8.3%","-3.8%", "0.0%"]),
        ("",                              False, [None]*5),
        ("BALANCE SHEET HEALTH",         True,  [None]*5),
        ("Net Debt ($M)",                 False, [4557,  6576,  6548,  5467,  5448]),
        ("Net Debt / EBITDA",             False, ["1.3x","2.3x","2.6x","1.9x","2.0x"]),
        ("Interest Coverage (EBIT/Int)",  False, ["15.6x","15.3x","11.0x","13.2x","13.8x"]),
        ("Current Ratio",                 False, ["1.35x","0.98x","0.75x","0.84x","0.81x"]),
    ]

    for i, (label, bold, vals) in enumerate(kpis):
        if not label:
            row += 1
            continue
        bg = LIGHT_GRAY if bold else alt(i)
        lbl(ws, row, 2, label, bg=bg)
        ws.cell(row=row, column=2).font = mf(bold=bold, size=FONT_SIZE)
        ws.cell(row=row, column=2).fill = mfill(bg)
        ws.cell(row=row, column=2).border = mborder()
        for col, val in enumerate(vals, start=3):
            dc(ws, row, col, val, bg=bg,
               num_fmt="#,##0" if isinstance(val, (int, float)) else None, bold=bold)
        row += 1

    # eBay is NOT SaaS — note this explicitly
    row += 1
    sec_hdr(ws, row, 2, "SaaS Rule of 50 — Not Applicable", span=6, bg=EBAY_YELLOW, fg=DARK_BG)
    row += 1
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
    wc(ws, row, 2,
       "eBay is a MARKETPLACE / E-COMMERCE platform, NOT a SaaS company. The Rule of 50 "
       "(Revenue Growth % + FCF Margin % >= 50) is not the primary framework for marketplace valuation. "
       "For reference: eBay's FY2025 figures are Revenue Growth 7.9% + FCF Margin 13.3% = 21.2 — "
       "well below Rule of 50, which is EXPECTED for a mature marketplace (not a high-growth SaaS). "
       "Appropriate valuation frameworks: GMV multiples, EV/EBITDA, P/FCF, and DCF.",
       size=FONT_SIZE, bg=WARN_YELLOW, align="left", wrap=True, border=True, italic=True)
    ws.row_dimensions[row].height = 56


# ═══════════════════════════════════════════════════════════════════════════════
def main():
    output_dir = "/Users/naji/WORK/github.com/AI/claude/Agent/MarketResearch/output"
    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(output_dir, "EBAY_Financial_Analysis_2026.xlsx")

    wb = openpyxl.Workbook()
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    print("Building Cover...")
    build_cover(wb)
    print("Building Business Overview...")
    build_business_overview(wb)
    print("Building Moat...")
    build_moat(wb)
    print("Building Income Statement...")
    build_income_statement(wb)
    print("Building Balance Sheet...")
    build_balance_sheet(wb)
    print("Building Cash Flow...")
    build_cash_flow(wb)
    print("Building Return on Capital...")
    build_return_on_capital(wb)
    print("Building Management...")
    build_management(wb)
    print("Building Risks...")
    build_risks(wb)
    print("Building Valuation...")
    build_valuation(wb)
    print("Building Market Sentiment...")
    build_market_sentiment(wb)
    print("Building Key Indicators...")
    build_key_indicators(wb)

    wb.save(output_path)
    print(f"\nSaved: {output_path}")
    return output_path

if __name__ == "__main__":
    main()
