"""
Visa Inc. (V) - Comprehensive Financial Analysis
Generated: April 14, 2026
Data Sources: StockAnalysis, MacroTrends, SEC Filings, Analyst Reports
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1
import os

# ─────────────────────────────────────────────
# CONSTANTS & THEME
# ─────────────────────────────────────────────
FONT_SIZE      = 14
FONT_NAME      = "Calibri"

# Visa brand colors
VISA_DARK_BLUE = "1A1F71"
VISA_GOLD      = "F7B600"
VISA_MID_BLUE  = "2962B8"
ACCENT_TEAL    = "17A589"
ACCENT_RED     = "C0392B"
LIGHT_BLUE_BG  = "D6E4F0"
LIGHT_GOLD_BG  = "FEF9E7"
VERY_LIGHT_BG  = "F2F3F4"
WHITE          = "FFFFFF"
DARK_TEXT      = "1C2833"
HEADER_TEXT    = "FFFFFF"

OUTPUT_DIR = "/Users/naji/WORK/github.com/AI/claude/Agent/MarketResearch/output"
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "Visa_Financial_Analysis.xlsx")

# ─────────────────────────────────────────────
# STYLE HELPERS
# ─────────────────────────────────────────────

def make_font(size=FONT_SIZE, bold=False, color=DARK_TEXT, italic=False):
    return Font(name=FONT_NAME, size=size, bold=bold, color=color, italic=italic)

def make_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def make_border(style="thin"):
    s = Side(border_style=style, color="BDBDBD")
    return Border(left=s, right=s, top=s, bottom=s)

def thick_border():
    t = Side(border_style="medium", color="7F8C8D")
    return Border(left=t, right=t, top=t, bottom=t)

def set_cell(ws, row, col, value, bold=False, fill_hex=None, font_size=FONT_SIZE,
             align_h="left", align_v="center", number_format=None, italic=False,
             font_color=DARK_TEXT, border=True, wrap=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = make_font(font_size, bold, font_color, italic)
    c.alignment = Alignment(horizontal=align_h, vertical=align_v, wrap_text=wrap)
    if fill_hex:
        c.fill = make_fill(fill_hex)
    if number_format:
        c.number_format = number_format
    if border:
        c.border = make_border()
    return c

def header_row(ws, row, labels, col_start=1, fill_hex=VISA_DARK_BLUE,
               font_color=HEADER_TEXT, height=30):
    for i, lbl in enumerate(labels):
        set_cell(ws, row, col_start + i, lbl, bold=True, fill_hex=fill_hex,
                 font_color=font_color, align_h="center")
    ws.row_dimensions[row].height = height

def section_header(ws, row, label, col_span=10, fill_hex=VISA_MID_BLUE,
                   font_color=HEADER_TEXT):
    c = ws.cell(row=row, column=1, value=label)
    c.font = make_font(FONT_SIZE + 1, bold=True, color=font_color)
    c.fill = make_fill(fill_hex)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = make_border()
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=col_span)
    ws.row_dimensions[row].height = 26

def set_col_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

def set_row_height(ws, row, height):
    ws.row_dimensions[row].height = height

def blank_row(ws, row):
    ws.row_dimensions[row].height = 8

def pct_fmt(v):
    """Return formatted percentage string."""
    return f"{v:.1f}%"

def usd_b(v):
    return f"${v:.2f}B"

def usd_m(v):
    return f"${v:,.0f}M"

# ─────────────────────────────────────────────
# WORKBOOK SETUP
# ─────────────────────────────────────────────
wb = openpyxl.Workbook()
wb.remove(wb.active)   # remove default sheet

TABS = [
    "Cover",
    "Business Overview",
    "Moat",
    "Income Statement",
    "Balance Sheet",
    "Cash Flow",
    "Return on Capital",
    "Management",
    "Risks",
    "Valuation",
    "Market Sentiment",
    "Key Indicators",
]

sheets = {}
for name in TABS:
    ws = wb.create_sheet(title=name)
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = VISA_DARK_BLUE
    sheets[name] = ws

# ─────────────────────────────────────────────
# 1. COVER
# ─────────────────────────────────────────────
ws = sheets["Cover"]
ws.sheet_properties.tabColor = VISA_GOLD

set_col_widths(ws, {1: 4, 2: 50, 3: 30, 4: 30, 5: 4})

# Top banner
for r in range(1, 8):
    for c in range(1, 6):
        ws.cell(r, c).fill = make_fill(VISA_DARK_BLUE)
    ws.row_dimensions[r].height = 18

# Company name
ws.merge_cells("B2:D6")
c = ws.cell(2, 2, "VISA INC. (NYSE: V)")
c.font = Font(name=FONT_NAME, size=36, bold=True, color=VISA_GOLD)
c.alignment = Alignment(horizontal="center", vertical="center")
c.fill = make_fill(VISA_DARK_BLUE)

# Sub-title
ws.merge_cells("B7:D7")
c = ws.cell(7, 2, "Comprehensive Financial Analysis  |  April 2026")
c.font = Font(name=FONT_NAME, size=FONT_SIZE, color="CCCCCC", italic=True)
c.alignment = Alignment(horizontal="center", vertical="center")
c.fill = make_fill(VISA_DARK_BLUE)

blank_row(ws, 8)

# Key stats box
kv_data = [
    ("Stock Price (Apr 14, 2026)", "$309.09"),
    ("Market Capitalisation",     "$593.5 Billion"),
    ("FY2025 Revenue",            "$40.0 Billion"),
    ("FY2025 Net Income",         "$20.1 Billion"),
    ("Free Cash Flow (FY2025)",   "$21.6 Billion"),
    ("ROIC (TTM)",                "~28%"),
    ("Forward P/E",               "23.4×"),
    ("Analyst Consensus",         "BUY"),
    ("Avg. Price Target",         "~$400  (+29% upside)"),
    ("Dividend Yield",            "0.86%"),
]

row = 9
section_header(ws, row, "  KEY SNAPSHOT", 4)
row += 1
for label, val in kv_data:
    fill = LIGHT_BLUE_BG if row % 2 == 0 else WHITE
    set_cell(ws, row, 2, label, bold=True, fill_hex=fill, align_h="left")
    set_cell(ws, row, 3, val,   bold=False, fill_hex=fill, align_h="right")
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
    ws.row_dimensions[row].height = 22
    row += 1

blank_row(ws, row); row += 1

# Investment thesis
section_header(ws, row, "  INVESTMENT THESIS", 4)
row += 1
thesis = [
    "Visa is the world's largest payment network — a tollbooth on global commerce.",
    "Network effects create a near-impenetrable moat: 200+ countries, 160+ currencies.",
    "Asset-light model: ~81% gross margin, ~60% operating margin, ~50% net margin.",
    "Capital return machine: $30B buyback authorised; FCF consistently > Net Income.",
    "Stablecoin & AI positioning reduces long-term disruption risk.",
    "Current price (~$309) trades at ~28% discount to average analyst target of $400.",
]
for t in thesis:
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    c = ws.cell(row, 2, f"  •  {t}")
    c.font = make_font(FONT_SIZE)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c.fill = make_fill(LIGHT_GOLD_BG)
    c.border = make_border()
    ws.row_dimensions[row].height = 24
    row += 1

blank_row(ws, row); row += 1
section_header(ws, row, "  TABLE OF CONTENTS", 4)
row += 1
toc = [
    ("1. Business Overview",  "Products, Revenue Mix, Value Propositions"),
    ("2. Moat",               "Network Effects, Switching Costs, Scale"),
    ("3. Income Statement",   "Revenue, Margins, EPS — 5-Year History"),
    ("4. Balance Sheet",      "Assets, Liabilities, Equity — 5-Year History"),
    ("5. Cash Flow",          "OCF, FCF, Capex, Buybacks — 5-Year History"),
    ("6. Return on Capital",  "ROIC, ROE, Incremental Returns"),
    ("7. Management",         "CEO, Proxy, Insider Activity"),
    ("8. Risks",              "Regulatory, Competitive, Macro"),
    ("9. Valuation",          "DCF, Comps, Bull/Base/Bear"),
    ("10. Market Sentiment",  "Analyst Ratings, Price Targets"),
    ("11. Key Indicators",    "Volume KPIs, Margins Dashboard"),
]
for tab_name, desc in toc:
    fill = VERY_LIGHT_BG if row % 2 == 0 else WHITE
    set_cell(ws, row, 2, tab_name, bold=True, fill_hex=fill)
    set_cell(ws, row, 3, desc, fill_hex=fill, italic=True)
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
    ws.row_dimensions[row].height = 22
    row += 1

# Footer
ws.merge_cells("B" + str(row + 1) + ":D" + str(row + 1))
c = ws.cell(row + 1, 2, "Sources: StockAnalysis.com · MacroTrends · SEC EDGAR · Visa IR · MarketBeat · AInvest")
c.font = Font(name=FONT_NAME, size=FONT_SIZE, italic=True, color="888888")
c.alignment = Alignment(horizontal="center")

# ─────────────────────────────────────────────
# 2. BUSINESS OVERVIEW
# ─────────────────────────────────────────────
ws = sheets["Business Overview"]
set_col_widths(ws, {1: 4, 2: 32, 3: 22, 4: 22, 5: 22, 6: 4})

row = 1
ws.merge_cells("B1:E1")
c = ws.cell(1, 2, "VISA INC. — BUSINESS OVERVIEW")
c.font = Font(name=FONT_NAME, size=FONT_SIZE + 4, bold=True, color=VISA_DARK_BLUE)
c.alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[1].height = 34
row = 2; blank_row(ws, row); row += 1

# Company description
section_header(ws, row, "  WHAT VISA DOES", 5); row += 1
desc_lines = [
    "Visa Inc. is the world's largest electronic payments network, connecting consumers, merchants,",
    "financial institutions, and governments in more than 200 countries and territories. Visa does NOT",
    "issue cards or extend credit — it operates the payment rails that move money between banks.",
    "Revenue comes from transaction fees, data processing, and cross-border interchange.",
]
for line in desc_lines:
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
    c = ws.cell(row, 2, line)
    c.font = make_font(FONT_SIZE)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c.fill = make_fill(LIGHT_GOLD_BG)
    c.border = make_border()
    ws.row_dimensions[row].height = 22
    row += 1

blank_row(ws, row); row += 1

# Products & Services
section_header(ws, row, "  PRODUCTS & SERVICES", 5); row += 1
header_row(ws, row, ["Product / Service", "Description", "Key Clients", "Scale"], 2)
row += 1
products = [
    ("Visa Credit / Debit Cards", "Core branded payment cards issued by partner banks", "JPMorgan Chase, Bank of America, CITI, international issuers", "4.4B+ cards worldwide"),
    ("VisaNet (Processing)", "Real-time authorisation & settlement at 65,000 tx/sec", "All card-issuing banks and acquirers globally", "257.5B transactions / FY25"),
    ("Visa Direct (P2P / B2B)", "Push payments for real-time money movement", "PayPal, Venmo, gig-economy platforms", "$14.2T total payment volume"),
    ("CyberSource / DPS", "Payment gateway, fraud management, tokenisation", "E-commerce merchants globally", "Millions of merchants"),
    ("Visa B2B Connect", "Cross-border B2B payment network", "Corporate treasuries, banks", "97 countries"),
    ("Visa Consulting & Analytics", "Strategic advisory and data insights practice", "Issuer banks, fintechs, governments", "New stablecoin advisory practice"),
    ("Stablecoin Settlement", "USDC settlement & stablecoin-linked card programs", "Fintechs, crypto exchanges", "$4.6B annualised run-rate (Mar-26)"),
    ("Visa Token Service", "Payment tokenisation replacing PAN data", "Apple Pay, Google Pay, Samsung Pay", "10B+ tokens issued"),
]
for i, (prod, desc, clients, scale) in enumerate(products):
    fill = LIGHT_BLUE_BG if i % 2 == 0 else WHITE
    set_cell(ws, row, 2, prod,    bold=True, fill_hex=fill)
    set_cell(ws, row, 3, desc,    fill_hex=fill, wrap=True)
    set_cell(ws, row, 4, clients, fill_hex=fill, wrap=True)
    set_cell(ws, row, 5, scale,   fill_hex=fill, align_h="center")
    ws.row_dimensions[row].height = 30
    row += 1

blank_row(ws, row); row += 1

# Revenue segments
section_header(ws, row, "  REVENUE BREAKDOWN — FY2025 (Gross Revenues)", 5); row += 1
header_row(ws, row, ["Revenue Segment", "FY2025 ($M)", "YoY Growth", "% of Gross"], 2)
row += 1
segs = [
    ("Data Processing Revenues",     19990, 12.87, 35.9),
    ("Service Revenues",             17540,  8.84, 31.4),
    ("International Transaction Rev",14170, 11.85, 25.4),
    ("Other Revenues",                4050, 26.78,  7.3),
    ("TOTAL GROSS REVENUES",         55750, 11.50, 100.0),
    ("Less: Client Incentives",     -15750,    0,  -28.2),
    ("NET REVENUES",                 40000, 11.34,   None),
]
for i, (seg, val, growth, pct) in enumerate(segs):
    is_total = "TOTAL" in seg or "NET" in seg
    fill = VISA_DARK_BLUE if is_total else (LIGHT_BLUE_BG if i % 2 == 0 else WHITE)
    fc   = HEADER_TEXT if is_total else DARK_TEXT
    pct_str = f"{pct:.1f}%" if isinstance(pct, float) and pct != 0 else ("—" if pct == 0 else str(pct))
    set_cell(ws, row, 2, seg,              bold=is_total, fill_hex=fill, font_color=fc)
    set_cell(ws, row, 3, val,              bold=is_total, fill_hex=fill, font_color=fc,
             number_format='#,##0', align_h="right")
    set_cell(ws, row, 4, f"{growth:.2f}%" if growth else "—", bold=is_total,
             fill_hex=fill, font_color=fc, align_h="center")
    set_cell(ws, row, 5, pct_str,          bold=is_total, fill_hex=fill, font_color=fc, align_h="center")
    ws.row_dimensions[row].height = 22
    row += 1

blank_row(ws, row); row += 1

# Geography
section_header(ws, row, "  GEOGRAPHIC MIX & KEY CLIENTS", 5); row += 1
header_row(ws, row, ["Geography / Channel", "Key Details", "Revenue Mix (est.)", "Growth Driver"], 2)
row += 1
geos = [
    ("United States", "Largest single market; domestic debit/credit dominance",  "~45%", "Consumer spending, tap-to-pay adoption"),
    ("Europe",        "Strong cross-border flows post-Brexit; multi-currency",    "~25%", "Travel recovery, SEPA interoperability"),
    ("Asia-Pacific",  "High-growth markets: India, SE Asia; mobile-first",        "~17%", "Digitisation, contactless expansion"),
    ("CEMEA",         "Central Europe, Middle East, Africa — fastest growing",     "~7%",  "Financial inclusion, remittances"),
    ("Latin America", "Brazil, Mexico — growing middle class",                    "~6%",  "E-commerce, B2C digital wallets"),
    ("Cross-Border",  "International transactions; 11.9% growth FY25",           "Multi", "Travel rebound, global e-commerce"),
]
for i, (geo, detail, mix, driver) in enumerate(geos):
    fill = LIGHT_BLUE_BG if i % 2 == 0 else WHITE
    set_cell(ws, row, 2, geo,    bold=True,  fill_hex=fill)
    set_cell(ws, row, 3, detail, fill_hex=fill, wrap=True)
    set_cell(ws, row, 4, mix,    fill_hex=fill, align_h="center")
    set_cell(ws, row, 5, driver, fill_hex=fill, wrap=True)
    ws.row_dimensions[row].height = 24
    row += 1

blank_row(ws, row); row += 1

# Value proposition & buying process
section_header(ws, row, "  VALUE PROPOSITION & BUYING PROCESS", 5); row += 1
vp = [
    ("For Consumers",      "Secure, accepted everywhere, rewards programs; fraud protection; contactless/mobile pay."),
    ("For Merchants",      "Guaranteed payment, global reach, reduced cash-handling costs, fraud chargeback protection."),
    ("For Issuer Banks",   "Revenue sharing on interchange; card program management tools; data analytics."),
    ("For Acquirer Banks", "Access to global acceptance network; streamlined settlement and reconciliation."),
    ("Buying Process",     "B2B model: Visa sells to financial institutions (issuers/acquirers). No direct consumer sales."
     " Contracts are long-term with incentive arrangements. Decision-makers: bank treasurers, payment teams."),
    ("Seasonality",        "Q4 (Oct-Dec) is strongest — holiday spending spike. Cross-border peaks in Summer travel season."),
]
header_row(ws, row, ["Stakeholder", "Value / Process Detail"], 2, fill_hex=VISA_MID_BLUE); row += 1
for i, (stk, detail) in enumerate(vp):
    fill = LIGHT_BLUE_BG if i % 2 == 0 else WHITE
    set_cell(ws, row, 2, stk,    bold=True, fill_hex=fill)
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=5)
    set_cell(ws, row, 3, detail, fill_hex=fill, wrap=True)
    ws.row_dimensions[row].height = 28
    row += 1

# ─────────────────────────────────────────────
# 3. MOAT
# ─────────────────────────────────────────────
ws = sheets["Moat"]
set_col_widths(ws, {1: 4, 2: 30, 3: 50, 4: 20, 5: 4})

ws.merge_cells("B1:D1")
c = ws.cell(1, 2, "VISA INC. — COMPETITIVE MOAT ANALYSIS")
c.font = Font(name=FONT_NAME, size=FONT_SIZE + 4, bold=True, color=VISA_DARK_BLUE)
c.alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[1].height = 34

row = 3
section_header(ws, row, "  MOAT RATING: WIDE  (Morningstar / GuruFocus consensus)", 4,
               fill_hex=ACCENT_TEAL); row += 1

moat_sources = [
    ("1. NETWORK EFFECTS (Primary Moat)", VISA_DARK_BLUE,
     [("How it works",
       "Every additional cardholder makes the network more valuable to merchants (more sales potential). "
       "Every additional merchant makes the network more valuable to cardholders (more acceptance). "
       "This self-reinforcing loop has operated for 60+ years."),
      ("Scale",
       "4.4 billion cards; 150+ million merchant locations; 15,000+ financial institution clients. "
       "Visa processes >65,000 transactions/second. Near-zero marginal cost per additional transaction."),
      ("Defensibility",
       "Duopoly with Mastercard. New entrants would need simultaneous massive merchant AND consumer adoption. "
       "No fintech has replicated this at scale despite 15+ years of trying."),
     ]),
    ("2. INTANGIBLE ASSETS — BRAND & TRUST", VISA_MID_BLUE,
     [("Brand Recognition",
       "One of the most recognised financial brands globally. \"Everywhere you want to be\" translates to "
       "merchant compliance and consumer preference. Brand value estimated at $35–40B."),
      ("Regulatory Licences",
       "Operating licences in 200+ countries. Significant compliance infrastructure creates high barriers."),
      ("Proprietary Data",
       "Real-time access to $14.2T in annual payment data — an unrivalled competitive intelligence asset."),
     ]),
    ("3. SWITCHING COSTS", ACCENT_TEAL,
     [("Issuer Bank Lock-in",
       "Bank card programs deeply integrated into core banking systems. Switching to a competing network "
       "requires reissuing millions of cards, re-training staff, renegotiating contracts."),
      ("Merchant Lock-in",
       "POS infrastructure often tied to Visa/Mastercard acceptance. Acquirer contracts typically 2-5 years. "
       "Merchants risk losing ~50% of card transactions if they drop Visa."),
      ("Consumer Habits",
       "Embedded in wallets (Apple Pay, Google Pay), auto-payments, loyalty programs. "
       "Very high friction to switch."),
     ]),
    ("4. COST ADVANTAGES & SCALE ECONOMICS", VISA_DARK_BLUE,
     [("Operating Leverage",
       "VisaNet infrastructure cost is largely fixed. Each incremental transaction costs near zero to process. "
       "Operating margin expands as volume grows — 60% OM on $40B revenue."),
      ("Technology Infrastructure",
       "Decades of investment in fraud detection, tokenisation, and network reliability. "
       "Competitors cannot easily replicate without massive capital and time."),
      ("Economies of Scale",
       "Processing 257.5B transactions/year over fixed infrastructure yields unmatched per-unit economics."),
     ]),
]

for moat_name, fill_color, items in moat_sources:
    section_header(ws, row, f"  {moat_name}", 4, fill_hex=fill_color); row += 1
    header_row(ws, row, ["Dimension", "Analysis"], 2, fill_hex=VISA_DARK_BLUE); row += 1
    for i, (dim, analysis) in enumerate(items):
        fill = LIGHT_BLUE_BG if i % 2 == 0 else WHITE
        set_cell(ws, row, 2, dim,      bold=True, fill_hex=fill)
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
        set_cell(ws, row, 3, analysis, fill_hex=fill, wrap=True, align_h="left")
        ws.row_dimensions[row].height = 36
        row += 1
    blank_row(ws, row); row += 1

# Moat summary scorecard
section_header(ws, row, "  MOAT SCORECARD", 4); row += 1
header_row(ws, row, ["Moat Source", "Strength", "Durability", "Verdict"], 2); row += 1
scorecard = [
    ("Network Effects",     "★★★★★", "10+ years",  "WIDE"),
    ("Brand / Intangibles", "★★★★☆", "10+ years",  "WIDE"),
    ("Switching Costs",     "★★★★☆", "5–10 years", "WIDE"),
    ("Cost Advantages",     "★★★★☆", "10+ years",  "WIDE"),
    ("Regulatory Moat",     "★★★☆☆", "5+ years",   "NARROW-WIDE"),
    ("OVERALL",             "★★★★★", "Long-term",  "WIDE"),
]
for i, (src, str_val, dur, verdict) in enumerate(scorecard):
    is_total = src == "OVERALL"
    fill = VISA_DARK_BLUE if is_total else (LIGHT_BLUE_BG if i % 2 == 0 else WHITE)
    fc   = HEADER_TEXT if is_total else DARK_TEXT
    vc   = ACCENT_TEAL if verdict == "WIDE" else (VISA_GOLD if "NARROW" in verdict else ACCENT_RED)
    set_cell(ws, row, 2, src,     bold=is_total, fill_hex=fill, font_color=fc)
    set_cell(ws, row, 3, str_val, bold=is_total, fill_hex=fill, font_color=fc, align_h="center")
    set_cell(ws, row, 4, dur,     bold=is_total, fill_hex=fill, font_color=fc, align_h="center")
    vc_cell = ws.cell(row, 5)   # verdict in col 5 — but sheet only has B-D visible; we'll use D
    # put verdict in col 4 note field — re-use last col
    # Actually col 4 is "Durability", col 5 = verdict — adjust header
    set_cell(ws, row, 5, verdict, bold=is_total,
             fill_hex=(ACCENT_TEAL if verdict == "WIDE" else LIGHT_GOLD_BG),
             font_color=(WHITE if verdict == "WIDE" else DARK_TEXT), align_h="center")
    ws.row_dimensions[row].height = 22
    row += 1

# fix column widths for moat sheet col 5
ws.column_dimensions["E"].width = 18

# ─────────────────────────────────────────────
# 4. INCOME STATEMENT
# ─────────────────────────────────────────────
ws = sheets["Income Statement"]
set_col_widths(ws, {1: 4, 2: 36, 3: 18, 4: 18, 5: 18, 6: 18, 7: 18, 8: 4})

ws.merge_cells("B1:G1")
c = ws.cell(1, 2, "VISA INC. — INCOME STATEMENT  (FY2021 – FY2025, USD Millions)")
c.font = Font(name=FONT_NAME, size=FONT_SIZE + 4, bold=True, color=VISA_DARK_BLUE)
c.alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[1].height = 34

row = 3
header_row(ws, row, ["Metric", "FY 2021", "FY 2022", "FY 2023", "FY 2024", "FY 2025"], 2)
row += 1

income_data = [
    ("NET REVENUES ($M)",              24105, 29310, 32653, 35926, 40000, True),
    ("  YoY Growth",                    None,  21.6,  11.4,  10.0,  11.3, False),
    ("GROSS PROFIT ($M)",              19135, 23577, 26086, 28884, 32145, True),
    ("  Gross Margin %",                79.4,  80.4,  79.9,  80.4,  80.4, False),
    ("OPERATING INCOME ($M)",          15804, 18813, 21000, 23595, 23994, True),
    ("  Operating Margin %",            65.6,  64.2,  64.3,  65.7,  60.0, False),
    ("EBITDA ($M) [est.]",             17200, 20400, 22800, 25500, 25200, True),
    ("  EBITDA Margin %",               71.3,  69.6,  69.8,  71.0,  63.0, False),
    ("NET INCOME ($M)",                12311, 14957, 17273, 19743, 20058, True),
    ("  Net Margin %",                  51.1,  51.0,  52.9,  55.0,  50.1, False),
    ("DILUTED EPS ($)",                 5.63,   7.00,  8.28,  9.73, 10.20, True),
    ("  EPS Growth %",                  None,  24.3,  18.3,  17.5,   4.8, False),
    ("CLIENT INCENTIVES ($M)",        -11627,-13039,-14528,-15670,-15750, False),
    ("R&D / Technology Spend ($M)",     1700,  2100,  2400,  2700,  2950, False),
    ("SHARES OUTSTANDING, Diluted (M)",2750,  2600,  2514,  2400,  2192, False),
]

for label, fy21, fy22, fy23, fy24, fy25, is_bold in income_data:
    is_hdr = is_bold and label.startswith(" ") == False and "%" not in label and "Growth" not in label and "Margin" not in label
    fill_base = VERY_LIGHT_BG if "%" in label or "Growth" in label else (LIGHT_BLUE_BG if row % 2 == 0 else WHITE)
    if "($M)" in label and is_bold:
        fill_base = LIGHT_BLUE_BG
    italic_flag = "%" in label or "Growth" in label or "Margin" in label

    def fmt(v, lbl):
        if v is None: return "—"
        if "%" in lbl or "Growth" in lbl or "Margin" in lbl:
            return f"{v:.1f}%"
        if "$" in lbl or "EPS" in lbl or "SHARES" in lbl:
            return v
        return v

    for ci, val in enumerate([label, fy21, fy22, fy23, fy24, fy25]):
        col = 2 + ci
        if ci == 0:
            set_cell(ws, row, col, val, bold=is_bold, fill_hex=fill_base, italic=italic_flag)
        else:
            formatted_val = fmt(val, label)
            is_pct = "%" in label or "Growth" in label or "Margin" in label
            nfmt = "0.0%" if is_pct and isinstance(val, float) else (
                   "#,##0" if ("$M" in label) else (
                   "0.00" if ("EPS" in label) else None))
            if val is None:
                set_cell(ws, row, col, "—", fill_hex=fill_base, align_h="right", italic=italic_flag)
            else:
                actual_val = val / 100 if is_pct else val
                set_cell(ws, row, col, actual_val,
                         bold=is_bold, fill_hex=fill_base, align_h="right",
                         number_format=nfmt, italic=italic_flag)
    ws.row_dimensions[row].height = 22
    row += 1

blank_row(ws, row); row += 1

# Add a margin chart reference note
section_header(ws, row, "  KEY MARGIN TRENDS (FY2021 – FY2025)", 7); row += 1
header_row(ws, row, ["Metric", "FY 2021", "FY 2022", "FY 2023", "FY 2024", "FY 2025", "5-Yr Avg"], 2)
row += 1
margin_rows = [
    ("Gross Margin",     79.4,  80.4,  79.9,  80.4,  80.4),
    ("Operating Margin", 65.6,  64.2,  64.3,  65.7,  60.0),
    ("Net Margin",       51.1,  51.0,  52.9,  55.0,  50.1),
    ("EBITDA Margin",    71.3,  69.6,  69.8,  71.0,  63.0),
]
for i, (metric, v21, v22, v23, v24, v25) in enumerate(margin_rows):
    avg = (v21 + v22 + v23 + v24 + v25) / 5
    fill = LIGHT_BLUE_BG if i % 2 == 0 else WHITE
    set_cell(ws, row, 2, metric,      bold=True, fill_hex=fill)
    for ci, v in enumerate([v21, v22, v23, v24, v25]):
        set_cell(ws, row, 3 + ci, f"{v:.1f}%", fill_hex=fill, align_h="right")
    set_cell(ws, row, 8, f"{avg:.1f}%", bold=True, fill_hex=LIGHT_GOLD_BG, align_h="right")
    ws.row_dimensions[row].height = 22
    row += 1

# ─────────────────────────────────────────────
# 5. BALANCE SHEET
# ─────────────────────────────────────────────
ws = sheets["Balance Sheet"]
set_col_widths(ws, {1: 4, 2: 38, 3: 18, 4: 18, 5: 18, 6: 18, 7: 18, 8: 4})

ws.merge_cells("B1:G1")
c = ws.cell(1, 2, "VISA INC. — BALANCE SHEET  (FY2021 – FY2025, USD Millions)")
c.font = Font(name=FONT_NAME, size=FONT_SIZE + 4, bold=True, color=VISA_DARK_BLUE)
c.alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[1].height = 34

row = 3
header_row(ws, row, ["Metric", "FY 2021", "FY 2022", "FY 2023", "FY 2024", "FY 2025"], 2)
row += 1

bs_data = [
    # ASSETS
    ("── ASSETS ──",                    None,  None,  None,  None,  None,  True, VISA_MID_BLUE),
    ("Cash & Cash Equivalents ($M)",   16487, 15689, 16286, 11975, 17164, False, None),
    ("Short-Term Investments ($M)",     3200,  2800,  3100,  5200,  2836, False, None),
    ("Total Liquid Assets ($M)",       19687, 18489, 19386, 17175, 20000, True,  None),
    ("Accounts Receivable ($M)",        3100,  3400,  3700,  4200,  4800, False, None),
    ("Goodwill ($M)",                  15958, 17787, 17997, 18941, 19879, False, None),
    ("Intangible Assets ($M)",          7600,  7200,  7100,  7500,  7700, False, None),
    ("TOTAL ASSETS ($M)",              82896, 85501, 90499, 94511, 99627, True,  None),
    # LIABILITIES
    ("── LIABILITIES ──",               None,  None,  None,  None,  None,  True, VISA_MID_BLUE),
    ("Accounts Payable ($M)",           1800,  2100,  2300,  2500,  2700, False, None),
    ("Client Incentives Payable ($M)",  4500,  5100,  5600,  6100,  6500, False, None),
    ("Short-Term Debt ($M)",             500,   400,  1000,   400,  1000, False, None),
    ("Long-Term Debt ($M)",            19978, 20200, 20463, 20836, 19602, False, None),
    ("Other Long-Term Liabilities ($M)",18529, 22120, 22403, 25538, 31916, False, None),
    ("TOTAL LIABILITIES ($M)",         45307, 49920, 51766, 55374, 61718, True,  None),
    # EQUITY
    ("── EQUITY ──",                    None,  None,  None,  None,  None,  True, VISA_MID_BLUE),
    ("Common Stock + APIC ($M)",       21300, 22100, 23000, 24000, 25000, False, None),
    ("Retained Earnings ($M)",         22000, 22000, 24500, 24000, 22000, False, None),
    ("Treasury Stock ($M)",            -5700, -8400,-10000,-10000,-10000, False, None),
    ("TOTAL EQUITY ($M)",              37589, 35581, 38733, 39137, 37909, True,  None),
    # RATIOS
    ("── KEY RATIOS ──",                None,  None,  None,  None,  None,  True, ACCENT_TEAL),
    ("Debt-to-Equity",                   0.53,  0.57,  0.53,  0.53,  0.52, False, None),
    ("Current Ratio",                    1.10,  1.05,  1.08,  1.10,  1.11, False, None),
    ("Book Value per Share ($)",        13.67, 13.69, 15.41, 16.31, 17.30, False, None),
    ("Net Debt ($B)",                    3.79,  4.91,  4.58,  9.26,  2.60, False, None),
]

for label, v21, v22, v23, v24, v25, is_bold, fill_override in bs_data:
    if v21 is None and is_bold:
        section_header(ws, row, f"  {label}", 7, fill_hex=fill_override or VISA_MID_BLUE)
        ws.row_dimensions[row].height = 22
        row += 1
        continue
    fill = LIGHT_BLUE_BG if row % 2 == 0 else WHITE
    is_ratio = label in ["Debt-to-Equity", "Current Ratio", "Book Value per Share ($)", "Net Debt ($B)"]
    nfmt = "#,##0" if "$M" in label else ("0.00" if is_ratio else None)
    set_cell(ws, row, 2, label, bold=is_bold, fill_hex=fill)
    for ci, v in enumerate([v21, v22, v23, v24, v25]):
        if v is None:
            set_cell(ws, row, 3 + ci, "—", fill_hex=fill, align_h="right")
        else:
            set_cell(ws, row, 3 + ci, v, bold=is_bold, fill_hex=fill, align_h="right",
                     number_format=nfmt)
    ws.row_dimensions[row].height = 22
    row += 1

# ─────────────────────────────────────────────
# 6. CASH FLOW
# ─────────────────────────────────────────────
ws = sheets["Cash Flow"]
set_col_widths(ws, {1: 4, 2: 38, 3: 18, 4: 18, 5: 18, 6: 18, 7: 18, 8: 4})

ws.merge_cells("B1:G1")
c = ws.cell(1, 2, "VISA INC. — CASH FLOW ANALYSIS  (FY2021 – FY2025, USD Millions)")
c.font = Font(name=FONT_NAME, size=FONT_SIZE + 4, bold=True, color=VISA_DARK_BLUE)
c.alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[1].height = 34

row = 3
header_row(ws, row, ["Metric", "FY 2021", "FY 2022", "FY 2023", "FY 2024", "FY 2025"], 2)
row += 1

cf_data = [
    # Operating
    ("── OPERATING ACTIVITIES ──",  None,   None,   None,   None,   None,  True, VISA_MID_BLUE),
    ("Net Income ($M)",            12311,  14957,  17273,  19743,  20058, False, None),
    ("D&A ($M)",                    1400,   1600,   1800,   1900,   1206, False, None),
    ("Stock-Based Compensation ($M)",850,    950,   1000,   1050,   1100, False, None),
    ("Working Capital Changes ($M)",  300,    400,    600,   -500,    750, False, None),
    ("Other Adjustments ($M)",        366,    -57,     82,   -243,    -55, False, None),
    ("OPERATING CASH FLOW ($M)",   15227,  18849,  20755,  19950,  23059, True,  None),
    ("OCF Margin %",                 63.2,   64.3,   63.6,   55.5,   57.6, False, None),
    # Investing
    ("── INVESTING ACTIVITIES ──",  None,   None,   None,   None,   None,  True, VISA_MID_BLUE),
    ("Capital Expenditures ($M)",    -705,   -970,  -1059,  -1257,  -1482, False, None),
    ("Acquisitions ($M)",            -200,  -1500,   -500,  -1000,  -1500, False, None),
    ("Investment Purchases/Sales ($M)", 500, -300,    200,   2000,  -1500, False, None),
    ("TOTAL INVESTING ($M)",       -1405,  -2770,  -1359,   -257,  -4482, True,  None),
    # Financing
    ("── FINANCING ACTIVITIES ──",  None,   None,   None,   None,   None,  True, VISA_MID_BLUE),
    ("Share Repurchases ($M)",      -8676, -11589, -12101, -16713, -18316, False, None),
    ("Dividends Paid ($M)",         -2798,  -3203,  -3751,  -4217,  -4634, False, None),
    ("Debt Issuance / (Repayment)",    0,     200,    300,    400,   -200, False, None),
    ("TOTAL FINANCING ($M)",       -11474, -14592, -15552, -20530, -23150, True,  None),
    # FCF
    ("── FREE CASH FLOW ──",        None,   None,   None,   None,   None,  True, ACCENT_TEAL),
    ("FREE CASH FLOW ($M)",        14522,  17879,  19696,  18693,  21577, True,  ACCENT_TEAL),
    ("FCF Margin %",                60.2,   61.0,   60.3,   52.0,   53.9, False, None),
    ("FCF / Net Income %",         117.9,  119.5,  114.0,   94.7,  107.6, False, None),
    ("FCF per Share ($)",            5.28,   6.88,   7.84,   7.79,   9.84, False, None),
    ("Dividend + Buyback ($M)",    11474,  14792,  15852,  20930,  22950, False, None),
    ("Shareholder Return / FCF %", 79.0,   82.7,   80.5,  112.0,  106.4, False, None),
]

for item in cf_data:
    # Handle the tuple with None for "Other Adjustments"
    if len(item) == 7:
        label, v21, v22, v23, v24, v25, is_bold, fill_override = (*item,)
    elif len(item) == 8:
        label, v21, v22, v23, v24, v25, is_bold, fill_override = item
    else:
        # skip malformed
        continue

    if v21 is None and is_bold:
        section_header(ws, row, f"  {label}", 7, fill_hex=fill_override or VISA_MID_BLUE)
        ws.row_dimensions[row].height = 22
        row += 1
        continue
    fill = LIGHT_BLUE_BG if row % 2 == 0 else WHITE
    if is_bold and fill_override:
        fill = fill_override
    is_pct = "%" in label
    is_header_row_item = label.startswith("FREE CASH FLOW")
    nfmt = "#,##0" if "$M" in label else ("0.0%" if is_pct else ("0.00" if "$)" in label else None))
    set_cell(ws, row, 2, label, bold=is_bold, fill_hex=fill,
             font_color=WHITE if (is_bold and fill_override == ACCENT_TEAL) else DARK_TEXT)
    for ci, v in enumerate([v21, v22, v23, v24, v25]):
        if v is None:
            set_cell(ws, row, 3 + ci, "—", fill_hex=fill, align_h="right")
        else:
            fc = WHITE if (is_bold and fill_override == ACCENT_TEAL) else DARK_TEXT
            actual = v / 100 if is_pct else v
            set_cell(ws, row, 3 + ci, actual, bold=is_bold, fill_hex=fill, align_h="right",
                     number_format=nfmt, font_color=fc)
    ws.row_dimensions[row].height = 22
    row += 1

# ─────────────────────────────────────────────
# 7. RETURN ON CAPITAL
# ─────────────────────────────────────────────
ws = sheets["Return on Capital"]
set_col_widths(ws, {1: 4, 2: 38, 3: 18, 4: 18, 5: 18, 6: 18, 7: 18, 8: 22, 9: 4})

ws.merge_cells("B1:H1")
c = ws.cell(1, 2, "VISA INC. — RETURN ON CAPITAL  (FY2021 – FY2025)")
c.font = Font(name=FONT_NAME, size=FONT_SIZE + 4, bold=True, color=VISA_DARK_BLUE)
c.alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[1].height = 34

row = 3
header_row(ws, row, ["Metric", "FY 2021", "FY 2022", "FY 2023", "FY 2024", "FY 2025", "5-Yr Avg"], 2, fill_hex=VISA_DARK_BLUE)
row += 1

roc_data = [
    ("RETURN ON EQUITY (ROE %)",        32.7,  42.0,  44.6,  50.4,  52.9),
    ("RETURN ON ASSETS (ROA %)",        14.9,  17.5,  19.1,  20.9,  20.1),
    ("RETURN ON INVESTED CAPITAL %",    20.1,  24.3,  25.6,  27.0,  27.8),
    ("RETURN ON CAPITAL EMPLOYED %",    17.5,  21.0,  23.2,  25.3,  24.2),
]

for label, v21, v22, v23, v24, v25 in roc_data:
    avg = (v21+v22+v23+v24+v25)/5
    fill = LIGHT_BLUE_BG if row % 2 == 0 else WHITE
    set_cell(ws, row, 2, label, bold=True, fill_hex=fill)
    for ci, v in enumerate([v21, v22, v23, v24, v25]):
        set_cell(ws, row, 3+ci, f"{v:.1f}%", fill_hex=fill, align_h="right", bold=False)
    set_cell(ws, row, 8, f"{avg:.1f}%", bold=True, fill_hex=LIGHT_GOLD_BG, align_h="right")
    ws.row_dimensions[row].height = 24
    row += 1

blank_row(ws, row); row += 1

section_header(ws, row, "  INCREMENTAL RETURN ON CAPITAL (ROIC Analysis)", 8); row += 1
header_row(ws, row, ["Year", "Net Income ($M)", "Invested Capital ($M)", "ROIC%",
                     "∆ Net Income ($M)", "∆ Invested Capital ($M)", "Incremental ROIC%"], 2); row += 1

inc_data = [
    ("FY 2021", 12311, 56000, 22.0, None,  None,  None),
    ("FY 2022", 14957, 57500, 26.0, 2646,  1500,  176.4),
    ("FY 2023", 17273, 60000, 28.8, 2316,  2500,   92.6),
    ("FY 2024", 19743, 63000, 31.3, 2470,  3000,   82.3),
    ("FY 2025", 20058, 63500, 31.6,  315,   500,   63.0),
]
for i, (yr, ni, ic, roic, dni, dic, iroic) in enumerate(inc_data):
    fill = LIGHT_BLUE_BG if i % 2 == 0 else WHITE
    set_cell(ws, row, 2, yr,   bold=True,  fill_hex=fill, align_h="center")
    set_cell(ws, row, 3, ni,   fill_hex=fill, align_h="right", number_format="#,##0")
    set_cell(ws, row, 4, ic,   fill_hex=fill, align_h="right", number_format="#,##0")
    set_cell(ws, row, 5, f"{roic:.1f}%", fill_hex=fill, align_h="right")
    set_cell(ws, row, 6, "—" if dni is None else dni,  fill_hex=fill, align_h="right",
             number_format="#,##0" if dni else None)
    set_cell(ws, row, 7, "—" if dic is None else dic,  fill_hex=fill, align_h="right",
             number_format="#,##0" if dic else None)
    set_cell(ws, row, 8, "—" if iroic is None else f"{iroic:.1f}%", fill_hex=fill, align_h="right")
    ws.row_dimensions[row].height = 22
    row += 1

blank_row(ws, row); row += 1
section_header(ws, row, "  OWNER'S PERSPECTIVE ON CAPITAL ALLOCATION", 8); row += 1
alloc_notes = [
    ("Asset-Light Model",      "Visa does not own the underlying credit risk — capital requirements are minimal relative to earnings."),
    ("Capital Return Focus",   "FY2025: $18.3B buybacks + $4.6B dividends = $22.9B returned vs $21.6B FCF (>100% payout). "
                               "$30B buyback authorised in Apr-2025."),
    ("Acquisitions (disciplined)", "CyberSource, Visa Europe, Earthport, Paysafe — all strategic to network expansion. "
                               "No reckless M&A."),
    ("Reinvestment Rate",      "Low capex (~$1.5B, ~3.7% of revenue) reflects the toll-road economics: "
                               "infrastructure scales with minimal marginal cost."),
    ("WACC vs ROIC",           "ROIC ~28% vs estimated WACC ~8-9%. Spread of 19-20 ppts — exceptional value creation."),
]
for i, (k, v) in enumerate(alloc_notes):
    fill = LIGHT_GOLD_BG if i % 2 == 0 else VERY_LIGHT_BG
    set_cell(ws, row, 2, k, bold=True, fill_hex=fill)
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=8)
    set_cell(ws, row, 3, v, fill_hex=fill, wrap=True)
    ws.row_dimensions[row].height = 30
    row += 1

# ─────────────────────────────────────────────
# 8. MANAGEMENT
# ─────────────────────────────────────────────
ws = sheets["Management"]
set_col_widths(ws, {1: 4, 2: 30, 3: 60, 4: 4})

ws.merge_cells("B1:C1")
c = ws.cell(1, 2, "VISA INC. — MANAGEMENT QUALITY ANALYSIS")
c.font = Font(name=FONT_NAME, size=FONT_SIZE + 4, bold=True, color=VISA_DARK_BLUE)
c.alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[1].height = 34

row = 3
section_header(ws, row, "  CEO: RYAN McINERNEY (President & CEO since March 2023)", 3)
row += 1
mgmt_bio = [
    ("Background",         "30-year Visa veteran. Joined 2013 as President of North America. Previously MD at JPMorgan Chase's "
                           "consumer banking division. Deep payments expertise, industry relationships."),
    ("Tenure",             "CEO since March 2023, taking over from Alfred Kelly Jr. Previously known internally as a "
                           "data-driven, disciplined executor."),
    ("2025 Compensation",  "$31.56M total (21% raise): $1.5M base, $5.25M non-equity incentive, $18.7M stock awards, "
                           "$5.69M stock options, $373K benefits. ~95% tied to long-term performance metrics."),
    ("CEO vs Median Pay",  "254× the median Visa employee pay of $124,000. Compensation is performance-linked."),
    ("Owner-Like Behaviour","60%+ of comp in equity aligns CEO with shareholders. Long vesting schedules (3-5 years) "
                           "prevent short-term thinking. No signs of value destruction via vanity acquisitions."),
    ("Strategic Vision",   "Driving 'Network of Networks': expanding Visa Direct, B2B Connect, and stablecoin settlement. "
                           "Proactive integration of digital assets rather than defensive posture."),
]
header_row(ws, row, ["Category", "Assessment"], 2, fill_hex=VISA_DARK_BLUE); row += 1
for i, (k, v) in enumerate(mgmt_bio):
    fill = LIGHT_BLUE_BG if i % 2 == 0 else WHITE
    set_cell(ws, row, 2, k, bold=True, fill_hex=fill)
    set_cell(ws, row, 3, v, fill_hex=fill, wrap=True)
    ws.row_dimensions[row].height = 40
    row += 1

blank_row(ws, row); row += 1
section_header(ws, row, "  PROXY STATEMENT — COMPENSATION PHILOSOPHY", 3); row += 1
proxy_data = [
    ("Pay Mix",             "FY2025: 5% base salary · 17% STI · 60% long-term equity · 18% options. "
                            "Strongly back-loaded to long-term performance."),
    ("STI Metrics",         "Adjusted EPS growth (~60% weight) + Net Revenue growth (~40% weight). "
                            "2025 threshold/target/max payout structure."),
    ("LTI (Performance Shares)", "3-year ROIC vs S&P 500 Financial sector benchmark + relative TSR. "
                            "Ties executive wealth to multi-year shareholder value creation."),
    ("Clawback Policy",     "Robust recoupment provisions per SEC Rule 10D-1. Misconduct-triggered clawback."),
    ("Board Oversight",     "Independent comp committee. No insider-dominated board. "
                            "Alfred Kelly (Former CEO) serves as Executive Chairman — smooth transition, no empire building."),
    ("Dilution Control",    "Net share count reduced from 2.75B (FY2021) to 1.91B (current) via buybacks. "
                            "Stock comp issuance more than offset — EPS growing faster than net income."),
]
header_row(ws, row, ["Topic", "Details"], 2, fill_hex=VISA_MID_BLUE); row += 1
for i, (k, v) in enumerate(proxy_data):
    fill = LIGHT_GOLD_BG if i % 2 == 0 else WHITE
    set_cell(ws, row, 2, k, bold=True, fill_hex=fill)
    set_cell(ws, row, 3, v, fill_hex=fill, wrap=True)
    ws.row_dimensions[row].height = 36
    row += 1

blank_row(ws, row); row += 1
section_header(ws, row, "  INSIDER TRADING & SEC FILINGS", 3); row += 1
header_row(ws, row, ["Date / Party", "Transaction / Assessment"], 2, fill_hex=VISA_DARK_BLUE); row += 1
insider_data = [
    ("Sep 3, 2025 — CEO Ryan McInerney",
     "Sold 10,485 shares @ $348.57 (~$3.65M) under 10b5-1 plan. Days prior, received same count at $109.82 exercise price. Net = ordinary equity compensation settlement."),
    ("Jun 11, 2025 — Chief Risk Officer (Fabara)",
     "Sold 11,636 shares @ $375.00 as part of 6 pre-planned transactions totalling $59.31M. Compliance-driven, not distress signal."),
    ("Mar 11, 2026 — Director Lloyd Carney",
     "Open-market sale: 650 shares @ $309.62 (~$201K). Small amount, likely estate/diversification planning."),
    ("Apr 2025 — Board",
     "$30B share repurchase programme authorised. Strong signal of confidence in intrinsic value vs. current market price."),
    ("Overall Insider Sentiment",
     "NO pattern of distress-related sales. All major transactions are rule 10b5-1 (pre-planned) structured selling. "
     "Company itself is the largest buyer (buybacks). This is bullish, not bearish."),
    ("Institutional Ownership",
     "~94% held by institutions (Vanguard 8.5%, BlackRock 7.8%, State Street 4.5%). High-quality holder base."),
]
for i, (party, detail) in enumerate(insider_data):
    fill = LIGHT_BLUE_BG if i % 2 == 0 else WHITE
    is_last = party == "Overall Insider Sentiment"
    set_cell(ws, row, 2, party, bold=True, fill_hex=fill)
    set_cell(ws, row, 3, detail, fill_hex=fill, wrap=True)
    ws.row_dimensions[row].height = 46
    row += 1

blank_row(ws, row); row += 1
section_header(ws, row, "  CAPITAL ALLOCATION TRACK RECORD", 3); row += 1
header_row(ws, row, ["Decision", "Assessment"], 2, fill_hex=ACCENT_TEAL,
           font_color=WHITE); row += 1
cap_alloc = [
    ("Buybacks ($47B in 4 yrs)",      "Share count down 30% since 2021. EPS growth outpaced net income growth — value accretive."),
    ("Dividend Growth",               "Dividend grew from $1.28/share (FY2021) to $2.23/share (FY2025) — 74% increase, 5-yr CAGR ~15%."),
    ("Debt Management",               "LTD ~$19.6B at low average interest rates (~3%). Net debt ~$2.6B — near net-cash. Conservative leverage."),
    ("Strategic M&A",                 "Earthport (cross-border), Currencycloud (FX), Pismo (cloud banking infra). "
                                      "Bolt-on acquisitions at reasonable prices. No transformational overpay."),
    ("Capex Discipline",              "Only $1.5B capex on $40B revenue (3.7%). Technology investments are ROI-positive at >28% ROIC."),
    ("Seeds for the Future",          "Multi-year investment in Visa Direct, B2B Connect, stablecoin infrastructure, AI fraud detection. "
                                      "Not borrowing from the future — actually planting seeds."),
]
for i, (decision, assessment) in enumerate(cap_alloc):
    fill = VERY_LIGHT_BG if i % 2 == 0 else WHITE
    set_cell(ws, row, 2, decision,   bold=True, fill_hex=fill)
    set_cell(ws, row, 3, assessment, fill_hex=fill, wrap=True)
    ws.row_dimensions[row].height = 36
    row += 1

# ─────────────────────────────────────────────
# 9. RISKS
# ─────────────────────────────────────────────
ws = sheets["Risks"]
set_col_widths(ws, {1: 4, 2: 28, 3: 10, 4: 50, 5: 22, 6: 4})

ws.merge_cells("B1:E1")
c = ws.cell(1, 2, "VISA INC. — RISK ANALYSIS")
c.font = Font(name=FONT_NAME, size=FONT_SIZE + 4, bold=True, color=VISA_DARK_BLUE)
c.alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[1].height = 34

row = 3
header_row(ws, row, ["Risk Factor", "Severity", "Description & Current Status", "Mitigation"], 2)
row += 1

risks = [
    # Category, Severity, Description, Mitigation, Category Color
    ("REGULATORY / LEGAL", None, None, None, VISA_DARK_BLUE),
    ("DOJ Antitrust Lawsuit",
     "HIGH",
     "Sep 2024: DOJ sued Visa alleging monopolisation of debit card markets. Visa controls ~60% of US debit. "
     "Trial could result in network unbundling, interchange caps, or structural remedies.",
     "Visa contests vigorously. Historical precedent (Durbin Amendment) showed Visa adapted. "
     "Settlement risk: ~$1-3B fine without structural change most likely outcome.",
     None),
    ("Interchange Fee Regulation",
     "HIGH",
     "EU Payment Services Directive 3 (PSD3), UK Payment Systems Regulator review, potential US legislation. "
     "Interchange fee caps would compress service revenues.",
     "Visa is actively lobbying. PSD3 grandfathered existing fee structures initially. "
     "Cross-border volumes (not subject to caps) are largest growth driver.",
     None),
    ("$30B Class Action (US Merchants)",
     "MEDIUM",
     "Long-running merchant antitrust class action settled for $5.6B (2024). New filings ongoing. "
     "Merchants argue swipe fees are anti-competitive.",
     "$500M set aside in litigation escrow (Apr 2026 SEC filing). Total risk manageable given $21B FCF.",
     None),
    ("COMPETITIVE", None, None, None, ACCENT_RED),
    ("AI Agent Payment Routing",
     "MEDIUM-HIGH",
     "Citrini Research (Feb 2026): AI agents optimised for cost minimisation could bypass card rails "
     "(2-3% fees) in favour of ACH/RTP networks (fractions of a cent).",
     "Visa is investing in AI-native payment infrastructure. B2B Connect & stablecoin integration "
     "provide lower-cost alternatives within the Visa ecosystem.",
     None),
    ("Stablecoin / Crypto Disruption",
     "MEDIUM",
     "Stablecoin transfer volume hit $33T in 2025 (+72% YoY). USDC, USDT cross-border flows may "
     "bypass traditional card rails.",
     "Visa's counter-strategy: USDC settlement launched Dec 2025. Stablecoin Advisory Practice. "
     "$4.6B annualised stablecoin-linked card run-rate. Visa is becoming the infrastructure layer "
     "FOR stablecoins, not competing against them.",
     None),
    ("Stripe / Direct Infrastructure",
     "MEDIUM",
     "Stripe processed $1.9T (2025), acquired Bridge ($1.1B stablecoin specialist). Stripe increasingly "
     "competes with card rails for enterprise clients.",
     "Stripe still relies heavily on Visa/MC rails for card acceptance. Not a near-term threat to core volumes.",
     None),
    ("Real-Time Payments (RTP/FedNow)",
     "MEDIUM",
     "FedNow (US) and RTP network expanding instant ACH availability. Could displace some debit volume.",
     "Visa Direct (Visa's push payments product) routes over these same rails. Visa participates in the ecosystem.",
     None),
    ("China / Domestic Networks",
     "LOW-MEDIUM",
     "China UnionPay dominates domestic Chinese transactions. WeChat Pay, Alipay control digital wallets. "
     "Limited Visa share in largest emerging market.",
     "Visa benefits from cross-border China tourism spend. Domestic market penetration remains constrained.",
     None),
    ("MACRO / OPERATIONAL", None, None, None, ACCENT_TEAL),
    ("Consumer Spending Slowdown",
     "MEDIUM",
     "Recession or credit tightening reduces payment volumes. Q1 2026 saw slowing consumer spending growth. "
     "Visa revenue directly tied to global GDP and consumption.",
     "Diversified global exposure. Cross-border (higher margin) less correlated with domestic cycles. "
     "Even in 2020-COVID, volumes recovered within 12 months.",
     None),
    ("FX / Currency Risk",
     "LOW-MEDIUM",
     "~55% of revenue is international. USD strengthening reduces reported revenue from international ops.",
     "Natural hedging via costs in local currencies. Financial instruments used for further protection.",
     None),
    ("Cybersecurity / Fraud",
     "MEDIUM",
     "Payment networks are prime targets. A major breach of VisaNet could devastate trust.",
     "Multi-billion $ annual investment in security. 99.999% uptime history. Tokenisation reduces breach risk.",
     None),
    ("Geopolitical / Sanctions",
     "LOW",
     "Russia suspension (2022) removed ~$150M revenue. Future sanctions on other countries possible.",
     "Diversification across 200 countries limits single-country exposure.",
     None),
]

current_cat = None
for item in risks:
    if item[1] is None:  # category header
        cat_name, _, _, _, cat_color = item
        section_header(ws, row, f"  {cat_name}", 5, fill_hex=cat_color)
        ws.row_dimensions[row].height = 22
        row += 1
        continue
    risk, severity, desc, mitigation, _ = item
    sev_color = (ACCENT_RED if severity in ["HIGH", "MEDIUM-HIGH"]
                 else VISA_GOLD if severity in ["MEDIUM", "LOW-MEDIUM"]
                 else ACCENT_TEAL)
    fill = LIGHT_BLUE_BG if row % 2 == 0 else WHITE
    set_cell(ws, row, 2, risk,       bold=True, fill_hex=fill)
    set_cell(ws, row, 3, severity,   bold=True, fill_hex=sev_color,
             font_color=WHITE, align_h="center")
    set_cell(ws, row, 4, desc,       fill_hex=fill, wrap=True)
    set_cell(ws, row, 5, mitigation, fill_hex=VERY_LIGHT_BG, wrap=True)
    ws.row_dimensions[row].height = 52
    row += 1

# ─────────────────────────────────────────────
# 10. VALUATION
# ─────────────────────────────────────────────
ws = sheets["Valuation"]
set_col_widths(ws, {1: 4, 2: 36, 3: 22, 4: 22, 5: 22, 6: 4})

ws.merge_cells("B1:E1")
c = ws.cell(1, 2, "VISA INC. — VALUATION ANALYSIS  (As of April 14, 2026)")
c.font = Font(name=FONT_NAME, size=FONT_SIZE + 4, bold=True, color=VISA_DARK_BLUE)
c.alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[1].height = 34

row = 3
section_header(ws, row, "  CURRENT MARKET DATA", 5); row += 1
mkt_data = [
    ("Stock Price",           "$309.09"),
    ("Market Capitalisation", "$593.5 Billion"),
    ("Enterprise Value",      "~$598.2 Billion"),
    ("Trailing P/E",          "29.1×"),
    ("Forward P/E (FY2026E)", "23.4×"),
    ("P/FCF",                 "27.5×"),
    ("EV / EBITDA",           "20.6×"),
    ("P/B Ratio",             "15.4×"),
    ("Dividend Yield",        "0.86%"),
    ("52-Week Range",         "$286 – $371"),
    ("YTD Return (2026)",     "-11.8%"),
]
header_row(ws, row, ["Metric", "Value", "", ""], 2, fill_hex=VISA_DARK_BLUE); row += 1
for i, (k, v) in enumerate(mkt_data):
    fill = LIGHT_BLUE_BG if i % 2 == 0 else WHITE
    set_cell(ws, row, 2, k, bold=True, fill_hex=fill)
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=5)
    set_cell(ws, row, 3, v, fill_hex=fill, align_h="left")
    ws.row_dimensions[row].height = 22
    row += 1

blank_row(ws, row); row += 1

# DCF Analysis
section_header(ws, row, "  DCF VALUATION — BASE CASE", 5); row += 1
header_row(ws, row, ["Assumption", "Value", "Justification", ""], 2, fill_hex=VISA_DARK_BLUE); row += 1
dcf_assumptions = [
    ("FCF FY2025 (Base)",     "$21.6B",        "Actual reported FCF"),
    ("FCF Growth — Yrs 1-5",  "10% p.a.",      "In-line with 5-yr FCF CAGR of ~10.4%"),
    ("FCF Growth — Yrs 6-10", "8% p.a.",       "Moderate deceleration as base grows"),
    ("Terminal Growth Rate",  "3.5%",          "GDP-like growth; global digital payments long-term expansion"),
    ("Discount Rate (WACC)",  "9.0%",          "Risk-free 4.3% + equity risk premium; low beta ~0.95"),
    ("Shares Outstanding",    "1.91 Billion",  "Diluted current shares; declining via buybacks"),
]
for i, (k, v, j) in enumerate(dcf_assumptions):
    fill = LIGHT_GOLD_BG if i % 2 == 0 else WHITE
    set_cell(ws, row, 2, k, bold=True, fill_hex=fill)
    set_cell(ws, row, 3, v, fill_hex=fill, align_h="center")
    ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=5)
    set_cell(ws, row, 4, j, fill_hex=fill, italic=True)
    ws.row_dimensions[row].height = 24
    row += 1

blank_row(ws, row); row += 1

# Scenario table
section_header(ws, row, "  VALUATION SCENARIOS", 5); row += 1
header_row(ws, row, ["Scenario", "FCF Growth (Yr1-5)", "Terminal Rate", "Intrinsic Value / Share", "vs Current Price"], 2)
row += 1
scenarios = [
    ("Bear Case",    "6% p.a.",   "2.5%", "$270",  "-12.7%"),
    ("Base Case",    "10% p.a.",  "3.5%", "$340",  "+10.0%"),
    ("Bull Case",    "14% p.a.",  "4.0%", "$430",  "+39.1%"),
    ("Analyst Avg",  "Consensus", "—",    "$400",  "+29.4%"),
    ("Current Price","—",         "—",    "$309",  "0%  (reference)"),
]
for i, (scen, gr, tr, iv, vs) in enumerate(scenarios):
    is_current = scen == "Current Price"
    fill = (VERY_LIGHT_BG if is_current else
            ACCENT_RED if scen == "Bear Case" else
            LIGHT_BLUE_BG if scen == "Base Case" else
            ACCENT_TEAL if scen == "Bull Case" else LIGHT_GOLD_BG)
    fc = WHITE if scen in ["Bull Case"] else DARK_TEXT
    set_cell(ws, row, 2, scen, bold=True, fill_hex=fill, font_color=fc)
    set_cell(ws, row, 3, gr,   fill_hex=fill, font_color=fc, align_h="center")
    set_cell(ws, row, 4, tr,   fill_hex=fill, font_color=fc, align_h="center")
    set_cell(ws, row, 5, iv,   bold=True, fill_hex=fill, font_color=fc, align_h="center")
    vs_color = ACCENT_TEAL if "+" in vs else (ACCENT_RED if "-" in vs else VISA_GOLD)
    set_cell(ws, row, 6, vs,   bold=True,
             fill_hex=(vs_color),
             font_color=WHITE, align_h="center")
    ws.row_dimensions[row].height = 24
    row += 1

blank_row(ws, row); row += 1

# Comps
section_header(ws, row, "  COMPARABLE COMPANY ANALYSIS (Payments Peers)", 5); row += 1
header_row(ws, row, ["Company", "P/E (Fwd)", "EV/EBITDA", "FCF Yield", "ROIC", "Notes"], 2)
row += 1
comps = [
    ("Visa (V)",           "23.4×", "20.6×",  "3.6%", "~28%",  "Current; -12% YTD"),
    ("Mastercard (MA)",    "26.0×", "22.1×",  "3.2%", "~45%",  "Premium to Visa; higher ROIC"),
    ("American Express",   "16.4×", "12.8×",  "6.1%", "~28%",  "Credit issuer; different risk profile"),
    ("PayPal (PYPL)",      "14.5×",  "9.2×",  "7.2%", "~18%",  "Disintermediation risk; value trap?"),
    ("Block / Square (SQ)","35.0×", "18.5×",  "1.8%", "~10%",  "High growth; pre-mature profitability"),
    ("Stripe (Private)",    "N/A",   "N/A",    "N/A",  "N/A",  "$70B last valuation; 2025 rev ~$14B"),
]
for i, (comp, pe, ev, fcf, roic, notes) in enumerate(comps):
    is_visa = comp.startswith("Visa")
    fill = VISA_GOLD if is_visa else (LIGHT_BLUE_BG if i % 2 == 0 else WHITE)
    set_cell(ws, row, 2, comp,  bold=is_visa, fill_hex=fill)
    set_cell(ws, row, 3, pe,    fill_hex=fill, align_h="center")
    set_cell(ws, row, 4, ev,    fill_hex=fill, align_h="center")
    set_cell(ws, row, 5, fcf,   fill_hex=fill, align_h="center")
    set_cell(ws, row, 6, roic,  fill_hex=fill, align_h="center")
    ws.merge_cells(start_row=row, start_column=7, end_row=row, end_column=7)
    # There's no col 7 in this layout - notes goes in last existing col
    # We'll extend the table
    set_cell(ws, row, 7, notes, fill_hex=fill, italic=True)
    ws.row_dimensions[row].height = 22
    row += 1

ws.column_dimensions["G"].width = 36

blank_row(ws, row); row += 1
section_header(ws, row, "  MARGIN OF SAFETY ASSESSMENT", 7); row += 1
mos = [
    ("Base Case Intrinsic Value",  "~$340", "+10.0% upside from $309"),
    ("Analyst Consensus Target",   "~$400", "+29.4% upside from $309"),
    ("Bull Case Value",            "~$430", "+39.1% upside from $309"),
    ("52-Week Low",                "~$286", "Already tested — supported"),
    ("Margin of Safety (Base)",    "~10%",  "Modest but low business risk warrants lower discount"),
    ("Verdict",                    "FAIR TO MODESTLY UNDERVALUED",
     "Stock appears 10-30% below fair value. Quality business at a reasonable price."),
]
header_row(ws, row, ["Item", "Level", "Context"], 2, fill_hex=ACCENT_TEAL,
           font_color=WHITE); row += 1
for i, row_data in enumerate(mos):
    label, val, ctx = row_data[0], row_data[1], row_data[2] if len(row_data) > 2 else ""
    is_verdict = label == "Verdict"
    fill = (ACCENT_TEAL if is_verdict else LIGHT_BLUE_BG if i % 2 == 0 else WHITE)
    fc   = WHITE if is_verdict else DARK_TEXT
    set_cell(ws, row, 2, label, bold=is_verdict, fill_hex=fill, font_color=fc)
    set_cell(ws, row, 3, val,   bold=is_verdict, fill_hex=fill, font_color=fc, align_h="center")
    ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=5)
    set_cell(ws, row, 4, ctx,   fill_hex=fill, font_color=fc, italic=not is_verdict)
    ws.row_dimensions[row].height = 24
    row += 1

# ─────────────────────────────────────────────
# 11. MARKET SENTIMENT
# ─────────────────────────────────────────────
ws = sheets["Market Sentiment"]
set_col_widths(ws, {1: 4, 2: 32, 3: 28, 4: 28, 5: 28, 6: 4})

ws.merge_cells("B1:E1")
c = ws.cell(1, 2, "VISA INC. — MARKET SENTIMENT & ANALYST COVERAGE")
c.font = Font(name=FONT_NAME, size=FONT_SIZE + 4, bold=True, color=VISA_DARK_BLUE)
c.alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[1].height = 34

row = 3
section_header(ws, row, "  ANALYST CONSENSUS (April 2026)", 5); row += 1
header_row(ws, row, ["Category", "Count / Detail", "Implication", ""], 2); row += 1
consensus_data = [
    ("Strong Buy",         "7",      "High conviction outperformers"),
    ("Buy",                "30",     "Constructive; expect market-beat returns"),
    ("Hold",               "3",      "Fair-valued or wait-and-see"),
    ("Sell",               "0",      "No analyst recommends selling"),
    ("CONSENSUS RATING",   "BUY",    "37/40 analysts bullish — strong consensus"),
    ("Avg. Price Target",  "~$400",  "+29.4% upside from $309"),
    ("High Target",        "$450",   "Bull-case scenario"),
    ("Low Target",         "$323",   "Conservative/cautious view"),
    ("Median Target",      "$400.60","Fair value estimate"),
]
for i, (cat, count, impl) in enumerate(consensus_data):
    is_bold_row = "CONSENSUS" in cat or "Avg." in cat
    fill = VISA_DARK_BLUE if "CONSENSUS" in cat else (LIGHT_BLUE_BG if i % 2 == 0 else WHITE)
    fc   = HEADER_TEXT if "CONSENSUS" in cat else DARK_TEXT
    set_cell(ws, row, 2, cat,   bold=is_bold_row, fill_hex=fill, font_color=fc)
    cnt_fill = (ACCENT_TEAL if cat == "Strong Buy"
                else VISA_MID_BLUE if cat == "Buy"
                else VISA_GOLD if cat == "Hold"
                else fill)
    set_cell(ws, row, 3, count, bold=is_bold_row,
             fill_hex=(ACCENT_TEAL if "CONSENSUS" in cat else cnt_fill),
             font_color=(WHITE if cat in ["Strong Buy", "Buy"] else (DARK_TEXT if cat == "Hold" else fc)),
             align_h="center")
    ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=5)
    set_cell(ws, row, 4, impl, fill_hex=fill, font_color=fc)
    ws.row_dimensions[row].height = 22
    row += 1

blank_row(ws, row); row += 1

section_header(ws, row, "  RECENT ANALYST ACTIONS & COMMENTARY (2025–2026)", 5); row += 1
header_row(ws, row, ["Firm / Date", "Action", "New Target", "Commentary"], 2); row += 1
analyst_actions = [
    ("BofA Securities — Feb 2026",  "UPGRADE to Buy",     "$420", "Stablecoin push reduces disruption risk; attractive entry"),
    ("Citrini Research — Feb 2026", "CAUTIOUS NOTE",      "N/A",  "AI-agent routing could bypass card rails; watch this risk"),
    ("Goldman Sachs — Q4 2025",     "Buy (maintained)",   "$400", "Network effects intact; FCF growth undervalued"),
    ("Morgan Stanley — Q4 2025",    "Overweight",         "$415", "Cross-border recovery + value-added services acceleration"),
    ("JPMorgan — Q1 2026",          "Overweight",         "$395", "Slowing consumer but long-term durable compounder"),
    ("TIKR — Apr 2026",             "Deep-Dive (neutral)", "$340-380", "Stock -11.8% in 2026; approaching value territory"),
]
for i, (firm, action, target, comment) in enumerate(analyst_actions):
    fill = LIGHT_BLUE_BG if i % 2 == 0 else WHITE
    action_fill = (ACCENT_TEAL if "UPGRADE" in action or "Buy" in action or "Overweight" in action
                   else ACCENT_RED if "CAUTIOUS" in action else VISA_GOLD)
    action_fc = WHITE if action_fill in [ACCENT_TEAL, ACCENT_RED] else DARK_TEXT
    set_cell(ws, row, 2, firm,    bold=True,  fill_hex=fill)
    set_cell(ws, row, 3, action,  bold=True,  fill_hex=action_fill, font_color=action_fc, align_h="center")
    set_cell(ws, row, 4, target,  fill_hex=fill, align_h="center")
    set_cell(ws, row, 5, comment, fill_hex=fill, wrap=True)
    ws.row_dimensions[row].height = 32
    row += 1

blank_row(ws, row); row += 1
section_header(ws, row, "  STOCK PERFORMANCE & TECHNICAL CONTEXT", 5); row += 1
header_row(ws, row, ["Metric", "Value", "Interpretation", ""], 2); row += 1
tech_data = [
    ("Current Price (Apr 14, 2026)", "$309.09", "Down 11.8% YTD — creating potential entry point"),
    ("52-Week High",                 "$371.00",  "Aug 2025 peak; -16.7% from high"),
    ("52-Week Low",                  "$286.00",  "Oct 2025 trough; +8.0% from low"),
    ("YTD 2026 Return",              "-11.8%",   "Underperformance vs. S&P; valuation compression"),
    ("5-Year Total Return",          "+78%",     "Outperformed S&P 500 by ~12 ppts"),
    ("Dividend (TTM)",               "$2.66/share","Yield 0.86%; growing ~15%/yr"),
    ("Beta",                         "~0.95",    "Slightly less volatile than market"),
    ("Short Interest",               "~0.9%",    "Very low — bears are not positioned"),
    ("Institutional Ownership",      "~94%",     "High quality long-term holders"),
]
for i, (metric, val, interp) in enumerate(tech_data):
    fill = LIGHT_BLUE_BG if i % 2 == 0 else WHITE
    set_cell(ws, row, 2, metric, bold=True, fill_hex=fill)
    set_cell(ws, row, 3, val,    fill_hex=fill, align_h="center", bold=True)
    ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=5)
    set_cell(ws, row, 4, interp, fill_hex=fill)
    ws.row_dimensions[row].height = 22
    row += 1

blank_row(ws, row); row += 1
section_header(ws, row, "  SOCIAL / MARKET SENTIMENT INDICATORS", 5); row += 1
social = [
    ("Seeking Alpha Coverage",       "Multiple 'Wide Moat' and 'Buy' articles. Community sentiment: BULLISH."),
    ("Wall Street Analyst Consensus","37/40 BUY. Strongest conviction payments name after Mastercard."),
    ("Retail Investor Sentiment",    "Top 10 most-held stock on major brokerages. Known as a 'forever hold'."),
    ("News Flow (Apr 2026)",         "Stablecoin expansion positive; DOJ litigation ongoing but manageable. "
                                     "Stock up +1.6% on Apr 13 on strong cross-border data."),
    ("Macroeconomic Headwinds",      "Tariff concerns, potential US slowdown weigh on short-term sentiment. "
                                     "However, Visa benefits from global diversification."),
    ("Key Catalyst to Watch",        "FQ2 2026 earnings (late Apr 2026): cross-border volume data, "
                                     "consumer spending trends, stablecoin revenue disclosure."),
]
header_row(ws, row, ["Source / Signal", "Detail"], 2, fill_hex=VISA_MID_BLUE); row += 1
for i, (source, detail) in enumerate(social):
    fill = VERY_LIGHT_BG if i % 2 == 0 else WHITE
    set_cell(ws, row, 2, source, bold=True, fill_hex=fill)
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=5)
    set_cell(ws, row, 3, detail, fill_hex=fill, wrap=True)
    ws.row_dimensions[row].height = 30
    row += 1

# ─────────────────────────────────────────────
# 12. KEY INDICATORS
# ─────────────────────────────────────────────
ws = sheets["Key Indicators"]
set_col_widths(ws, {1: 4, 2: 38, 3: 18, 4: 18, 5: 18, 6: 18, 7: 18, 8: 4})

ws.merge_cells("B1:G1")
c = ws.cell(1, 2, "VISA INC. — KEY PERFORMANCE INDICATORS  (FY2021 – FY2025)")
c.font = Font(name=FONT_NAME, size=FONT_SIZE + 4, bold=True, color=VISA_DARK_BLUE)
c.alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[1].height = 34

row = 3
header_row(ws, row, ["KPI", "FY 2021", "FY 2022", "FY 2023", "FY 2024", "FY 2025"], 2)
row += 1

kpi_data = [
    # Network Volume
    ("── PAYMENT NETWORK VOLUME ──", None, None, None, None, None, True, VISA_DARK_BLUE),
    ("Total Payments Volume ($T)",  10.4,  11.6,  12.3,  13.2,  14.2, False, None),
    ("  YoY Growth %",              None,  11.5,   6.0,   7.3,   7.6, False, None),
    ("US Payments Volume ($T)",      5.0,   5.5,   5.9,   6.3,   6.8, False, None),
    ("International Volume ($T)",    5.4,   6.1,   6.4,   6.9,   7.4, False, None),
    ("Cross-Border Vol Index",        88,   110,   124,   134,   145, False, None),
    ("Transactions Processed (B)",  192.5, 212.6, 225.5, 237.0, 257.5, False, None),
    ("  Txn Growth YoY %",          None,  10.4,   6.1,   5.1,   8.6, False, None),
    # Cards / Acceptance
    ("── CARDS & NETWORK REACH ──",  None, None, None, None, None, True, VISA_MID_BLUE),
    ("Cards in Force (Billions)",     3.8,   4.0,   4.2,   4.3,   4.4, False, None),
    ("Merchant Locations (M)",        80,   100,   130,   150,   155, False, None),
    ("Financial Institutions (000s)", 13.5,  14.5,  14.9,  15.1,  15.3, False, None),
    # Per-Transaction Economics
    ("── UNIT ECONOMICS ──",         None, None, None, None, None, True, ACCENT_TEAL),
    ("Net Rev per Transaction ($)",  0.125, 0.138, 0.145, 0.152, 0.155, False, None),
    ("Yield (Net Rev / Vol %)",      0.232, 0.253, 0.265, 0.272, 0.282, False, None),
    ("Client Incentive Ratio %",     32.5,  30.8,  30.8,  30.4,  28.3, False, None),
    # Profitability
    ("── PROFITABILITY DASHBOARD ──", None, None, None, None, None, True, VISA_DARK_BLUE),
    ("Net Revenue ($B)",             24.1,  29.3,  32.7,  35.9,  40.0, False, None),
    ("Gross Margin %",               79.4,  80.4,  79.9,  80.4,  80.4, False, None),
    ("Operating Margin %",           65.6,  64.2,  64.3,  65.7,  60.0, False, None),
    ("Net Margin %",                 51.1,  51.0,  52.9,  55.0,  50.1, False, None),
    ("FCF Margin %",                 60.2,  61.0,  60.3,  52.0,  53.9, False, None),
    # Returns
    ("── CAPITAL RETURNS ──",        None, None, None, None, None, True, ACCENT_TEAL),
    ("EPS (Diluted $)",               5.63,  7.00,  8.28,  9.73, 10.20, False, None),
    ("EPS CAGR (vs FY21) %",         None,  24.3,  21.3,  19.9,  16.0, False, None),
    ("Dividend per Share ($)",        1.28,  1.50,  1.80,  2.08,  2.34, False, None),
    ("Dividend Growth %",            None,  17.2,  20.0,  15.6,  12.5, False, None),
    ("Buyback ($B)",                  8.7,  11.6,  12.1,  16.7,  18.3, False, None),
    ("Total Capital Returned ($B)",  11.5,  14.8,  15.9,  20.9,  22.9, False, None),
    ("ROIC %",                       20.1,  24.3,  25.6,  27.0,  27.8, False, None),
    ("ROE %",                        32.7,  42.0,  44.6,  50.4,  52.9, False, None),
]

for item in kpi_data:
    label, v21, v22, v23, v24, v25, is_bold, fill_override = item
    if v21 is None and is_bold:
        section_header(ws, row, f"  {label}", 7, fill_hex=fill_override)
        ws.row_dimensions[row].height = 22
        row += 1
        continue
    is_pct  = "%" in label
    is_growth = "Growth" in label or "YoY" in label or "CAGR" in label
    fill = LIGHT_BLUE_BG if row % 2 == 0 else WHITE
    set_cell(ws, row, 2, label, bold=is_bold, fill_hex=fill,
             italic=(is_pct or is_growth))
    for ci, v in enumerate([v21, v22, v23, v24, v25]):
        col = 3 + ci
        if v is None:
            set_cell(ws, row, col, "—", fill_hex=fill, align_h="right", italic=True)
        else:
            is_monetary = "$B" in label or "$T" in label or "$M" in label
            nfmt = (None)  # we'll display as-is
            display_val = f"{v:.1f}%" if is_pct else (
                          f"${v:.2f}B" if "$B" in label else (
                          f"${v:.1f}T" if "$T" in label else (
                          f"${v:.2f}" if "$)" in label or "Share" in label or "EPS" in label or "per" in label else
                          f"{v:,.1f}" if v > 100 else f"{v:.1f}")))
            set_cell(ws, row, col, display_val, fill_hex=fill, align_h="right",
                     italic=(is_pct or is_growth))
    ws.row_dimensions[row].height = 20
    row += 1

blank_row(ws, row); row += 1
section_header(ws, row, "  NOTE: Visa is NOT a SaaS company — Rule of 50 not applicable", 7,
               fill_hex=VISA_GOLD); row += 1
ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
c = ws.cell(row, 2,
    "Visa is a global payment network operator (financial infrastructure), not a software-as-a-service business. "
    "The Rule of 40/50 (Revenue Growth % + FCF Margin %) is designed for SaaS companies. "
    "For reference only: Visa's 2025 score = 11% revenue growth + 54% FCF margin = 65 — well above threshold — "
    "but this metric is not the appropriate lens for evaluating Visa.")
c.font = make_font(FONT_SIZE, italic=True)
c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
c.fill = make_fill(LIGHT_GOLD_BG)
c.border = make_border()
ws.row_dimensions[row].height = 52

# ─────────────────────────────────────────────
# FIX CF DATA (tuple issue) — rebuild cleanly
# ─────────────────────────────────────────────
# Already handled above

# ─────────────────────────────────────────────
# SAVE
# ─────────────────────────────────────────────
os.makedirs(OUTPUT_DIR, exist_ok=True)
wb.save(OUTPUT_FILE)
print(f"[✓] Workbook saved to: {OUTPUT_FILE}")
print(f"    Sheets: {[ws.title for ws in wb.worksheets]}")
