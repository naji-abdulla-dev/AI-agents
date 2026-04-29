"""
Novo Nordisk A/S (NYSE ADR: NVO) - Comprehensive Financial & Investment Analysis
Generated: April 2026
Fiscal Year: January 1 – December 31 | Reporting Currency: DKK (USD ~6.6 DKK/USD)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

OUTPUT_PATH = "/Users/naji/WORK/github.com/AI/claude/Agent/claude/MarketResearch/output/NVO_Financial_Analysis.xlsx"

# ── Colour Palette (Healthcare / Pharma) ──────────────────────────────────────
DARK_NAVY    = "0A2342"
NVO_BLUE     = "004F9E"
LIGHT_BLUE   = "BBDEFB"
NVO_RED      = "C8102E"
GREEN        = "1B5E20"
LIGHT_GREEN  = "E8F5E9"
RED          = "B71C1C"
LIGHT_RED    = "FFCDD2"
YELLOW       = "FFF9C4"
GREY         = "F5F5F5"
MID_GREY     = "90A4AE"
WHITE        = "FFFFFF"
ACCENT_TEAL  = "00838F"
LIGHT_TEAL   = "E0F7FA"

FONT_SIZE = 14

# ── Style helpers ─────────────────────────────────────────────────────────────
def _font(bold=False, colour=None, size=FONT_SIZE, italic=False):
    return Font(name="Calibri", bold=bold,
                color=colour or "000000", size=size, italic=italic)

def _fill(hex_colour):
    return PatternFill("solid", fgColor=hex_colour)

def _border(style="thin"):
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _hdr_cell(ws, row, col, value, bg=DARK_NAVY, fg=WHITE, bold=True,
              align="center", span=1):
    c = ws.cell(row=row, column=col, value=value)
    c.font = _font(bold=bold, colour=fg, size=FONT_SIZE)
    c.fill = _fill(bg)
    c.alignment = _align(align, "center")
    c.border = _border()
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col + span - 1)
    return c

def _data_cell(ws, row, col, value, bg=WHITE, bold=False,
               align="right", fmt=None, colour=None, wrap=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = _font(bold=bold, colour=colour or "000000", size=FONT_SIZE)
    c.fill = _fill(bg)
    c.alignment = _align(align, "center", wrap=wrap)
    c.border = _border()
    if fmt:
        c.number_format = fmt
    return c

def _section_hdr(ws, row, col, text, span=7):
    _hdr_cell(ws, row, col, text, bg=NVO_BLUE, fg=WHITE, span=span)
    ws.row_dimensions[row].height = 26

def _set_col_widths(ws, widths):
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = w

def _note_row(ws, row, col, text, span=7):
    ws.merge_cells(start_row=row, start_column=col,
                   end_row=row, end_column=col + span - 1)
    c = ws.cell(row, col, text)
    c.font = _font(italic=True, colour="5D4037", size=FONT_SIZE - 1)
    c.fill = _fill(YELLOW)
    c.alignment = _align("left", "center", wrap=True)
    c.border = _border()
    ws.row_dimensions[row].height = 28


# ─────────────────────────────────────────────────────────────────────────────
# TAB 1 — COVER
# ─────────────────────────────────────────────────────────────────────────────
def build_cover(wb):
    ws = wb.create_sheet("Cover")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, [2, 26, 22, 22, 22, 22, 22, 2])

    for r in range(1, 55):
        for c in range(1, 9):
            ws.cell(r, c).fill = _fill(WHITE)

    ws.merge_cells("B1:G4")
    hdr = ws["B1"]
    hdr.value = "Novo Nordisk A/S (NYSE ADR: NVO)"
    hdr.font = Font(name="Calibri", bold=True, size=28, color=WHITE)
    hdr.fill = _fill(DARK_NAVY)
    hdr.alignment = _align("center", "center")
    ws.row_dimensions[1].height = 40

    ws.merge_cells("B5:G6")
    sub = ws["B5"]
    sub.value = "Comprehensive Financial & Investment Analysis  |  April 2026"
    sub.font = Font(name="Calibri", bold=False, size=17, color=WHITE)
    sub.fill = _fill(NVO_BLUE)
    sub.alignment = _align("center", "center")
    ws.row_dimensions[5].height = 30

    ws.merge_cells("B7:G8")
    tag = ws["B7"]
    tag.value = "GLP-1 Diabetes & Obesity Leader · Danish Pharma Pioneer · Post-Peak Inflection Point"
    tag.font = Font(name="Calibri", italic=True, size=FONT_SIZE, color=DARK_NAVY)
    tag.fill = _fill(LIGHT_BLUE)
    tag.alignment = _align("center", "center")
    ws.row_dimensions[7].height = 26

    def kv(row, label, value):
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=7)
        lc = ws.cell(row, 2, label)
        lc.font = _font(bold=True, colour=WHITE, size=FONT_SIZE)
        lc.fill = _fill(DARK_NAVY)
        lc.alignment = _align("right", "center")
        lc.border = _border()
        vc = ws.cell(row, 4, value)
        vc.font = _font(colour=DARK_NAVY, size=FONT_SIZE)
        vc.fill = _fill(LIGHT_BLUE)
        vc.alignment = _align("left", "center")
        vc.border = _border()
        ws.row_dimensions[row].height = 22

    kv_rows = [
        (10, "Company",            "Novo Nordisk A/S"),
        (11, "Ticker",             "NYSE ADR: NVO  |  Primary: NOVO B (Nasdaq Copenhagen)"),
        (12, "Sector",             "Healthcare – Pharmaceuticals (GLP-1 Specialist)"),
        (13, "Headquarters",       "Bagsværd, Denmark"),
        (14, "Founded",            "1923"),
        (15, "CEO",                "Maziar Mike Doustdar (since August 7, 2025)"),
        (16, "CFO",                "Karsten Munk Knudsen"),
        (17, "Analysis Date",      "April 29, 2026"),
        (18, "Stock Price (ADR)",  "$41.28 USD (April 27, 2026)"),
        (19, "Market Cap",         "~$183.5B USD"),
        (20, "Enterprise Value",   "~$189B USD (Net Debt ~$15.8B USD)"),
        (21, "52-Wk Range",        "$35.12 – $81.44 USD"),
        (22, "Analyst Consensus",  "HOLD (2 Strong Buy, 0 Buy, 7 Hold, 2 Sell — 11 analysts)"),
        (23, "Avg Price Target",   "$51 USD avg | Median $47 | High $70 | Low $41"),
        (24, "Revenue FY2025",     "DKK 309,064M (~$46.8B USD)  |  +6.4% YoY (+~15% CER)"),
        (25, "EBIT Margin FY2025", "41.3%  |  Net Margin 33.1%  |  FCF DKK 58,962M"),
        (26, "FY2026 Guidance",    "Sales: -5% to -13% CER (first-ever annual decline)"),
        (27, "Valuation",          "P/E (TTM): ~12x  |  EV/EBITDA: ~8x  |  Div. Yield: ~2.6%"),
    ]
    for r, lbl, val in kv_rows:
        kv(r, lbl, val)

    ws.merge_cells("B29:G29")
    th = ws["B29"]
    th.value = "INVESTMENT THESIS"
    th.font = _font(bold=True, colour=WHITE, size=FONT_SIZE + 1)
    th.fill = _fill(GREEN)
    th.alignment = _align("center", "center")
    ws.row_dimensions[29].height = 26

    thesis = [
        "• GLP-1 SECULAR GROWTH INTACT: Obesity affects 650M+ people globally with <5% on treatment. Novo dominates this market with semaglutide (Ozempic, Wegovy) and has a 100-year scientific heritage in diabetes. The addressable market is projected at $150-200B by 2030 — structural demand is unbroken despite near-term competitive headwinds.",
        "• DEEP VALUATION RESET: NVO has fallen ~57% from its all-time high (~$92 USD). At ~12x forward earnings and ~8x EV/EBITDA — well below its 5-year avg 25x P/E and peer median 15x P/E — the stock prices in prolonged competitive pressure. Meaningful pessimism is already embedded at current levels.",
        "• PIPELINE OPTIONALITY (AMYCRETIN): Amycretin, a dual GLP-1/amylin oral agent, is in Phase 3 with results expected 2026-2027. Early Phase 1/2 data showed ~13% weight loss in 12 weeks — superior to Ozempic. If Phase 3 confirms, Novo could leapfrog Lilly's tirzepatide and trigger a major re-rating from current beaten-down levels.",
        "• KEY HEADWINDS: Eli Lilly's tirzepatide (Zepbound) demonstrated clinical superiority (~25.5% vs ~23% weight loss) in head-to-head, winning ~60% US market share. IRA Medicare pricing from Jan 2027 cuts Ozempic net revenue ~$274/month. CagriSema Phase 3 failed to beat Zepbound (Feb 2026). FY2026 marks the first ever guidance for negative sales growth.",
        "• FOUNDATION OWNERSHIP = PATIENT CAPITAL: The Novo Nordisk Foundation controls ~28% economic / ~67% voting rights via A-shares. This prevents activist/short-term pressure and ensures the company can invest through the cycle. Long-term R&D orientation is structurally preserved. Dividend grew every year and capex investment in manufacturing is strategic, not distress.",
        "• VERDICT: NVO is a world-class pharma franchise at a deep value price, but near-term earnings headwinds are real. The stock is a HOLD at $41 — value investors accumulate here for the 2027-2028 recovery. The amycretin Phase 3 readout is the single most important catalyst. If it succeeds, the stock could double; if it fails, further downside to ~$28-33 (bear case).",
    ]
    for i, line in enumerate(thesis):
        ws.merge_cells(start_row=30 + i, start_column=2, end_row=30 + i, end_column=7)
        c = ws.cell(30 + i, 2, line)
        c.font = _font(size=FONT_SIZE)
        c.fill = _fill(GREY if i % 2 == 0 else WHITE)
        c.alignment = _align("left", "center", wrap=True)
        ws.row_dimensions[30 + i].height = 40

    ws.merge_cells("B37:G37")
    rat = ws["B37"]
    rat.value = "RATING: HOLD  |  12-Month Target: $48 (Base)  |  Bull: $85  |  Bear: $30  |  Current: $41.28"
    rat.font = Font(name="Calibri", bold=True, size=FONT_SIZE + 1, color=WHITE)
    rat.fill = _fill(ACCENT_TEAL)
    rat.alignment = _align("center", "center")
    ws.row_dimensions[37].height = 30


# ─────────────────────────────────────────────────────────────────────────────
# TAB 2 — BUSINESS OVERVIEW
# ─────────────────────────────────────────────────────────────────────────────
def build_business(wb):
    ws = wb.create_sheet("Business Overview")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, [2, 28, 20, 18, 18, 18, 18, 2])

    for r in range(1, 80):
        for c in range(1, 9):
            ws.cell(r, c).fill = _fill(WHITE)

    _hdr_cell(ws, 1, 2, "Novo Nordisk A/S — Business Overview", bg=DARK_NAVY, fg=WHITE, span=6)
    ws.row_dimensions[1].height = 32

    # Company snapshot
    _section_hdr(ws, 3, 2, "COMPANY SNAPSHOT", span=6)
    snapshot = [
        ("Business Model",      "Prescription pharmaceutical manufacturer. Develops, manufactures, and markets GLP-1 agonists (Ozempic, Wegovy, Rybelsus), insulins, and rare disease treatments. Revenue is 100% from drug sales to distributors, PBMs, and health systems. No services, no devices."),
        ("Core Products",       "Ozempic (semaglutide inj., T2D), Wegovy (semaglutide 2.4mg, obesity), Rybelsus (oral semaglutide, T2D), Victoza (liraglutide, legacy), Tresiba (insulin), NovoLog/Fiasp (insulin), NovoSeven RT (haemophilia), Norditropin (growth hormone)"),
        ("Revenue Model",       "Commercial insurance (co-pay cards), Medicare Part D / Medicare Advantage, Medicaid, international national health systems. ~55-60% US revenue. Net prices after rebates are 40-60% of list prices in the US."),
        ("Key Customers",       "US: Express Scripts, CVS Caremark, OptumRx (PBMs control formulary access). Also: health systems, retail pharmacies, specialty pharmacies. International: national health insurance programs in EU, Japan, China, EM."),
        ("Seasonality",         "Minimal — chronic prescription drug; steady refill patterns. Slight Q1 drag from deductible resets (patients delay prescriptions in Jan). Q4 pull-forward is possible as patients hit out-of-pocket maximums."),
        ("TAM",                 "Global obesity market: ~650M adults obese; <5% treated; ~$150-200B market by 2030. Global diabetes: ~537M people with T2D; significant untreated population. Rare disease: smaller but stable with high barriers."),
        ("Founded / History",   "1923 (Denmark). First insulin manufacturer outside Canada. 100+ years of diabetes expertise. IPO not applicable (continuously listed; Foundation owner). Ozempic approved US 2017; Wegovy approved 2021."),
        ("Employees",           "~72,000 globally (2025). Major manufacturing sites: Denmark, US (Catalent sites), China, Brazil, France, Algeria."),
    ]
    for i, (label, value) in enumerate(snapshot):
        row = 4 + i
        lc = ws.cell(row, 2, label)
        lc.font = _font(bold=True, colour=WHITE, size=FONT_SIZE)
        lc.fill = _fill(NVO_BLUE)
        lc.alignment = _align("right", "center")
        lc.border = _border()
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=7)
        vc = ws.cell(row, 3, value)
        vc.font = _font(size=FONT_SIZE)
        vc.fill = _fill(GREY if i % 2 == 0 else WHITE)
        vc.alignment = _align("left", "center", wrap=True)
        vc.border = _border()
        ws.row_dimensions[row].height = 30

    # Segment breakdown
    _section_hdr(ws, 13, 2, "SEGMENT REVENUE BREAKDOWN", span=6)
    seg_hdrs = ["Segment", "FY2024 Revenue (DKK M)", "% of Total", "FY2025 Revenue (DKK M)", "YoY Growth (CER)", "Key Products"]
    for ci, h in enumerate(seg_hdrs):
        _hdr_cell(ws, 14, 2 + ci, h, bg=NVO_BLUE, fg=WHITE)

    seg_data = [
        ("Diabetes Care",    "206,618", "71.1%", "~223,000",  "+8% CER",  "Ozempic, Rybelsus, Victoza, Tresiba, NovoLog"),
        ("  — Ozempic",      "120,342", "41.4%", "~130,000+", "+~10%",    "Semaglutide once-weekly injection (T2D)"),
        ("Obesity Care",     "65,146",  "22.4%", "82,347",    "+31% CER", "Wegovy (semaglutide 2.4mg), Saxenda"),
        ("  — Wegovy",       "~58,200", "~20.0%","~73,000+",  "+~25%",    "Semaglutide 2.4mg once-weekly (obesity)"),
        ("Rare Disease",     "~17,500", "~6.0%", "~18,000",   "+5%",      "NovoSeven RT, Esperoct, Sogroya, Norditropin"),
        ("Other",            "~1,139",  "~0.4%", "~1,200",    "~+5%",     "Other biopharmaceuticals"),
        ("TOTAL",            "290,403", "100%",  "309,064",   "+6.4% DKK","All segments"),
    ]
    bg_cycle = [GREY, WHITE, LIGHT_TEAL, WHITE, GREY, WHITE, LIGHT_BLUE]
    for i, row_data in enumerate(seg_data):
        r = 15 + i
        bgs = [bg_cycle[i]] * 6
        for ci, val in enumerate(row_data):
            bold = (i == 6)
            _data_cell(ws, r, 2 + ci, val, bg=bgs[ci], bold=bold, align="center")
        ws.row_dimensions[r].height = 20

    # Geographic split
    _section_hdr(ws, 23, 2, "GEOGRAPHIC REVENUE SPLIT (FY2024)", span=6)
    geo_hdrs = ["Geography", "Revenue Share", "Key Notes"]
    _hdr_cell(ws, 24, 2, "Geography", bg=NVO_BLUE, fg=WHITE)
    _hdr_cell(ws, 24, 3, "Revenue Share", bg=NVO_BLUE, fg=WHITE)
    ws.merge_cells(start_row=24, start_column=4, end_row=24, end_column=7)
    _hdr_cell(ws, 24, 4, "Key Notes", bg=NVO_BLUE, fg=WHITE, span=4)

    geo_data = [
        ("North America (US/Canada)", "~55-60%", "Largest GLP-1 market; IRA Medicare pricing from Jan 2027 is key risk. Ozempic + Wegovy dominant. Compounders exiting 2026."),
        ("Europe",                    "~20-22%", "Growing obesity reimbursement; Novo's home region; less USD FX exposure on DKK reporting."),
        ("International (APAC/EM)",   "~18-22%", "China, Japan growing; lower net prices. EM expansion is a hedge vs US market share loss to Lilly."),
        ("Rest of World",             "~2-3%",   "Smaller markets; opportunistic."),
    ]
    for i, (geo, share, note) in enumerate(geo_data):
        r = 25 + i
        bg = GREY if i % 2 == 0 else WHITE
        _data_cell(ws, r, 2, geo, bg=bg, align="left")
        _data_cell(ws, r, 3, share, bg=bg, align="center")
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=7)
        nc = ws.cell(r, 4, note)
        nc.font = _font(size=FONT_SIZE)
        nc.fill = _fill(bg)
        nc.alignment = _align("left", "center", wrap=True)
        nc.border = _border()
        ws.row_dimensions[r].height = 26

    _note_row(ws, 30, 2, "NOTE: FX exposure is significant — ~70% of revenue USD-denominated but reported in DKK. Strong USD relative to DKK creates translation headwind for USD ADR investors. Constant-exchange-rate (CER) growth better reflects business performance.", span=6)


# ─────────────────────────────────────────────────────────────────────────────
# TAB 3 — MOAT
# ─────────────────────────────────────────────────────────────────────────────
def build_moat(wb):
    ws = wb.create_sheet("Moat")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, [2, 24, 14, 14, 30, 2])

    for r in range(1, 60):
        for c in range(1, 7):
            ws.cell(r, c).fill = _fill(WHITE)

    _hdr_cell(ws, 1, 2, "Novo Nordisk A/S — Competitive Moat Analysis", bg=DARK_NAVY, fg=WHITE, span=4)
    ws.row_dimensions[1].height = 32

    _section_hdr(ws, 3, 2, "MOAT RATING: NARROW-TO-MODERATE (was WIDE; eroding due to Lilly competition)", span=4)

    col_hdrs = ["Moat Element", "Width", "Trend", "Evidence"]
    for ci, h in enumerate(col_hdrs):
        _hdr_cell(ws, 4, 2 + ci, h, bg=NVO_BLUE, fg=WHITE)

    moat_rows = [
        ("Semaglutide Patent Portfolio", "MODERATE", "→ Stable",    "Core compound patent protected into ~mid-2030s; manufacturing process IP also filed; formulation patents add layers."),
        ("Manufacturing Scale (post-Catalent)", "MODERATE", "↑ Growing", "$11.7B Catalent acquisition (2024) gives world's largest GLP-1 fill-finish capacity in the US. Hard to replicate in <5 years."),
        ("Brand / HCP Loyalty (Diabetes)", "WIDE",     "→ Stable",   "100+ years in diabetes. Endocrinologists trust Ozempic implicitly. Patients on semaglutide rarely switch. First-mover in once-weekly GLP-1."),
        ("Scientific Peptide Chemistry Expertise", "WIDE", "→ Stable","Deep GLP-1/GIP/amylin peptide expertise developed over decades. Amycretin (GLP-1/amylin) is a product of this platform. Hard to replicate."),
        ("Regulatory Track Record", "MODERATE", "→ Stable",         "Proven FDA approval track record across T2D, obesity, CKD (SELECT trial), pediatric. Trust with FDA reduces approval risk for pipeline."),
        ("Pricing Power (US)",       "NARROWING", "↓ Declining",    "IRA Medicare negotiation forces Ozempic to $274/month and Wegovy to $385/month from Jan 2027. PBM rebates also increasing. Structural pressure."),
        ("US GLP-1 Market Share",    "NARROWING", "↓ Declining",    "Novo ~40% US market share vs Lilly ~60% (2026 est.). Tirzepatide head-to-head superiority in weight loss (25.5% vs 23%). Lilly winning new patients."),
        ("Geographic Diversification","MODERATE", "→ Stable",       "55%+ US revenue, but growing EM/Asia. EU + International partially offsets US market share loss. Less exposed to Lilly in EU/EM."),
    ]
    bg_c = [GREY, WHITE]
    for i, (elem, width, trend, evidence) in enumerate(moat_rows):
        r = 5 + i
        bg = bg_c[i % 2]
        width_col = GREEN if "WIDE" in width else (NVO_BLUE if "MODERATE" in width else NVO_RED)
        _data_cell(ws, r, 2, elem, bg=bg, align="left", bold=True)
        _data_cell(ws, r, 3, width, bg=bg, align="center", colour=width_col, bold=True)
        _data_cell(ws, r, 4, trend, bg=bg, align="center")
        ev = ws.cell(r, 5, evidence)
        ev.font = _font(size=FONT_SIZE)
        ev.fill = _fill(bg)
        ev.alignment = _align("left", "center", wrap=True)
        ev.border = _border()
        ws.row_dimensions[r].height = 30

    _section_hdr(ws, 14, 2, "PRIMARY COMPETITORS", span=4)
    comp_hdrs = ["Competitor", "Ticker", "Key Threat", "Novo Advantage vs Threat"]
    for ci, h in enumerate(comp_hdrs):
        _hdr_cell(ws, 15, 2 + ci, h, bg=NVO_BLUE, fg=WHITE)

    comp_rows = [
        ("Eli Lilly",      "LLY",    "Zepbound/Mounjaro (tirzepatide, dual GIP/GLP-1): clinical superiority, ~60% US mkt share",
                                     "Oral pipeline (amycretin); EM geographic expansion; brand loyalty in diabetes; 100yr heritage"),
        ("AstraZeneca",    "AZN",    "Farxiga (SGLT-2) in diabetes; no current obesity GLP-1",
                                     "Not a direct GLP-1 obesity competitor; Novo stronger in core diabetes market"),
        ("Amgen",          "AMGN",   "MariTide (GLP-1/GIPR monthly injection); Phase 3 ongoing",
                                     "Not yet proven; monthly dosing less convenient; years from commercialisation"),
        ("Roche",          "RHHBY",  "CT-996 (oral GLP-1); very early stage",
                                     "Multiple years from market; Novo already in Phase 3 with oral program"),
        ("Pfizer",         "PFE",    "Danuglipron (oral GLP-1) — dropped 2025",
                                     "Programme abandoned; no current head-to-head threat"),
    ]
    for i, (comp, ticker, threat, adv) in enumerate(comp_rows):
        r = 16 + i
        bg = bg_c[i % 2]
        _data_cell(ws, r, 2, comp, bg=bg, align="left", bold=True)
        _data_cell(ws, r, 3, ticker, bg=bg, align="center")
        tc = ws.cell(r, 4, threat)
        tc.font = _font(size=FONT_SIZE)
        tc.fill = _fill(LIGHT_RED if i == 0 else bg)
        tc.alignment = _align("left", "center", wrap=True)
        tc.border = _border()
        ac = ws.cell(r, 5, adv)
        ac.font = _font(size=FONT_SIZE)
        ac.fill = _fill(LIGHT_GREEN if i == 0 else bg)
        ac.alignment = _align("left", "center", wrap=True)
        ac.border = _border()
        ws.row_dimensions[r].height = 30

    _section_hdr(ws, 22, 2, "PIPELINE — KEY PROGRAMMES", span=4)
    pipe_hdrs = ["Programme", "Stage", "Indication", "Significance"]
    for ci, h in enumerate(pipe_hdrs):
        _hdr_cell(ws, 23, 2 + ci, h, bg=NVO_BLUE, fg=WHITE)

    pipe_rows = [
        ("Amycretin (oral, GLP-1/amylin)", "Phase 3", "Obesity",        "CRITICAL — Phase 1/2 showed ~13% wt loss at 12wks. Phase 3 readout 2026-2027 is the primary re-rating catalyst. Success could restore Novo's leadership over tirzepatide."),
        ("Amycretin (injectable)",          "Phase 3", "Obesity",        "Injectable version; backup/complement to oral programme. Early data strong."),
        ("Mim8 (anti-TFPI)",                "Phase 3", "Haemophilia A/B","Non-factor therapy with best-in-class data. Prophylactic once-weekly sub-cut. Strong revenue potential in rare disease."),
        ("CagriSema",                       "Failed",  "Obesity",        "Feb 2026: failed non-inferiority vs Zepbound in Phase 3. 23% vs 25.5% weight loss. Key setback removing re-rating catalyst."),
        ("Monlunabant (CBR1)",              "Phase 2", "Obesity",        "Endocannabinoid system modulator; differentiated MOA; early data. Long-term optionality."),
        ("Ozempic (semaglutide)",           "Approved","CKD / NASH",     "SELECT trial: 20% reduction in CV events. Label expansion ongoing. FLOW trial: CKD benefit confirmed. Expands TAM."),
        ("Oral Semaglutide (Rybelsus)",     "Approved","T2D (expanding)","Growing prescriptions; label expansion for obesity being studied. Lower efficacy than injectable but convenient."),
    ]
    for i, (prog, stage, ind, sig) in enumerate(pipe_rows):
        r = 24 + i
        stage_bg = LIGHT_GREEN if stage == "Phase 3" else (LIGHT_RED if stage == "Failed" else GREY)
        bg = GREY if i % 2 == 0 else WHITE
        _data_cell(ws, r, 2, prog, bg=bg, align="left", bold=True)
        _data_cell(ws, r, 3, stage, bg=stage_bg, align="center", bold=True)
        _data_cell(ws, r, 4, ind, bg=bg, align="center")
        sc = ws.cell(r, 5, sig)
        sc.font = _font(size=FONT_SIZE)
        sc.fill = _fill(bg)
        sc.alignment = _align("left", "center", wrap=True)
        sc.border = _border()
        ws.row_dimensions[r].height = 32

    _note_row(ws, 32, 2, "MOAT VERDICT: Strong patent and scientific moat remain but are narrower than peak. Clinical inferiority of semaglutide vs tirzepatide is the key moat breach. Oral amycretin Phase 3 is the single most important catalyst for moat restoration. Novo Nordisk Foundation's patient capital provides structural protection for long-term R&D investment through this cycle.", span=4)


# ─────────────────────────────────────────────────────────────────────────────
# TAB 4 — INCOME STATEMENTS
# ─────────────────────────────────────────────────────────────────────────────
def build_income(wb):
    ws = wb.create_sheet("Income Statements")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, [2, 32, 14, 14, 14, 14, 14, 2])

    for r in range(1, 55):
        for c in range(1, 9):
            ws.cell(r, c).fill = _fill(WHITE)

    _hdr_cell(ws, 1, 2, "Novo Nordisk A/S — Income Statement (DKK Millions)", bg=DARK_NAVY, fg=WHITE, span=6)
    ws.row_dimensions[1].height = 32

    _section_hdr(ws, 3, 2, "ANNUAL INCOME STATEMENT — DKK MILLIONS (As Reported)", span=6)

    years = ["FY2021", "FY2022", "FY2023", "FY2024", "FY2025"]
    _hdr_cell(ws, 4, 2, "Metric", bg=NVO_BLUE, fg=WHITE, align="left")
    for ci, yr in enumerate(years):
        _hdr_cell(ws, 4, 3 + ci, yr, bg=NVO_BLUE, fg=WHITE)
    ws.row_dimensions[4].height = 22

    income_data = [
        ("Revenue",              ["140,800",  "176,954",  "232,261",  "290,403",  "309,064"],  False, GREY),
        ("Gross Profit",         ["117,142",  "148,506",  "196,496",  "245,881",  "250,276"],  False, WHITE),
        ("Gross Margin %",       ["83.2%",    "83.9%",    "84.6%",    "84.7%",    "81.0%"],    False, GREY),
        ("EBITDA",               ["64,669",   "82,171",   "111,987",  "147,446",  "149,640"],  False, WHITE),
        ("EBITDA Margin %",      ["45.9%",    "46.4%",    "48.2%",    "50.8%",    "48.4%"],    False, GREY),
        ("Operating Income (EBIT)",["58,644", "74,809",   "102,574",  "128,339",  "127,658"],  False, WHITE),
        ("EBIT Margin %",        ["41.7%",    "42.3%",    "44.2%",    "44.2%",    "41.3%"],    False, GREY),
        ("Net Income",           ["47,757",   "55,525",   "83,683",   "100,988",  "102,434"],  True,  LIGHT_GREEN),
        ("Net Margin %",         ["33.9%",    "31.4%",    "36.0%",    "34.8%",    "33.1%"],    False, GREY),
        ("EPS Diluted (DKK)",    ["10.37",    "12.22",    "18.62",    "22.63",    "23.03"],    True,  WHITE),
        ("YoY Revenue Growth",   ["—",        "+25.7%",   "+31.3%",   "+25.0%",   "+6.4%"],    False, YELLOW),
    ]
    for i, (label, vals, bold, bg) in enumerate(income_data):
        r = 5 + i
        lc = ws.cell(r, 2, label)
        lc.font = _font(bold=bold, size=FONT_SIZE)
        lc.fill = _fill(bg)
        lc.alignment = _align("left", "center")
        lc.border = _border()
        for ci, v in enumerate(vals):
            _data_cell(ws, r, 3 + ci, v, bg=bg, bold=bold, align="right")
        ws.row_dimensions[r].height = 20

    _section_hdr(ws, 17, 2, "USD APPROXIMATE (at ~6.6 DKK/USD, illustrative)", span=6)
    _hdr_cell(ws, 18, 2, "Metric (USD ~)", bg=NVO_BLUE, fg=WHITE, align="left")
    for ci, yr in enumerate(years):
        _hdr_cell(ws, 18, 3 + ci, yr, bg=NVO_BLUE, fg=WHITE)

    usd_data = [
        ("Revenue ($B)",      ["~$21.3B", "~$26.5B", "~$35.2B", "~$42.1B", "~$46.8B"]),
        ("EBITDA ($B)",       ["~$9.8B",  "~$12.3B", "~$17.0B", "~$22.3B", "~$22.7B"]),
        ("Net Income ($B)",   ["~$7.2B",  "~$8.3B",  "~$12.7B", "~$15.3B", "~$15.5B"]),
        ("EPS Diluted (USD)", ["~$1.57",  "~$1.84",  "~$2.70",  "~$3.28",  "~$3.49"]),
    ]
    for i, (label, vals) in enumerate(usd_data):
        r = 19 + i
        bg = GREY if i % 2 == 0 else WHITE
        lc = ws.cell(r, 2, label)
        lc.font = _font(size=FONT_SIZE)
        lc.fill = _fill(bg)
        lc.alignment = _align("left", "center")
        lc.border = _border()
        for ci, v in enumerate(vals):
            _data_cell(ws, r, 3 + ci, v, bg=bg, align="right")
        ws.row_dimensions[r].height = 20

    _section_hdr(ws, 24, 2, "SEGMENT REVENUE — FY2024 (DKK Millions, Most Granular Available)", span=6)
    seg_hdrs2 = ["Segment", "FY2024 (DKK M)", "% of Total", "YoY Growth (CER)", "Key Product", ""]
    for ci, h in enumerate(seg_hdrs2):
        _hdr_cell(ws, 25, 2 + ci, h, bg=NVO_BLUE, fg=WHITE)

    seg24 = [
        ("Diabetes Care",        "206,618", "71.1%", "+20% CER", "Ozempic, Rybelsus, Victoza", ""),
        ("  of which Ozempic",   "120,342", "41.4%", "+26% CER", "Semaglutide injection",      ""),
        ("Obesity Care",         "65,146",  "22.4%", "+57% CER", "Wegovy, Saxenda",             ""),
        ("  of which Wegovy",    "~58,200", "~20.0%","~+86%",    "Semaglutide 2.4mg",          ""),
        ("Rare Disease",         "~17,500", "~6.0%", "+9% CER",  "NovoSeven, Norditropin",      ""),
        ("Other",                "~1,139",  "~0.4%", "—",        "Other biopharmaceuticals",    ""),
        ("TOTAL",                "290,403", "100%",  "+25% CER", "All segments",                ""),
    ]
    for i, row_d in enumerate(seg24):
        r = 26 + i
        bg = GREY if i % 2 == 0 else WHITE
        bg = LIGHT_GREEN if i == 6 else bg
        bold = (i == 6)
        for ci, v in enumerate(row_d):
            _data_cell(ws, r, 2 + ci, v, bg=bg, bold=bold, align="center" if ci > 0 else "left")
        ws.row_dimensions[r].height = 20

    _section_hdr(ws, 34, 2, "FY2026 GUIDANCE & CONSENSUS", span=6)
    guidance_rows = [
        ("FY2026 Sales Growth (CER)", "-5% to -13%", "First-ever guidance for annual sales decline since 2017"),
        ("FY2026 Op. Profit Growth",  "Similar range","Comparable decline to sales guidance"),
        ("FY2026 Revenue (consensus)","DKK 291,890M", "~-5.6% vs FY2025 (DKK 309,064M)"),
        ("FY2026 EPS (consensus)",    "DKK 21.37",    "~-7.2% vs FY2025 (DKK 23.03)"),
        ("FY2026 EPS USD (est.)",     "~$3.20",       "At 6.6 DKK/USD; Forward P/E ~12-13x at $41.28"),
        ("Key headwinds",             "IRA pricing + Lilly competition + CagriSema failure", "Triple headwind in 2026"),
    ]
    for ci in range(3):
        _hdr_cell(ws, 35, 2 + ci, ["Metric", "Value", "Notes"][ci], bg=NVO_BLUE, fg=WHITE)
    for i, (m, v, n) in enumerate(guidance_rows):
        r = 36 + i
        bg = GREY if i % 2 == 0 else WHITE
        _data_cell(ws, r, 2, m, bg=bg, align="left", bold=True)
        _data_cell(ws, r, 3, v, bg=bg, align="center")
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=7)
        nc = ws.cell(r, 4, n)
        nc.font = _font(size=FONT_SIZE)
        nc.fill = _fill(bg)
        nc.alignment = _align("left", "center", wrap=True)
        nc.border = _border()
        ws.row_dimensions[r].height = 22

    _note_row(ws, 43, 2, "Source: StockAnalysis.com; Novo Nordisk annual reports; macrotrends (USD EPS verified: $3.49/2025, $3.28/2024, $2.70/2023). USD ~DKK 6.6. Segment FY2025 Obesity DKK 82,347M confirmed; Diabetes FY2025 estimated.", span=6)


# ─────────────────────────────────────────────────────────────────────────────
# TAB 5 — BALANCE SHEET
# ─────────────────────────────────────────────────────────────────────────────
def build_balance_sheet(wb):
    ws = wb.create_sheet("Balance Sheet")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, [2, 32, 14, 14, 14, 14, 14, 2])

    for r in range(1, 55):
        for c in range(1, 9):
            ws.cell(r, c).fill = _fill(WHITE)

    _hdr_cell(ws, 1, 2, "Novo Nordisk A/S — Balance Sheet (DKK Millions)", bg=DARK_NAVY, fg=WHITE, span=6)
    ws.row_dimensions[1].height = 32

    years = ["FY2021", "FY2022", "FY2023", "FY2024", "FY2025"]

    _section_hdr(ws, 3, 2, "BALANCE SHEET — DKK MILLIONS", span=6)
    _hdr_cell(ws, 4, 2, "Metric", bg=NVO_BLUE, fg=WHITE, align="left")
    for ci, yr in enumerate(years):
        _hdr_cell(ws, 4, 3 + ci, yr, bg=NVO_BLUE, fg=WHITE)
    ws.row_dimensions[4].height = 22

    bs_data = [
        ("Cash & Equivalents",    ["10,720",  "12,653",  "14,392",  "15,655",   "26,464"],  False, LIGHT_GREEN),
        ("Intangible Assets",     ["43,171",  "50,939",  "60,406",  "90,804",   "110,208"], False, GREY),
        ("Goodwill",              ["—",       "—",       "—",       "20,017",   "19,845"],  False, WHITE),
        ("Total Assets",          ["194,508", "241,257", "314,486", "465,630",  "542,902"], True,  LIGHT_BLUE),
        ("Total Debt",            ["26,645",  "25,784",  "27,006",  "102,787",  "130,958"], False, LIGHT_RED),
        ("Shareholders' Equity",  ["70,746",  "83,486",  "106,561", "143,486",  "194,047"], True,  LIGHT_GREEN),
        ("Net Debt (Debt−Cash)",  ["15,925",  "13,131",  "12,614",  "87,132",   "104,494"], False, YELLOW),
    ]
    for i, (label, vals, bold, bg) in enumerate(bs_data):
        r = 5 + i
        lc = ws.cell(r, 2, label)
        lc.font = _font(bold=bold, size=FONT_SIZE)
        lc.fill = _fill(bg)
        lc.alignment = _align("left", "center")
        lc.border = _border()
        for ci, v in enumerate(vals):
            _data_cell(ws, r, 3 + ci, v, bg=bg, bold=bold, align="right")
        ws.row_dimensions[r].height = 20

    _section_hdr(ws, 13, 2, "DERIVED BALANCE SHEET METRICS", span=6)
    derived_hdrs = ["Metric", "FY2021", "FY2022", "FY2023", "FY2024", "FY2025"]
    for ci, h in enumerate(derived_hdrs):
        _hdr_cell(ws, 14, 2 + ci, h, bg=NVO_BLUE, fg=WHITE, align="left" if ci == 0 else "center")

    derived_data = [
        ("Net Debt / EBITDA",   ["0.25x", "0.16x", "0.11x", "0.59x", "0.70x"], GREY),
        ("Debt / Equity",       ["0.38x", "0.31x", "0.25x", "0.72x", "0.67x"], WHITE),
        ("Current Ratio",       ["~1.6x", "~1.5x", "~1.5x", "~1.4x", "~1.3x"], GREY),
        ("Book Value/Share DKK",["~15.9", "~18.7", "~23.9", "~32.2", "~43.6"], WHITE),
    ]
    for i, (label, vals, bg) in enumerate(derived_data):
        r = 15 + i
        lc = ws.cell(r, 2, label)
        lc.font = _font(size=FONT_SIZE)
        lc.fill = _fill(bg)
        lc.alignment = _align("left", "center")
        lc.border = _border()
        for ci, v in enumerate(vals):
            _data_cell(ws, r, 3 + ci, v, bg=bg, align="right")
        ws.row_dimensions[r].height = 20

    _note_row(ws, 20, 2, "KEY NOTE: Total debt jumped from DKK 27B (FY2023) to DKK 103B (FY2024) due to the Catalent fill-finish acquisition (~$11.7B USD ≈ DKK 80B). This strategic move secured GLP-1 manufacturing capacity but significantly increased balance sheet leverage. Net Debt/EBITDA 0.70x remains conservative. Catalent integration is still in progress.", span=6)

    _section_hdr(ws, 22, 2, "BALANCE SHEET QUALITY ASSESSMENT", span=6)
    quality = [
        ("Liquidity",        "ADEQUATE",     "Cash DKK 26.5B (FY2025); current ratio ~1.3x; ample short-term liquidity despite higher debt"),
        ("Leverage",         "LOW",          "Net Debt/EBITDA 0.70x — conservative for an investment-grade pharma. Catalent debt well within comfort range."),
        ("Intangibles",      "WATCH",        "Intangibles jumped from DKK 60B (FY2023) to DKK 110B (FY2025) — largely from Catalent acquisition + existing IP. Goodwill $19.8B from Catalent."),
        ("Equity Growth",    "STRONG",       "Equity grew from DKK 71B (FY2021) to DKK 194B (FY2025) — retained earnings accumulation; healthy."),
        ("Capital Structure","CONSERVATIVE", "Predominantly equity-financed with strategic debt for manufacturing capacity. No financial stress risk."),
    ]
    for ci in range(3):
        _hdr_cell(ws, 23, 2 + ci, ["Category", "Assessment", "Commentary"][ci], bg=NVO_BLUE, fg=WHITE)
    for i, (cat, assess, comment) in enumerate(quality):
        r = 24 + i
        bg = GREY if i % 2 == 0 else WHITE
        assess_col = GREEN if assess in ("ADEQUATE", "STRONG", "CONSERVATIVE", "LOW") else (NVO_RED if assess == "HIGH" else "000000")
        _data_cell(ws, r, 2, cat, bg=bg, align="left", bold=True)
        _data_cell(ws, r, 3, assess, bg=bg, align="center", colour=assess_col, bold=True)
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=7)
        cc = ws.cell(r, 4, comment)
        cc.font = _font(size=FONT_SIZE)
        cc.fill = _fill(bg)
        cc.alignment = _align("left", "center", wrap=True)
        cc.border = _border()
        ws.row_dimensions[r].height = 26


# ─────────────────────────────────────────────────────────────────────────────
# TAB 6 — CASH FLOW ANALYSIS
# ─────────────────────────────────────────────────────────────────────────────
def build_cashflow(wb):
    ws = wb.create_sheet("Cash Flow Analysis")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, [2, 32, 14, 14, 14, 14, 14, 2])

    for r in range(1, 55):
        for c in range(1, 9):
            ws.cell(r, c).fill = _fill(WHITE)

    _hdr_cell(ws, 1, 2, "Novo Nordisk A/S — Cash Flow Analysis (DKK Millions)", bg=DARK_NAVY, fg=WHITE, span=6)
    ws.row_dimensions[1].height = 32

    years = ["FY2021", "FY2022", "FY2023", "FY2024", "FY2025"]

    _section_hdr(ws, 3, 2, "CASH FLOW STATEMENT — DKK MILLIONS", span=6)
    _hdr_cell(ws, 4, 2, "Metric", bg=NVO_BLUE, fg=WHITE, align="left")
    for ci, yr in enumerate(years):
        _hdr_cell(ws, 4, 3 + ci, yr, bg=NVO_BLUE, fg=WHITE)
    ws.row_dimensions[4].height = 22

    cf_data = [
        ("Operating Cash Flow (CFO)", ["55,000",   "78,887",   "108,908",  "120,968",  "119,102"],  True,  LIGHT_GREEN),
        ("Capital Expenditures",      ["(6,335)",  "(12,146)", "(25,806)", "(47,164)", "(60,140)"],  False, LIGHT_RED),
        ("Free Cash Flow (FCF)",      ["48,665",   "66,741",   "83,102",   "73,804",   "58,962"],   True,  LIGHT_GREEN),
        ("FCF Margin %",              ["34.6%",    "37.7%",    "35.8%",    "25.4%",    "19.1%"],    False, GREY),
        ("Dividends Paid",            ["(21,517)", "(25,303)", "(31,767)", "(44,140)", "(51,763)"],  False, LIGHT_RED),
        ("Share Buybacks",            ["(19,447)", "(24,086)", "(29,924)", "(20,181)", "(1,388)"],   False, YELLOW),
        ("Capex / Revenue %",         ["4.5%",     "6.9%",     "11.1%",    "16.2%",    "19.5%"],    False, GREY),
    ]
    for i, (label, vals, bold, bg) in enumerate(cf_data):
        r = 5 + i
        lc = ws.cell(r, 2, label)
        lc.font = _font(bold=bold, size=FONT_SIZE)
        lc.fill = _fill(bg)
        lc.alignment = _align("left", "center")
        lc.border = _border()
        for ci, v in enumerate(vals):
            _data_cell(ws, r, 3 + ci, v, bg=bg, bold=bold, align="right")
        ws.row_dimensions[r].height = 20

    _section_hdr(ws, 13, 2, "FREE CASH FLOW QUALITY ANALYSIS", span=6)
    obs_hdrs = ["Observation", "Detail"]
    _hdr_cell(ws, 14, 2, "Observation", bg=NVO_BLUE, fg=WHITE)
    ws.merge_cells(start_row=14, start_column=3, end_row=14, end_column=7)
    _hdr_cell(ws, 14, 3, "Detail", bg=NVO_BLUE, fg=WHITE, span=5)

    observations = [
        ("CFO Growth",          "CFO grew from DKK 55B (2021) to DKK 121B (2024), then flat in 2025. Excellent cash conversion from operations — Net Income to CFO ratio consistently >1.0x indicating high earnings quality and non-cash add-backs."),
        ("Capex Surge",         "Capex 10x from DKK 6.3B (2021) to DKK 60.1B (2025). This is manufacturing buildout (new factories across Denmark/US) + Catalent site integration. Peak capex is likely 2025-2026; expected to taper post-2027 as facilities come online."),
        ("FCF Compression",     "FCF peaked at DKK 83.1B (FY2023), declined 29% to DKK 59.0B by FY2025. FCF margin compressed from 35.8% to 19.1%. This is purely capex-driven — underlying cash generation remains strong."),
        ("Buybacks Suspended",  "Share buybacks nearly stopped in 2025 (DKK 1.4B vs DKK 30B in 2023). Company prioritizing Catalent integration and dividend. Returns to buybacks likely post-2026 as FCF recovers."),
        ("Dividend Payout",     "Dividends grew every year: DKK 21.5B → 51.8B (2021-2025). FY2025 payout from FCF = 88% — very high but sustainable given the growth trajectory of underlying CFO."),
        ("FCF Outlook",         "FCF expected to remain compressed in FY2026 (capex + revenue decline). Recovery anticipated FY2027+ as capex tapers, manufacturing efficiencies improve, and oral amycretin (if approved) drives revenue growth."),
    ]
    for i, (obs, detail) in enumerate(observations):
        r = 15 + i
        bg = GREY if i % 2 == 0 else WHITE
        _data_cell(ws, r, 2, obs, bg=bg, align="left", bold=True)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)
        dc = ws.cell(r, 3, detail)
        dc.font = _font(size=FONT_SIZE)
        dc.fill = _fill(bg)
        dc.alignment = _align("left", "center", wrap=True)
        dc.border = _border()
        ws.row_dimensions[r].height = 32

    _note_row(ws, 22, 2, "FCF VERDICT: Despite the near-term compression, Novo's cash generation engine is fundamentally intact. The capex investment is deliberate (manufacturing capacity for 650M+ person GLP-1 market) and management has been transparent. FCF should recover 2027+ as the capex cycle peaks. The dividend remains safe at current CFO levels.", span=6)


# ─────────────────────────────────────────────────────────────────────────────
# TAB 7 — RETURN ON CAPITAL
# ─────────────────────────────────────────────────────────────────────────────
def build_returns(wb):
    ws = wb.create_sheet("Return on Capital")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, [2, 32, 14, 14, 14, 14, 14, 2])

    for r in range(1, 45):
        for c in range(1, 9):
            ws.cell(r, c).fill = _fill(WHITE)

    _hdr_cell(ws, 1, 2, "Novo Nordisk A/S — Return on Capital", bg=DARK_NAVY, fg=WHITE, span=6)
    ws.row_dimensions[1].height = 32

    years = ["FY2021", "FY2022", "FY2023", "FY2024", "FY2025"]

    _section_hdr(ws, 3, 2, "RETURN METRICS — 5 YEAR HISTORY", span=6)
    _hdr_cell(ws, 4, 2, "Metric", bg=NVO_BLUE, fg=WHITE, align="left")
    for ci, yr in enumerate(years):
        _hdr_cell(ws, 4, 3 + ci, yr, bg=NVO_BLUE, fg=WHITE)
    ws.row_dimensions[4].height = 22

    returns_data = [
        ("ROE (NI / Avg Equity)",     ["~74%",   "~72%",   "~88%",   "~83%",  "~62%"],  LIGHT_GREEN),
        ("ROA (NI / Avg Assets)",     ["~27%",   "~26%",   "~30%",   "~25%",  "~20%"],  GREY),
        ("ROIC (est.)",               ["~50%+",  "~50%+",  "~55%+",  "~45%",  "~35%"],  LIGHT_GREEN),
        ("WACC (estimate)",           ["~8%",    "~8%",    "~8-9%",  "~8-9%", "~8-9%"], GREY),
        ("Economic Spread (ROIC-WACC)",["~42%+", "~42%+",  "~46%+",  "~36%+", "~26%+"], YELLOW),
    ]
    for i, (label, vals, bg) in enumerate(returns_data):
        r = 5 + i
        lc = ws.cell(r, 2, label)
        lc.font = _font(bold=(i in (2, 4)), size=FONT_SIZE)
        lc.fill = _fill(bg)
        lc.alignment = _align("left", "center")
        lc.border = _border()
        for ci, v in enumerate(vals):
            _data_cell(ws, r, 3 + ci, v, bg=bg, bold=(i in (2, 4)), align="right")
        ws.row_dimensions[r].height = 20

    _section_hdr(ws, 11, 2, "CAPITAL RETURNS CONTEXT", span=6)
    context_rows = [
        ("Why ROE is Extraordinarily High",
         "Novo's net margins of 33-36% on an equity base built from retained earnings (not large equity raises) produces outsized ROE. The Foundation structure means no dilutive secondary offerings; equity grows organically from profits."),
        ("Why Returns Are Declining",
         "Equity base expanding from retained earnings + Catalent acquisition leverage. Higher capex intensity (19.5% of revenue in FY2025) reduces ROIC. IRA pricing impact will further compress EBIT margins. Still exceptional at 35% ROIC vs ~9% WACC (26pp economic spread)."),
        ("Benchmark Comparison",
         "Global pharma median ROIC: ~12-15%. Novo at ~35% (FY2025) is still ~2-3x the industry median even after the decline. This reflects the exceptional cash-generative nature of GLP-1 blockbusters with limited manufacturing capex (historically)."),
        ("Outlook for Returns",
         "ROIC likely to stay in 25-35% range through 2026-2027 as revenue declines and capex remains elevated. Recovery expected 2027-2028 as capex peaks and manufacturing reaches full utilization. Oral amycretin success would add a new high-ROIC revenue stream."),
    ]
    for ci in range(2):
        _hdr_cell(ws, 12, 2 + ci * 3, ["Topic", "Commentary"][ci], bg=NVO_BLUE, fg=WHITE, span=3)
    for i, (topic, comment) in enumerate(context_rows):
        r = 13 + i
        bg = GREY if i % 2 == 0 else WHITE
        _data_cell(ws, r, 2, topic, bg=bg, align="left", bold=True)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)
        cc = ws.cell(r, 3, comment)
        cc.font = _font(size=FONT_SIZE)
        cc.fill = _fill(bg)
        cc.alignment = _align("left", "center", wrap=True)
        cc.border = _border()
        ws.row_dimensions[r].height = 38

    _note_row(ws, 18, 2, "ROIC VERDICT: Even at the post-decline level of ~35%, Novo's ROIC exceeds its ~9% WACC by ~26 percentage points — generating substantial economic value per incremental invested dollar. The business is not broken; the decline is cyclical (pricing + capex). For context, Lilly's ROIC is ~25-30% while benefiting from GLP-1 market share gains.", span=6)


# ─────────────────────────────────────────────────────────────────────────────
# TAB 8 — MANAGEMENT
# ─────────────────────────────────────────────────────────────────────────────
def build_management(wb):
    ws = wb.create_sheet("Management")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, [2, 24, 18, 16, 30, 2])

    for r in range(1, 65):
        for c in range(1, 7):
            ws.cell(r, c).fill = _fill(WHITE)

    _hdr_cell(ws, 1, 2, "Novo Nordisk A/S — Management & Capital Allocation", bg=DARK_NAVY, fg=WHITE, span=4)
    ws.row_dimensions[1].height = 32

    _section_hdr(ws, 3, 2, "KEY EXECUTIVES", span=4)
    exec_hdrs = ["Name", "Role", "Tenure", "Key Contributions"]
    for ci, h in enumerate(exec_hdrs):
        _hdr_cell(ws, 4, 2 + ci, h, bg=NVO_BLUE, fg=WHITE)

    execs = [
        ("Maziar Mike Doustdar", "President & CEO",             "Since Aug 7, 2025",        "Former EVP International Operations. Deep insider knowledge. Took helm during GLP-1 competitive crisis. Driving pivot toward oral pipeline and EM expansion."),
        ("Lars Fruergaard Jørgensen","Former CEO",              "2017–Aug 2025 (34 yrs at NN)","Oversaw Ozempic/Wegovy launch and 10× revenue growth (DKK 107B→290B). Left during competitive headwinds. Strong legacy of value creation."),
        ("Karsten Munk Knudsen", "CFO",                         "Multiple years",            "Capital allocation architect. Managed Catalent acquisition financing. Conservative balance sheet management. Maintained dividend growth through cycle."),
        ("Martin Holst Lange",   "Chief Scientific Officer",    "Since Aug 2025",            "R&D strategy. Leads amycretin, oral GLP-1, and rare disease pipeline. Key executive for the re-rating catalyst pipeline."),
        ("Emil Kongshøj Larsen", "EVP International Operations","Since Aug 2025",            "Succeeds Doustdar in international role. EM expansion strategy execution — key for geographic diversification away from US pricing pressure."),
    ]
    for i, (name, role, tenure, contrib) in enumerate(execs):
        r = 5 + i
        bg = GREY if i % 2 == 0 else WHITE
        _data_cell(ws, r, 2, name, bg=bg, align="left", bold=True)
        _data_cell(ws, r, 3, role, bg=bg, align="left")
        _data_cell(ws, r, 4, tenure, bg=bg, align="center")
        cc = ws.cell(r, 5, contrib)
        cc.font = _font(size=FONT_SIZE)
        cc.fill = _fill(bg)
        cc.alignment = _align("left", "center", wrap=True)
        cc.border = _border()
        ws.row_dimensions[r].height = 32

    _section_hdr(ws, 11, 2, "COMPENSATION & INCENTIVE STRUCTURE", span=4)
    comp_rows = [
        ("CEO Base Salary",   "~DKK 15-18M (~$2.1-2.7M USD)",  "Moderate by US pharma standards; Danish governance culture prioritizes restraint"),
        ("CEO Total Comp",    "~DKK 40-60M (~$6-9M USD)",       "Including LTI (PSUs); well below US pharma CEO peers (~$20-40M USD)"),
        ("LTI Structure",     "Performance Share Units (PSUs), 3-year vest", "Aligns executives with multi-year performance, not quarterly EPS"),
        ("PSU Metrics",       "Revenue growth (CER), Op. profit growth, TSR vs peers, ESG/patient outcomes", "Strong alignment with long-term value creation; ESG component unusual but reflects Danish culture"),
        ("CEO Ownership",     "Doustdar: ~0.002% (~$4.67M)",   "Small by US standards; typical for Danish/European executive culture. Foundation alignment substitutes."),
        ("Insider Activity",  "0 open-market purchases in last 24 months", "NEUTRAL — No open-market buys during -57% drawdown. Typical of European executive culture, not necessarily bearish."),
        ("Say-on-Pay",        "Strong historical shareholder support",  "Danish governance culture; Foundation (67% votes) provides alignment rather than pure ownership stake."),
    ]
    for ci in range(3):
        _hdr_cell(ws, 12, 2 + ci, ["Item", "Value", "Commentary"][ci], bg=NVO_BLUE, fg=WHITE)
    for i, (item, value, comment) in enumerate(comp_rows):
        r = 13 + i
        bg = GREY if i % 2 == 0 else WHITE
        _data_cell(ws, r, 2, item, bg=bg, align="left", bold=True)
        _data_cell(ws, r, 3, value, bg=bg, align="left")
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=5)
        nc = ws.cell(r, 4, comment)
        nc.font = _font(size=FONT_SIZE)
        nc.fill = _fill(bg)
        nc.alignment = _align("left", "center", wrap=True)
        nc.border = _border()
        ws.row_dimensions[r].height = 26

    _section_hdr(ws, 21, 2, "CAPITAL ALLOCATION TRACK RECORD", span=4)
    alloc_hdrs = ["Year", "Decision", "Outcome", "Assessment"]
    for ci, h in enumerate(alloc_hdrs):
        _hdr_cell(ws, 22, 2 + ci, h, bg=NVO_BLUE, fg=WHITE)

    alloc_rows = [
        ("2017–2023", "Ozempic/Wegovy buildout — $5B+ capex and R&D",                          "Created $400B+ market cap at peak; world-class GLP-1 franchise",        "EXCELLENT",  LIGHT_GREEN),
        ("2024–2025", "Catalent acquisition ($11.7B) — fill-finish manufacturing capacity",      "Secured supply chain; ~80B DKK debt; now burden during revenue downturn","POSITIVE / NEUTRAL", YELLOW),
        ("2024–2025", "Suspended buybacks; maintained growing dividend",                          "Preserved cash during Catalent integration; dividend grew 17.3%",       "NEUTRAL",    GREY),
        ("2023–2026", "CagriSema Phase 3 programme continuation",                                 "Failed non-inferiority vs Zepbound (Feb 2026); significant setback",    "WARNING",    LIGHT_RED),
        ("2025",      "CEO succession — appointed insider Doustdar",                              "Continuity vs fresh thinking debate; decision made during crisis",       "NEUTRAL",    GREY),
        ("2024+",     "Amycretin Phase 3 programme (GLP-1/amylin oral)",                          "Phase 3 ongoing; Phase 1/2 data very promising; key catalyst",           "POSITIVE",   LIGHT_GREEN),
    ]
    for i, (yr, decision, outcome, assess, bg) in enumerate(alloc_rows):
        r = 23 + i
        assess_col = GREEN if "EXCELLENT" in assess or "POSITIVE" in assess else (NVO_RED if "WARNING" in assess else "000000")
        _data_cell(ws, r, 2, yr, bg=bg, align="center")
        _data_cell(ws, r, 3, decision, bg=bg, align="left")
        dc = ws.cell(r, 4, outcome)
        dc.font = _font(size=FONT_SIZE)
        dc.fill = _fill(bg)
        dc.alignment = _align("left", "center", wrap=True)
        dc.border = _border()
        _data_cell(ws, r, 5, assess, bg=bg, align="center", colour=assess_col, bold=True)
        ws.row_dimensions[r].height = 28

    _section_hdr(ws, 30, 2, "NOVO NORDISK FOUNDATION — DOMINANT SHAREHOLDER", span=4)
    fnd_rows = [
        ("Owner Type",     "Non-profit foundation with dual purpose: long-term development of Novo Nordisk + funding Danish scientific research"),
        ("Economic Stake", "~28% of total economic interest (B shares equivalent)"),
        ("Voting Rights",  "~67% of total votes — via A-share structure (A shares carry 10× voting rights vs B shares)"),
        ("Strategic Value","Prevents hostile takeover; insulates management from short-term activist pressure; enables 10-20 year R&D planning horizons"),
        ("Capital Impact", "Foundation alignment means no pressure to maximize short-term EPS. Div growth and R&D investment are prioritized — POSITIVE for long-term holders"),
    ]
    for i, (item, detail) in enumerate(fnd_rows):
        r = 31 + i
        bg = LIGHT_TEAL if i % 2 == 0 else WHITE
        _data_cell(ws, r, 2, item, bg=bg, align="left", bold=True)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
        dc = ws.cell(r, 3, detail)
        dc.font = _font(size=FONT_SIZE)
        dc.fill = _fill(bg)
        dc.alignment = _align("left", "center", wrap=True)
        dc.border = _border()
        ws.row_dimensions[r].height = 28

    _note_row(ws, 37, 2, "MANAGEMENT VERDICT: NEUTRAL. New CEO Doustdar is a deep insider — continuity is good but the question is whether a more transformational external leader was needed. Capital allocation has been strong historically (Ozempic buildout); the Catalent bet is strategic but timing was unfortunate. No insider buying during the -57% decline is a mild negative, explained by Danish governance norms rather than bearish insider views.", span=4)


# ─────────────────────────────────────────────────────────────────────────────
# TAB 9 — RISKS
# ─────────────────────────────────────────────────────────────────────────────
def build_risks(wb):
    ws = wb.create_sheet("Risks")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, [2, 26, 12, 12, 24, 16, 2])

    for r in range(1, 70):
        for c in range(1, 8):
            ws.cell(r, c).fill = _fill(WHITE)

    _hdr_cell(ws, 1, 2, "Novo Nordisk A/S — Risk Assessment", bg=DARK_NAVY, fg=WHITE, span=5)
    ws.row_dimensions[1].height = 32

    def risk_section(title, start_row, risks):
        _section_hdr(ws, start_row, 2, title, span=5)
        hdrs = ["Risk", "Probability", "Impact", "Mitigation", "Net Risk"]
        for ci, h in enumerate(hdrs):
            _hdr_cell(ws, start_row + 1, 2 + ci, h, bg=NVO_BLUE, fg=WHITE)
        prob_map = {"HIGH": (LIGHT_RED, RED), "MEDIUM": (YELLOW, "000000"), "LOW": (LIGHT_GREEN, GREEN)}
        for i, (risk, prob, impact, mitigation, net) in enumerate(risks):
            r = start_row + 2 + i
            bg = GREY if i % 2 == 0 else WHITE
            pc = ws.cell(r, 2, risk)
            pc.font = _font(size=FONT_SIZE)
            pc.fill = _fill(bg)
            pc.alignment = _align("left", "center", wrap=True)
            pc.border = _border()
            for col_i, (val, cat) in enumerate([(prob, prob), (impact, impact)]):
                fill_c, font_c = prob_map.get(cat, (GREY, "000000"))
                _data_cell(ws, r, 3 + col_i, val, bg=fill_c, align="center", colour=font_c, bold=True)
            mc = ws.cell(r, 5, mitigation)
            mc.font = _font(size=FONT_SIZE)
            mc.fill = _fill(bg)
            mc.alignment = _align("left", "center", wrap=True)
            mc.border = _border()
            fill_n, font_n = prob_map.get(net, (GREY, "000000"))
            _data_cell(ws, r, 6, net, bg=fill_n, align="center", colour=font_n, bold=True)
            ws.row_dimensions[r].height = 30

    financial_risks = [
        ("IRA Medicare Price Negotiation", "HIGH", "HIGH",
         "Ozempic: $274/mo max fair price; Wegovy: $385/mo (vs ~$900+ gross) from Jan 2027. Already priced in guidance.",
         "HIGH"),
        ("FCF Compression (capex cycle)", "HIGH", "MEDIUM",
         "Manufacturing capacity will reduce capex intensity post-2027; operating leverage to return as factories utilise.",
         "MEDIUM"),
        ("FX Headwinds (USD vs DKK)", "MEDIUM", "MEDIUM",
         "~70% USD revenue reported in DKK; no formal FX hedging disclosed. Natural hedge via USD-cost US manufacturing sites.",
         "MEDIUM"),
        ("Dividend Sustainability", "LOW", "MEDIUM",
         "Dividend payout from FCF = 88% (FY2025); high but CFO remains at DKK 119B. Foundation prefers growing dividends; risk is if FCF falls further.",
         "LOW"),
    ]
    risk_section("FINANCIAL RISKS", 3, financial_risks)

    operational_risks = [
        ("Market Share Loss to Eli Lilly", "HIGH", "HIGH",
         "Oral GLP-1 amycretin pipeline; next-gen injectables; EM geographic expansion; brand loyalty in diabetes segment.",
         "HIGH"),
        ("Amycretin Phase 3 Failure", "MEDIUM", "HIGH",
         "Phase 1/2 data very promising. If fails: no clear second pipeline candidate; stock likely to re-test bear case $28-33.",
         "MEDIUM"),
        ("Manufacturing/Catalent Integration", "MEDIUM", "MEDIUM",
         "Company took direct control to manage quality/supply; US FDA compliance risk at Catalent sites. Internal expertise growing.",
         "MEDIUM"),
        ("CEO Transition Risk", "MEDIUM", "MEDIUM",
         "Doustdar was internal — deep institutional knowledge. Foundation provides stability. Strategy continuity high.",
         "LOW"),
    ]
    risk_section("OPERATIONAL RISKS", 10, operational_risks)

    macro_risks = [
        ("US Healthcare Policy", "HIGH", "HIGH",
         "IRA already in effect; further policy risk possible. Novo actively lobbying. List price transparency legislation could further pressure.",
         "HIGH"),
        ("GLP-1 Market Growth Slowdown", "LOW", "HIGH",
         "Still 80%+ of adults with obesity are untreated; massive TAM runway globally. Market penetration remains <5%.",
         "LOW"),
        ("Compounding GLP-1 (outsourced copies)", "MEDIUM", "MEDIUM",
         "FDA designated semaglutide off shortage list; compounders required to exit by 2026. Enforcement ongoing.",
         "LOW"),
        ("Generic Entry (pre-patent expiry)", "LOW", "HIGH",
         "Core semaglutide patents protected into 2030s+. Manufacturing complexity provides additional barrier beyond patents.",
         "LOW"),
    ]
    risk_section("MACRO / MARKET RISKS", 17, macro_risks)

    regulatory_risks = [
        ("FDA Label / Safety Action", "LOW", "HIGH",
         "Extensive real-world data 10+ years; ongoing CVOT/FLOW trials positive. Thyroid cancer signal (label warning) remains but no new concerns.",
         "LOW"),
        ("EU / International Pricing Pressure", "MEDIUM", "MEDIUM",
         "National health systems negotiating harder; parallel trade risk in EU. But lower list prices internationally limit downside.",
         "MEDIUM"),
    ]
    risk_section("REGULATORY RISKS", 24, regulatory_risks)

    _note_row(ws, 28, 2, "OVERALL RISK VERDICT: NVO faces an unusually concentrated set of simultaneous headwinds — IRA pricing, Lilly market share, pipeline failure, and management transition. However, these risks are largely known and reflected in the -57% stock decline and ~12x P/E. The key asymmetric risk to the upside is amycretin Phase 3 success. The key downside risk is if oral pipeline also disappoints.", span=5)


# ─────────────────────────────────────────────────────────────────────────────
# TAB 10 — VALUATION
# ─────────────────────────────────────────────────────────────────────────────
def build_valuation(wb):
    ws = wb.create_sheet("Valuation")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, [2, 28, 18, 16, 26, 2])

    for r in range(1, 70):
        for c in range(1, 7):
            ws.cell(r, c).fill = _fill(WHITE)

    _hdr_cell(ws, 1, 2, "Novo Nordisk A/S — Valuation Analysis", bg=DARK_NAVY, fg=WHITE, span=4)
    ws.row_dimensions[1].height = 32

    _section_hdr(ws, 3, 2, "CURRENT TRADING MULTIPLES (April 27, 2026)", span=4)
    mult_hdrs = ["Metric", "NVO Value", "Notes", ""]
    for ci, h in enumerate(mult_hdrs):
        _hdr_cell(ws, 4, 2 + ci, h, bg=NVO_BLUE, fg=WHITE)

    current_multiples = [
        ("Stock Price (ADR)", "$41.28 USD",  "April 27, 2026",                                                                       ""),
        ("Market Cap",        "~$183.5B USD","4,450M diluted shares × $41.28",                                                       ""),
        ("Enterprise Value",  "~$189B USD",  "Market Cap + Net Debt ~$15.8B USD (DKK 104.5B / 6.6)",                                  ""),
        ("P/E (TTM GAAP)",    "~11.8x",      "NI ~$15.5B USD at $183.5B market cap",                                                  ""),
        ("EV/EBITDA (TTM)",   "~8.3x",       "$189B EV / ~$22.7B EBITDA",                                                             ""),
        ("EV/Revenue (TTM)",  "~4.0x",       "$189B / ~$46.8B Revenue",                                                               ""),
        ("Forward P/E (FY2026E)", "~12-13x", "EPS ~$3.20E (DKK 21.37 / 6.6); revenue decline year",                                  ""),
        ("Price/FCF (TTM)",   "~21x",        "$183.5B / $8.9B FCF (~DKK 59B / 6.6)",                                                  ""),
        ("Price/Book",        "~2.3x",       "$183.5B / ($194B equity / 6.6 = $29.4B)",                                               ""),
        ("Dividend Yield",    "~2.6%",       "~$1.76 USD div (~DKK 11.62 / 6.6) at $41.28",                                          ""),
    ]
    for i, (m, v, n, _) in enumerate(current_multiples):
        r = 5 + i
        bg = GREY if i % 2 == 0 else WHITE
        _data_cell(ws, r, 2, m, bg=bg, align="left", bold=True)
        _data_cell(ws, r, 3, v, bg=bg, align="center", bold=True, colour=NVO_BLUE)
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=5)
        nc = ws.cell(r, 4, n)
        nc.font = _font(size=FONT_SIZE)
        nc.fill = _fill(bg)
        nc.alignment = _align("left", "center")
        nc.border = _border()
        ws.row_dimensions[r].height = 20

    _section_hdr(ws, 16, 2, "HISTORICAL MULTIPLE COMPARISON", span=4)
    hist_hdrs = ["Metric", "5-Year Avg", "Current", "Discount to Historical Avg"]
    for ci, h in enumerate(hist_hdrs):
        _hdr_cell(ws, 17, 2 + ci, h, bg=NVO_BLUE, fg=WHITE)

    hist_data = [
        ("P/E",       "~25x",  "~12x",  "-55% discount"),
        ("EV/EBITDA", "~20x",  "~8.3x", "-59% discount"),
        ("EV/Revenue","~8-10x","~4.0x", "-50-60% discount"),
    ]
    for i, (m, avg, curr, disc) in enumerate(hist_data):
        r = 18 + i
        bg = GREY if i % 2 == 0 else WHITE
        _data_cell(ws, r, 2, m, bg=bg, align="left", bold=True)
        _data_cell(ws, r, 3, avg, bg=bg, align="center")
        _data_cell(ws, r, 4, curr, bg=LIGHT_GREEN, align="center", bold=True, colour=GREEN)
        _data_cell(ws, r, 5, disc, bg=LIGHT_RED, align="center", colour=RED)
        ws.row_dimensions[r].height = 20

    _section_hdr(ws, 22, 2, "PEER VALUATION COMPARISON (April 2026)", span=4)
    peer_hdrs = ["Company", "Ticker", "P/E (NTM)", "EV/EBITDA", "Revenue ($B)", "Notes"]
    for ci, h in enumerate(peer_hdrs):
        _hdr_cell(ws, 23, 2 + ci, h, bg=NVO_BLUE, fg=WHITE)

    peer_data = [
        ("Novo Nordisk",  "NVO",    "~12x",  "~8x",  "~$47B", "Subject — deeply discounted vs peers and history"),
        ("Eli Lilly",     "LLY",    "~26x",  "~25x", "~$45B", "GLP-1 market share winner; premium valuation"),
        ("Roche",         "RHHBY",  "~14x",  "~11x", "~$60B", "Diversified pharma; oncology leader"),
        ("Novartis",      "NVS",    "~15x",  "~13x", "~$50B", "Growing branded pharma"),
        ("J&J",           "JNJ",    "~14x",  "~15x", "~$90B", "Diversified; MedTech + Pharma"),
        ("AstraZeneca",   "AZN",    "~18x",  "~15x", "~$55B", "Strong oncology/CV pipeline"),
        ("Peer Median",   "—",      "~15x",  "~13x", "—",     "NVO discount: -20% P/E, -38% EV/EBITDA"),
    ]
    for i, (comp, ticker, pe, eveb, rev, note) in enumerate(peer_data):
        r = 24 + i
        bg = LIGHT_TEAL if i == 0 else (GREY if i == 6 else (GREY if i % 2 == 0 else WHITE))
        bold = (i in (0, 6))
        _data_cell(ws, r, 2, comp, bg=bg, align="left", bold=bold)
        _data_cell(ws, r, 3, ticker, bg=bg, align="center")
        _data_cell(ws, r, 4, pe, bg=bg, align="center", bold=bold)
        _data_cell(ws, r, 5, eveb, bg=bg, align="center", bold=bold)
        _data_cell(ws, r, 6, rev, bg=bg, align="center")
        ws.row_dimensions[r].height = 20

    _section_hdr(ws, 32, 2, "DCF / SCENARIO ANALYSIS (Base Currency DKK; USD at 6.6 DKK/USD)", span=4)
    dcf_hdrs = ["Scenario", "2026-2030 Rev CAGR", "2030 EBIT Margin", "Terminal EV/EBITDA", "Implied USD/Share", "vs Current $41.28"]
    for ci, h in enumerate(dcf_hdrs):
        _hdr_cell(ws, 33, 2 + ci, h, bg=NVO_BLUE, fg=WHITE)

    dcf_scenarios = [
        ("BEAR: Oral pipeline fails; Lilly dominance entrenched",     "-8% then flat",   "35%", "8x",  "~$28-33", "-20% to -33%", LIGHT_RED),
        ("BASE: Revenue recovers 2027+; amycretin partial success",   "-6% then +8% CAGR","40%","12x", "~$48-58", "+16% to +40%", YELLOW),
        ("BULL: Amycretin wins; next growth cycle unlocked",           "Oral wins; +12% CAGR 2027+","44%","16x","~$80-100","+94% to +142%", LIGHT_GREEN),
    ]
    for i, (scenario, cagr, margin, ev, share, vs, bg) in enumerate(dcf_scenarios):
        r = 34 + i
        _data_cell(ws, r, 2, scenario, bg=bg, align="left", bold=True)
        _data_cell(ws, r, 3, cagr, bg=bg, align="center")
        _data_cell(ws, r, 4, margin, bg=bg, align="center")
        _data_cell(ws, r, 5, ev, bg=bg, align="center")
        _data_cell(ws, r, 6, share, bg=bg, align="center", bold=True)
        _data_cell(ws, r, 7, vs, bg=bg, align="center", colour=GREEN if "+" in vs else RED, bold=True) if False else None
        ws.row_dimensions[r].height = 30

    _note_row(ws, 38, 2, "VALUATION VERDICT: At ~12x forward P/E and ~8x EV/EBITDA — 55% and 59% below 5-year averages — NVO embeds substantial pessimism. The stock sits near the base/bear border (~$41 vs base case $48-58), implying limited downside from current levels if the business stabilises. The risk/reward favours patient long-term holders. Amycretin Phase 3 success (expected 2026-2027) is the primary re-rating catalyst. Discount rate: 9% WACC; shares outstanding: 4,450M.", span=4)


# ─────────────────────────────────────────────────────────────────────────────
# TAB 11 — MARKET SENTIMENT
# ─────────────────────────────────────────────────────────────────────────────
def build_sentiment(wb):
    ws = wb.create_sheet("Market Sentiment")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, [2, 24, 14, 14, 26, 2])

    for r in range(1, 70):
        for c in range(1, 7):
            ws.cell(r, c).fill = _fill(WHITE)

    _hdr_cell(ws, 1, 2, "Novo Nordisk A/S — Market Sentiment & Ownership", bg=DARK_NAVY, fg=WHITE, span=4)
    ws.row_dimensions[1].height = 32

    _section_hdr(ws, 3, 2, "ANALYST CONSENSUS (April 2026)", span=4)
    cons_hdrs = ["Rating", "Count", "% of Coverage", "Implication"]
    for ci, h in enumerate(cons_hdrs):
        _hdr_cell(ws, 4, 2 + ci, h, bg=NVO_BLUE, fg=WHITE)

    consensus = [
        ("Strong Buy", "2",  "18%", "2 analysts remain bullish — likely citing valuation + amycretin optionality"),
        ("Buy",        "0",  "0%",  "No pure Buy ratings"),
        ("Hold",       "7",  "64%", "Consensus: wait-and-see; near-term headwinds priced in but no catalysts yet"),
        ("Sell",       "2",  "18%", "Bearish on market share loss and IRA pricing structural damage"),
        ("Strong Sell","0",  "0%",  "No extreme bearish views"),
        ("TOTAL",      "11", "100%",""),
    ]
    rating_bgs = [LIGHT_GREEN, LIGHT_GREEN, YELLOW, LIGHT_RED, LIGHT_RED, LIGHT_BLUE]
    for i, (rating, cnt, pct, impl) in enumerate(consensus):
        r = 5 + i
        bg = rating_bgs[i]
        _data_cell(ws, r, 2, rating, bg=bg, align="center", bold=(i == 5))
        _data_cell(ws, r, 3, cnt, bg=bg, align="center", bold=(i == 5))
        _data_cell(ws, r, 4, pct, bg=bg, align="center", bold=(i == 5))
        ic = ws.cell(r, 5, impl)
        ic.font = _font(size=FONT_SIZE)
        ic.fill = _fill(bg)
        ic.alignment = _align("left", "center", wrap=True)
        ic.border = _border()
        ws.row_dimensions[r].height = 24

    _section_hdr(ws, 12, 2, "PRICE TARGET SUMMARY", span=4)
    pt_data = [
        ("Average Target",   "$51 USD",  "+23.6% upside"),
        ("Median Target",    "$47 USD",  "+13.9% upside"),
        ("High Target",      "$70 USD",  "+69.6% upside (stale Cantor target)"),
        ("Low Target",       "$41 USD",  "Goldman Sachs (Hold) — at current price"),
        ("Current Price",    "$41.28",   "April 27, 2026"),
        ("52-Week Range",    "$35.12 – $81.44", "High was pre-CagriSema failure; low was panic bottom"),
        ("1-Year Return",    "-36.9%",   "vs April 2025 price ~$65"),
        ("YTD 2026 Return",  "approx -30%", "Bottomed near $35 in early 2026"),
    ]
    _hdr_cell(ws, 13, 2, "Metric", bg=NVO_BLUE, fg=WHITE)
    _hdr_cell(ws, 13, 3, "Value",  bg=NVO_BLUE, fg=WHITE)
    _hdr_cell(ws, 13, 4, "Commentary", bg=NVO_BLUE, fg=WHITE, span=2)
    for i, (m, v, c) in enumerate(pt_data):
        r = 14 + i
        bg = GREY if i % 2 == 0 else WHITE
        _data_cell(ws, r, 2, m, bg=bg, align="left", bold=True)
        _data_cell(ws, r, 3, v, bg=bg, align="center", bold=True, colour=NVO_BLUE)
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=5)
        cc = ws.cell(r, 4, c)
        cc.font = _font(size=FONT_SIZE)
        cc.fill = _fill(bg)
        cc.alignment = _align("left", "center")
        cc.border = _border()
        ws.row_dimensions[r].height = 22

    _section_hdr(ws, 23, 2, "RECENT ANALYST ACTIONS (2026)", span=4)
    action_hdrs = ["Firm", "Analyst", "Rating", "Price Target", "Action"]
    for ci, h in enumerate(action_hdrs):
        _hdr_cell(ws, 24, 2 + ci, h, bg=NVO_BLUE, fg=WHITE)

    analyst_actions = [
        ("Goldman Sachs",    "James Quigley",  "Hold",        "$41",              "Downgraded from Strong Buy; drastic cut triggered by CagriSema failure"),
        ("JP Morgan",        "Richard Vosser", "Hold",        "~$45-50",          "Downgraded from Buy on competitive headwinds"),
        ("Argus Research",   "John Eade",      "Hold",        "~$45",             "Downgraded from Strong Buy post CagriSema/2026 guidance"),
        ("Sell-side (2)",    "Various",        "Sell",        "~$35-36",          "Bearish on structural market share loss to Lilly"),
        ("Cantor Fitzgerald","—",              "Buy (stale)", "$160 (Nov 2024)",   "Issued at peak; not updated — stale, not actionable"),
    ]
    for i, (firm, analyst, rating, pt, action) in enumerate(analyst_actions):
        r = 25 + i
        bg = GREY if i % 2 == 0 else WHITE
        _data_cell(ws, r, 2, firm, bg=bg, align="left", bold=True)
        _data_cell(ws, r, 3, analyst, bg=bg, align="center")
        rating_c = GREEN if "Buy" in rating and "stale" not in rating.lower() else (RED if "Sell" in rating else "000000")
        _data_cell(ws, r, 4, rating, bg=bg, align="center", colour=rating_c)
        _data_cell(ws, r, 5, pt, bg=bg, align="center")
        ws.row_dimensions[r].height = 24

    _section_hdr(ws, 31, 2, "MARKET THEMES — TAILWINDS & HEADWINDS", span=4)
    theme_hdrs = ["Theme", "Type", "Detail"]
    for ci in range(3):
        _hdr_cell(ws, 32, 2 + ci, theme_hdrs[ci], bg=NVO_BLUE, fg=WHITE)

    themes = [
        ("GLP-1 Massive TAM",             "TAILWIND",  "Obesity affects 650M+ globally; penetration <5%; $150-200B market by 2030. Secular structural demand intact."),
        ("Eli Lilly Market Share Capture", "HEADWIND",  "Lilly holds ~60% US GLP-1 market. Zepbound clinical superiority proven in head-to-head trial. Novo losing new patients."),
        ("CagriSema Pipeline Failure",     "HEADWIND",  "Feb 2026: failed to beat Zepbound. 23% vs 25.5% weight loss. Triggered analyst downgrades and guidance cuts."),
        ("IRA Medicare Drug Pricing",      "HEADWIND",  "Ozempic: $274/month max fair price from Jan 2027; Wegovy: $385/month. ~60%+ list price cut to negotiated max."),
        ("Oral Amycretin (Phase 3)",       "CATALYST",  "Phase 3 underway; Phase 1/2 showed ~13% weight loss at 12 weeks. Results 2026-2027. Primary re-rating catalyst."),
        ("Compounding Ban Enforcement",    "CATALYST",  "FDA enforcing ban on compounded semaglutide copies. Novo recaptures US volumes lost to compounders in 2025-2026."),
        ("Pediatric/Adolescent Approvals", "CATALYST",  "Semaglutide approved for children with obesity. Expanding addressable market to younger demographic."),
        ("Manufacturing Self-Sufficiency", "TAILWIND",  "Catalent sites secured. Less supply disruption vs 2024 shortages. Long-term margin tailwind as utilization grows."),
        ("CEO Transition",                 "NEUTRAL",   "New CEO Doustdar taking over during competitive crisis. Continuity but strategic clarity still needed."),
        ("DKK/USD FX",                     "HEADWIND",  "~70% revenue USD-denominated but reported in DKK. DKK strength hurts ADR investor returns beyond business fundamentals."),
    ]
    type_bg = {"TAILWIND": LIGHT_GREEN, "HEADWIND": LIGHT_RED, "CATALYST": LIGHT_TEAL, "NEUTRAL": YELLOW}
    for i, (theme, ttype, detail) in enumerate(themes):
        r = 33 + i
        bg = GREY if i % 2 == 0 else WHITE
        _data_cell(ws, r, 2, theme, bg=bg, align="left", bold=True)
        _data_cell(ws, r, 3, ttype, bg=type_bg.get(ttype, GREY), align="center", bold=True)
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=5)
        dc = ws.cell(r, 4, detail)
        dc.font = _font(size=FONT_SIZE)
        dc.fill = _fill(bg)
        dc.alignment = _align("left", "center", wrap=True)
        dc.border = _border()
        ws.row_dimensions[r].height = 26

    _section_hdr(ws, 44, 2, "OWNERSHIP STRUCTURE", span=4)
    own_hdrs = ["Category", "% Ownership", "Notes"]
    for ci, h in enumerate(own_hdrs):
        _hdr_cell(ws, 45, 2 + ci, h, bg=NVO_BLUE, fg=WHITE, span=1 if ci < 2 else 2)

    ownership = [
        ("Novo Nordisk Foundation (Novo Holdings)", "~28% economic / ~67% voting", "Via A-share structure (10x votes); foundation purpose = long-term stewardship of Novo Nordisk"),
        ("Institutional (B-shares/ADRs)",           "~9.87% (1,782 institutions)",  "Top holders: Morgan Stanley, UBS; major US/EU pension funds"),
        ("Retail & Other",                           "~62% free float",              "B-shares and ADR retail investors"),
        ("Insider (executives/board)",               "~0.01%",                       "Very low by US standards; normal for Danish/European governance. New CEO: 0.002% (~$4.67M)"),
    ]
    for i, (cat, pct, note) in enumerate(ownership):
        r = 46 + i
        bg = LIGHT_TEAL if i == 0 else (GREY if i % 2 == 1 else WHITE)
        _data_cell(ws, r, 2, cat, bg=bg, align="left", bold=True)
        _data_cell(ws, r, 3, pct, bg=bg, align="center")
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=5)
        nc = ws.cell(r, 4, note)
        nc.font = _font(size=FONT_SIZE)
        nc.fill = _fill(bg)
        nc.alignment = _align("left", "center", wrap=True)
        nc.border = _border()
        ws.row_dimensions[r].height = 28

    _note_row(ws, 51, 2, "Short Interest: 27.58M ADR shares (~0.6-1% of ADR float). Low short interest. Foundation shares not in free float and not lendable — structural floor for the stock. ADR ratio: 1 ADR = 1 B share.", span=4)


# ─────────────────────────────────────────────────────────────────────────────
# TAB 12 — KEY INDICATORS
# ─────────────────────────────────────────────────────────────────────────────
def build_key_indicators(wb):
    ws = wb.create_sheet("Key Indicators")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, [2, 30, 14, 14, 14, 14, 14, 2])

    for r in range(1, 55):
        for c in range(1, 9):
            ws.cell(r, c).fill = _fill(WHITE)

    _hdr_cell(ws, 1, 2, "Novo Nordisk A/S — Key Indicators & KPI Scorecard", bg=DARK_NAVY, fg=WHITE, span=6)
    ws.row_dimensions[1].height = 32

    years = ["FY2022", "FY2023", "FY2024", "FY2025", "FY2026E"]

    _section_hdr(ws, 3, 2, "KPI SCORECARD — 5 YEAR TREND", span=6)
    _hdr_cell(ws, 4, 2, "KPI", bg=NVO_BLUE, fg=WHITE, align="left")
    _hdr_cell(ws, 4, 3, "FY2022", bg=NVO_BLUE, fg=WHITE)
    _hdr_cell(ws, 4, 4, "FY2023", bg=NVO_BLUE, fg=WHITE)
    _hdr_cell(ws, 4, 5, "FY2024", bg=NVO_BLUE, fg=WHITE)
    _hdr_cell(ws, 4, 6, "FY2025", bg=NVO_BLUE, fg=WHITE)
    _hdr_cell(ws, 4, 7, "FY2026E", bg=NVO_RED, fg=WHITE)
    ws.row_dimensions[4].height = 22

    kpi_data = [
        ("Revenue Growth % (DKK)",       "+25.7%", "+31.3%", "+25.0%", "+6.4%",  "~-5 to -13%", "WATCH"),
        ("Revenue Growth % (CER)",        "~+26%",  "~+32%",  "~+26%",  "~+15%",  "~-5 to -13%","WATCH"),
        ("EBIT Margin %",                "42.3%",  "44.2%",  "44.2%",  "41.3%",  "~38-40%",     "WATCH"),
        ("Gross Margin %",               "83.9%",  "84.6%",  "84.7%",  "81.0%",  "~79-81%",     "WATCH"),
        ("FCF Margin %",                 "37.7%",  "35.8%",  "25.4%",  "19.1%",  "~15-18%",     "WATCH"),
        ("Capex / Revenue %",            "6.9%",   "11.1%",  "16.2%",  "19.5%",  "~20-22%",     "WATCH (peak)"),
        ("Net Debt / EBITDA",            "0.16x",  "0.11x",  "0.59x",  "0.70x",  "~0.8-1.0x",  "NEUTRAL"),
        ("ROE",                          "~72%",   "~88%",   "~83%",   "~62%",   "~50-55%",     "WATCH"),
        ("ROIC",                         "~50%+",  "~55%+",  "~45%",   "~35%",   "~25-30%",     "WATCH"),
        ("Dividend Growth %",            "+17.5%", "+25.5%", "+38.9%", "+17.3%", "~+5-10%",     "POSITIVE"),
        ("Buyback (DKK B equiv.)",       "3.6B",   "4.5B",   "3.1B",   "0.2B",   "~0-0.5B",     "WATCH (suspended)"),
        ("Obesity Care Rev Growth (CER)","n/a",    "+154%",  "+57%",   "+31%",   "+15-20%",     "POSITIVE"),
        ("EPS DKK (reported)",           "12.22",  "18.62",  "22.63",  "23.03",  "~21.37E",     "WATCH"),
    ]
    trend_colours = {"POSITIVE": GREEN, "WATCH": NVO_RED, "NEUTRAL": NVO_BLUE, "WATCH (peak)": NVO_RED, "WATCH (suspended)": NVO_RED}
    for i, (kpi, *vals_trend) in enumerate(kpi_data):
        vals = vals_trend[:5]
        trend = vals_trend[5]
        r = 5 + i
        bg = GREY if i % 2 == 0 else WHITE
        lc = ws.cell(r, 2, kpi)
        lc.font = _font(size=FONT_SIZE)
        lc.fill = _fill(bg)
        lc.alignment = _align("left", "center")
        lc.border = _border()
        for ci, v in enumerate(vals):
            cell_bg = LIGHT_RED if ci == 4 and "WATCH" in trend else bg
            _data_cell(ws, r, 3 + ci, v, bg=cell_bg, align="right")
        trend_c = trend_colours.get(trend, "000000")
        tc = ws.cell(r, 8, "")
        ws.row_dimensions[r].height = 20

    _section_hdr(ws, 19, 2, "MANAGEMENT GUIDANCE & TARGETS", span=6)
    guidance_hdrs = ["Timeframe", "Target/Guidance", "Commentary"]
    for ci, h in enumerate(guidance_hdrs):
        _hdr_cell(ws, 20, 2 + ci, h, bg=NVO_BLUE, fg=WHITE, span=1 if ci < 2 else 4)

    guidance_data = [
        ("FY2026 (Near-term)", "Sales: -5% to -13% CER; Op. profit: similar range", "First-ever guidance for annual sales decline. IRA + Lilly + CagriSema headwinds. Consensus: DKK 291.9B, EPS DKK 21.37."),
        ("FY2027-2028 (Med-term)", "Return to growth; manufacturing reaches efficiency; oral pipeline data", "Capex cycle expected to peak; facilities reach full utilization; amycretin Phase 3 readout is the catalyst."),
        ("FY2030+ (Long-term)", "GLP-1 market leadership; $150-200B global obesity market participation", "Cardio-renal-metabolic expansion (SELECT/FLOW trials positive); rare disease gene therapy; EM market growth."),
        ("Dividend Policy", "Progressive dividend — grew every year since listed", "Management targets growing dividend; Foundation ownership means payout continuity is prioritized. FY2025: DKK ~11.62/share (~$1.76 USD)"),
        ("Capex Guidance", "~DKK 55-70B range expected FY2026; taper 2027+", "Manufacturing investment continuing for GLP-1 capacity. Will begin to taper as Catalent integration completes."),
    ]
    for i, (tf, target, comment) in enumerate(guidance_data):
        r = 21 + i
        bg = GREY if i % 2 == 0 else WHITE
        _data_cell(ws, r, 2, tf, bg=bg, align="left", bold=True)
        _data_cell(ws, r, 3, target, bg=bg, align="left")
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=7)
        cc = ws.cell(r, 4, comment)
        cc.font = _font(size=FONT_SIZE)
        cc.fill = _fill(bg)
        cc.alignment = _align("left", "center", wrap=True)
        cc.border = _border()
        ws.row_dimensions[r].height = 30

    _section_hdr(ws, 27, 2, "INVESTMENT THESIS SUMMARY", span=6)
    thesis_rows = [
        ("1. Near-term Pain",       "FY2026 is a trough year: IRA pricing cuts, Lilly market share, CagriSema failure, CEO transition. First-ever negative guidance."),
        ("2. Valuation Embeds Pain","~12x forward P/E and ~8x EV/EBITDA — 55-60% below historical averages. Most pessimism is priced in at $41."),
        ("3. Secular Demand Intact","GLP-1 global TAM: $150-200B by 2030. <5% penetration. Obesity and T2D are structural, not cyclical."),
        ("4. Amycretin Optionality","Phase 3 results 2026-2027. Phase 1/2 data very promising (~13% weight loss in 12 weeks). Success = massive re-rating."),
        ("5. Foundation = Patient Capital","67% voting control prevents short-term forced decisions. Dividend and R&D investment protected through cycle."),
        ("6. RATING: HOLD",         "Not demanding valuation, but near-term earnings headwinds are real. Monitor amycretin Phase 3 as the primary catalyst. Accumulate below $38 for risk/reward cushion. Bull case $85-100 if oral pipeline succeeds."),
    ]
    for i, (label, desc) in enumerate(thesis_rows):
        r = 28 + i
        bg = LIGHT_TEAL if i == 5 else (GREY if i % 2 == 0 else WHITE)
        bold = (i == 5)
        _data_cell(ws, r, 2, label, bg=bg, align="left", bold=True, colour=NVO_BLUE if i == 5 else "000000")
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)
        dc = ws.cell(r, 3, desc)
        dc.font = _font(bold=bold, size=FONT_SIZE, colour=DARK_NAVY if i == 5 else "000000")
        dc.fill = _fill(bg)
        dc.alignment = _align("left", "center", wrap=True)
        dc.border = _border()
        ws.row_dimensions[r].height = 30

    _note_row(ws, 35, 2, "All financial data in DKK millions from StockAnalysis.com / Novo Nordisk annual reports. USD equivalents at ~6.6 DKK/USD (April 2026). EPS USD verified: $3.49 (FY2025), $3.28 (FY2024), $2.70 (FY2023). Report Date: April 29, 2026.", span=6)


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────
def main():
    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    build_cover(wb)
    build_business(wb)
    build_moat(wb)
    build_income(wb)
    build_balance_sheet(wb)
    build_cashflow(wb)
    build_returns(wb)
    build_management(wb)
    build_risks(wb)
    build_valuation(wb)
    build_sentiment(wb)
    build_key_indicators(wb)

    wb.save(OUTPUT_PATH)
    print(f"Saved: {OUTPUT_PATH}")
    size = os.path.getsize(OUTPUT_PATH)
    print(f"File size: {size:,} bytes ({size / 1024:.1f} KB)")
    print(f"Tabs: {[s.title for s in wb.worksheets]}")


if __name__ == "__main__":
    main()
