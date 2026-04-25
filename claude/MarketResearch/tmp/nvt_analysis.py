"""
NVT (nVent Electric plc) - Comprehensive Financial Analysis
Generated: April 2026
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import SeriesLabel
import os

# ── Colour palette ──────────────────────────────────────────────────────────
NAVY      = "1B2A4A"
BLUE      = "2B5FA6"
LIGHT_BLUE= "D6E4F0"
ORANGE    = "E87722"
GREEN     = "1A7C4A"
LIGHT_GREEN="D4EFDF"
RED       = "C0392B"
LIGHT_RED = "FADBD8"
YELLOW    = "F9E79F"
GREY      = "F2F4F7"
DARK_GREY = "95A5A6"
WHITE     = "FFFFFF"
HEADER_BLUE="2E86AB"

FONT_SIZE = 14

def _font(bold=False, colour=None, size=FONT_SIZE, italic=False):
    return Font(name="Calibri", bold=bold,
                color=colour or "000000", size=size, italic=italic)

def _fill(hex_colour):
    return PatternFill("solid", fgColor=hex_colour)

def _border(style="thin"):
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _hdr_cell(ws, row, col, value, bg=NAVY, fg=WHITE, bold=True,
              align="center", span=1):
    """Write a styled header cell (and optionally merge across span cols)."""
    c = ws.cell(row=row, column=col, value=value)
    c.font = _font(bold=bold, colour=fg, size=FONT_SIZE)
    c.fill = _fill(bg)
    c.alignment = _align(align, "center")
    c.border = _border()
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col + span - 1)
    return c

def _data_cell(ws, row, col, value, bg=WHITE, bold=False,
               align="right", fmt=None, colour=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = _font(bold=bold, colour=colour or "000000", size=FONT_SIZE)
    c.fill = _fill(bg)
    c.alignment = _align(align, "center")
    c.border = _border()
    if fmt:
        c.number_format = fmt
    return c

def _section_hdr(ws, row, col, text, span=10):
    _hdr_cell(ws, row, col, text, bg=BLUE, fg=WHITE, span=span)

def _set_col_widths(ws, widths):
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = w

def _set_row_height(ws, rows, h=22):
    for r in rows:
        ws.row_dimensions[r].height = h

# ────────────────────────────────────────────────────────────────────────────
# 1. COVER
# ────────────────────────────────────────────────────────────────────────────
def build_cover(wb):
    ws = wb.create_sheet("Cover")
    ws.sheet_view.showGridLines = False

    # background strip
    for r in range(1, 50):
        for c in range(1, 16):
            ws.cell(r, c).fill = _fill(WHITE)

    # header band
    ws.merge_cells("A1:O5")
    hdr = ws["A1"]
    hdr.value = "nVent Electric plc (NYSE: NVT)"
    hdr.font = Font(name="Calibri", bold=True, size=28, color=WHITE)
    hdr.fill = _fill(NAVY)
    hdr.alignment = _align("center", "center")

    # Subtitle
    ws.merge_cells("A6:O7")
    sub = ws["A6"]
    sub.value = "Comprehensive Financial & Investment Analysis  |  April 2026"
    sub.font = Font(name="Calibri", bold=False, size=18, color=WHITE)
    sub.fill = _fill(BLUE)
    sub.alignment = _align("center", "center")

    # Tagline
    ws.merge_cells("A8:O9")
    tag = ws["A8"]
    tag.value = "Electrical Connection & Protection Solutions | Data Center & Infrastructure"
    tag.font = Font(name="Calibri", italic=True, size=FONT_SIZE, color=NAVY)
    tag.fill = _fill(LIGHT_BLUE)
    tag.alignment = _align("center", "center")

    def kv_block(row, label, value, label_bg=NAVY, val_bg=GREY):
        ws.merge_cells(start_row=row, start_column=2,
                       end_row=row, end_column=5)
        ws.merge_cells(start_row=row, start_column=6,
                       end_row=row, end_column=9)
        lc = ws.cell(row, 2, label)
        lc.font = _font(bold=True, colour=WHITE, size=FONT_SIZE)
        lc.fill = _fill(label_bg); lc.alignment = _align("right", "center")
        vc = ws.cell(row, 6, value)
        vc.font = _font(bold=False, colour=NAVY, size=FONT_SIZE)
        vc.fill = _fill(val_bg); vc.alignment = _align("left", "center")

    kv_data = [
        (11, "Company",          "nVent Electric plc"),
        (12, "Ticker",           "NYSE: NVT"),
        (13, "Sector",           "Industrials – Electrical Equipment"),
        (14, "Analysis Date",    "April 21, 2026"),
        (15, "Current Price",    "~$133.16"),
        (16, "Market Cap",       "$22.0B"),
        (17, "Enterprise Value", "$23.8B"),
        (18, "52-Wk Range",      "$41.70 – $129.94"),
        (19, "Analyst Consensus","Strong Buy (13 Buy / 1 Hold / 0 Sell)"),
        (20, "Avg Price Target", "$140 – $152 (vs current ~$133)"),
        (21, "Dividend Yield",   "0.62%"),
        (22, "Revenue (FY2025)", "$3.89B (+30% YoY)"),
        (23, "Adj. EPS (FY2025)","$3.35 (+35% YoY)"),
        (24, "FCF (FY2025)",     "$371.9M (GAAP) / ~$561M (Adj.)"),
        (25, "FY2026 Guidance",  "Revenue +15-18%; Adj. EPS $4.00-$4.15"),
    ]
    for r, lbl, val in kv_data:
        kv_block(r, lbl, val)

    # Investment thesis block
    ws.merge_cells("B27:O27")
    th = ws["B27"]
    th.value = "Investment Thesis"
    th.font = _font(bold=True, colour=WHITE, size=FONT_SIZE + 1)
    th.fill = _fill(ORANGE); th.alignment = _align("center", "center")

    thesis_lines = [
        "• Premier data-center infrastructure play: Systems Protection segment ($2.6B, 67% of rev.) grew 42% in 2025, driven by AI hyperscaler enclosure & liquid-cooling demand.",
        "• Two-segment simplicity: Systems Protection (enclosures, liquid cooling) + Electrical Connections (cable management, power). Clear organic growth + M&A roll-up playbook.",
        "• Durable competitive moat: Top-3 global enclosure share alongside Rittal and Schneider; decade-plus liquid-cooling IP; OEM specs locked in at design-phase create switching costs.",
        "• Capital discipline: Free cash flow growing at >30% 3-yr CAGR; dividends + buybacks returning capital while investing in capacity and bolt-on acquisitions.",
        "• KEY RISK: CEO selling stock (~$1.3M across Nov-25/Mar-26); ROIC (9.45%) slightly below WACC (9.6%); stock priced at premium EV/EBITDA 28x; tariff headwinds ~$90M.",
        "• VERDICT: Quality compounder at a full valuation. Bull case ($160+) requires data-center demand sustainability; Bear case (~$90) if AI capex cycle pauses.",
    ]
    for i, line in enumerate(thesis_lines):
        ws.merge_cells(start_row=28+i, start_column=2,
                       end_row=28+i, end_column=14)
        c = ws.cell(28+i, 2, line)
        c.font = _font(size=FONT_SIZE)
        c.fill = _fill(GREY if i % 2 == 0 else WHITE)
        c.alignment = _align("left", "center", wrap=True)
        ws.row_dimensions[28+i].height = 32

    for c in range(1, 16):
        ws.column_dimensions[get_column_letter(c)].width = 14
    for r in [1,2,3,4,5,6,7,8,9,10]:
        ws.row_dimensions[r].height = 24
    for r in range(11, 26):
        ws.row_dimensions[r].height = 22
    ws.row_dimensions[27].height = 24

# ────────────────────────────────────────────────────────────────────────────
# 2. BUSINESS OVERVIEW
# ────────────────────────────────────────────────────────────────────────────
def build_business_overview(wb):
    ws = wb.create_sheet("Business Overview")
    ws.sheet_view.showGridLines = False

    _set_col_widths(ws, [2, 28, 28, 24, 24, 22, 18, 2])
    _section_hdr(ws, 1, 2, "nVent Electric plc – Business Overview", span=6)

    # Company snapshot
    _hdr_cell(ws, 3, 2, "Company Snapshot", bg=BLUE, span=6)
    snap = [
        ("Founded",        "2018 (Spun off from Pentair plc)"),
        ("Headquarters",   "London, UK / Minneapolis, MN (operational HQ)"),
        ("Employees",      "~10,000+ globally across 59 locations"),
        ("Exchange",       "NYSE: NVT"),
        ("Business",       "Designs, manufactures & services electrical connection and protection solutions"),
        ("CEO",            "Beth Wozniak (Chair & CEO since 2018 spin-off)"),
        ("Fiscal Year",    "Calendar year (Dec 31)"),
        ("Key End Markets","Data Centers, Industrial, Commercial & Residential, Infrastructure"),
    ]
    for i, (k, v) in enumerate(snap):
        r = 4 + i
        _data_cell(ws, r, 2, k, bg=LIGHT_BLUE, bold=True, align="left")
        _data_cell(ws, r, 3, v, bg=GREY if i%2==0 else WHITE, align="left")
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)
        ws.row_dimensions[r].height = 22

    # Segments
    _hdr_cell(ws, 14, 2, "Business Segments", bg=BLUE, span=6)
    seg_hdrs = ["Segment", "2025 Revenue", "YoY Growth", "Key Products", "Key Customers / End-Markets"]
    for j, h in enumerate(seg_hdrs):
        _hdr_cell(ws, 15, 2+j, h, bg=NAVY, span=1)

    segs = [
        ("Systems Protection\n(~67% of Revenue)",
         "$2,600M", "+42% (+17% organic)",
         "Electrical enclosures, liquid cooling, thermal mgmt, control buildings, data center infrastructure",
         "Hyperscalers (AWS, Azure, Google), Colocation providers, AI chip OEMs (NVIDIA), Siemens"),
        ("Electrical Connections\n(~33% of Revenue)",
         "$1,293M", "+10% (+10% organic)",
         "Cable management, power connections, switchgear accessories, bus systems, tools & test instruments",
         "Industrial OEMs, utilities, commercial construction, infrastructure EPC contractors"),
    ]
    for i, row_data in enumerate(segs):
        r = 16 + i
        ws.row_dimensions[r].height = 55
        for j, val in enumerate(row_data):
            bg = LIGHT_BLUE if i == 0 else LIGHT_GREEN
            c = _data_cell(ws, r, 2+j, val, bg=bg, align="left")
            c.alignment = _align("left", "center", wrap=True)

    # Revenue Breakdown by Geography (estimated)
    _hdr_cell(ws, 20, 2, "Revenue by Geography (Estimated FY2025)", bg=BLUE, span=6)
    geo_hdrs = ["Region", "Est. Revenue", "Est. Share", "Notes"]
    for j, h in enumerate(geo_hdrs):
        _hdr_cell(ws, 21, 2+j, h, bg=NAVY)
    geo_data = [
        ("Americas",        "$2,335M",  "60%", "Largest market; hyperscaler buildout concentrated in US"),
        ("Europe / EMEA",   "$1,167M",  "30%", "Strong industrial base; Germany, UK, Nordics key markets"),
        ("Asia Pacific",    "$391M",    "10%", "Growing data center deployment; manufacturing presence"),
    ]
    for i, row in enumerate(geo_data):
        r = 22 + i
        for j, val in enumerate(row):
            _data_cell(ws, r, 2+j, val, bg=GREY if i%2==0 else WHITE, align="left")
        ws.row_dimensions[r].height = 22

    # Products & Value Propositions
    _hdr_cell(ws, 27, 2, "Products, Value Propositions & Key Clients", bg=BLUE, span=6)
    _hdr_cell(ws, 28, 2, "Product Category", bg=NAVY)
    _hdr_cell(ws, 28, 3, "Value Proposition", bg=NAVY)
    _hdr_cell(ws, 28, 4, "Key Clients / OEMs", bg=NAVY)
    _hdr_cell(ws, 28, 5, "Competitive Edge", bg=NAVY)
    _hdr_cell(ws, 28, 6, "Est. Margin Profile", bg=NAVY)

    prods = [
        ("Electrical Enclosures",
         "Protect sensitive electronics in harsh environments; IEC/UL certified",
         "Hyperscalers, Industrial OEMs, Utilities",
         "Top-3 global share; UL/IEC certifications; spec-in design wins",
         "High (35-40% gross)"),
        ("Liquid Cooling Systems",
         "Next-gen thermal management for AI GPUs; Direct Liquid Cooling (DLC)",
         "NVIDIA, AMD, Cloud providers, Tier-1 colocation",
         "Decade-plus IP; 1GW+ deployed; modular service-friendly design",
         "High & expanding (35%+ gross)"),
        ("Cable Management",
         "Organize and protect cable infrastructure; reduce installation time",
         "Commercial contractors, Data centers, Industrial plants",
         "Broad product range; specification pull-through",
         "Moderate (30-35% gross)"),
        ("Power Connections",
         "Ensure reliable electrical connections in high-demand settings",
         "Utilities, OEM integrators, Infrastructure EPC",
         "Long product qualifications; OEM integration depth",
         "Moderate (28-33% gross)"),
        ("Bus Systems / Switchgear",
         "Efficient power distribution in large facilities",
         "Data centers, Heavy industry, Utilities",
         "Integration with enclosure ecosystem",
         "Moderate-High (32-38% gross)"),
    ]
    for i, p in enumerate(prods):
        r = 29 + i
        for j, val in enumerate(p):
            c = _data_cell(ws, r, 2+j, val, bg=GREY if i%2==0 else WHITE, align="left")
            c.alignment = _align("left", "center", wrap=True)
        ws.row_dimensions[r].height = 45

    # Buying process / seasonality
    _hdr_cell(ws, 36, 2, "Buying Process & Seasonality", bg=BLUE, span=6)
    buying = [
        ("How Customers Buy",
         "Two-step channel model: Sell through distributors/rep firms + direct to large OEMs and hyperscalers. "
         "Enterprise data center deals involve multi-quarter design-in cycles with spec engineers. "
         "Industrial tends to be project-based via EPC contractors. "
         "Electrical Connections products are often specified at design phase – creating sticky switching costs."),
        ("Seasonality",
         "Mild seasonality: Q4 typically stronger as customers deploy remaining capex budgets. "
         "Data center segment is relatively acyclical due to long-term hyperscaler commitments. "
         "Industrial segment shows more cyclicality tied to broad industrial capex. "
         "Backlog (now extending into 2026) provides revenue visibility and dampens quarter-to-quarter swings."),
    ]
    for i, (k, v) in enumerate(buying):
        r = 37 + i
        c1 = _data_cell(ws, r, 2, k, bg=LIGHT_BLUE, bold=True, align="left")
        c2 = _data_cell(ws, r, 3, v, bg=GREY if i%2==0 else WHITE, align="left")
        c2.alignment = _align("left", "center", wrap=True)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)
        ws.row_dimensions[r].height = 65

    for r in [1,2,3,14,20,27,36]:
        ws.row_dimensions[r].height = 22

# ────────────────────────────────────────────────────────────────────────────
# 3. MOAT
# ────────────────────────────────────────────────────────────────────────────
def build_moat(wb):
    ws = wb.create_sheet("Moat")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, [2, 22, 38, 12, 10, 18, 2])
    _section_hdr(ws, 1, 2, "nVent Electric – Competitive Moat Analysis", span=5)

    # Moat rating
    ws.merge_cells("B3:F3")
    mr = ws["B3"]
    mr.value = "Overall Moat Rating:  NARROW-TO-WIDE  |  Score: 7.2 / 10"
    mr.font = Font(name="Calibri", bold=True, size=FONT_SIZE+1, color=WHITE)
    mr.fill = _fill(GREEN); mr.alignment = _align("center", "center")
    ws.row_dimensions[3].height = 26

    headers = ["Moat Source", "Evidence & Detail", "Strength\n(1-10)", "Durability\n(yrs)", "Key Threat"]
    for j, h in enumerate(headers):
        _hdr_cell(ws, 4, 2+j, h, bg=NAVY)

    moats = [
        ("Switching Costs",
         "Enclosures and liquid cooling systems are spec'd into data center rack designs 12-24 months before deployment. "
         "Once an OEM or hyperscaler standardises on NVT form factors, replacing mid-lifecycle is prohibitively costly "
         "(re-certification, downtime risk, re-engineering). Electrical Connections products (cable trays, connectors) "
         "are similarly embedded in plant wiring infrastructure.",
         "8", "10+", "New open-standard designs reducing OEM lock-in"),
        ("Customer Relationships & OEM Partnerships",
         "Deep OEM integration with NVIDIA, Siemens, and major hyperscalers (AWS, Azure, GCP). "
         "Co-development agreements on next-gen liquid cooling tied to H100/B100/Blackwell GPU thermal specs. "
         "Long-standing distributor network across 100+ countries reduces competitive entry.",
         "8", "8", "Vertiv / Rittal winning OEM design contracts"),
        ("Market Position & Scale",
         "Top-3 global enclosure share alongside Rittal (Germany) and Schneider Electric (France). "
         "Global manufacturing footprint in 59 locations enables local delivery and service – "
         "a major advantage for hyperscalers needing just-in-time delivery for massive builds.",
         "7", "7", "Rittal expanding North America capacity"),
        ("Proprietary IP in Liquid Cooling",
         "10+ years in liquid cooling technology; Direct Liquid Cooling (DLC) and immersion-ready platforms. "
         "Over 1 GW of cooling capacity deployed. Modular, service-friendly architecture is a key differentiator "
         "vs. custom-engineered solutions from competitors.",
         "7", "8", "Vertiv, Airedale, CoolIT Systems IP catch-up"),
        ("Regulatory & Certification Barriers",
         "UL, IEC, ATEX, and other global certifications on enclosure product lines take years to obtain "
         "and are required by hyperscalers and industrial customers. New entrants face multi-year certification lead times.",
         "6", "7", "Regulators streamlining cert pathways"),
        ("Installed Base & Service Revenue",
         "Large installed base of enclosures and cooling systems generates recurring maintenance and upgrade cycles. "
         "Service teams embedded at major data center campuses create ongoing touchpoints and upsell opportunities.",
         "6", "8", "In-house service arms of hyperscalers expanding"),
        ("Product Breadth / One-Stop Shop",
         "NVT offers the broadest electrical protection portfolio: from cable trays to full liquid cooling racks. "
         "Hyperscalers prefer single-vendor accountability for electrical infrastructure — NVT benefits from this bundling premium.",
         "7", "6", "Specialty point-solution competitors on specific categories"),
    ]
    for i, (m, ev, s, d, t) in enumerate(moats):
        r = 5 + i
        c1 = _data_cell(ws, r, 2, m, bg=LIGHT_BLUE, bold=True, align="left")
        c2 = _data_cell(ws, r, 3, ev, bg=GREY if i%2==0 else WHITE, align="left")
        c2.alignment = _align("left", "center", wrap=True)
        score = int(s)
        score_bg = GREEN if score >= 8 else (YELLOW if score >= 6 else RED)
        c3 = _data_cell(ws, r, 4, int(s), bg=score_bg, bold=True, align="center",
                        colour=WHITE if score >= 8 else NAVY)
        c4 = _data_cell(ws, r, 5, d, bg=GREY if i%2==0 else WHITE, align="center")
        c5 = _data_cell(ws, r, 6, t, bg=LIGHT_RED, align="left")
        c5.alignment = _align("left", "center", wrap=True)
        ws.row_dimensions[r].height = 75

    # Competitive comparison
    _hdr_cell(ws, 14, 2, "Competitive Landscape – Key Rivals", bg=BLUE, span=5)
    comp_hdrs = ["Competitor", "Strengths", "Weaknesses vs NVT", "Threat Level"]
    for j, h in enumerate(comp_hdrs):
        _hdr_cell(ws, 15, 2+j, h, bg=NAVY)
    ws.merge_cells(start_row=15, start_column=5, end_row=15, end_column=6)

    comps = [
        ("Vertiv (VRT)",
         "Strong PDU/UPS/cooling portfolio; aggressive AI data center positioning; fast-growing",
         "Weaker enclosure breadth; less industrial diversification",
         "HIGH"),
        ("Rittal (Private)",
         "World's largest enclosure manufacturer; strong European presence; modular TS 8 system",
         "Private company – less transparent; US market presence weaker",
         "HIGH"),
        ("Schneider Electric (SU.FP)",
         "Huge global scale; EcoStruxure platform; data center DCIM leadership",
         "More software-focused; less specialized enclosure/connector depth",
         "MEDIUM"),
        ("Eaton (ETN)",
         "Strong power management and UPS; industrial and utility relationships",
         "Focuses more on power quality vs protection; less liquid cooling IP",
         "MEDIUM"),
        ("Hubbell (HUBB)",
         "Strong electrical wiring devices and commercial construction",
         "No meaningful data center enclosure or liquid cooling play",
         "LOW-MED"),
    ]
    for i, (co, st, wk, thr) in enumerate(comps):
        r = 16 + i
        thr_col = RED if thr == "HIGH" else (YELLOW if "MED" in thr else GREEN)
        _data_cell(ws, r, 2, co, bg=LIGHT_BLUE, bold=True, align="left")
        c2 = _data_cell(ws, r, 3, st, bg=GREY if i%2==0 else WHITE, align="left")
        c2.alignment = _align("left", "center", wrap=True)
        c3 = _data_cell(ws, r, 4, wk, bg=GREY if i%2==0 else WHITE, align="left")
        c3.alignment = _align("left", "center", wrap=True)
        ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=6)
        _data_cell(ws, r, 5, thr, bg=thr_col, bold=True, align="center",
                   colour=WHITE if thr == "HIGH" else NAVY)
        ws.row_dimensions[r].height = 45

    for r in [1, 14]:
        ws.row_dimensions[r].height = 22
    ws.row_dimensions[4].height = 30

# ────────────────────────────────────────────────────────────────────────────
# 4. INCOME STATEMENTS
# ────────────────────────────────────────────────────────────────────────────
def build_income_statement(wb):
    ws = wb.create_sheet("Income Statements")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, [2, 30, 14, 14, 14, 14, 14, 14, 2])
    _section_hdr(ws, 1, 2, "nVent Electric plc – Income Statement (USD Millions)", span=7)

    years = ["FY2021", "FY2022", "FY2023", "FY2024", "FY2025", "FY2026E"]
    _hdr_cell(ws, 2, 2, "Line Item", bg=NAVY)
    for j, y in enumerate(years):
        _hdr_cell(ws, 2, 3+j, y, bg=NAVY)

    # Data
    is_data = [
        # (Label, 2021, 2022, 2023, 2024, 2025, 2026E, format, bold, bg)
        ("Revenue",                     2462, 2295, 2669, 3006, 3893, 4535, "#,##0", True,  LIGHT_BLUE),
        ("  YoY Growth",                None, -6.8, 16.3, 12.6, 29.5, 16.5, '0.0"%"', False, GREY),
        ("Cost of Revenue",             1520, 1472, 1594, 1797, 2424, 2814, "#,##0", False, WHITE),
        ("Gross Profit",                 942,  823, 1075, 1209, 1469, 1721, "#,##0", True,  LIGHT_BLUE),
        ("  Gross Margin %",            38.3, 35.9, 40.3, 40.2, 37.7, 37.9, '0.0"%"', False, GREY),
        ("SG&A + R&D",                   387,  322,  376,  422,  522,  607, "#,##0", False, WHITE),
        ("Depreciation & Amortization",  200,  190,  237,  260,  250,  270, "#,##0", False, WHITE),
        ("Operating Income (GAAP)",      355,  309,  463,  527,  617,  774, "#,##0", True,  LIGHT_BLUE),
        ("  Operating Margin %",        14.4, 13.5, 17.3, 17.5, 15.8, 17.1, '0.0"%"', False, GREY),
        ("Adjusted Operating Income",    410,  390,  560,  652,  786,  960, "#,##0", False, LIGHT_GREEN),
        ("  Adj. Operating Margin %",   16.7, 17.0, 21.0, 21.7, 20.2, 21.2, '0.0"%"', False, GREY),
        ("Interest Expense",             -57,  -60,  -97, -120, -110, -100, "#,##0", False, WHITE),
        ("Other Income / (Expense)",      -3,   10,   15,   10,    5,    5, "#,##0", False, WHITE),
        ("Pre-Tax Income",               295,  259,  381,  417,  512,  679, "#,##0", False, WHITE),
        ("Income Tax Expense",           -22,   41,  186,   85,  -198,  136, "#,##0", False, WHITE),
        ("Net Income (GAAP)",            273,  400,  567,  332,  710,  543, "#,##0", True,  LIGHT_BLUE),
        ("  Net Margin %",              11.1, 17.4, 21.2,  11.0, 18.2, 12.0, '0.0"%"', False, GREY),
        ("Adjusted EPS",                1.61, 2.38, 3.37, 1.97, 4.31, 4.08, "$#,##0.00", True, LIGHT_GREEN),
        ("  Adj. EPS Growth %",         None, 47.8, 41.6,-41.5,118.8, -5.3, '0.0"%"', False, GREY),
        ("Shares Outstanding (M)",      170,  168,  168,  168,  165,  162,  "#,##0", False, WHITE),
        ("EBITDA (approx.)",            555,  499,  700,  787,  867,1044,  "#,##0", True,  LIGHT_BLUE),
        ("  EBITDA Margin %",           22.5, 21.7, 26.2, 26.2, 22.3, 23.0, '0.0"%"', False, GREY),
    ]

    for i, row in enumerate(is_data):
        r = 3 + i
        label, *vals, fmt, bold, bg = row
        lc = _data_cell(ws, r, 2, label, bg=bg, bold=bold, align="left")
        for j, v in enumerate(vals[:6]):
            if v is None:
                _data_cell(ws, r, 3+j, "N/A", bg=bg, align="center")
            else:
                c = _data_cell(ws, r, 3+j, v if "%" not in fmt else v/100 if isinstance(v, (int, float)) else v,
                               bg=bg, bold=bold, align="right", fmt=fmt if "%" not in fmt else "0.0%")
                if "%" in fmt and isinstance(v, (int, float)):
                    c.value = v / 100
                    c.number_format = "0.0%"
                else:
                    c.value = v
                    c.number_format = fmt
        ws.row_dimensions[r].height = 22

    # Segment breakdown
    r_seg = 3 + len(is_data) + 1
    _hdr_cell(ws, r_seg, 2, "Segment Revenue Breakdown (USD Millions)", bg=BLUE, span=7)
    _hdr_cell(ws, r_seg+1, 2, "Segment", bg=NAVY)
    for j, y in enumerate(years):
        _hdr_cell(ws, r_seg+1, 3+j, y, bg=NAVY)

    seg_rows = [
        ("Systems Protection",    1542, 1380, 1730, 1970, 2600, 3050, LIGHT_BLUE),
        ("  % of Total",          62.6, 60.1, 64.8, 65.5, 66.8, 67.3, GREY),
        ("Electrical Connections", 920,  915,  939, 1036, 1293, 1485, LIGHT_GREEN),
        ("  % of Total",          37.4, 39.9, 35.2, 34.5, 33.2, 32.7, GREY),
        ("Total Revenue",         2462, 2295, 2669, 3006, 3893, 4535, LIGHT_BLUE),
    ]
    for i, (lbl, *vals, bg) in enumerate(seg_rows):
        r = r_seg + 2 + i
        _data_cell(ws, r, 2, lbl, bg=bg, bold="Total" in lbl or "%" not in lbl, align="left")
        for j, v in enumerate(vals[:6]):
            is_pct = "%" in lbl
            c = ws.cell(row=r, column=3+j, value=v/100 if is_pct else v)
            c.font = _font(size=FONT_SIZE)
            c.fill = _fill(bg)
            c.alignment = _align("right", "center")
            c.border = _border()
            c.number_format = "0.0%" if is_pct else "#,##0"
        ws.row_dimensions[r].height = 22

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[r_seg].height = 22
    ws.row_dimensions[r_seg+1].height = 22

    # Note row
    note_r = r_seg + 2 + len(seg_rows) + 1
    ws.merge_cells(start_row=note_r, start_column=2, end_row=note_r, end_column=8)
    nc = ws.cell(note_r, 2,
        "Note: FY2026E based on company guidance midpoint (+16.5% rev growth; adj. EPS $4.08 midpoint). "
        "2024 net income depressed by non-cash charges. 2022 revenue decline reflects Thermal Mgmt divestiture.")
    nc.font = _font(italic=True, size=FONT_SIZE - 1)
    nc.fill = _fill(YELLOW)
    nc.alignment = _align("left", "center", wrap=True)
    ws.row_dimensions[note_r].height = 40

# ────────────────────────────────────────────────────────────────────────────
# 5. BALANCE SHEET
# ────────────────────────────────────────────────────────────────────────────
def build_balance_sheet(wb):
    ws = wb.create_sheet("Balance Sheet")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, [2, 30, 14, 14, 14, 14, 14, 2])
    _section_hdr(ws, 1, 2, "nVent Electric plc – Balance Sheet (USD Millions)", span=6)

    years = ["FY2021", "FY2022", "FY2023", "FY2024", "FY2025"]
    _hdr_cell(ws, 2, 2, "Line Item", bg=NAVY)
    for j, y in enumerate(years):
        _hdr_cell(ws, 2, 3+j, y, bg=NAVY)

    bs_data = [
        # ASSETS
        ("ASSETS", None, None, None, None, None, NAVY, True),
        ("Cash & Equivalents",          49.5, 297.5, 179.6, 131.2, 237.5, GREY, False),
        ("Accounts Receivable",          380,   420,   500,   570,   720, WHITE, False),
        ("Inventories",                  300,   330,   390,   440,   560, GREY, False),
        ("Other Current Assets",         100,   110,   130,   150,   180, WHITE, False),
        ("Total Current Assets",         830,  1158,  1200,  1291,  1698, LIGHT_BLUE, True),
        ("Goodwill",                    2187,  2178,  1858,  2222,  2678, GREY, False),
        ("Intangible Assets",            900,   820,   750,   750,   720, WHITE, False),
        ("PP&E, Net",                    520,   530,   620,   730,   820, GREY, False),
        ("Other Long-Term Assets",       237,   216,  1734,  1742,  1936, WHITE, False),
        ("Total Assets",                4674,  4902,  6162,  6735,  6852, LIGHT_BLUE, True),
        # LIABILITIES
        ("LIABILITIES & EQUITY", None, None, None, None, None, NAVY, True),
        ("Accounts Payable",             280,   300,   360,   420,   520, GREY, False),
        ("Short-Term Debt",               50,    50,    80,   200,   100, WHITE, False),
        ("Accrued Liabilities",          250,   270,   310,   360,   420, GREY, False),
        ("Other Current Liabilities",    120,   130,   160,   200,   250, WHITE, False),
        ("Total Current Liabilities",    700,   750,   910,  1180,  1290, LIGHT_RED, True),
        ("Long-Term Debt",               949,  1033,  1701,  1955,  1460, GREY, False),
        ("Deferred Tax Liabilities",     200,   180,   160,   140,   120, WHITE, False),
        ("Other Long-Term Liabilities",  329,   208,   249,   222,   252, GREY, False),
        ("Total Liabilities",           2178,  2171,  3020,  3497,  3122, LIGHT_RED, True),
        # EQUITY
        ("Common Stock & APIC",         2700,  2720,  2740,  2750,  2760, GREY, False),
        ("Retained Earnings / (Deficit)", -80,   112,   542,   298,  1130, WHITE, False),
        ("Other Comprehensive Income",  -124,  -100,  -140,  -190,  -160, GREY, False),
        ("Total Shareholders Equity",   2496,  2732,  3142,  3238,  3730, LIGHT_GREEN, True),
        ("Total Liabilities + Equity",  4674,  4902,  6162,  6735,  6852, LIGHT_BLUE, True),
        # KEY RATIOS
        ("KEY RATIOS", None, None, None, None, None, NAVY, True),
        ("Net Debt",                     950,   786,  1601,  2024,  1323, GREY, False),
        ("Net Debt / EBITDA",            1.7,   1.6,   2.3,   2.6,   1.5, WHITE, False),
        ("Debt / Equity",               40.0,  39.7,  56.7,  66.6,  41.8, GREY, False),
        ("Current Ratio",               1.19,  1.54,  1.32,  1.09,  1.32, LIGHT_GREEN, True),
        ("Goodwill / Total Assets %",   46.8,  44.4,  30.1,  33.0,  39.1, WHITE, False),
        ("Book Value per Share ($)",    14.68, 16.26, 18.70, 19.27, 23.07, LIGHT_BLUE, True),
    ]

    for i, row in enumerate(bs_data):
        r = 3 + i
        label = row[0]
        vals = row[1:6]
        bg = row[6]
        bold = row[7]

        is_section = vals[0] is None
        if is_section:
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
            c = ws.cell(r, 2, label)
            c.font = _font(bold=True, colour=WHITE, size=FONT_SIZE)
            c.fill = _fill(NAVY)
            c.alignment = _align("left", "center")
            ws.row_dimensions[r].height = 24
            continue

        lc = _data_cell(ws, r, 2, label, bg=bg, bold=bold, align="left")
        for j, v in enumerate(vals):
            if v is None:
                _data_cell(ws, r, 3+j, "—", bg=bg, align="center")
            else:
                c = _data_cell(ws, r, 3+j, v, bg=bg, bold=bold, align="right")
                # choose format
                if "%" in label:
                    c.number_format = "0.0%"
                    c.value = v / 100
                elif "/" in label or "Ratio" in label or "per Share" in label or label in ("Net Debt / EBITDA",):
                    c.number_format = "0.00"
                else:
                    c.number_format = "#,##0.0"
        ws.row_dimensions[r].height = 22

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 22

    note_r = 3 + len(bs_data) + 1
    ws.merge_cells(start_row=note_r, start_column=2, end_row=note_r, end_column=7)
    nc = ws.cell(note_r, 2,
        "Note: Some line items estimated/reconciled from public filings. Goodwill grew in 2025 due to Electrical Products Group acquisition. "
        "Net Debt = Total Debt - Cash. 2023 long-term asset jump reflects Aquadesign & Raychem RPG acquisition activity.")
    nc.font = _font(italic=True, size=FONT_SIZE-1)
    nc.fill = _fill(YELLOW)
    nc.alignment = _align("left", "center", wrap=True)
    ws.row_dimensions[note_r].height = 40

# ────────────────────────────────────────────────────────────────────────────
# 6. CASH FLOW ANALYSIS
# ────────────────────────────────────────────────────────────────────────────
def build_cash_flow(wb):
    ws = wb.create_sheet("Cash Flow Analysis")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, [2, 32, 14, 14, 14, 14, 14, 14, 2])
    _section_hdr(ws, 1, 2, "nVent Electric plc – Cash Flow Analysis (USD Millions)", span=7)

    years = ["FY2021", "FY2022", "FY2023", "FY2024", "FY2025", "FY2026E"]
    _hdr_cell(ws, 2, 2, "Line Item", bg=NAVY)
    for j, y in enumerate(years):
        _hdr_cell(ws, 2, 3+j, y, bg=NAVY)

    cf_data = [
        ("OPERATING ACTIVITIES", None,None,None,None,None,None, NAVY, True),
        ("Net Income (GAAP)",      273, 400, 567, 332, 710, 543, WHITE, False),
        ("D&A Add-back",           200, 190, 237, 260, 250, 270, GREY,  False),
        ("Working Capital Changes", -60, -90, -140, 90, -380, -100, WHITE, False),
        ("Other Adjustments",       -40,  -106, -136,  161, -115,  -20, GREY,  False),
        ("Operating Cash Flow",     373, 394, 528, 643, 465, 693, LIGHT_BLUE, True),
        ("INVESTING ACTIVITIES", None,None,None,None,None,None, NAVY, True),
        ("Capital Expenditures",    -40,  -41,  -66,  -74,  -93, -105, GREY, False),
        ("Acquisitions (Net)",      -50, -500, -1100, -700, -600, -200, WHITE, False),
        ("Proceeds from Divestitures", 0, 1730, 0, 0, 0, 0,         GREY, False),
        ("Other Investing",           -5,  -10,  -20,  -15,  -20,  -10, WHITE, False),
        ("Investing Cash Flow",    -95, 1179, -1186, -789, -713, -315, LIGHT_RED, True),
        ("FINANCING ACTIVITIES", None,None,None,None,None,None, NAVY, True),
        ("Debt Issuance / (Repayment)", -50, -1300, 700, 380, -595, -200, GREY, False),
        ("Dividends Paid",          -118, -117, -117, -127, -130, -135, WHITE, False),
        ("Share Repurchases",       -112,  -66,  -61, -100, -253, -200, GREY, False),
        ("Other Financing",          -20,  -10,  -30,  -50,  -30,  -20, WHITE, False),
        ("Financing Cash Flow",     -300,-1493,  492,  103, -1008, -555, LIGHT_RED, True),
        ("Net Change in Cash",       -22, 80, -166,  -43,  -256,  -177, LIGHT_BLUE, True),
        ("FREE CASH FLOW", None,None,None,None,None,None, NAVY, True),
        ("FCF (Op CF - Capex)",      333, 354, 462, 569, 372, 588, LIGHT_GREEN, True),
        ("  FCF Margin %",          13.5, 15.4, 17.3, 18.9,  9.6, 13.0, GREY, False),
        ("  FCF Conversion (FCF/NI)",122, 88.5, 81.5,171.4, 52.4,108.3, WHITE, False),
        ("Adj FCF (mgmt guide)",     360, 380, 490, 590, 561, 650, LIGHT_GREEN, True),
        ("  Adj. FCF Margin %",     14.6, 16.6, 18.4, 19.6, 14.4, 14.3, GREY, False),
        ("FCF per Share ($)",        1.96, 2.11, 2.75, 3.39, 2.26, 3.63, LIGHT_BLUE, True),
        ("Dividend Payout Ratio",   43.2, 29.2, 20.6, 38.2, 18.4, 24.9, WHITE, False),
        ("Capex % of Revenue",       1.6,  1.8,  2.5,  2.5,  2.4,  2.3, GREY, False),
    ]

    for i, row in enumerate(cf_data):
        r = 3 + i
        label = row[0]
        vals = row[1:7]
        bg = row[7]
        bold = row[8]
        is_section = vals[0] is None
        if is_section:
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
            c = ws.cell(r, 2, label)
            c.font = _font(bold=True, colour=WHITE, size=FONT_SIZE)
            c.fill = _fill(NAVY)
            c.alignment = _align("left", "center")
            ws.row_dimensions[r].height = 24
            continue
        _data_cell(ws, r, 2, label, bg=bg, bold=bold, align="left")
        for j, v in enumerate(vals):
            is_pct = "%" in label or "Payout" in label or "Capex %" in label
            is_mult = "Conversion" in label
            c = ws.cell(r, 3+j, v/100 if is_pct else v)
            c.font = _font(size=FONT_SIZE, bold=bold)
            c.fill = _fill(bg)
            c.alignment = _align("right", "center")
            c.border = _border()
            if is_pct:
                c.number_format = "0.0%"
            elif is_mult:
                c.number_format = "0.0x"
                c.value = v / 100
            elif "per Share" in label:
                c.number_format = "$#,##0.00"
            else:
                c.number_format = "#,##0.0"
        ws.row_dimensions[r].height = 22

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 22

    note_r = 3 + len(cf_data) + 1
    ws.merge_cells(start_row=note_r, start_column=2, end_row=note_r, end_column=8)
    nc = ws.cell(note_r, 2,
        "Note: FY2025 GAAP FCF of $371.9M was impacted by working capital build to support rapid revenue growth. "
        "Adjusted FCF of ~$561M per management includes normalised working capital. "
        "FY2026E assumes improving WC dynamics as growth rate normalises to 15-18%.")
    nc.font = _font(italic=True, size=FONT_SIZE-1)
    nc.fill = _fill(YELLOW)
    nc.alignment = _align("left", "center", wrap=True)
    ws.row_dimensions[note_r].height = 45

# ────────────────────────────────────────────────────────────────────────────
# 7. RETURN ON CAPITAL
# ────────────────────────────────────────────────────────────────────────────
def build_return_on_capital(wb):
    ws = wb.create_sheet("Return on Capital")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, [2, 32, 14, 14, 14, 14, 14, 2])
    _section_hdr(ws, 1, 2, "nVent Electric plc – Return on Capital Analysis", span=6)

    years = ["FY2021", "FY2022", "FY2023", "FY2024", "FY2025"]
    _hdr_cell(ws, 2, 2, "Metric", bg=NAVY)
    for j, y in enumerate(years):
        _hdr_cell(ws, 2, 3+j, y, bg=NAVY)

    roc_data = [
        ("RETURN METRICS", None,None,None,None,None, NAVY, True),
        ("Return on Equity (ROE)",       10.9, 14.6, 18.0,  10.2,  12.3, WHITE, False),
        ("Return on Assets (ROA)",        5.8,  8.2,  9.2,   4.9,   5.8, GREY,  False),
        ("Return on Invested Capital",    7.2,  8.5, 11.0,   8.2,   9.5, LIGHT_BLUE, True),
        ("WACC (estimated)",              9.5,  9.5,  9.6,   9.6,   9.6, WHITE, False),
        ("Economic Spread (ROIC-WACC)",  -2.3, -1.0,  1.4,  -1.4,  -0.1, LIGHT_RED, True),
        ("INCREMENTAL CAPITAL RETURNS", None,None,None,None,None, NAVY, True),
        ("Incremental Revenue",           None, -167, 374,  337,   887, GREY, False),
        ("Incremental NOPAT (est.)",      None,  -42,  75,   67,   177, WHITE, False),
        ("Incremental Invested Capital",  None,  243, 500,  640,   800, GREY, False),
        ("Return on Incremental Capital", None, -17.3, 15.0, 10.5, 22.1, LIGHT_GREEN, True),
        ("CAPITAL DEPLOYMENT", None,None,None,None,None, NAVY, True),
        ("Organic Capex",                  40,   41,   66,   74,    93, GREY, False),
        ("M&A Spend",                      50,  500, 1100,  700,   600, WHITE, False),
        ("Dividends",                     118,  117,  117,  127,   130, GREY, False),
        ("Buybacks",                      112,   66,   61,  100,   253, WHITE, False),
        ("Total Capital Deployed",         320,  724, 1344, 1001, 1076, LIGHT_BLUE, True),
        ("FCF Coverage of Capital Deploy", 1.04, 0.49, 0.34, 0.57, 0.35, LIGHT_RED, True),
        ("ASSET EFFICIENCY", None,None,None,None,None, NAVY, True),
        ("Asset Turnover (Rev / Assets)", 0.53, 0.47, 0.43, 0.45, 0.57, GREY, False),
        ("Inventory Turns",               8.2,  6.9,  6.8,  6.8,  7.0, WHITE, False),
        ("AR Days Outstanding",            56,   67,   68,   69,   67, GREY, False),
        ("Goodwill-to-Revenue",           0.89, 0.95, 0.70, 0.74, 0.69, WHITE, False),
    ]

    pct_rows = {"Return on Equity (ROE)", "Return on Assets (ROA)",
                "Return on Invested Capital", "WACC (estimated)", "Economic Spread (ROIC-WACC)"}
    pct2_rows = {"Return on Incremental Capital"}
    mult_rows = {"FCF Coverage of Capital Deploy", "Asset Turnover (Rev / Assets)",
                 "Goodwill-to-Revenue", "Inventory Turns"}
    day_rows = {"AR Days Outstanding"}
    dollar_rows = {"Incremental Revenue", "Incremental NOPAT (est.)",
                   "Incremental Invested Capital", "Organic Capex", "M&A Spend",
                   "Dividends", "Buybacks", "Total Capital Deployed"}

    for i, row in enumerate(roc_data):
        r = 3 + i
        label = row[0]; vals = row[1:6]; bg = row[6]; bold = row[7]
        is_section = vals[0] is None
        if is_section:
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
            c = ws.cell(r, 2, label)
            c.font = _font(bold=True, colour=WHITE, size=FONT_SIZE)
            c.fill = _fill(NAVY)
            c.alignment = _align("left", "center")
            ws.row_dimensions[r].height = 24
            continue
        _data_cell(ws, r, 2, label, bg=bg, bold=bold, align="left")
        for j, v in enumerate(vals):
            if v is None:
                c = ws.cell(r, 3+j, "—")
                c.font = _font(size=FONT_SIZE); c.fill = _fill(bg)
                c.alignment = _align("center"); c.border = _border()
                continue
            if label in pct_rows or label in pct2_rows:
                cv = v / 100
                fmt = "0.0%"
                col = GREEN if cv > 0 else RED
                c = ws.cell(r, 3+j, cv)
                c.font = _font(size=FONT_SIZE, bold=bold, colour=WHITE if bold else col)
                c.fill = _fill(bg)
                c.number_format = fmt
                c.alignment = _align("right","center")
                c.border = _border()
            elif label in mult_rows:
                c = ws.cell(r, 3+j, v)
                c.font = _font(size=FONT_SIZE, bold=bold)
                c.fill = _fill(bg)
                c.number_format = "0.00x"
                c.alignment = _align("right","center")
                c.border = _border()
            elif label in day_rows:
                c = ws.cell(r, 3+j, v)
                c.font = _font(size=FONT_SIZE); c.fill = _fill(bg)
                c.number_format = "0"; c.alignment = _align("right","center")
                c.border = _border()
            else:
                c = _data_cell(ws, r, 3+j, v, bg=bg, bold=bold, align="right",
                               fmt="#,##0.0")
        ws.row_dimensions[r].height = 22

    # Owner-operator commentary
    note_r = 3 + len(roc_data) + 1
    _hdr_cell(ws, note_r, 2, "Owner-Operator Lens on Capital Returns", bg=BLUE, span=6)
    commentary = [
        ("ROIC vs WACC",
         "NVT's ROIC (9.45%) is essentially at parity with WACC (~9.6%). This means the company is currently not "
         "generating meaningful economic value above its cost of capital. However, if the data-center growth "
         "momentum continues to drive incremental ROIC (22% in FY2025), intrinsic value creation will follow. "
         "The 2023 peak ROIC of 11% after Thermal Management divestiture shows capital discipline is possible."),
        ("M&A Quality",
         "The 2022 Thermal Management divestiture ($1.73B to Thomas & Betts) was astute – capital recycled into "
         "higher-growth enclosures and data center markets. Subsequent acquisitions (Electrical Products Group for "
         "$975M) at ~12x EBITDA look reasonable given the synergy potential with existing channels. "
         "Risk: goodwill now 39% of total assets – any impairment could hit reported equity hard."),
        ("Capital Allocation Score",
         "✓ Dividends: Consistent, growing ~3% p.a. (yield 0.62%) – shareholder-friendly but not a reason to own.\n"
         "✓ Buybacks: $253M in FY2025 (up from $60M in FY2023) – accelerating, signals confidence.\n"
         "✗ M&A: Premium valuations; goodwill intensity rising; integration risk.\n"
         "Overall Capital Allocation Score: 6.5 / 10."),
    ]
    for i, (k, v) in enumerate(commentary):
        r = note_r + 1 + i
        c1 = _data_cell(ws, r, 2, k, bg=LIGHT_BLUE, bold=True, align="left")
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)
        c2 = ws.cell(r, 3, v)
        c2.font = _font(size=FONT_SIZE)
        c2.fill = _fill(GREY if i%2==0 else WHITE)
        c2.alignment = _align("left", "center", wrap=True)
        c2.border = _border()
        ws.row_dimensions[r].height = 75

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[note_r].height = 22

# ────────────────────────────────────────────────────────────────────────────
# 8. MANAGEMENT
# ────────────────────────────────────────────────────────────────────────────
def build_management(wb):
    ws = wb.create_sheet("Management")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, [2, 26, 42, 20, 2])
    _section_hdr(ws, 1, 2, "nVent Electric plc – Management Analysis", span=3)

    # Key Executives
    _hdr_cell(ws, 3, 2, "Key Executives", bg=BLUE, span=3)
    _hdr_cell(ws, 4, 2, "Name / Title", bg=NAVY)
    _hdr_cell(ws, 4, 3, "Background & Assessment", bg=NAVY)
    _hdr_cell(ws, 4, 4, "Owner Score (1-10)", bg=NAVY)

    mgmt = [
        ("Beth Wozniak\nChair & CEO",
         "Wozniak has led NVT since its 2018 Pentair spin-off. 30+ years in electrical/industrial markets "
         "(previously EVP at Honeywell). She engineered the strategic pivot toward data center infrastructure "
         "and successfully divested the Thermal Management segment in 2022, unlocking capital for higher-growth "
         "investments. Under her leadership: revenue CAGR ~12%, adj. EPS CAGR ~22% (2018-2025). "
         "CONCERN: She has sold $1.3M+ of stock in the past 6 months – may signal valuation concern at current levels.",
         "7 / 10"),
        ("Sara Zawoyski\nEVP & CFO",
         "Joined NVT in 2018 from Pentair. Conservative balance sheet manager – reduced net debt post-Thermal "
         "divestiture proceeds. Emphasis on FCF conversion and disciplined M&A. Consistent dividend growth. "
         "POSITIVE: Clear communicator with institutional investors.",
         "8 / 10"),
        ("Operations / Segment Leads",
         "Segment leaders for Systems Protection (data center focus) and Electrical Connections have deep "
         "operational experience in electrification markets. NVT promotes organic talent from engineering & "
         "sales backgrounds – customer-centric culture.",
         "7 / 10"),
    ]
    for i, (n, b, s) in enumerate(mgmt):
        r = 5 + i
        c1 = _data_cell(ws, r, 2, n, bg=LIGHT_BLUE, bold=True, align="left")
        c1.alignment = _align("left","center", wrap=True)
        c2 = _data_cell(ws, r, 3, b, bg=GREY if i%2==0 else WHITE, align="left")
        c2.alignment = _align("left","center", wrap=True)
        c3 = _data_cell(ws, r, 4, s, bg=LIGHT_GREEN, bold=True, align="center")
        ws.row_dimensions[r].height = 90

    # Compensation / Incentives
    _hdr_cell(ws, 10, 2, "Executive Compensation & Incentives (Proxy 2026)", bg=BLUE, span=3)
    comp_rows = [
        ("Base Salary",
         "CEO base ~$1.2M; CFO ~$700K. Competitive with mid-cap industrial peers. Not excessive."),
        ("Annual Cash Bonus",
         "Tied to organic sales growth, adjusted EPS, and FCF conversion. "
         "~50% revenue/EPS metrics, 50% FCF conversion – aligns management with shareholder FCF generation."),
        ("Long-Term Equity",
         "PSUs (Performance Share Units) and RSUs. CEO received 45,221 stock options + 18,708 RSUs "
         "in March 2026 vesting ratable over 3 years. PSU payout tied to relative TSR and ROIC vs. peers. "
         "ROIC metric is a POSITIVE signal – it's the right metric for owner-operator mentality."),
        ("Total CEO Comp (est.)",
         "~$12-14M all-in (salary + bonus + equity). Reasonable for a $22B market cap company. "
         "Pay-for-performance linkage is credible."),
        ("Insider Ownership",
         "CEO owns ~0.1% of shares outstanding. Board and management collectively own <1%. "
         "Low insider ownership is a NEGATIVE for owner-operator culture – management may be "
         "more focused on short-term metrics tied to compensation than long-term value creation."),
    ]
    for i, (k, v) in enumerate(comp_rows):
        r = 11 + i
        _data_cell(ws, r, 2, k, bg=LIGHT_BLUE, bold=True, align="left")
        c2 = _data_cell(ws, r, 3, v, bg=GREY if i%2==0 else WHITE, align="left")
        c2.alignment = _align("left","center", wrap=True)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
        ws.row_dimensions[r].height = 60

    # Insider Activity
    _hdr_cell(ws, 18, 2, "Insider Buying / Selling Activity (SEC Form 4)", bg=BLUE, span=3)
    _hdr_cell(ws, 19, 2, "Date", bg=NAVY)
    _hdr_cell(ws, 19, 3, "Insider", bg=NAVY)
    _hdr_cell(ws, 19, 4, "Activity", bg=NAVY)
    insider_data = [
        ("Mar 30, 2026", "Beth Wozniak (CEO)",    "SOLD 7,597 shares @ ~$116.50 → ~$884K"),
        ("Feb 2026",     "Beth Wozniak (CEO)",    "SOLD 4,137 shares → ~$456K (Form 4 filing)"),
        ("Nov 2025",     "Beth Wozniak (CEO)",    "SOLD ~4,137 shares → ~$456K"),
        ("Mar 2026",     "Board / Compensation",  "Granted 45,221 options + 18,708 RSUs to CEO (routine annual grant)"),
        ("2025 Full Yr", "Various Directors",     "No significant insider PURCHASES reported in FY2025-2026"),
    ]
    for i, (dt, who, act) in enumerate(insider_data):
        r = 20 + i
        _data_cell(ws, r, 2, dt, bg=GREY if i%2==0 else WHITE, align="center")
        _data_cell(ws, r, 3, who, bg=GREY if i%2==0 else WHITE, bold=True, align="left")
        c = _data_cell(ws, r, 4, act,
                       bg=LIGHT_RED if "SOLD" in act else (LIGHT_GREEN if "BOUGHT" in act else WHITE),
                       align="left")
        c.alignment = _align("left","center", wrap=True)
        ws.row_dimensions[r].height = 30

    # CEO Owner-Operator Assessment
    _hdr_cell(ws, 27, 2, "CEO Owner-Operator Assessment", bg=BLUE, span=3)
    owner_checks = [
        ("Plants seeds for future?",
         "YES – Launched 86 new products in 2025 contributing ~10pts to growth. "
         "Invested in liquid cooling capacity ahead of demand. Heavy R&D in next-gen thermal solutions.", GREEN),
        ("Acts like an owner?",
         "PARTIAL – Compensation tied to ROIC (positive), but recent stock sales and "
         "low insider ownership (<1%) suggest limited personal financial exposure to long-term outcomes.", YELLOW),
        ("Capital allocation discipline?",
         "GOOD – Thermal Management divestiture (2022) was strategically astute. "
         "M&A bolt-ons at reasonable prices. Buybacks increasing as FCF grows.", LIGHT_GREEN),
        ("Borrows from future?",
         "MODERATE RISK – Working capital build in 2025 depressed GAAP FCF; "
         "heavy goodwill from acquisitions requires strong execution to justify. "
         "Tariff impact of ~$90M partially offset by price increases – some pricing risk ahead.", YELLOW),
        ("Overall Owner-Operator Score",
         "6.5 / 10 – Competent, strategically sharp CEO who has driven strong value creation "
         "since the 2018 spin-off. Key concern is recent stock sales at a moment when the stock "
         "appears fully valued. Management is good but not exceptional owner-operators.", LIGHT_BLUE),
    ]
    for i, (k, v, bg) in enumerate(owner_checks):
        r = 28 + i
        _data_cell(ws, r, 2, k, bg=LIGHT_BLUE, bold=True, align="left")
        c2 = _data_cell(ws, r, 3, v, bg=bg, align="left")
        c2.alignment = _align("left","center", wrap=True)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
        ws.row_dimensions[r].height = 60

    for r in [1,2,3,10,18,27]:
        ws.row_dimensions[r].height = 22
    ws.row_dimensions[4].height = 22
    ws.row_dimensions[19].height = 22

# ────────────────────────────────────────────────────────────────────────────
# 9. RISKS
# ────────────────────────────────────────────────────────────────────────────
def build_risks(wb):
    ws = wb.create_sheet("Risks")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, [2, 24, 42, 12, 12, 16, 2])
    _section_hdr(ws, 1, 2, "nVent Electric plc – Risk Analysis", span=5)

    _hdr_cell(ws, 3, 2, "Risk Factor", bg=NAVY)
    _hdr_cell(ws, 3, 3, "Description & Nuance", bg=NAVY)
    _hdr_cell(ws, 3, 4, "Probability\n(1-5)", bg=NAVY)
    _hdr_cell(ws, 3, 5, "Impact\n(1-5)", bg=NAVY)
    _hdr_cell(ws, 3, 6, "Overall\n(P×I, max 25)", bg=NAVY)

    risks = [
        ("AI/Data Center Capex Cycle Slowdown",
         "~25% of NVT revenue is data center. If hyperscalers (AWS, Azure, Google, Meta) pull back "
         "on AI infrastructure spending – even temporarily – Systems Protection growth could decelerate "
         "sharply. Precedent: 2016 data center pause. NVT's 1B+ backlog provides 6-12mo buffer but not "
         "permanent protection. This is the single largest risk to the bull thesis.",
         4, 5, 20),
        ("Tariff & Trade Policy Headwinds",
         "Management guided ~$90M tariff impact in 2025. Further escalation of US-China tariffs or "
         "reciprocal tariffs on European goods (NVT sources from EU, Mexico, Asia) could further pressure "
         "gross margins. Price increases can offset but create volume risk in competitive bids.",
         4, 3, 12),
        ("Competition Intensification (Vertiv / Rittal)",
         "Vertiv (VRT) is aggressively expanding into the same AI data center market with power and "
         "cooling solutions. Rittal (private) expanding North American capacity. Schneider Electric "
         "deploying large capital. NVT's top-3 position is not a guarantee against spec losses.",
         3, 3, 9),
        ("CEO Insider Selling",
         "Beth Wozniak has sold $1.3M+ in NVT stock since Nov 2025. While routine RSU/option exercises "
         "are normal, net selling by the CEO at a high-multiple stock is a caution flag. Does not confirm "
         "fundamentals are weak, but suggests limited insider conviction at current valuations.",
         3, 2, 6),
        ("ROIC Below WACC",
         "Current ROIC of ~9.5% is at/below estimated WACC of ~9.6%. The company must sustain "
         "above-market growth to justify premium valuation multiples. If growth decelerates, "
         "the market will re-rate the multiple from 28x EV/EBITDA toward industrial peers at 15-18x.",
         3, 4, 12),
        ("Goodwill Impairment Risk",
         "Goodwill at $2.68B (39% of total assets) reflects historical acquisition premiums. "
         "Any segment disappointing (especially if data center enthusiasm fades) could trigger "
         "a non-cash impairment charge, eroding reported net income and equity. "
         "2024 net income was already distorted by similar charges.",
         2, 4, 8),
        ("Execution Risk: Liquid Cooling Scale-Up",
         "Liquid cooling is an emerging product requiring significant manufacturing scale-up, "
         "supply chain development, and field service capability. Early-stage technology always carries "
         "execution risk. Any high-profile data center cooling failure could damage reputation.",
         2, 3, 6),
        ("Interest Rate Sensitivity",
         "$1.56B in debt (net debt $1.32B). Rising rates increase interest expense. "
         "However, at 1.5x Net Debt/EBITDA, leverage is manageable. "
         "If rates rise significantly, re-financing costs could pressure EPS.",
         2, 2, 4),
        ("Currency / FX Exposure",
         "~40% of revenue from outside Americas (Europe + APAC). A strong USD depresses reported "
         "revenue and earnings. This is a recurring but manageable risk given hedging programs.",
         3, 2, 6),
        ("Valuation De-rating Risk",
         "Stock at EV/EBITDA 28x vs industrial average ~15-18x. Any disappointment in growth "
         "or margin could cause rapid multiple compression. DCF intrinsic value at conservative "
         "assumptions suggests ~$64-90/share range – well below current ~$133.",
         3, 5, 15),
    ]

    for i, (rsk, desc, prob, imp, score) in enumerate(risks):
        r = 4 + i
        score_bg = RED if score >= 15 else (ORANGE if score >= 10 else (YELLOW if score >= 6 else LIGHT_GREEN))
        _data_cell(ws, r, 2, rsk, bg=LIGHT_BLUE if i%2==0 else WHITE, bold=True, align="left")
        c2 = _data_cell(ws, r, 3, desc, bg=GREY if i%2==0 else WHITE, align="left")
        c2.alignment = _align("left", "center", wrap=True)
        _data_cell(ws, r, 4, prob, bg=GREY if i%2==0 else WHITE, align="center", bold=True)
        _data_cell(ws, r, 5, imp, bg=GREY if i%2==0 else WHITE, align="center", bold=True)
        _data_cell(ws, r, 6, score, bg=score_bg, bold=True, align="center",
                   colour=WHITE if score >= 15 else NAVY)
        ws.row_dimensions[r].height = 80

    # Risk legend
    leg_r = 4 + len(risks) + 1
    _hdr_cell(ws, leg_r, 2, "Risk Severity Legend", bg=BLUE, span=5)
    legends = [
        ("15-25  CRITICAL",  RED,    WHITE),
        ("10-14  HIGH",       ORANGE, WHITE),
        ("6-9    MEDIUM",     YELLOW, NAVY),
        ("1-5    LOW",        LIGHT_GREEN, NAVY),
    ]
    for i, (lbl, bg, fg) in enumerate(legends):
        r = leg_r + 1 + i
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        c = ws.cell(r, 2, lbl)
        c.font = _font(bold=True, colour=fg, size=FONT_SIZE)
        c.fill = _fill(bg)
        c.alignment = _align("center", "center")
        ws.row_dimensions[r].height = 22

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[3].height = 30
    ws.row_dimensions[leg_r].height = 22

# ────────────────────────────────────────────────────────────────────────────
# 10. VALUATION
# ────────────────────────────────────────────────────────────────────────────
def build_valuation(wb):
    ws = wb.create_sheet("Valuation")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, [2, 30, 18, 18, 18, 18, 2])
    _section_hdr(ws, 1, 2, "nVent Electric plc – Valuation Analysis", span=5)

    # Current trading multiples
    _hdr_cell(ws, 3, 2, "Current Market Snapshot (April 21, 2026)", bg=BLUE, span=5)
    snap = [
        ("Current Stock Price",     "$133.16",  "Analyst Avg Target", "$140 – $152"),
        ("Market Capitalization",   "$22.0B",   "Enterprise Value",   "$23.8B"),
        ("Shares Outstanding",      "161.7M",   "52-Wk Range",        "$41.70 – $129.94"),
        ("Dividend Yield",          "0.62%",    "Beta",               "~1.15"),
        ("Consensus Rating",        "Strong Buy","# Analysts",        "14 Buy / 1 Hold / 0 Sell"),
    ]
    for i, (l1, v1, l2, v2) in enumerate(snap):
        r = 4 + i
        _data_cell(ws, r, 2, l1, bg=LIGHT_BLUE, bold=True, align="left")
        _data_cell(ws, r, 3, v1, bg=GREY, bold=True, align="center")
        _data_cell(ws, r, 4, l2, bg=LIGHT_BLUE, bold=True, align="left")
        _data_cell(ws, r, 5, v2, bg=GREY, bold=True, align="center")
        ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=6)
        ws.row_dimensions[r].height = 22

    # Multiples table
    _hdr_cell(ws, 11, 2, "Valuation Multiples vs Peers", bg=BLUE, span=5)
    _hdr_cell(ws, 12, 2, "Multiple", bg=NAVY)
    _hdr_cell(ws, 12, 3, "NVT Current", bg=NAVY)
    _hdr_cell(ws, 12, 4, "Industrial Avg", bg=NAVY)
    _hdr_cell(ws, 12, 5, "Premium / (Discount)", bg=NAVY)
    ws.merge_cells(start_row=12, start_column=5, end_row=12, end_column=6)

    mults = [
        ("P/E (Trailing)",          "31.6x",  "22x",   "+44% premium"),
        ("P/E (Forward FY2026E)",   "32.5x",  "20x",   "+63% premium"),
        ("EV/EBITDA",               "28.3x",  "16x",   "+77% premium"),
        ("EV/Revenue",              "6.1x",   "2.5x",  "+144% premium"),
        ("P/S",                     "5.7x",   "2.2x",  "+159% premium"),
        ("P/B",                     "5.9x",   "3.0x",  "+97% premium"),
        ("EV/FCF",                  "55.4x",  "25x",   "+122% premium"),
        ("FCF Yield",               "1.8%",   "4.0%",  "-2.2pp"),
    ]
    for i, (m, nvt, avg, comp) in enumerate(mults):
        r = 13 + i
        is_premium = "premium" in comp
        comp_bg = LIGHT_RED if is_premium else LIGHT_GREEN
        _data_cell(ws, r, 2, m, bg=LIGHT_BLUE if i%2==0 else WHITE, bold=True, align="left")
        _data_cell(ws, r, 3, nvt, bg=GREY, bold=True, align="center", colour=RED)
        _data_cell(ws, r, 4, avg, bg=GREY, align="center")
        c = _data_cell(ws, r, 5, comp, bg=comp_bg, bold=True, align="center")
        ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=6)
        ws.row_dimensions[r].height = 22

    # DCF Valuation
    dcf_r = 13 + len(mults) + 2
    _hdr_cell(ws, dcf_r, 2, "DCF Valuation – 5-Year Explicit Period", bg=BLUE, span=5)
    _hdr_cell(ws, dcf_r+1, 2, "Assumption", bg=NAVY)
    _hdr_cell(ws, dcf_r+1, 3, "Base Case", bg=NAVY)
    _hdr_cell(ws, dcf_r+1, 4, "Bull Case", bg=NAVY)
    _hdr_cell(ws, dcf_r+1, 5, "Bear Case", bg=NAVY)
    ws.merge_cells(start_row=dcf_r+1, start_column=5, end_row=dcf_r+1, end_column=6)

    dcf_inputs = [
        ("Starting Adj. FCF (FY2025)",  "$561M",    "$561M",    "$561M"),
        ("Year 1-2 FCF Growth",          "15%",      "22%",      "8%"),
        ("Year 3-5 FCF Growth",          "10%",      "15%",      "4%"),
        ("Terminal Growth Rate",         "3.0%",     "3.5%",     "2.5%"),
        ("WACC",                         "9.5%",     "8.5%",     "10.5%"),
        ("Net Debt (FY2025)",            "$1,323M",  "$1,323M",  "$1,323M"),
        ("Shares Outstanding",           "161.7M",   "161.7M",   "161.7M"),
    ]
    dcf_outputs = [
        ("PV of Explicit FCFs",          "$2,986M",  "$3,521M",  "$2,291M"),
        ("Terminal Value (PV)",          "$9,220M",  "$12,180M", "$6,145M"),
        ("Enterprise Value (DCF)",       "$12,206M", "$15,701M", "$8,436M"),
        ("Equity Value",                 "$10,883M", "$14,378M", "$7,113M"),
        ("Intrinsic Value per Share",    "$67.30",   "$88.90",   "$44.00"),
        ("Current Price",                "$133.16",  "$133.16",  "$133.16"),
        ("Upside / (Downside)",          "-49%",     "-33%",     "-67%"),
    ]

    for i, (lbl, b, bu, be) in enumerate(dcf_inputs):
        r = dcf_r + 2 + i
        _data_cell(ws, r, 2, lbl, bg=LIGHT_BLUE, bold=True, align="left")
        _data_cell(ws, r, 3, b, bg=GREY, align="center")
        _data_cell(ws, r, 4, bu, bg=LIGHT_GREEN, align="center")
        c = _data_cell(ws, r, 5, be, bg=LIGHT_RED, align="center")
        ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=6)
        ws.row_dimensions[r].height = 22

    sep_r = dcf_r + 2 + len(dcf_inputs)
    ws.merge_cells(start_row=sep_r, start_column=2, end_row=sep_r, end_column=6)
    c = ws.cell(sep_r, 2, "DCF Output")
    c.font = _font(bold=True, colour=WHITE, size=FONT_SIZE)
    c.fill = _fill(BLUE); c.alignment = _align("center")
    ws.row_dimensions[sep_r].height = 22

    for i, (lbl, b, bu, be) in enumerate(dcf_outputs):
        r = sep_r + 1 + i
        is_final = "Intrinsic" in lbl or "Upside" in lbl
        bg_b = YELLOW if is_final else WHITE
        bg_bu = LIGHT_GREEN
        bg_be = LIGHT_RED
        _data_cell(ws, r, 2, lbl, bg=LIGHT_BLUE, bold=is_final, align="left")
        _data_cell(ws, r, 3, b, bg=bg_b, bold=is_final, align="center",
                   colour=RED if "Downside" in lbl or "-" in b else NAVY)
        _data_cell(ws, r, 4, bu, bg=bg_bu, bold=is_final, align="center")
        c = _data_cell(ws, r, 5, be, bg=bg_be, bold=is_final, align="center")
        ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=6)
        ws.row_dimensions[r].height = 22

    # Valuation Commentary
    vc_r = sep_r + 1 + len(dcf_outputs) + 1
    _hdr_cell(ws, vc_r, 2, "Valuation Commentary & Margin of Safety", bg=BLUE, span=5)
    commentary = (
        "VERDICT: On a discounted cash flow basis, nVent Electric appears significantly OVERVALUED at ~$133/share. "
        "The DCF intrinsic value ranges from ~$44 (bear) to ~$89 (bull) vs the current price, implying a 33-67% "
        "downside to intrinsic value.\n\n"
        "HOWEVER – the DCF model may be conservative. The market is pricing NVT as a structural data-center "
        "beneficiary with sustained 15-20% growth for a decade. If AI infrastructure spending accelerates and "
        "NVT captures increasing share of liquid-cooling TAM, intrinsic value could reach $100-130/share.\n\n"
        "MARGIN OF SAFETY: At $133, there is NO margin of safety on a fundamental DCF basis. "
        "The stock is priced for perfection. Analyst consensus target of $140-152 reflects a momentum/growth "
        "premium rather than a value-investor entry point.\n\n"
        "OWNER LENS: A rational owner would not pay 28x EV/EBITDA for a business with ROIC at its WACC, "
        "a CEO selling stock, and execution risk in a nascent liquid-cooling market. "
        "A better entry price would be $80-100/share (15-18x forward EV/EBITDA), which would provide "
        "a margin of safety while still capturing data center upside."
    )
    ws.merge_cells(start_row=vc_r+1, start_column=2, end_row=vc_r+1, end_column=6)
    c = ws.cell(vc_r+1, 2, commentary)
    c.font = _font(size=FONT_SIZE)
    c.fill = _fill(YELLOW)
    c.alignment = _align("left", "top", wrap=True)
    c.border = _border()
    ws.row_dimensions[vc_r+1].height = 180

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[3].height = 22
    ws.row_dimensions[11].height = 22
    ws.row_dimensions[12].height = 22
    ws.row_dimensions[dcf_r].height = 22
    ws.row_dimensions[vc_r].height = 22

# ────────────────────────────────────────────────────────────────────────────
# 11. MARKET SENTIMENT
# ────────────────────────────────────────────────────────────────────────────
def build_market_sentiment(wb):
    ws = wb.create_sheet("Market Sentiment")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, [2, 24, 48, 16, 2])
    _section_hdr(ws, 1, 2, "nVent Electric plc – Market Sentiment & Trends", span=3)

    # Analyst Summary
    _hdr_cell(ws, 3, 2, "Analyst Ratings Summary", bg=BLUE, span=3)
    _hdr_cell(ws, 4, 2, "Metric", bg=NAVY)
    _hdr_cell(ws, 4, 3, "Detail", bg=NAVY)
    _hdr_cell(ws, 4, 4, "Signal", bg=NAVY)

    analyst_data = [
        ("Consensus Rating",    "Strong Buy – 13 Buy, 1 Hold, 0 Sell (of 14 analysts)", "BULLISH",  GREEN),
        ("Avg Price Target",    "$140.00 (range: $68 – $160); +5.1% upside from current $133", "MODEST UPSIDE", YELLOW),
        ("High Target",         "$160.00 (Citigroup, Apr 13 2026 – 12.7% upside)",       "BULLISH",  GREEN),
        ("Low Target",          "$68.04 (outlier bear case – significant downside concern)", "BEARISH", RED),
        ("Citigroup Latest",    "Price target raised to $152 on Apr 13, 2026 – citing data center demand acceleration", "BULLISH", GREEN),
        ("Q4 2025 Reaction",    "Stock dipped after Q4 2025 earnings despite strong results – market had expected more; "
                                 "high expectations are now built into the stock price", "CAUTIOUS", ORANGE),
    ]
    for i, (k, v, sig, sig_bg) in enumerate(analyst_data):
        r = 5 + i
        _data_cell(ws, r, 2, k, bg=LIGHT_BLUE if i%2==0 else WHITE, bold=True, align="left")
        c = _data_cell(ws, r, 3, v, bg=GREY if i%2==0 else WHITE, align="left")
        c.alignment = _align("left","center", wrap=True)
        _data_cell(ws, r, 4, sig, bg=sig_bg, bold=True, align="center",
                   colour=WHITE if sig_bg in [GREEN, RED, ORANGE] else NAVY)
        ws.row_dimensions[r].height = 40

    # Market trends & sentiment
    _hdr_cell(ws, 13, 2, "Key Market Trends & Competitive Dynamics (April 2026)", bg=BLUE, span=3)
    trends = [
        ("AI Data Center Boom",
         "Hyperscaler AI capex remains at record levels in early 2026. AWS, Azure, Google, Meta have collectively "
         "committed $500B+ in data center infrastructure over 2024-2026. NVT's enclosure & liquid-cooling products "
         "are mission-critical for AI GPU clusters. Demand pipeline extends well into 2027.",
         "TAILWIND"),
        ("Liquid Cooling Adoption Acceleration",
         "NVIDIA H100/B100/Blackwell GPUs exceed air-cooling thermal limits. Liquid cooling (DLC, immersion) is "
         "transitioning from specialty to mainstream. NVT is positioned in the 'picks-and-shovels' of this transition. "
         "Market for liquid cooling systems in data centers is expected to grow at 20%+ CAGR through 2030.",
         "STRONG TAILWIND"),
        ("Electrification & Grid Modernization",
         "US infrastructure spending (IRA, CHIPS Act) driving industrial electrification. NVT's Electrical "
         "Connections segment benefits from grid upgrades, EV charging infrastructure, and reshoring of manufacturing. "
         "Government spending is relatively insulated from economic cycles.",
         "TAILWIND"),
        ("Tariff & Trade Policy Uncertainty",
         "US tariff environment remains uncertain as of April 2026. $90M impact in 2025; 2026 guidance "
         "assumes partial mitigation through price increases and supply chain adjustments. "
         "Ongoing US-China trade tensions could further affect component costs.",
         "HEADWIND"),
        ("Competitor Expansion",
         "Vertiv (VRT) – closest AI data center infrastructure competitor – is aggressively expanding liquid "
         "cooling and power distribution. Rittal (private German co.) announced North American capacity expansion. "
         "Price pressure could emerge if capacity overshoots demand in enclosures.",
         "HEADWIND"),
        ("M&A Environment",
         "NVT acquired Electrical Products Group from Avail Infrastructure Solutions for $975M – a bolt-on "
         "that adds switchgear accessories and cable management products. Integration execution will be key. "
         "Industrial consolidation is ongoing; NVT could be both acquirer and acquisition target.",
         "NEUTRAL"),
        ("Investor Sentiment",
         "NVT has re-rated sharply from ~$41 (52-wk low) to $133 on AI infrastructure enthusiasm. "
         "The stock is now a 'consensus AI pick' in industrials – which means any disappointment will "
         "trigger outsized selling. Sentiment is HIGH but positioning is CROWDED.",
         "CAUTIOUS"),
    ]
    for i, (topic, detail, signal) in enumerate(trends):
        r = 14 + i
        sig_bg = GREEN if "TAILWIND" in signal and "STRONG" in signal else \
                 (LIGHT_GREEN if "TAILWIND" in signal else \
                 (LIGHT_RED if "HEADWIND" in signal else \
                 (YELLOW if "CAUTIOUS" in signal else GREY)))
        sig_fg = NAVY
        _data_cell(ws, r, 2, topic, bg=LIGHT_BLUE if i%2==0 else WHITE, bold=True, align="left")
        c = _data_cell(ws, r, 3, detail, bg=GREY if i%2==0 else WHITE, align="left")
        c.alignment = _align("left","center", wrap=True)
        _data_cell(ws, r, 4, signal, bg=sig_bg, bold=True, align="center", colour=sig_fg)
        ws.row_dimensions[r].height = 65

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[3].height = 22
    ws.row_dimensions[4].height = 22
    ws.row_dimensions[13].height = 22

# ────────────────────────────────────────────────────────────────────────────
# 12. KEY INDICATORS
# ────────────────────────────────────────────────────────────────────────────
def build_key_indicators(wb):
    ws = wb.create_sheet("Key Indicators")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, [2, 28, 16, 16, 16, 16, 14, 2])
    _section_hdr(ws, 1, 2, "nVent Electric plc – Key Indicators Dashboard", span=6)

    def kpi_block(start_row, title, items, hdr_bg=BLUE):
        _hdr_cell(ws, start_row, 2, title, bg=hdr_bg, span=6)
        for i, (lbl, val, chg, chg_bg) in enumerate(items):
            r = start_row + 1 + i
            _data_cell(ws, r, 2, lbl, bg=LIGHT_BLUE if i%2==0 else WHITE, bold=True, align="left")
            _data_cell(ws, r, 3, val, bg=GREY, bold=True, align="center", colour=NAVY)
            ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
            _data_cell(ws, r, 6, chg, bg=chg_bg, bold=True, align="center",
                       colour=WHITE if chg_bg in [RED, GREEN] else NAVY)
            ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=7)
            ws.row_dimensions[r].height = 22
        ws.row_dimensions[start_row].height = 22

    kpi_block(3, "GROWTH METRICS", [
        ("Revenue (FY2025)",          "$3,893M",         "+29.5% YoY",       LIGHT_GREEN),
        ("Revenue (FY2024)",          "$3,006M",         "+12.6% YoY",       LIGHT_GREEN),
        ("Rev 3-yr CAGR (2022-2025)", "19.3%",           "Accelerating",     GREEN),
        ("Adj. EPS (FY2025)",         "$3.35 (reported)", "+35% YoY",        GREEN),
        ("FCF Growth 2022-2025",      "+5% CAGR (GAAP)", "WC headwind 2025", YELLOW),
        ("Data Center Rev (2025)",    "~$1,000M",        "+270% orders YoY", GREEN),
        ("Organic Growth (FY2025)",   "~17% Systems / 10% Conn.", "Strong", LIGHT_GREEN),
        ("FY2026E Revenue",           "$4,480-4,590M",   "+15-18% guided",   LIGHT_GREEN),
    ])

    kpi_block(14, "PROFITABILITY METRICS", [
        ("Gross Margin (FY2025)",     "37.7%",  "vs 40.2% FY2024 (mix/tariffs)", YELLOW),
        ("Adj. Operating Margin",     "20.2%",  "vs 21.7% FY2024",              YELLOW),
        ("EBITDA Margin (FY2025)",    "~22.3%", "vs 26.2% FY2024",              YELLOW),
        ("Net Margin (FY2025)",       "18.2%",  "Elevated by tax items",         LIGHT_GREEN),
        ("FCF Margin (GAAP)",         "9.6%",   "WC build depressed; adj ~14%",  YELLOW),
        ("Target Adj. Op. Margin",    "21-22%", "Management 2026 guide",         LIGHT_GREEN),
    ])

    kpi_block(22, "BALANCE SHEET HEALTH", [
        ("Total Debt (FY2025)",       "$1,560M",  "Down from $2,155M FY2024", LIGHT_GREEN),
        ("Net Debt",                  "$1,323M",  "1.5x Net Debt/EBITDA",     LIGHT_GREEN),
        ("Cash",                      "$237.5M",  "Adequate liquidity",        LIGHT_GREEN),
        ("Current Ratio",             "1.32x",    "Adequate",                  LIGHT_GREEN),
        ("Goodwill",                  "$2,678M",  "39% of total assets",       YELLOW),
        ("Total Equity",              "$3,730M",  "+15% YoY",                  LIGHT_GREEN),
        ("Debt / Equity",             "41.8%",    "Manageable",                LIGHT_GREEN),
    ])

    kpi_block(31, "RETURN ON CAPITAL", [
        ("ROE",                       "12.3%",    "Below 15% threshold",        YELLOW),
        ("ROA",                       "5.8%",     "Improving",                  LIGHT_GREEN),
        ("ROIC",                      "9.45%",    "At WACC – no economic moat?", YELLOW),
        ("WACC (est.)",               "~9.6%",    "NVT not creating EVA",        YELLOW),
        ("Incr. ROIC (FY2025)",       "~22%",     "Strong on new capital",       GREEN),
        ("Div. Yield",                "0.62%",    "Growing ~3% p.a.",            LIGHT_GREEN),
        ("Buyback Yield (FY2025)",    "~1.1%",    "Accelerating",                LIGHT_GREEN),
    ])

    kpi_block(40, "VALUATION SNAPSHOT", [
        ("Stock Price (Apr 21, 2026)","$133.16",  "Near all-time highs",         YELLOW),
        ("Market Cap",                "$22.0B",   "Mid-large cap",               GREY),
        ("Enterprise Value",          "$23.8B",   "Includes net debt",           GREY),
        ("P/E (Trailing)",            "31.6x",    "Premium vs sector",           LIGHT_RED),
        ("EV/EBITDA",                 "28.3x",    "+77% vs industrial avg 16x",  LIGHT_RED),
        ("EV/FCF",                    "55.4x",    "Very expensive on FCF basis", RED),
        ("DCF Intrinsic Value",       "~$67 (base)", "49% downside to DCF",     LIGHT_RED),
        ("Analyst Avg Target",        "$140",     "Strong Buy consensus",        LIGHT_GREEN),
    ])

    ws.row_dimensions[1].height = 22

# ────────────────────────────────────────────────────────────────────────────
# MAIN
# ────────────────────────────────────────────────────────────────────────────
def main():
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    print("Building Cover...")
    build_cover(wb)
    print("Building Business Overview...")
    build_business_overview(wb)
    print("Building Moat...")
    build_moat(wb)
    print("Building Income Statements...")
    build_income_statement(wb)
    print("Building Balance Sheet...")
    build_balance_sheet(wb)
    print("Building Cash Flow Analysis...")
    build_cash_flow(wb)
    print("Building Return on Capital...")
    build_return_on_capital(wb)
    print("Building Management...")
    build_management(wb)
    print("Building Risks...")
    build_risks(wb)
    print("Building Valuation...")
    build_valuation(wb)
    print("Building Market Sentiment...")
    build_market_sentiment(wb)
    print("Building Key Indicators...")
    build_key_indicators(wb)

    out_path = "/Users/naji/WORK/github.com/AI/claude/Agent/MarketResearch/output/NVT_Financial_Analysis.xlsx"
    wb.save(out_path)
    print(f"\nSaved: {out_path}")
    return out_path

if __name__ == "__main__":
    main()
