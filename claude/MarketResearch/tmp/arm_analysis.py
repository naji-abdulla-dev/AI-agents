"""
Arm Holdings plc (NASDAQ: ARM) - Comprehensive Financial & Investment Analysis
Generated: April 2026
Fiscal Year: April 1 – March 31
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

OUTPUT_PATH = "/Users/naji/WORK/github.com/AI/claude/Agent/claude/MarketResearch/output/ARM_Financial_Analysis.xlsx"

# ── Colour Palette (Technology / Semiconductor) ───────────────────────────────
DARK_NAVY   = "0D1B2A"
ARM_BLUE    = "1565C0"
LIGHT_BLUE  = "BBDEFB"
ARM_ORANGE  = "F57C00"
GREEN       = "1B5E20"
LIGHT_GREEN = "E8F5E9"
RED         = "B71C1C"
LIGHT_RED   = "FFCDD2"
YELLOW      = "FFF9C4"
GREY        = "F5F5F5"
MID_GREY    = "90A4AE"
WHITE       = "FFFFFF"
ACCENT_GOLD = "F9A825"

FONT_SIZE = 14

# ── Style helpers ─────────────────────────────────────────────────────────────
def _font(bold=False, colour=None, size=FONT_SIZE, italic=False):
    return Font(name="Calibri", bold=bold,
                color=colour or "000000", size=size, italic=italic)

def _fill(hex_colour):
    return PatternFill("solid", fgColor=hex_colour)

def _border(style="thin"):
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _hdr_cell(ws, row, col, value, bg=DARK_NAVY, fg=WHITE, bold=True,
              align="center", span=1):
    c = ws.cell(row=row, column=col, value=value)
    c.font = _font(bold=bold, colour=fg, size=FONT_SIZE)
    c.fill = _fill(bg)
    c.alignment = _align(align, "center")
    c.border = _border()
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col + span - 1)
    return c

def _data_cell(ws, row, col, value, bg=WHITE, bold=False,
               align="right", fmt=None, colour=None, wrap=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = _font(bold=bold, colour=colour or "000000", size=FONT_SIZE)
    c.fill = _fill(bg)
    c.alignment = _align(align, "center", wrap=wrap)
    c.border = _border()
    if fmt:
        c.number_format = fmt
    return c

def _section_hdr(ws, row, col, text, span=7):
    _hdr_cell(ws, row, col, text, bg=ARM_BLUE, fg=WHITE, span=span)
    ws.row_dimensions[row].height = 26

def _set_col_widths(ws, widths):
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = w

def _note_row(ws, row, col, text, span=7):
    ws.merge_cells(start_row=row, start_column=col,
                   end_row=row, end_column=col + span - 1)
    c = ws.cell(row, col, text)
    c.font = _font(italic=True, colour="5D4037", size=FONT_SIZE - 1)
    c.fill = _fill(YELLOW)
    c.alignment = _align("left", "center", wrap=True)
    c.border = _border()
    ws.row_dimensions[row].height = 28

# ─────────────────────────────────────────────────────────────────────────────
# TAB 1 — COVER
# ─────────────────────────────────────────────────────────────────────────────
def build_cover(wb):
    ws = wb.create_sheet("Cover")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, [2, 26, 22, 22, 22, 22, 22, 2])

    for r in range(1, 55):
        for c in range(1, 9):
            ws.cell(r, c).fill = _fill(WHITE)

    # Banner
    ws.merge_cells("B1:G4")
    hdr = ws["B1"]
    hdr.value = "Arm Holdings plc (NASDAQ: ARM)"
    hdr.font = Font(name="Calibri", bold=True, size=28, color=WHITE)
    hdr.fill = _fill(DARK_NAVY)
    hdr.alignment = _align("center", "center")
    ws.row_dimensions[1].height = 40

    ws.merge_cells("B5:G6")
    sub = ws["B5"]
    sub.value = "Comprehensive Financial & Investment Analysis  |  April 2026"
    sub.font = Font(name="Calibri", bold=False, size=17, color=WHITE)
    sub.fill = _fill(ARM_BLUE)
    sub.alignment = _align("center", "center")
    ws.row_dimensions[5].height = 30

    ws.merge_cells("B7:G8")
    tag = ws["B7"]
    tag.value = "Semiconductor IP Licensor · Fabless Architecture · AI & Data Center Tailwinds"
    tag.font = Font(name="Calibri", italic=True, size=FONT_SIZE, color=DARK_NAVY)
    tag.fill = _fill(LIGHT_BLUE)
    tag.alignment = _align("center", "center")
    ws.row_dimensions[7].height = 26

    def kv(row, label, value):
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=7)
        lc = ws.cell(row, 2, label)
        lc.font = _font(bold=True, colour=WHITE, size=FONT_SIZE)
        lc.fill = _fill(DARK_NAVY)
        lc.alignment = _align("right", "center")
        lc.border = _border()
        vc = ws.cell(row, 4, value)
        vc.font = _font(colour=DARK_NAVY, size=FONT_SIZE)
        vc.fill = _fill(LIGHT_BLUE)
        vc.alignment = _align("left", "center")
        vc.border = _border()
        ws.row_dimensions[row].height = 22

    kv_rows = [
        (10, "Company",           "Arm Holdings plc"),
        (11, "Ticker",            "NASDAQ: ARM"),
        (12, "Sector",            "Technology – Semiconductor IP"),
        (13, "Headquarters",      "Cambridge, United Kingdom"),
        (14, "Founded",           "1990 (IPO September 2023)"),
        (15, "CEO",               "Rene Haas (CEO since February 2022)"),
        (16, "CFO",               "Jason Child"),
        (17, "Analysis Date",     "April 27, 2026"),
        (18, "Stock Price",       "~$205 (April 27, 2026)"),
        (19, "Market Cap",        "~$217B (1.06B diluted shares)"),
        (20, "Enterprise Value",  "~$215B (net cash: ~$2.4B)"),
        (21, "52-Wk Range",       "~$115 – ~$210"),
        (22, "Analyst Consensus", "82% Buy / 13% Hold / 4% Sell  (30 analysts)"),
        (23, "Avg Price Target",  "~$180 avg; High $240 (Guggenheim); Low $120 (BofA)"),
        (24, "Revenue FY2025",    "$4.01B (+24% YoY)  |  Royalties $2.2B + Licensing $1.8B"),
        (25, "Non-GAAP Op. Mgn",  "~41%  |  Armv9 now >50% of royalty revenue"),
        (26, "FY2026 Trend",      "4 consecutive $1B+ quarters; Q3 FY26: $1.24B (+26% YoY)"),
        (27, "Valuation",         "P/E (GAAP): ~336x  |  EV/EBITDA: ~107x  |  Non-GAAP P/E: ~70x"),
    ]
    for r, lbl, val in kv_rows:
        kv(r, lbl, val)

    # Investment Thesis header
    ws.merge_cells("B29:G29")
    th = ws["B29"]
    th.value = "INVESTMENT THESIS"
    th.font = _font(bold=True, colour=WHITE, size=FONT_SIZE + 1)
    th.fill = _fill(GREEN)
    th.alignment = _align("center", "center")
    ws.row_dimensions[29].height = 26

    thesis = [
        "• ARCHITECTURAL TOLL-BOOTH: Arm's ISA (Instruction Set Architecture) is the dominant standard in mobile computing (>99% of smartphones) and rapidly penetrating data centers. Every chip shipped using Arm IP pays a royalty. No chip can be made without design tools – Arm's moat is structural, not cyclical.",
        "• ARMV9 UPGRADE CYCLE: Arm's v9 architecture commands ~2× the royalty rate of the previous generation and now accounts for >50% of royalty revenue. As the installed base upgrades (smartphones, servers, IoT), average royalty per chip expands even without volume growth.",
        "• AI & DATA CENTER ACCELERATION: Custom Arm-based CPUs are winning at hyperscalers (AWS Graviton, Google Axion, Microsoft Cobalt, Nvidia Grace). Data-center royalties doubled YoY in FY2026. AI inference workloads favor Arm's energy-efficient architecture over x86.",
        "• COMPUTE SUBSYSTEM (CSS) LICENSING: CSS products allow chip designers to license complete, pre-validated subsystems rather than just IP. This significantly shortens time-to-market for customers and increases per-chip licensing revenue for Arm, shifting the revenue mix toward higher-value contracts.",
        "• KEY RISKS: Premium valuation leaves minimal margin of safety (GAAP P/E ~336x). FCF turned negative in FY2025 (–$36M) due to elevated capex and working capital changes. RISC-V open-source ISA gaining traction at Qualcomm and Meta. SoftBank (87% owner) represents a concentration/overhang risk.",
        "• VERDICT: World-class, durable IP franchise with secular tailwinds from AI and data center Arm adoption. The business quality is undeniable; the valuation demands flawless execution for years. HOLD at current price; initiating at ACCUMULATE below $160 for a 15–20% margin of safety.",
    ]
    for i, line in enumerate(thesis):
        ws.merge_cells(start_row=30 + i, start_column=2, end_row=30 + i, end_column=7)
        c = ws.cell(30 + i, 2, line)
        c.font = _font(size=FONT_SIZE)
        c.fill = _fill(GREY if i % 2 == 0 else WHITE)
        c.alignment = _align("left", "center", wrap=True)
        ws.row_dimensions[30 + i].height = 36

    ws.merge_cells("B37:G37")
    rat = ws["B37"]
    rat.value = "RATING: HOLD / ACCUMULATE  |  12-Month Target: $195  |  Upside (from $205): ~Flat; Entry: $160–175"
    rat.font = Font(name="Calibri", bold=True, size=FONT_SIZE + 1, color=WHITE)
    rat.fill = _fill(ARM_ORANGE)
    rat.alignment = _align("center", "center")
    ws.row_dimensions[37].height = 30


# ─────────────────────────────────────────────────────────────────────────────
# TAB 2 — BUSINESS OVERVIEW
# ─────────────────────────────────────────────────────────────────────────────
def build_business(wb):
    ws = wb.create_sheet("Business Overview")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, [2, 30, 22, 22, 22, 22, 22, 2])

    _section_hdr(ws, 1, 2, "Arm Holdings plc – Business Overview", span=6)
    ws.row_dimensions[1].height = 28

    # Revenue segments
    _hdr_cell(ws, 3, 2, "Revenue Segments (FY2025)", bg=ARM_BLUE, span=6)
    hdrs = ["Segment", "FY2024 ($M)", "FY2025 ($M)", "YoY Growth", "% of FY2025 Rev", "Description"]
    for i, h in enumerate(hdrs):
        _hdr_cell(ws, 4, 2 + i, h, bg=DARK_NAVY)
    ws.row_dimensions[4].height = 28

    seg_data = [
        ("Royalty Revenue",   "$1,819",  "$2,200",  "+21.0%", "54.9%",  "Per-chip royalties paid by chip designers on every Arm-based chip shipped; rate scales with ASP and architecture generation"),
        ("Licensing Revenue", "$1,414",  "$1,807",  "+27.8%", "45.1%",  "Upfront and annual fees from semiconductor companies and hyperscalers licensing Arm IP (TLA, ATA, CSS); higher-value, strategic contracts"),
        ("TOTAL",             "$3,233",  "$4,007",  "+24.0%", "100.0%", "Record FY2025 revenue; Royalties crossed $2B milestone for first time; Licensing hit record driven by AI/data-center custom silicon demand"),
    ]
    for i, (seg, fy24, fy25, yoy, pct, desc) in enumerate(seg_data):
        bg = LIGHT_BLUE if i < 2 else GREY
        bold = i == 2
        _data_cell(ws, 5 + i, 2, seg, bg=bg, align="left", bold=bold)
        _data_cell(ws, 5 + i, 3, fy24, bg=bg, align="right", bold=bold)
        _data_cell(ws, 5 + i, 4, fy25, bg=bg, align="right", bold=bold)
        _data_cell(ws, 5 + i, 5, yoy, bg=bg, align="center", bold=bold)
        _data_cell(ws, 5 + i, 6, pct, bg=bg, align="center", bold=bold)
        _data_cell(ws, 5 + i, 7, desc, bg=bg, align="left", bold=bold, wrap=True)
        ws.row_dimensions[5 + i].height = 36

    # Business model
    _section_hdr(ws, 9, 2, "Business Model & How Arm Makes Money", span=6)
    model = [
        ("IP Licensing (TLA/ATA)", "Semiconductor companies pay upfront/annual fees for the right to design chips using Arm IP. Arm Total Access (ATA) and Technology Licensing Agreements (TLA) are the primary vehicles. High-value, multi-year contracts. Revenue recognized over contract term."),
        ("Per-Chip Royalties",     "Every chip designed using Arm IP that is shipped commercially triggers a royalty. Rate = fixed % of chip ASP (typically 1–2%). Armv9 architecture commands ~2× higher royalty rate than Armv8 due to SVE2 AI extensions and TrustZone security."),
        ("Compute Subsystems (CSS)", "Pre-validated 'chip building blocks' delivered as a subsystem (CPU + interconnect + debug infrastructure). Customers license the full subsystem rather than individual IP. Significantly shortens design cycles for hyperscaler custom silicon."),
        ("Revenue Recognition",    "Licensing revenue = recognized ratably over TLA/ATA contract life or at point of delivery of IP. Royalty revenue = recognized in the quarter after chip shipment (one-quarter lag). This creates a natural delay between market share gains and royalty realization."),
    ]
    for i, (cat, desc) in enumerate(model):
        bg = GREY if i % 2 == 0 else WHITE
        _data_cell(ws, 11 + i, 2, cat, bg=bg, align="left", bold=True)
        ws.merge_cells(start_row=11 + i, start_column=3, end_row=11 + i, end_column=7)
        c = ws.cell(11 + i, 3, desc)
        c.font = _font(size=FONT_SIZE)
        c.fill = _fill(bg)
        c.alignment = _align("left", "center", wrap=True)
        c.border = _border()
        ws.row_dimensions[11 + i].height = 42

    # End-market breakdown
    _section_hdr(ws, 16, 2, "Revenue by End Market (FY2025 Royalties, Estimated)", span=6)
    hdrs2 = ["End Market", "Est. % of Royalties", "Key Customers", "Architecture Used", "Growth Driver", "Notes"]
    for i, h in enumerate(hdrs2):
        _hdr_cell(ws, 17, 2 + i, h, bg=DARK_NAVY)

    mkt_data = [
        ("Mobile",         "~45%",  "Apple, Qualcomm, MediaTek, Samsung", "Armv9 / Cortex-A / Apple M-series", "Armv9 upgrade cycle; AI NPU integration",    "Stable volume; higher ASP as Armv9 penetrates"),
        ("Data Center / Cloud", "~25%", "AWS, Google, Microsoft, NVIDIA, Ampere", "Neoverse V/N platforms; CSS", "Custom Arm CPUs replacing x86 for workloads", "Fastest-growing segment; royalties doubled YoY"),
        ("Automotive",     "~12%",  "NXP, Renesas, Tesla, Rivian",        "Cortex-R / Cortex-A",               "Autonomy, ADAS, EV compute demand",          "Long design cycles; high royalty rates"),
        ("IoT / Embedded", "~12%",  "Microchip, ST, TI, Nordic",          "Cortex-M series",                   "Smart devices, industrial automation",       "Largest volume but lowest per-chip royalties"),
        ("Networking",     "~6%",   "Marvell, Intel (infrastructure)",    "Neoverse / Cortex-A",               "5G infrastructure build-out",                "Moderate growth; specialized applications"),
    ]
    for i, row in enumerate(mkt_data):
        bg = GREY if i % 2 == 0 else WHITE
        for j, val in enumerate(row):
            _data_cell(ws, 18 + i, 2 + j, val, bg=bg, align="left" if j != 1 else "center", wrap=True)
        ws.row_dimensions[18 + i].height = 36

    # Geographic revenue
    _section_hdr(ws, 24, 2, "Geographic Revenue Distribution (FY2025, Approximate)", span=6)
    hdrs3 = ["Region", "Est. % of Revenue", "Primary Customers", "Notes"]
    for i, h in enumerate(hdrs3):
        _hdr_cell(ws, 25, 2 + i, h, bg=DARK_NAVY)
        if i == 2:
            ws.merge_cells(start_row=25, start_column=4, end_row=25, end_column=6)

    geo_data = [
        ("Asia Pacific (excl. China)", "~45%", "Samsung (Korea), Apple/TSMC fab chain (Taiwan), Qualcomm design (India)"),
        ("China",                       "~23%", "Huawei HiSilicon, MediaTek, Unisoc – moderate licensing restrictions from US export controls"),
        ("United States",               "~20%", "Apple (SoC design), AWS, Google, NVIDIA, Qualcomm (Snapdragon) – key hyperscaler CSS contracts"),
        ("Europe & Rest of World",       "~12%", "NXP (auto), STMicro (IoT), Renesas (auto) – royalties from automotive and industrial chips"),
    ]
    for i, (geo, pct, cust) in enumerate(geo_data):
        bg = GREY if i % 2 == 0 else WHITE
        _data_cell(ws, 26 + i, 2, geo, bg=bg, align="left", bold=True)
        _data_cell(ws, 26 + i, 3, pct, bg=bg, align="center")
        ws.merge_cells(start_row=26 + i, start_column=4, end_row=26 + i, end_column=7)
        c = ws.cell(26 + i, 4, cust)
        c.font = _font(size=FONT_SIZE)
        c.fill = _fill(bg)
        c.alignment = _align("left", "center", wrap=True)
        c.border = _border()
        ws.row_dimensions[26 + i].height = 34

    _note_row(ws, 31, 2, "NOTE: Geographic breakdown based on management disclosures and analyst estimates; Arm does not report by geography with precision. China represents ~23% of royalties but is subject to export control restrictions on advanced AI chip architectures.", span=6)


# ─────────────────────────────────────────────────────────────────────────────
# TAB 3 — MOAT
# ─────────────────────────────────────────────────────────────────────────────
def build_moat(wb):
    ws = wb.create_sheet("Moat")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, [2, 28, 14, 14, 36, 2])

    _section_hdr(ws, 1, 2, "Arm Holdings – Competitive Moat Analysis", span=4)

    hdrs = ["Moat Element", "Moat Width", "Category", "Evidence / Commentary"]
    _hdr_cell(ws, 3, 2, hdrs[0], bg=DARK_NAVY)
    _hdr_cell(ws, 3, 3, hdrs[1], bg=DARK_NAVY)
    _hdr_cell(ws, 3, 4, hdrs[2], bg=DARK_NAVY)
    _hdr_cell(ws, 3, 5, hdrs[3], bg=DARK_NAVY)
    ws.row_dimensions[3].height = 28

    moat_rows = [
        ("Standard / ISA Monopoly",     "WIDE",     "Network Effect",   "Arm ISA is the de-facto standard for mobile CPUs. 99%+ of smartphones use Arm. 325B+ chips shipped to date. Every OS, compiler, toolchain, and SDK is optimized for Arm – the switching cost for the entire ecosystem is prohibitive."),
        ("Developer Ecosystem",         "WIDE",     "Switching Cost",   "22M+ developers trained on Arm architecture. Software compiled for Arm (iOS, Android, most embedded firmware) does not run natively on RISC-V or x86 without recompilation. Large established ecosystems take decades to recreate."),
        ("Armv9 Royalty Rate Premium",  "WIDE",     "Pricing Power",    "Armv9 architecture commands approximately 2× the royalty rate per chip vs. Armv8. SVE2 (AI inference), TrustZone security, and Confidential Compute extensions justify the premium. Armv9 >50% of royalty revenue in Q3 FY2026."),
        ("Fabless Model / Zero Capex",  "WIDE",     "Cost Structure",   "Arm designs IP but fabricates nothing. No foundry capex, no inventory risk, no materials exposure. Gross margins 96%+. Every incremental dollar of royalty revenue flows almost entirely to pre-tax profit."),
        ("Multi-Year Licensing Contracts", "MODERATE", "Switching Cost", "TLA/ATA agreements are multi-year (typically 5–10 years) with large upfront payments. Chip designers plan on ARM IP years before tapeout. Switching to RISC-V mid-program would require re-design, re-verification, and re-tooling – cost runs to hundreds of millions per SoC."),
        ("Neoverse Data Center Platform","MODERATE", "Network Effect",   "AWS Graviton4, Google Axion, Microsoft Cobalt, Nvidia Grace are all built on Arm Neoverse. As hyperscalers standardize internal tooling on Arm, the cost to switch to x86 or RISC-V rises. Data-center royalties doubled YoY."),
        ("Compute Subsystems (CSS)",    "MODERATE", "Pricing Power",    "CSS allows customers to license pre-validated complete CPU subsystems. This raises the engineering value delivered per license and justifies higher revenue per chip design. Customers with tighter roadmaps (automotive, cloud) pay premium for reduced time-to-silicon."),
        ("China Exposure Risk",         "NARROW",   "Regulatory",       "~23% of royalties from China. US export controls restrict shipment of advanced Arm IP to certain Chinese entities. Huawei's HiSilicon is under restriction. This is a moat-erosion risk: if China accelerates domestic RISC-V adoption, Arm could lose a significant royalty pool."),
        ("RISC-V Threat",               "NARROW",   "Competition",      "RISC-V is a free, open-source ISA. Qualcomm acquired Ventana Micro (RISC-V) to develop royalty-free alternatives. RISC-V has reached 25% market penetration in IoT/microcontrollers. Threat is primarily in low-cost, low-royalty segments (not yet in high-performance smartphone SoCs)."),
    ]

    color_map = {"WIDE": GREEN, "MODERATE": ACCENT_GOLD, "NARROW": RED}
    for i, (element, width, cat, evidence) in enumerate(moat_rows):
        bg = GREY if i % 2 == 0 else WHITE
        _data_cell(ws, 4 + i, 2, element, bg=bg, bold=True, align="left")
        badge = color_map.get(width, GREY)
        c = ws.cell(4 + i, 3, width)
        c.font = _font(bold=True, colour=WHITE, size=FONT_SIZE - 1)
        c.fill = _fill(badge)
        c.alignment = _align("center", "center")
        c.border = _border()
        _data_cell(ws, 4 + i, 4, cat, bg=bg, align="center")
        _data_cell(ws, 4 + i, 5, evidence, bg=bg, align="left", wrap=True)
        ws.row_dimensions[4 + i].height = 40

    # Competitor comparison
    _section_hdr(ws, 14, 2, "Competitive Landscape", span=4)
    hdrs2 = ["Competitor / Threat", "ISA Type", "Relative Threat Level", "Assessment vs. Arm"]
    for i, h in enumerate(hdrs2):
        _hdr_cell(ws, 15, 2 + i, h, bg=DARK_NAVY)

    comp_data = [
        ("RISC-V (Open Source / SiFive / Ventana)", "RISC-V (free, open)",  "MEDIUM",  "Real but long-dated threat. Zero royalty cost is compelling. Ecosystem is immature for high-performance applications. Arm's 30-year head start in tooling, OS support, and developer skills creates a decade+ runway."),
        ("Qualcomm (RISC-V via Ventana Micro)",      "RISC-V (proprietary)", "MEDIUM-HIGH", "Qualcomm's acquisition of Ventana Micro (2025) is a direct challenge. Qualcomm has the scale to build a viable alternative. Primary motivation: reduce Arm royalty costs. Timeline: 2–4 years to production-grade RISC-V Snapdragon."),
        ("Intel (x86)",                              "x86 (CISC)",           "LOW",     "x86 competes in data centers but is losing share to Arm Neoverse. Power-efficiency disadvantage vs. Arm. Intel is not a credible mobile competitor. Intel has invested in RISC-V but from a different angle (infrastructure tooling)."),
        ("Apple (in-house Arm design)",              "Arm (licensed)",       "N/A",     "Apple designs its own Arm-based chips (Apple Silicon) under a perpetual TLA. Apple is a partner, not a competitor. However, Apple's in-house design capability means future renegotiation of royalty rates is a risk."),
        ("AMD / Marvell / Broadcom (network ASICs)", "x86 / Custom",         "LOW",     "Niche competition in specific data-center workloads. Arm's Neoverse is winning server CPU market share from all these players. Marvell and Broadcom are net customers of Arm IP for networking chips."),
    ]
    for i, (comp, isa, threat, assess) in enumerate(comp_data):
        bg = GREY if i % 2 == 0 else WHITE
        _data_cell(ws, 16 + i, 2, comp, bg=bg, bold=True, align="left")
        _data_cell(ws, 16 + i, 3, isa, bg=bg, align="center")
        threat_color = {"MEDIUM": ACCENT_GOLD, "MEDIUM-HIGH": ARM_ORANGE, "LOW": GREEN, "HIGH": RED, "N/A": MID_GREY}
        tc = ws.cell(16 + i, 4, threat)
        tc.font = _font(bold=True, colour=WHITE, size=FONT_SIZE - 1)
        tc.fill = _fill(threat_color.get(threat, GREY))
        tc.alignment = _align("center", "center")
        tc.border = _border()
        _data_cell(ws, 16 + i, 5, assess, bg=bg, align="left", wrap=True)
        ws.row_dimensions[16 + i].height = 42

    # Moat verdict
    _section_hdr(ws, 22, 2, "MOAT VERDICT", span=4, )
    ws.row_dimensions[22].height = 26
    ws.merge_cells("B23:E27")
    v = ws["B23"]
    v.value = (
        "NARROW-TO-WIDE MOAT — Arm's architectural monopoly in mobile and rapidly growing data-center presence "
        "represent a genuinely durable competitive advantage. The developer ecosystem (22M+ engineers), backward-compatible "
        "toolchain, and Armv9 royalty rate expansion create compounding pricing power.\n\n"
        "The RISC-V risk is real but over-stated in the near term: rebuilding a credible competitor to Arm's 30+ year "
        "ecosystem would require a decade and hundreds of billions in developer investment. In IoT/microcontrollers "
        "RISC-V is already gaining; in high-performance SoCs and data-center CPUs, the switching cost remains prohibitive.\n\n"
        "OVERALL MOAT RATING: NARROW (trending to WIDE as data-center royalties compound)."
    )
    v.font = _font(size=FONT_SIZE)
    v.fill = _fill(LIGHT_GREEN)
    v.alignment = _align("left", "top", wrap=True)
    v.border = _border()
    for r in range(23, 28):
        ws.row_dimensions[r].height = 28


# ─────────────────────────────────────────────────────────────────────────────
# TAB 4 — INCOME STATEMENTS
# ─────────────────────────────────────────────────────────────────────────────
def build_income(wb):
    ws = wb.create_sheet("Income Statements")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, [2, 34, 14, 14, 14, 14, 14, 2])

    _section_hdr(ws, 1, 2, "Arm Holdings – Income Statement (GAAP, USD $M, FY Apr–Mar)", span=6)

    years = ["Metric", "FY2021", "FY2022", "FY2023", "FY2024", "FY2025"]
    for i, y in enumerate(years):
        _hdr_cell(ws, 3, 2 + i, y, bg=DARK_NAVY)
    ws.row_dimensions[3].height = 28

    # Revenue section
    _hdr_cell(ws, 4, 2, "REVENUE", bg=ARM_BLUE, fg=WHITE, span=6)
    is_data = [
        ("Royalty Revenue",              ["~$1,150", "~$1,568", "~$1,682", "$1,819", "$2,200"], False, WHITE),
        ("Licensing & Other Revenue",    ["~$877",   "~$1,135", "~$997",  "$1,414", "$1,807"], False, WHITE),
        ("Total Revenue",                ["$2,027",  "$2,703",  "$2,679", "$3,233", "$4,007"], True,  LIGHT_BLUE),
        ("YoY Revenue Growth (%)",       ["—",       "+33.4%",  "–0.9%",  "+20.7%", "+24.0%"], False, GREY),
    ]
    for i, (metric, vals, bold, bg) in enumerate(is_data):
        _data_cell(ws, 5 + i, 2, metric, bg=bg, bold=bold, align="left")
        for j, v in enumerate(vals):
            _data_cell(ws, 5 + i, 3 + j, v, bg=bg, bold=bold, align="right")
        ws.row_dimensions[5 + i].height = 22

    # Profitability section
    _hdr_cell(ws, 10, 2, "PROFITABILITY", bg=ARM_BLUE, fg=WHITE, span=6)
    prof_data = [
        ("Gross Profit",             ["$1,882",  "$2,572",  "$2,573",  "$3,079",  "$3,884"],  True,  LIGHT_BLUE),
        ("Gross Margin (%)",         ["92.8%",   "95.1%",   "96.0%",   "95.2%",   "96.9%"],   False, GREY),
        ("R&D Expense",              ["~$(920)", "~$(1,095)","~$(1,198)","~$(1,801)","~$(2,098)"], False, WHITE),
        ("SG&A Expense",             ["~$(723)", "~$(844)", "~$(704)", "~$(1,167)","~$(978)"],  False, WHITE),
        ("Operating Income (GAAP)",  ["$239",    "$633",    "$671",    "$111",    "$808"],     True,  LIGHT_BLUE),
        ("Operating Margin (GAAP %)",["11.8%",   "23.4%",   "25.1%",   "3.4%",    "20.2%"],   False, GREY),
        ("Non-GAAP Op. Margin (%)",  ["~38%",    "~39%",    "~40%",    "~39%",    "~41%"],     False, WHITE),
        ("Net Income (GAAP)",        ["$388",    "$549",    "$524",    "$306",    "$653"],     True,  LIGHT_BLUE),
        ("Net Margin (GAAP %)",      ["19.1%",   "20.3%",   "19.6%",   "9.5%",    "16.3%"],   False, GREY),
        ("EPS Diluted (GAAP)",       ["$0.38",   "$0.54",   "$0.51",   "$0.29",   "$0.61"],   True,  LIGHT_BLUE),
        ("EPS Adj. (Non-GAAP est.)", ["~$0.82",  "~$1.04",  "~$1.11",  "~$1.24",  "~$1.56"],  False, WHITE),
    ]
    for i, (metric, vals, bold, bg) in enumerate(prof_data):
        _data_cell(ws, 11 + i, 2, metric, bg=bg, bold=bold, align="left")
        for j, v in enumerate(vals):
            _data_cell(ws, 11 + i, 3 + j, v, bg=bg, bold=bold, align="right")
        ws.row_dimensions[11 + i].height = 22

    _note_row(ws, 23, 2,
              "NOTE: FY2024 GAAP operating income fell to $111M due to a surge in R&D expense and SBC costs linked to IPO-related equity awards. "
              "Non-GAAP operating margin (excl. SBC, amortization of acquired intangibles) held at ~39–41% throughout. "
              "FY2025 GAAP recovery reflects higher revenue and normalising SBC expense post-IPO lockup.", span=6)

    # FY2026 quarterly update
    _section_hdr(ws, 25, 2, "FY2026 Quarterly Progression (Fiscal Year Apr 2025 – Mar 2026)", span=6)
    hdrs2 = ["Quarter", "Revenue", "YoY Growth", "Royalty Rev", "Licensing Rev", "Comment"]
    for i, h in enumerate(hdrs2):
        _hdr_cell(ws, 26, 2 + i, h, bg=DARK_NAVY)
    q_data = [
        ("Q1 FY2026 (Apr–Jun 2025)", "$939M",  "+14% YoY", "~$530M", "~$409M",  "First quarter above $900M; growth driven by data-center licensing"),
        ("Q2 FY2026 (Jul–Sep 2025)", "$1,024M", "+19% YoY", "$620M",  "~$404M",  "First $1B quarter in company history; Armv9 >50% of royalties"),
        ("Q3 FY2026 (Oct–Dec 2025)", "$1,240M", "+26% YoY", "$737M",  "~$503M",  "Fourth consecutive $1B quarter; royalties record $737M (+27% YoY)"),
        ("Q4 FY2026 (Jan–Mar 2026)", "~$1,225M","~+25% est","~$710M", "~$515M",  "Guidance $1.225B ±$50M; AGI CPU roadmap announcement March 26, 2026"),
        ("FY2026E Full Year",        "~$4,428M", "~+10.5%", "~$2,597M","~$1,831M","Management withheld full-year guidance citing tariff/macro uncertainty"),
    ]
    for i, row in enumerate(q_data):
        bg = GREY if i % 2 == 0 else WHITE
        bold = i == 4
        for j, val in enumerate(row):
            _data_cell(ws, 27 + i, 2 + j, val, bg=LIGHT_BLUE if bold else bg,
                       bold=bold, align="left" if j in (0, 5) else "right", wrap=True)
        ws.row_dimensions[27 + i].height = 28


# ─────────────────────────────────────────────────────────────────────────────
# TAB 5 — BALANCE SHEET
# ─────────────────────────────────────────────────────────────────────────────
def build_balance_sheet(wb):
    ws = wb.create_sheet("Balance Sheet")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, [2, 32, 16, 16, 16, 16, 2])

    _section_hdr(ws, 1, 2, "Arm Holdings – Balance Sheet (GAAP, USD $M, as of March 31)", span=5)

    years = ["Metric", "FY2022", "FY2023", "FY2024", "FY2025"]
    for i, y in enumerate(years):
        _hdr_cell(ws, 3, 2 + i, y, bg=DARK_NAVY)

    # ASSETS
    _hdr_cell(ws, 4, 2, "ASSETS", bg=ARM_BLUE, span=5)
    asset_data = [
        ("Cash & Cash Equivalents",    ["$1,004",  "$1,554",  "$1,925",  "$2,807"],  False, WHITE),
        ("Short-term Investments",     ["~$200",   "~$150",   "~$100",   "~$90"],    False, GREY),
        ("Accounts Receivable",        ["~$480",   "~$510",   "~$680",   "~$820"],   False, WHITE),
        ("Total Current Assets",       ["~$2,100", "~$2,500", "~$3,000", "~$4,200"], True,  LIGHT_BLUE),
        ("Goodwill",                   ["$1,636",  "$1,620",  "$1,611",  "$1,625"],  False, WHITE),
        ("Intangible Assets (net)",    ["$205",    "$138",    "$152",    "$240"],    False, GREY),
        ("PP&E (net)",                 ["~$140",   "~$160",   "~$200",   "~$320"],   False, WHITE),
        ("Other Long-term Assets",     ["~$429",   "~$448",   "~$1,088", "~$1,791"], False, GREY),
        ("TOTAL ASSETS",               ["$6,510",  "$6,866",  "$8,051",  "$10,176"], True,  LIGHT_BLUE),
    ]
    for i, (metric, vals, bold, bg) in enumerate(asset_data):
        _data_cell(ws, 5 + i, 2, metric, bg=bg, bold=bold, align="left")
        for j, v in enumerate(vals):
            _data_cell(ws, 5 + i, 3 + j, v, bg=bg, bold=bold, align="right")
        ws.row_dimensions[5 + i].height = 22

    # LIABILITIES
    _hdr_cell(ws, 15, 2, "LIABILITIES & EQUITY", bg=ARM_BLUE, span=5)
    liab_data = [
        ("Accounts Payable",            ["~$90",    "~$95",    "~$120",   "~$150"],   False, WHITE),
        ("Deferred Revenue (current)",  ["~$480",   "~$510",   "~$680",   "~$720"],   False, GREY),
        ("Short-term Debt",             ["~$40",    "~$30",    "~$28",    "~$50"],    False, WHITE),
        ("Total Current Liabilities",   ["~$1,200", "~$1,300", "~$1,550", "~$1,750"], True,  LIGHT_BLUE),
        ("Long-term Debt",              ["$221",    "$189",    "$200",    "$387"],    False, WHITE),
        ("Deferred Revenue (LT)",       ["~$600",   "~$700",   "~$900",   "~$1,100"], False, GREY),
        ("Other Long-term Liabilities", ["~$941",   "~$626",   "~$325",   "~$628"],   False, WHITE),
        ("TOTAL LIABILITIES",           ["~$2,962", "~$2,815", "~$2,775", "~$2,378"], True,  LIGHT_BLUE),
        ("TOTAL SHAREHOLDERS' EQUITY",  ["$3,548",  "$4,051",  "$5,276",  "$7,798"],  True,  LIGHT_BLUE),
    ]
    for i, (metric, vals, bold, bg) in enumerate(liab_data):
        _data_cell(ws, 16 + i, 2, metric, bg=bg, bold=bold, align="left")
        for j, v in enumerate(vals):
            _data_cell(ws, 16 + i, 3 + j, v, bg=bg, bold=bold, align="right")
        ws.row_dimensions[16 + i].height = 22

    # Key ratios
    _section_hdr(ws, 26, 2, "Key Balance Sheet Metrics", span=5)
    ratio_hdrs = ["Metric", "FY2022", "FY2023", "FY2024", "FY2025"]
    for i, h in enumerate(ratio_hdrs):
        _hdr_cell(ws, 27, 2 + i, h, bg=DARK_NAVY)

    ratio_data = [
        ("Net Cash (Cash – Total Debt)",   ["$743",   "$1,335",  "$1,697",  "$2,370"],  True,  LIGHT_GREEN),
        ("Total Debt / Equity",            ["0.07×",  "0.05×",   "0.04×",   "0.06×"],   False, WHITE),
        ("Net Debt / EBITDA",              ["n/m",    "n/m",     "n/m",     "n/m"],     False, GREY),
        ("Book Value per Share",           ["~$3.38", "~$3.86",  "~$5.03",  "~$7.43"],  False, WHITE),
        ("Current Ratio",                  ["~1.75×", "~1.92×",  "~1.94×",  "~2.40×"],  False, GREY),
        ("Tangible Book Value",            ["$1,707", "$2,293",  "$3,513",  "$5,933"],  True,  LIGHT_BLUE),
    ]
    for i, (metric, vals, bold, bg) in enumerate(ratio_data):
        _data_cell(ws, 28 + i, 2, metric, bg=bg, bold=bold, align="left")
        for j, v in enumerate(vals):
            _data_cell(ws, 28 + i, 3 + j, v, bg=bg, bold=bold, align="right")
        ws.row_dimensions[28 + i].height = 22

    _note_row(ws, 35, 2,
              "NOTE: Arm is effectively DEBT-FREE with a large net cash position ($2.4B as of FY2025). Total debt of $437M represents primarily "
              "lease liabilities and deferred obligations. The $1.6B of goodwill reflects Arm's pre-IPO ownership history under SoftBank. "
              "Strong balance sheet with no financial leverage risk.", span=5)


# ─────────────────────────────────────────────────────────────────────────────
# TAB 6 — CASH FLOW ANALYSIS
# ─────────────────────────────────────────────────────────────────────────────
def build_cashflow(wb):
    ws = wb.create_sheet("Cash Flow Analysis")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, [2, 34, 14, 14, 14, 14, 14, 2])

    _section_hdr(ws, 1, 2, "Arm Holdings – Cash Flow Statement (GAAP, USD $M, FY Apr–Mar)", span=6)

    years_hdr = ["Metric", "FY2021", "FY2022", "FY2023", "FY2024", "FY2025"]
    for i, y in enumerate(years_hdr):
        _hdr_cell(ws, 3, 2 + i, y, bg=DARK_NAVY)

    _hdr_cell(ws, 4, 2, "OPERATING CASH FLOW", bg=ARM_BLUE, span=6)
    ocf_data = [
        ("Net Income (GAAP)",              ["$388",   "$549",  "$524",  "$306",  "$653"],   False, WHITE),
        ("D&A and Amortization",           ["~$120",  "~$110", "~$100", "~$120", "~$140"],  False, GREY),
        ("Stock-Based Compensation (SBC)", ["~$130",  "~$180", "~$240", "~$890", "~$740"],  False, WHITE),
        ("Changes in Working Capital",     ["~$595",  "~$(381)","~$(125)","~$(483)","~$(1,354)"], False, LIGHT_RED),
        ("Cash from Operations (CFO)",     ["$1,233", "$458",  "$739",  "$833",  "$179"],   True,  LIGHT_BLUE),
        ("CFO Margin (%)",                 ["60.8%",  "16.9%", "27.6%", "25.8%", "4.5%"],   False, GREY),
    ]
    for i, (metric, vals, bold, bg) in enumerate(ocf_data):
        _data_cell(ws, 5 + i, 2, metric, bg=bg, bold=bold, align="left")
        for j, v in enumerate(vals):
            _data_cell(ws, 5 + i, 3 + j, v, bg=bg, bold=bold, align="right")
        ws.row_dimensions[5 + i].height = 22

    _note_row(ws, 12, 2,
              "NOTE: FY2025 CFO collapsed to $179M despite record revenue because large working-capital outflows (~$1.35B) "
              "offset net income + SBC. The outflow reflects: (a) large cash payments under new Arm Total Access (ATA) contracts "
              "where cash is paid upfront but recognized ratably; (b) a significant increase in contract assets under the revenue "
              "recognition standard. This is a TIMING issue, not a structural deterioration.", span=6)

    _hdr_cell(ws, 14, 2, "CAPITAL EXPENDITURE & FREE CASH FLOW", bg=ARM_BLUE, span=6)
    capex_data = [
        ("Capital Expenditures",          ["$(104)", "$(34)",  "$(64)",  "$(81)",  "$(215)"],  False, WHITE),
        ("as % of Revenue",               ["5.1%",   "1.3%",   "2.4%",   "2.5%",   "5.4%"],   False, GREY),
        ("Free Cash Flow (FCF = CFO-Capex)",["$1,129","$424",  "$675",   "$752",   "$(36)"],   True,  LIGHT_BLUE),
        ("FCF Margin (%)",                 ["55.7%",  "15.7%", "25.2%",  "23.3%",  "N/M"],    False, GREY),
        ("FCF Conversion (FCF / NI)",      ["291%",   "77%",   "129%",   "246%",   "N/M"],    False, WHITE),
    ]
    for i, (metric, vals, bold, bg) in enumerate(capex_data):
        # Highlight negative FCF row
        actual_bg = LIGHT_RED if (i == 2 and vals[-1] == "$(36)") else bg
        _data_cell(ws, 15 + i, 2, metric, bg=bg, bold=bold, align="left")
        for j, v in enumerate(vals):
            cell_bg = LIGHT_RED if (i == 2 and j == 4) else bg
            _data_cell(ws, 15 + i, 3 + j, v, bg=cell_bg, bold=bold, align="right")
        ws.row_dimensions[15 + i].height = 22

    _note_row(ws, 21, 2,
              "NOTE: FY2025 FCF turned negative (-$36M) primarily due to: (1) $179M CFO (timing issue as noted above) and "
              "(2) $215M capex – a 3× increase vs FY2024 – reflecting investment in engineering facilities, servers, and "
              "infrastructure for Arm's expanding CSS and data-center IP design teams. Normalized FCF (adjusting for working "
              "capital timing) would likely be $800M–$1.0B+. Management expects FCF to normalize in FY2026.", span=6)

    _hdr_cell(ws, 23, 2, "CAPITAL ALLOCATION", bg=ARM_BLUE, span=6)
    alloc_data = [
        ("Dividends Paid",             ["$(750)", "$0",  "$0",  "$0",  "$0"],   False, WHITE),
        ("Share Repurchases",          ["$0",     "$0",  "$0",  "$0",  "$0"],   False, GREY),
        ("Acquisitions",               ["$0",     "$0",  "$0",  "$0",  "$0"],   False, WHITE),
        ("Net Cash Change (end of yr)",["—",       "$454","$550","$371","$882"], False, GREY),
        ("Ending Cash Balance",        ["—",       "$1,004","$1,554","$1,925","$2,807"], True, LIGHT_BLUE),
    ]
    for i, (metric, vals, bold, bg) in enumerate(alloc_data):
        _data_cell(ws, 24 + i, 2, metric, bg=bg, bold=bold, align="left")
        for j, v in enumerate(vals):
            _data_cell(ws, 24 + i, 3 + j, v, bg=bg, bold=bold, align="right")
        ws.row_dimensions[24 + i].height = 22

    # Quality assessment
    _section_hdr(ws, 30, 2, "Cash Flow Quality Assessment", span=6)
    quality = [
        ("Gross Margin Quality",      "POSITIVE", "96.9% gross margins demonstrate near-pure IP business with essentially zero cost of goods. Every dollar of new royalty revenue is ~97¢ gross profit."),
        ("FCF Volatility",            "WATCH",    "FY2025 FCF turned negative due to timing of contract asset recognition. This creates quarterly noise but is not a structural problem. FY2021's $1.13B FCF demonstrates the underlying cash generation power."),
        ("SBC as % of Revenue",       "WATCH",    "SBC was ~$740M in FY2025 (18.5% of revenue). This is high but declining post-IPO. Non-GAAP adjustments exclude SBC, which explains the large gap between GAAP and non-GAAP metrics."),
        ("Capital Efficiency",        "POSITIVE", "Arm requires minimal capex to grow revenue. FY2025 capex of $215M (5.4% of revenue) is temporarily elevated. Normal capex run-rate is ~2–3% of revenue for a fabless IP licensor."),
        ("No Debt Service",           "POSITIVE", "Zero meaningful long-term debt. No interest expense burden. Growing net cash position ($2.4B at FY2025 year-end). Balance sheet is a source of strength, not a constraint."),
    ]
    for i, (aspect, signal, comment) in enumerate(quality):
        bg = GREY if i % 2 == 0 else WHITE
        sig_colors = {"POSITIVE": GREEN, "WATCH": ACCENT_GOLD, "NEGATIVE": RED}
        _data_cell(ws, 31 + i, 2, aspect, bg=bg, bold=True, align="left")
        sc = ws.cell(31 + i, 3, signal)
        sc.font = _font(bold=True, colour=WHITE, size=FONT_SIZE - 1)
        sc.fill = _fill(sig_colors.get(signal, GREY))
        sc.alignment = _align("center", "center")
        sc.border = _border()
        ws.merge_cells(start_row=31 + i, start_column=4, end_row=31 + i, end_column=7)
        c = ws.cell(31 + i, 4, comment)
        c.font = _font(size=FONT_SIZE)
        c.fill = _fill(bg)
        c.alignment = _align("left", "center", wrap=True)
        c.border = _border()
        ws.row_dimensions[31 + i].height = 36


# ─────────────────────────────────────────────────────────────────────────────
# TAB 7 — RETURN ON CAPITAL
# ─────────────────────────────────────────────────────────────────────────────
def build_roc(wb):
    ws = wb.create_sheet("Return on Capital")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, [2, 32, 14, 14, 14, 14, 14, 2])

    _section_hdr(ws, 1, 2, "Arm Holdings – Return on Capital Analysis", span=6)

    years = ["Metric", "FY2021", "FY2022", "FY2023", "FY2024", "FY2025"]
    for i, y in enumerate(years):
        _hdr_cell(ws, 3, 2 + i, y, bg=DARK_NAVY)

    _hdr_cell(ws, 4, 2, "GAAP RETURNS (Reported)", bg=ARM_BLUE, span=6)
    gaap_roc = [
        ("Return on Equity (ROE)",     ["~22%",  "~15%",  "~14%",  "~6.5%", "~9.9%"],  False, WHITE),
        ("Return on Assets (ROA)",     ["~6.8%", "~8.4%", "~7.8%", "~4.2%", "~7.2%"],  False, GREY),
        ("Return on Invested Capital", ["~11%",  "~14%",  "~15%",  "~2.5%", "~10%"],   False, WHITE),
        ("Est. WACC",                  ["—",     "—",     "~8.5%", "~9.0%", "~9.5%"],  False, GREY),
        ("Economic Spread (ROIC-WACC)","—",      "—",     "~+6.5%","~-6.5%","~+0.5%"], False, WHITE),
    ]
    for i, (metric, *rest) in enumerate(gaap_roc):
        vals = rest[0] if isinstance(rest[0], list) else rest
        bg = rest[1] if len(rest) > 1 else WHITE
        bold = rest[2] if len(rest) > 2 else False
        _data_cell(ws, 5 + i, 2, metric, bg=bg, align="left")
        for j, v in enumerate(vals):
            _data_cell(ws, 5 + i, 3 + j, v, bg=bg, align="right")
        ws.row_dimensions[5 + i].height = 22

    _note_row(ws, 11, 2,
              "NOTE: GAAP ROIC is suppressed by $1.6B in goodwill from Arm's pre-IPO SoftBank-era valuation. "
              "On a tangible capital basis (ex-goodwill), Arm's ROIC would be substantially higher (est. 30%+). "
              "The FY2024 dip in ROIC reflects the spike in R&D/SBC expense and $111M GAAP operating income.", span=6)

    _hdr_cell(ws, 13, 2, "NON-GAAP / ADJUSTED RETURNS", bg=ARM_BLUE, span=6)
    adj_roc = [
        ("Non-GAAP Operating Income",       ["~$771M","~$1,054M","~$1,067M","~$1,264M","~$1,643M"], True,  LIGHT_BLUE),
        ("Non-GAAP Operating Margin (%)",   ["~38%",  "~39%",    "~40%",    "~39%",    "~41%"],     False, WHITE),
        ("Adj. EBITDA (est.)",              ["~$1,000M","~$1,180M","~$1,200M","~$1,400M","~$1,800M"], True, LIGHT_BLUE),
        ("Adj. EBITDA Margin (%)",          ["~49%",  "~44%",    "~45%",    "~43%",    "~45%"],     False, GREY),
        ("Incremental Op. Margin (adj.)",   ["—",     "~38%",    "~37%",    "~—",      "~41%"],     False, WHITE),
    ]
    for i, (metric, vals, bold, bg) in enumerate(adj_roc):
        _data_cell(ws, 14 + i, 2, metric, bg=bg, bold=bold, align="left")
        for j, v in enumerate(vals):
            _data_cell(ws, 14 + i, 3 + j, v, bg=bg, bold=bold, align="right")
        ws.row_dimensions[14 + i].height = 22

    # Return on capital verdict
    _section_hdr(ws, 20, 2, "RETURN ON CAPITAL VERDICT", span=6)
    ws.row_dimensions[20].height = 26
    ws.merge_cells("B21:G25")
    v = ws["B21"]
    v.value = (
        "EXCEPTIONAL UNDERLYING ECONOMICS, OBSCURED BY ACCOUNTING\n\n"
        "1. Arm's core business generates 96%+ gross margins and 41% non-GAAP operating margins — "
        "among the highest of any technology company globally. These are the economics of a pure IP monopoly.\n\n"
        "2. Tangible ROIC (excluding goodwill) is estimated at 30%+, well above WACC (~9.5%). "
        "This is genuine value creation from IP licensing.\n\n"
        "3. GAAP ROIC appears modest (9–10%) because $1.6B of goodwill inflates the denominator. "
        "This accounting artefact should not be confused with business quality.\n\n"
        "4. The Armv9 royalty rate premium means that incremental royalty revenue carries higher margins "
        "than the existing base — margins should expand as Armv9 penetrates the installed base.\n\n"
        "OVERALL: World-class returns on tangible capital. GAAP metrics understate business quality."
    )
    v.font = _font(size=FONT_SIZE)
    v.fill = _fill(LIGHT_GREEN)
    v.alignment = _align("left", "top", wrap=True)
    v.border = _border()
    for r in range(21, 26):
        ws.row_dimensions[r].height = 30


# ─────────────────────────────────────────────────────────────────────────────
# TAB 8 — MANAGEMENT
# ─────────────────────────────────────────────────────────────────────────────
def build_management(wb):
    ws = wb.create_sheet("Management")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, [2, 26, 20, 14, 32, 2])

    _section_hdr(ws, 1, 2, "Arm Holdings – Management Analysis", span=4)

    # Key executives
    _hdr_cell(ws, 3, 2, "Key Executives", bg=ARM_BLUE, span=4)
    hdrs = ["Name", "Role", "Tenure", "Key Contribution"]
    for i, h in enumerate(hdrs):
        _hdr_cell(ws, 4, 2 + i, h, bg=DARK_NAVY)

    execs = [
        ("Rene Haas",       "President & CEO",              "CEO since Feb 2022 (~4 yrs); at Arm since 2013", "Led Arm through transformative IPO (Sept 2023); drove Armv9 adoption; launched CSS strategy; pivoted Arm's narrative from mobile IP licensor to AI infrastructure platform. Previously President, Intellectual Property (IP)."),
        ("Jason Child",     "CFO",                           "CFO since March 2022",                           "Joined from SoftBank portfolio (prev. CFO at Limeade and Kabam). Oversaw the IPO process and post-IPO investor relations. Manages TLA/ATA contract structures and revenue recognition."),
        ("Chris Bergey",    "SVP & GM, Client Products",    "At Arm since 2018",                               "Leads mobile and edge compute product group (Cortex-A, Cortex-M). Responsible for maintaining market leadership in smartphone SoC IP. Key relationships with Apple, Qualcomm, and Samsung design teams."),
        ("Mohamed Awad",    "SVP & GM, Infrastructure",     "At Arm since 2016",                               "Leads Neoverse (data center/cloud) and networking. Driving AWS Graviton, Google Axion, Microsoft Cobalt adoption. Key person for data-center royalty growth thesis."),
        ("Erin Brandin",    "Chief Legal Officer",           "At Arm since 2021",                               "Oversees IP licensing strategy, litigation (incl. Qualcomm dispute), and export control compliance. Critical role given Arm's IP-centric business model and China exposure."),
    ]
    for i, row in enumerate(execs):
        bg = GREY if i % 2 == 0 else WHITE
        for j, val in enumerate(row):
            align = "left"
            _data_cell(ws, 5 + i, 2 + j, val, bg=bg, align=align, wrap=True)
        ws.row_dimensions[5 + i].height = 44

    # Executive compensation
    _section_hdr(ws, 11, 2, "Executive Compensation & Incentive Alignment", span=4)
    _hdr_cell(ws, 12, 2, "Incentive Element", bg=DARK_NAVY)
    _hdr_cell(ws, 12, 3, "Structure", bg=DARK_NAVY)
    _hdr_cell(ws, 12, 4, "Alignment", bg=DARK_NAVY)
    _hdr_cell(ws, 12, 5, "Commentary", bg=DARK_NAVY)

    comp_data = [
        ("CEO Base Salary",        "~$875K annual",                        "MODERATE",  "Below-median for large-cap tech CEO; majority of comp is equity-linked."),
        ("Annual Bonus (STIP)",    "Revenue, EPS, strategic milestones",   "HIGH",      "Linked to financial performance and strategic goals (e.g., CSS adoption rate, Armv9 royalty %). Objective and verifiable."),
        ("Long-Term Equity (RSU)", "3–4 year vesting; partly performance-based", "HIGH", "Significant RSU awards at IPO and annually. Haas holds ~283K shares/ADSs directly + unvested RSUs. Long vesting aligns incentives with shareholders."),
        ("SoftBank Alignment",     "SoftBank owns ~87% of Arm",            "MODERATE",  "Haas operates with SoftBank's backing but must also serve minority public shareholders. Masayoshi Son's interests may not always align with minority shareholders (e.g., potential secondary sale pressure)."),
        ("IPO Lockup Expiry",      "Multiple tranches post-Sept 2023",     "WATCH",     "Post-IPO lockup expirations created selling pressure in late 2023/2024. Most lockups have now expired. Haas has sold under Rule 10b5-1 plans (pre-arranged schedule sales) – not a bearish insider signal."),
    ]
    for i, (elem, struct, align_rating, comment) in enumerate(comp_data):
        bg = GREY if i % 2 == 0 else WHITE
        _data_cell(ws, 13 + i, 2, elem, bg=bg, bold=True, align="left")
        _data_cell(ws, 13 + i, 3, struct, bg=bg, align="left", wrap=True)
        rating_colors = {"HIGH": GREEN, "MODERATE": ACCENT_GOLD, "WATCH": ARM_ORANGE, "LOW": RED}
        rc = ws.cell(13 + i, 4, align_rating)
        rc.font = _font(bold=True, colour=WHITE, size=FONT_SIZE - 1)
        rc.fill = _fill(rating_colors.get(align_rating, GREY))
        rc.alignment = _align("center", "center")
        rc.border = _border()
        _data_cell(ws, 13 + i, 5, comment, bg=bg, align="left", wrap=True)
        ws.row_dimensions[13 + i].height = 38

    # Insider activity
    _section_hdr(ws, 19, 2, "Insider Activity (SEC Form 4 – Key Transactions)", span=4)
    _hdr_cell(ws, 20, 2, "Date", bg=DARK_NAVY)
    _hdr_cell(ws, 20, 3, "Insider", bg=DARK_NAVY)
    _hdr_cell(ws, 20, 4, "Transaction", bg=DARK_NAVY)
    _hdr_cell(ws, 20, 5, "Details & Signal", bg=DARK_NAVY)

    insider = [
        ("Mar 25–26, 2026", "Rene Haas (CEO)",     "SOLD 31,853 shares @ ~$160–$163", "Pre-arranged Rule 10b5-1 plan (adopted June 2025). After sale holds 282,979 shares directly. Proceeds ~$5.2M. SIGNAL: NEUTRAL — pre-planned, no discretionary timing."),
        ("Ongoing 2025",    "Rene Haas (CEO)",     "SOLD ~9,299 shares (periodic)",   "Multiple tranches of 6,152 ADSs each under 10b5-1 plan. Total disclosed proceeds ~$2.5M across Q1 FY2026. Pattern: systematic quarterly liquidation of vested RSUs."),
        ("IPO Sept 2023",   "SoftBank Group",      "Offered 95.5M shares in IPO",     "Primary offering raised ~$4.7B. SoftBank retained ~87% stake. No secondary proceeds to SoftBank in IPO – shares were new issuance. This was constructive for minority shareholders."),
        ("No open-market purchases detected", "All insiders", "NO open-market BUYS", "No discretionary open-market purchases by any insider as of April 2026. All insider sales appear pre-planned. SIGNAL: NEUTRAL — not uncommon for recently-IPO'd tech company with large equity compensation packages."),
    ]
    for i, (date, insider_name, txn, details) in enumerate(insider):
        bg = GREY if i % 2 == 0 else WHITE
        _data_cell(ws, 21 + i, 2, date, bg=bg, align="left")
        _data_cell(ws, 21 + i, 3, insider_name, bg=bg, align="left", bold=True)
        _data_cell(ws, 21 + i, 4, txn, bg=bg, align="left")
        _data_cell(ws, 21 + i, 5, details, bg=bg, align="left", wrap=True)
        ws.row_dimensions[21 + i].height = 38

    # Management verdict
    _section_hdr(ws, 26, 2, "DOES MANAGEMENT ACT LIKE AN OWNER?", span=4)
    ws.merge_cells("B27:E31")
    v = ws["B27"]
    v.value = (
        "VERDICT: COMPETENT OPERATOR — ALIGNMENT ADEQUATE BUT NOT EXCEPTIONAL\n\n"
        "Rene Haas has been the right CEO for Arm's transition from a niche IP licensor to an AI infrastructure platform. "
        "His execution on the IPO, the Armv9 royalty rate expansion, and the CSS product strategy have all been high-quality strategic moves.\n\n"
        "INSIDER SIGNAL: NEUTRAL. Haas sells shares periodically under pre-arranged 10b5-1 plans — this is standard RSU "
        "monetization behavior, not a signal of concern. No board members or executives have made discretionary open-market "
        "purchases, which would be a more bullish signal.\n\n"
        "SOFTBANK OVERHANG: The primary alignment risk is SoftBank's 87% ownership. Masayoshi Son has historically used "
        "SoftBank's stake in portfolio companies to fund Vision Fund investments. Any indication SoftBank might sell a "
        "significant stake at below-market prices would be materially negative for minority shareholders.\n\n"
        "OVERALL: Management quality is solid. Compensation structure is reasonably aligned. The SoftBank concentration "
        "is the key governance risk to monitor."
    )
    v.font = _font(size=FONT_SIZE)
    v.fill = _fill(LIGHT_GREEN)
    v.alignment = _align("left", "top", wrap=True)
    v.border = _border()
    for r in range(27, 32):
        ws.row_dimensions[r].height = 28


# ─────────────────────────────────────────────────────────────────────────────
# TAB 9 — RISKS
# ─────────────────────────────────────────────────────────────────────────────
def build_risks(wb):
    ws = wb.create_sheet("Risks")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, [2, 28, 12, 12, 28, 14, 2])

    _section_hdr(ws, 1, 2, "Arm Holdings – Risk Analysis", span=5)

    hdrs = ["Risk Factor", "Probability", "Impact", "Mitigation", "Net Risk"]
    for i, h in enumerate(hdrs):
        _hdr_cell(ws, 3, 2 + i, h, bg=DARK_NAVY)

    color_map = {
        "HIGH": RED, "MEDIUM": ACCENT_GOLD, "LOW": GREEN,
        "LOW-MED": "92D050", "MED-HIGH": ARM_ORANGE, "N/A": MID_GREY
    }

    def risk_row(ws, row, factor, prob, impact, mitigation, net_risk, bg=WHITE):
        _data_cell(ws, row, 2, factor, bg=bg, bold=True, align="left", wrap=True)
        for col, rating in [(3, prob), (4, impact), (6, net_risk)]:
            c = ws.cell(row, col, rating)
            c.font = _font(bold=True, colour=WHITE, size=FONT_SIZE - 1)
            c.fill = _fill(color_map.get(rating, GREY))
            c.alignment = _align("center", "center")
            c.border = _border()
        _data_cell(ws, row, 5, mitigation, bg=bg, align="left", wrap=True)
        ws.row_dimensions[row].height = 40

    # Financial Risks
    _hdr_cell(ws, 4, 2, "FINANCIAL RISKS", bg=MID_GREY, fg=WHITE, span=5)
    fin_risks = [
        ("Premium Valuation / Multiple Compression",  "HIGH",   "HIGH",   "Strong revenue growth cadence required to justify ~336× GAAP P/E. Any guidance miss or macro slowdown could trigger 20–40% price correction.", "HIGH"),
        ("FCF Volatility (Working Capital)",           "MEDIUM", "MEDIUM", "FY2025 FCF was –$36M due to TLA timing. Management expects normalization in FY2026. Monitor CFO trends in quarterly reports.", "MEDIUM"),
        ("SBC Dilution",                              "MEDIUM", "MEDIUM", "SBC was ~18% of FY2025 revenue. Share count grows if buybacks are not implemented. Monitor diluted share count trend.", "LOW-MED"),
    ]
    for i, (f, p, imp, mit, nr) in enumerate(fin_risks):
        risk_row(ws, 5 + i, f, p, imp, mit, nr, bg=GREY if i % 2 == 0 else WHITE)

    # Operational Risks
    _hdr_cell(ws, 9, 2, "OPERATIONAL / COMPETITIVE RISKS", bg=MID_GREY, fg=WHITE, span=5)
    op_risks = [
        ("RISC-V ISA Disruption",                     "MEDIUM", "HIGH",   "RISC-V is growing in IoT and microcontrollers. Qualcomm's acquisition of Ventana Micro signals intent to develop high-performance RISC-V. Timeline to viability in smartphone SoCs: 3–5 years minimum.", "MED-HIGH"),
        ("Qualcomm Licensing Dispute",                "MEDIUM", "HIGH",   "Ongoing litigation between Arm and Qualcomm over Nuvia acquisition licenses. An adverse ruling could reduce Qualcomm licensing fees. Qualcomm is ~20% of Arm royalty revenue.", "MEDIUM"),
        ("Customer Concentration (Apple, Qualcomm)",  "LOW",    "HIGH",   "Apple and Qualcomm together represent a significant portion of royalties. Loss of either would be material. Mitigated by multi-year contracts and lack of alternatives.", "LOW-MED"),
        ("SoftBank Governance / Overhang",            "MEDIUM", "MEDIUM", "SoftBank controls 87% of votes. Could sell stake at discount, diluting minority shareholders. Vision Fund liquidity needs could accelerate secondary offerings.", "MEDIUM"),
    ]
    for i, (f, p, imp, mit, nr) in enumerate(op_risks):
        risk_row(ws, 10 + i, f, p, imp, mit, nr, bg=GREY if i % 2 == 0 else WHITE)

    # Macro / Market Risks
    _hdr_cell(ws, 15, 2, "MACRO / MARKET RISKS", bg=MID_GREY, fg=WHITE, span=5)
    macro_risks = [
        ("Semiconductor Cycle Downturn",              "MEDIUM", "HIGH",   "Royalties lag chip shipments by one quarter. A downturn in semiconductor demand (smartphones, consumer electronics) directly reduces royalty revenue.", "MEDIUM"),
        ("Tariff / Trade Policy Uncertainty",         "HIGH",   "MEDIUM", "Management withheld FY2026 full-year guidance citing tariff-related uncertainty (April 2026). US-China trade friction affects both Arm's China revenue and customer supply chains.", "MED-HIGH"),
        ("AI Capex Cycle Pause",                     "LOW",    "HIGH",   "Data-center royalties doubled YoY and are priced into valuation. Any pause in hyperscaler AI infrastructure spending would reduce Arm's fastest-growing revenue stream.", "LOW-MED"),
    ]
    for i, (f, p, imp, mit, nr) in enumerate(macro_risks):
        risk_row(ws, 16 + i, f, p, imp, mit, nr, bg=GREY if i % 2 == 0 else WHITE)

    # Regulatory Risks
    _hdr_cell(ws, 20, 2, "REGULATORY RISKS", bg=MID_GREY, fg=WHITE, span=5)
    reg_risks = [
        ("China Export Controls / Revenue Restriction", "MEDIUM", "HIGH",   "~23% of royalties from China. US export controls restrict advanced Arm IP (e.g., v9 AI extensions) to certain Chinese entities. If controls expand, China royalties could decline materially.", "MEDIUM"),
        ("IP / Antitrust Regulation",                   "LOW",    "MEDIUM", "Arm's architectural dominance (99% smartphone market share) could attract regulatory scrutiny. An antitrust ruling requiring royalty caps would structurally impair Arm's revenue model.", "LOW"),
    ]
    for i, (f, p, imp, mit, nr) in enumerate(reg_risks):
        risk_row(ws, 21 + i, f, p, imp, mit, nr, bg=GREY if i % 2 == 0 else WHITE)

    # Overall risk verdict
    _section_hdr(ws, 24, 2, "Overall Risk Verdict", span=5)
    ws.merge_cells("B25:F28")
    v = ws["B25"]
    v.value = (
        "MEDIUM-HIGH RISK PROFILE — primarily driven by valuation, not business quality.\n\n"
        "The BUSINESS risk is LOW: Arm has a near-monopoly in mobile CPU IP, secular tailwinds from AI/data center, "
        "and a widening competitive moat via Armv9 royalty rates and CSS.\n\n"
        "The VALUATION risk is HIGH: At GAAP P/E ~336× and EV/EBITDA ~107×, there is virtually no margin of safety. "
        "Any disappointment in revenue growth, royalty rate trajectory, or macro environment could trigger severe multiple compression.\n\n"
        "The GEOPOLITICAL risk is MEDIUM: China (~23% of royalties) faces increasing export control headwinds. "
        "RISC-V is a long-dated but real existential risk to the royalty model in lower-end markets.\n\n"
        "CONCLUSION: The business deserves a premium multiple. The current multiple demands near-perfection. "
        "Risk-adjusted return is more attractive at $160–175 vs. $205."
    )
    v.font = _font(size=FONT_SIZE)
    v.fill = _fill(LIGHT_BLUE)
    v.alignment = _align("left", "top", wrap=True)
    v.border = _border()
    for r in range(25, 29):
        ws.row_dimensions[r].height = 30


# ─────────────────────────────────────────────────────────────────────────────
# TAB 10 — VALUATION
# ─────────────────────────────────────────────────────────────────────────────
def build_valuation(wb):
    ws = wb.create_sheet("Valuation")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, [2, 30, 16, 16, 22, 2])

    _section_hdr(ws, 1, 2, "Arm Holdings – Valuation Analysis (as of April 27, 2026)", span=4)

    # Market snapshot
    _hdr_cell(ws, 3, 2, "CURRENT MARKET SNAPSHOT", bg=ARM_BLUE, span=4)
    snap_hdrs = ["Metric", "Value", "Note"]
    for i, h in enumerate(snap_hdrs):
        _hdr_cell(ws, 4, 2 + i if i < 2 else 4, h, bg=DARK_NAVY, span=1 if i < 2 else 2)
    ws.merge_cells("D4:E4")

    snap_data = [
        ("Stock Price (April 27, 2026)",  "~$205",     "Near 52-week high; up ~45% from Dec 2025 low of ~$141"),
        ("Market Capitalisation",         "~$217B",    "1.06B diluted shares × $205; SoftBank owns 87% (~$189B of mkt cap)"),
        ("Enterprise Value",              "~$215B",    "Mkt Cap - Net Cash ($2.4B); net cash = $2.807B cash - $437M debt"),
        ("P/E (GAAP, trailing FY2025)",   "~336×",     "Net income $653M GAAP; GAAP is severely impacted by SBC (~$740M)"),
        ("P/E (Non-GAAP, trailing)",      "~70×",      "Adjusted for SBC & acquired intangible amortization; adj. EPS ~$1.56"),
        ("EV / Adj. EBITDA (trailing)",   "~107×",     "Adj. EBITDA ~$1.8B (41% non-GAAP op margin + D&A)"),
        ("EV / Revenue (trailing)",       "~54×",      "FY2025 revenue $4.0B; premium reflects IP royalty model quality"),
        ("Price / Book",                  "~27.8×",    "Book value $7.8B; tangible book $5.9B → P/tangible book ~36.7×"),
        ("Dividend Yield",                "0%",        "No dividend; no buyback program. Growth-phase capital allocation."),
    ]
    for i, (metric, val, note) in enumerate(snap_data):
        bg = GREY if i % 2 == 0 else WHITE
        _data_cell(ws, 5 + i, 2, metric, bg=bg, bold=True, align="left")
        _data_cell(ws, 5 + i, 3, val, bg=bg, align="right", bold=True)
        ws.merge_cells(start_row=5 + i, start_column=4, end_row=5 + i, end_column=5)
        c = ws.cell(5 + i, 4, note)
        c.font = _font(size=FONT_SIZE)
        c.fill = _fill(bg)
        c.alignment = _align("left", "center", wrap=True)
        c.border = _border()
        ws.row_dimensions[5 + i].height = 28

    # Peer comparison
    _section_hdr(ws, 15, 2, "Peer Valuation Comparison", span=4)
    peer_hdrs = ["Company / Ticker", "P/E (Trailing)", "EV/EBITDA", "Revenue / AUM", "Comment"]
    for i, h in enumerate(peer_hdrs):
        _hdr_cell(ws, 16, 2 + i, h, bg=DARK_NAVY)

    peer_data = [
        ("ARM Holdings (ARM)",            "~336× (GAAP)",  "~107×",  "$4.0B FY2025",  "Subject company; GAAP distorted by SBC; non-GAAP P/E ~70×"),
        ("NVIDIA (NVDA)",                 "~42×",          "~35×",   "$130B TTM",     "GPU/AI dominance; similar AI premium but far larger revenue base"),
        ("Broadcom (AVGO)",               "~80×",          "~28×",   "$52B TTM",     "Semiconductor + software; high margins; different model (chips + IP)"),
        ("Marvell Technology (MRVL)",     "~52×",          "~27×",   "$7.5B TTM",    "Custom AI ASICs; high growth but capital-intensive vs. ARM"),
        ("Qualcomm (QCOM)",              "~26×",          "~10×",   "$40B TTM",     "Cheapest of peers; RISC-V hedge; ARM litigation risk"),
        ("Cadence Design Systems (CDNS)", "~55×",          "~45×",   "$4.6B TTM",    "EDA software monopoly; closest analog to Arm's IP model; compelling comp"),
        ("Synopsys (SNPS)",              "~38×",          "~32×",   "$6.1B TTM",    "EDA/IP; merger with Ansys; strong IP analog"),
        ("Peer Median (excl. ARM)",       "~47×",          "~31×",   "—",            "Arm trades at ~2× premium to semiconductor IP peers on P/E and EV/EBITDA"),
    ]
    for i, row in enumerate(peer_data):
        is_arm = i == 0
        is_median = i == 7
        bg = LIGHT_BLUE if is_arm else (LIGHT_GREEN if is_median else (GREY if i % 2 == 0 else WHITE))
        bold = is_arm or is_median
        for j, val in enumerate(row):
            _data_cell(ws, 17 + i, 2 + j, val, bg=bg, bold=bold,
                       align="left" if j in (0, 4) else "center", wrap=(j == 4))
        ws.row_dimensions[17 + i].height = 28

    # DCF scenario analysis
    _section_hdr(ws, 26, 2, "DCF / Scenario Analysis (5-Year Revenue CAGR Approach)", span=4)
    dcf_hdrs = ["Scenario", "Rev. CAGR (5yr)", "Non-GAAP Margin", "EV/Revenue Exit", "Implied Price"]
    for i, h in enumerate(dcf_hdrs):
        _hdr_cell(ws, 27, 2 + i, h, bg=DARK_NAVY)

    dcf_data = [
        ("Bear Case",  "+12% CAGR",  "39%",  "25×", "~$80 (–61% from $205)"),
        ("Base Case",  "+22% CAGR",  "43%",  "35×", "~$165 (–20% from $205)"),
        ("Bull Case",  "+32% CAGR",  "48%",  "50×", "~$310 (+51% from $205)"),
        ("Current price", "$205",    "",     "",    "Already pricing in Bull-leaning Base case"),
    ]
    dcf_colors = [LIGHT_RED, LIGHT_BLUE, LIGHT_GREEN, GREY]
    for i, row in enumerate(dcf_data):
        bg = dcf_colors[i]
        bold = i == 3
        for j, val in enumerate(row):
            _data_cell(ws, 28 + i, 2 + j, val, bg=bg, bold=bold, align="center" if j > 0 else "left")
        ws.row_dimensions[28 + i].height = 26

    # Margin of safety
    _section_hdr(ws, 33, 2, "MARGIN OF SAFETY ASSESSMENT", span=4)
    ws.merge_cells("B34:E38")
    v = ws["B34"]
    v.value = (
        "VERDICT: NO MARGIN OF SAFETY AT CURRENT PRICE — HOLD / ACCUMULATE BELOW $175\n\n"
        "1. At $205, Arm is trading at ~2× the peer median EV/EBITDA and ~7× the revenue of FY2025. "
        "The stock is pricing in the Base-to-Bull scenario with essentially zero downside buffer.\n\n"
        "2. The business quality justifies a premium over peers — Arm's 96%+ gross margins and structural "
        "IP monopoly in mobile are genuinely rare. A fair valuation premium of 50% over semiconductor peers "
        "would imply EV/EBITDA of ~45× → fair value ~$150–175.\n\n"
        "3. Catalysts that could justify $200+: (a) FY2026 royalty growth exceeds 30%, (b) CSS/hyperscaler "
        "contracts accelerate materially, (c) AGI CPU architecture gains mainstream traction, "
        "(d) non-GAAP EPS exits FY2026 at $2.50+.\n\n"
        "4. ENTRY POINT: At $160–175, downside is protected by the $150 bear case with ~77% probability. "
        "At $205, risk-reward is asymmetric to the downside.\n\n"
        "RATING: HOLD at $205. ACCUMULATE at $160–175. TARGET: $195 (12-month); $280 (3-year bull case)."
    )
    v.font = _font(size=FONT_SIZE)
    v.fill = _fill(LIGHT_GREEN)
    v.alignment = _align("left", "top", wrap=True)
    v.border = _border()
    for r in range(34, 39):
        ws.row_dimensions[r].height = 32


# ─────────────────────────────────────────────────────────────────────────────
# TAB 11 — MARKET SENTIMENT
# ─────────────────────────────────────────────────────────────────────────────
def build_sentiment(wb):
    ws = wb.create_sheet("Market Sentiment")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, [2, 28, 16, 16, 30, 2])

    _section_hdr(ws, 1, 2, "Arm Holdings – Market Sentiment & Analyst Coverage", span=4)

    # Analyst consensus
    _hdr_cell(ws, 3, 2, "ANALYST CONSENSUS SUMMARY", bg=ARM_BLUE, span=4)
    cons_hdrs = ["Rating", "Count", "% of Coverage", "Detail"]
    for i, h in enumerate(cons_hdrs):
        _hdr_cell(ws, 4, 2 + i, h, bg=DARK_NAVY)

    cons_data = [
        ("Strong Buy",  "11",  "39%",  "Bullish on Armv9 royalty expansion and data-center adoption momentum"),
        ("Buy",         "12",  "43%",  "Constructive on AI/server tailwinds; note premium valuation risk"),
        ("Hold",        "4",   "13%",  "Neutral; valuation leaves limited upside from current levels"),
        ("Sell",        "1",   "4%",   "One analyst (BofA) set $120 PT in January 2026 citing valuation"),
        ("BUY TOTAL",   "23",  "82%",  "Strong overall buy consensus; reflects Arm's dominant market position"),
    ]
    for i, (rating, count, pct, detail) in enumerate(cons_data):
        is_total = i == 4
        bg = LIGHT_GREEN if is_total else (GREY if i % 2 == 0 else WHITE)
        c_bg = {"Strong Buy": GREEN, "Buy": ARM_BLUE, "Hold": ACCENT_GOLD, "Sell": RED, "BUY TOTAL": GREEN}
        rc = ws.cell(5 + i, 2, rating)
        rc.font = _font(bold=True, colour=WHITE, size=FONT_SIZE - 1)
        rc.fill = _fill(c_bg.get(rating, DARK_NAVY))
        rc.alignment = _align("center", "center")
        rc.border = _border()
        _data_cell(ws, 5 + i, 3, count, bg=bg, align="center", bold=is_total)
        _data_cell(ws, 5 + i, 4, pct, bg=bg, align="center", bold=is_total)
        ws.merge_cells(start_row=5 + i, start_column=5, end_row=5 + i, end_column=5)
        c = ws.cell(5 + i, 5, detail)
        c.font = _font(size=FONT_SIZE, bold=is_total)
        c.fill = _fill(bg)
        c.alignment = _align("left", "center")
        c.border = _border()
        ws.row_dimensions[5 + i].height = 24

    _hdr_cell(ws, 11, 2, "Avg Price Target: ~$180  |  High: $240 (Guggenheim, Mar 2026)  |  Low: $120 (BofA, Jan 2026)",
              bg=DARK_NAVY, span=4, fg=WHITE)

    # Analyst actions
    _section_hdr(ws, 13, 2, "Recent Analyst Actions", span=4)
    action_hdrs = ["Firm", "Rating", "Price Target", "Commentary"]
    for i, h in enumerate(action_hdrs):
        _hdr_cell(ws, 14, 2 + i, h, bg=DARK_NAVY)

    analyst_actions = [
        ("Guggenheim",        "BUY",      "$240",   "Most bullish on Wall St. AGI CPU roadmap + CSS hyperscaler wins are the thesis. March 2026 PT raise."),
        ("Susquehanna",       "BUY",      "$210",   "Most recent action (April 16, 2026). Sees data-center royalty momentum sustaining. AT $210 slightly above current."),
        ("Morgan Stanley",    "BUY",      "$200",   "Constructive on Armv9 royalty rate expansion; bullish on server market penetration."),
        ("Barclays",          "BUY",      "$185",   "Sees strong licensing pipeline; notes RISC-V risk as 'real but overblown near-term'."),
        ("J.P. Morgan",       "BUY",      "$180",   "Consensus avg; solid franchise; recommends buying on dips below $170."),
        ("Deutsche Bank",     "HOLD",     "$165",   "Valuation concern overrides business quality; would turn constructive below $150."),
        ("Goldman Sachs",     "HOLD",     "$155",   "Notes FCF deterioration in FY2025; awaiting FY2026 FCF recovery proof."),
        ("BofA Securities",   "UNDERPERFORM","$120","Most bearish; flags RISC-V existential risk and GAAP P/E over 300×. Set Jan 13, 2026."),
    ]
    for i, (firm, rating, pt, comment) in enumerate(analyst_actions):
        bg = GREY if i % 2 == 0 else WHITE
        _data_cell(ws, 15 + i, 2, firm, bg=bg, bold=True, align="left")
        r_colors = {"BUY": GREEN, "HOLD": ACCENT_GOLD, "UNDERPERFORM": RED, "SELL": RED}
        rc = ws.cell(15 + i, 3, rating)
        rc.font = _font(bold=True, colour=WHITE, size=FONT_SIZE - 1)
        rc.fill = _fill(r_colors.get(rating, DARK_NAVY))
        rc.alignment = _align("center", "center")
        rc.border = _border()
        _data_cell(ws, 15 + i, 4, pt, bg=bg, align="center", bold=True)
        _data_cell(ws, 15 + i, 5, comment, bg=bg, align="left", wrap=True)
        ws.row_dimensions[15 + i].height = 32

    # Market themes
    _section_hdr(ws, 24, 2, "Market Themes & Catalysts", span=4)
    theme_hdrs = ["Theme", "Type", "Explanation"]
    for i, h in enumerate(theme_hdrs):
        col = 2 + i if i < 2 else 4
        span = 1 if i < 2 else 2
        _hdr_cell(ws, 25, col, h, bg=DARK_NAVY, span=span)

    themes = [
        ("Armv9 Royalty Rate Expansion", "TAILWIND",  "Armv9 chips pay ~2× the royalty rate. As the device installed base upgrades, average royalties/chip compound upward even without volume growth. Armv9 >50% of royalties in Q3 FY2026."),
        ("AI Data Center CPU Adoption",  "TAILWIND",  "AWS Graviton4, Google Axion, Microsoft Cobalt, NVIDIA Grace all run on Arm Neoverse. Data-center royalties doubled YoY. Custom silicon preferred over x86 for AI inference workloads on efficiency grounds."),
        ("AGI CPU Strategy (Mar 2026)",  "CATALYST",  "ARM announced its first in-house production silicon ('Arm AGI CPU') on March 25, 2026. Stock surged 16% in one day. This is a strategic pivot from pure IP licensor to active chip competitor — a new revenue layer."),
        ("Tariff / Trade War Uncertainty","HEADWIND",  "Management withdrew FY2026 guidance citing tariff-related macro uncertainty. US-China trade frictions affect both Arm's China revenue and customer supply chains. Adds near-term earnings variability."),
        ("RISC-V Open-Source Threat",    "HEADWIND",  "Qualcomm acquired Ventana Micro (RISC-V startup) in 2025. RISC-V has achieved 25% market penetration in IoT/microcontrollers. Direct royalty-free alternative long-term; immature in high-performance applications today."),
        ("Semiconductor Cycle Recovery", "TAILWIND",  "After a digestion period in 2023–2024, smartphone and server chip demand is recovering. ARM's royalty revenue is a direct derivative of semiconductor unit volumes."),
        ("Qualcomm Litigation Resolution","CATALYST",  "Resolution of Arm-Qualcomm license dispute (over Nuvia acquisition) could unlock or lock in a significant royalty stream. Qualcomm is ~20% of Arm's royalties. Outcome expected 2026."),
        ("China Regulatory Headwinds",   "HEADWIND",  "Huawei and other Chinese entities face restrictions on advanced Arm IP. If restrictions expand to Armv9 AI extensions, China (~23% of royalties) revenue could be structurally impaired."),
    ]
    for i, (theme, type_, detail) in enumerate(themes):
        bg = GREY if i % 2 == 0 else WHITE
        _data_cell(ws, 26 + i, 2, theme, bg=bg, bold=True, align="left")
        t_colors = {"TAILWIND": GREEN, "HEADWIND": RED, "CATALYST": ARM_BLUE}
        tc = ws.cell(26 + i, 3, type_)
        tc.font = _font(bold=True, colour=WHITE, size=FONT_SIZE - 1)
        tc.fill = _fill(t_colors.get(type_, DARK_NAVY))
        tc.alignment = _align("center", "center")
        tc.border = _border()
        ws.merge_cells(start_row=26 + i, start_column=4, end_row=26 + i, end_column=5)
        c = ws.cell(26 + i, 4, detail)
        c.font = _font(size=FONT_SIZE)
        c.fill = _fill(bg)
        c.alignment = _align("left", "center", wrap=True)
        c.border = _border()
        ws.row_dimensions[26 + i].height = 36

    # Ownership
    _section_hdr(ws, 35, 2, "Ownership Structure & Short Interest", span=4)
    own_hdrs = ["Category", "% Ownership", "Notes"]
    for i, h in enumerate(own_hdrs):
        col = 2 + i if i < 2 else 4
        span = 1 if i < 2 else 2
        _hdr_cell(ws, 36, col, h, bg=DARK_NAVY, span=span)

    own_data = [
        ("SoftBank Group Corp.",       "~87%",  "Controlling shareholder; acquired Arm for $32B in 2016. Holds stake via SB Cayman entities. No announced plan to reduce further. Key governance risk."),
        ("Public Float / Institutions","~7.5%", "Very thin public float (~$16B of $217B mkt cap) for such a large-cap stock. Institutional ownership limited by float constraint."),
        ("Management / Insiders",      "~0.5%", "Rene Haas holds ~283K shares ($58M at $205). Modest relative to company size; primary comp is RSU-based."),
        ("Short Interest (% of float)","~10–11%", "~16M shares short. High short interest as % of float reflects both valuation skeptics and short-sellers hedging volatility strategies."),
        ("Days to Cover",              "~4–5 days","Based on average daily volume. Moderate squeeze risk if positive catalysts emerge."),
    ]
    for i, (cat, pct, notes) in enumerate(own_data):
        bg = GREY if i % 2 == 0 else WHITE
        _data_cell(ws, 37 + i, 2, cat, bg=bg, bold=True, align="left")
        _data_cell(ws, 37 + i, 3, pct, bg=bg, align="center")
        ws.merge_cells(start_row=37 + i, start_column=4, end_row=37 + i, end_column=5)
        c = ws.cell(37 + i, 4, notes)
        c.font = _font(size=FONT_SIZE)
        c.fill = _fill(bg)
        c.alignment = _align("left", "center", wrap=True)
        c.border = _border()
        ws.row_dimensions[37 + i].height = 32


# ─────────────────────────────────────────────────────────────────────────────
# TAB 12 — KEY INDICATORS
# ─────────────────────────────────────────────────────────────────────────────
def build_key_indicators(wb):
    ws = wb.create_sheet("Key Indicators")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, [2, 32, 16, 16, 16, 16, 2])

    _section_hdr(ws, 1, 2, "Arm Holdings – Key Performance Indicators & Scorecard", span=5)

    # Financial scorecard
    _hdr_cell(ws, 3, 2, "Financial KPI Scorecard", bg=ARM_BLUE, span=5)
    kpi_hdrs = ["KPI", "FY2023", "FY2024", "FY2025", "Trend / Signal"]
    for i, h in enumerate(kpi_hdrs):
        _hdr_cell(ws, 4, 2 + i, h, bg=DARK_NAVY)

    signal_colors = {"POSITIVE": GREEN, "NEUTRAL": ACCENT_GOLD, "WATCH": ARM_ORANGE, "NEGATIVE": RED}

    kpi_data = [
        ("Total Revenue ($M)",             "$2,679",  "$3,233",  "$4,007",  "POSITIVE"),
        ("YoY Revenue Growth (%)",         "–0.9%",   "+20.7%",  "+24.0%",  "POSITIVE"),
        ("Royalty Revenue ($M)",           "~$1,682", "$1,819",  "$2,200",  "POSITIVE"),
        ("Royalty YoY Growth (%)",         "~+8%",    "+8.1%",   "+21.0%",  "POSITIVE"),
        ("Licensing Revenue ($M)",         "~$997",   "$1,414",  "$1,807",  "POSITIVE"),
        ("Gross Margin (%)",               "96.0%",   "95.2%",   "96.9%",   "POSITIVE"),
        ("Non-GAAP Operating Margin (%)",  "~40%",    "~39%",    "~41%",    "POSITIVE"),
        ("GAAP Net Income ($M)",           "$524",    "$306",    "$653",    "POSITIVE"),
        ("GAAP EPS (diluted)",             "$0.51",   "$0.29",   "$0.61",   "POSITIVE"),
        ("Adj. EPS (non-GAAP est.)",       "~$1.11",  "~$1.24",  "~$1.56",  "POSITIVE"),
        ("Free Cash Flow ($M)",            "$675",    "$752",    "–$36",    "WATCH"),
        ("Cash & Equivalents ($M)",        "$1,554",  "$1,925",  "$2,807",  "POSITIVE"),
        ("Armv9 % of Royalty Revenue",     "~10%",    "~25%",    ">50% (Q3FY26)", "POSITIVE"),
        ("SBC as % of Revenue",            "~9%",     "~28%",    "~18%",    "NEUTRAL"),
        ("Shares Outstanding (diluted, B)","~1.02",   "~1.04",   "~1.06",   "NEUTRAL"),
    ]
    for i, (kpi, y23, y24, y25, signal) in enumerate(kpi_data):
        bg = GREY if i % 2 == 0 else WHITE
        _data_cell(ws, 5 + i, 2, kpi, bg=bg, bold=(signal == "POSITIVE"), align="left")
        _data_cell(ws, 5 + i, 3, y23, bg=bg, align="right")
        _data_cell(ws, 5 + i, 4, y24, bg=bg, align="right")
        _data_cell(ws, 5 + i, 5, y25, bg=bg, align="right")
        sc = ws.cell(5 + i, 6, signal)
        sc.font = _font(bold=True, colour=WHITE, size=FONT_SIZE - 1)
        sc.fill = _fill(signal_colors.get(signal, GREY))
        sc.alignment = _align("center", "center")
        sc.border = _border()
        ws.row_dimensions[5 + i].height = 22

    # Operating metrics & guidance
    _section_hdr(ws, 21, 2, "Operating Metrics & Management Targets", span=5)
    op_hdrs = ["Metric", "Current / FY2025", "FY2026E Target", "3–5 Year Target", "Source"]
    for i, h in enumerate(op_hdrs):
        _hdr_cell(ws, 22, 2 + i, h, bg=DARK_NAVY)

    targets = [
        ("Annual Revenue",              "$4.0B (FY2025)",  "~$4.4B (mgmt Q4 guidance)",        ">$8B by FY2030 (analyst est.)",      "Arm Q3 FY2026 earnings release"),
        ("Royalty Revenue",             "$2.2B (FY2025)",  "~$2.6B (FY2026E est.)",             "$4–5B long term",                    "Armv9 upgrade cycle extrapolation"),
        ("Royalty Rate per Chip",       "~2× Armv8 for v9","Rising; CSS adds incremental value", "2.5–3× Armv8 as v9 matures",        "Management commentary"),
        ("Non-GAAP Op. Margin",         "~41% (FY2025)",   "~42–44% (FY2026E)",                 "45%+ long term",                     "Management guidance; analyst consensus"),
        ("Armv9 % of Royalties",        ">50% (Q3 FY2026)","~60% (FY2026E)",                    "~80–90% by FY2028",                  "Shipment mix data"),
        ("Data Center Royalty Growth",  "+100% YoY (FY26E)","Continued >25% growth expected",   "DC becomes 35%+ of royalties",       "Neoverse adoption trajectory"),
        ("FCF Margin",                  "N/M (–0.9%)",     "~18–20% (normalized FY2026E)",      "25–30% long-term target",            "Management normalized FCF guidance"),
        ("Chips Shipped/Year (ARM-IP)", "~30–35B",         "~35–40B (FY2026E est.)",            ">50B by FY2028",                     "Industry data; Arm presentations"),
    ]
    for i, row in enumerate(targets):
        bg = GREY if i % 2 == 0 else WHITE
        for j, val in enumerate(row):
            _data_cell(ws, 23 + i, 2 + j, val, bg=bg, align="left" if j in (0, 4) else "center", wrap=True)
        ws.row_dimensions[23 + i].height = 30

    # Investment summary
    _section_hdr(ws, 32, 2, "INVESTMENT SUMMARY & FINAL RECOMMENDATION", span=5)
    ws.merge_cells("B33:F42")
    v = ws["B33"]
    v.value = (
        "FINAL RECOMMENDATION: HOLD / ACCUMULATE ON WEAKNESS\n"
        "Rating: HOLD at $205  |  Target: $195 (12-month)  |  Accumulate: $160–175  |  Bull Target: $280 (3-year)\n\n"
        "INVESTMENT THESIS:\n"
        "• Arm is one of the most defensible technology businesses in the world — a structural IP monopoly in mobile "
        "computing with a deepening presence in data centers and AI infrastructure.\n"
        "• The Armv9 upgrade cycle provides a clear, multi-year royalty rate expansion mechanism that is independent "
        "of chip volume growth — this is a genuine pricing power story, not just a market share story.\n"
        "• Data-center royalties doubling YoY (FY2026) validate the thesis that Arm is the CPU architecture "
        "of choice for hyperscaler AI infrastructure.\n"
        "• The AGI CPU announcement (March 2026) opens a new direct chip revenue stream that is not yet priced.\n\n"
        "VALUATION VERDICT:\n"
        "• At $205, the stock prices in near-perfection (Base-to-Bull scenario). GAAP P/E ~336×, EV/EBITDA ~107×.\n"
        "• There is no margin of safety at this price. A business this good deserves a premium, not unlimited premium.\n"
        "• Fair value range: $165–195 (Base case). Entry zone: $160–175.\n\n"
        "KEY RISKS TO WATCH:\n"
        "• RISC-V penetration in high-performance SoCs (2–4 year horizon)\n"
        "• Qualcomm litigation outcome\n"
        "• China revenue restriction expansion\n"
        "• SoftBank secondary offering creating price pressure\n\n"
        "CATALYSTS (UPSIDE):\n"
        "• Armv9 exceeds 70% of royalties by FY2027\n"
        "• AGI CPU gains design wins at major hyperscalers\n"
        "• Qualcomm litigation settles favorably\n"
        "• FY2026 FCF recovers to $1B+, validating business model\n"
    )
    v.font = _font(size=FONT_SIZE)
    v.fill = _fill(LIGHT_GREEN)
    v.alignment = _align("left", "top", wrap=True)
    v.border = _border()
    for r in range(33, 43):
        ws.row_dimensions[r].height = 28

    ws.row_dimensions[32].height = 28


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────
def main():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    build_cover(wb)
    build_business(wb)
    build_moat(wb)
    build_income(wb)
    build_balance_sheet(wb)
    build_cashflow(wb)
    build_roc(wb)
    build_management(wb)
    build_risks(wb)
    build_valuation(wb)
    build_sentiment(wb)
    build_key_indicators(wb)

    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"Saved: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
