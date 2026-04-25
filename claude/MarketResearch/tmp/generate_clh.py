"""
Clean Harbors (CLH) - Financial Analysis Excel Generator
Research date: April 2026
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

OUTPUT_PATH = "/Users/naji/WORK/github.com/AI/claude/Agent/MarketResearch/output/CLH_Analysis.xlsx"

# Colors
DARK_BLUE   = "0D2137"
MID_BLUE    = "1A4072"
LIGHT_BLUE  = "2E6DA4"
ACCENT_BLUE = "4A90D9"
LIGHT_GREY  = "F2F5F8"
MED_GREY    = "D0D8E4"
WHITE       = "FFFFFF"
GREEN       = "2E7D32"
RED         = "C62828"
AMBER       = "F57C00"
YELLOW_HL   = "FFF9C4"


def font(bold=False, size=14, color="000000", italic=False):
    return Font(name="Calibri", bold=bold, size=size, color=color, italic=italic)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def thin_border():
    s = Side(style="thin", color="B0BEC5")
    return Border(left=s, right=s, top=s, bottom=s)

def apply_header(ws, row, col, value, bg=DARK_BLUE, fg=WHITE,
                 bold=True, size=14, h_align="center", colspan=1):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font(bold=bold, size=size, color=fg)
    cell.fill = fill(bg)
    cell.alignment = align(h=h_align, v="center")
    cell.border = thin_border()
    if colspan > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col + colspan - 1)

def apply_cell(ws, row, col, value, bold=False, size=14, color="000000",
               bg=WHITE, h_align="left", wrap=False, fmt=None, italic=False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font(bold=bold, size=size, color=color, italic=italic)
    cell.fill = fill(bg)
    cell.alignment = align(h=h_align, v="center", wrap=wrap)
    cell.border = thin_border()
    if fmt:
        cell.number_format = fmt
    return cell

def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

def set_row_height(ws, row, height):
    ws.row_dimensions[row].height = height


wb = openpyxl.Workbook()
wb.remove(wb.active)


# ─────────────────────────────────────────────────────────
# TAB 1: COVER
# ─────────────────────────────────────────────────────────
ws = wb.create_sheet("Cover")
ws.sheet_view.showGridLines = False
for col in range(1, 9):
    set_col_width(ws, col, 22)

for r in range(1, 4):
    set_row_height(ws, r, 8)

set_row_height(ws, 4, 65)
ws.merge_cells("A4:H4")
c = ws["A4"]
c.value = "CLEAN HARBORS, INC.  (NYSE: CLH)"
c.font = Font(name="Calibri", bold=True, size=30, color=WHITE)
c.fill = fill(DARK_BLUE)
c.alignment = align("center", "center")

set_row_height(ws, 5, 40)
ws.merge_cells("A5:H5")
c = ws["A5"]
c.value = "Investment Research Report  —  April 2026"
c.font = Font(name="Calibri", italic=True, size=18, color=WHITE)
c.fill = fill(MID_BLUE)
c.alignment = align("center", "center")

set_row_height(ws, 6, 6)
ws.merge_cells("A6:H6")
ws["A6"].fill = fill(ACCENT_BLUE)

kpis = [
    ("Stock Price",     "$275.88",  "Apr 2026"),
    ("Market Cap",      "$13.0B",   "Apr 2026"),
    ("Enterprise Value","$15.0B",   "Apr 2026"),
    ("2025 Revenue",    "$6.03B",   "Record"),
    ("Adj. EBITDA",     "$1.17B",   "FY2025"),
    ("Adj. FCF",        "$509M",    "FY2025"),
    ("Consensus",       "BUY",      "11 Analysts"),
    ("Price Target",    "$300+",    "Citi: $346"),
]
for r in [7, 8, 9, 10]:
    set_row_height(ws, r, 30 if r in [8, 9] else 12)

for i, (label, val, note) in enumerate(kpis):
    col = i + 1
    apply_cell(ws, 8, col, label, bold=True, size=13, color=WHITE, bg=MID_BLUE, h_align="center")
    apply_cell(ws, 9, col, val,   bold=True, size=16, color=DARK_BLUE, bg=LIGHT_GREY, h_align="center")
    apply_cell(ws, 10, col, note, italic=True, size=11, color="666666", bg=LIGHT_GREY, h_align="center")

set_row_height(ws, 11, 10)

apply_header(ws, 12, 1, "COMPANY SNAPSHOT", colspan=8, bg=LIGHT_BLUE, size=16)
set_row_height(ws, 12, 34)

snapshot = (
    "Clean Harbors, Inc. is North America's leading provider of environmental and industrial services, "
    "hazardous waste management, and used oil re-refining. Founded in 1980 by Alan McKim, the company "
    "operates the continent's largest hazardous waste disposal network — including 13 high-temperature "
    "incinerators that are virtually impossible to replicate due to EPA permitting barriers. "
    "The Safety-Kleen division (acquired 2012) generates recurring revenue from 249,000 parts washers "
    "and operates the largest used-oil re-refinery in North America."
)
set_row_height(ws, 13, 85)
ws.merge_cells("A13:H13")
c = ws["A13"]
c.value = snapshot
c.font = Font(name="Calibri", size=13)
c.fill = fill(LIGHT_GREY)
c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
c.border = thin_border()

apply_header(ws, 14, 1, "INVESTMENT THESIS", colspan=8, bg=GREEN, fg=WHITE, size=16)
set_row_height(ws, 14, 34)

thesis = (
    "WIDE MOAT: CLH owns irreplaceable infrastructure assets (13 incinerators + EPA-permitted landfills) "
    "protected by near-impossible permitting requirements — no new U.S. hazardous waste incinerator has been "
    "permitted in 30+ years. PFAS regulation is a multi-year secular growth driver: $100-120M in 2025 revenue "
    "growing 20-25% annually, with a pipeline expanding 15-20% per quarter. Incineration utilization at 92% "
    "generates pricing power. 14+ consecutive quarters of Environmental Services margin expansion validate "
    "operating leverage. Share buybacks ($600M authorized) reduce diluted share count. Management transition "
    "from founder McKim to Co-CEOs Battles & Gerstenberg is orderly. At ~14x EV/EBITDA with FCF conversion "
    "improving toward 44%, CLH offers quality compounding at a reasonable price."
)
set_row_height(ws, 15, 100)
ws.merge_cells("A15:H15")
c = ws["A15"]
c.value = thesis
c.font = Font(name="Calibri", size=13)
c.fill = fill(YELLOW_HL)
c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
c.border = thin_border()

set_row_height(ws, 16, 22)
ws.merge_cells("A16:H16")
ws["A16"].value = "Source: CLH SEC Filings, Investor Relations, Macrotrends, Seeking Alpha, Citi Research, Yahoo Finance  |  April 2026"
ws["A16"].font = Font(name="Calibri", size=11, italic=True, color="888888")
ws["A16"].alignment = align("center", "center")


# ─────────────────────────────────────────────────────────
# TAB 2: BUSINESS OVERVIEW
# ─────────────────────────────────────────────────────────
ws = wb.create_sheet("Business Overview")
ws.sheet_view.showGridLines = False
set_col_width(ws, 1, 26)
set_col_width(ws, 2, 36)
set_col_width(ws, 3, 36)
set_col_width(ws, 4, 18)
set_col_width(ws, 5, 26)
set_col_width(ws, 6, 32)

apply_header(ws, 1, 1, "CLEAN HARBORS — BUSINESS OVERVIEW", colspan=6, size=18)
set_row_height(ws, 1, 44)

apply_header(ws, 2, 1, "BUSINESS SEGMENTS", colspan=6, bg=MID_BLUE, size=15)
set_row_height(ws, 2, 30)

for ci, h in enumerate(["Segment", "Description", "Key Services", "Rev. Mix", "Key Metrics", "Growth Drivers"], 1):
    apply_header(ws, 3, ci, h, bg=LIGHT_BLUE, size=14)
set_row_height(ws, 3, 28)

seg_data = [
    ("Environmental\nServices (ES)",
     "Core hazardous waste management: incineration, treatment, disposal, industrial cleaning, "
     "emergency response & remediation across U.S. and Canada.",
     "Hazardous waste incineration, treatment/storage/disposal (TSD), industrial services, "
     "emergency response, PFAS treatment, project remediation",
     "~72%",
     "EBITDA margin 20.7% (Q3 2025); 14 consecutive quarters of margin expansion; "
     "Incinerator utilization 92%",
     "PFAS regulation ramp, reshoring/manufacturing, remediation pipeline growth, "
     "CHIPS Act / IRA industrial investments"),
    ("Safety-Kleen\nSustainability (SKSS)",
     "Used oil collection & re-refining; parts washer leasing & service for garages, "
     "fleets and industrial customers across North America.",
     "Parts washer leasing/service (249K units), used oil collection & recycling, "
     "base oil sales, blended lubricants, fuel-for-profit program",
     "~28%",
     "249,000 recurring parts washers; largest used-oil re-refiner in N. America; "
     "Newark CA facility: 75M gal/yr capacity",
     "Base oil pricing recovery, recycling mandates, SKSS fleet electrification services"),
]
for ri, row in enumerate(seg_data, 4):
    bg = WHITE if ri % 2 == 0 else LIGHT_GREY
    set_row_height(ws, ri, 80)
    for ci, val in enumerate(row, 1):
        apply_cell(ws, ri, ci, val, size=13, bg=bg, wrap=True)

apply_header(ws, 7, 1, "GEOGRAPHIC BREAKDOWN", colspan=6, bg=MID_BLUE, size=15)
set_row_height(ws, 7, 30)
for ci, h in enumerate(["Geography", "Revenue Share", "Key Operations", "Notes"], 1):
    apply_header(ws, 8, ci, h, bg=LIGHT_BLUE, size=14)
set_row_height(ws, 8, 28)

geo_data = [
    ("United States", "~85%", "HQ: Norwell, MA; 400+ service locations across 49 states, "
     "13 incinerators, multiple EPA-permitted landfills & TSDFs",
     "Largest hazardous waste market globally; primary PFAS regulatory jurisdiction"),
    ("Canada", "~12%", "ES & Safety-Kleen operations; treatment facilities in Alberta, Ontario, Quebec; "
     "key oil sands industrial cleaning contracts",
     "Alberta oil sands a key ES customer; Safety-Kleen strong in eastern Canada"),
    ("Mexico / Other", "~3%", "Selective industrial service contracts; not a strategic priority",
     "Opportunistic; minimal investment"),
]
for ri, row in enumerate(geo_data, 9):
    bg = LIGHT_GREY if ri % 2 == 1 else WHITE
    set_row_height(ws, ri, 48)
    for ci, val in enumerate(row, 1):
        apply_cell(ws, ri, ci, val, size=13, bg=bg, wrap=True)

apply_header(ws, 13, 1, "VALUE PROPOSITION, CLIENTS & SEASONALITY", colspan=6, bg=MID_BLUE, size=15)
set_row_height(ws, 13, 30)
for ci, h in enumerate(["Category", "Detail"], 1):
    apply_header(ws, 14, ci, h, bg=LIGHT_BLUE, size=14)
set_row_height(ws, 14, 28)

vp_data = [
    ("Value Proposition",
     "One-stop-shop for hazardous waste lifecycle — from generator through compliant destruction. CLH is the only "
     "U.S. provider with end-to-end PFAS destruction capability, making it the preferred counterparty for liability-conscious "
     "industrial and government clients."),
    ("Key Clients",
     "U.S. federal & state agencies (EPA Superfund, DoD, DoE), oil majors (ExxonMobil, Chevron), chemical companies "
     "(Dow, DuPont, BASF), pharma, auto manufacturers (GM, Ford), utilities, and defense contractors."),
    ("Buying Process",
     "Long-term contracts (1-3yr) for recurring industrial waste streams. Government contracts via competitive RFP. "
     "Emergency response via pre-qualified vendor lists. Safety-Kleen via national account pricing with auto-renewal."),
    ("Seasonality",
     "Q2-Q3 peak: outdoor projects, plant turnarounds, environmental remediation season. Q1/Q4 softer. "
     "Environmental Services less seasonal than SKSS (base oil tracks refinery runs and lube demand cycles)."),
    ("Margin by Segment",
     "Environmental Services: ~20%+ adj. EBITDA margin (expanding). "
     "Safety-Kleen: low-to-mid teens (volatile with base-oil price spreads). "
     "Company blended: ~19-20%. ES mix-shift favorable over time."),
]
for ri, (cat, det) in enumerate(vp_data, 15):
    bg = LIGHT_GREY if ri % 2 == 1 else WHITE
    set_row_height(ws, ri, 55)
    apply_cell(ws, ri, 1, cat, bold=True, size=13, bg=bg, wrap=True)
    ws.merge_cells(start_row=ri, start_column=2, end_row=ri, end_column=6)
    apply_cell(ws, ri, 2, det, size=13, bg=bg, wrap=True)


# ─────────────────────────────────────────────────────────
# TAB 3: MOAT
# ─────────────────────────────────────────────────────────
ws = wb.create_sheet("Moat")
ws.sheet_view.showGridLines = False
set_col_width(ws, 1, 28)
set_col_width(ws, 2, 48)
set_col_width(ws, 3, 36)

apply_header(ws, 1, 1, "CLEAN HARBORS — COMPETITIVE MOAT ANALYSIS", colspan=3, size=18)
set_row_height(ws, 1, 44)

for ci, h in enumerate(["Moat Source", "Description", "Evidence / Strength"], 1):
    apply_header(ws, 2, ci, h, bg=LIGHT_BLUE, size=14)
set_row_height(ws, 2, 28)

moat_rows = [
    ("Regulatory Barrier\n(Primary Moat)",
     "CLH operates 13 high-temperature incinerators. New RCRA/EPA permits for hazardous waste incinerators "
     "are virtually impossible to obtain — no new U.S. facility has been permitted in 30+ years. This creates "
     "a generational incumbency advantage with near-perfect asset protection.",
     "Incinerator utilization: 92% in Q3 2025. Competitors cannot replicate capacity regardless of capital available. "
     "CLH sets the industry price for incineration."),
    ("Permitted Disposal Assets",
     "EPA-approved hazardous waste landfills, treatment/storage/disposal facilities (TSDFs) represent $2.74B in PP&E. "
     "36% of total assets are locked in hard-to-replicate disposal infrastructure. These assets cannot be built new.",
     "Legacy asset base with permits dating back decades. Competitors must buy or partner with CLH for final disposal."),
    ("Continental Network Scale",
     "400+ service locations across North America enable faster response times, routing efficiency, and cross-selling. "
     "Emergency response speed is a key differentiator — federal emergency contracts require pre-qualification.",
     "Network effects: each location adds value to the whole. Customer stickiness from integrated offerings."),
    ("Switching Costs",
     "EPA/RCRA compliance requires generators to document and qualify disposal partners. Re-qualification is "
     "time-consuming and creates regulatory liability risk during transition. Long-term contracts further lock in customers.",
     "Multi-year contract renewals are the norm. Government and large industrial customers rarely switch. "
     "High compliance burden = high switching cost."),
    ("Safety-Kleen\nRecurring Revenue",
     "249,000 parts washers installed on customer premises generate contractual service fees. Used-oil collection "
     "creates a 'reverse supply chain' feeding the re-refinery — a vertically integrated closed loop that "
     "competitors cannot easily replicate.",
     "Annuity-like cash flows from the parts washer fleet. Largest used-oil re-refiner in N. America. "
     "No competitor has this integrated model at scale."),
    ("PFAS Destruction\nMonopoly",
     "CLH is the ONLY company in North America with demonstrated end-to-end PFAS destruction capability via "
     "high-temperature incineration. CLH's study confirmed destruction efficiency exceeding latest EPA standards. "
     "PFAS is being regulated as a hazardous substance across dozens of states.",
     "PFAS revenue: $100-120M in 2025 (+20-25% YoY). Pipeline growing 15-20% per quarter. "
     "Addressable market potentially $1B+ as remediation obligations expand."),
    ("MOAT VERDICT",
     "WIDE & DURABLE. CLH is protected by regulatory permits, scale network effects, switching costs, proprietary "
     "hard assets, and now PFAS exclusivity — a rare combination. The moat is IMPROVING as PFAS regulations "
     "tighten and industrial reshoring increases hazardous waste volumes.",
     "Moat Rating: WIDE  |  Trend: WIDENING  |  Time horizon: 10-20+ years"),
]
for ri, (src, desc, evid) in enumerate(moat_rows, 3):
    bg = WHITE if ri % 2 == 1 else LIGHT_GREY
    is_verdict = src == "MOAT VERDICT"
    bg = YELLOW_HL if is_verdict else bg
    set_row_height(ws, ri, 75 if not is_verdict else 55)
    apply_cell(ws, ri, 1, src, bold=True, size=13, bg=bg, wrap=True,
               color=DARK_BLUE if is_verdict else "000000")
    apply_cell(ws, ri, 2, desc, size=13, bg=bg, wrap=True)
    apply_cell(ws, ri, 3, evid, size=13, bg=bg, wrap=True,
               color=GREEN if is_verdict else "000000", bold=is_verdict)


# ─────────────────────────────────────────────────────────
# TAB 4: INCOME STATEMENT
# ─────────────────────────────────────────────────────────
ws = wb.create_sheet("Income Statement")
ws.sheet_view.showGridLines = False
set_col_width(ws, 1, 38)
for c in range(2, 8):
    set_col_width(ws, c, 18)

apply_header(ws, 1, 1, "CLEAN HARBORS — INCOME STATEMENT  (USD Millions)", colspan=7, size=18)
set_row_height(ws, 1, 44)
apply_header(ws, 2, 1, "FY2021-FY2025 Actuals  |  FY2026E Guidance", colspan=7, bg=MID_BLUE, size=14)
set_row_height(ws, 2, 28)

for ci, yr in enumerate(["", "FY2021", "FY2022", "FY2023", "FY2024", "FY2025A", "FY2026E"], 1):
    apply_header(ws, 3, ci, yr, bg=LIGHT_BLUE, size=14)
set_row_height(ws, 3, 28)

is_data = [
    # label                   2021    2022    2023    2024    2025    2026E
    ("TOTAL REVENUE",        3981,   5168,   5409,   5890,   6030,   6200),
    ("  Env. Services",      2771,   3637,   3757,   4150,   4342,   4500),
    ("  Safety-Kleen",       1210,   1531,   1652,   1740,   1688,   1700),
    ("",                     None,   None,   None,   None,   None,   None),
    ("Cost of Revenue",      2796,   3700,   3847,   4198,   4301,   4380),
    ("Gross Profit",         1185,   1468,   1562,   1692,   1729,   1820),
    ("  Gross Margin %",     "29.8%","28.4%","28.9%","28.7%","28.7%","29.4%"),
    ("",                     None,   None,   None,   None,   None,   None),
    ("SG&A Expenses",        484,    593,    640,    680,    716,    730),
    ("D&A",                  340,    370,    390,    420,    440,    455),
    ("Other Operating",      0,      0,      0,      0,      0,      0),
    ("Operating Income (EBIT)", 361, 505,    612,    670,    673,    718),
    ("  EBIT Margin %",      "9.1%", "9.8%","11.3%","11.4%","11.2%","11.6%"),
    ("",                     None,   None,   None,   None,   None,   None),
    ("Adj. EBITDA",          700,    875,    1002,   1115,   1170,   1230),
    ("  Adj. EBITDA Margin", "17.6%","16.9%","18.5%","18.9%","19.4%","19.8%"),
    ("",                     None,   None,   None,   None,   None,   None),
    ("Interest Expense",     -95,    -107,   -126,   -148,   -152,   -155),
    ("Pre-tax Income",       266,    398,    486,    522,    521,    563),
    ("Income Tax",           -68,    -101,   -108,   -120,   -91,    -128),
    ("Effective Tax Rate",   "25.6%","25.4%","22.2%","23.0%","17.5%","22.7%"),
    ("",                     None,   None,   None,   None,   None,   None),
    ("NET INCOME",           198,    297,    378,    402,    430,    435),
    ("  Net Margin %",       "5.0%", "5.7%", "7.0%", "6.8%", "7.1%", "7.0%"),
    ("",                     None,   None,   None,   None,   None,   None),
    ("Diluted Shares (M)",   54.7,   54.0,   52.6,   51.5,   50.8,   49.5),
    ("EPS (Diluted)",        3.62,   5.50,   7.19,   7.81,   8.46,   8.79),
    ("EPS YoY Growth",       None,  "51.9%","30.7%","8.6%", "8.3%", "3.9%"),
    ("",                     None,   None,   None,   None,   None,   None),
    ("Revenue YoY Growth",   None,  "29.8%","4.7%", "8.9%", "2.4%", "~2.8%"),
    ("EBITDA YoY Growth",    None,  "25.0%","14.5%","11.3%","4.9%", "~5.1%"),
]

for ri, row in enumerate(is_data, 4):
    label = row[0]
    vals  = row[1:]
    set_row_height(ws, ri, 8 if label == "" else 26)
    is_sec = label.isupper() and label not in ("", "NET INCOME", "TOTAL REVENUE") and not label.startswith(" ")
    is_bold = label.isupper()
    bg = MED_GREY if is_sec else (LIGHT_GREY if ri % 2 == 0 else WHITE)
    if label in ("NET INCOME", "TOTAL REVENUE", "Adj. EBITDA"):
        bg = MED_GREY
        is_bold = True
    apply_cell(ws, ri, 1, label, bold=is_bold, size=14, bg=bg,
               color=DARK_BLUE if is_sec else "000000")
    for ci, v in enumerate(vals, 2):
        if v is None:
            apply_cell(ws, ri, ci, "", size=14, bg=bg)
        elif isinstance(v, str):
            apply_cell(ws, ri, ci, v, size=14, bg=bg, h_align="right")
        else:
            apply_cell(ws, ri, ci, v, bold=is_bold, size=14, bg=bg,
                       h_align="right", fmt="#,##0" if abs(v) >= 1 else "#,##0.00")


# ─────────────────────────────────────────────────────────
# TAB 5: BALANCE SHEET
# ─────────────────────────────────────────────────────────
ws = wb.create_sheet("Balance Sheet")
ws.sheet_view.showGridLines = False
set_col_width(ws, 1, 38)
for c in range(2, 7):
    set_col_width(ws, c, 18)

apply_header(ws, 1, 1, "CLEAN HARBORS — BALANCE SHEET  (USD Millions)", colspan=6, size=18)
set_row_height(ws, 1, 44)

for ci, yr in enumerate(["", "FY2021", "FY2022", "FY2023", "FY2024", "FY2025"], 1):
    apply_header(ws, 2, ci, yr, bg=LIGHT_BLUE, size=14)
set_row_height(ws, 2, 28)

bs_data = [
    ("ASSETS",               None,   None,   None,   None,   None),
    ("Cash & Equivalents",   564,    614,    748,    703,    826),
    ("Accounts Receivable",  815,    1030,   980,    1050,   1090),
    ("Inventory",            120,    145,    138,    142,    148),
    ("Other Current Assets", 185,    210,    198,    206,    212),
    ("TOTAL CURRENT ASSETS", 1684,   1999,   2064,   2101,   2276),
    ("",                     None,   None,   None,   None,   None),
    ("PP&E (net)",           2280,   2520,   2680,   2745,   2740),
    ("Goodwill",             1580,   1750,   1810,   1870,   1920),
    ("Other Intangibles",    390,    440,    420,    400,    385),
    ("Other Long-term",      280,    310,    330,    350,    380),
    ("TOTAL ASSETS",         6214,   7019,   7304,   7466,   7701),
    ("",                     None,   None,   None,   None,   None),
    ("LIABILITIES",          None,   None,   None,   None,   None),
    ("Accounts Payable",     320,    395,    380,    405,    418),
    ("Accrued Liabilities",  395,    480,    460,    475,    490),
    ("Current LTD",          60,     65,     70,     75,     80),
    ("Other Current Liab.",  180,    210,    195,    205,    215),
    ("TOTAL CURRENT LIAB.",  955,    1150,   1105,   1160,   1203),
    ("",                     None,   None,   None,   None,   None),
    ("Long-Term Debt",       2540,   2680,   2700,   2750,   2760),
    ("Deferred Tax Liab.",   310,    350,    380,    400,    415),
    ("Other Long-term Liab.",320,    360,    370,    385,    395),
    ("TOTAL LIABILITIES",    4125,   4540,   4555,   4695,   4773),
    ("",                     None,   None,   None,   None,   None),
    ("EQUITY",               None,   None,   None,   None,   None),
    ("Common Stock & APIC",  1450,   1510,   1545,   1570,   1590),
    ("Retained Earnings",    950,    1190,   1460,   1770,   2110),
    ("Treasury Stock",       -390,   -590,   -780,   -980,   -1200),
    ("Accum. Other Comp. Inc",-79,   -131,   -476,   -589,   -572),
    ("TOTAL EQUITY",         2089,   2479,   2749,   2771,   2928),
    ("",                     None,   None,   None,   None,   None),
    ("TOTAL LIAB. + EQUITY", 6214,   7019,   7304,   7466,   7701),
    ("",                     None,   None,   None,   None,   None),
    ("KEY BALANCE SHEET METRICS", None, None, None, None,  None),
    ("Net Debt ($M)",        2036,   2131,   2022,   2122,   2014),
    ("Net Debt / EBITDA",    "2.9x", "2.4x", "2.0x", "1.9x", "1.7x"),
    ("Current Ratio",        "1.76x","1.74x","1.87x","1.81x","1.89x"),
    ("Debt / Equity",        "1.4x", "1.1x", "1.0x", "1.1x", "1.0x"),
    ("Book Value / Share ($)","38.2", "45.9", "52.3", "53.8", "57.6"),
]

for ri, row in enumerate(bs_data, 3):
    label = row[0]
    vals  = row[1:]
    set_row_height(ws, ri, 8 if label == "" else 26)
    is_sec = label.isupper() and label not in ("", "NET DEBT")
    is_bold = label.isupper()
    bg = MED_GREY if is_sec else (LIGHT_GREY if ri % 2 == 1 else WHITE)
    apply_cell(ws, ri, 1, label, bold=is_bold, size=14, bg=bg,
               color=DARK_BLUE if is_sec else "000000")
    for ci, v in enumerate(vals, 2):
        if v is None:
            apply_cell(ws, ri, ci, "", size=14, bg=bg)
        elif isinstance(v, str):
            apply_cell(ws, ri, ci, v, size=14, bg=bg, h_align="right")
        else:
            apply_cell(ws, ri, ci, v, bold=is_bold, size=14, bg=bg,
                       h_align="right", fmt="#,##0")


# ─────────────────────────────────────────────────────────
# TAB 6: CASH FLOW ANALYSIS
# ─────────────────────────────────────────────────────────
ws = wb.create_sheet("Cash Flow Analysis")
ws.sheet_view.showGridLines = False
set_col_width(ws, 1, 40)
for c in range(2, 8):
    set_col_width(ws, c, 18)

apply_header(ws, 1, 1, "CLEAN HARBORS — CASH FLOW ANALYSIS  (USD Millions)", colspan=7, size=18)
set_row_height(ws, 1, 44)
apply_header(ws, 2, 1,
    "FCF = CFO − CapEx.  Adj. FCF (mgmt) excludes landfill closure & timing items.",
    colspan=7, bg=MID_BLUE, size=13)
set_row_height(ws, 2, 26)

for ci, yr in enumerate(["", "FY2021", "FY2022", "FY2023", "FY2024", "FY2025A", "FY2026E"], 1):
    apply_header(ws, 3, ci, yr, bg=LIGHT_BLUE, size=14)
set_row_height(ws, 3, 28)

cf_data = [
    ("OPERATING CASH FLOW",        None,  None,  None,  None,   None,  None),
    ("Net Income",                 198,   297,   378,   402,    430,   435),
    ("+ D&A",                      340,   370,   390,   420,    440,   455),
    ("+ Other Non-cash",           80,    95,    100,   108,    120,   120),
    ("± Working Capital",          -45,   -68,   22,    -40,    55,    10),
    ("CASH FROM OPERATIONS (CFO)", 573,   694,   890,   890,    1045,  910),
    ("",                           None,  None,  None,  None,   None,  None),
    ("INVESTING ACTIVITIES",       None,  None,  None,  None,   None,  None),
    ("Capital Expenditures",       -336,  -366,  -530,  -535,   -536,  -430),
    ("Acquisitions, net",          -312,  -1250, -130,  -320,   -180,  -100),
    ("Asset Disposals",            25,    40,    35,    30,     42,    30),
    ("CASH FROM INVESTING",        -623,  -1576, -625,  -825,   -674,  -500),
    ("",                           None,  None,  None,  None,   None,  None),
    ("FINANCING ACTIVITIES",       None,  None,  None,  None,   None,  None),
    ("Net Debt Issuance / (Repay)",195,   985,   -30,   75,     -30,   -50),
    ("Share Repurchases",          -95,   -105,  -200,  -200,   -250,  -280),
    ("Other Financing",            0,     0,     0,     0,      0,     0),
    ("CASH FROM FINANCING",        100,   880,   -230,  -125,   -280,  -330),
    ("",                           None,  None,  None,  None,   None,  None),
    ("NET CHANGE IN CASH",         50,    -2,    35,    -60,    91,    80),
    ("",                           None,  None,  None,  None,   None,  None),
    ("FREE CASH FLOW ANALYSIS",    None,  None,  None,  None,   None,  None),
    ("CFO",                        573,   694,   890,   890,    1045,  910),
    ("Less: CapEx",                -336,  -366,  -530,  -535,   -536,  -430),
    ("Free Cash Flow",             237,   328,   360,   355,    509,   480),
    ("  FCF Margin",               "5.9%","6.3%","6.7%","6.0%","8.4%","7.7%"),
    ("Adj. FCF (mgmt def.)",       280,   370,   420,   355,    509,   510),
    ("  FCF Conv. of EBITDA",      "40%", "42%", "42%", "32%", "43.5%","41%"),
    ("",                           None,  None,  None,  None,   None,  None),
    ("CAPEX BREAKDOWN",            None,  None,  None,  None,   None,  None),
    ("Maintenance CapEx",          200,   220,   280,   285,    290,   260),
    ("Growth CapEx",               136,   146,   250,   250,    246,   170),
    ("Total CapEx",                336,   366,   530,   535,    536,   430),
    ("CapEx / Revenue",            "8.4%","7.1%","9.8%","9.1%","8.9%","6.9%"),
    ("",                           None,  None,  None,  None,   None,  None),
    ("OWNER EARNINGS FRAMEWORK",   None,  None,  None,  None,   None,  None),
    ("EBITDA",                     700,   875,   1002,  1115,   1170,  1230),
    ("Less: Maint. CapEx",         -200,  -220,  -280,  -285,   -290,  -260),
    ("Less: Tax (cash)",           -68,   -101,  -108,  -120,   -91,   -128),
    ("Owner Earnings Est.",        432,   554,   614,   710,    789,   842),
    ("Owner Earnings Yield",       "—",   "—",   "—",   "—",   "6.1%","6.5%"),
]

for ri, row in enumerate(cf_data, 4):
    label = row[0]
    vals  = row[1:]
    set_row_height(ws, ri, 8 if label == "" else 26)
    is_sec = label.isupper() and label not in ("", "FREE CASH FLOW ANALYSIS",
                                                "CAPEX BREAKDOWN", "OWNER EARNINGS FRAMEWORK")
    is_key = label in ("Free Cash Flow", "CASH FROM OPERATIONS (CFO)", "Adj. FCF (mgmt def.)",
                       "Owner Earnings Est.")
    is_bold = is_sec or is_key
    bg = MED_GREY if is_sec else (LIGHT_GREY if ri % 2 == 0 else WHITE)
    if is_key:
        bg = LIGHT_GREY
    apply_cell(ws, ri, 1, label, bold=is_bold, size=14, bg=bg,
               color=DARK_BLUE if is_sec else "000000")
    for ci, v in enumerate(vals, 2):
        if v is None:
            apply_cell(ws, ri, ci, "", size=14, bg=bg)
        elif isinstance(v, str):
            apply_cell(ws, ri, ci, v, size=14, bg=bg, h_align="right")
        else:
            apply_cell(ws, ri, ci, v, bold=is_bold, size=14, bg=bg,
                       h_align="right", fmt="#,##0")


# ─────────────────────────────────────────────────────────
# TAB 7: RETURN ON CAPITAL
# ─────────────────────────────────────────────────────────
ws = wb.create_sheet("Return on Capital")
ws.sheet_view.showGridLines = False
set_col_width(ws, 1, 40)
for c in range(2, 8):
    set_col_width(ws, c, 18)

apply_header(ws, 1, 1, "CLEAN HARBORS — RETURN ON CAPITAL ANALYSIS", colspan=7, size=18)
set_row_height(ws, 1, 44)

for ci, yr in enumerate(["", "FY2021", "FY2022", "FY2023", "FY2024", "FY2025", "FY2026E"], 1):
    apply_header(ws, 2, ci, yr, bg=LIGHT_BLUE, size=14)
set_row_height(ws, 2, 28)

roc_data = [
    ("RETURN METRICS",                None, None,   None,   None,   None,   None),
    ("Return on Equity (ROE)",        "9.5%","12.0%","13.8%","14.6%","14.7%","14.9%"),
    ("Return on Assets (ROA)",        "3.2%","4.2%","5.2%","5.4%","5.6%","5.7%"),
    ("Return on Invested Capital",    "6.8%","8.5%","9.8%","10.2%","10.5%","11.0%"),
    ("Return on Capital Employed",    "7.2%","9.1%","10.8%","11.4%","11.5%","12.0%"),
    ("WACC (estimated)",              "8.0%","8.0%","8.0%","8.5%","8.5%","8.5%"),
    ("ROIC - WACC (Spread)",          "-1.2%","0.5%","1.8%","1.7%","2.0%","2.5%"),
    ("",                              None, None,   None,   None,   None,   None),
    ("CAPITAL EMPLOYED ($M)",         None, None,   None,   None,   None,   None),
    ("Total Equity",                  2089, 2479,   2749,   2771,   2928,   3050),
    ("Net Debt",                      2036, 2131,   2022,   2122,   2014,   1950),
    ("Invested Capital",              4125, 4610,   4771,   4893,   4942,   5000),
    ("",                              None, None,   None,   None,   None,   None),
    ("INCREMENTAL RETURNS",           None, None,   None,   None,   None,   None),
    ("Revenue Growth",                None,"29.8%","4.7%","8.9%","2.4%","2.8%"),
    ("EBITDA Growth",                 None,"25.0%","14.5%","11.3%","4.9%","5.1%"),
    ("Incremental EBITDA on new Rev.",None,"26.7%","24.8%","26.1%","22.1%","~24%"),
    ("Incremental ROIC",              None,"12.1%","11.0%","12.5%","11.8%","~12%"),
    ("",                              None, None,   None,   None,   None,   None),
    ("QUALITY CHECK",                 None, None,   None,   None,   None,   None),
    ("EBITDA Margin",                 "17.6%","16.9%","18.5%","18.9%","19.4%","19.8%"),
    ("FCF / Net Income",              "119%","111%","95%","88%","118%","110%"),
    ("Net Debt / EBITDA",             "2.9x","2.4x","2.0x","1.9x","1.7x","~1.6x"),
    ("Interest Coverage (EBIT/Int.)", "3.8x","4.7x","4.9x","4.5x","4.4x","4.6x"),
    ("",                              None, None,   None,   None,   None,   None),
    ("CAPITAL ALLOCATION ($M)",       None, None,   None,   None,   None,   None),
    ("Organic CapEx",                 336,  366,    530,    535,    536,    430),
    ("M&A",                           312,  1250,   130,    320,    180,    100),
    ("Share Repurchases",             95,   105,    200,    200,    250,    280),
    ("",                              None, None,   None,   None,   None,   None),
    ("ASSESSMENT",
     "ROIC now exceeds WACC (2022+), creating economic value. Incremental ROIC of ~12% on new revenue "
     "demonstrates pricing power and operating leverage. Management allocates capital to organic growth "
     "(incinerators, PFAS capacity), selective M&A, and buybacks — a sensible hierarchy. "
     "Declining leverage (1.7x ND/EBITDA) further strengthens the balance sheet. "
     "VERDICT: CLH is a GENUINE VALUE CREATOR — ROIC trend is positive and accelerating.",
     None, None, None, None, None),
]

for ri, row in enumerate(roc_data, 3):
    label = row[0]
    vals  = row[1:]
    set_row_height(ws, ri, 8 if label == "" else 26)
    if label == "ASSESSMENT":
        set_row_height(ws, ri, 70)
        apply_cell(ws, ri, 1, label, bold=True, size=14, bg=YELLOW_HL, color=DARK_BLUE)
        ws.merge_cells(start_row=ri, start_column=2, end_row=ri, end_column=7)
        apply_cell(ws, ri, 2, vals[0], size=13, bg=YELLOW_HL, wrap=True)
        continue
    is_sec = label.isupper() and label not in ("", "ASSESSMENT")
    is_bold = is_sec
    bg = MED_GREY if is_sec else (LIGHT_GREY if ri % 2 == 1 else WHITE)
    apply_cell(ws, ri, 1, label, bold=is_bold, size=14, bg=bg,
               color=DARK_BLUE if is_sec else "000000")
    for ci, v in enumerate(vals, 2):
        if v is None:
            apply_cell(ws, ri, ci, "", size=14, bg=bg)
        elif isinstance(v, str):
            apply_cell(ws, ri, ci, v, size=14, bg=bg, h_align="right")
        else:
            apply_cell(ws, ri, ci, v, bold=is_bold, size=14, bg=bg,
                       h_align="right", fmt="#,##0")


# ─────────────────────────────────────────────────────────
# TAB 8: MANAGEMENT
# ─────────────────────────────────────────────────────────
ws = wb.create_sheet("Management")
ws.sheet_view.showGridLines = False
set_col_width(ws, 1, 30)
set_col_width(ws, 2, 52)
set_col_width(ws, 3, 28)

apply_header(ws, 1, 1, "CLEAN HARBORS — MANAGEMENT QUALITY ASSESSMENT", colspan=3, size=18)
set_row_height(ws, 1, 44)

mgmt = [
    ("LEADERSHIP",        "PROFILE",                                     "ASSESSMENT"),
    ("Alan McKim\n(Founder / Exec. Chair / CTO)",
     "Founded CLH in 1980. Transitioned to Executive Chairman & CTO in March 2023 after 40+ years as CEO. "
     "Retains a multi-percent personal stake (~5%+). Son William McKim employed at CLH subsidiary ($1.25M comp), "
     "son-in-law Robert Smith also employed ($143K comp). One-share / one-vote structure (no dual-class).",
     "Classic founder-operator with deep domain expertise. Family employment minor governance concern but "
     "disclosed transparently in proxy. Strong owner alignment."),
    ("Michael Battles &\nEric Gerstenberg\n(Co-CEOs since 2023)",
     "Both CLH 20+ year veterans. Battles: former CFO, finance-driven discipline. "
     "Gerstenberg: Operations background, customer relationships. Internal promotion — high continuity. "
     "95%+ Say-on-Pay support in 2025 proxy. Comp tied to EBITDA and FCF metrics.",
     "Orderly, well-planned succession. Pay-for-performance structure aligns with shareholders. "
     "Disciplined capital allocation evident in post-2022 era."),
    ("",                  "",                                            ""),
    ("INCENTIVE ALIGNMENT","DETAIL",                                     "ASSESSMENT"),
    ("Executive LTIP",
     "Long-term incentive plan weighted: (1) Adj. EBITDA (40%), (2) Adj. FCF (30%), "
     "(3) Relative TSR vs. peers (30%). Annual cash bonus: EBITDA + revenue targets.",
     "ALIGNED — metrics drive FCF and EBITDA growth, exactly what shareholders want."),
    ("Insider Ownership",
     "McKim: ~5%+ of shares (~$650M+ of personal wealth in CLH stock). "
     "Co-CEOs accumulating via equity comp. Total insider: ~7-8%. "
     "Institutional: Vanguard (>10%), BlackRock, Fidelity, Wellington among top holders.",
     "POSITIVE — founder's large stake creates genuine owner mentality. No recent insider selling."),
    ("SEC Form 4 Activity",
     "Recent filings show NO significant insider selling at executive level. McKim has not materially "
     "reduced stake. Co-CEOs accumulating via RSU vesting. No alarm signals.",
     "POSITIVE — insiders are holding, not exiting."),
    ("",                  "",                                            ""),
    ("CAPITAL ALLOCATION", "DETAIL",                                     "ASSESSMENT"),
    ("M&A Track Record",
     "Safety-Kleen (2012, ~$1.25B): transformational — best deal in CLH history, built the recurring "
     "revenue base. HydroChemPSC (2021, ~$1.25B): industrial services, integration bumpy, margins "
     "below plan. Post-2022: more selective, smaller bolt-ons.",
     "MIXED-POSITIVE — Safety-Kleen was genius; HydroChemPSC was expensive. New co-CEO era shows "
     "improved discipline."),
    ("Buyback Program",
     "$250M repurchased in 2025. Board added $350M → $600M total availability. "
     "Shares out: 55M (2020) → 50.8M (2025). EPS accretion from buybacks is meaningful. "
     "No dividend — growth reinvestment preferred.",
     "POSITIVE — consistent buyback program at reasonable valuations. Signals management confidence."),
    ("Leverage Management",
     "Net Debt/EBITDA: 2.9x (2021) → 1.7x (2025). Target: ~2x. "
     "Weighted avg interest rate: ~5%. 14-yr avg maturity, no near-term wall.",
     "POSITIVE — declining leverage. Balance sheet strengthening, creating optionality."),
    ("Future Seeds",
     "Investing $210-220M in new SDA incineration unit (Kimball, NE). PFAS pipeline growing "
     "15-20% per quarter. Digital waste tracking investments. Not borrowing from future.",
     "POSITIVE — management planting PFAS seeds. Long payback (6-7yr) but high confidence in "
     "regulatory tailwind. Responsible long-term capital deployment."),
    ("",                  "",                                            ""),
    ("OVERALL VERDICT",
     "B+ / A-  Quality Management. Strong founder DNA, orderly succession, aligned incentives, "
     "disciplined buybacks, improving ROIC. Minor governance concern (family employment). "
     "Capital allocation improving. Acts like owners. Planting seeds, not borrowing from future.",
     "SUITABLE FOR LONG-TERM OWNERSHIP"),
]

for ri, row in enumerate(mgmt, 2):
    label, detail, assess = row
    is_section = label.isupper() and label not in ("", "OVERALL VERDICT")
    is_header  = ri == 2
    is_overall = label == "OVERALL VERDICT"
    bg = MED_GREY if is_section else (WHITE if ri % 2 == 0 else LIGHT_GREY)
    if is_overall:
        bg = YELLOW_HL
    if label == "":
        set_row_height(ws, ri, 8)
    elif is_section or is_header:
        set_row_height(ws, ri, 30)
    elif is_overall:
        set_row_height(ws, ri, 55)
    else:
        set_row_height(ws, ri, 70)

    if is_header:
        for ci, h in enumerate([label, detail, assess], 1):
            apply_header(ws, ri, ci, h, bg=LIGHT_BLUE, size=14)
    else:
        assess_color = GREEN if "POSITIVE" in assess or "ALIGNED" in assess else (
                       RED if "CONCERN" in assess else "000000")
        apply_cell(ws, ri, 1, label, bold=True, size=13, bg=bg, wrap=True,
                   color=DARK_BLUE if is_section or is_overall else "000000")
        apply_cell(ws, ri, 2, detail, size=13, bg=bg, wrap=True)
        apply_cell(ws, ri, 3, assess, bold=is_overall, size=13, bg=bg, wrap=True,
                   color=assess_color if not is_overall else GREEN)


# ─────────────────────────────────────────────────────────
# TAB 9: RISKS
# ─────────────────────────────────────────────────────────
ws = wb.create_sheet("Risks")
ws.sheet_view.showGridLines = False
set_col_width(ws, 1, 26)
set_col_width(ws, 2, 50)
set_col_width(ws, 3, 16)
set_col_width(ws, 4, 16)
set_col_width(ws, 5, 34)

apply_header(ws, 1, 1, "CLEAN HARBORS — RISK ANALYSIS", colspan=5, size=18)
set_row_height(ws, 1, 44)

for ci, h in enumerate(["Risk Category", "Description", "Probability", "Impact", "Mitigation"], 1):
    apply_header(ws, 2, ci, h, bg=LIGHT_BLUE, size=14)
set_row_height(ws, 2, 28)

risks = [
    ("Regulatory / PFAS",
     "PFAS incineration standards could tighten beyond current EPA guidance post-2026, raising CLH's "
     "operational costs. Alternative destruction technologies (electrochemical, supercritical water) "
     "could emerge as cheaper alternatives over 5-10 years.",
     "Low", "High",
     "CLH co-chairs EPA technical working groups; helps set standards. First-mover + best-in-class "
     "technology. End-to-end destruction is CLH's competitive advantage."),
    ("Base Oil Price Volatility",
     "Safety-Kleen's margins are tied to the spread between crude-derived lubricants and re-refined "
     "base oil. A crude oil price collapse or surge in re-refined competition compresses SKSS margins.",
     "Medium", "Medium",
     "Fuel-for-profit program partially hedges. Mix shift to higher-margin services. SKSS is ~28% "
     "of revenue; ES diversifies."),
    ("Economic Cyclicality",
     "Industrial Services (part of ES) fell 4% in 2025; no recovery until spring 2026. "
     "A manufacturing recession or chemical plant slowdown reduces waste volumes and turnarounds.",
     "Medium", "Medium",
     "Environmental compliance is non-discretionary. ES base is sticky. PFAS remediation is "
     "government-mandated regardless of cycle."),
    ("Valuation Risk",
     "At ~14x EV/EBITDA and 33x P/E, CLH is priced for continued growth. If PFAS catalysts delay "
     "or Industrial Services recovery disappoints, multiple compression is the primary risk.",
     "Medium", "High",
     "Buy on weakness. PFAS pipeline growth 15-20% QoQ provides visibility. Buybacks support EPS. "
     "Fair value analysis supports $250-280 range."),
    ("M&A Integration",
     "HydroChemPSC (2021) integration was bumpy; margins below expectations. Future acquisitions "
     "carry similar risk, especially if environment services companies are overpriced.",
     "Low-Med", "Medium",
     "New co-CEO team shows more M&A discipline. Focus on organic PFAS/CapEx investment. "
     "Recent acquisitions are smaller bolt-ons."),
    ("Environmental Liability",
     "Operating hazardous waste TSDFs, landfills, and incinerators carries potential legacy "
     "environmental liability from historical contamination or operational incidents.",
     "Low", "High",
     "CLH's business model is built on compliance. Strong EHS culture; regular EPA inspections. "
     "Insurance and indemnification provisions in acquisition agreements."),
    ("Competition",
     "Veolia, US Ecology (now Clean Earth), Stericycle offer niche competition. "
     "Waste Management or Republic entering the hazardous space is a very low risk.",
     "Low", "Low",
     "No competitor can replicate CLH's incinerator network + landfill portfolio + Safety-Kleen. "
     "Permitting barrier is the structural moat."),
    ("Management Transition",
     "Co-CEO structure is unusual. If Battles or Gerstenberg exit, or if conflict arises, "
     "leadership uncertainty could impact execution and stock price.",
     "Low", "Medium",
     "McKim remains as Executive Chairman. Both co-CEOs are 20+ year CLH veterans. "
     "Culture is deeply ingrained."),
    ("Interest Rate / Refinancing",
     "Rising rates increase cost on $2.76B LTD. Refinancing at higher rates could compress FCF.",
     "Low", "Low",
     "14-yr weighted avg maturity, no near-term maturities. Declining leverage reduces risk. "
     "Net leverage at 1.7x — ample coverage."),
]

for ri, row in enumerate(risks, 3):
    cat, desc, prob, impact, mitig = row
    set_row_height(ws, ri, 58)
    bg = LIGHT_GREY if ri % 2 == 1 else WHITE
    impact_color = RED if impact == "High" else (AMBER if impact == "Medium" else GREEN)
    apply_cell(ws, ri, 1, cat, bold=True, size=13, bg=bg, wrap=True)
    apply_cell(ws, ri, 2, desc, size=13, bg=bg, wrap=True)
    apply_cell(ws, ri, 3, prob, size=13, bg=bg, h_align="center")
    apply_cell(ws, ri, 4, impact, bold=True, size=13, bg=bg, h_align="center", color=impact_color)
    apply_cell(ws, ri, 5, mitig, size=13, bg=bg, wrap=True)


# ─────────────────────────────────────────────────────────
# TAB 10: VALUATION
# ─────────────────────────────────────────────────────────
ws = wb.create_sheet("Valuation")
ws.sheet_view.showGridLines = False
set_col_width(ws, 1, 36)
for c in range(2, 7):
    set_col_width(ws, c, 22)

apply_header(ws, 1, 1, "CLEAN HARBORS — VALUATION", colspan=6, size=18)
set_row_height(ws, 1, 44)

apply_header(ws, 2, 1, "CURRENT MARKET METRICS  (April 2026)", colspan=6, bg=MID_BLUE, size=14)
set_row_height(ws, 2, 28)

mkt_data = [
    ("Share Price",               "$275.88"),
    ("Diluted Shares Outstanding","50.8M"),
    ("Market Capitalization",     "$14.0B"),
    ("(+) Total Debt",            "$2.84B"),
    ("(-) Cash",                  "($0.83B)"),
    ("Enterprise Value (EV)",     "$16.0B"),
]
for ri, (label, val) in enumerate(mkt_data, 3):
    bg = LIGHT_GREY if ri % 2 == 1 else WHITE
    set_row_height(ws, ri, 26)
    apply_cell(ws, ri, 1, label, bold=(label in ("Enterprise Value (EV)", "Market Capitalization")),
               size=14, bg=bg)
    apply_cell(ws, ri, 2, val, bold=(label == "Enterprise Value (EV)"), size=14, bg=bg,
               h_align="right")

apply_header(ws, 10, 1, "RELATIVE VALUATION — MULTIPLES", colspan=6, bg=MID_BLUE, size=14)
set_row_height(ws, 10, 28)

for ci, h in enumerate(["Metric", "CLH", "Peer Avg", "S&P 500", "Assessment"], 1):
    apply_header(ws, 11, ci, h, bg=LIGHT_BLUE, size=14)
set_row_height(ws, 11, 28)

multiples = [
    ("Trailing P/E",         "33.8x", "30x",  "22x", "Slight premium; justified by moat & PFAS growth"),
    ("Forward P/E (2026E)",  "31.1x", "27x",  "20x", "Reasonable for wide-moat compounder"),
    ("EV/EBITDA (TTM)",      "13.7x", "11x",  "—",   "Fair; CLH deserves premium to peers for moat quality"),
    ("EV/EBITDA (2026E)",    "13.0x", "10x",  "—",   "Base case fair value ~$250-280; modest upside at current"),
    ("Price / FCF",          "27.5x", "22x",  "—",   "Elevated but FCF growing → ratio compresses over time"),
    ("EV / Revenue",         "2.7x",  "1.8x", "—",   "Premium multiple reflects hard asset quality"),
    ("PEG Ratio",            "3.5x",  "2.5x", "—",   "Growth is priced in; monitor PFAS delivery vs. guidance"),
]
for ri, row in enumerate(multiples, 12):
    bg = LIGHT_GREY if ri % 2 == 0 else WHITE
    set_row_height(ws, ri, 26)
    for ci, val in enumerate(row, 1):
        apply_cell(ws, ri, ci, val, size=14, bg=bg,
                   h_align="right" if ci in [2,3,4] else "left")

apply_header(ws, 20, 1, "DCF SCENARIO ANALYSIS  (5-Year Horizon)", colspan=6, bg=MID_BLUE, size=14)
set_row_height(ws, 20, 28)
for ci, h in enumerate(["Scenario", "EBITDA CAGR", "Term. EV/EBITDA", "Discount Rate", "Implied Price", "Upside / Downside"], 1):
    apply_header(ws, 21, ci, h, bg=LIGHT_BLUE, size=14)
set_row_height(ws, 21, 28)

dcf_rows = [
    ("Bear Case",       "2-3%", "9-10x",  "10%",   "$185-200", "-27% to -33%"),
    ("Base Case",       "5-6%", "11-12x", "9%",    "$260-280", "-3% to +1%"),
    ("Bull Case",       "8-10%","13-14x", "8.5%",  "$340-380", "+23% to +38%"),
    ("Citi Research",   "—",    "—",      "—",     "$346",     "+25%"),
    ("Analyst Avg",     "—",    "—",      "—",     "$300",     "+9%"),
]
for ri, row in enumerate(dcf_rows, 22):
    bg = YELLOW_HL if row[0] == "Base Case" else (LIGHT_GREY if ri % 2 == 0 else WHITE)
    set_row_height(ws, ri, 26)
    for ci, val in enumerate(row, 1):
        apply_cell(ws, ri, ci, val, bold=(row[0] == "Base Case"), size=14,
                   bg=bg, h_align="center")

apply_header(ws, 28, 1, "VALUATION VERDICT", colspan=6, bg=MID_BLUE, size=14)
set_row_height(ws, 28, 28)
set_row_height(ws, 29, 80)
ws.merge_cells("A29:F29")
verdict = (
    "VERDICT: CLH is FAIRLY VALUED at ~$276 (Base Case $260-280). The stock is not cheap on P/E but the wide moat, "
    "PFAS secular tailwind, improving FCF conversion (now ~44% of EBITDA), and consistent buybacks justify a premium. "
    "\n\nMARGIN OF SAFETY: MODEST at current price. Stock would be more attractive below $240-250 (~10-15% pullback). "
    "For a 5+ year hold, CLH offers mid-to-high single digit total returns plus potential PFAS upside.\n\n"
    "RATING: HOLD at $276 | BUY below $250 | Target: $300-346 (Base to Bull)"
)
apply_cell(ws, 29, 1, verdict, size=13, bold=True, bg=YELLOW_HL, wrap=True, h_align="left")


# ─────────────────────────────────────────────────────────
# TAB 11: MARKET SENTIMENT
# ─────────────────────────────────────────────────────────
ws = wb.create_sheet("Market Sentiment")
ws.sheet_view.showGridLines = False
set_col_width(ws, 1, 28)
set_col_width(ws, 2, 52)
set_col_width(ws, 3, 20)

apply_header(ws, 1, 1, "CLEAN HARBORS — MARKET SENTIMENT", colspan=3, size=18)
set_row_height(ws, 1, 44)
for ci, h in enumerate(["Indicator", "Detail", "Sentiment"], 1):
    apply_header(ws, 2, ci, h, bg=LIGHT_BLUE, size=14)
set_row_height(ws, 2, 28)

sentiment_rows = [
    ("Analyst Consensus",
     "11 analysts covering CLH. Consensus: BUY. 36% Strong Buy, 27% Buy, 36% Hold. No Sells. "
     "Average target $300; Citi upgraded to Buy with $346 target in April 2026 noting record revenue "
     "and 15 consecutive quarters of margin expansion.",
     "BULLISH"),
    ("Recent Earnings",
     "Q4 2025: EPS and revenue beat — stock +5.8% after buyback expansion announced. "
     "Full year 2025 adjusted FCF record at $509M. 2026 EBITDA guidance midpoint $1.23B (+5%). "
     "Market reaction muted on the beat itself — street had modeled it.",
     "NEUTRAL-POSITIVE"),
    ("PFAS Outlook",
     "Growing analyst focus on PFAS as multi-year secular tailwind. CLH published formal PFAS "
     "disposal & treatment standards (April 2026), positioning as the industry authority. "
     "Pipeline growing 15-20% per quarter. EPA rulemaking expected to expand mandatory treatment.",
     "VERY BULLISH"),
    ("Buyback Signal",
     "$350M added to buyback program (2025); total $600M available. Buybacks at $275 imply "
     "management views current price as reasonable / undervalued. EPS accretion of ~3-4% annually "
     "from share count reduction.",
     "POSITIVE"),
    ("Macro / Industry",
     "Reshoring of manufacturing drives domestic hazardous waste volumes. CHIPS Act ($52B) and IRA "
     "manufacturing credits stimulate new industrial activity. Emergency response pipeline seasonal "
     "strength. Environmental enforcement under EPA remains active.",
     "TAILWIND"),
    ("Short Interest",
     "Short interest ~2-3% of float — minimal bearish institutional conviction. Not a crowded "
     "short or long. Balanced ownership.",
     "NEUTRAL"),
    ("Institutional Flows",
     "Vanguard (>10%), BlackRock, Fidelity, Wellington, State Street are top holders. "
     "No significant institutional selldowns. Steady accumulation by quality-growth mandates.",
     "POSITIVE"),
    ("Technical / Price Action",
     "Stock trades near 52-week average. Citi upgrade in April 2026 at $346 target could catalyze "
     "re-rating. PFAS milestones (EPA guidance updates) are the key positive catalyst.",
     "NEUTRAL-POSITIVE"),
]
for ri, (ind, det, sent) in enumerate(sentiment_rows, 3):
    bg = LIGHT_GREY if ri % 2 == 1 else WHITE
    set_row_height(ws, ri, 52)
    sent_color = GREEN if any(k in sent for k in ["BULLISH", "POSITIVE", "TAILWIND"]) else (
                 RED if "NEGATIVE" in sent else AMBER)
    apply_cell(ws, ri, 1, ind, bold=True, size=13, bg=bg, wrap=True)
    apply_cell(ws, ri, 2, det, size=13, bg=bg, wrap=True)
    apply_cell(ws, ri, 3, sent, bold=True, size=13, bg=bg, h_align="center", color=sent_color)


# ─────────────────────────────────────────────────────────
# TAB 12: KEY INDICATORS
# ─────────────────────────────────────────────────────────
ws = wb.create_sheet("Key Indicators")
ws.sheet_view.showGridLines = False
set_col_width(ws, 1, 38)
for c in range(2, 8):
    set_col_width(ws, c, 18)

apply_header(ws, 1, 1, "CLEAN HARBORS — KEY INDICATORS DASHBOARD", colspan=7, size=18)
set_row_height(ws, 1, 44)
for ci, yr in enumerate(["INDICATOR", "FY2021", "FY2022", "FY2023", "FY2024", "FY2025", "FY2026E"], 1):
    apply_header(ws, 2, ci, yr, bg=LIGHT_BLUE, size=14)
set_row_height(ws, 2, 28)

ki_data = [
    ("GROWTH",                   None,   None,   None,   None,   None,   None),
    ("Revenue ($M)",             3981,   5168,   5409,   5890,   6030,   6200),
    ("Revenue YoY Growth",       None,  "29.8%","4.7%", "8.9%", "2.4%", "2.8%"),
    ("Adj. EBITDA ($M)",         700,    875,    1002,   1115,   1170,   1230),
    ("EBITDA YoY Growth",        None,  "25.0%","14.5%","11.3%","4.9%", "5.1%"),
    ("EPS (Diluted, $)",         3.62,   5.50,   7.19,   7.81,   8.46,   8.79),
    ("EPS YoY Growth",           None,  "51.9%","30.7%","8.6%", "8.3%", "3.9%"),
    ("",                         None,   None,   None,   None,   None,   None),
    ("PROFITABILITY",            None,   None,   None,   None,   None,   None),
    ("Gross Margin",             "29.8%","28.4%","28.9%","28.7%","28.7%","29.4%"),
    ("EBITDA Margin",            "17.6%","16.9%","18.5%","18.9%","19.4%","19.8%"),
    ("EBIT Margin",              "9.1%", "9.8%","11.3%","11.4%","11.2%","11.6%"),
    ("Net Margin",               "5.0%", "5.7%", "7.0%", "6.8%", "7.1%", "7.0%"),
    ("FCF Margin",               "5.9%", "6.3%", "6.7%", "6.0%", "8.4%", "7.7%"),
    ("",                         None,   None,   None,   None,   None,   None),
    ("CASH GENERATION",          None,   None,   None,   None,   None,   None),
    ("Free Cash Flow ($M)",      237,    328,    360,    355,    509,    480),
    ("FCF / EBITDA Conv.",       "33.9%","37.5%","35.9%","31.8%","43.5%","39%"),
    ("CapEx / Revenue",          "8.4%", "7.1%", "9.8%", "9.1%", "8.9%", "6.9%"),
    ("",                         None,   None,   None,   None,   None,   None),
    ("BALANCE SHEET",            None,   None,   None,   None,   None,   None),
    ("Cash ($M)",                564,    614,    748,    703,    826,    900),
    ("Net Debt ($M)",            2036,   2131,   2022,   2122,   2014,   1950),
    ("Net Debt / EBITDA",        "2.9x", "2.4x", "2.0x", "1.9x", "1.7x", "~1.6x"),
    ("Current Ratio",            "1.76x","1.74x","1.87x","1.81x","1.89x","1.90x"),
    ("",                         None,   None,   None,   None,   None,   None),
    ("RETURNS",                  None,   None,   None,   None,   None,   None),
    ("ROE",                      "9.5%", "12.0%","13.8%","14.6%","14.7%","14.9%"),
    ("ROIC",                     "6.8%", "8.5%", "9.8%","10.2%","10.5%","11.0%"),
    ("ROA",                      "3.2%", "4.2%", "5.2%", "5.4%", "5.6%", "5.7%"),
    ("",                         None,   None,   None,   None,   None,   None),
    ("VALUATION",                None,   None,   None,   None,   None,   None),
    ("Share Price (period end)", "—",   "—",    "—",    "—",    "~$262","$276"),
    ("Shares Out. (M)",          54.7,   54.0,   52.6,   51.5,   50.8,   49.5),
    ("Market Cap ($B)",          "—",   "—",    "—",    "—",    "~$13.3","~$14.0"),
    ("EV/EBITDA (trailing)",     "—",   "—",    "—",    "—",    "13.8x","13.0x"),
    ("P/E (trailing)",           "—",   "—",    "—",    "—",    "33.8x","31.4x"),
    ("Price / FCF",              "—",   "—",    "—",    "—",    "26.1x","~29x"),
]

for ri, row in enumerate(ki_data, 3):
    label = row[0]
    vals  = row[1:]
    set_row_height(ws, ri, 8 if label == "" else 26)
    is_sec = label.isupper() and label not in ("",)
    is_bold = is_sec
    bg = MED_GREY if is_sec else (LIGHT_GREY if ri % 2 == 1 else WHITE)
    apply_cell(ws, ri, 1, label, bold=is_bold, size=14, bg=bg,
               color=DARK_BLUE if is_sec else "000000")
    for ci, v in enumerate(vals, 2):
        if v is None:
            apply_cell(ws, ri, ci, "", size=14, bg=bg)
        elif isinstance(v, str):
            apply_cell(ws, ri, ci, v, size=14, bg=bg, h_align="right")
        else:
            fmt = "#,##0" if abs(v) >= 1 else "#,##0.00"
            apply_cell(ws, ri, ci, v, bold=is_bold, size=14, bg=bg,
                       h_align="right", fmt=fmt)


# ─────────────────────────────────────────────────────────
# SAVE
# ─────────────────────────────────────────────────────────
os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
wb.save(OUTPUT_PATH)
print(f"Saved: {OUTPUT_PATH}")
