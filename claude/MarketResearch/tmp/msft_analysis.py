"""
Microsoft Corporation (MSFT) Financial Analysis - Excel Generator
Data as of April 2026 | FY2025 Annual (ended June 30, 2025)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

# ── Color palette ──────────────────────────────────────────────────────────────
MS_BLUE     = "0078D4"
MS_DARK     = "003366"
MS_GREEN    = "107C10"
MS_ORANGE   = "D83B01"
HEADER_BG   = "0078D4"
HEADER_FG   = "FFFFFF"
SUBHDR_BG   = "CCE4F7"
SUBHDR_FG   = "003366"
ALT_ROW     = "EBF5FB"
WHITE       = "FFFFFF"
LIGHT_GRAY  = "F4F6F7"
RED         = "C0392B"
GREEN       = "27AE60"
DARK_BLUE   = "003366"
GOLD        = "D4AC0D"
WARN_YELLOW = "FEF9E7"

FONT_SIZE = 14

def mf(bold=False, size=FONT_SIZE, color="000000", italic=False):
    return Font(name="Calibri", bold=bold, size=size, color=color, italic=italic)

def mfill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def mborder(style="thin"):
    s = Side(border_style=style, color="BBBBBB")
    return Border(left=s, right=s, top=s, bottom=s)

def cal(h="left", wrap=False):
    return Alignment(horizontal=h, vertical="center", wrap_text=wrap)

def wc(ws, row, col, value, bold=False, size=FONT_SIZE, fg="000000", bg=None,
       align="left", wrap=False, border=False, italic=False, num_fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = mf(bold=bold, size=size, color=fg, italic=italic)
    if bg:
        c.fill = mfill(bg)
    c.alignment = cal(align, wrap)
    if border:
        c.border = mborder()
    if num_fmt:
        c.number_format = num_fmt
    return c

def sec_hdr(ws, row, col, text, span=1, bg=HEADER_BG, fg=HEADER_FG):
    c = wc(ws, row, col, text, bold=True, size=FONT_SIZE+2, fg=fg, bg=bg, align="center")
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+span-1)
    ws.row_dimensions[row].height = 30
    return c

def sub_hdr(ws, row, col, text, span=1, bg=SUBHDR_BG, fg=SUBHDR_FG):
    c = wc(ws, row, col, text, bold=True, size=FONT_SIZE, fg=fg, bg=bg, align="center", border=True)
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+span-1)
    ws.row_dimensions[row].height = 22
    return c

def lbl(ws, row, col, text, bg=None, indent=0):
    val = ("   " * indent) + text
    c = wc(ws, row, col, val, size=FONT_SIZE, bg=bg, align="left", border=True)
    ws.row_dimensions[row].height = 20
    return c

def dc(ws, row, col, value, bg=None, num_fmt=None, bold=False, color="000000"):
    c = wc(ws, row, col, value, bold=bold, size=FONT_SIZE, fg=color, bg=bg, align="right", border=True, num_fmt=num_fmt)
    ws.row_dimensions[row].height = 20
    return c

def alt(i):
    return ALT_ROW if i % 2 == 0 else WHITE

def scw(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

# ═══════════════════════════════════════════════════════════════════════════════
def build_cover(wb):
    ws = wb.create_sheet("Cover")
    ws.sheet_view.showGridLines = False
    for c in range(1, 9):
        scw(ws, c, 18)
    scw(ws, 1, 4)
    scw(ws, 2, 30)
    scw(ws, 3, 22)

    ws.merge_cells("B3:G3")
    wc(ws, 3, 2, "MICROSOFT CORPORATION (MSFT)", bold=True, size=32, fg=WHITE,
       bg=MS_BLUE, align="center")
    ws.row_dimensions[3].height = 55

    ws.merge_cells("B4:G4")
    wc(ws, 4, 2, "Comprehensive Investment Analysis | April 2026", bold=False,
       size=16, fg=WHITE, bg=DARK_BLUE, align="center")
    ws.row_dimensions[4].height = 28

    ws.merge_cells("B6:G6")
    wc(ws, 6, 2, "Company Profile", bold=True, size=16, fg=WHITE, bg=DARK_BLUE, align="center")
    ws.row_dimensions[6].height = 25

    info = [
        ("Ticker",               "MSFT"),
        ("Exchange",             "NASDAQ"),
        ("Sector",               "Technology / Software"),
        ("Headquarters",         "Redmond, Washington, USA"),
        ("Founded",              "1975 by Bill Gates & Paul Allen"),
        ("CEO",                  "Satya Nadella (since 2014)"),
        ("CFO",                  "Amy Hood"),
        ("Fiscal Year End",      "June 30"),
        ("Market Cap",           "~$2.77 Trillion (April 2026)"),
        ("Stock Price",          "~$373 (April 13, 2026)"),
        ("P/E Ratio (TTM)",      "~23x | Forward ~21x"),
        ("Shares Outstanding",   "~7.43 Billion"),
        ("Dividend Yield",       "~0.8% ($3.32/share annually)"),
        ("Credit Rating",        "Aaa / AAA (Moody's / S&P) — One of 2 US AAA-rated cos"),
    ]

    for i, (k, v) in enumerate(info):
        r = 7 + i
        bg = ALT_ROW if i % 2 == 0 else WHITE
        wc(ws, r, 2, k, bold=True, size=FONT_SIZE, bg=bg, align="left", border=True)
        wc(ws, r, 3, v, size=FONT_SIZE, bg=bg, align="left", border=True)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)
        ws.row_dimensions[r].height = 22

    ws.merge_cells("B22:G22")
    wc(ws, 22, 2, "Investment Thesis", bold=True, size=16, fg=WHITE, bg=DARK_BLUE, align="center")
    ws.row_dimensions[22].height = 25

    thesis = (
        "Microsoft is the world's most important enterprise technology platform, operating as the "
        "'operating system' of the digital economy. Azure is the world's #2 cloud platform (24% market "
        "share), growing 34–40% annually. The OpenAI partnership creates an exclusive AI moat that "
        "competitors cannot immediately replicate. Microsoft 365's deep enterprise penetration (hundreds "
        "of millions of users) and Copilot monetization ($30/user/month) represent a massive hidden "
        "revenue layer. FY2025 revenue was $281.7B (+15%), operating income $128.5B (+17%), and FCF "
        "$71.6B — all record highs. Trading at ~21x forward earnings — its lowest valuation in 3 years "
        "— against a consensus price target of ~$595 (60% upside), MSFT offers a compelling risk-adjusted "
        "entry for long-term investors. One of only two AAA-rated US companies, with a fortress balance sheet."
    )
    ws.merge_cells("B23:G27")
    wc(ws, 23, 2, thesis, size=FONT_SIZE, bg=LIGHT_GRAY, align="left", wrap=True, border=True)
    for r in range(23, 28):
        ws.row_dimensions[r].height = 22

    ws.merge_cells("B29:G29")
    wc(ws, 29, 2, "Key Financial Highlights (FY2025 — Year Ended June 30, 2025)", bold=True,
       size=16, fg=WHITE, bg=DARK_BLUE, align="center")
    ws.row_dimensions[29].height = 25

    highlights = [
        ("Total Revenue",               "$281.7B",   "+15.1% YoY"),
        ("Operating Income",            "$128.5B",   "+17% YoY, 45.6% margin"),
        ("Net Income",                  "$101.8B",   "+22% YoY, 36.1% margin"),
        ("Gross Profit",                "$193.9B",   "68.8% margin"),
        ("Operating Cash Flow",         "~$136B",    "48.3% OCF margin"),
        ("Free Cash Flow",              "$71.6B",    "After $64.6B CapEx (AI infra)"),
        ("Capital Expenditures",        "$64.6B",    "+45% YoY — AI data center buildout"),
        ("Azure Revenue",               "$75B+",     "+34% YoY; first time >$75B"),
        ("Microsoft 365 Commercial",    "High-teens%", "Cloud revenue growing 18%"),
        ("Share Repurchases + Divs",    "~$44B",     "Consistent capital return"),
    ]

    sub_hdr(ws, 30, 2, "Metric", bg=SUBHDR_BG)
    sub_hdr(ws, 30, 3, "Value", bg=SUBHDR_BG)
    sub_hdr(ws, 30, 4, "Comment", bg=SUBHDR_BG)
    ws.merge_cells(start_row=30, start_column=4, end_row=30, end_column=7)

    for i, (m, v, c) in enumerate(highlights):
        r = 31 + i
        bg = ALT_ROW if i % 2 == 0 else WHITE
        lbl(ws, r, 2, m, bg=bg)
        dc(ws, r, 3, v, bg=bg)
        lbl(ws, r, 4, c, bg=bg)
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=7)

    ws.merge_cells("B42:G42")
    wc(ws, 42, 2, "Report Date: April 13, 2026  |  Data Sources: Microsoft 10-K, SEC Filings, Bloomberg, Analyst Reports",
       italic=True, size=12, fg="666666", align="center")


# ═══════════════════════════════════════════════════════════════════════════════
def build_business(wb):
    ws = wb.create_sheet("Business Overview")
    ws.sheet_view.showGridLines = False
    scw(ws, 1, 3); scw(ws, 2, 32); scw(ws, 3, 20); scw(ws, 4, 20)
    scw(ws, 5, 20); scw(ws, 6, 20); scw(ws, 7, 20)

    sec_hdr(ws, 1, 2, "BUSINESS OVERVIEW — MICROSOFT CORPORATION (MSFT)", span=6)

    sub_hdr(ws, 3, 2, "What Microsoft Does", span=6)
    desc = ("Microsoft is a global technology company that develops, licenses, and supports software, "
            "services, devices, and solutions. Founded in 1975, it has evolved from a desktop OS/productivity "
            "suite company into one of the world's largest cloud computing and enterprise AI platforms. "
            "Key businesses include Azure (cloud infrastructure), Microsoft 365 (productivity SaaS), "
            "Dynamics 365 (ERP/CRM), GitHub, LinkedIn, Xbox/Gaming, Bing/Search, and Windows. The company "
            "serves over 1 billion Windows users, 345M+ Microsoft 365 subscribers, and hosts millions of "
            "enterprise applications on Azure. The OpenAI partnership is central to Microsoft's AI strategy, "
            "integrating GPT-4 and o1 models across Copilot, Azure AI, and its entire product portfolio.")
    ws.merge_cells("B4:G7")
    wc(ws, 4, 2, desc, size=FONT_SIZE, bg=LIGHT_GRAY, align="left", wrap=True, border=True)
    for r in range(4, 8):
        ws.row_dimensions[r].height = 22

    # Revenue Segments
    sub_hdr(ws, 9, 2, "Revenue Segments (FY2025 — Year Ended June 30, 2025)", span=6)

    seg_headers = ["Segment", "FY2025 Rev ($B)", "% of Total", "YoY Growth", "Key Products / Services", "Margin Profile"]
    for i, h in enumerate(seg_headers):
        sub_hdr(ws, 10, 2+i, h)

    segments = [
        ("Productivity & Business Processes", "$120.9B", "42.9%", "+13%",
         "Microsoft 365 (Office), Exchange, SharePoint, Teams, LinkedIn, Dynamics 365",
         "High margins; ~68% gross margin on cloud"),
        ("Intelligent Cloud", "$107.2B", "38.1%", "+21%",
         "Azure, SQL Server, Windows Server, GitHub, Visual Studio, Azure AI Services",
         "Highest growth; data center scale economies"),
        ("More Personal Computing", "$53.6B", "19.0%", "+9%",
         "Windows OEM, Surface devices, Xbox, Bing/Search, PC gaming",
         "Lower margin; hardware headwinds"),
        ("TOTAL", "$281.7B", "100%", "+15%",
         "All three segments growing", "Blended ~68.8% gross margin"),
    ]

    for i, row_data in enumerate(segments):
        r = 11 + i
        bg = ALT_ROW if i % 2 == 0 else WHITE
        if i == len(segments) - 1:
            bg = SUBHDR_BG
        for j, val in enumerate(row_data):
            bold = (i == len(segments) - 1)
            lbl(ws, r, 2+j, val, bg=bg)
            if bold:
                ws.cell(row=r, column=2+j).font = mf(bold=True, size=FONT_SIZE)

    # Geographic Revenue
    sub_hdr(ws, 17, 2, "Geographic Revenue Breakdown (FY2025)", span=6)

    geo_headers = ["Region", "Revenue ($B)", "% of Total", "Growth", "Notes"]
    for i, h in enumerate(geo_headers):
        sub_hdr(ws, 18, 2+i, h)

    geo = [
        ("United States",          "$153.0B", "54.3%", "+16%", "Largest market; enterprise IT spending"),
        ("Other Americas",         "$18.2B",  "6.5%",  "+12%", "Latin America growth; USD headwinds"),
        ("Europe, Middle East & Africa", "$70.1B", "24.9%", "+14%", "GDPR compliance; strong Azure adoption"),
        ("Asia Pacific",           "$40.4B",  "14.3%", "+12%", "China licensing + India Azure growth"),
        ("TOTAL",                  "$281.7B", "100%",  "+15%", "Diversified global footprint"),
    ]

    for i, row_data in enumerate(geo):
        r = 19 + i
        bg = ALT_ROW if i % 2 == 0 else WHITE
        if i == len(geo) - 1:
            bg = SUBHDR_BG
        for j, val in enumerate(row_data):
            lbl(ws, r, 2+j, val, bg=bg)

    # Key Products & Value Propositions
    sub_hdr(ws, 26, 2, "Key Products & Value Propositions", span=6)

    products = [
        ("Azure",                "Cloud Infrastructure & PaaS",
         "200+ cloud services; #2 cloud globally; 40% growth; hosts OpenAI models"),
        ("Microsoft 365",        "Enterprise Productivity SaaS",
         "345M+ users; Teams, Word, Excel, Outlook, OneDrive; $30/mo Copilot add-on"),
        ("Copilot (AI)",         "AI Assistant across M365/Azure",
         "GPT-4 powered; integrated in Word, Excel, Teams; 3.5% penetration with huge upside"),
        ("Dynamics 365",         "ERP/CRM Business Applications",
         "+23% growth; competing with SAP and Salesforce in enterprise"),
        ("GitHub",               "Developer Platform",
         "100M+ developers; GitHub Copilot monetized; source code intelligence"),
        ("LinkedIn",             "Professional Network",
         "1B+ members; learning, hiring, marketing solutions; AI job matching"),
        ("Xbox / Gaming",        "Consumer Gaming Platform",
         "Series X/S; Game Pass; Activision Blizzard integration"),
        ("Bing / Search",        "Search + Advertising",
         "AI-powered Bing; 21% growth in ads; still 3% global share vs Google 90%"),
    ]

    prod_hdrs = ["Product", "Category", "Key Value Proposition"]
    for i, h in enumerate(prod_hdrs):
        sub_hdr(ws, 27, 2+i, h)
    ws.merge_cells(start_row=27, start_column=4, end_row=27, end_column=7)

    for i, (prod, cat, vp) in enumerate(products):
        r = 28 + i
        bg = ALT_ROW if i % 2 == 0 else WHITE
        lbl(ws, r, 2, prod, bg=bg)
        lbl(ws, r, 3, cat, bg=bg)
        lbl(ws, r, 4, vp, bg=bg)
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=7)

    # Who Buys & How
    sub_hdr(ws, 38, 2, "Who Buys Microsoft Products & Buying Process", span=6)

    buyers = [
        ("Enterprise IT Buyers", "CIOs, CTOs, IT Procurement",
         "Multi-year enterprise agreements (EA); per-seat and consumption-based Azure pricing; MSFT field sales + partner channels"),
        ("SMBs",                 "Business Owners, Office Managers",
         "Monthly Microsoft 365 subscriptions through resellers and direct online; cloud migration support"),
        ("Developers",           "Software Engineers, DevOps",
         "GitHub, Azure Dev Tools, VS Code; freemium to paid conversion; consumption-based cloud billing"),
        ("Consumers",            "Individuals, Gamers, Students",
         "Microsoft 365 Personal/Family, Xbox Game Pass, Surface devices via retail and online"),
        ("Government/Public Sector", "Federal, State, Local, Defense",
         "FedRAMP-certified Azure Government; long-term contracts; DoD JEDI/JWCC wins"),
    ]

    buy_hdrs = ["Buyer Type", "Decision Makers", "Buying Process"]
    for i, h in enumerate(buy_hdrs):
        sub_hdr(ws, 39, 2+i, h)
    ws.merge_cells(start_row=39, start_column=4, end_row=39, end_column=7)

    for i, (bt, dm, bp) in enumerate(buyers):
        r = 40 + i
        bg = ALT_ROW if i % 2 == 0 else WHITE
        lbl(ws, r, 2, bt, bg=bg)
        lbl(ws, r, 3, dm, bg=bg)
        lbl(ws, r, 4, bp, bg=bg)
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=7)

    # Seasonality
    sub_hdr(ws, 47, 2, "Seasonality & Margin Structure by Segment", span=6)
    seas_data = [
        ("Fiscal Year",         "July 1 – June 30; Q4 (Apr-Jun) seasonally strongest quarter"),
        ("Productivity & BP",   "Relatively stable; enterprise renewals spread across year; Teams usage spikes in Q2 (Jan-Mar)"),
        ("Intelligent Cloud",   "Q3/Q4 highest as enterprises close annual cloud commitments; Azure consumption accelerates in H2"),
        ("More Personal Computing", "Q2 (Oct-Dec) peaks with holiday Xbox and Surface; Windows OEM tied to PC manufacturing cycles"),
        ("Gross Margin",        "Overall ~68.8%; Cloud ~70%+; Software near 80%; Hardware <30%"),
        ("Operating Margin",    "45.6% blended; Intelligent Cloud ~50%; Gaming margins dilutive to blended rate"),
    ]

    for i, (k, v) in enumerate(seas_data):
        r = 48 + i
        bg = ALT_ROW if i % 2 == 0 else WHITE
        wc(ws, r, 2, k, bold=True, size=FONT_SIZE, bg=bg, align="left", border=True)
        wc(ws, r, 3, v, size=FONT_SIZE, bg=bg, align="left", border=True, wrap=True)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)
        ws.row_dimensions[r].height = 22


# ═══════════════════════════════════════════════════════════════════════════════
def build_moat(wb):
    ws = wb.create_sheet("Moat")
    ws.sheet_view.showGridLines = False
    scw(ws, 1, 3); scw(ws, 2, 30); scw(ws, 3, 26); scw(ws, 4, 22)
    scw(ws, 5, 22); scw(ws, 6, 22); scw(ws, 7, 22)

    sec_hdr(ws, 1, 2, "COMPETITIVE MOAT ANALYSIS — MICROSOFT (MSFT)", span=6)

    sub_hdr(ws, 3, 2, "Moat Assessment: WIDE MOAT (Morningstar Rating: Wide)", span=6)

    intro = ("Microsoft possesses one of the widest competitive moats in the technology sector, built on "
             "overlapping sources: ecosystem lock-in, network effects, switching costs, intangible assets "
             "(IP, brands, data), and scale advantages. The emergence of AI as a competitive dimension "
             "has deepened—not eroded—Microsoft's moat, given its exclusive OpenAI partnership and "
             "$200B+ committed AI infrastructure investment.")
    ws.merge_cells("B4:G6")
    wc(ws, 4, 2, intro, size=FONT_SIZE, bg=LIGHT_GRAY, align="left", wrap=True, border=True)
    for r in range(4, 7):
        ws.row_dimensions[r].height = 22

    moat_sources = [
        ("1. Ecosystem Lock-in & Switching Costs",
         "Microsoft 365 + Azure integration creates a unified enterprise platform. Replacing Microsoft "
         "means re-training staff, migrating petabytes of SharePoint data, rebuilding Active Directory "
         "integrations, and re-engineering applications — costs that vastly exceed annual license fees. "
         "Enterprises typically spend 3-5x the annual Microsoft bill to migrate away.",
         "VERY HIGH"),
        ("2. OpenAI Exclusive Integration (AI Moat)",
         "GPT-4 and o1 reasoning models are only available natively through Azure OpenAI Service. "
         "Enterprises building AI applications on Azure access exclusive model capabilities unavailable "
         "on AWS or Google Cloud. This creates AI model lock-in on top of existing cloud infrastructure "
         "lock-in — compounding switching costs materially.",
         "HIGH"),
        ("3. Network Effects",
         "Microsoft Teams (300M+ DAU), LinkedIn (1B+ members), and GitHub (100M+ developers) all benefit "
         "from network effects where the platform becomes more valuable as more users join. Azure's developer "
         "ecosystem (millions of certified professionals) creates a self-reinforcing cycle of supply and demand.",
         "HIGH"),
        ("4. Intangible Assets (IP, Brands, Data)",
         "Microsoft owns 60,000+ patents, the Windows and Office brands (trusted by 1 billion users), "
         "and vast proprietary datasets (GitHub code repositories, LinkedIn professional data, Microsoft "
         "productivity telemetry) that train superior AI models. These data advantages compound over time.",
         "HIGH"),
        ("5. Scale & Capital Efficiency",
         "Azure's $200B+ physical infrastructure (data centers, fiber, silicon) creates cost advantages "
         "that prevent new entrants. Only AWS and Google can match Azure's scale. Microsoft's ability to "
         "amortize R&D ($29.5B/year) across billions of users creates structural cost advantages.",
         "HIGH"),
        ("6. Enterprise Relationships",
         "Microsoft has deep multi-decade relationships with virtually every Fortune 500 company through "
         "Enterprise Agreements, Microsoft's field sales organization (10,000+ enterprise account managers), "
         "and a 400,000+ partner ecosystem. These relationships are self-reinforcing and sticky.",
         "MODERATE-HIGH"),
    ]

    sub_hdr(ws, 8, 2, "Moat Source", bg=SUBHDR_BG)
    sub_hdr(ws, 8, 3, "Description", bg=SUBHDR_BG)
    sub_hdr(ws, 8, 4, "Strength", bg=SUBHDR_BG)
    ws.merge_cells(start_row=8, start_column=3, end_row=8, end_column=6)

    for i, (name, desc, strength) in enumerate(moat_sources):
        r = 9 + i * 2
        bg = ALT_ROW if i % 2 == 0 else WHITE
        wc(ws, r, 2, name, bold=True, size=FONT_SIZE, bg=bg, align="left", border=True, wrap=True)
        wc(ws, r, 3, desc, size=FONT_SIZE, bg=bg, align="left", border=True, wrap=True)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
        wc(ws, r, 7, strength, bold=True, size=FONT_SIZE, fg=MS_BLUE, bg=bg, align="center", border=True)
        ws.row_dimensions[r].height = 50

    # Competitive positioning
    row_start = 9 + len(moat_sources) * 2 + 1
    sub_hdr(ws, row_start, 2, "Competitive Positioning vs. Key Rivals", span=6)

    comp_hdrs = ["Dimension", "Microsoft (MSFT)", "Amazon AWS", "Google Cloud", "Assessment"]
    for i, h in enumerate(comp_hdrs):
        sub_hdr(ws, row_start+1, 2+i, h)

    competitors = [
        ("Cloud Market Share",      "24% (#2)",           "30% (#1)",  "12% (#3)",  "MSFT gaining share"),
        ("AI Platform",             "Azure + OpenAI (excl)", "AWS Bedrock", "Google Gemini", "MSFT differentiates via OpenAI exclusivity"),
        ("Enterprise Software",     "Dominant (M365, Dynamics)", "Limited", "G Suite (growing)", "MSFT dominant by far"),
        ("Developer Ecosystem",     "GitHub #1 (100M devs)", "AWS CodeBuild", "Google Cloud IDE", "MSFT leads with GitHub"),
        ("Revenue Growth (Cloud)",  "+34-40%",            "+17%",      "+28%",      "MSFT fastest growing vs AWS, below Google"),
        ("Margin Profile",          "45.6% op margin",    "~30% AWS",  "~16% GCP",  "MSFT most profitable"),
        ("Enterprise Relationships","Decades of trust",   "Growing",   "Growing",   "MSFT has structural advantage"),
    ]

    for i, row_data in enumerate(competitors):
        r = row_start + 2 + i
        bg = ALT_ROW if i % 2 == 0 else WHITE
        for j, val in enumerate(row_data):
            lbl(ws, r, 2+j, val, bg=bg)


# ═══════════════════════════════════════════════════════════════════════════════
def build_income(wb):
    ws = wb.create_sheet("Income Statements")
    ws.sheet_view.showGridLines = False
    scw(ws, 1, 3); scw(ws, 2, 34); scw(ws, 3, 17); scw(ws, 4, 17)
    scw(ws, 5, 17); scw(ws, 6, 17); scw(ws, 7, 17)

    sec_hdr(ws, 1, 2, "INCOME STATEMENTS — MICROSOFT CORPORATION (MSFT)", span=6)
    wc(ws, 2, 2, "All figures in USD Billions | Fiscal Year ends June 30",
       italic=True, size=12, fg="666666", align="left")

    years = ["FY2021", "FY2022", "FY2023", "FY2024", "FY2025"]
    sub_hdr(ws, 4, 2, "Metric / Year", bg=SUBHDR_BG)
    for i, y in enumerate(years):
        sub_hdr(ws, 4, 3+i, y, bg=SUBHDR_BG)

    # Income Statement Data
    is_data = [
        ("REVENUE", None, None, None, None, None),
        ("  Productivity & Business Processes", 46.2, 63.4, 69.3, 77.7, 120.9),
        ("  Intelligent Cloud",                 60.1, 75.3, 87.9, 105.4, 107.2),
        ("  More Personal Computing",           54.1, 59.6, 54.7, 62.0, 53.6),
        ("TOTAL REVENUE",                       168.1, 198.3, 211.9, 245.1, 281.7),
        ("", None, None, None, None, None),
        ("COSTS & EXPENSES", None, None, None, None, None),
        ("  Cost of Revenue",                   52.2, 62.7, 65.9, 74.1, 87.8),
        ("GROSS PROFIT",                        115.9, 135.6, 146.1, 171.0, 193.9),
        ("  Gross Margin %",                    "69.0%", "68.4%", "68.9%", "69.8%", "68.8%"),
        ("", None, None, None, None, None),
        ("  Research & Development",            20.7, 24.5, 27.2, 29.5, 31.4),
        ("  Sales & Marketing",                 20.1, 21.8, 22.8, 24.3, 26.0),
        ("  General & Administrative",          5.9, 5.9, 7.6, 7.9, 8.0),
        ("TOTAL OPERATING EXPENSES",            99.0, 114.9, 123.5, 135.7, 153.2),
        ("", None, None, None, None, None),
        ("OPERATING INCOME",                    69.9, 83.4, 88.5, 109.4, 128.5),
        ("  Operating Margin %",                "41.6%", "42.1%", "41.8%", "44.6%", "45.6%"),
        ("", None, None, None, None, None),
        ("  Other Income / (Expense), Net",     1.2, 0.3, 1.3, 3.3, 3.2),
        ("INCOME BEFORE TAX",                   71.1, 83.7, 89.8, 112.7, 131.7),
        ("  Income Tax Provision",              9.8, 10.9, 17.4, 24.6, 29.9),
        ("NET INCOME",                          61.3, 72.7, 72.4, 88.1, 101.8),
        ("  Net Margin %",                      "36.5%", "36.7%", "34.2%", "35.9%", "36.1%"),
        ("", None, None, None, None, None),
        ("PER SHARE DATA", None, None, None, None, None),
        ("  Diluted EPS ($)",                   8.05, 9.65, 9.72, 11.80, 13.70),
        ("  Weighted Avg Diluted Shares (B)",   7.62, 7.54, 7.45, 7.46, 7.43),
        ("  Dividends Per Share ($)",           2.24, 2.48, 2.72, 2.94, 3.32),
        ("", None, None, None, None, None),
        ("SEGMENT OPERATING METRICS", None, None, None, None, None),
        ("  Azure Revenue ($B, est.)",          "~$30B", "~$42B", "~$55B", "~$65B", "$75B+"),
        ("  Azure Growth YoY",                  "~50%", "~40%", "~29%", "~29%", "~34%"),
        ("  Microsoft 365 Commercial Cloud",    "~$18B", "~$22B", "~$28B", "~$36B", "~$42B"),
        ("  LinkedIn Revenue ($B)",             "~$9.4B", "~$13.8B", "~$15.1B", "~$17.0B", "~$18.5B"),
    ]

    for i, row_data in enumerate(is_data):
        r = 5 + i
        label = row_data[0]
        values = row_data[1:]

        # Style logic
        is_total = label.startswith("TOTAL") or label in ("NET INCOME", "GROSS PROFIT", "OPERATING INCOME", "INCOME BEFORE TAX")
        is_section = label.isupper() and not is_total and label != ""
        is_pct = "%" in str(label)
        is_empty = label == "" and all(v is None for v in values)

        if is_empty:
            ws.row_dimensions[r].height = 8
            continue

        if is_section:
            bg = DARK_BLUE
            fg = WHITE
            wc(ws, r, 2, label, bold=True, size=FONT_SIZE, fg=fg, bg=bg, align="left", border=True)
            for j in range(5):
                wc(ws, r, 3+j, "", bg=bg, border=True)
            ws.row_dimensions[r].height = 22
            continue

        if is_total:
            bg = SUBHDR_BG
            bold = True
        elif is_pct:
            bg = WARN_YELLOW
            bold = False
        else:
            bg = ALT_ROW if i % 2 == 0 else WHITE
            bold = False

        lbl(ws, r, 2, label, bg=bg)
        if bold:
            ws.cell(row=r, column=2).font = mf(bold=True, size=FONT_SIZE)

        for j, val in enumerate(values):
            if val is None:
                dc(ws, r, 3+j, "", bg=bg)
            elif isinstance(val, float) or isinstance(val, int):
                c = ws.cell(row=r, column=3+j, value=val)
                c.font = mf(bold=bold, size=FONT_SIZE)
                c.fill = mfill(bg)
                c.alignment = cal("right")
                c.border = mborder()
                c.number_format = '#,##0.0'
            else:
                dc(ws, r, 3+j, val, bg=bg, bold=bold)
        ws.row_dimensions[r].height = 20

    # YoY Growth Section
    last_data_row = 5 + len(is_data) + 2
    sub_hdr(ws, last_data_row, 2, "Year-over-Year Revenue Growth Analysis", span=6)
    yoy_data = [
        ("Productivity & BP Revenue Growth", "N/A", "+37.2%", "+9.3%", "+12.1%", "+55.6%"),
        ("Intelligent Cloud Revenue Growth",  "N/A", "+25.3%", "+16.7%", "+19.9%", "+1.7%"),
        ("More Personal Computing Growth",    "N/A", "+10.2%", "-8.2%", "+13.4%", "-13.5%"),
        ("Total Revenue Growth",              "N/A", "+18.0%", "+6.9%", "+15.7%", "+14.9%"),
        ("Net Income Growth",                 "N/A", "+18.6%", "-0.4%", "+21.7%", "+15.5%"),
        ("EPS Growth",                        "N/A", "+19.9%", "+0.7%", "+21.4%", "+16.1%"),
    ]

    yoy_hdrs = ["Metric", "FY2021", "FY2022", "FY2023", "FY2024", "FY2025"]
    for i, h in enumerate(yoy_hdrs):
        sub_hdr(ws, last_data_row+1, 2+i, h)

    for i, row_data in enumerate(yoy_data):
        r = last_data_row + 2 + i
        bg = ALT_ROW if i % 2 == 0 else WHITE
        lbl(ws, r, 2, row_data[0], bg=bg)
        for j in range(5):
            dc(ws, r, 3+j, row_data[j+1], bg=bg)


# ═══════════════════════════════════════════════════════════════════════════════
def build_balance_sheet(wb):
    ws = wb.create_sheet("Balance Sheet")
    ws.sheet_view.showGridLines = False
    scw(ws, 1, 3); scw(ws, 2, 34); scw(ws, 3, 18); scw(ws, 4, 18)
    scw(ws, 5, 18); scw(ws, 6, 18); scw(ws, 7, 18)

    sec_hdr(ws, 1, 2, "BALANCE SHEET — MICROSOFT CORPORATION (MSFT)", span=6)
    wc(ws, 2, 2, "All figures in USD Billions | As of June 30 each year",
       italic=True, size=12, fg="666666", align="left")

    years = ["FY2021", "FY2022", "FY2023", "FY2024", "FY2025"]
    sub_hdr(ws, 4, 2, "Line Item", bg=SUBHDR_BG)
    for i, y in enumerate(years):
        sub_hdr(ws, 4, 3+i, y, bg=SUBHDR_BG)

    bs_data = [
        ("ASSETS", None, None, None, None, None),
        ("Current Assets", None, None, None, None, None),
        ("  Cash & Cash Equivalents", 14.2, 13.9, 34.7, 18.3, 30.2),
        ("  Short-term Investments", 116.4, 90.8, 76.6, 76.5, 68.3),
        ("  Accounts Receivable, Net", 38.0, 44.3, 48.7, 54.0, 55.5),
        ("  Other Current Assets", 11.6, 14.2, 17.0, 19.1, 20.0),
        ("TOTAL CURRENT ASSETS", 184.4, 169.7, 184.4, 171.9, 174.0),
        ("", None, None, None, None, None),
        ("Non-Current Assets", None, None, None, None, None),
        ("  Property, Plant & Equip., Net", 59.7, 74.6, 95.6, 131.4, 189.0),
        ("  Operating Lease ROU Assets", 13.1, 14.3, 14.6, 17.4, 20.0),
        ("  Goodwill", 67.9, 67.5, 67.9, 119.8, 120.5),
        ("  Intangible Assets, Net", 7.6, 8.5, 9.4, 18.5, 16.0),
        ("  Long-term Investments", 5.9, 6.9, 9.9, 13.1, 14.0),
        ("  Other Non-Current Assets", 26.5, 29.3, 35.2, 48.4, 85.5),
        ("TOTAL NON-CURRENT ASSETS", 174.2, 196.5, 232.8, 348.6, 445.0),
        ("", None, None, None, None, None),
        ("TOTAL ASSETS", 333.8, 364.8, 411.9, 523.0, 619.0),
        ("", None, None, None, None, None),
        ("LIABILITIES & EQUITY", None, None, None, None, None),
        ("Current Liabilities", None, None, None, None, None),
        ("  Accounts Payable", 15.1, 19.0, 19.0, 24.5, 26.0),
        ("  Short-term Debt", 8.1, 2.8, 5.2, 3.2, 3.1),
        ("  Deferred Revenue (Current)", 41.7, 45.0, 50.9, 57.5, 64.0),
        ("  Other Current Liabilities", 24.9, 31.6, 38.0, 43.6, 47.0),
        ("TOTAL CURRENT LIABILITIES", 88.7, 95.1, 115.1, 125.3, 140.1),
        ("", None, None, None, None, None),
        ("Non-Current Liabilities", None, None, None, None, None),
        ("  Long-term Debt", 50.1, 47.0, 41.9, 38.5, 37.3),
        ("  Deferred Revenue (LT)", 3.5, 3.7, 4.5, 5.0, 5.5),
        ("  Operating Lease Liabilities (LT)", 11.8, 12.7, 12.7, 15.3, 18.5),
        ("  Other Non-Current Liabilities", 30.7, 27.1, 31.7, 30.2, 72.5),
        ("TOTAL NON-CURRENT LIABILITIES", 96.1, 90.5, 90.8, 89.0, 133.8),
        ("", None, None, None, None, None),
        ("TOTAL LIABILITIES", 191.8, 198.3, 213.5, 225.2, 273.9),
        ("", None, None, None, None, None),
        ("STOCKHOLDERS' EQUITY", None, None, None, None, None),
        ("  Common Stock + APIC", 86.9, 86.1, 93.5, 101.2, 108.0),
        ("  Retained Earnings (Accum. Deficit)", 57.0, 84.3, 118.8, 202.0, 242.5),
        ("  Treasury Stock", None, None, None, None, None),
        ("  AOCI",            -1.9, -3.9, -5.9, -5.4, -5.4),
        ("TOTAL STOCKHOLDERS' EQUITY", 142.0, 166.5, 206.2, 297.8, 345.1),
        ("", None, None, None, None, None),
        ("TOTAL LIABILITIES & EQUITY", 333.8, 364.8, 411.9, 523.0, 619.0),
        ("", None, None, None, None, None),
        ("KEY BALANCE SHEET METRICS", None, None, None, None, None),
        ("  Net Cash (Cash + ST Inv - All Debt)", 72.3, 54.9, 64.2, 53.1, 58.1),
        ("  Debt-to-Equity Ratio", "0.40x", "0.30x", "0.23x", "0.14x", "0.12x"),
        ("  Current Ratio", "2.08x", "1.78x", "1.60x", "1.37x", "1.24x"),
        ("  Book Value Per Share ($)", 18.7, 22.1, 27.7, 40.0, 46.5),
        ("  Goodwill as % of Total Assets", "20.3%", "18.5%", "16.5%", "22.9%", "19.5%"),
    ]

    for i, row_data in enumerate(bs_data):
        r = 5 + i
        label = row_data[0]
        values = row_data[1:]

        is_total = label.startswith("TOTAL") or label in ("STOCKHOLDERS' EQUITY",)
        is_section = label.isupper() and not label.startswith("TOTAL") and not label.startswith("KEY") and label not in ("ASSETS", "LIABILITIES & EQUITY") and "EQUITY" not in label
        is_subsection = any(label.startswith(x) for x in ("Current Assets", "Non-Current Assets", "Current Liabilities", "Non-Current Liabilities"))
        is_metrics = label.startswith("KEY")
        is_empty = label == "" and all(v is None for v in values)

        if is_empty:
            ws.row_dimensions[r].height = 8
            continue

        if label in ("ASSETS", "LIABILITIES & EQUITY", "STOCKHOLDERS' EQUITY", "KEY BALANCE SHEET METRICS"):
            wc(ws, r, 2, label, bold=True, size=FONT_SIZE, fg=WHITE, bg=DARK_BLUE, align="left", border=True)
            for j in range(5):
                wc(ws, r, 3+j, "", bg=DARK_BLUE, border=True)
            ws.row_dimensions[r].height = 22
            continue

        if is_subsection:
            bg = "D6EAF8"
            wc(ws, r, 2, label, bold=True, size=FONT_SIZE, bg=bg, align="left", border=True)
            for j in range(5):
                wc(ws, r, 3+j, "", bg=bg, border=True)
            ws.row_dimensions[r].height = 20
            continue

        if is_total or is_metrics:
            bg = SUBHDR_BG
            bold = True
        else:
            bg = ALT_ROW if i % 2 == 0 else WHITE
            bold = False

        lbl(ws, r, 2, label, bg=bg)
        if bold:
            ws.cell(row=r, column=2).font = mf(bold=True, size=FONT_SIZE)

        for j, val in enumerate(values):
            if val is None:
                dc(ws, r, 3+j, "", bg=bg)
            elif isinstance(val, (float, int)):
                c = ws.cell(row=r, column=3+j, value=val)
                c.font = mf(bold=bold, size=FONT_SIZE)
                c.fill = mfill(bg)
                c.alignment = cal("right")
                c.border = mborder()
                c.number_format = '#,##0.0'
            else:
                dc(ws, r, 3+j, val, bg=bg, bold=bold)
        ws.row_dimensions[r].height = 20


# ═══════════════════════════════════════════════════════════════════════════════
def build_cashflow(wb):
    ws = wb.create_sheet("Cash Flow Analysis")
    ws.sheet_view.showGridLines = False
    scw(ws, 1, 3); scw(ws, 2, 36); scw(ws, 3, 17); scw(ws, 4, 17)
    scw(ws, 5, 17); scw(ws, 6, 17); scw(ws, 7, 17)

    sec_hdr(ws, 1, 2, "CASH FLOW ANALYSIS — MICROSOFT CORPORATION (MSFT)", span=6)
    wc(ws, 2, 2, "All figures in USD Billions | Fiscal Year ends June 30",
       italic=True, size=12, fg="666666", align="left")

    years = ["FY2021", "FY2022", "FY2023", "FY2024", "FY2025"]
    sub_hdr(ws, 4, 2, "Line Item", bg=SUBHDR_BG)
    for i, y in enumerate(years):
        sub_hdr(ws, 4, 3+i, y, bg=SUBHDR_BG)

    cf_data = [
        ("OPERATING ACTIVITIES", None, None, None, None, None),
        ("  Net Income",                              61.3, 72.7, 72.4, 88.1, 101.8),
        ("  Depreciation & Amortization",             11.0, 14.7, 18.0, 22.3, 31.0),
        ("  Stock-Based Compensation",                6.1, 7.5, 9.6, 10.7, 11.0),
        ("  Changes in Working Capital",             -1.5, -4.6, 3.3, 4.5, -7.0),
        ("  Other Operating Adjustments",            -1.1, -1.7, -3.2, -4.4, -0.8),
        ("OPERATING CASH FLOW",                       76.7, 89.0, 87.6, 118.5, 136.0),
        ("  OCF Margin %",                           "45.6%", "44.9%", "41.3%", "48.3%", "48.3%"),
        ("", None, None, None, None, None),
        ("INVESTING ACTIVITIES", None, None, None, None, None),
        ("  Capital Expenditures",                   -20.6, -23.9, -28.1, -44.5, -64.6),
        ("  Acquisitions, Net",                      -8.7, -22.0, -68.7, -1.4, -3.2),
        ("  Purchases of Investments",               -77.6, -74.5, -62.0, -74.7, -70.0),
        ("  Maturities / Sales of Investments",      67.4, 89.5, 63.1, 72.5, 65.0),
        ("  Other Investing Activities",             -0.3, -1.5, -2.1, 1.7, -0.5),
        ("NET INVESTING ACTIVITIES",                 -39.8, -32.4, -97.8, -46.4, -73.3),
        ("", None, None, None, None, None),
        ("FINANCING ACTIVITIES", None, None, None, None, None),
        ("  Share Repurchases",                      -27.4, -32.7, -22.2, -23.4, -22.5),
        ("  Dividends Paid",                         -17.3, -18.7, -20.0, -21.8, -24.0),
        ("  Debt Issuance / (Repayment), Net",       -4.5, -10.9, -2.6, -2.8, -1.5),
        ("  Other Financing Activities",             -2.8, -3.4, -4.2, -2.1, -1.0),
        ("NET FINANCING ACTIVITIES",                 -52.0, -65.7, -49.0, -50.1, -49.0),
        ("", None, None, None, None, None),
        ("NET CHANGE IN CASH",                       -15.1, -9.1, -59.2, 22.0, 13.7),
        ("  Beginning Cash Balance",                 13.6, 14.2, 13.9, 34.7, 18.3),
        ("  Ending Cash Balance",                    14.2, 13.9, 34.7, 18.3, 30.2),
        ("", None, None, None, None, None),
        ("FREE CASH FLOW ANALYSIS", None, None, None, None, None),
        ("  Operating Cash Flow",                    76.7, 89.0, 87.6, 118.5, 136.0),
        ("  Less: Capital Expenditures",             -20.6, -23.9, -28.1, -44.5, -64.6),
        ("FREE CASH FLOW (FCF)",                     56.1, 65.1, 59.5, 74.0, 71.4),
        ("  FCF Margin %",                           "33.4%", "32.8%", "28.1%", "30.2%", "25.3%"),
        ("  FCF Conversion (FCF/Net Income)",        "91.5%", "89.5%", "82.2%", "84.0%", "70.1%"),
        ("", None, None, None, None, None),
        ("CAPITAL ALLOCATION", None, None, None, None, None),
        ("  Share Repurchases ($B)",                 27.4, 32.7, 22.2, 23.4, 22.5),
        ("  Dividends ($B)",                         17.3, 18.7, 20.0, 21.8, 24.0),
        ("  Total Returned to Shareholders ($B)",    44.7, 51.4, 42.2, 45.2, 46.5),
        ("  % of FCF Returned",                      "79.7%", "78.9%", "70.9%", "61.1%", "65.1%"),
        ("  Capex as % of Revenue",                  "12.3%", "12.1%", "13.3%", "18.2%", "22.9%"),
        ("  Capex Growth YoY",                       "N/A", "+16.0%", "+17.6%", "+58.4%", "+45.2%"),
    ]

    for i, row_data in enumerate(cf_data):
        r = 5 + i
        label = row_data[0]
        values = row_data[1:]

        is_total = any(label.startswith(x) for x in ("OPERATING CASH FLOW", "NET INVESTING", "NET FINANCING",
                                                       "FREE CASH FLOW (FCF)", "NET CHANGE", "TOTAL RETURNED"))
        is_section = label.isupper() and label not in ("", "OPERATING CASH FLOW") and not is_total
        is_pct = "%" in str(label)
        is_empty = label == "" and all(v is None for v in values)

        if is_empty:
            ws.row_dimensions[r].height = 8
            continue

        if is_section:
            wc(ws, r, 2, label, bold=True, size=FONT_SIZE, fg=WHITE, bg=DARK_BLUE, align="left", border=True)
            for j in range(5):
                wc(ws, r, 3+j, "", bg=DARK_BLUE, border=True)
            ws.row_dimensions[r].height = 22
            continue

        if is_total:
            bg = SUBHDR_BG
            bold = True
        elif is_pct:
            bg = WARN_YELLOW
            bold = False
        else:
            bg = ALT_ROW if i % 2 == 0 else WHITE
            bold = False

        lbl(ws, r, 2, label, bg=bg)
        if bold:
            ws.cell(row=r, column=2).font = mf(bold=True, size=FONT_SIZE)

        for j, val in enumerate(values):
            if val is None:
                dc(ws, r, 3+j, "", bg=bg)
            elif isinstance(val, (float, int)):
                c = ws.cell(row=r, column=3+j, value=val)
                c.font = mf(bold=bold, size=FONT_SIZE)
                c.fill = mfill(bg)
                c.alignment = cal("right")
                c.border = mborder()
                c.number_format = '#,##0.0'
            else:
                dc(ws, r, 3+j, val, bg=bg, bold=bold)
        ws.row_dimensions[r].height = 20


# ═══════════════════════════════════════════════════════════════════════════════
def build_roic(wb):
    ws = wb.create_sheet("Return on Capital")
    ws.sheet_view.showGridLines = False
    scw(ws, 1, 3); scw(ws, 2, 36); scw(ws, 3, 17); scw(ws, 4, 17)
    scw(ws, 5, 17); scw(ws, 6, 17); scw(ws, 7, 17)

    sec_hdr(ws, 1, 2, "RETURN ON CAPITAL — MICROSOFT CORPORATION (MSFT)", span=6)
    wc(ws, 2, 2, "All figures in USD Billions unless noted | Thinking like an owner",
       italic=True, size=12, fg="666666", align="left")

    years = ["FY2021", "FY2022", "FY2023", "FY2024", "FY2025"]
    sub_hdr(ws, 4, 2, "Metric", bg=SUBHDR_BG)
    for i, y in enumerate(years):
        sub_hdr(ws, 4, 3+i, y, bg=SUBHDR_BG)

    roic_data = [
        ("PROFITABILITY RETURNS", None, None, None, None, None),
        ("  Return on Equity (ROE)",          "43.2%", "43.7%", "35.1%", "29.6%", "29.5%"),
        ("  Return on Assets (ROA)",          "18.4%", "19.9%", "17.6%", "16.8%", "16.5%"),
        ("  Return on Invested Capital (ROIC)","28.3%", "30.2%", "27.0%", "25.0%", "24.5%"),
        ("  Return on Capital Employed",      "35.1%", "36.0%", "31.2%", "29.5%", "29.8%"),
        ("", None, None, None, None, None),
        ("INCREMENTAL ROIC ANALYSIS", None, None, None, None, None),
        ("  Revenue Growth ($B)",             "N/A", 30.2, 13.6, 33.2, 36.6),
        ("  Incremental Gross Profit ($B)",   "N/A", 19.7, 10.5, 24.9, 22.9),
        ("  Incremental Op Income ($B)",      "N/A", 13.5, 5.1, 20.9, 19.1),
        ("  Incremental ROIC",                "N/A", "44.7%", "37.5%", "63.0%", "52.2%"),
        ("  Capital Deployed (CapEx + Acq)",  29.3, 45.9, 96.8, 45.9, 67.8),
        ("", None, None, None, None, None),
        ("MARGINAL RETURNS ON AI INVESTMENT", None, None, None, None, None),
        ("  AI Capex (est. % of total Capex)", "~20%", "~25%", "~35%", "~55%", "~65%"),
        ("  Implied AI Capex ($B)",           "~4.1B", "~6.0B", "~9.8B", "~24.5B", "~42.0B"),
        ("  Azure Revenue Uplift (AI contrib)","N/A", "N/A", "~$5B", "~$12B", "~$22B"),
        ("  Implied AI Capex Payback",        "N/A", "N/A", "2-3 yrs", "2 yrs", "2-3 yrs"),
        ("", None, None, None, None, None),
        ("ECONOMIC VALUE ADDED", None, None, None, None, None),
        ("  NOPAT (Net Op Profit After Tax)",  56.1, 65.6, 67.5, 83.1, 97.7),
        ("  Invested Capital ($B)",            178.3, 192.5, 215.0, 306.0, 376.0),
        ("  WACC (est.)",                      "8.0%", "8.5%", "8.3%", "8.0%", "7.8%"),
        ("  Cost of Capital ($B)",            14.3, 16.4, 17.8, 24.5, 29.3),
        ("  Economic Value Added (EVA, $B)",  41.8, 49.2, 49.7, 58.6, 68.4),
        ("", None, None, None, None, None),
        ("RULE OF 50 ANALYSIS (SaaS Check)", None, None, None, None, None),
        ("  Revenue Growth Rate",              "N/A", "18.0%", "6.9%", "15.7%", "14.9%"),
        ("  FCF Margin",                       "33.4%", "32.8%", "28.1%", "30.2%", "25.3%"),
        ("  Rule of 50 Score (Growth+FCF%)",   "N/A", "50.8%", "35.0%", "45.9%", "40.2%"),
        ("  Assessment",                       "N/A", "PASS (>50)", "BELOW 50", "BELOW 50", "BELOW 50"),
        ("  Note: FCF margin lower due to surge in AI infrastructure CapEx ($64.6B)",
         None, None, None, None, None),
        ("  Normalized FCF margin (excl. AI growth CapEx) est. ~35%+ → Rule of 50 would PASS",
         None, None, None, None, None),
    ]

    for i, row_data in enumerate(roic_data):
        r = 5 + i
        label = row_data[0]
        values = row_data[1:]

        is_section = label.isupper() and label != "" and all(v is None for v in values)
        is_empty = label == "" and all(v is None for v in values)
        is_note = label.strip().startswith("Note:")
        is_normalized = "Normalized" in label

        if is_empty:
            ws.row_dimensions[r].height = 8
            continue

        if is_section:
            wc(ws, r, 2, label, bold=True, size=FONT_SIZE, fg=WHITE, bg=DARK_BLUE, align="left", border=True)
            for j in range(5):
                wc(ws, r, 3+j, "", bg=DARK_BLUE, border=True)
            ws.row_dimensions[r].height = 22
            continue

        if is_note or is_normalized:
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
            wc(ws, r, 2, label, italic=True, size=12, fg="555555", bg=WARN_YELLOW,
               align="left", border=True, wrap=True)
            ws.row_dimensions[r].height = 22
            continue

        is_total = any(x in label for x in ("Economic Value Added", "NOPAT", "EVA"))
        bg = SUBHDR_BG if is_total else (ALT_ROW if i % 2 == 0 else WHITE)
        bold = is_total

        lbl(ws, r, 2, label, bg=bg)
        if bold:
            ws.cell(row=r, column=2).font = mf(bold=True, size=FONT_SIZE)

        for j, val in enumerate(values):
            if val is None:
                dc(ws, r, 3+j, "", bg=bg)
            elif isinstance(val, (float, int)):
                c = ws.cell(row=r, column=3+j, value=val)
                c.font = mf(bold=bold, size=FONT_SIZE)
                c.fill = mfill(bg)
                c.alignment = cal("right")
                c.border = mborder()
                c.number_format = '#,##0.0'
            else:
                dc(ws, r, 3+j, val, bg=bg, bold=bold)
        ws.row_dimensions[r].height = 20


# ═══════════════════════════════════════════════════════════════════════════════
def build_management(wb):
    ws = wb.create_sheet("Management")
    ws.sheet_view.showGridLines = False
    scw(ws, 1, 3); scw(ws, 2, 32); scw(ws, 3, 26)
    scw(ws, 4, 26); scw(ws, 5, 26); scw(ws, 6, 20); scw(ws, 7, 20)

    sec_hdr(ws, 1, 2, "MANAGEMENT ANALYSIS — MICROSOFT CORPORATION (MSFT)", span=6)

    # Key Leadership
    sub_hdr(ws, 3, 2, "Key Leadership Team", span=6)

    leaders = [
        ("Satya Nadella",  "Chairman & CEO",  "Joined 1992; CEO since Feb 2014. Led transformation from PC-centric to cloud-first company. Architect of $75B Azure business. Board Chair since 2021.",  "$96.5M total (FY2025, +22% YoY)"),
        ("Amy Hood",       "EVP & CFO",       "Joined 2002; CFO since 2013. Oversaw Microsoft's capital allocation through cloud transformation. Key architect of shareholder return program.",         "$29.5M total (FY2025)"),
        ("Brad Smith",     "Vice Chair & President", "Chief legal officer; oversees government affairs and policy. Manages EU/US regulatory relationships.",                                          "$26M total (FY2025)"),
        ("Kevin Scott",    "EVP & CTO",       "Chief Technology Officer; oversees technical strategy, AI infrastructure, and OpenAI partnership execution.",                                          "$23M total (FY2025)"),
        ("Judson Althoff", "EVP & Chief Commercial Officer", "Leads Microsoft's commercial sales globally; oversees enterprise and SMB go-to-market.",                                             "$18M total (FY2025)"),
    ]

    ldr_hdrs = ["Name", "Title", "Background & Role", "FY2025 Compensation"]
    for i, h in enumerate(ldr_hdrs):
        sub_hdr(ws, 4, 2+i, h)

    for i, (name, title, bg_text, comp) in enumerate(leaders):
        r = 5 + i
        bg = ALT_ROW if i % 2 == 0 else WHITE
        lbl(ws, r, 2, name, bg=bg)
        lbl(ws, r, 3, title, bg=bg)
        lbl(ws, r, 4, bg_text, bg=bg)
        wc(ws, r, 4, bg_text, size=FONT_SIZE, bg=bg, align="left", border=True, wrap=True)
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=5)
        lbl(ws, r, 6, comp, bg=bg)
        ws.row_dimensions[r].height = 36

    # Compensation Philosophy
    sub_hdr(ws, 12, 2, "Executive Compensation Philosophy (From 2025 Proxy)", span=6)

    comp_items = [
        ("Pay-for-Performance",     ">95% of CEO compensation is performance-based; >50% for other executives. Short-term cash incentives tied to revenue, operating income growth, and ESG metrics."),
        ("Stock-Linked Incentives", "Long-term equity awards (stock awards, performance-based shares) tied to relative TSR vs. S&P 500 and specific business metrics including Azure growth, M365 adoption."),
        ("Security Metric Added",   "FY2025 first time security included as executive compensation metric — response to Microsoft's 2023-24 high-profile security incidents and regulatory pressure."),
        ("Alignment with Shareholders", "Executives must hold 6x annual base salary in MSFT stock. CEO Nadella holds $840M+ in MSFT shares personally, ensuring strong alignment."),
        ("Clawback Policy",         "Robust clawback provisions allow board to recover compensation in cases of financial restatements or misconduct."),
    ]

    for i, (k, v) in enumerate(comp_items):
        r = 13 + i
        bg = ALT_ROW if i % 2 == 0 else WHITE
        wc(ws, r, 2, k, bold=True, size=FONT_SIZE, bg=bg, align="left", border=True)
        wc(ws, r, 3, v, size=FONT_SIZE, bg=bg, align="left", border=True, wrap=True)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)
        ws.row_dimensions[r].height = 36

    # Capital Allocation Track Record
    sub_hdr(ws, 20, 2, "Capital Allocation Decisions", span=6)

    cap_alloc = [
        ("Cloud Transformation (2014-2019)",  "Nadella pivoted Microsoft from Windows/Office licenses to Azure+SaaS. Sold Nokia mobile ($7.6B write-off — accepting past mistake). Acquired LinkedIn ($26.2B), GitHub ($7.5B) at decisive moments."),
        ("Activision Blizzard Acquisition",   "Acquired for $68.7B (FY2023) — largest gaming deal in history. Expanding Xbox Game Pass to 100M+ subscribers. Strategic bet on gaming/metaverse, though margin-dilutive near term."),
        ("OpenAI Partnership ($13B)",         "Multi-year $13B investment in OpenAI (2019-2023); exclusive Azure hosting rights for GPT-4 and o1 models. Best ROI decision in MSFT history — creates AI moat no competitor can replicate."),
        ("AI Infrastructure CapEx ($64.6B)",  "FY2025 CapEx surged 45% to $64.6B — massive data center buildout for AI workloads. Management views as 'planting seeds'; initial payback visible in Azure 40% growth, though FCF margin temporarily compressed."),
        ("Shareholder Returns ($46.5B/yr)",   "Consistently returns $44-50B annually through buybacks and dividends. Has repurchased $200B+ in shares since 2013. Dividend raised every year since 2003 — 20+ consecutive years of increases."),
        ("Leverage Policy",                   "Microsoft maintains AAA credit rating (only 2 US companies). Net cash position $58B+. Deliberately under-leveraged — management prioritizes financial flexibility for strategic M&A over financial engineering."),
    ]

    for i, (k, v) in enumerate(cap_alloc):
        r = 21 + i
        bg = ALT_ROW if i % 2 == 0 else WHITE
        wc(ws, r, 2, k, bold=True, size=FONT_SIZE, bg=bg, align="left", border=True, wrap=True)
        wc(ws, r, 3, v, size=FONT_SIZE, bg=bg, align="left", border=True, wrap=True)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)
        ws.row_dimensions[r].height = 42

    # Owner-Operator Assessment
    sub_hdr(ws, 29, 2, "Does Nadella Act Like an Owner?", span=6)

    owner_criteria = [
        ("Long-term Orientation", "YES", "Consistently articulates 10+ year vision. Azure, OpenAI investments had multi-year payback horizons; not driven by quarterly targets."),
        ("Skin in the Game",      "YES", "Holds $840M+ in MSFT shares; significant personal wealth tied to MSFT performance."),
        ("Capital Discipline",    "YES", "Maintained AAA rating; refused excessive leverage. But $64.6B CapEx requires monitoring — must monetize AI investments."),
        ("Planting Seeds",        "YES", "OpenAI bet, Azure datacenter buildout, Copilot are all 'seeds' with 3-5 year payback horizons. Not borrowing from the future."),
        ("Admitting Mistakes",    "YES", "Wrote off Nokia mobile ($7.6B) and pivoted decisively. Publicly acknowledged security failures in 2024."),
        ("Insider Trading",       "NEUTRAL", "Nadella has sold shares as part of pre-planned 10b5-1 plans (tax diversification) but retains massive stake. Net holdings growing with stock awards."),
    ]

    own_hdrs = ["Criteria", "Assessment", "Evidence"]
    for i, h in enumerate(own_hdrs):
        sub_hdr(ws, 30, 2+i, h)
    ws.merge_cells(start_row=30, start_column=4, end_row=30, end_column=7)

    for i, (crit, assess, ev) in enumerate(owner_criteria):
        r = 31 + i
        bg = ALT_ROW if i % 2 == 0 else WHITE
        lbl(ws, r, 2, crit, bg=bg)
        color = "27AE60" if assess == "YES" else ("E67E22" if assess == "NEUTRAL" else "C0392B")
        wc(ws, r, 3, assess, bold=True, size=FONT_SIZE, fg=color, bg=bg, align="center", border=True)
        lbl(ws, r, 4, ev, bg=bg)
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=7)
        ws.row_dimensions[r].height = 28


# ═══════════════════════════════════════════════════════════════════════════════
def build_risks(wb):
    ws = wb.create_sheet("Risks")
    ws.sheet_view.showGridLines = False
    scw(ws, 1, 3); scw(ws, 2, 30); scw(ws, 3, 16); scw(ws, 4, 16)
    scw(ws, 5, 30); scw(ws, 6, 20); scw(ws, 7, 20)

    sec_hdr(ws, 1, 2, "RISK ANALYSIS — MICROSOFT CORPORATION (MSFT)", span=6)

    risk_hdrs = ["Risk Factor", "Likelihood", "Impact", "Description", "Mitigation"]
    for i, h in enumerate(risk_hdrs):
        sub_hdr(ws, 3, 2+i, h)

    risks = [
        ("AI CapEx Monetization Risk",
         "MEDIUM", "HIGH",
         "Microsoft invested $64.6B in CapEx (FY2025, +45%). If AI revenue growth disappoints (e.g. Copilot adoption stays at 3.5%), FCF margins could remain depressed and investors could reprice the stock.",
         "Azure AI already growing 40%; expanding Copilot to M365 base ($30/seat adds $5-10B ARR potential)"),
        ("Regulatory / Antitrust Risk",
         "MEDIUM-HIGH", "MEDIUM",
         "FTC probe into 'cloud tying' (bundling AI/security with Azure). EU AI Act compliance costs. OpenAI partnership scrutiny. History: EU fined Microsoft repeatedly for bundling.",
         "Microsoft has AAA government relationships; Brad Smith (President) manages regulatory risk as core competency"),
        ("OpenAI Relationship Risk",
         "LOW-MEDIUM", "VERY HIGH",
         "OpenAI may pursue IPO or strategic diversification, reducing Microsoft's exclusivity advantage. GPT-5+ models may be distributed more broadly. Sam Altman departure scenario re-emerged in 2023.",
         "$13B+ investment + exclusive Azure hosting makes defection extremely costly for OpenAI; Microsoft holds board observer seat"),
        ("Cloud Competition (AWS/Google)",
         "HIGH", "MEDIUM",
         "AWS remains #1 with 30% market share; Google Cloud growing faster (28%+) in AI-native workloads. Google Gemini directly competes with Copilot in enterprise. Neoclouds (CoreWeave) threatening Azure GPU share.",
         "Azure differentiated by enterprise trust + M365 integration + OpenAI exclusivity; Azure share gains confirmed by management"),
        ("Cybersecurity / Reputation Risk",
         "MEDIUM", "HIGH",
         "Microsoft suffered multiple high-profile breaches 2023-24 (including nation-state attack on Exchange Online). US government review found security lapses. A major breach could damage enterprise trust.",
         "Microsoft Secure Future Initiative launched; security now a compensation metric; $1B+/year incremental security investment"),
        ("Gaming / Activision Integration Risk",
         "MEDIUM", "LOW-MEDIUM",
         "$68.7B Activision deal is margin-dilutive. Xbox hardware business structurally declining vs Sony PS5. Game Pass subscriber growth must accelerate to justify acquisition premium.",
         "Activision titles (COD, Overwatch) drive Game Pass adoption; mobile gaming (King) provides growing revenue"),
        ("Macroeconomic / IT Spending Slowdown",
         "MEDIUM", "MEDIUM",
         "Enterprise IT budgets sensitive to recession. Azure is consumption-based — revenue can slow if customers optimize cloud spend. Windows OEM directly tied to PC market cycles.",
         "Azure 'optimization cycles' transitory; cloud is now mission-critical capex not discretionary opex; M365 very sticky"),
        ("AI Commoditization",
         "HIGH", "MEDIUM",
         "Open-source LLMs (Meta Llama, Mistral) are closing the performance gap with GPT-4, potentially reducing enterprise willingness to pay Copilot premiums. DeepSeek disruption in Jan 2026.",
         "Microsoft's moat is integration (Copilot in Office, Azure AI) not model superiority alone; network effects and data moat persist"),
        ("Concentration Risk (Azure)",
         "LOW", "HIGH",
         "Increasing revenue concentration in Azure (>38% of revenue) creates dependency on cloud infrastructure spending. Any major Azure outage could have significant revenue/reputational impact.",
         "Azure uptime SLAs (99.99%); geographically distributed data centers; Azure outages historically brief and recovering"),
    ]

    for i, (risk, likelihood, impact, desc, mit) in enumerate(risks):
        r = 4 + i
        bg = ALT_ROW if i % 2 == 0 else WHITE

        if impact in ("HIGH", "VERY HIGH"):
            imp_color = "C0392B"
        elif impact == "MEDIUM":
            imp_color = "E67E22"
        else:
            imp_color = "27AE60"

        lbl(ws, r, 2, risk, bg=bg)
        wc(ws, r, 3, likelihood, bold=False, size=FONT_SIZE, bg=bg, align="center", border=True)
        wc(ws, r, 4, impact, bold=True, size=FONT_SIZE, fg=imp_color, bg=bg, align="center", border=True)
        lbl(ws, r, 5, desc, bg=bg)
        wc(ws, r, 5, desc, size=FONT_SIZE, bg=bg, align="left", border=True, wrap=True)
        lbl(ws, r, 6, mit, bg=bg)
        wc(ws, r, 6, mit, size=FONT_SIZE, bg=bg, align="left", border=True, wrap=True)
        ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=7)
        ws.row_dimensions[r].height = 50


# ═══════════════════════════════════════════════════════════════════════════════
def build_valuation(wb):
    ws = wb.create_sheet("Valuation")
    ws.sheet_view.showGridLines = False
    scw(ws, 1, 3); scw(ws, 2, 34); scw(ws, 3, 22); scw(ws, 4, 22)
    scw(ws, 5, 22); scw(ws, 6, 22); scw(ws, 7, 22)

    sec_hdr(ws, 1, 2, "VALUATION ANALYSIS — MICROSOFT CORPORATION (MSFT)", span=6)
    wc(ws, 2, 2, "Stock Price: ~$373 (April 13, 2026) | Market Cap: ~$2.77T | Fiscal Year: June 30",
       italic=True, size=12, fg="666666", align="left")

    # Current Multiples
    sub_hdr(ws, 4, 2, "Current Trading Multiples vs. Historical Average", span=6)

    mult_hdrs = ["Multiple", "Current (Apr 2026)", "5-Year Historical Avg", "Industry Avg", "Assessment"]
    for i, h in enumerate(mult_hdrs):
        sub_hdr(ws, 5, 2+i, h)

    multiples = [
        ("Price / Earnings (TTM)",          "~23x",  "~33x",  "~28x (Tech)",    "CHEAP vs. history — lowest in 3 years"),
        ("Forward P/E (FY2026E)",            "~21x",  "~29x",  "~25x (Tech)",    "Significant discount to historical average"),
        ("EV / EBITDA",                      "~18x",  "~24x",  "~22x (Tech)",    "Attractive vs. peers"),
        ("Price / Free Cash Flow",           "~39x",  "~35x",  "~30x (Tech)",    "Premium due to CapEx cycle; normalizes to ~25x"),
        ("Price / Sales",                    "~9.8x", "~11x",  "~8x (Tech)",     "Fair; cloud businesses command premium"),
        ("PEG Ratio (Forward)",              "~1.4x", "~2.0x", "~1.8x",          "Attractive — growth not fully priced in"),
        ("EV / Revenue",                     "~10x",  "~12x",  "~9x",            "Reasonable for 15%+ revenue growth"),
        ("Dividend Yield",                   "~0.9%", "~1.0%", "~1.5% (S&P500)", "Low yield; return via buybacks + div growth"),
    ]

    for i, row_data in enumerate(multiples):
        r = 6 + i
        bg = ALT_ROW if i % 2 == 0 else WHITE
        for j, val in enumerate(row_data):
            lbl(ws, r, 2+j, val, bg=bg)

    # DCF Valuation
    sub_hdr(ws, 16, 2, "DCF Valuation Model (Owner's Perspective)", span=6)

    dcf_hdrs = ["Assumption", "Bear Case", "Base Case", "Bull Case", "Notes"]
    for i, h in enumerate(dcf_hdrs):
        sub_hdr(ws, 17, 2+i, h)

    dcf_data = [
        ("Revenue Growth (Yrs 1-5)",        "10%", "15%", "20%",  "Base: consistent cloud + Copilot ramp"),
        ("Revenue Growth (Yrs 6-10)",       "8%",  "12%", "15%",  "Deceleration as base grows"),
        ("Terminal Growth Rate",            "3%",  "4%",  "4.5%", "Above GDP given secular cloud tailwinds"),
        ("FCF Margin (normalized, Yrs 1-5)","28%", "33%", "38%",  "Base: FCF improves as CapEx cycle peaks in 2025"),
        ("WACC / Discount Rate",            "9%",  "8%",  "7%",   "Low WACC justified by AAA rating + stability"),
        ("Terminal FCF Multiple",           "25x", "30x", "35x",  "Comp: Visa ~30x, Apple ~35x"),
        ("", None, None, None, None),
        ("Implied Intrinsic Value ($/share)","$390", "$560", "$700", "Range of outcomes"),
        ("Current Price (Apr 13, 2026)",    "$373", "$373", "$373", ""),
        ("Implied Upside / Downside",       "+4.6%", "+50.1%", "+87.7%", "Base case = significant upside"),
        ("Margin of Safety",                "Modest", "STRONG (34%+ margin)", "N/A", "Base case offers >30% upside at current price"),
    ]

    for i, row_data in enumerate(dcf_data):
        r = 18 + i
        label = row_data[0]
        vals = row_data[1:]

        if label == "":
            ws.row_dimensions[r].height = 8
            continue

        is_key = label.startswith("Implied") or label.startswith("Current") or label.startswith("Margin")
        bg = SUBHDR_BG if is_key else (ALT_ROW if i % 2 == 0 else WHITE)
        bold = is_key

        lbl(ws, r, 2, label, bg=bg)
        if bold:
            ws.cell(row=r, column=2).font = mf(bold=True, size=FONT_SIZE)
        for j, val in enumerate(vals):
            if val is None:
                dc(ws, r, 3+j, "", bg=bg)
            else:
                wc(ws, r, 3+j, val, bold=bold, size=FONT_SIZE, bg=bg, align="center", border=True)

    # Analyst Consensus
    row_ac = 31
    sub_hdr(ws, row_ac, 2, "Wall Street Analyst Consensus (April 2026)", span=6)

    analyst_data = [
        ("Number of Analysts Covering",     "33 analysts"),
        ("Consensus Rating",                "STRONG BUY"),
        ("Average Price Target",            "$595.58"),
        ("Median Price Target",             "$600.00"),
        ("Low Price Target",                "$392.00 (base bear)"),
        ("High Price Target",               "$675.00 (Goldman Sachs bull)"),
        ("Current Price",                   "~$373"),
        ("Implied Upside to Avg Target",    "+59.7%"),
        ("% of Analysts with Buy Rating",   "94%"),
        ("Key Bull Arguments",              "Azure AI moat, Copilot monetization ramp, OpenAI exclusivity, FCF normalization post-CapEx cycle"),
        ("Key Bear Arguments",              "CapEx burn elevated, Copilot adoption slow (3.5%), regulatory risks, valuation premium to tech sector"),
        ("Safety of Margin Assessment",     "STRONG — trading at lowest P/E in 3 years with 60% consensus upside; $58B net cash; AAA rating"),
    ]

    for i, (k, v) in enumerate(analyst_data):
        r = row_ac + 1 + i
        bg = ALT_ROW if i % 2 == 0 else WHITE
        wc(ws, r, 2, k, bold=True, size=FONT_SIZE, bg=bg, align="left", border=True)
        wc(ws, r, 3, v, size=FONT_SIZE, bg=bg, align="left", border=True, wrap=True)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)
        ws.row_dimensions[r].height = 24


# ═══════════════════════════════════════════════════════════════════════════════
def build_sentiment(wb):
    ws = wb.create_sheet("Market Sentiment")
    ws.sheet_view.showGridLines = False
    scw(ws, 1, 3); scw(ws, 2, 32); scw(ws, 3, 28)
    scw(ws, 4, 20); scw(ws, 5, 20); scw(ws, 6, 20); scw(ws, 7, 20)

    sec_hdr(ws, 1, 2, "MARKET SENTIMENT — MICROSOFT CORPORATION (MSFT)", span=6)

    sentiment_items = [
        ("CURRENT STOCK PERFORMANCE", None),
        ("Stock Price (Apr 13, 2026)",            "~$373 — approximately flat YTD; down ~7-8% in last 30 days"),
        ("52-Week Range",                         "~$340 (low) – ~$468 (high)"),
        ("YTD Performance (2026)",                "-7% to -10% — underperforming broader market"),
        ("vs. S&P 500 YTD",                       "Lagging — MSFT has underperformed the market in early 2026"),
        ("Market Cap",                            "~$2.77 Trillion"),
        ("", None),
        ("INSTITUTIONAL OWNERSHIP", None),
        ("Institutional Ownership",               "~72% — held by large funds as 'core holding'"),
        ("Top Shareholders",                      "Vanguard (~9%), BlackRock (~7%), State Street (~4%), FMR/Fidelity (~3%)"),
        ("Institutional Stance",                  "Largely holding — Microsoft viewed as 'must-own utility for AI age'"),
        ("Short Interest",                        "Low (<1% of float) — minimal bearish positioning"),
        ("", None),
        ("SENTIMENT DRIVERS (BEARISH FACTORS)", None),
        ("AI ROI Skepticism",                     "Growing investor concern that $200B+ AI capex is not yet translating to proportional revenue; questioning ROI timeline"),
        ("Copilot Adoption Disappointment",       "Only ~3.5% of M365 users adopted Copilot vs. higher expectations; enterprise rollout slower than anticipated"),
        ("CapEx Compression of FCF",              "FCF declined to $71.6B (-3% YoY) vs. rising net income; FCF margin compressed from 30% to 25%"),
        ("OpenAI Scrutiny",                       "Questions about exclusivity as OpenAI considers IPO and broader distribution of models"),
        ("DeepSeek Disruption (Jan 2026)",        "Chinese AI model DeepSeek R1 demonstrated competitive performance at fraction of cost, causing broad AI stock selloff"),
        ("", None),
        ("SENTIMENT DRIVERS (BULLISH FACTORS)", None),
        ("Azure Growth Acceleration",             "Azure grew 40% in Q4 FY2025; management sees sustained 28%+ growth in FY2026; largest AI workload hosting provider"),
        ("Valuation Reset to 3-Year Lows",        "Trading at ~21x forward P/E — lowest valuation in 3 years; creates attractive entry for patient investors"),
        ("Monopoly-Level Enterprise Position",    "M365 in every Fortune 500; switching costs preventing churn; renewal rates near 100%"),
        ("Copilot Inflection Ahead",              "Enterprise Copilot penetration has long runway; each 1% increase in M365 base adds $1.5B+ ARR"),
        ("AI Infrastructure Supply Expanding",   "New Azure data centers coming online; GPU availability improving; should reduce capacity constraints in H1 2026"),
        ("", None),
        ("COMPETITIVE DYNAMICS", None),
        ("Azure Market Share",                    "24% and gaining — AWS at 30% (losing share), Google Cloud at 12% (growing)"),
        ("Enterprise AI Platform Race",           "Microsoft leads with Copilot + OpenAI integration; Google's Gemini competitive in specific verticals"),
        ("Key Trends to Watch",                   "Copilot seats growth, Azure AI share, FTE vs. AI agent licensing dynamics, regulatory outcomes"),
        ("", None),
        ("OVERALL SENTIMENT ASSESSMENT",          "CAUTIOUSLY BULLISH — Short-term headwinds (CapEx, Copilot adoption) overshadow medium-term AI monetization story; provides patient investors an unusually attractive entry into a world-class franchise at a discount to historical valuations."),
    ]

    for i, (k, v) in enumerate(sentiment_items):
        r = 3 + i
        if v is None and k.isupper():
            wc(ws, r, 2, k, bold=True, size=FONT_SIZE, fg=WHITE, bg=DARK_BLUE, align="left", border=True)
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
            ws.row_dimensions[r].height = 22
        elif k == "" and v is None:
            ws.row_dimensions[r].height = 8
        else:
            bg = ALT_ROW if i % 2 == 0 else WHITE
            wc(ws, r, 2, k, bold=True, size=FONT_SIZE, bg=bg, align="left", border=True)
            wc(ws, r, 3, v, size=FONT_SIZE, bg=bg, align="left", border=True, wrap=True)
            ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)
            ws.row_dimensions[r].height = 28


# ═══════════════════════════════════════════════════════════════════════════════
def build_key_indicators(wb):
    ws = wb.create_sheet("Key Indicators")
    ws.sheet_view.showGridLines = False
    scw(ws, 1, 3); scw(ws, 2, 34); scw(ws, 3, 17); scw(ws, 4, 17)
    scw(ws, 5, 17); scw(ws, 6, 17); scw(ws, 7, 17)

    sec_hdr(ws, 1, 2, "KEY INDICATORS — MICROSOFT CORPORATION (MSFT)", span=6)
    wc(ws, 2, 2, "Summary of key performance and financial metrics | April 2026",
       italic=True, size=12, fg="666666", align="left")

    years = ["FY2021", "FY2022", "FY2023", "FY2024", "FY2025"]
    sub_hdr(ws, 4, 2, "KPI", bg=SUBHDR_BG)
    for i, y in enumerate(years):
        sub_hdr(ws, 4, 3+i, y, bg=SUBHDR_BG)

    kpis = [
        ("GROWTH METRICS", None, None, None, None, None),
        ("  Revenue ($B)",                          168.1, 198.3, 211.9, 245.1, 281.7),
        ("  Revenue Growth %",                      "N/A", "18.0%", "6.9%", "15.7%", "14.9%"),
        ("  Azure Growth %",                        "~50%", "~40%", "~29%", "~29%", "~34%"),
        ("  Microsoft 365 Commercial Cloud Growth", "~27%", "~20%", "~17%", "~15%", "~18%"),
        ("", None, None, None, None, None),
        ("PROFITABILITY", None, None, None, None, None),
        ("  Gross Margin %",                        "69.0%", "68.4%", "68.9%", "69.8%", "68.8%"),
        ("  Operating Margin %",                    "41.6%", "42.1%", "41.8%", "44.6%", "45.6%"),
        ("  Net Margin %",                          "36.5%", "36.7%", "34.2%", "35.9%", "36.1%"),
        ("  FCF Margin %",                          "33.4%", "32.8%", "28.1%", "30.2%", "25.3%"),
        ("", None, None, None, None, None),
        ("RULE OF 50 (SaaS CHECK)", None, None, None, None, None),
        ("  Revenue Growth Rate",                   "N/A", "18.0%", "6.9%", "15.7%", "14.9%"),
        ("  FCF Margin",                            "33.4%", "32.8%", "28.1%", "30.2%", "25.3%"),
        ("  Rule of 50 Score",                      "N/A", "50.8", "35.0", "45.9", "40.2"),
        ("  Rule of 50 Result",                     "N/A", "PASS", "FAIL", "FAIL", "FAIL*"),
        ("  *Note: Normalized FCF (ex-AI CapEx surge) est. ~35%, giving Rule of 50 score ~50. Temporarily below threshold.", None, None, None, None, None),
        ("", None, None, None, None, None),
        ("RETURNS & CAPITAL", None, None, None, None, None),
        ("  ROE",                                   "43.2%", "43.7%", "35.1%", "29.6%", "29.5%"),
        ("  ROIC",                                  "28.3%", "30.2%", "27.0%", "25.0%", "24.5%"),
        ("  Free Cash Flow ($B)",                   56.1, 65.1, 59.5, 74.0, 71.4),
        ("  CapEx ($B)",                            20.6, 23.9, 28.1, 44.5, 64.6),
        ("  Dividends Per Share ($)",               2.24, 2.48, 2.72, 2.94, 3.32),
        ("", None, None, None, None, None),
        ("BALANCE SHEET HEALTH", None, None, None, None, None),
        ("  Total Assets ($B)",                     333.8, 364.8, 411.9, 523.0, 619.0),
        ("  Net Cash Position ($B)",                72.3, 54.9, 64.2, 53.1, 58.1),
        ("  Debt-to-Equity",                        "0.40x", "0.30x", "0.23x", "0.14x", "0.12x"),
        ("  Credit Rating",                         "Aaa/AAA", "Aaa/AAA", "Aaa/AAA", "Aaa/AAA", "Aaa/AAA"),
        ("", None, None, None, None, None),
        ("VALUATION", None, None, None, None, None),
        ("  EPS Diluted ($)",                       8.05, 9.65, 9.72, 11.80, 13.70),
        ("  P/E (TTM) at Period End",               "~35x", "~32x", "~34x", "~36x", "~23x"),
        ("  Forward P/E",                           "~30x", "~26x", "~28x", "~30x", "~21x"),
        ("  Market Cap ($T)",                       "~2.4T", "~1.8T", "~2.8T", "~3.3T", "~2.77T"),
        ("  Analyst Consensus",                     "N/A", "N/A", "N/A", "N/A", "STRONG BUY"),
        ("  Avg Price Target",                      "N/A", "N/A", "N/A", "N/A", "$595.58"),
    ]

    for i, row_data in enumerate(kpis):
        r = 5 + i
        label = row_data[0]
        values = row_data[1:]

        is_section = label.isupper() and label != "" and all(v is None for v in values)
        is_empty = label == "" and all(v is None for v in values)
        is_note = "*Note:" in label

        if is_empty:
            ws.row_dimensions[r].height = 8
            continue

        if is_section:
            wc(ws, r, 2, label, bold=True, size=FONT_SIZE, fg=WHITE, bg=DARK_BLUE, align="left", border=True)
            for j in range(5):
                wc(ws, r, 3+j, "", bg=DARK_BLUE, border=True)
            ws.row_dimensions[r].height = 22
            continue

        if is_note:
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
            wc(ws, r, 2, label, italic=True, size=12, fg="555555", bg=WARN_YELLOW,
               align="left", border=True, wrap=True)
            ws.row_dimensions[r].height = 28
            continue

        bg = ALT_ROW if i % 2 == 0 else WHITE
        lbl(ws, r, 2, label, bg=bg)

        for j, val in enumerate(values):
            if val is None:
                dc(ws, r, 3+j, "", bg=bg)
            elif isinstance(val, (float, int)):
                c = ws.cell(row=r, column=3+j, value=val)
                c.font = mf(size=FONT_SIZE)
                c.fill = mfill(bg)
                c.alignment = cal("right")
                c.border = mborder()
                c.number_format = '#,##0.0'
            else:
                dc(ws, r, 3+j, val, bg=bg)
        ws.row_dimensions[r].height = 20


# ═══════════════════════════════════════════════════════════════════════════════
def main():
    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    build_cover(wb)
    build_business(wb)
    build_moat(wb)
    build_income(wb)
    build_balance_sheet(wb)
    build_cashflow(wb)
    build_roic(wb)
    build_management(wb)
    build_risks(wb)
    build_valuation(wb)
    build_sentiment(wb)
    build_key_indicators(wb)

    out_dir = "/Users/naji/WORK/github.com/AI/claude/Agent/MarketResearch/output"
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, "MSFT_Financial_Analysis.xlsx")
    wb.save(out_path)
    print(f"Saved: {out_path}")


if __name__ == "__main__":
    main()
