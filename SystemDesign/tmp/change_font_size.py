from openpyxl import load_workbook
from openpyxl.styles import Font
import copy

input='CEG_Financial_Analysis'
output='CEG_Financial_Analysis_Resized'

wb = load_workbook(f'{input}.xlsx')

NEW_FONT_SIZE = 16  # Change this to whatever size you want

for sheet in wb.worksheets:
    for row in sheet.iter_rows():
        for cell in row:
            if cell.font:
                # Preserve existing font properties, only change size
                existing = cell.font
                cell.font = Font(
                    name=existing.name,
                    size=NEW_FONT_SIZE,
                    bold=existing.bold,
                    italic=existing.italic,
                    underline=existing.underline,
                    color=existing.color,
                    strike=existing.strike
                )
            else:
                cell.font = Font(size=NEW_FONT_SIZE)

wb.save(f'{output}.xlsx')
